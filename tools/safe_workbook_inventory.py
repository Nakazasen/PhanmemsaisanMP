#!/usr/bin/env python
"""Công cụ kiểm kê workbook Excel an toàn với giới hạn thời gian.

Mỗi workbook được mở trong một tiến trình con để tệp chậm hoặc hỏng không làm treo toàn bộ lô.
"""
from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.utils.cli import VietnameseArgumentParser

SHEET_NAME = "内訳ﾘｽﾄ(4～3月)"
EXCLUDED_ROWS = {1, 2, 3, 4, 5, 9, 17, 25}


def _norm(value):
    return "" if value is None else str(value).strip()


def _scan_one(path: Path) -> dict[str, object]:
    from openpyxl import load_workbook

    result: dict[str, object] = {
        "path": str(path.resolve()),
        "file_name": path.name,
        "contains_sheet": False,
        "business_row_count": 0,
        "has_5005026371": False,
        "has_5005066281": False,
        "has_5005066282": False,
        "sheet_names": "",
        "read_status": "OK",
    }
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        result["sheet_names"] = "|".join(wb.sheetnames)
        if SHEET_NAME not in wb.sheetnames:
            result["read_status"] = "NO_SHEET"
            return result
        result["contains_sheet"] = True
        ws = wb[SHEET_NAME]
        business = 0
        has_6371 = has_6281 = has_6282 = False
        for row in range(1, ws.max_row + 1):
            if row in EXCLUDED_ROWS:
                continue
            account = _norm(ws.cell(row, 2).value)
            description = _norm(ws.cell(row, 19).value)
            months = [_norm(ws.cell(row, col).value) for col in range(6, 18)]
            if account or description or any(months):
                business += 1
            has_6371 = has_6371 or account == "5005026371"
            has_6281 = has_6281 or account == "5005066281"
            has_6282 = has_6282 or account == "5005066282"
        result.update(
            business_row_count=business,
            has_5005026371=has_6371,
            has_5005066281=has_6281,
            has_5005066282=has_6282,
        )
        return result
    finally:
        wb.close()


def child_main() -> int:
    path = Path(sys.argv[2])
    try:
        print(json.dumps(_scan_one(path), ensure_ascii=False))
        return 0
    except Exception:  # pragma: no cover - child safety path
        print(json.dumps({"path": str(path.resolve()), "file_name": path.name, "read_status": "READ_ERROR:WORKBOOK_UNREADABLE"}, ensure_ascii=False))
        return 0


def discover(path: Path) -> list[Path]:
    if path.is_file():
        return [path]
    return sorted([p for p in path.rglob("*.xlsx") if p.is_file()])


def main() -> int:
    if len(sys.argv) > 1 and sys.argv[1] == "--child-scan":
        return child_main()
    parser = VietnameseArgumentParser(description="Kiểm kê workbook an toàn với giới hạn thời gian")
    parser.add_argument("path", nargs="?", help="Đường dẫn workbook hoặc thư mục")
    parser.add_argument("--input", dest="input_path", help="Đường dẫn workbook hoặc thư mục")
    parser.add_argument("--output", required=True, help="Đường dẫn tệp CSV kết quả")
    parser.add_argument("--timeout", "--timeout-seconds", dest="timeout", type=float, default=20.0, help="Giới hạn thời gian cho mỗi workbook, tính bằng giây")
    parser.add_argument("--limit", type=int, default=0, help="Số workbook tối đa cần quét; 0 là không giới hạn")
    args = parser.parse_args()

    input_path = args.input_path or args.path
    if not input_path:
        parser.error("bắt buộc phải có path hoặc --input")
    files = discover(Path(input_path))
    if args.limit:
        files = files[: args.limit]
    rows: list[dict[str, object]] = []
    for workbook in files:
        try:
            proc = subprocess.run(
                [sys.executable, str(Path(__file__).resolve()), "--child-scan", str(workbook)],
                text=True,
                capture_output=True,
                timeout=args.timeout,
            )
            if proc.stdout.strip():
                row = json.loads(proc.stdout.strip().splitlines()[-1])
            else:
                row = {"path": str(workbook.resolve()), "file_name": workbook.name, "read_status": "READ_ERROR:NO_STDOUT"}
        except subprocess.TimeoutExpired:
            row = {"path": str(workbook.resolve()), "file_name": workbook.name, "read_status": "READ_TIMEOUT"}
        row.setdefault("contains_sheet", "")
        row.setdefault("business_row_count", "")
        row.setdefault("has_5005026371", "")
        row.setdefault("has_5005066281", "")
        row.setdefault("has_5005066282", "")
        row.setdefault("sheet_names", "")
        rows.append(row)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["path", "file_name", "contains_sheet", "business_row_count", "has_5005026371", "has_5005066281", "has_5005066282", "sheet_names", "read_status"]
    with out.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})
    print(json.dumps({"scanned": len(rows), "output": str(out), "timeouts": sum(1 for r in rows if r.get("read_status") == "READ_TIMEOUT")}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
