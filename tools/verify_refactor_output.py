"""Compare Excel outputs from a baseline and a refactored MP2027 build."""
from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from hashlib import sha256
import json
from pathlib import Path
import shutil
import sqlite3
import subprocess
import sys
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _configure_console_encoding() -> None:
    """Avoid cp932/cp1252 failures when this Vietnamese CLI prints help text."""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except (OSError, ValueError):
                pass


_configure_console_encoding()

DETAIL_SHEET = "内訳ﾘｽﾄ(4～3月)"
OUTPUT_PATTERN = "MP_CC_*.xlsx"
PAYLOAD_START_ROW = 38
LOCKED_ROWS = range(31, 37)
LOCKED_COLUMNS = range(2, 21)  # B:T


@dataclass(frozen=True)
class CellDifference:
    coordinate: str
    baseline: Any
    candidate: Any


@dataclass
class CostCenterComparison:
    cost_center: str
    baseline_file: str | None
    candidate_file: str | None
    baseline_data_rows: int | None = None
    candidate_data_rows: int | None = None
    row_count_passed: bool = False
    locked_range_passed: bool = False
    locked_differences: list[CellDifference] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.errors and self.row_count_passed and self.locked_range_passed


@dataclass
class ComparisonEvidence:
    baseline_output: str
    candidate_output: str
    generated_at: str
    results: list[CostCenterComparison]
    baseline_ref: str | None = None
    baseline_commit: str | None = None
    input_fingerprints: dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return bool(self.results) and all(result.passed for result in self.results)


def _safe_value(value: Any) -> Any:
    return value if value is None or isinstance(value, (str, int, float, bool)) else str(value)


def _cost_center(path: Path) -> str:
    return path.stem.removeprefix("MP_CC_")


def _find_outputs(directory: Path) -> dict[str, Path]:
    if not directory.is_dir():
        raise FileNotFoundError(f"Không tìm thấy thư mục output: {directory}")
    outputs = {_cost_center(path): path for path in sorted(directory.glob(OUTPUT_PATTERN))}
    if not outputs:
        raise FileNotFoundError(f"Không có file {OUTPUT_PATTERN} trong: {directory}")
    return outputs


def _generated_row_count(worksheet) -> int:
    """Count populated generated rows, from row 38, including formulas."""
    return sum(
        any(worksheet.cell(row=row, column=column).value not in (None, "") for column in LOCKED_COLUMNS)
        for row in range(PAYLOAD_START_ROW, max(PAYLOAD_START_ROW - 1, worksheet.max_row) + 1)
    )


def _locked_differences(baseline_ws, candidate_ws) -> list[CellDifference]:
    differences: list[CellDifference] = []
    for row in LOCKED_ROWS:
        for column in LOCKED_COLUMNS:
            baseline = baseline_ws.cell(row=row, column=column).value
            candidate = candidate_ws.cell(row=row, column=column).value
            if baseline != candidate:
                differences.append(CellDifference(
                    f"{get_column_letter(column)}{row}", _safe_value(baseline), _safe_value(candidate)
                ))
    return differences


def _compare_pair(cost_center: str, baseline_path: Path | None, candidate_path: Path | None) -> CostCenterComparison:
    result = CostCenterComparison(
        cost_center, str(baseline_path) if baseline_path else None, str(candidate_path) if candidate_path else None
    )
    if baseline_path is None:
        result.errors.append("Thiếu workbook ở baseline.")
        return result
    if candidate_path is None:
        result.errors.append("Thiếu workbook ở bản refactor.")
        return result
    baseline_wb = candidate_wb = None
    try:
        baseline_wb = openpyxl.load_workbook(baseline_path, data_only=False)
        candidate_wb = openpyxl.load_workbook(candidate_path, data_only=False)
        for label, workbook in (("baseline", baseline_wb), ("refactor", candidate_wb)):
            if DETAIL_SHEET not in workbook.sheetnames:
                result.errors.append(f"Workbook {label} thiếu sheet {DETAIL_SHEET}.")
        if result.errors:
            return result
        baseline_ws, candidate_ws = baseline_wb[DETAIL_SHEET], candidate_wb[DETAIL_SHEET]
        result.baseline_data_rows = _generated_row_count(baseline_ws)
        result.candidate_data_rows = _generated_row_count(candidate_ws)
        result.row_count_passed = result.baseline_data_rows == result.candidate_data_rows
        result.locked_differences = _locked_differences(baseline_ws, candidate_ws)
        result.locked_range_passed = not result.locked_differences
        return result
    except Exception as exc:
        result.errors.append(f"Không thể đọc workbook: {type(exc).__name__}: {exc}")
        return result
    finally:
        if baseline_wb:
            baseline_wb.close()
        if candidate_wb:
            candidate_wb.close()


def compare_output_directories(
    baseline_output: Path, candidate_output: Path, *, baseline_ref: str | None = None,
    baseline_commit: str | None = None, input_fingerprints: dict[str, Any] | None = None,
) -> ComparisonEvidence:
    """Compare already-generated output folders without running the pipeline."""
    baseline_files, candidate_files = _find_outputs(baseline_output), _find_outputs(candidate_output)
    codes = sorted(set(baseline_files) | set(candidate_files), key=lambda value: (len(value), value))
    return ComparisonEvidence(
        str(baseline_output), str(candidate_output), datetime.now(timezone.utc).isoformat(),
        [_compare_pair(code, baseline_files.get(code), candidate_files.get(code)) for code in codes],
        baseline_ref, baseline_commit, input_fingerprints or {},
    )


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _fingerprint(path: Path) -> dict[str, Any]:
    if path.is_file():
        return {"kind": "file", "path": str(path), "size": path.stat().st_size, "sha256": _sha256_file(path)}
    digest, count = sha256(), 0
    for file_path in sorted(item for item in path.rglob("*") if item.is_file()):
        digest.update(file_path.relative_to(path).as_posix().encode("utf-8"))
        digest.update(_sha256_file(file_path).encode("ascii"))
        count += 1
    return {"kind": "directory", "path": str(path), "files": count, "sha256": digest.hexdigest()}


def _snapshot_sqlite(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_connection = sqlite3.connect(f"file:{source.resolve().as_posix()}?mode=ro", uri=True)
    target_connection = sqlite3.connect(destination)
    try:
        source_connection.backup(target_connection)
        target_connection.commit()
    finally:
        target_connection.close()
        source_connection.close()


def _snapshot_input(source: Path, destination: Path) -> Path:
    """Copy an input into the evidence workspace without ever editing its source."""
    if source.is_file():
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
    else:
        shutil.copytree(source, destination)
    return destination


def _git(project_root: Path, *args: str) -> str:
    completed = subprocess.run(
        ["git", "-C", str(project_root), *args], check=True, text=True,
        encoding="utf-8", stdout=subprocess.PIPE, stderr=subprocess.PIPE,
    )
    return completed.stdout.strip()


def _pipeline_command(
    *, fiscal_year: int, template: Path, source: Path, headcount_source: Path | None,
    uniform_policy: Path | None, operational_db: Path, manual_input_store: Path,
    output_dir: Path, history_root: Path, exchange_rate: float, target_cc: str | None,
    simulate_baseline_t3_from_t4: bool = False,
) -> list[str]:
    command = [
        sys.executable, "scripts/run_e2e.py", "--fy", str(fiscal_year),
        "--template", str(template), "--source", str(source),
        "--operational-db", str(operational_db), "--manual-input-store", str(manual_input_store),
        "--output-dir", str(output_dir), "--run-history-root", str(history_root),
        "--exchange-rate", str(exchange_rate), "--no-run-history",
    ]
    if headcount_source:
        command.extend(("--headcount-source", str(headcount_source)))
    if uniform_policy:
        command.extend(("--uniform-policy", str(uniform_policy)))
    if target_cc:
        command.extend(("--target-cc", target_cc))
    return command


def _populate_audit_baselines_from_april(
    operational_db: Path,
    manual_input_store: Path,
    fiscal_year: int,
    target_cc: str | None,
) -> int:
    """Copy T04 to missing T03 manual rows inside one evidence-only DB snapshot."""
    baseline_period, april_period = f"{fiscal_year - 1}03", f"{fiscal_year - 1}04"
    connection = sqlite3.connect(operational_db)
    try:
        target_cc_clause = "AND CAST(april.cc_code AS TEXT)=?" if target_cc else ""
        params: list[Any] = [april_period]
        if target_cc:
            params.append(str(target_cc).strip())
        april_rows = connection.execute(
            f"""SELECT april.cc_code, april.headcount_all, april.headcount_expat,
                       april.headcount_staff, april.headcount_worker, april.headcount_male,
                       april.headcount_female, april.split_status, april.headcount_local_total,
                       april.source_file, april.source_sheet
                FROM fact_monthly_headcount AS april
                WHERE april.period=? AND april.source='department_plan' {target_cc_clause}
                ORDER BY CAST(april.cc_code AS TEXT)""",
            params,
        ).fetchall()
        rows = []
        for row in april_rows:
            existing_manual = connection.execute(
                """SELECT 1 FROM fact_monthly_headcount
                   WHERE period=? AND cc_code=? AND source='manual'""",
                (baseline_period, row[0]),
            ).fetchone()
            if not existing_manual:
                rows.append((
                    baseline_period, row[0], *row[1:9], "manual",
                    "SIMULATED_BASELINE_T3_FROM_T4_AUDIT_ONLY", row[9], row[10],
                ))
        connection.executemany(
            """INSERT INTO fact_monthly_headcount
               (period,cc_code,headcount_all,headcount_expat,headcount_staff,headcount_worker,
                headcount_male,headcount_female,split_status,headcount_local_total,source,
                description,source_file,source_sheet,imported_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
            rows,
        )
        connection.commit()
        if rows:
            annual_store = sqlite3.connect(manual_input_store)
            try:
                annual_store.executemany(
                    """INSERT OR IGNORE INTO fact_manual_headcount_baseline_override
                       (period,cc_code,fiscal_year,headcount_all,headcount_expat,headcount_staff,
                        headcount_worker,headcount_male,headcount_female,split_status,
                        headcount_local_total,description,source_file,source_sheet,updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                    [(period, cc, fiscal_year, total, expat, staff, worker, male, female,
                      status, local_total, description, source_file, source_sheet)
                     for (period, cc, total, expat, staff, worker, male, female, status,
                          local_total, _source, description, source_file, source_sheet) in rows],
                )
                annual_store.commit()
            finally:
                annual_store.close()
        return len(rows)
    finally:
        connection.close()




def _append_audit_baselines_to_manual_csv(
    operational_db: Path,
    source_dir: Path,
    fiscal_year: int,
    target_cc: str | None,
) -> int:
    """Append missing audit-only T03 rows to the isolated manual CSV input."""
    csv_path = source_dir / "headcount_manual.csv"
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        existing = {(str(row.get("cc_code", "")).strip(), str(row.get("period", "")).strip()) for row in reader}
    required = ["cc_code", "period", "headcount_staff", "headcount_worker", "headcount_male", "headcount_female", "description", "headcount_expat"]
    if fieldnames != required:
        raise ValueError(f"Tệp nhân sự thủ công tạm có cột không đúng: {csv_path}")
    baseline, april = f"{fiscal_year - 1}03", f"{fiscal_year - 1}04"
    conn = sqlite3.connect(operational_db)
    try:
        clause = "AND CAST(cc_code AS TEXT)=?" if target_cc else ""
        params: list[Any] = [april]
        if target_cc:
            params.append(str(target_cc).strip())
        april_rows = conn.execute(
            f"""SELECT CAST(cc_code AS TEXT),headcount_staff,headcount_worker,
                       headcount_male,headcount_female,headcount_expat
                FROM fact_monthly_headcount
                WHERE period=? AND source='department_plan' {clause}
                ORDER BY CAST(cc_code AS TEXT)""",
            params,
        ).fetchall()
    finally:
        conn.close()
    new_rows = [
        {
            "cc_code": cc, "period": baseline,
            "headcount_staff": staff, "headcount_worker": worker,
            "headcount_male": male or 0, "headcount_female": female or 0,
            "description": "SIMULATED_BASELINE_T3_FROM_T4_AUDIT_ONLY",
            "headcount_expat": expat or 0,
        }
        for cc, staff, worker, male, female, expat in april_rows
        if (cc, baseline) not in existing
    ]
    if new_rows:
        with csv_path.open("a", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=required)
            writer.writerows(new_rows)
    return len(new_rows)


def _run_pipeline(
    project_root: Path, command: list[str], log_path: Path, *, allow_failure: bool = False,
) -> None:
    with log_path.open("w", encoding="utf-8") as log_file:
        completed = subprocess.run(command, cwd=project_root, stdout=log_file, stderr=subprocess.STDOUT, text=True)
    if completed.returncode and not allow_failure:
        raise RuntimeError(f"Pipeline thất bại. Xem log: {log_path}")


def run_two_version_comparison(args: argparse.Namespace) -> ComparisonEvidence:
    """Run baseline/refactor pipelines in isolated workspaces, then compare them."""
    project_root, report_dir = args.project_root.resolve(), args.report_dir.resolve()
    run_root = report_dir / "run_workspaces"
    baseline_root = run_root / "baseline_worktree"
    baseline_output, candidate_output = run_root / "baseline_output", run_root / "candidate_output"
    if run_root.exists():
        raise FileExistsError(f"Evidence workspace đã tồn tại: {run_root}")
    inputs = {
        "template": args.template.resolve(), "source": args.source.resolve(),
        "operational_db": args.operational_db.resolve(), "manual_input_store": args.manual_input_store.resolve(),
    }
    if args.headcount_source:
        inputs["headcount_source"] = args.headcount_source.resolve()
    if args.uniform_policy:
        inputs["uniform_policy"] = args.uniform_policy.resolve()
    missing = [name for name, path in inputs.items() if not path.exists()]
    if missing:
        raise FileNotFoundError("Thiếu input: " + ", ".join(missing))
    run_root.mkdir(parents=True)
    baseline_ref, baseline_commit = args.baseline_ref, _git(project_root, "rev-parse", args.baseline_ref)
    added = False
    try:
        subprocess.run(["git", "-C", str(project_root), "worktree", "add", "--detach", str(baseline_root), baseline_ref], check=True)
        added = True
        for label, root, output in (("baseline", baseline_root, baseline_output), ("refactor", project_root, candidate_output)):
            data = run_root / f"{label}_data"
            snapshot_inputs = {
                "template": _snapshot_input(inputs["template"], data / "FORM.xlsx"),
                "source": _snapshot_input(inputs["source"], data / "source"),
                "operational_db": data / "mp2027.db",
                "manual_input_store": data / "manual_inputs.db",
            }
            _snapshot_sqlite(inputs["operational_db"], snapshot_inputs["operational_db"])
            _snapshot_sqlite(inputs["manual_input_store"], snapshot_inputs["manual_input_store"])
            # The receipt records a migration into the original annual store.  It must
            # not suppress migration into this new isolated copy, otherwise approved
            # FY2027 T3 manual baselines from the operational snapshot are absent.
            if args.fy == 2027:
                receipt = snapshot_inputs["manual_input_store"].with_name("manual_inputs_migration_fy2027.json")
                receipt.unlink(missing_ok=True)
            if "headcount_source" in inputs:
                snapshot_inputs["headcount_source"] = _snapshot_input(
                    inputs["headcount_source"], data / "headcount_source"
                )
            if "uniform_policy" in inputs:
                snapshot_inputs["uniform_policy"] = _snapshot_input(
                    inputs["uniform_policy"], data / inputs["uniform_policy"].name
                )
            command = _pipeline_command(
                fiscal_year=args.fy, template=snapshot_inputs["template"], source=snapshot_inputs["source"],
                headcount_source=snapshot_inputs.get("headcount_source"),
                uniform_policy=snapshot_inputs.get("uniform_policy"),
                operational_db=snapshot_inputs["operational_db"], manual_input_store=snapshot_inputs["manual_input_store"],
                output_dir=output, history_root=run_root / f"{label}_history",
                exchange_rate=args.exchange_rate, target_cc=args.target_cc,
            )
            if args.simulate_baseline_t3_from_t4:
                # First pass imports the genuine T04 staffing source into the isolated
                # snapshot. It may stop at the expected missing-T03 safeguard.
                _run_pipeline(root, command, report_dir / f"pipeline_{label}_seed.log", allow_failure=True)
                copied = _populate_audit_baselines_from_april(
                    snapshot_inputs["operational_db"], snapshot_inputs["manual_input_store"],
                    args.fy, args.target_cc,
                )
                csv_rows = _append_audit_baselines_to_manual_csv(
                    snapshot_inputs["operational_db"], snapshot_inputs["source"],
                    args.fy, args.target_cc,
                )
                if copied == 0 and csv_rows == 0:
                    raise RuntimeError(
                        "Không tạo được baseline T03 tạm từ dữ liệu T04 sau lượt nạp nguồn thật."
                    )
                command = _pipeline_command(
                    fiscal_year=args.fy, template=snapshot_inputs["template"], source=snapshot_inputs["source"],
                    headcount_source=snapshot_inputs.get("headcount_source"), uniform_policy=snapshot_inputs.get("uniform_policy"),
                    operational_db=snapshot_inputs["operational_db"], manual_input_store=snapshot_inputs["manual_input_store"],
                    output_dir=output, history_root=run_root / f"{label}_history",
                    exchange_rate=args.exchange_rate, target_cc=args.target_cc,
                )
            _run_pipeline(root, command, report_dir / f"pipeline_{label}.log")
        return compare_output_directories(
            baseline_output, candidate_output, baseline_ref=baseline_ref, baseline_commit=baseline_commit,
            input_fingerprints={name: _fingerprint(path) for name, path in inputs.items()},
        )
    finally:
        if added:
            subprocess.run(["git", "-C", str(project_root), "worktree", "remove", "--force", str(baseline_root)], check=False)


def _write_excel(path: Path, evidence: ComparisonEvidence) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Tóm_tắt"
    for row in (
        ("Kết quả", "PASS" if evidence.passed else "FAIL"),
        ("Baseline ref", evidence.baseline_ref or "(output có sẵn)"),
        ("Baseline commit", evidence.baseline_commit or ""),
        ("Output baseline", evidence.baseline_output),
        ("Output refactor", evidence.candidate_output),
        ("Thời điểm", evidence.generated_at), (),
        ("CC", "Dòng baseline", "Dòng refactor", "Số dòng", "B31:T36", "Kết quả"),
    ):
        summary.append(row)
    for result in evidence.results:
        summary.append((result.cost_center, result.baseline_data_rows, result.candidate_data_rows,
                        "PASS" if result.row_count_passed else "FAIL",
                        "PASS" if result.locked_range_passed else "FAIL", "PASS" if result.passed else "FAIL"))
    diff_sheet = workbook.create_sheet("Khác_biệt_B31_T36")
    diff_sheet.append(("CC", "Ô", "Baseline", "Refactor"))
    for result in evidence.results:
        for difference in result.locked_differences:
            diff_sheet.append((result.cost_center, difference.coordinate, difference.baseline, difference.candidate))
        for error in result.errors:
            diff_sheet.append((result.cost_center, "LỖI", error, ""))
    source_sheet = workbook.create_sheet("Dấu_vết_input")
    source_sheet.append(("Input", "Fingerprint"))
    for name, fingerprint in sorted(evidence.input_fingerprints.items()):
        source_sheet.append((name, json.dumps(fingerprint, ensure_ascii=False)))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cells in sheet.columns:
            letter = get_column_letter(cells[0].column)
            sheet.column_dimensions[letter].width = min(80, max(12, max(len(str(cell.value or "")) for cell in cells) + 2))
    workbook.save(path)
    workbook.close()


def write_evidence(report_dir: Path, evidence: ComparisonEvidence) -> tuple[Path, Path]:
    report_dir.mkdir(parents=True, exist_ok=True)
    json_path, xlsx_path = report_dir / "refactor_output_verification.json", report_dir / "refactor_output_verification.xlsx"
    json_path.write_text(json.dumps(asdict(evidence), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_excel(xlsx_path, evidence)
    return json_path, xlsx_path


def _print_summary(evidence: ComparisonEvidence) -> None:
    for result in evidence.results:
        print(f"CC {result.cost_center}")
        if result.baseline_data_rows is not None:
            print(f"  Số dòng dữ liệu: {result.baseline_data_rows} → {result.candidate_data_rows}  {'PASS' if result.row_count_passed else 'FAIL'}")
        if result.locked_range_passed:
            print("  B31:T36: giống hoàn toàn  PASS")
        else:
            for difference in result.locked_differences:
                print(f"  B31:T36: {difference.coordinate} khác: {difference.baseline!r} → {difference.candidate!r}  FAIL")
        for error in result.errors:
            print(f"  Lỗi: {error}  FAIL")
    print("\nKẾT QUẢ: " + ("PASS" if evidence.passed else "FAIL"))


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="So sánh output Excel baseline Git và bản refactor.")
    parser.add_argument("--baseline-output", type=Path, help="Thư mục output đã tạo từ baseline")
    parser.add_argument("--candidate-output", type=Path, help="Thư mục output từ bản refactor")
    parser.add_argument("--report-dir", required=True, type=Path, help="Thư mục ghi evidence")
    parser.add_argument("--run-pipelines", action="store_true", help="Chạy hai pipeline trong workspace cách ly")
    parser.add_argument("--project-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--baseline-ref", default="HEAD")
    parser.add_argument("--fy", type=int, default=2027)
    parser.add_argument("--template", type=Path)
    parser.add_argument("--source", type=Path)
    parser.add_argument("--headcount-source", type=Path)
    parser.add_argument("--uniform-policy", type=Path)
    parser.add_argument("--operational-db", type=Path)
    parser.add_argument("--manual-input-store", type=Path)
    parser.add_argument("--exchange-rate", type=float, default=26273.0)
    parser.add_argument("--target-cc", type=str)
    parser.add_argument(
        "--simulate-baseline-t3-from-t4", action="store_true",
        help="Chỉ chạy kiểm chứng: tạm dùng T04/2026 làm baseline T03/2026 trong vùng cách ly; không sửa dữ liệu thật.",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    try:
        if args.run_pipelines:
            missing = [f"--{name.replace('_', '-')}" for name in ("template", "source", "operational_db", "manual_input_store") if getattr(args, name) is None]
            if missing:
                raise ValueError("Khi dùng --run-pipelines phải có: " + ", ".join(missing))
            evidence = run_two_version_comparison(args)
        else:
            if not args.baseline_output or not args.candidate_output:
                raise ValueError("Cần --baseline-output và --candidate-output, hoặc dùng --run-pipelines.")
            evidence = compare_output_directories(args.baseline_output, args.candidate_output)
        json_path, xlsx_path = write_evidence(args.report_dir, evidence)
        _print_summary(evidence)
        print(f"JSON evidence: {json_path}")
        print(f"Excel evidence: {xlsx_path}")
        return 0 if evidence.passed else 2
    except Exception as exc:
        print(f"LỖI: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
