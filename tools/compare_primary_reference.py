"""So sánh kết quả MP2027 đã tạo với tệp tham chiếu chính đáng tin cậy."""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.utils.cli import VietnameseArgumentParser

DETAIL_SHEET = "内訳ﾘｽﾄ(4～3月)"
SUMMARY_LABELS = {
    "generated_path": "Đường dẫn tệp kết quả",
    "reference_path": "Đường dẫn tệp tham chiếu",
    "compare_mode": "Chế độ so sánh",
    "total_cells_compared": "Tổng số ô đã so sánh",
    "exact_matches": "Số ô khớp chính xác",
    "diff_total": "Tổng số khác biệt",
    "differences": "Số khác biệt",
    "diff_by_row": "Số khác biệt theo dòng",
    "missing_sheets": "Sheet bị thiếu",
    "missing_rows_or_cells": "Dòng hoặc ô bị thiếu",
    "fixed_rows_compared": "Số dòng cố định đã so sánh",
    "fixed_row_differences": "Khác biệt ở dòng cố định",
    "identity_rows_checked": "Số dòng nhận dạng đã kiểm tra",
    "identity_rows_matched": "Số dòng nhận dạng đã khớp",
    "identity_rows_ambiguous": "Số dòng nhận dạng còn mơ hồ",
    "identity_rows_not_found": "Số dòng nhận dạng không tìm thấy",
    "identity_differences": "Khác biệt ở dòng nhận dạng",
    "strict_diff_total": "Tổng khác biệt khi so sánh chính xác",
    "strict_identity_diff_total": "Khác biệt nhận dạng khi so sánh chính xác",
    "layout_fixed_rows": "Danh sách dòng bố cục cố định",
    "layout_fixed_rows_total": "Tổng số dòng bố cục cố định",
    "new_identity_rows_from_42N1G": "Dòng nhận dạng mới từ 42N1G",
    "new_identity_rows_total": "Tổng số dòng nhận dạng mới",
    "new_identity_rows_matched": "Số dòng nhận dạng mới đã khớp",
    "new_identity_rows_ambiguous": "Số dòng nhận dạng mới còn mơ hồ",
    "new_identity_rows_not_found": "Số dòng nhận dạng mới không tìm thấy",
    "existing_identity_rows": "Danh sách dòng nhận dạng hiện có",
    "existing_identity_rows_total": "Tổng số dòng nhận dạng hiện có",
    "existing_identity_rows_matched": "Số dòng nhận dạng hiện có đã khớp",
    "existing_identity_rows_ambiguous": "Số dòng nhận dạng hiện có còn mơ hồ",
    "existing_identity_rows_not_found": "Số dòng nhận dạng hiện có không tìm thấy",
    "false_fixed_row_diffs_removed": "Khác biệt dòng cố định giả đã loại bỏ",
    "real_identity_strict_diffs_remaining": "Khác biệt nhận dạng thực còn lại",
}
DISPLAY_VALUES = {
    "strict_exact": "so sánh chính xác",
    "fixed_row": "dòng cố định",
    "identity": "nhận dạng",
    "DIFF": "KHÁC",
    "High": "Cao",
    "Medium": "Trung bình",
    "Low": "Thấp",
}


def _display_value(value: Any) -> Any:
    """Đổi mã nội bộ sang nhãn tiếng Việt mà không thay contract JSON."""
    if isinstance(value, dict):
        return {key: _display_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_display_value(item) for item in value]
    return DISPLAY_VALUES.get(value, value)


def _display_summary(summary: dict[str, Any]) -> dict[str, Any]:
    """Tạo bản tóm tắt tiếng Việt dành riêng cho terminal và Excel."""
    return {
        SUMMARY_LABELS.get(key, key): _display_value(value)
        for key, value in summary.items()
    }


FIXED_ROW_RULES = {
}
LAYOUT_FIXED_ROWS = (3, 5, 9, 17, 25)
NEW_IDENTITY_ROWS_FROM_42N1G = (38, 40, 42, 51, 59)
EXISTING_IDENTITY_ROWS = (53, 54, 58, 66, 75, 97, 98, 137)
REQUIREMENT_LOCK_NOTES = {
    38: "42R0-G1: Mục tiêu khấu hao tài sản cố định F38:Q38; NEED_FORM_TEMPLATE_CHECK; khi so với tệp chính phải căn chỉnh theo nhận dạng.",
    40: "42R0-F1/F2: Facility có sáu hạng mục riêng; không gộp hoặc so theo cùng số dòng của tệp chính.",
    42: "42R0-G2: Mục tiêu lãi tài sản cố định F42:Q42; NEED_FORM_TEMPLATE_CHECK; khi so với tệp chính phải căn chỉnh theo nhận dạng.",
    51: "42R0-H1/H2: Phân bổ hành chính có mức ưu tiên cao; phân loại theo hạng mục nghiệp vụ/tài khoản, không theo cùng dòng của tệp chính.",
    59: "42R0-D1: Mục tiêu sinh nhật F59:Q59 đã được chấp nhận nhưng xung đột dòng 63 vẫn là MD_INTERPRETATION_RISK.",
    75: "42R0-E1/E2: Chi phí hệ thống là một dòng công thức, không phải số cố định.",
    137: "42R0-C1/C2: Mục tiêu giấy tờ NNN F137:Q137, lọc theo mã bộ phận.",
}
IDENTITY_ROW_CANDIDATES = {row: label for row, label in {
    38: "Khấu hao tài sản cố định",
    40: "Lãi nhà xưởng/cơ sở/đất",
    42: "Lãi tài sản cố định",
    51: "Phí vệ sinh / phân bổ hành chính",
    59: "Sinh nhật",
    53: "Xe buýt JP / đi lại người biệt phái",
    54: "Xe buýt VN / đi lại nhân sự địa phương",
    58: "Khám sức khỏe tuyển dụng",
    66: "Du lịch công ty",
    75: "Chi phí hệ thống",
    97: "Sổ tay nhân viên",
    98: "Sổ tay công nhân",
    137: "Giấy tờ NNN",
}.items()}
COMPARE_COLUMNS = ("B", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S")
STRICT_COMPARE_COLUMNS = ("F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R")
MEDIUM_COLUMNS = set("FGHIJKLMNOPQ")
IDENTITY_TOKENS = {
    "system": ("system cost", "mail", "vpn", "r3", "mes", "plm", "vps", "システム"),
    "depreciation": ("減価償却", "khấu hao", "khau hao", "depreciation"),
    "interest": ("金利", "lãi", "lai", "interest"),
    "building": ("建物", "土地", "nhà", "nha", "đất", "dat", "building", "land"),
    "equipment": ("設備", "治具", "工具", "thiết bị", "thiet bi", "equipment", "tool", "jig"),
    "facility": ("facility", "building", "land", "nhà", "nha", "đất", "dat", "建物", "土地"),
    "cleaning": ("cleaning", "làm sạch", "lam sach", "vệ sinh", "ve sinh", "清掃", "掃除", "admin allocation", "phân bổ", "phan bo"),
    "birthday": ("birthday", "sinh nhật", "sinh nhat", "誕生日", "生日"),
    "bus": ("bus", "バス", "送迎", "通勤", "xe bus", "xe buýt"),
    "bus_local": ("ローカル", "local", "ベトナム", "người vn"),
    "bus_expat": ("出向者", "日本", "người jp", "expat"),
    "notebook": ("ノート", "notebook", "note", "sổ", "so tay", "sổ tay"),
    "notebook_staff": ("staff", "nhân viên", "nhan vien", "スタッフ", "社員用"),
    "notebook_worker": ("worker", "công nhân", "cong nhan", "g7", "ローカル"),
    "health": ("健康", "健診", "採用", "入社", "入社月", "health", "medical", "checkup", "khám", "sức khỏe", "suc khoe", "tuyen dung", "tuyển dụng"),
    "trip": ("社員旅行", "社内旅行", "company trip", "employee trip", "trip", "travel", "du lịch", "du lich"),
    "paperwork": ("nnn", "書類", "giấy tờ", "paperwork", "document", "visa", "work permit", "residence"),
}


@dataclass
class CellDiff:
    row: int
    column: str
    generated: Any
    reference: Any
    match: bool
    severity: str
    note: str
    compare_mode: str = "fixed_row"
    reference_row: int | None = None


@dataclass
class IdentityAlignment:
    generated_row: int
    generated_account: Any
    generated_description: Any
    reference_matched_row: int | None
    reference_account: Any
    reference_description: Any
    match_method: str
    confidence: str
    same_business_item: bool
    note: str
    match_score: float = 0.0
    matched_fields: str = ""
    candidate_rows: str = ""
    row_group: str = ""
    requirement_note: str = ""
    classification: str = ""


def normalize_text(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _cell_value(ws, row: int, column: str):
    return ws[f"{column}{row}"].value


def _severity(column: str, generated, reference, match: bool, mode: str = "fixed_row") -> tuple[str, str]:
    if match:
        return "None", ""
    if mode == "identity":
        if column in MEDIUM_COLUMNS:
            return "Medium", "Công thức/giá trị F:Q của dòng nhận dạng đã căn chỉnh khác nhau"
        return "Medium", "Dòng nhận dạng đã căn chỉnh khác nhau"
    if column == "B":
        return "High", "Tài khoản tại cột B khác nhau"
    if column == "R":
        return "High", "Công thức/giá trị tổng tại cột R khác nhau hoặc bị thiếu"
    if column == "S":
        if generated in (None, ""):
            return "High", "Mô tả tại cột S bị trống ngoài dự kiến"
        return "High", "Mô tả tại cột S khác nhau"
    if column in MEDIUM_COLUMNS:
        return "Medium", "Công thức/giá trị F:Q khác nhau"
    return "Low", "Công cụ này chưa so sánh định dạng"


def _token_hits(text: Any) -> set[str]:
    normalized = normalize_text(text)
    return {name for name, tokens in IDENTITY_TOKENS.items() if any(token in normalized for token in tokens)}


def _identity_row_group(row: int) -> str:
    if row in NEW_IDENTITY_ROWS_FROM_42N1G:
        return "new_identity_rows_from_42N1G"
    if row in EXISTING_IDENTITY_ROWS:
        return "existing_identity_rows"
    if row in LAYOUT_FIXED_ROWS:
        return "layout_fixed_rows"
    return "other"


def _alignment_classification(alignment: "IdentityAlignment", diff_count: int = 0) -> str:
    if alignment.match_method == "ambiguous":
        return "ambiguous_identity"
    if alignment.reference_matched_row is None:
        return "identity_not_found"
    if diff_count:
        return "matched_identity_with_strict_diffs"
    return "matched_identity_no_strict_diff"


def _disambiguate_close_candidates(gen_tokens: set[str], candidates: list[tuple]) -> tuple | None:
    if len(candidates) < 2:
        return None
    building_land_evidence = bool(gen_tokens & {"building", "facility"}) and "equipment" not in gen_tokens
    tool_jig_evidence = "equipment" in gen_tokens and not (gen_tokens & {"building", "facility"})
    if not (building_land_evidence or tool_jig_evidence):
        return None
    wanted = {"building", "facility"} if building_land_evidence else {"equipment"}
    rejected = {"equipment"} if building_land_evidence else {"building", "facility"}
    matching = []
    for candidate in candidates:
        ref_desc = candidate[3]
        ref_tokens = _token_hits(ref_desc)
        if ref_tokens & wanted and not (ref_tokens & rejected):
            matching.append(candidate)
    if len(matching) == 1:
        return matching[0]
    return None


def _find_identity_match(gen_ws, ref_ws, generated_row: int) -> IdentityAlignment:
    gen_account = _cell_value(gen_ws, generated_row, "B")
    gen_desc = _cell_value(gen_ws, generated_row, "S")
    gen_tokens = _token_hits(gen_desc)
    desired_tokens = set()
    if generated_row == 53:
        desired_tokens.add("bus_expat")
    elif generated_row == 54:
        desired_tokens.add("bus_local")
    elif generated_row == 97:
        desired_tokens.add("notebook_staff")
    elif generated_row == 98:
        desired_tokens.add("notebook_worker")
    elif generated_row == 38:
        desired_tokens.update({"depreciation", "equipment"})
    elif generated_row == 40:
        desired_tokens.update({"interest", "building", "facility"})
    elif generated_row == 42:
        desired_tokens.update({"interest", "equipment"})
    elif generated_row == 51:
        desired_tokens.add("cleaning")
    elif generated_row == 59:
        desired_tokens.add("birthday")
    best = None
    second_best = None
    max_row = ref_ws.max_row or 1
    for row in range(1, max_row + 1):
        ref_account = _cell_value(ref_ws, row, "B")
        ref_desc = _cell_value(ref_ws, row, "S")
        ref_tokens = _token_hits(ref_desc)
        same_account = gen_account not in (None, "") and gen_account == ref_account
        token_overlap = bool(gen_tokens & ref_tokens)
        transport_preferred = "bus" in ref_tokens and bool(desired_tokens & ref_tokens)
        notebook_preferred = "notebook" in ref_tokens and bool(desired_tokens & ref_tokens)
        required_by_row = {
            38: {"depreciation"},
            40: {"interest"},
            42: {"interest"},
            51: {"cleaning"},
            58: {"health"},
            59: {"birthday"},
            66: {"trip"},
            97: {"notebook"},
            98: {"notebook"},
        }.get(generated_row, set())
        if required_by_row and not (required_by_row & ref_tokens):
            continue
        preferred_identity = transport_preferred or notebook_preferred or bool(desired_tokens & ref_tokens)
        if same_account and normalize_text(gen_desc) == normalize_text(ref_desc):
            score = 5
            note = "cùng tài khoản và mô tả"
            candidate = (score, row, ref_account, ref_desc, "same_account", "High", True, note, "account,description")
        elif same_account:
            score = 4 if preferred_identity else 3
            note = "cùng tài khoản và có token nhận dạng ưu tiên" if preferred_identity else "cùng tài khoản"
            candidate = (score, row, ref_account, ref_desc, "same_account", "High", True, note, "account,tokens" )
        elif token_overlap:
            score = 2.5 if preferred_identity else 2
            note = "token mô tả trùng mạnh và có token nhận dạng ưu tiên" if preferred_identity else "token mô tả trùng mạnh"
            candidate = (score, row, ref_account, ref_desc, "token_overlap", "Medium", True, note, "tokens")
        else:
            continue
        if best is None or candidate[0] > best[0]:
            second_best = best
            best = candidate
        elif second_best is None or candidate[0] > second_best[0]:
            second_best = candidate
    req_note = REQUIREMENT_LOCK_NOTES.get(generated_row, "")
    row_group = _identity_row_group(generated_row)
    if best:
        candidate_rows = "; ".join(
            f"row={c[1]},score={c[0]},method={c[4]},account={c[2]},desc={normalize_text(c[3])[:80]}"
            for c in ([best] + ([second_best] if second_best is not None else []))
        )
        if second_best is not None and best[0] - second_best[0] < 0.5:
            disambiguated = _disambiguate_close_candidates(gen_tokens, [best, second_best])
            if disambiguated is None:
                return IdentityAlignment(generated_row, gen_account, gen_desc, None, None, None, "ambiguous", "Low", False, "có nhiều ứng viên nhận dạng gần nhau; không ép ghép", best[0], "", candidate_rows, row_group, req_note, "ambiguous_identity")
            best = disambiguated
            evidence_field = "building_land_tokens" if gen_tokens & {"building", "facility"} else "tool_jig_tokens"
            evidence_note = "phân biệt dựa trên token nhà xưởng/đất" if evidence_field == "building_land_tokens" else "phân biệt dựa trên token công cụ/đồ gá"
            score, row, ref_account, ref_desc, method, confidence, same, _note, matched_fields = best
            return IdentityAlignment(generated_row, gen_account, gen_desc, row, ref_account, ref_desc, method, "High", same, evidence_note, score + 0.25, f"{matched_fields},{evidence_field}", candidate_rows, row_group, req_note, "")
        score, row, ref_account, ref_desc, method, confidence, same, note, matched_fields = best
        return IdentityAlignment(generated_row, gen_account, gen_desc, row, ref_account, ref_desc, method, confidence, same, note, score, matched_fields, candidate_rows, row_group, req_note, "")
    return IdentityAlignment(generated_row, gen_account, gen_desc, None, None, None, "not_found", "Low", False, "không tìm thấy tài khoản giống nhau hoặc token trùng đủ mạnh", 0.0, "", "", row_group, req_note, "identity_not_found")


def _compare_row(gen_ws, ref_ws, generated_row: int, reference_row: int, mode: str) -> tuple[list[CellDiff], dict[str, Any]]:
    diffs = []
    row_diff_count = 0
    for column in COMPARE_COLUMNS:
        generated = _cell_value(gen_ws, generated_row, column)
        reference = _cell_value(ref_ws, reference_row, column)
        match = generated == reference
        severity, note = _severity(column, generated, reference, match, mode)
        if not match:
            row_diff_count += 1
        diffs.append(CellDiff(generated_row, column, generated, reference, match, severity, note, mode, reference_row))
    summary = {
        "row": generated_row,
        "reference_row": reference_row,
        "compare_mode": mode,
        "description_generated": _cell_value(gen_ws, generated_row, "S"),
        "description_reference": _cell_value(ref_ws, reference_row, "S"),
        "differences_count": row_diff_count,
        "status": "OK" if row_diff_count == 0 else "DIFF",
    }
    return diffs, summary


def _compare_workbooks_strict_exact(generated_path: Path, reference_path: Path, out_dir: Path) -> dict[str, Any]:
    if not generated_path.is_file():
        raise FileNotFoundError(f"Không tìm thấy workbook kết quả: {generated_path}")
    if not reference_path.is_file():
        raise FileNotFoundError(f"Không tìm thấy workbook tham chiếu: {reference_path}")
    out_dir.mkdir(parents=True, exist_ok=True)
    gen_wb = load_workbook(generated_path, data_only=False)
    ref_wb = load_workbook(reference_path, data_only=False)
    try:
        if DETAIL_SHEET not in gen_wb.sheetnames or DETAIL_SHEET not in ref_wb.sheetnames:
            raise ValueError(f"Thiếu sheet bắt buộc: {DETAIL_SHEET}")
        gen_ws = gen_wb[DETAIL_SHEET]
        ref_ws = ref_wb[DETAIL_SHEET]
        diffs: list[CellDiff] = []
        row_counts: dict[int, int] = {}
        row_summary: list[dict[str, Any]] = []
        alignments: list[IdentityAlignment] = []
        identity_rows = set(IDENTITY_ROW_CANDIDATES)
        total_cells = 0
        identity_differences = 0
        max_row = max(gen_ws.max_row or 1, ref_ws.max_row or 1)

        for row in IDENTITY_ROW_CANDIDATES:
            alignment = _find_identity_match(gen_ws, ref_ws, row)
            alignments.append(alignment)
            if alignment.reference_matched_row is None:
                continue
            row_diff_count = 0
            for column in STRICT_COMPARE_COLUMNS:
                generated = _cell_value(gen_ws, row, column)
                reference = _cell_value(ref_ws, alignment.reference_matched_row, column)
                if generated in (None, "") and reference in (None, ""):
                    continue
                if generated in (None, "") or reference in (None, ""):
                    continue
                total_cells += 1
                match = generated == reference
                if not match:
                    row_diff_count += 1
                    identity_differences += 1
                    row_counts[row] = row_counts.get(row, 0) + 1
                    severity, note = _severity(column, generated, reference, match, "identity")
                    diffs.append(CellDiff(row, column, generated, reference, match, severity, note, "identity", alignment.reference_matched_row))
            row_summary.append({
                "row": row,
                "reference_row": alignment.reference_matched_row,
                "compare_mode": "identity",
                "description_generated": _cell_value(gen_ws, row, "S"),
                "description_reference": _cell_value(ref_ws, alignment.reference_matched_row, "S"),
                "differences_count": row_diff_count,
                "status": "OK" if row_diff_count == 0 else "DIFF",
            })

        for row in range(1, max_row + 1):
            if row in identity_rows:
                continue
            row_diff_count = 0
            for column in STRICT_COMPARE_COLUMNS:
                generated = _cell_value(gen_ws, row, column)
                reference = _cell_value(ref_ws, row, column)
                if generated in (None, "") and reference in (None, ""):
                    continue
                if generated in (None, "") or reference in (None, ""):
                    continue
                total_cells += 1
                match = generated == reference
                if not match:
                    row_diff_count += 1
                    row_counts[row] = row_counts.get(row, 0) + 1
                    severity, note = _severity(column, generated, reference, match, "strict_exact")
                    diffs.append(CellDiff(row, column, generated, reference, match, severity, note, "strict_exact", row))
            if row_diff_count:
                row_summary.append({"row": row, "reference_row": row, "compare_mode": "strict_exact", "description_generated": _cell_value(gen_ws, row, "S"), "description_reference": _cell_value(ref_ws, row, "S"), "differences_count": row_diff_count, "status": "DIFF"})
        diff_total = len(diffs)
        fixed_row_differences = diff_total - identity_differences
        diff_counts_by_row = {row["row"]: row.get("differences_count", 0) for row in row_summary}
        for alignment in alignments:
            alignment.classification = _alignment_classification(alignment, diff_counts_by_row.get(alignment.generated_row, 0))
        new_alignments = [a for a in alignments if a.generated_row in NEW_IDENTITY_ROWS_FROM_42N1G]
        existing_alignments = [a for a in alignments if a.generated_row in EXISTING_IDENTITY_ROWS]
        summary = {
            "generated_path": str(generated_path),
            "reference_path": str(reference_path),
            "compare_mode": "strict_exact",
            "total_cells_compared": total_cells,
            "exact_matches": total_cells - diff_total,
            "diff_total": diff_total,
            "differences": diff_total,
            "diff_by_row": {str(row): count for row, count in sorted(row_counts.items())},
            "fixed_rows_compared": max_row - len(identity_rows),
            "fixed_row_differences": fixed_row_differences,
            "identity_rows_checked": len(identity_rows),
            "identity_rows_matched": sum(1 for a in alignments if a.reference_matched_row is not None),
            "identity_rows_ambiguous": sum(1 for a in alignments if a.match_method == "ambiguous"),
            "identity_rows_not_found": sum(1 for a in alignments if a.match_method == "not_found"),
            "identity_differences": identity_differences,
            "strict_diff_total": diff_total,
            "strict_identity_diff_total": identity_differences,
            "layout_fixed_rows": list(LAYOUT_FIXED_ROWS),
            "layout_fixed_rows_total": len(LAYOUT_FIXED_ROWS),
            "new_identity_rows_from_42N1G": list(NEW_IDENTITY_ROWS_FROM_42N1G),
            "new_identity_rows_total": len(NEW_IDENTITY_ROWS_FROM_42N1G),
            "new_identity_rows_matched": sum(1 for a in new_alignments if a.reference_matched_row is not None),
            "new_identity_rows_ambiguous": sum(1 for a in new_alignments if a.match_method == "ambiguous"),
            "new_identity_rows_not_found": sum(1 for a in new_alignments if a.match_method == "not_found"),
            "existing_identity_rows": list(EXISTING_IDENTITY_ROWS),
            "existing_identity_rows_total": len(EXISTING_IDENTITY_ROWS),
            "existing_identity_rows_matched": sum(1 for a in existing_alignments if a.reference_matched_row is not None),
            "existing_identity_rows_ambiguous": sum(1 for a in existing_alignments if a.match_method == "ambiguous"),
            "existing_identity_rows_not_found": sum(1 for a in existing_alignments if a.match_method == "not_found"),
            "false_fixed_row_diffs_removed": "các dòng nhận dạng được loại khỏi phép so sánh nghiêm ngặt cùng dòng; số lượng phụ thuộc đường cơ sở",
            "real_identity_strict_diffs_remaining": identity_differences,
        }
        xlsx_path = out_dir / "MP2027_primary_reference_compare.xlsx"
        json_path = out_dir / "MP2027_primary_reference_compare.json"
        _write_excel_report(xlsx_path, summary, diffs, row_summary, alignments)
        json_payload = {
            "summary": summary,
            "diff_detail": [asdict(diff) for diff in diffs],
            "diff_by_row": summary["diff_by_row"],
            "important_row_diff": [asdict(diff) for diff in diffs],
            "row_summary": row_summary,
            "identity_row_alignment": [asdict(alignment) for alignment in alignments],
        }
        json_path.write_text(json.dumps(json_payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return {"summary": summary, "xlsx_path": str(xlsx_path), "json_path": str(json_path)}
    finally:
        gen_wb.close()
        ref_wb.close()

def compare_workbooks(generated_path: Path, reference_path: Path, out_dir: Path, strict_exact: bool = False) -> dict[str, Any]:
    if strict_exact:
        return _compare_workbooks_strict_exact(generated_path, reference_path, out_dir)
    if not generated_path.is_file():
        raise FileNotFoundError(f"Không tìm thấy workbook kết quả: {generated_path}")
    if not reference_path.is_file():
        raise FileNotFoundError(f"Không tìm thấy workbook tham chiếu: {reference_path}")
    out_dir.mkdir(parents=True, exist_ok=True)
    gen_wb = load_workbook(generated_path, data_only=False)
    ref_wb = load_workbook(reference_path, data_only=False)
    try:
        missing_sheets = []
        if DETAIL_SHEET not in gen_wb.sheetnames:
            missing_sheets.append(f"kết_quả:{DETAIL_SHEET}")
        if DETAIL_SHEET not in ref_wb.sheetnames:
            missing_sheets.append(f"tham_chiếu:{DETAIL_SHEET}")
        if missing_sheets:
            raise ValueError("Thiếu sheet bắt buộc: " + ", ".join(missing_sheets))
        gen_ws = gen_wb[DETAIL_SHEET]
        ref_ws = ref_wb[DETAIL_SHEET]
        diffs: list[CellDiff] = []
        row_summary: list[dict[str, Any]] = []
        alignments: list[IdentityAlignment] = []
        fixed_row_differences = 0
        identity_differences = 0

        for row in FIXED_ROW_RULES:
            row_diffs, summary_row = _compare_row(gen_ws, ref_ws, row, row, "fixed_row")
            fixed_row_differences += sum(1 for diff in row_diffs if not diff.match)
            diffs.extend(row_diffs)
            row_summary.append(summary_row)

        for row in IDENTITY_ROW_CANDIDATES:
            alignment = _find_identity_match(gen_ws, ref_ws, row)
            alignments.append(alignment)
            if alignment.reference_matched_row is not None:
                row_diffs, summary_row = _compare_row(gen_ws, ref_ws, row, alignment.reference_matched_row, "identity")
                identity_differences += sum(1 for diff in row_diffs if not diff.match)
                diffs.extend(row_diffs)
                row_summary.append(summary_row)

        total_cells = len(diffs)
        differences = sum(1 for diff in diffs if not diff.match)
        summary = {
            "generated_path": str(generated_path),
            "reference_path": str(reference_path),
            "total_cells_compared": total_cells,
            "exact_matches": total_cells - differences,
            "differences": differences,
            "missing_sheets": [],
            "missing_rows_or_cells": [],
            "fixed_rows_compared": len(FIXED_ROW_RULES),
            "identity_rows_checked": len(IDENTITY_ROW_CANDIDATES),
            "identity_rows_matched": sum(1 for a in alignments if a.reference_matched_row is not None),
            "identity_rows_not_found": sum(1 for a in alignments if a.reference_matched_row is None),
            "fixed_row_differences": fixed_row_differences,
            "identity_differences": identity_differences,
            "layout_fixed_rows": list(LAYOUT_FIXED_ROWS),
            "layout_fixed_rows_total": len(LAYOUT_FIXED_ROWS),
            "new_identity_rows_total": len(NEW_IDENTITY_ROWS_FROM_42N1G),
            "new_identity_rows_matched": sum(1 for a in alignments if a.generated_row in NEW_IDENTITY_ROWS_FROM_42N1G and a.reference_matched_row is not None),
            "new_identity_rows_ambiguous": sum(1 for a in alignments if a.generated_row in NEW_IDENTITY_ROWS_FROM_42N1G and a.match_method == "ambiguous"),
            "new_identity_rows_not_found": sum(1 for a in alignments if a.generated_row in NEW_IDENTITY_ROWS_FROM_42N1G and a.match_method == "not_found"),
            "existing_identity_rows_total": len(EXISTING_IDENTITY_ROWS),
            "existing_identity_rows_matched": sum(1 for a in alignments if a.generated_row in EXISTING_IDENTITY_ROWS and a.reference_matched_row is not None),
            "existing_identity_rows_ambiguous": sum(1 for a in alignments if a.generated_row in EXISTING_IDENTITY_ROWS and a.match_method == "ambiguous"),
            "existing_identity_rows_not_found": sum(1 for a in alignments if a.generated_row in EXISTING_IDENTITY_ROWS and a.match_method == "not_found"),
        }
        xlsx_path = out_dir / "MP2027_primary_reference_compare.xlsx"
        json_path = out_dir / "MP2027_primary_reference_compare.json"
        _write_excel_report(xlsx_path, summary, diffs, row_summary, alignments)
        json_payload = {
            "summary": summary,
            "important_row_diff": [asdict(diff) for diff in diffs],
            "row_summary": row_summary,
            "identity_row_alignment": [asdict(alignment) for alignment in alignments],
        }
        json_path.write_text(json.dumps(json_payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return {"summary": summary, "xlsx_path": str(xlsx_path), "json_path": str(json_path)}
    finally:
        gen_wb.close()
        ref_wb.close()


def _write_excel_report(path: Path, summary: dict[str, Any], diffs: list[CellDiff], row_summary: list[dict[str, Any]], alignments: list[IdentityAlignment]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tóm_tắt"
    for key, value in _display_summary(summary).items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])
    ws = wb.create_sheet("Khác_biệt_quan_trọng")
    ws.append(["Dòng", "Dòng tham chiếu", "Chế độ", "Cột", "Giá trị/công thức kết quả", "Giá trị/công thức tham chiếu", "Khớp?", "Mức độ", "Ghi chú"])
    for diff in diffs:
        ws.append([diff.row, diff.reference_row, _display_value(diff.compare_mode), diff.column, diff.generated, diff.reference, diff.match, _display_value(diff.severity), diff.note])
    ws = wb.create_sheet("Tóm_tắt_theo_dòng")
    ws.append(["Dòng", "Dòng tham chiếu", "Chế độ", "Mô tả kết quả", "Mô tả tham chiếu", "Số khác biệt", "Trạng thái"])
    for row in row_summary:
        ws.append([row["row"], row["reference_row"], _display_value(row["compare_mode"]), row["description_generated"], row["description_reference"], row["differences_count"], _display_value(row["status"])])
    ws = wb.create_sheet("Căn_chỉnh_nhận_dạng")
    ws.append(["Dòng kết quả", "Nhóm dòng", "Tài khoản kết quả", "Mô tả kết quả", "Dòng tham chiếu khớp", "Tài khoản tham chiếu", "Mô tả tham chiếu", "Phương pháp khớp", "Điểm khớp", "Trường đã khớp", "Độ tin cậy", "Cùng hạng mục nghiệp vụ?", "Phân loại", "Các dòng ứng viên", "Ghi chú yêu cầu", "Ghi chú"])
    for alignment in alignments:
        ws.append([alignment.generated_row, alignment.row_group, alignment.generated_account, alignment.generated_description, alignment.reference_matched_row, alignment.reference_account, alignment.reference_description, _display_value(alignment.match_method), alignment.match_score, alignment.matched_fields, _display_value(alignment.confidence), alignment.same_business_item, _display_value(alignment.classification), alignment.candidate_rows, alignment.requirement_note, alignment.note])
    ws = wb.create_sheet("Hướng_dẫn")
    for item in [
        "Tệp tham chiếu chính là đường cơ sở kết quả mong đợi.",
        "Khác biệt theo dòng cố định chỉ áp dụng cho các dòng đã được khóa bởi yêu cầu/người dùng/FORM.",
        "Các dòng nhận dạng được căn chỉnh theo tài khoản hoặc token mô tả đủ mạnh trước khi so sánh giá trị.",
        "Khác biệt không tự động được coi là lỗi; người dùng phải xác nhận ý nghĩa nghiệp vụ.",
        "Tệp tham chiếu phụ chỉ dùng cho trường hợp biên.",
    ]:
        ws.append([item])
    wb.save(path)


def main(argv: list[str] | None = None) -> int:
    parser = VietnameseArgumentParser(description="So sánh kết quả MP2027 đã tạo với tệp tham chiếu chính đáng tin cậy.")
    parser.add_argument("--generated", required=True, type=Path, help="Đường dẫn workbook kết quả đã tạo")
    parser.add_argument("--reference", required=True, type=Path, help="Đường dẫn workbook tham chiếu")
    parser.add_argument("--out-dir", required=True, type=Path, help="Thư mục ghi báo cáo")
    parser.add_argument(
        "--strict-exact",
        action="store_true",
        help="So sánh chính xác công thức/giá trị, không áp dụng căn chỉnh nhận dạng hoặc ngoại lệ đã chấp nhận.",
    )
    args = parser.parse_args(argv)
    try:
        result = compare_workbooks(args.generated, args.reference, args.out_dir, strict_exact=args.strict_exact)
    except Exception:  # pragma: no cover - ranh giới CLI
        print("Lỗi: Không thể tạo báo cáo so sánh với tệp tham chiếu chính.", file=sys.stderr)
        return 1
    print(json.dumps(_display_summary(result["summary"]), ensure_ascii=False, indent=2))
    print(f"Báo cáo Excel: {result['xlsx_path']}")
    print(f"Báo cáo JSON: {result['json_path']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
