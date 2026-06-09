from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src.engine.hub_builder import IT_COMPONENT_ORDER, IT_SYSTEM_ACCOUNT_CODES
from src.utils.excel_helpers import find_hub_sheet_name, get_fy_months

DEFAULT_DB = ROOT / "mp2027.db"
DEFAULT_OUTPUT_DIR = ROOT / "OUTPUT_FY2027"
MONTH_START_COL = 6  # F
DESCRIPTION_COL = 19  # S
TOTAL_COL = 18  # R


def _format_number(value: float) -> str:
    if abs(value - int(value)) < 1e-9:
        return str(int(value))
    return (f"{value:.10f}").rstrip("0").rstrip(".")


def _parse_component_term(description: str) -> tuple[str, float, float] | None:
    parts = str(description or "").split("|")
    if len(parts) < 5 or parts[0:2] != ["it_sim", "component_term"]:
        return None
    component_key = parts[2]
    qty = 0.0
    unit_usd = 0.0
    for part in parts[3:]:
        if part.startswith("qty="):
            qty = float(part.split("=", 1)[1] or 0)
        elif part.startswith("unit_usd="):
            unit_usd = float(part.split("=", 1)[1] or 0)
    if qty <= 0 or unit_usd <= 0:
        return None
    return component_key, qty, unit_usd


def _load_expected_from_db(db_path: Path, cc_code: str, fiscal_year: int) -> dict[str, str]:
    fy_months = get_fy_months(fiscal_year)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT period, amount_vnd, amount_usd, description
            FROM fact_input_data
            WHERE source = 'it_sim' AND CAST(cc_code AS TEXT) = ?
            """,
            (str(cc_code),),
        ).fetchall()
    finally:
        conn.close()

    terms_by_period: dict[str, dict[str, list[tuple[float, float]]]] = defaultdict(lambda: defaultdict(list))
    values_by_period: dict[str, dict[str, float]] = defaultdict(dict)
    totals_by_period: dict[str, float] = {}

    for row in rows:
        period = str(row["period"])
        description = str(row["description"] or "")
        parsed = _parse_component_term(description)
        if parsed:
            component, qty, unit_usd = parsed
            terms_by_period[period][component].append((qty, unit_usd))
            continue
        if description.startswith("it_sim|component|"):
            parts = description.split("|")
            if len(parts) >= 3:
                values_by_period[period][parts[2]] = float(row["amount_usd"] or 0)
            continue
        if description.startswith("it_sim|system_usage_total"):
            totals_by_period[period] = float(row["amount_vnd"] or 0)

    expected: dict[str, str] = {}
    for period in fy_months:
        ordered_terms: list[str] = []
        for key in IT_COMPONENT_ORDER:
            for qty, unit_usd in terms_by_period.get(period, {}).get(key, []):
                ordered_terms.append(f"{_format_number(qty)}*{_format_number(unit_usd)}")
        if ordered_terms:
            expected[period] = f"=ROUND(({'+'.join(ordered_terms)})*$B$2,0)"
            continue

        ordered_values = [
            values_by_period.get(period, {}).get(key, 0.0)
            for key in IT_COMPONENT_ORDER
            if float(values_by_period.get(period, {}).get(key, 0.0)) > 0
        ]
        if ordered_values:
            expected[period] = f"=ROUND(({'+'.join(_format_number(v) for v in ordered_values)})*$B$2,0)"
            continue

        total = totals_by_period.get(period, 0.0)
        expected[period] = str(int(round(total))) if total else ""
    return expected


def _find_system_row(ws) -> int | None:
    for row in range(1, ws.max_row + 1):
        account = ws.cell(row=row, column=2).value
        desc = str(ws.cell(row=row, column=DESCRIPTION_COL).value or "")
        if account in IT_SYSTEM_ACCOUNT_CODES or "System Cost" in desc or "KDC" in desc:
            # Prefer rows that actually have a monthly system formula/value.
            monthly_values = [ws.cell(row=row, column=col).value for col in range(MONTH_START_COL, MONTH_START_COL + 12)]
            if any(value not in (None, "") for value in monthly_values):
                return row
    return None


def _formula_terms(formula: str) -> list[str]:
    text = str(formula or "").replace(" ", "")
    match = re.match(r"^=ROUND\(\((.*)\)\*\$B\$2,0\)$", text)
    if not match:
        return []
    return [part for part in match.group(1).split("+") if part]


def verify(cc_code: str, fiscal_year: int, db_path: Path, output_dir: Path) -> int:
    output_path = output_dir / f"MP_CC_{cc_code}.xlsx"
    print(f"CC: {cc_code}")
    print(f"DB: {db_path}")
    print(f"Output: {output_path}")

    if not db_path.exists():
        print("NG: Không tìm thấy database mp2027.db")
        return 2
    if not output_path.exists():
        print("NG: Không tìm thấy file output. Hãy chạy chương trình xuất CC này trước.")
        return 2

    expected = _load_expected_from_db(db_path, cc_code, fiscal_year)
    wb = load_workbook(output_path, data_only=False)
    try:
        ws = wb[find_hub_sheet_name(wb)]
        row = _find_system_row(ws)
        if row is None:
            print("NG: Không tìm thấy dòng Chi phí hệ thống trong file output.")
            return 1

        print(f"Dòng System Cost trong output: {row}")
        ok = True
        fy_months = get_fy_months(fiscal_year)
        for offset, period in enumerate(fy_months):
            col = MONTH_START_COL + offset
            actual = ws.cell(row=row, column=col).value
            expected_formula = expected.get(period, "")
            if expected_formula == "":
                expected_display = "<trống>"
                match = actual in (None, "")
            else:
                expected_display = expected_formula
                match = str(actual or "") == expected_formula

            status = "OK" if match else "NG"
            if not match:
                ok = False
            print(f"{status} {period}: output={actual!r} | expected={expected_display}")

            if actual and str(actual).startswith("=ROUND"):
                terms = _formula_terms(str(actual))
                print(f"    terms trong công thức: {len(terms)} -> {', '.join(terms[:8])}{' ...' if len(terms) > 8 else ''}")

        total_formula = ws.cell(row=row, column=TOTAL_COL).value
        expected_total = f"=SUM(F{row}:Q{row})"
        if total_formula == expected_total:
            print(f"OK Total: {total_formula}")
        else:
            ok = False
            print(f"NG Total: output={total_formula!r} | expected={expected_total}")

        print("\nKẾT LUẬN:")
        if ok:
            print("OK: Chi phí hệ thống trong output khớp dữ liệu đã import và dùng đúng công thức/value mong đợi.")
            return 0
        print("NG: Có tháng chưa khớp. Xem các dòng NG phía trên.")
        return 1
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Kiểm chứng Chi phí hệ thống trong file MP output theo CC.")
    parser.add_argument("cc_code", help="Mã cost center, ví dụ 1412000006")
    parser.add_argument("--fy", type=int, default=2027, help="Fiscal year, mặc định 2027")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB, help="Đường dẫn mp2027.db")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Thư mục OUTPUT_FY2027")
    return verify(str(parser.parse_args().cc_code), parser.parse_args().fy, parser.parse_args().db, parser.parse_args().output_dir)


if __name__ == "__main__":
    raise SystemExit(main())
