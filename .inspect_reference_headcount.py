from pathlib import Path
import openpyxl
base=Path('reference_outputs/secondary/FY2027')
for prefix in ('14.','15.','16.','72.','73.'):
 p=next(base.glob(prefix+'*.xlsx')); wbv=openpyxl.load_workbook(p,data_only=False,read_only=True); wbd=openpyxl.load_workbook(p,data_only=True,read_only=True)
 try:
  sn=next((x for x in wbv.sheetnames if '4' in x and '3' in x),wbv.sheetnames[0]); wsv=wbv[sn]; wsd=wbd[sn]
  print('\n###',p.name,'sheet',sn,'B5',wsv['B5'].value)
  for r in range(1,min(wsv.max_row,220)+1):
   b=wsv.cell(r,2).value; s=wsv.cell(r,19).value; e=wsv.cell(r,5).value
   texts=' '.join(str(x or '') for x in (b,e,s)).lower()
   if any(t in texts for t in ('headcount','人員','nhân','staff','worker','定年','健康診断','system cost')):
    months=[wsd.cell(r,c).value for c in range(6,18)]
    print(r,'B=',b,'E=',e,'S=',s,'F:Q=',months)
 finally: wbv.close(); wbd.close()
