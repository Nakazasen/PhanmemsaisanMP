from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.services.fiscal_run import RunPreflightReport

from scripts.run_e2e import run_universal_pipeline
from src.engine.mp_saisan_complete_export import (
    PROVENANCE_PRIORITY,
    _append_secondary_skeleton_deduped,
    canonical_description,
    collect_existing_identities,
    count_provenance_rows,
    resolve_reference_path,
    row_identity_key,
)

SHEET = "内訳ﾘｽﾄ(4～3月)"


def _wb(path: Path, rows=None):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for idx, row in enumerate(rows or [], start=1):
        ws.cell(idx, 2).value = row.get("account")
        ws.cell(idx, 19).value = row.get("description")
        ws.cell(idx, 20).value = row.get("note")
        ws.cell(idx, 6).value = row.get("f")
        ws.cell(idx, 17).value = row.get("q")
    wb.save(path)
    wb.close()
    return path


def _csv(path: Path, descriptions):
    path.write_text(
        "classification,account,description,pattern_signature,month_F_sample,month_Q_sample\n"
        + "".join(
            f"REFERENCE_ASSISTED_FILL_CANDIDATE,5005026371,{desc},sig{i},1,\n"
            for i, desc in enumerate(descriptions)
        ),
        encoding="utf-8",
    )
    return path


def test_priority_source_derived_wins_over_reference():
    assert PROVENANCE_PRIORITY["SOURCE_DERIVED"] < PROVENANCE_PRIORITY["REFERENCE_FILLED_FROM_PRIMARY"]


def test_primary_reference_wins_over_secondary_skeleton():
    assert PROVENANCE_PRIORITY["REFERENCE_FILLED_FROM_PRIMARY"] < PROVENANCE_PRIORITY["REFERENCE_ASSISTED_SECONDARY_SKELETON"]


def test_secondary_skeleton_writes_only_without_duplicate_identity(tmp_path):
    target = _wb(tmp_path / "target.xlsx", [{"account": "5005026371", "description": "A", "note": "REFERENCE_FILLED_FROM_PRIMARY"}])
    csvp = _csv(tmp_path / "candidates.csv", ["A", "B"])
    result = _append_secondary_skeleton_deduped(target, csvp, 1412000040)
    assert result["written"] == 1
    assert result["skipped_duplicate"] == 1
    counts = count_provenance_rows(target)
    assert counts["fixed_assets_skeleton_rows"] == 1


def test_electricity_water_aliases_do_not_duplicate():
    assert canonical_description("電気代") == canonical_description("electricity")
    assert canonical_description("水道代") == canonical_description("water")
    assert canonical_description("Điện") == canonical_description("electricity")
    assert canonical_description("Nước") == canonical_description("water")


def test_complete_mode_requires_valid_source_inputs(tmp_path):
    ok, message = run_universal_pipeline(2027, str(_wb(tmp_path / "target.xlsx")), str(tmp_path), target_cc=1412000040)
    assert ok is False
    assert "nhân sự" in message and "FY2027" in message


def test_complete_mode_uses_reference_map_for_1412000040(tmp_path):
    ref = _wb(tmp_path / "ref.xlsx")
    m = tmp_path / "map.csv"
    m.write_text("target_cc,reference_role,reference_path\n1412000040,primary_reference,ref.xlsx\n", encoding="utf-8")
    assert resolve_reference_path(1412000040, reference_map_path=m).name == ref.name


def test_other_cc_without_reference_fails_clearly(tmp_path):
    with pytest.raises(ValueError, match="cần đường dẫn hoặc ánh xạ đến tệp tham khảo chính"):
        resolve_reference_path(999, reference_map_path=tmp_path / "missing.csv")


def test_provenance_counts_are_returned(tmp_path):
    target = _wb(
        tmp_path / "target.xlsx",
        [
            {"account": "1", "description": "source", "note": ""},
            {"account": "2", "description": "primary", "note": "REFERENCE_FILLED_FROM_PRIMARY"},
            {"account": "3", "description": "secondary", "note": "REFERENCE_ASSISTED_SECONDARY_SKELETON"},
        ],
    )
    assert count_provenance_rows(target) == {"source_rows": 1, "primary_reference_rows": 1, "fixed_assets_skeleton_rows": 1}


def test_no_overwrite_rows_200_212(tmp_path):
    target = _wb(tmp_path / "target.xlsx")
    wb = load_workbook(target)
    ws = wb[SHEET]
    ws.cell(200, 2).value = "KEEP"
    ws.cell(200, 19).value = "source-layer"
    wb.save(target)
    wb.close()
    csvp = _csv(tmp_path / "candidates.csv", ["new"])
    result = _append_secondary_skeleton_deduped(target, csvp, 1412000040)
    assert result["written"] == 1
    wb = load_workbook(target)
    ws = wb[SHEET]
    assert ws.cell(200, 2).value == "KEEP"
    assert ws.cell(213, 2).value == "5005026371"
    wb.close()


def test_complete_v1_missing_reference_map_orchestration_contract(monkeypatch, tmp_path):
    """Fast mocked orchestration coverage; this is not real pipeline acceptance."""
    import scripts.run_e2e as run_e2e

    export_calls = []
    complete_reference_calls = []

    class Cursor:
        def execute(self, *args):
            return None

        def fetchone(self):
            return (0,)

        def fetchall(self):
            return []

    class Conn:
        def cursor(self):
            return Cursor()

        def execute(self, *args, **kwargs):
            return Cursor()

        def executemany(self, *args, **kwargs):
            return Cursor()

        def commit(self):
            return None

        def rollback(self):
            return None

        def close(self):
            return None

    class Builder:
        def __init__(self, *args, **kwargs):
            pass

        def export_to_template(self, *args, **kwargs):
            output_path = args[1]
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            Path(output_path).write_text("placeholder", encoding="utf-8")
            export_calls.append((args, kwargs))
            return True

    class Engine:
        def __init__(self, *args, **kwargs):
            pass

        def run_allocation(self):
            return None

    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(run_e2e, "get_connection", lambda path: Conn())
    monkeypatch.setattr(
        run_e2e,
        "preflight_fiscal_run",
        lambda context: RunPreflightReport(
            fiscal_year=2027,
            resolved_sources={},
        ),
    )
    monkeypatch.setattr(
        run_e2e,
        "import_headcount_time_sources",
        lambda *args, **kwargs: {
            "files": 1,
            "imported_files": 1,
            "skipped": [],
            "errors": [],
            "results": [],
        },
    )
    monkeypatch.setattr(run_e2e, "_staffing_preflight", lambda *args, **kwargs: [])
    monkeypatch.setattr(run_e2e, "create_schema", lambda conn: None)
    monkeypatch.setattr(run_e2e, "init_sys_params", lambda conn, exchange_rate, fiscal_year, **kwargs: None)
    monkeypatch.setattr(run_e2e, "load_all", lambda **kwargs: None)
    monkeypatch.setattr(run_e2e, "describe_manifest", lambda source_dir: [])
    monkeypatch.setattr(run_e2e, "parse_facility", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_fixed_assets", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_it_simulation", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_ga", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_birthday_workbook", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_manual_headcount", lambda conn, source_dir, base_dir=None: {})
    monkeypatch.setattr(run_e2e, "parse_manual_special_costs", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_manual_event_drivers", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "parse_nnn_paperwork", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "AllocationEngine", Engine)
    monkeypatch.setattr(run_e2e, "HubBuilder", Builder)
    monkeypatch.setattr(run_e2e, "write_pipeline_audit_report", lambda **kwargs: {"report_path": "audit.md", "missing_csv_path": "missing.csv"})
    monkeypatch.setattr(run_e2e, "audit_exchange_rate_workbook", lambda *args, **kwargs: {})
    monkeypatch.setattr(run_e2e, "write_exchange_rate_audit_report", lambda *args, **kwargs: "exchange-audit.xlsx")
    monkeypatch.setattr(run_e2e, "apply_mp_saisan_complete_v1", lambda **kwargs: complete_reference_calls.append(kwargs) or {})
    monkeypatch.setattr(run_e2e, "_apply_complete_v1_source_order", lambda workbook_path, log_callback, phase, **kwargs: {"phase": phase})

    ok, message = run_e2e.run_universal_pipeline(
        2027,
        str(tmp_path / "template.xlsx"),
        str(tmp_path),
        target_cc=1412000006,
        mp_saisan_complete_v1=True,
        reference_map_path=str(tmp_path / "missing-map.csv"),
    )

    assert ok is True, message
    assert export_calls
    assert complete_reference_calls == []
