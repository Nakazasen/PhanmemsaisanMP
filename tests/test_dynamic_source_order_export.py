import sqlite3
from pathlib import Path

import openpyxl
import pytest

from src.db.schema import create_schema
from src.engine.dynamic_source_order_export import (
    DynamicExportError,
    export_dynamic_source_order,
    resolve_form_layout,
)


PERIODS = ["202604", "202605", "202606", "202607", "202608", "202609", "202610", "202611", "202612", "202701", "202702", "202703"]


def _workbook(path: Path, *, column_shift: int = 0, row_shift: int = 0) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail"
    account = 2 + column_shift
    name = account + 1
    group = account + 2
    item = account + 3
    month_start = account + 4
    total = month_start + 12
    description = total + 1
    wbs = total + 2
    period_row = 4 + row_shift
    header_row = 7 + row_shift
    seed_row = 9 + row_shift
    output_row = seed_row + 1
    output_end = output_row + 30

    ws.cell(1 + row_shift, account, "原価センター")
    ws.cell(2 + row_shift, account, 1412000006)
    for offset, month in enumerate((4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3)):
        ws.cell(period_row, month_start + offset, month)
    ws.cell(period_row, total, "合計")
    ws.cell(header_row, name, "勘定科目名称")
    ws.cell(header_row, group, "採算科目")
    ws.cell(header_row, item, "ﾃｰﾏｺｰﾄﾞ")
    ws.cell(header_row, description, "備考")
    ws.cell(header_row, wbs, "WBS")
    ws.cell(seed_row, account, 5000000000)
    ws.cell(3 + row_shift, month_start, f"=SUM({openpyxl.utils.get_column_letter(month_start)}{seed_row}:{openpyxl.utils.get_column_letter(month_start)}{output_end})")
    for row in range(output_row, output_end + 1):
        account_letter = openpyxl.utils.get_column_letter(account)
        ws.cell(row, name, f'=IF({account_letter}{row}="","","name")')
        ws.cell(row, group, f'=IF({account_letter}{row}="","","group")')
        ws.cell(row, total, f"=SUM({openpyxl.utils.get_column_letter(month_start)}{row}:{openpyxl.utils.get_column_letter(month_start + 11)}{row})")
    wb.save(path)
    wb.close()


def _connection() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    create_schema(conn)
    return conn


def _insert(conn, source, period, amount, account, description):
    conn.execute(
        """INSERT INTO fact_input_data
        (source,period,amount_vnd,cc_code,account_code,description)
        VALUES(?,?,?,?,?,?)""",
        (source, period, amount, 1412000006, account, description),
    )


def test_layout_is_discovered_after_rows_and_columns_move(tmp_path):
    path = tmp_path / "shifted.xlsx"
    _workbook(path, column_shift=3, row_shift=4)
    wb = openpyxl.load_workbook(path, data_only=False)
    layout = resolve_form_layout(wb, "Detail")
    wb.close()

    assert layout.output_start_row == 14
    assert layout.account_col == 5
    assert layout.description_col == 22
    assert layout.wbs_col == 23
    assert layout.output_end_row == 44


def test_layout_fails_closed_when_required_header_is_missing(tmp_path):
    path = tmp_path / "missing-header.xlsx"
    _workbook(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Detail"]
    ws.cell(7, 19).value = None
    wb.save(path)
    wb.close()
    wb = openpyxl.load_workbook(path, data_only=False)
    with pytest.raises(DynamicExportError, match="tiêu đề"):
        resolve_form_layout(wb, "Detail")
    wb.close()


def test_layout_fails_closed_when_required_header_is_duplicated(tmp_path):
    path = tmp_path / "duplicate-header.xlsx"
    _workbook(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Detail"]
    ws.cell(7, 21).value = "備考"
    wb.save(path)
    wb.close()
    wb = openpyxl.load_workbook(path, data_only=False)
    with pytest.raises(DynamicExportError, match="tiêu đề"):
        resolve_form_layout(wb, "Detail")
    wb.close()


def test_single_pass_export_groups_rows_and_leaves_one_blank_separator(tmp_path):
    template = tmp_path / "FORM.xlsx"
    output = tmp_path / "result.xlsx"
    _workbook(template)
    conn = _connection()
    _insert(conn, "facility", PERIODS[0], 100, 5001, "electric")
    _insert(conn, "facility", PERIODS[1], 200, 5001, "electric")
    _insert(conn, "it_sim", PERIODS[0], 300, 5002, "mail")
    _insert(conn, "alloc_1", PERIODS[0], 400, 5003, "Alloc: Event A")
    conn.commit()
    manifest = [
        {"order": "1", "category": "facility", "filename": "facility-renamed.xlsx", "enabled": "1"},
        {"order": "2", "category": "it_simulation", "filename": "system-part-a.xls", "enabled": "1"},
        {"order": "3", "category": "it_simulation", "filename": "system-part-b.xls", "enabled": "1"},
        {"order": "4", "category": "allocation_rules", "filename": "master-renamed.xlsx", "enabled": "1"},
    ]

    result = export_dynamic_source_order(
        conn,
        fiscal_year=2027,
        template_path=template,
        output_path=output,
        cc_code=1412000006,
        manifest_entries=manifest,
        sheet_name="Detail",
        require_staffing=False,
    )

    assert result == {
        "source_blocks_written": 3,
        "rows_written": 3,
        "blank_rows_written": 2,
        "start_row": 10,
        "end_row": 14,
    }
    wb = openpyxl.load_workbook(output, data_only=False)
    ws = wb["Detail"]
    assert ws.cell(10, 2).value == 5001
    assert ws.cell(11, 2).value is None
    assert ws.cell(12, 2).value == 5002
    assert ws.cell(13, 2).value is None
    assert ws.cell(14, 2).value == 5003
    assert ws.cell(10, 6).value == 100
    assert ws.cell(10, 7).value == 200
    audit = list(wb["_mp2027_output_audit"].iter_rows(min_row=2, values_only=True))
    assert "system-part-a.xls" in audit[1][5]
    assert "system-part-b.xls" in audit[1][5]
    wb.close()
    conn.close()


def test_export_is_idempotent_and_clears_stale_rows(tmp_path):
    template = tmp_path / "FORM.xlsx"
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _workbook(template)
    conn = _connection()
    _insert(conn, "facility", PERIODS[0], 100, 5001, "electric")
    _insert(conn, "alloc_1", PERIODS[0], 400, 5003, "Alloc: Event A")
    conn.commit()
    manifest = [{"order": "1", "category": "facility", "filename": "a.xlsx", "enabled": "1"}]
    export_dynamic_source_order(conn, fiscal_year=2027, template_path=template, output_path=first, cc_code=1412000006, manifest_entries=manifest, sheet_name="Detail", require_staffing=False)
    conn.execute("DELETE FROM fact_input_data WHERE source='alloc_1'")
    conn.commit()
    export_dynamic_source_order(conn, fiscal_year=2027, template_path=first, output_path=second, cc_code=1412000006, manifest_entries=manifest, sheet_name="Detail", require_staffing=False)

    wb = openpyxl.load_workbook(second, data_only=False)
    ws = wb["Detail"]
    assert ws.cell(10, 2).value == 5001
    assert all(ws.cell(row, 2).value is None for row in range(11, 41))
    wb.close()
    conn.close()


def test_repository_form_has_clean_dynamic_start_area():
    wb = openpyxl.load_workbook("docs/MP2027/FORM.xlsx", read_only=False, data_only=False)
    layout = resolve_form_layout(wb)
    ws = wb[layout.sheet_name]
    assert layout.output_start_row == 30
    assert all(
        ws.cell(row, col).value in (None, "")
        for row in range(layout.output_start_row, layout.output_end_row + 1)
        for col in (layout.account_col, layout.description_col, layout.wbs_col)
    )
    wb.close()


def test_repository_form_declares_dynamic_output_metadata():
    wb = openpyxl.load_workbook("docs/MP2027/FORM.xlsx", data_only=False)
    layout = resolve_form_layout(wb)
    output_name = wb.defined_names["MP_OUTPUT_AREA"]
    template_name = wb.defined_names["MP_OUTPUT_ROW_TEMPLATE"]
    output_destinations = list(output_name.destinations)
    template_destinations = list(template_name.destinations)
    assert output_destinations[0][0] == layout.sheet_name
    assert template_destinations[0][0] == layout.sheet_name
    assert str(layout.output_start_row) in output_destinations[0][1]
    assert str(layout.output_end_row) in output_destinations[0][1]
    assert str(layout.template_row) in template_destinations[0][1]
    wb.close()


def test_new_production_export_has_no_legacy_destination_rows_or_source_filenames():
    source = Path("src/engine/dynamic_source_order_export.py").read_text(encoding="utf-8")
    pipeline = Path("scripts/run_e2e.py").read_text(encoding="utf-8")
    assert "form_row" not in source
    assert "SOURCE_ROW_GROUPS" not in source
    assert 'os.path.join(source_dir, "施設課' not in pipeline
    assert "export_dynamic_source_order(" in pipeline
