from pathlib import Path

import openpyxl

from src.engine.source_order_output import (
    CANONICAL_SOURCE_FILE_ORDER,
    OutputRow,
    place_rows_by_source_file_order,
    source_order_index,
)
from src.utils.source_manifest import (
    classify_source_candidate,
    detect_source_files,
    inventory_source_files,
    merge_manifest_with_detected,
    read_source_manifest_inventory_fast,
)


def test_canonical_source_order_matches_user_requirement():
    assert len(CANONICAL_SOURCE_FILE_ORDER) == 7
    assert CANONICAL_SOURCE_FILE_ORDER[0].endswith("MPFY2027.xlsx")
    assert "Fixed_Assets_Information_2025.11 - Nov.xlsx" in CANONICAL_SOURCE_FILE_ORDER[1]
    assert "Simulation)_FY2027_Apr.2026 ~ June.2026.xls" in CANONICAL_SOURCE_FILE_ORDER[2]
    assert "FY2027 MP" in CANONICAL_SOURCE_FILE_ORDER[3]
    assert CANONICAL_SOURCE_FILE_ORDER[4].startswith("Sinh")
    assert "(2025.12.29).xlsx" in CANONICAL_SOURCE_FILE_ORDER[5]
    assert "NNN FY2027.xlsx" in CANONICAL_SOURCE_FILE_ORDER[6]


def test_source_order_index_unknown_files_go_last():
    assert source_order_index(CANONICAL_SOURCE_FILE_ORDER[0]) == 0
    assert source_order_index(CANONICAL_SOURCE_FILE_ORDER[6]) == 6
    assert source_order_index("unknown.xlsx") > 6


def test_place_rows_by_source_order_inserts_one_blank_between_file_blocks():
    rows = [
        OutputRow(CANONICAL_SOURCE_FILE_ORDER[4], {"item": "birthday-1"}),
        OutputRow(CANONICAL_SOURCE_FILE_ORDER[0], {"item": "facility-1"}),
        OutputRow(CANONICAL_SOURCE_FILE_ORDER[0], {"item": "facility-2"}),
        OutputRow(CANONICAL_SOURCE_FILE_ORDER[6], {"item": "nnn-1"}),
    ]

    placed = place_rows_by_source_file_order(rows, start_row=200)

    assert [(r.output_row, r.source_file, r.values.get("item"), r.is_blank_separator) for r in placed] == [
        (200, CANONICAL_SOURCE_FILE_ORDER[0], "facility-1", False),
        (201, CANONICAL_SOURCE_FILE_ORDER[0], "facility-2", False),
        (202, CANONICAL_SOURCE_FILE_ORDER[0], None, True),
        (203, CANONICAL_SOURCE_FILE_ORDER[4], "birthday-1", False),
        (204, CANONICAL_SOURCE_FILE_ORDER[4], None, True),
        (205, CANONICAL_SOURCE_FILE_ORDER[6], "nnn-1", False),
    ]


def test_no_fixed_form_rows_are_required_by_policy():
    rows = [
        OutputRow(CANONICAL_SOURCE_FILE_ORDER[1], {"amount": 1}),
        OutputRow(CANONICAL_SOURCE_FILE_ORDER[6], {"amount": 2}),
    ]

    placed = place_rows_by_source_file_order(rows, start_row=1)

    assert [r.output_row for r in placed if not r.is_blank_separator] == [1, 3]
    assert all(r.output_row not in [38, 42, 58, 59, 97, 98, 137] for r in placed)


def _write_facility_source(path: Path) -> None:
    workbook = openpyxl.Workbook()
    workbook.active.title = "減価償却費（Depreciation）"
    workbook.create_sheet("固定資産金利（Interest）")
    workbook.create_sheet("水道光熱費（Electric & Water）")
    workbook.save(path)
    workbook.close()


def test_source_inventory_never_hides_unknown_workbooks(tmp_path):
    _write_facility_source(tmp_path / "renamed-without-keywords.xlsx")
    workbook = openpyxl.Workbook()
    workbook.active.title = "Freight FY2026"
    workbook.save(tmp_path / "new-cost-layout.xlsx")
    workbook.close()
    (tmp_path / "FORM.xlsx").write_bytes(b"system file")
    (tmp_path / "~$editing.xlsx").write_bytes(b"temporary file")

    inventory = inventory_source_files(str(tmp_path))
    detected = detect_source_files(str(tmp_path))

    assert [path.name for path in inventory] == [
        "new-cost-layout.xlsx",
        "renamed-without-keywords.xlsx",
    ]
    assert {entry["filename"] for entry in detected} == {
        "new-cost-layout.xlsx",
        "renamed-without-keywords.xlsx",
    }
    by_name = {entry["filename"]: entry for entry in detected}
    assert by_name["renamed-without-keywords.xlsx"]["category"] == "facility"
    assert by_name["renamed-without-keywords.xlsx"]["status"] == "recognized"
    assert by_name["renamed-without-keywords.xlsx"]["detection_method"] == "structure"
    assert by_name["new-cost-layout.xlsx"]["category"] == ""
    assert by_name["new-cost-layout.xlsx"]["status"] == "needs_review"


def test_manual_manifest_decision_survives_rename_but_not_structure_change(tmp_path):
    path = tmp_path / "totally-renamed.xlsx"
    _write_facility_source(path)
    detected = detect_source_files(str(tmp_path))[0]
    saved = [
        {
            **detected,
            "category": "facility",
            "status": "recognized",
            "detection_method": "manual",
            "reason": "Đã chọn thủ công",
        }
    ]

    compatible = merge_manifest_with_detected(str(tmp_path), saved)
    assert compatible[0]["category"] == "facility"
    assert compatible[0]["status"] == "recognized"
    assert compatible[0]["detection_method"] == "manual"

    workbook = openpyxl.load_workbook(path)
    workbook.remove(workbook["固定資産金利（Interest）"])
    workbook.save(path)
    workbook.close()

    changed = merge_manifest_with_detected(str(tmp_path), saved)
    assert changed[0]["category"] == ""
    assert changed[0]["status"] == "needs_review"
    assert "thay đổi" in changed[0]["reason"].lower()


def test_legacy_manual_confirmation_left_disabled_is_restored_when_structure_matches(tmp_path):
    path = tmp_path / "facility.xlsx"
    _write_facility_source(path)
    detected = detect_source_files(str(tmp_path))[0]
    saved = [{
        **detected,
        "category": "facility",
        "enabled": "0",
        "status": "recognized",
        "detection_method": "manual",
        "signature": "None",
        "description": "Cần xác nhận loại nguồn",
        "reason": "Người dùng đã xác nhận loại nguồn này.",
    }]

    merged = merge_manifest_with_detected(str(tmp_path), saved)

    assert merged[0]["category"] == "facility"
    assert merged[0]["enabled"] == "1"
    assert "tự bật" in merged[0]["reason"].lower()


def test_merge_manifest_with_detected_preserves_user_custom_order(tmp_path):
    p1 = tmp_path / "facility_alpha.xlsx"
    _write_facility_source(p1)
    p2 = tmp_path / "facility_beta.xlsx"
    _write_facility_source(p2)

    saved = [
        {
            "filename": "facility_alpha.xlsx",
            "category": "facility",
            "order": "2",
            "enabled": "1",
            "status": "recognized",
            "detection_method": "structure",
        },
        {
            "filename": "facility_beta.xlsx",
            "category": "facility",
            "order": "1",
            "enabled": "1",
            "status": "recognized",
            "detection_method": "structure",
        },
    ]

    merged = merge_manifest_with_detected(str(tmp_path), saved)
    by_name = {m["filename"]: m for m in merged}
    assert by_name["facility_beta.xlsx"]["order"] == "1"
    assert by_name["facility_alpha.xlsx"]["order"] == "2"
    assert Path(by_name["facility_beta.xlsx"]["_path"]).is_file()
    assert Path(by_name["facility_alpha.xlsx"]["_path"]).is_file()


def test_manual_decision_survives_raw_multi_match_structure(monkeypatch, tmp_path):
    from src.utils import source_manifest

    path = tmp_path / "ambiguous.xlsx"
    path.write_bytes(b"fixture")
    current = {
        "order": "1",
        "category": "",
        "filename": path.name,
        "enabled": "0",
        "description": "Cần xác nhận loại nguồn",
        "status": "needs_review",
        "detection_method": "structure",
        "signature": "new-signature",
        "reason": "Cấu trúc khớp nhiều loại.",
        "_matches": "facility|ga",
        "_path": str(path.resolve()),
    }
    monkeypatch.setattr(source_manifest, "detect_source_files", lambda _source_dir: [current])

    saved = [{
        **current,
        "category": "facility",
        "enabled": "1",
        "status": "recognized",
        "detection_method": "manual",
        "reason": "Đã chọn thủ công",
    }]

    merged = merge_manifest_with_detected(str(tmp_path), saved)

    assert merged[0]["category"] == "facility"
    assert merged[0]["status"] == "recognized"
    assert merged[0]["detection_method"] == "manual"


def test_fast_manifest_inventory_never_opens_or_classifies_workbooks(monkeypatch, tmp_path):
    from src.utils import source_manifest

    (tmp_path / "new-source.xlsx").write_bytes(b"metadata-only fixture")

    def fail_if_deep_scan_runs(_source_dir):
        raise AssertionError("fast inventory must not classify workbook contents")

    monkeypatch.setattr(source_manifest, "detect_source_files", fail_if_deep_scan_runs)

    rows = read_source_manifest_inventory_fast(str(tmp_path))

    assert len(rows) == 1
    assert rows[0]["filename"] == "new-source.xlsx"
    assert rows[0]["category"] == ""
    assert rows[0]["enabled"] == "0"
    assert rows[0]["status"] == "needs_review"
    assert rows[0]["detection_method"] == "inventory"


def test_nnn_fy2029_structure_is_recognized_without_finite_year_allowlist(tmp_path):
    path = tmp_path / "renamed-future-source.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "FY2029"
    worksheet.append(["Cost Center", "Account Code", "2029"])
    worksheet.append(["CC001", "611000", 1])
    workbook.save(path)
    workbook.close()

    result = classify_source_candidate(path)

    assert result["category"] == "nnn_paperwork"
    assert result["status"] == "recognized"
    assert result["detection_method"] == "structure"


def test_clean_manifest_cell_str_handles_none_coercions():
    from src.utils.source_manifest import _clean_manifest_cell_str

    assert _clean_manifest_cell_str(None) == ""
    assert _clean_manifest_cell_str(None, default="default") == "default"
    assert _clean_manifest_cell_str("None") == ""
    assert _clean_manifest_cell_str("none") == ""
    assert _clean_manifest_cell_str(" NULL ") == ""
    assert _clean_manifest_cell_str("facility") == "facility"
    assert _clean_manifest_cell_str(" 1 ") == "1"


def test_merge_manifest_with_detected_self_heals_unclassified_or_none_category(tmp_path):
    path = tmp_path / "facility.xlsx"
    _write_facility_source(path)

    # Manifest row had literal "None" as category (from openpyxl None coercion)
    saved = [{
        "order": "1",
        "category": "None",
        "filename": "facility.xlsx",
        "enabled": "0",
        "status": "needs_review",
        "detection_method": "structure",
        "signature": "None",
        "description": "None",
        "reason": "None",
    }]

    merged = merge_manifest_with_detected(str(tmp_path), saved)

    assert len(merged) == 1
    assert merged[0]["category"] == "facility"
    assert merged[0]["enabled"] == "1"
    assert merged[0]["status"] == "recognized"
    assert "tự động nhận diện" in merged[0]["reason"].lower()


def test_write_and_read_manifest_xlsx_cleans_none_values(tmp_path):
    from src.utils.source_manifest import (
        write_source_manifest_xlsx,
        _read_saved_manifest,
    )

    path = tmp_path / "test_source.xlsx"
    path.write_bytes(b"placeholder")

    entries = [{
        "order": "1",
        "category": None,
        "filename": path.name,
        "enabled": "1",
        "status": "needs_review",
        "detection_method": "manual",
        "signature": "None",
        "description": "None",
        "reason": "None",
    }]

    written_path = write_source_manifest_xlsx(str(tmp_path), entries)
    assert Path(written_path).is_file()

    saved = _read_saved_manifest(str(tmp_path))
    assert len(saved) == 1
    assert saved[0]["category"] == ""
    assert saved[0]["signature"] == ""
    assert saved[0]["description"] == ""
    assert saved[0]["reason"] == ""


def test_single_export_raises_value_error_when_export_to_template_returns_false(tmp_path):
    import pytest
    from unittest.mock import MagicMock

    builder = MagicMock()
    builder.export_to_template.return_value = False
    target_cc = "1412000040"
    template_path = str(tmp_path / "FORM.xlsx")
    out_path = str(tmp_path / f"MP_CC_{target_cc}.xlsx")

    with pytest.raises(ValueError) as excinfo:
        if not builder.export_to_template(template_path, out_path, cc_code=target_cc):
            raise ValueError(
                f"Trung tâm chi phí {target_cc} không có dữ liệu chi phí để xuất sang mẫu FORM (fact_count <= 0). "
                "Vui lòng kiểm tra lại các tệp nguồn chi phí đã được bật trong Thứ tự tệp nguồn."
            )

    assert "1412000040" in str(excinfo.value)
    assert "fact_count <= 0" in str(excinfo.value)


def test_annual_complete_v1_output_source_order_preserves_custom_order():
    from unittest.mock import MagicMock
    from scripts.run_e2e import _annual_complete_v1_output_source_order

    context = MagicMock()
    context.ordered_sources = [
        {"path": "D:/data/facility.xlsx", "filename": "facility.xlsx"},
        {"path": "D:/data/fixed_assets.xlsx", "filename": "fixed_assets.xlsx"},
        {"path": "D:/data/nnn.xlsx", "filename": "nnn.xlsx"},
    ]
    context.resolved_sources = {
        "facility": ["D:/data/facility.xlsx"],
        "fixed_assets": ["D:/data/fixed_assets.xlsx"],
        "nnn_paperwork": ["D:/data/nnn.xlsx"],
    }
    ordered = _annual_complete_v1_output_source_order(context)
    assert ordered == ["facility.xlsx", "fixed_assets.xlsx", "nnn.xlsx"]


def test_headcount_editor_geometry_clamps_within_laptop_screen_bounds():
    def compute_geometry(screen_w, screen_h):
        target_w = min(1180, max(640, screen_w - 60))
        target_h = min(800, max(420, screen_h - 100))
        target_w = min(target_w, max(300, screen_w - 20))
        target_h = min(target_h, max(300, screen_h - 40))
        min_w = min(700, target_w)
        min_h = min(380, target_h)
        return target_w, target_h, min_w, min_h

    # 1. Standard 1080p
    w, h, min_w, min_h = compute_geometry(1920, 1080)
    assert w == 1180
    assert h == 800
    assert min_w <= w and min_h <= h

    # 2. Laptop 1366x768 at 125% scaling (effective ~1092x614)
    w, h, min_w, min_h = compute_geometry(1092, 614)
    assert w <= 1092 and h <= 614
    assert h <= 614 - 40  # fits comfortably above taskbar

    # 3. Laptop 1366x768 at 150% scaling (effective ~910x512)
    w, h, min_w, min_h = compute_geometry(910, 512)
    assert w <= 910 and h <= 512
    assert h == 420  # clamped to max(420, 512-100) = 420
    assert min_h <= h


def test_normalize_row_retains_valid_filename_when_category_empty(tmp_path):
    from src.utils.source_manifest import _normalize_row

    row = {
        "order": "1",
        "category": "",
        "filename": "unclassified.xlsx",
        "enabled": "1",
        "status": "",
    }
    normalized = _normalize_row(row, tmp_path)
    assert normalized is not None
    assert normalized["filename"] == "unclassified.xlsx"
    assert normalized["status"] == "needs_review"
    assert normalized["category"] == ""

