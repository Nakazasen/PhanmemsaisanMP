import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.db.schema import create_schema
from src.engine.allocator import AllocationEngine
from src.parsers.extracted_headcount_time_plan import parse_extracted_headcount_time_plan
from src.services.headcount_source_importer import import_headcount_time_sources
from src.services.headcount_source_policy import HeadcountSourceError, load_canonical_headcount
from src.utils.excel_helpers import get_fy_months


OUTPUT_SHEET = "人員・時間計画"
INDEX_SHEET = "Sheet1"


def create_split_required_company_form(path: Path) -> None:
    """Create the smallest valid 28-row version of the company Master Plan form."""
    workbook = Workbook()
    plan = workbook.active
    plan.title = OUTPUT_SHEET
    workbook.create_sheet(INDEX_SHEET)
    plan["A1"] = "FY2027 マスタープラン人員・時間計画表"
    plan["A5"] = "1412000001"
    plan["B5"] = "Test"

    for column, period in enumerate(get_fy_months(2027), start=3):
        plan.cell(8, column, int(period[-2:]))
        plan.cell(10, column, 1)
        plan.cell(13, column, 9)
        plan.cell(14, column, 10)
        plan.cell(17, column, 100)
        plan.cell(20, column, 900)
        plan.cell(21, column, 1000)
        plan.cell(24, column, 10)
        plan.cell(27, column, 90)
        plan.cell(28, column, 100)
    workbook.save(path)
    workbook.close()


class ExtractedImportTests(unittest.TestCase):
    def test_split_required_preserves_total_and_rejects_category_driver(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "01_Test_FY2027_staffing_truth.xlsx"
            create_split_required_company_form(path)
            conn = sqlite3.connect(":memory:")
            conn.row_factory = sqlite3.Row
            create_schema(conn)
            conn.execute(
                "INSERT INTO dim_cost_centers(code,name_jp,saisan_type,cost_type) "
                "VALUES('1412000001','Test','MFG','Fixed')"
            )

            result = import_headcount_time_sources(conn, tmp, 2027)

            self.assertEqual(result["split_required_files"], 1)
            row = conn.execute("SELECT * FROM fact_monthly_headcount LIMIT 1").fetchone()
            self.assertEqual(row["headcount_all"], 10)
            self.assertEqual(row["headcount_local_total"], 9)
            self.assertIsNone(row["headcount_staff"])
            canonical = load_canonical_headcount(conn, 2027)
            self.assertEqual(canonical[("1412000001", "202604")].headcount_all, 10)
            self.assertIsNone(canonical[("1412000001", "202604")].headcount_staff)
            engine = AllocationEngine.__new__(AllocationEngine)
            engine.hc_cache = {("1412000001", "202604"): {"headcount_all": 10, "headcount_staff": None}}
            with self.assertRaises(HeadcountSourceError):
                engine._get_canonical_monthly_hc("1412000001", "202604", "headcount_staff")

    def test_nonzero_local_hours_without_headcount_is_split_required(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "01_Test_FY2027_staffing_truth.xlsx"
            create_split_required_company_form(path)
            workbook = __import__("openpyxl").load_workbook(path)
            plan = workbook[OUTPUT_SHEET]
            plan["C13"] = 0
            plan["C14"] = 1
            workbook.save(path)
            workbook.close()

            parsed = parse_extracted_headcount_time_plan(str(path), 2027)

            self.assertEqual(parsed.status, "valid", parsed.errors)
            self.assertEqual(parsed.rows[0]["split_status"], "SPLIT_REQUIRED")
            self.assertIsNone(parsed.rows[0]["headcount_staff"])
            self.assertEqual(parsed.rows[0]["fixed_hours_local"], 900)


if __name__ == "__main__":
    unittest.main()
