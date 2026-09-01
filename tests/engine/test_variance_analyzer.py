import pytest
import pandas as pd
from src.engine.variance_analyzer import (
    safe_load_mp_form,
    VarianceStatus,
    CostLineVariance,
    VarianceReport,
    calculate_variance,
    map_and_analyze_variances,
    ComparisonContext,
    is_variance_alert,
    scan_directories_and_pair_files,
    _is_cost_row,
)

# --- Helpers ---
ACC_1 = "9114120018"
ACC_2 = "5005246282"
ACC_3 = "6005146628"
ACC_4 = "6005146542"

def _make_ctx(**overrides):
    defaults = dict(
        fiscal_year_base=2026,
        fiscal_year_current=2027,
        cost_center_code="1001",
        base_file_path="",
        current_file_path="",
    )
    defaults.update(overrides)
    return ComparisonContext(**defaults)


# --- Unit tests ---

def test_is_cost_row_filter():
    assert _is_cost_row(40, ACC_1) is True          # 10-digit
    assert _is_cost_row(39, "5005246") is True       # 7-digit minimum
    assert _is_cost_row(50, 9114120018) is True      # int
    assert _is_cost_row(60, 9114120018.0) is True    # float
    assert _is_cost_row(40, "1001") is False         # 4-digit -> NOT cost
    assert _is_cost_row(40, "FY2027") is False       # text
    assert _is_cost_row(40, None) is False
    assert _is_cost_row(40, float("nan")) is False
    assert _is_cost_row(40, "") is False
    assert _is_cost_row(2, ACC_1) is False           # row 2 is department header (not an input row and < 38)
    assert _is_cost_row(8, ACC_1) is True            # row 8 is an input row

def test_calculate_variance_edge_cases():
    assert is_variance_alert(60000000, 15.0, 50000000, 10.0) is True
    assert is_variance_alert(40000000, 5.0, 50000000, 10.0) is False

def test_scan_directories_regex(tmp_path):
    base_dir = tmp_path / "base"
    curr_dir = tmp_path / "curr"
    base_dir.mkdir()
    curr_dir.mkdir()
    (base_dir / "MP_1412000040_FY2026.xlsx").touch()
    (curr_dir / "MP_1412000040_FY2027.xlsx").touch()
    (base_dir / "CostCenter_1001.xlsx").touch()
    (curr_dir / "CostCenter_1001_v2.xlsx").touch()
    (base_dir / "Unmatched_2002.xlsx").touch()
    (curr_dir / "Unmatched_3003.xlsx").touch()
    pairs, unmatched = scan_directories_and_pair_files(str(base_dir), str(curr_dir))
    assert len(pairs) == 2
    assert len(unmatched) == 2

# --- Blank/NaN validation tests ---

def test_empty_string_cell_raises_error():
    ctx = _make_ctx(base_file_path="D:/data/MP_1001_FY2026.xlsx",
                    current_file_path="D:/data/MP_1001_FY2027.xlsx")
    df_base = pd.DataFrame([{1: ACC_1, 2: "Chi phi tiep khach", 17: ""}])
    df_base.index = [39]  # make it a cost row
    df_curr = pd.DataFrame([{1: ACC_1, 2: "Chi phi tiep khach", 17: 100.0}])
    df_curr.index = [39]
    with pytest.raises(ValueError) as excinfo:
        map_and_analyze_variances(df_base, df_curr, ctx)
    msg = str(excinfo.value)
    assert ACC_1 in msg
    assert "Chi phi tiep khach" in msg
    assert "Ô chi phí đang để trống" in msg
    assert "Nguyên nhân:" in msg
    assert "Cách xử lý:" in msg
    assert "1. Mở tệp" in msg

def test_whitespace_only_cell_raises_error():
    ctx = _make_ctx(base_file_path="D:/data/MP_1001_FY2026.xlsx",
                    current_file_path="D:/data/MP_1001_FY2027.xlsx")
    df_base = pd.DataFrame([{1: ACC_2, 2: "Chi phi van phong", 17: 200.0}])
    df_base.index = [40]
    df_curr = pd.DataFrame([{1: ACC_2, 2: "Chi phi van phong", 17: "   "}])
    df_curr.index = [40]
    with pytest.raises(ValueError) as excinfo:
        map_and_analyze_variances(df_base, df_curr, ctx)
    msg = str(excinfo.value)
    assert ACC_2 in msg
    assert "Ô chi phí đang để trống" in msg
    assert "Nguyên nhân:" in msg
    assert "Cách xử lý:" in msg

def test_none_and_nan_cells_raise_error():
    ctx = _make_ctx(base_file_path="D:/data/MP_1001_FY2026.xlsx",
                    current_file_path="D:/data/MP_1001_FY2027.xlsx")
    df_base_none = pd.DataFrame([{1: ACC_3, 2: "Chi phi xang dau", 17: None}])
    df_base_none.index = [41]
    df_curr_valid = pd.DataFrame([{1: ACC_3, 2: "Chi phi xang dau", 17: 50.0}])
    df_curr_valid.index = [41]
    with pytest.raises(ValueError) as excinfo:
        map_and_analyze_variances(df_base_none, df_curr_valid, ctx)
    msg = str(excinfo.value)
    assert "Ô chi phí đang để trống" in msg
    assert "Nguyên nhân:" in msg
    assert "Cách xử lý:" in msg

    df_base_valid = pd.DataFrame([{1: ACC_3, 2: "Chi phi xang dau", 17: 50.0}])
    df_base_valid.index = [41]
    df_curr_nan = pd.DataFrame([{1: ACC_3, 2: "Chi phi xang dau", 17: float("nan")}])
    df_curr_nan.index = [41]
    with pytest.raises(ValueError) as excinfo:
        map_and_analyze_variances(df_base_valid, df_curr_nan, ctx)
    assert "Ô chi phí đang để trống" in str(excinfo.value)

def test_explicit_zero_is_valid():
    ctx = _make_ctx()
    df_base = pd.DataFrame([
        {1: ACC_1, 2: "Item 1", 17: 0},
        {1: ACC_2, 2: "Item 2", 17: "0.0"}
    ])
    df_base.index = [40, 41]
    df_curr = pd.DataFrame([
        {1: ACC_1, 2: "Item 1", 17: 0.0},
        {1: ACC_2, 2: "Item 2", 17: "0"}
    ])
    df_curr.index = [40, 41]
    report = map_and_analyze_variances(df_base, df_curr, ctx)
    assert report.total_base == 0.0
    assert report.total_current == 0.0
    assert len(report.lines) == 2
    assert all(l.status == VarianceStatus.UNCHANGED for l in report.lines)

def test_non_numeric_invalid_text_raises_error():
    ctx = _make_ctx()
    df_base_bad = pd.DataFrame([{1: ACC_1, 2: "Chi phi", 17: "Loi"}])
    df_base_bad.index = [40]
    df_curr = pd.DataFrame([{1: ACC_1, 2: "Chi phi", 17: 100.0}])
    df_curr.index = [40]
    with pytest.raises(ValueError) as excinfo:
        map_and_analyze_variances(df_base_bad, df_curr, ctx)

# --- Non-cost row skipping tests ---

def test_header_and_layout_rows_are_skipped():
    """Rows without a valid 7+ digit account code OR not meeting cost row line criteria must be silently skipped."""
    ctx = _make_ctx()
    df_base = pd.DataFrame([
        {1: "Header",    2: "Phong Ke Toan",   17: None},       # index 0 (row 1)
        {1: "",          2: "",                17: None},       # index 1 (row 2)
        {1: ACC_1,       2: "Chi phi A",       17: 100.0},      # index 2 (row 3) -> NOT cost because row < 38
        {1: ACC_1,       2: "Chi phi A",       17: 100.0},      # index 40 (row 41) -> Valid cost
        {1: "Total",     2: "Tong cong",       17: 100.0},      # index 41 (row 42) -> text
    ])
    df_base.index = [0, 1, 2, 40, 41]
    df_curr = pd.DataFrame([
        {1: "Header",    2: "Phong Ke Toan",   17: None},
        {1: "",          2: "",                17: None},
        {1: ACC_1,       2: "Chi phi A",       17: 120.0},
        {1: ACC_1,       2: "Chi phi A",       17: 120.0},
        {1: "Total",     2: "Tong cong",       17: 120.0},
    ])
    df_curr.index = [0, 1, 2, 40, 41]
    report = map_and_analyze_variances(df_base, df_curr, ctx)
    assert len(report.lines) == 1
    assert report.lines[0].account_code == ACC_1
    assert report.lines[0].base_value == 100.0
    assert report.lines[0].current_value == 120.0

# --- Batch test ---

def test_batch_with_one_error_and_one_valid_file(tmp_path):
    import openpyxl
    from src.engine.variance_analyzer import batch_analyze_variances

    base_dir = tmp_path / "base"
    curr_dir = tmp_path / "curr"
    base_dir.mkdir()
    curr_dir.mkdir()

    # File 1: CC 1001 -> Cost row has blank value
    wb_b1 = openpyxl.Workbook()
    ws_b1 = wb_b1.active
    ws_b1.title = "内訳Detail"
    ws_b1.cell(row=40, column=2, value=int(ACC_1))
    ws_b1.cell(row=40, column=3, value="Item 1")
    ws_b1.cell(row=40, column=18, value=1000.0)
    b1_path = str(base_dir / "MP_1001_FY2026.xlsx")
    wb_b1.save(b1_path)

    wb_c1 = openpyxl.Workbook()
    ws_c1 = wb_c1.active
    ws_c1.title = "内訳Detail"
    ws_c1.cell(row=40, column=2, value=int(ACC_1))
    ws_c1.cell(row=40, column=3, value="Item 1")
    ws_c1.cell(row=40, column=18, value=None)  # Blank
    c1_path = str(curr_dir / "MP_1001_FY2027.xlsx")
    wb_c1.save(c1_path)

    # File 2: CC 1002 -> Valid file with explicit 0
    wb_b2 = openpyxl.Workbook()
    ws_b2 = wb_b2.active
    ws_b2.title = "内訳Detail"
    ws_b2.cell(row=40, column=2, value=int(ACC_2))
    ws_b2.cell(row=40, column=3, value="Item 2")
    ws_b2.cell(row=40, column=18, value=2000.0)
    b2_path = str(base_dir / "MP_1002_FY2026.xlsx")
    wb_b2.save(b2_path)

    wb_c2 = openpyxl.Workbook()
    ws_c2 = wb_c2.active
    ws_c2.title = "内訳Detail"
    ws_c2.cell(row=40, column=2, value=int(ACC_2))
    ws_c2.cell(row=40, column=3, value="Item 2")
    ws_c2.cell(row=40, column=18, value=0.0)
    c2_path = str(curr_dir / "MP_1002_FY2027.xlsx")
    wb_c2.save(c2_path)

    pairs = [("1001", b1_path, c1_path), ("1002", b2_path, c2_path)]
    reports, errors = batch_analyze_variances(pairs=pairs, base_fy=2026, curr_fy=2027,
                                              thresh_pct=10.0, thresh_abs=50000000.0)
    assert len(reports) == 1
    assert reports[0].context.cost_center_code == "1002"
    assert len(errors) == 1
    assert "1001" in errors[0]

# --- Core variance tests ---

def test_calculate_variance_original():
    assert calculate_variance(100.0, 110.0)[1] == 10.0
    assert calculate_variance(100.0, 90.0)[1] == -10.0
    assert calculate_variance(0.0, 100.0)[1] == 100.0
    assert calculate_variance(0.0, 0.0)[1] == 0.0
    assert calculate_variance(100.0, 110.0)[2] == VarianceStatus.INCREASE
    assert calculate_variance(100.0, 90.0)[2] == VarianceStatus.DECREASE
    assert calculate_variance(100.0, 100.0)[2] == VarianceStatus.UNCHANGED
    assert calculate_variance(0.0, 100.0)[2] == VarianceStatus.NEW_ITEM
    assert calculate_variance(100.0, 0.0)[2] == VarianceStatus.REMOVED

def test_duplicate_keys_aggregation():
    ctx = _make_ctx()
    df_base = pd.DataFrame({
        1: [ACC_1, ACC_1],
        2: ["Item A", "Item A"],
        17: [100.0, 50.0]
    })
    df_base.index = [40, 41]
    df_current = pd.DataFrame({
        1: [ACC_1],
        2: ["Item A"],
        17: [180.0]
    })
    df_current.index = [40]
    report = map_and_analyze_variances(df_base, df_current, ctx)
    assert len(report.lines) == 1
    assert report.lines[0].base_value == 150.0
    assert report.lines[0].current_value == 180.0
    assert report.lines[0].variance_absolute == 30.0

def test_export_variance_report(tmp_path):
    from src.utils.excel_variance_writer import export_variance_report
    ctx = _make_ctx(base_file_path="a.xlsx", current_file_path="b.xlsx")
    report = VarianceReport(context=ctx)
    report.lines.append(CostLineVariance(ACC_1, "Item A", 100, 110, 10, 10.0, VarianceStatus.INCREASE, True))
    export_path = tmp_path / "export.xlsx"
    export_variance_report(report, str(export_path))
    assert export_path.exists()
    import openpyxl
    wb = openpyxl.load_workbook(str(export_path))
    ws = wb.active
    assert ws["A1"].value == "Mã Tài Khoản"
    assert ws["A2"].value == ACC_1
    assert len(ws._charts) == 0
    assert len(ws._images) == 1


def test_export_variance_report_uses_the_active_interface_language(tmp_path):
    from src.services.i18n import set_current_language, t
    from src.utils.excel_variance_writer import export_variance_report

    report = VarianceReport(context=_make_ctx())
    report.lines.append(CostLineVariance(ACC_1, "Item A", 100, 110, 10, 10.0, VarianceStatus.INCREASE, True))
    try:
        for language in ("en", "ja"):
            set_current_language(language)
            export_path = tmp_path / f"export-{language}.xlsx"
            export_variance_report(report, str(export_path))
            import openpyxl
            workbook = openpyxl.load_workbook(export_path)
            worksheet = workbook.active
            assert worksheet.title == t("variance_export_sheet_name")
            assert worksheet["A1"].value == t("col_account_code")
            assert worksheet["G2"].value == t("variance_status_increase")
            workbook.close()
    finally:
        set_current_language("vi")

# --- Integration test: realistic multi-sheet workbook ---

def test_integration_realistic_mp_workbook(tmp_path):
    """
    Integration test with workbook structure similar to real MP FORM:
    - Sheet 0: summary sheet (should NOT be read)
    - Sheet 1: hub sheet with header/dept rows + cost data rows

    Proves:
    1. safe_load_mp_form selects hub sheet, not sheet 0.
    2. Header/layout rows are not falsely blocked (including department code 1412000040 in row 2).
    3. Blank cost cells on valid cost rows are still blocked.
    4. Valid file comparison succeeds.
    """
    import openpyxl
    from src.engine.variance_analyzer import batch_analyze_variances

    def _create_realistic_mp_workbook(filepath, cost_values, blank_row_index=None):
        wb = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Summary"
        ws0.cell(row=1, column=1, value="Summary Sheet")
        ws0.cell(row=2, column=1, value="Department: Test")
        ws0.cell(row=2, column=2, value=None)  # blank cell on summary

        ws1 = wb.create_sheet(title="内訳Detail")
        ws1.cell(row=1, column=1, value="Account Code")
        ws1.cell(row=1, column=2, value="Description")
        # Row 2 (Cost center row - should be skipped)
        ws1.cell(row=2, column=1, value="Cost Center:")
        ws1.cell(row=2, column=2, value=1412000040)

        cost_row_start = 40
        for i, (acc, name, val) in enumerate(cost_values):
            row = cost_row_start + i
            actual_val = val
            if blank_row_index is not None and i == blank_row_index:
                actual_val = None
            ws1.cell(row=row, column=2, value=acc)
            ws1.cell(row=row, column=3, value=name)
            ws1.cell(row=row, column=18, value=actual_val)

        wb.save(filepath)

    # Test 1: safe_load_mp_form reads hub sheet
    valid_path = str(tmp_path / "valid_mp.xlsx")
    _create_realistic_mp_workbook(valid_path, [
        (int(ACC_1), "Depreciation", 500000.0),
        (int(ACC_2), "Labor Cost",   300000.0),
    ])
    df = safe_load_mp_form(valid_path)
    # The first column extracted (col index 0)
    col0_values = df.iloc[:, 0].astype(str).tolist()
    assert "Summary Sheet" not in col0_values
    assert "Account Code" in col0_values

    # Test 2: Header/layout rows not falsely blocked (using default columns)
    ctx = _make_ctx()
    # Mocking the raw df from realistic sheet.
    # Note: row index 1 = excel row 2
    df_base = pd.DataFrame({
        1: ["Account Code", 1412000040, int(ACC_1), int(ACC_2)],
        2: ["Description",  "Test Department", "Depreciation", "Labor Cost"],
        17: ["Month_12",    None, 500000.0, 300000.0],
    })
    df_base.index = [0, 1, 39, 40]
    df_curr = pd.DataFrame({
        1: ["Account Code", 1412000040, int(ACC_1), int(ACC_2)],
        2: ["Description",  "Test Department", "Depreciation", "Labor Cost"],
        17: ["Month_12",    None, 550000.0, 280000.0],
    })
    df_curr.index = [0, 1, 39, 40]
    report = map_and_analyze_variances(df_base, df_curr, ctx)
    assert len(report.lines) == 2
    assert report.total_base == 800000.0
    assert report.total_current == 830000.0

    # Test 3: Blank cost cell on valid cost row still blocked
    df_base_blank = pd.DataFrame({
        1: ["Account Code", int(ACC_1)],
        2: ["Description",  "Depreciation"],
        17: ["Total",        None],
    })
    df_base_blank.index = [0, 40]
    df_curr_ok = pd.DataFrame({
        1: ["Account Code", int(ACC_1)],
        2: ["Description",  "Depreciation"],
        17: ["Total",        550000.0],
    })
    df_curr_ok.index = [0, 40]
    with pytest.raises(ValueError) as excinfo:
        map_and_analyze_variances(df_base_blank, df_curr_ok, ctx)
    assert "trống" in str(excinfo.value) or "de trong" in str(excinfo.value).lower()

    # Test 4: Batch with realistic workbooks
    base_dir = tmp_path / "batch_base"
    curr_dir = tmp_path / "batch_curr"
    base_dir.mkdir()
    curr_dir.mkdir()

    _create_realistic_mp_workbook(str(base_dir / "MP_1001_FY2026.xlsx"), [
        (int(ACC_1), "Depreciation", 500000.0),
        (int(ACC_2), "Labor Cost",   300000.0),
    ])
    _create_realistic_mp_workbook(str(curr_dir / "MP_1001_FY2027.xlsx"), [
        (int(ACC_1), "Depreciation", 550000.0),
        (int(ACC_2), "Labor Cost",   0.0),
    ])

    _create_realistic_mp_workbook(str(base_dir / "MP_1002_FY2026.xlsx"), [
        (int(ACC_3), "Utility", 200000.0),
    ])
    _create_realistic_mp_workbook(str(curr_dir / "MP_1002_FY2027.xlsx"), [
        (int(ACC_3), "Utility", 200000.0),
    ], blank_row_index=0)

    pairs = [
        ("1001", str(base_dir / "MP_1001_FY2026.xlsx"), str(curr_dir / "MP_1001_FY2027.xlsx")),
        ("1002", str(base_dir / "MP_1002_FY2026.xlsx"), str(curr_dir / "MP_1002_FY2027.xlsx")),
    ]
    reports, errors = batch_analyze_variances(pairs=pairs, base_fy=2026, curr_fy=2027,
                                              thresh_pct=10.0, thresh_abs=50000000.0)
    assert len(reports) == 1
    assert reports[0].context.cost_center_code == "1001"
    assert reports[0].total_base == 800000.0
    assert reports[0].total_current == 550000.0

    assert len(errors) == 1
    assert "1002" in errors[0]


def test_yoy_reads_totals_from_a_freshly_exported_formula_workbook(tmp_path, monkeypatch):
    """YoY must work before Excel has opened and cached generated formulas."""
    import openpyxl

    workbook_path = tmp_path / "MP_CC_1412000086.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Hub"
    worksheet["B2"] = 26273
    # FORM summary cells can contain unrelated formulas outside the cost area.
    # They must not block YoY from hydrating a real cost row below row 38.
    worksheet["V29"] = 1.5
    worksheet["F29"] = "=ROUNDUP(V29*$B$2,0)"
    worksheet["R29"] = "=SUM(F29:Q29)"
    worksheet["R5"] = "=SUM(R29:R1000)"
    worksheet.cell(row=40, column=2, value=int(ACC_1))
    worksheet.cell(row=40, column=3, value='=VLOOKUP($B40,AccountMaster!$A:$B,2,0)')
    worksheet.cell(row=40, column=19, value="Generated source description")
    worksheet.cell(row=40, column=6, value="=ROUND(1.5*$B$2,0)")
    worksheet.cell(row=40, column=17, value="=ROUND(2*$B$2,0)")
    worksheet.cell(row=40, column=18, value="=SUM(F40:Q40)")
    workbook.save(workbook_path)
    workbook.close()

    monkeypatch.setattr(
        "src.engine.variance_analyzer._resolve_hub_sheet_name",
        lambda _path: "Hub",
    )
    dataframe = safe_load_mp_form(str(workbook_path))
    report = map_and_analyze_variances(
        dataframe,
        dataframe,
        _make_ctx(base_file_path=str(workbook_path), current_file_path=str(workbook_path)),
    )

    assert len(report.lines) == 1
    assert report.lines[0].item_name == "Generated source description"
    assert report.total_current == 91956.0
    assert report.total_variance_absolute == 0.0
