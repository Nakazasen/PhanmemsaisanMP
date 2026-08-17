from pathlib import Path
import shutil

import openpyxl

from src.services.project_config import ProjectConfig
from src.services.template_confirmation import confirmation_status, inspect_form
from src.universal_app import MPManagerApp


ROOT = Path(__file__).resolve().parents[1]
FORM = ROOT / "docs" / "MP2027" / "FORM.xlsx"


def test_new_valid_form_requires_confirmation_and_then_is_known(tmp_path):
    form = tmp_path / "FORM.xlsx"
    shutil.copy2(FORM, form)

    inspection = inspect_form(form)

    assert inspection.is_valid
    assert confirmation_status({}, inspection) == "new"
    assert confirmation_status({inspection.checksum: inspection.as_confirmation()}, inspection) == "known"


def test_changed_form_at_same_path_requires_confirmation_again(tmp_path):
    form = tmp_path / "FORM.xlsx"
    shutil.copy2(FORM, form)
    first = inspect_form(form)
    workbook = openpyxl.load_workbook(form)
    workbook["内訳ﾘｽﾄ(4～3月)"]["B2"] = 26274
    workbook.save(form)
    workbook.close()
    changed = inspect_form(form)

    assert changed.is_valid
    assert changed.checksum != first.checksum
    assert confirmation_status({first.checksum: first.as_confirmation()}, changed) == "changed"


def test_form_with_department_payload_cannot_be_confirmed(tmp_path):
    form = tmp_path / "FORM.xlsx"
    shutil.copy2(FORM, form)
    workbook = openpyxl.load_workbook(form)
    workbook["内訳ﾘｽﾄ(4～3月)"]["B5"] = 1412000004
    workbook.save(form)
    workbook.close()

    inspection = inspect_form(form)

    assert not inspection.is_valid
    assert inspection.issue_cells == ("B5",)
    assert confirmation_status({}, inspection) == "invalid"


def test_project_persists_form_confirmation(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path))
    record = {"checksum": "a" * 64, "path": str(tmp_path / "FORM.xlsx")}

    project.confirm_form(2027, record)
    project.save()
    restored = ProjectConfig.load(project.config_path)

    assert restored.form_confirmations(2027)[record["checksum"]] == record


def test_ui_allows_new_valid_form_without_confirmation(monkeypatch, tmp_path):
    form = tmp_path / "FORM.xlsx"
    shutil.copy2(FORM, form)
    project = ProjectConfig.create_legacy_compatible(str(tmp_path))
    app = object.__new__(MPManagerApp)
    app.project = project
    monkeypatch.setattr(
        "src.universal_app.messagebox.askyesno",
        lambda *_args: (_ for _ in ()).throw(AssertionError("Valid FORM must not prompt")),
    )

    assert app._confirm_selected_form(str(form), 2027)
    assert project.form_confirmations(2027) == {}


def test_ui_allows_changed_valid_form_without_confirmation(monkeypatch, tmp_path):
    form = tmp_path / "FORM.xlsx"
    shutil.copy2(FORM, form)
    project = ProjectConfig.create_legacy_compatible(str(tmp_path))
    project.confirm_form(2027, inspect_form(form).as_confirmation())
    hub_sheet_name = inspect_form(form).hub_sheet_name
    workbook = openpyxl.load_workbook(form)
    workbook[hub_sheet_name]["B2"] = "changed"
    workbook.save(form)
    workbook.close()
    app = object.__new__(MPManagerApp)
    app.project = project
    monkeypatch.setattr(
        "src.universal_app.messagebox.askyesno",
        lambda *_args: (_ for _ in ()).throw(AssertionError("Valid FORM must not prompt")),
    )

    assert app._confirm_selected_form(str(form), 2027)


def test_ui_reports_invalid_form_and_stops(monkeypatch, tmp_path):
    form = tmp_path / "FORM.xlsx"
    shutil.copy2(FORM, form)
    hub_sheet_name = inspect_form(form).hub_sheet_name
    workbook = openpyxl.load_workbook(form)
    workbook[hub_sheet_name]["B5"] = 1412000004
    workbook.save(form)
    workbook.close()
    app = object.__new__(MPManagerApp)
    errors = []
    monkeypatch.setattr(
        "src.universal_app.messagebox.showerror",
        lambda title, message: errors.append((title, message)),
    )

    assert not app._confirm_selected_form(str(form), 2027)
    assert errors == [
        (
            "FORM không đúng cấu trúc",
            "FORM không đúng cấu trúc.\n"
            "Vui lòng chọn lại tệp hoặc kiểm tra các trang tính bắt buộc.",
        )
    ]
