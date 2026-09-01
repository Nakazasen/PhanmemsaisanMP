from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from src.engine.output_cost_row_ordering import (
    OUTPUT_COST_ROW_ORDER_METADATA_SHEET,
    OutputCostRowOrderError,
    read_cost_rows,
    restore_cost_layout,
    save_cost_row_order,
)
from src.engine.manual_special_cost_sections import preserve_manual_special_cost_section


CC_CODE = "1412000030"
SHEET_NAME = "Chi tiết MP"


def _make_workbook(path: Path, common_labels: list[str]) -> Path:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    for offset, label in enumerate(common_labels, start=38):
        worksheet.cell(offset, 2).value = 5000000000 + offset
        worksheet.cell(offset, 6).value = offset * 1_000
        worksheet.cell(offset, 18).value = f"=SUM(F{offset}:Q{offset})"
        worksheet.cell(offset, 19).value = label
    workbook.save(path)
    workbook.close()
    return path


def _add_legacy_manual_rows(path: Path, start_row: int = 40) -> None:
    workbook = openpyxl.load_workbook(path)
    try:
        worksheet = workbook[SHEET_NAME]
        worksheet.cell(start_row, 2).value = 5005246286
        worksheet.cell(start_row, 6).value = 1_250_000
        worksheet.cell(start_row, 18).value = f"=SUM(F{start_row}:Q{start_row})"
        worksheet.cell(start_row, 19).value = "manual-A"
        worksheet.cell(start_row, 19).fill = PatternFill("solid", fgColor="00FF00")
        worksheet.cell(start_row + 1, 2).value = 5005246288
        worksheet.cell(start_row + 1, 6).value = 2_500_000
        worksheet.cell(start_row + 1, 18).value = f"=SUM(F{start_row + 1}:Q{start_row + 1})"
        worksheet.cell(start_row + 1, 19).value = "manual-B"

        metadata = workbook.create_sheet("_mp2027_manual_special_cost_meta")
        metadata.sheet_state = "veryHidden"
        metadata.append((
            "sheet_name", "fiscal_year", "cc_code", "common_end_row",
            "manual_start_row", "manual_end_row", "schema_version",
        ))
        metadata.append((SHEET_NAME, None, CC_CODE, start_row - 1, start_row, start_row + 1, 1))
        workbook.save(path)
    finally:
        workbook.close()


def _rows_by_label(path: Path) -> dict[str, tuple[int, object, object, object]]:
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        worksheet = workbook[SHEET_NAME]
        return {
            str(worksheet.cell(row, 19).value): (
                row,
                worksheet.cell(row, 2).value,
                worksheet.cell(row, 6).value,
                worksheet.cell(row, 18).value,
            )
            for row in range(38, worksheet.max_row + 1)
            if worksheet.cell(row, 19).value not in (None, "")
        }
    finally:
        workbook.close()


def _save_mixed_layout(path: Path) -> None:
    _make_workbook(path, ["common-A", "common-B"])
    _add_legacy_manual_rows(path)
    rows = read_cost_rows(path, CC_CODE)
    by_label = {row.description: row.row_id for row in rows}
    save_cost_row_order(
        path,
        CC_CODE,
        [by_label[label] for label in ("common-A", "manual-A", "common-B", "manual-B")],
    )


def test_save_order_marks_legacy_manual_rows_and_keeps_mixed_order(tmp_path):
    workbook = tmp_path / "MP_CC_1412000030.xlsx"
    _save_mixed_layout(workbook)

    rows = read_cost_rows(workbook, CC_CODE)

    assert [(row.description, row.row_kind) for row in rows] == [
        ("common-A", "common"),
        ("manual-A", "manual"),
        ("common-B", "common"),
        ("manual-B", "manual"),
    ]
    saved = openpyxl.load_workbook(workbook, data_only=False)
    try:
        assert saved[OUTPUT_COST_ROW_ORDER_METADATA_SHEET].sheet_state == "veryHidden"
    finally:
        saved.close()


def test_rerun_keeps_manual_money_when_rows_are_mixed(tmp_path):
    source = tmp_path / "FY2027.xlsx"
    generated = _make_workbook(tmp_path / "FY2027-rerun.xlsx", ["common-A", "common-B", "common-C"])
    _save_mixed_layout(source)

    result = restore_cost_layout(generated, CC_CODE, source, source_kind="current_fiscal_year")

    rows = _rows_by_label(generated)
    assert list(rows) == ["common-A", "manual-A", "common-B", "manual-B", "common-C"]
    assert rows["manual-A"][2:] == (1_250_000, "=SUM(F39:Q39)")
    assert result["manual_rows_preserved"] == 2
    assert result["new_common_rows"] == 1


def test_new_fy_keeps_manual_code_description_and_order_but_clears_money(tmp_path):
    source = tmp_path / "FY2026.xlsx"
    generated = _make_workbook(tmp_path / "FY2027.xlsx", ["common-A", "common-B", "common-C"])
    _save_mixed_layout(source)

    restore_cost_layout(generated, CC_CODE, source, source_kind="previous_fiscal_year")

    rows = _rows_by_label(generated)
    assert list(rows) == ["common-A", "manual-A", "common-B", "manual-B", "common-C"]
    assert rows["manual-A"][1:] == (5005246286, None, None)
    workbook = openpyxl.load_workbook(generated, data_only=False)
    try:
        assert workbook[SHEET_NAME]["S39"].fill.fgColor.rgb.endswith("00FF00")
    finally:
        workbook.close()


def test_pipeline_wrapper_restores_saved_mixed_layout(tmp_path):
    source = tmp_path / "FY2027.xlsx"
    generated = _make_workbook(tmp_path / "FY2027-rerun.xlsx", ["common-A", "common-B", "common-C"])
    _save_mixed_layout(source)

    result = preserve_manual_special_cost_section(
        generated,
        CC_CODE,
        source_workbook_path=source,
        source_kind="current_fiscal_year",
    )

    assert list(_rows_by_label(generated)) == ["common-A", "manual-A", "common-B", "manual-B", "common-C"]
    assert result["new_common_rows"] == 1


def test_unmarked_workbook_requires_manual_boundary_confirmation(tmp_path):
    workbook = _make_workbook(tmp_path / "unmarked.xlsx", ["common-A"])

    with pytest.raises(OutputCostRowOrderError, match="dấu mốc"):
        read_cost_rows(workbook, CC_CODE)
