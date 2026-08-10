from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from src.engine.complete_v1_source_order_writer import apply_complete_v1_source_order_to_workbook
from src.engine.source_order_output import CANONICAL_SOURCE_FILE_ORDER

SHEET = "蜀・ｨｳ・假ｽｽ・・4・・譛・"
BLUE_FILL = PatternFill(fill_type="solid", fgColor="CCFFFF")


def _workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for row, account, description in [
        (36, 5006016260, "facility building depreciation"),
        (38, 5006016244, "fixed asset depreciation"),
        (42, 9114120007, "fixed asset interest"),
        (59, 5004086291, "birthday"),
        (137, 5005246286, "NNN paperwork"),
        (175, None, "admin toilet paper"),
        (179, None, "system cost"),
    ]:
        ws.cell(row, 2).value = account
        ws.cell(row, 6).value = row
        ws.cell(row, 19).value = description
        ws.cell(row, 20).value = "SOURCE_DERIVED"
    wb.save(path)
    wb.close()
    return path


def _fill_rgb(cell):
    fill = cell.fill
    if fill is None or not fill.fill_type:
        return None
    color = fill.fgColor
    if color is not None and color.type == "rgb":
        return color.rgb
    return fill.fill_type


def _note_text(cell):
    wb = cell.parent.parent
    if "_mp2027_source_order_meta" not in wb.sheetnames:
        return ""
    sheet = wb["_mp2027_source_order_meta"]
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == cell.parent.title and sheet.cell(row, 2).value == cell.row:
            return sheet.cell(row, 3).value or ""
    return ""


def _assert_output_layout_clean(ws, start_row: int = 30, end_row: int | None = None):
    end = end_row or ws.max_row
    for row in range(start_row, end + 1):
        for col in range(1, 5):
            assert _fill_rgb(ws.cell(row, col)) in (None, "00000000", "00FFFFFF", "FFFFFFFF")
        assert ws.cell(row, 5).value is None


def test_complete_v1_writer_rewrites_legacy_rows_to_source_order_blocks(tmp_path):
    workbook = _workbook(tmp_path / "out.xlsx")

    result = apply_complete_v1_source_order_to_workbook(workbook, start_row=168, clear_until_row=190)

    assert result["rows_written"] == 7
    assert result["blank_rows_written"] == 5

    wb = load_workbook(workbook)
    try:
        ws = wb[SHEET]
        assert ws.cell(168, 19).value == "facility building depreciation"
        assert ws.cell(168, 18).value == "=SUM(F168:Q168)"
        assert ws.cell(169, 19).value is None
        assert ws.cell(170, 19).value == "fixed asset depreciation"
        assert ws.cell(171, 19).value == "fixed asset interest"
        assert ws.cell(172, 19).value is None
        assert ws.cell(173, 19).value == "system cost"
        assert ws.cell(175, 19).value == "admin toilet paper"
        assert ws.cell(177, 19).value == "birthday"
        assert ws.cell(179, 19).value == "NNN paperwork"

        for legacy_row in [38, 42, 59, 137]:
            assert ws.cell(legacy_row, 19).value is None
            assert ws.cell(legacy_row, 6).value is None

        assert ws.cell(168, 20).value is None
        assert CANONICAL_SOURCE_FILE_ORDER[0] in _note_text(ws.cell(168, 20))
        assert CANONICAL_SOURCE_FILE_ORDER[1] in _note_text(ws.cell(170, 20))
        assert CANONICAL_SOURCE_FILE_ORDER[6] in _note_text(ws.cell(179, 20))
    finally:
        wb.close()


def test_complete_v1_writer_default_output_starts_at_row_38(tmp_path):
    workbook = _workbook(tmp_path / "out.xlsx")

    result = apply_complete_v1_source_order_to_workbook(workbook, clear_until_row=190)

    assert result["start_row"] == 38
    assert result["rows_written"] == 7
    assert result["blank_rows_written"] == 5

    wb = load_workbook(workbook)
    try:
        ws = wb[SHEET]
        assert ws.cell(38, 19).value == "facility building depreciation"
        assert ws.cell(38, 18).value == "=SUM(F38:Q38)"
        assert ws.cell(39, 19).value is None
        assert ws.cell(40, 19).value == "fixed asset depreciation"
        assert ws.cell(41, 19).value == "fixed asset interest"
        assert ws.cell(42, 19).value is None
        assert ws.cell(43, 19).value == "system cost"
        _assert_output_layout_clean(ws, 38, 48)
    finally:
        wb.close()


def test_complete_v1_writer_preserves_qlnn_rows_30_to_37(tmp_path):
    path = tmp_path / "out_qlnn.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    protected = {
        31: (9114120018, "部内間接経費", "部内間接経費"),
        32: (9114120029, "部外間接経費1", "部外間接経費1"),
        33: (9114120030, "部外間接経費2", "部外間接経費2"),
        34: (9114120021, "工場間接経費", "工場間接経費"),
        36: (9114120009, "社内金利（在庫）", "在庫金利"),
    }
    for row, values in protected.items():
        for column, value in zip((2, 3, 4), values):
            ws.cell(row, column).value = value
    ws.cell(154, 2).value = 5006016260
    ws.cell(154, 6).value = "=ROUND(1.5*$B$2,0)"
    ws.cell(154, 19).value = "facility building depreciation"
    wb.save(path)
    wb.close()

    apply_complete_v1_source_order_to_workbook(path, clear_until_row=190)

    wb = load_workbook(path, data_only=False)
    try:
        ws = wb[SHEET]
        for row, values in protected.items():
            assert tuple(ws.cell(row, column).value for column in (2, 3, 4)) == values
        assert ws.cell(38, 2).value == 5006016260
        assert ws.cell(38, 6).value == "=ROUND(1.5*$B$2,0)"
        assert ws.cell(38, 19).value == "facility building depreciation"
        assert ws.cell(154, 2).value is None
    finally:
        wb.close()


def test_complete_v1_writer_groups_unmarked_facility_depreciation_with_facility_rows(tmp_path):
    """Late facility depreciation rows must not be left after the facility block."""
    path = tmp_path / "out_facility_grouped.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    # This is the current FORM layout, where rows 30–37 belong to QLLN and
    # facility depreciation is normally staged at rows 154–155.
    ws.cell(31, 2).value = 9114120018
    for row, account, description in [
        (40, 9114120007, "facility building interest"),
        (41, 9114120007, "facility land interest"),
        (44, 5005066281, "facility electric"),
        (45, 5005066282, "facility water"),
        # These rows model the late reference-layer output from the reported
        # workbook: their physical positions are far away and have no source
        # metadata, but the account codes are the fixed facility contract.
        (80, 5006016260, "facility building depreciation"),
        (81, 5006016261, "facility land depreciation"),
    ]:
        ws.cell(row, 2).value = account
        ws.cell(row, 6).value = f"={row}*100"
        ws.cell(row, 19).value = description
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, clear_until_row=90)

    assert result["rows_written"] == 6
    assert result["preserved_rows_written"] == 0
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb[SHEET]
        assert [ws.cell(row, 19).value for row in range(38, 44)] == [
            "facility building depreciation",
            "facility land depreciation",
            "facility building interest",
            "facility land interest",
            "facility electric",
            "facility water",
        ]
        assert ws.cell(44, 19).value is None
        assert ws.cell(80, 19).value is None
        assert ws.cell(81, 19).value is None
        for row in range(38, 44):
            assert CANONICAL_SOURCE_FILE_ORDER[0] in _note_text(ws.cell(row, 20))
    finally:
        wb.close()


def test_complete_v1_writer_preserves_missing_separate_count_red_fill(tmp_path):
    workbook = tmp_path / "out_missing_count.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    wb.save(workbook)
    wb.close()

    periods = [
        "202604",
        "202605",
        "202606",
        "202607",
        "202608",
        "202609",
        "202610",
        "202611",
        "202612",
        "202701",
        "202702",
        "202703",
    ]
    result = apply_complete_v1_source_order_to_workbook(
        workbook,
        start_row=30,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "account_code": 600001,
                "description": "Alloc: separate count",
                "terms": {"202606": ["0*100"]},
                "highlight_periods": {"202606"},
            }
        ],
        fiscal_periods=periods,
    )

    assert result["rows_written"] == 1
    wb = load_workbook(workbook)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 8).value == "=0*100"
        assert (_fill_rgb(ws.cell(30, 8)) or "").endswith("FFC7CE")
    finally:
        wb.close()


def test_complete_v1_writer_sanitizes_visible_description_and_hides_wbs_metadata(tmp_path):
    path = tmp_path / "out_visible_clean.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    wb.save(path)
    wb.close()
    periods = [
        "202604",
        "202605",
        "202606",
        "202607",
        "202608",
        "202609",
        "202610",
        "202611",
        "202612",
        "202701",
        "202702",
        "202703",
    ]

    apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "account_code": 5004086291,
                "description": "Alloc: 社員旅行 Du lịch công ty|source_month=202605|driver_month=2026-05",
                "terms": {"202605": ["23*2061000"]},
            }
        ],
        fiscal_periods=periods,
    )
    result = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "account_code": 5004086291,
                "description": "Alloc: 社員旅行 Du lịch công ty|source_month=202605|driver_month=2026-05",
                "terms": {"202605": ["23*2061000"]},
            }
        ],
        fiscal_periods=periods,
    )

    assert result["rows_written"] == 1
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 19).value == "社員旅行 Du lịch công ty"
        assert ws.cell(30, 20).value is None
        assert ws.cell(30, 20).comment is None
        assert "source_file=" in _note_text(ws.cell(30, 20))
    finally:
        wb.close()


def test_complete_v1_writer_clears_ad_fill_and_column_e_item_ids_from_final_output(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for col in range(1, 5):
        ws.cell(36, col).fill = BLUE_FILL
    ws.cell(36, 2).value = 5006016260
    ws.cell(36, 5).value = "building_depreciation"
    ws.cell(36, 6).value = 100
    ws.cell(36, 19).value = "facility building depreciation"
    ws.cell(36, 20).value = "SOURCE_DERIVED"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=90)

    assert result["rows_written"] == 1
    assert result["layout_fills_cleared"] >= 4
    assert result["item_ids_cleared"] >= 1

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 2).value == 5006016260
        assert ws.cell(30, 6).value == 100
        assert ws.cell(30, 19).value == "facility building depreciation"
        assert ws.cell(30, 20).value is None
        assert CANONICAL_SOURCE_FILE_ORDER[0] in _note_text(ws.cell(30, 20))
        _assert_output_layout_clean(ws, 30, 36)
    finally:
        wb.close()


def test_complete_v1_writer_manages_bus_rows_in_admin_source_block(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(179, 2).value = 5004086291
    ws.cell(179, 6).value = 6111363
    ws.cell(179, 19).value = "system cost"
    ws.cell(53, 2).value = 5004086291
    ws.cell(53, 6).value = "=3*856107"
    ws.cell(53, 18).value = "=SUM(F53:Q53)"
    ws.cell(53, 19).value = "出向者BUS送迎費/Chi phí xe bus người JP"
    ws.cell(54, 2).value = 5004086291
    ws.cell(54, 6).value = "=20*1031546"
    ws.cell(54, 18).value = "=SUM(F54:Q54)"
    ws.cell(54, 19).value = "ローカル社BUS送迎費/Chi phí xe bus người VN"
    ws.cell(175, 2).value = 5004086291
    ws.cell(175, 6).value = 18895
    ws.cell(175, 19).value = "トイレットペーパー"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=199)

    assert result["rows_written"] == 4
    assert result["preserved_rows_written"] == 0
    assert result["blank_rows_written"] == 1

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 19).value == "system cost"
        assert ws.cell(31, 19).value is None
        assert ws.cell(32, 19).value == "出向者BUS送迎費/Chi phí xe bus người JP"
        assert ws.cell(32, 6).value == "=3*856107"
        assert ws.cell(32, 18).value == "=SUM(F32:Q32)"
        assert ws.cell(32, 20).value is None
        assert CANONICAL_SOURCE_FILE_ORDER[3] in _note_text(ws.cell(32, 20))
        assert "original_row=53" in _note_text(ws.cell(32, 20))
        assert ws.cell(33, 19).value == "ローカル社BUS送迎費/Chi phí xe bus người VN"
        assert ws.cell(33, 6).value == "=20*1031546"
        assert ws.cell(33, 18).value == "=SUM(F33:Q33)"
        assert "original_row=54" in _note_text(ws.cell(33, 20))
        assert ws.cell(34, 19).value == "トイレットペーパー"
        assert ws.cell(53, 19).value is None
        assert ws.cell(54, 19).value is None
    finally:
        wb.close()


def test_complete_v1_writer_keeps_source_rows_without_month_costs(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for row, account, description, amount in [
        (36, 5006016260, "facility building depreciation", 100),
        (175, 5004086291, "admin toilet paper", 200),
        (177, 5004086291, "admin alcohol placeholder", None),
        (59, 5004086291, "birthday", 300),
    ]:
        ws.cell(row, 2).value = account
        ws.cell(row, 5).value = "placeholder" if row == 177 else None
        ws.cell(row, 6).value = amount
        ws.cell(row, 19).value = description
        ws.cell(row, 20).value = "SOURCE_DERIVED"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=168, clear_until_row=190)

    assert result["rows_written"] == 4
    assert result["blank_rows_written"] == 2

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(168, 19).value == "facility building depreciation"
        assert ws.cell(169, 19).value is None
        assert ws.cell(170, 19).value == "admin toilet paper"
        assert ws.cell(171, 19).value == "admin alcohol placeholder"
        assert ws.cell(172, 19).value is None
        assert ws.cell(173, 19).value == "birthday"
        assert ws.cell(177, 19).value is None
        assert ws.cell(177, 5).value is None
    finally:
        wb.close()


def test_complete_v1_writer_reorders_existing_source_order_rows_idempotently(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(168, 6).value = 100
    ws.cell(168, 19).value = "facility building depreciation"
    ws.cell(168, 20).value = f"source_file={CANONICAL_SOURCE_FILE_ORDER[0]}; original_row=36; source-order-complete-v1"
    ws.cell(170, 5).value = "alcohol_disinfectant"
    ws.cell(170, 19).value = "admin alcohol placeholder"
    ws.cell(170, 20).value = f"source_file={CANONICAL_SOURCE_FILE_ORDER[3]}; original_row=177; source-order-complete-v1"
    ws.cell(171, 6).value = 300
    ws.cell(171, 19).value = "birthday"
    ws.cell(171, 20).value = f"source_file={CANONICAL_SOURCE_FILE_ORDER[4]}; original_row=59; source-order-complete-v1"
    for col in range(1, 5):
        ws.cell(170, col).fill = BLUE_FILL
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=168, clear_until_row=190)

    assert result["rows_written"] == 3
    assert result["blank_rows_written"] == 2

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(168, 19).value == "facility building depreciation"
        assert ws.cell(169, 19).value is None
        assert ws.cell(170, 19).value == "admin alcohol placeholder"
        assert ws.cell(171, 19).value is None
        assert ws.cell(172, 19).value == "birthday"
        assert ws.cell(172, 18).value == "=SUM(F172:Q172)"
        _assert_output_layout_clean(ws, 168, 172)
    finally:
        wb.close()


def test_complete_v1_writer_groups_split_accounts_inside_source_block(tmp_path):
    path = tmp_path / "out_split_accounts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    source_file = CANONICAL_SOURCE_FILE_ORDER[5]
    for row, account, original_row, description in [
        (30, 5004086291, 57, "welfare first"),
        (31, 5005246288, 97, "office supplies"),
        (32, 5004086291, 10001, "welfare second"),
    ]:
        ws.cell(row, 2).value = account
        ws.cell(row, 6).value = row
        ws.cell(row, 19).value = description
        ws.cell(row, 20).value = f"source_file={source_file}; original_row={original_row}; source-order-complete-v1"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=90)

    assert result["rows_written"] == 3
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert [ws.cell(row, 2).value for row in range(30, 33)] == [5004086291, 5004086291, 5005246288]
        assert [ws.cell(row, 19).value for row in range(30, 33)] == [
            "welfare first",
            "welfare second",
            "office supplies",
        ]
        assert ws.cell(30, 20).value is None
        assert "original_row=57" in _note_text(ws.cell(30, 20))
        assert "original_row=10001" in _note_text(ws.cell(31, 20))
        assert "original_row=97" in _note_text(ws.cell(32, 20))
    finally:
        wb.close()


def test_complete_v1_writer_groups_legacy_and_dynamic_allocation_accounts(tmp_path):
    path = tmp_path / "out_dynamic_split_accounts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(97, 2).value = 5005246288
    ws.cell(97, 7).value = "=1*9100"
    ws.cell(97, 18).value = "=SUM(F97:Q97)"
    ws.cell(97, 19).value = "legacy office supplies"
    ws.cell(97, 20).value = "SOURCE_DERIVED"
    wb.save(path)
    wb.close()

    periods = [
        "202604",
        "202605",
        "202606",
        "202607",
        "202608",
        "202609",
        "202610",
        "202611",
        "202612",
        "202701",
        "202702",
        "202703",
    ]
    result = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=120,
        dynamic_allocation_rows=[
            {
                "account_code": 5004086291,
                "description": "dynamic welfare",
                "terms": {"202607": ["22*100000"]},
            },
            {
                "account_code": 5005246288,
                "description": "dynamic office supplies",
                "terms": {"202608": ["2*3000"]},
            },
        ],
        fiscal_periods=periods,
    )

    assert result["rows_written"] == 3
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert [ws.cell(row, 2).value for row in range(30, 33)] == [5004086291, 5005246288, 5005246288]
        assert [ws.cell(row, 19).value for row in range(30, 33)] == [
            "dynamic welfare",
            "legacy office supplies",
            "dynamic office supplies",
        ]
        assert ws.cell(30, 20).value is None
        assert "original_row=10001" in _note_text(ws.cell(30, 20))
        assert "original_row=97" in _note_text(ws.cell(31, 20))
        assert "original_row=10002" in _note_text(ws.cell(32, 20))
    finally:
        wb.close()


def test_complete_v1_writer_drops_legacy_staged_row_when_dynamic_allocation_replaces_it(tmp_path):
    path = tmp_path / "out_dynamic_duplicate.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    description = "Alloc: 月餅 Bánh Trung Thu|source_month=202609|driver_month=2026-09|driver_type=headcount_all|driver_value=24"
    ws.cell(176, 2).value = 5004086291
    ws.cell(176, 11).value = "=24*56000"
    ws.cell(176, 18).value = "=SUM(F176:Q176)"
    ws.cell(176, 19).value = description
    ws.cell(176, 20).value = "SOURCE_DERIVED"
    wb.save(path)
    wb.close()

    periods = [
        "202604",
        "202605",
        "202606",
        "202607",
        "202608",
        "202609",
        "202610",
        "202611",
        "202612",
        "202701",
        "202702",
        "202703",
    ]
    result = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=190,
        dynamic_allocation_rows=[
            {
                "account_code": 5004086291,
                "description": description,
                "terms": {"202609": ["24*56000"]},
            },
        ],
        fiscal_periods=periods,
    )

    assert result["rows_written"] == 1
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        matching_rows = [
            row
            for row in range(30, 191)
            if ws.cell(row, 19).value == "月餅 Bánh Trung Thu"
        ]
        assert matching_rows == [30]
        assert ws.cell(30, 11).value == "=24*56000"
        assert ws.cell(30, 20).value is None
        assert CANONICAL_SOURCE_FILE_ORDER[5] in _note_text(ws.cell(30, 20))
        assert "original_row=10001" in _note_text(ws.cell(30, 20))
    finally:
        wb.close()


def test_complete_v1_writer_replaces_blank_and_stale_nnn_source_block_idempotently(tmp_path):
    path = tmp_path / "out_dynamic_nnn.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(137, 10).value = "=1820000+130000+50000"
    ws.cell(137, 12).value = "=1000000+1000000+50000+400000"
    ws.cell(137, 19).value = "legacy merged NNN"
    wb.save(path)
    wb.close()

    periods = ["202604", "202605", "202606", "202607", "202608", "202609", "202610", "202611", "202612", "202701", "202702", "202703"]
    dynamic_rows = [
        {
            "source_group_index": 6,
            "account_code": 5005246286,
            "description": "NNN paperwork: Worker A",
            "terms": {"202608": ["1820000+130000+50000"]},
        },
        {
            "source_group_index": 6,
            "account_code": 5005246286,
            "description": "NNN paperwork: Worker B",
            "terms": {"202610": ["1000000+1000000+50000+400000"]},
        },
    ]

    first = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=190,
        dynamic_allocation_rows=dynamic_rows,
        fiscal_periods=periods,
    )

    # Simulate an artifact created before account canonicalization changed:
    # the managed NNN row has the same source/description but a stale account.
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        stale_row = next(
            row
            for row in range(30, 191)
            if ws.cell(row, 19).value == "NNN paperwork: Worker B"
        )
        ws.cell(stale_row, 2).value = 6005246673
        wb.save(path)
    finally:
        wb.close()

    second = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=190,
        dynamic_allocation_rows=dynamic_rows,
        fiscal_periods=periods,
    )

    assert first["nnn_rows_written"] == 2
    assert second["nnn_rows_written"] == 2
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        nnn_rows = [
            row
            for row in range(30, 191)
            if CANONICAL_SOURCE_FILE_ORDER[6] in _note_text(ws.cell(row, 20))
        ]
        assert len(nnn_rows) == 2
        assert {ws.cell(row, 2).value for row in nnn_rows} == {5005246286}
        assert all(ws.cell(row, 2).value for row in nnn_rows)
        rows_by_description = {ws.cell(row, 19).value: row for row in nnn_rows}
        assert ws.cell(rows_by_description["NNN paperwork: Worker A"], 10).value == "=1820000+130000+50000"
        assert ws.cell(rows_by_description["NNN paperwork: Worker B"], 12).value == "=1000000+1000000+50000+400000"
        assert all(ws.cell(row, 19).value != "legacy merged NNN" for row in range(30, 191))
    finally:
        wb.close()


def test_complete_v1_writer_replaces_legacy_asset_totals_but_preserves_manual_asset_row(tmp_path):
    path = tmp_path / "out_fixed_assets_dynamic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for row, description in ((38, "legacy depreciation total"), (42, "legacy interest total")):
        ws.cell(row, 2).value = 5006016244 if row == 38 else 9114120007
        ws.cell(row, 6).value = 999
        ws.cell(row, 19).value = description
    ws.cell(50, 2).value = 5006016242
    ws.cell(50, 6).value = 123
    ws.cell(50, 19).value = "manual planned machine"
    ws.cell(50, 20).value = (
        f"source_file={CANONICAL_SOURCE_FILE_ORDER[1]}; original_row=500; source-order-complete-v1"
    )
    wb.save(path)
    wb.close()

    periods = ["202604", "202605", "202606", "202607", "202608", "202609", "202610", "202611", "202612", "202701", "202702", "202703"]
    result = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "source_file": CANONICAL_SOURCE_FILE_ORDER[1],
                "account_code": 5006016242,
                "description": "Khấu hao tài sản cố định - Machinery and Equipment",
                "terms": {"202604": ["ROUND(12.75*$B$2,0)"]},
            },
            {
                "source_file": CANONICAL_SOURCE_FILE_ORDER[1],
                "account_code": 9114120007,
                "description": "Lãi tài sản cố định - Machinery and Equipment",
                "terms": {"202604": ["ROUND(1.5*$B$2,0)"]},
            },
        ],
        fiscal_periods=periods,
    )

    assert result["rows_written"] == 3
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        descriptions = [ws.cell(row, 19).value for row in range(30, 33)]
        assert "manual planned machine" in descriptions
        assert "Khấu hao tài sản cố định - Machinery and Equipment" in descriptions
        assert "Lãi tài sản cố định - Machinery and Equipment" in descriptions
        assert "legacy depreciation total" not in descriptions
        assert "legacy interest total" not in descriptions
        depreciation_row = descriptions.index("Khấu hao tài sản cố định - Machinery and Equipment") + 30
        assert ws.cell(depreciation_row, 6).value == "=ROUND(12.75*$B$2,0)"
        assert ws.cell(depreciation_row, 18).value == f"=SUM(F{depreciation_row}:Q{depreciation_row})"
    finally:
        wb.close()


def test_complete_v1_writer_keeps_explicit_zero_distinct_from_post_terminal_blank(tmp_path):
    path = _workbook(tmp_path / "out.xlsx")
    periods = ["202604", "202605", "202606"]
    apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "source_file": CANONICAL_SOURCE_FILE_ORDER[1],
                "account_code": 5006016242,
                "description": "fixed asset explicit terminal zero",
                "terms": {"202604": ["ROUND(10*$B$2,0)"]},
                "explicit_zero_periods": {"202605"},
                "audit_trail": "fixed_assets_audit_table=audit_fixed_asset_import_rows; fiscal_year=2027",
            }
        ],
        fiscal_periods=periods,
    )
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        row = next(row for row in range(30, 90) if ws.cell(row, 19).value == "fixed asset explicit terminal zero")
        assert ws.cell(row, 6).value == "=ROUND(10*$B$2,0)"
        assert ws.cell(row, 7).value == 0
        assert ws.cell(row, 8).value is None
        assert "fixed_assets_audit_table" in _note_text(ws.cell(row, 20))
    finally:
        wb.close()


def test_complete_v1_writer_does_not_leak_legacy_row_58_red_fill(tmp_path):
    workbook = tmp_path / "out_row_58_style_leak.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for col in range(6, 18):
        ws.cell(58, col).fill = BLUE_FILL
    # These alerts belonged to the old business rule staged on physical row 58.
    ws.cell(58, 7).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    ws.cell(58, 10).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    wb.save(workbook)
    wb.close()

    periods = [
        "202604", "202605", "202606", "202607", "202608", "202609",
        "202610", "202611", "202612", "202701", "202702", "202703",
    ]
    apply_complete_v1_source_order_to_workbook(
        workbook,
        start_row=58,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "account_code": 600001,
                "description": "new item occupying physical row 58",
                "terms": {"202604": ["10*100"]},
                "highlight_periods": set(),
            }
        ],
        fiscal_periods=periods,
    )

    wb = load_workbook(workbook)
    try:
        ws = wb[SHEET]
        assert all((_fill_rgb(ws.cell(58, col)) or "").endswith("CCFFFF") for col in range(6, 18))
    finally:
        wb.close()


def test_complete_v1_writer_reapplies_only_current_rows_red_alerts(tmp_path):
    workbook = tmp_path / "out_row_58_current_alert.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for col in range(6, 18):
        ws.cell(58, col).fill = BLUE_FILL
    ws.cell(58, 7).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    ws.cell(58, 10).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    wb.save(workbook)
    wb.close()

    periods = [
        "202604", "202605", "202606", "202607", "202608", "202609",
        "202610", "202611", "202612", "202701", "202702", "202703",
    ]
    apply_complete_v1_source_order_to_workbook(
        workbook,
        start_row=58,
        clear_until_row=90,
        dynamic_allocation_rows=[
            {
                "account_code": 600001,
                "description": "current separate-count item",
                "terms": {"202606": ["0*100"]},
                "highlight_periods": {"202606"},
            }
        ],
        fiscal_periods=periods,
    )

    wb = load_workbook(workbook)
    try:
        ws = wb[SHEET]
        assert (_fill_rgb(ws["H58"]) or "").endswith("FFC7CE")
        assert all(
            (_fill_rgb(ws.cell(58, col)) or "").endswith("CCFFFF")
            for col in range(6, 18)
            if col != 8
        )
    finally:
        wb.close()


def test_complete_v1_writer_rejects_formula_beyond_excel_limit(tmp_path):
    path = _workbook(tmp_path / "out.xlsx")
    periods = ["202604"]
    oversized_terms = ["1"] * 4097

    with pytest.raises(ValueError, match="8.192"):
        apply_complete_v1_source_order_to_workbook(
            path,
            start_row=30,
            clear_until_row=90,
            dynamic_allocation_rows=[
                {
                    "source_file": CANONICAL_SOURCE_FILE_ORDER[1],
                    "account_code": 5006016242,
                    "description": "oversized fixed asset formula",
                    "terms": {"202604": oversized_terms},
                }
            ],
            fiscal_periods=periods,
        )


def test_complete_v1_writer_preserves_unmanaged_business_rows_inside_clear_range(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(36, 6).value = 100
    ws.cell(36, 19).value = "facility building depreciation"
    ws.cell(52, 2).value = 5000000052
    ws.cell(52, 6).value = "=$B52*100"
    ws.cell(52, 18).value = "=SUM(F52:Q52)"
    ws.cell(52, 19).value = "preserved legacy row"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=90)

    assert result["rows_written"] == 1
    assert result["preserved_rows_written"] == 1

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 19).value == "facility building depreciation"
        assert ws.cell(32, 19).value == "preserved legacy row"
        assert ws.cell(32, 6).value == "=$B32*100"
        assert ws.cell(32, 18).value == "=SUM(F32:Q32)"
        assert ws.cell(52, 19).value is None
    finally:
        wb.close()


def test_complete_v1_writer_clears_generated_allocations_below_legacy_row_199(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(36, 2).value = 5006016260
    ws.cell(36, 6).value = 100
    ws.cell(36, 19).value = "facility building depreciation"
    ws.cell(205, 2).value = 5005246288
    ws.cell(205, 6).value = "=4*3000"
    ws.cell(205, 18).value = "=SUM(F205:Q205)"
    ws.cell(205, 19).value = "Alloc: Pen"
    # The reusable FORM contains lookup/total formulas far below the visible
    # output.  They must not prevent cleanup of the actual append payload.
    ws.cell(1015, 3).value = '=IFERROR(VLOOKUP($B1015,A:B,2,0),"")'
    ws.cell(1015, 18).value = "=SUM(F1015:Q1015)"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(
        path,
        start_row=30,
        clear_until_row=199,
        dynamic_allocation_rows=[{
            "account_code": 5005246288,
            "description": "Alloc: Pen",
            "terms": {"202604": ["4*3000"]},
        }],
        fiscal_periods=["202604"],
    )

    assert result["cleared_through_row"] == 205
    assert result["tail_formulas_cleared"] == 2
    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        descriptions = [ws.cell(row, 19).value for row in range(30, 1016)]
        assert descriptions.count("Pen") == 1
        assert ws.cell(205, 2).value is None
        assert ws.cell(205, 6).value is None
        assert ws.cell(205, 19).value is None
        assert ws.cell(1015, 3).value is None
        assert ws.cell(1015, 18).value is None
    finally:
        wb.close()


def test_complete_v1_writer_does_not_preserve_generated_file_order_duplicates(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(36, 2).value = 5006016260
    ws.cell(36, 6).value = "=ROUND(244.09*$B$2,0)"
    ws.cell(36, 19).value = "facility building depreciation"
    ws.cell(52, 2).value = 5004086291
    ws.cell(52, 6).value = "=ROUND(244.09*$B$2,0)"
    ws.cell(52, 18).value = "=SUM(F52:Q52)"
    ws.cell(52, 19).value = "Khấu hao nhà"
    ws.cell(52, 20).value = "ROUND_USD_BY_B2"
    ws.cell(64, 2).value = 5004086291
    ws.cell(64, 6).value = 932009
    ws.cell(64, 18).value = "=SUM(F64:Q64)"
    ws.cell(64, 19).value = "Điện"
    ws.cell(64, 20).value = "COPY_VND_MONTHLY"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=90)

    assert result["rows_written"] == 1
    assert result["preserved_rows_written"] == 0

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        descriptions = [ws.cell(row, 19).value for row in range(30, 60)]
        assert "facility building depreciation" in descriptions
        assert "Khấu hao nhà" not in descriptions
        assert "Điện" not in descriptions
    finally:
        wb.close()


def test_complete_v1_writer_drops_generated_source_order_summary_rows_on_rerun(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(30, 2).value = 5004086291
    ws.cell(30, 6).value = 120399176
    ws.cell(30, 18).value = "=SUM(F30:Q30)"
    ws.cell(30, 19).value = "System Cost / システム課金"
    ws.cell(30, 20).value = (
        "COPY_SUMMARY_VND_TOTAL_BY_PERIOD | "
        f"source_file={CANONICAL_SOURCE_FILE_ORDER[2]}; original_row=179; source-order-complete-v1"
    )
    ws.cell(75, 2).value = 5005246282
    ws.cell(75, 6).value = "=ROUND((10*3)*$B$2,0)"
    ws.cell(75, 18).value = "=SUM(F75:Q75)"
    ws.cell(75, 19).value = "System Cost (Mail,VPN,R3, Mes,PLM,VPS,...)"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=199)

    assert result["rows_written"] == 0
    assert result["preserved_rows_written"] == 1

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 2).value == 5005246282
        assert ws.cell(30, 19).value == "System Cost (Mail,VPN,R3, Mes,PLM,VPS,...)"
        descriptions = [ws.cell(row, 19).value for row in range(30, 80)]
        assert "System Cost / システム課金" not in descriptions
    finally:
        wb.close()


def test_complete_v1_writer_fills_lookup_formulas_for_preserved_cost_rows(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(89, 2).value = 5005246288
    ws.cell(89, 6).value = "=4*3000"
    ws.cell(89, 18).value = "=SUM(F89:Q89)"
    ws.cell(89, 19).value = "ペン Bút"
    ws.cell(89, 20).value = "allocation_rule_row=89; exact_identity=ペン Bút"
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=199)

    assert result["rows_written"] == 0
    assert result["preserved_rows_written"] == 1

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 2).value == 5005246288
        assert ws.cell(30, 3).value.startswith("=IFERROR(IF(VLOOKUP($B30,")
        assert "勘定科目" in ws.cell(30, 3).value
        assert ws.cell(30, 4).value == '=IF(C30="","",VLOOKUP($B30,勘定科目!$A:$E,4,0))'
        assert ws.cell(30, 6).value == "=4*3000"
        assert ws.cell(30, 18).value == "=SUM(F30:Q30)"
    finally:
        wb.close()


def test_complete_v1_writer_drops_account_only_template_rows(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(36, 2).value = 5006016260
    ws.cell(36, 6).value = 100
    ws.cell(36, 19).value = "facility building depreciation"
    for row in (61, 62, 63):
        ws.cell(row, 2).value = 5004086291
        ws.cell(row, 3).value = f'=IFERROR(VLOOKUP($B{row},勘定科目!$A:$E,2,0),"")'
        ws.cell(row, 4).value = f'=IF(C{row}="","",VLOOKUP($B{row},勘定科目!$A:$E,4,0))'
    wb.save(path)
    wb.close()

    result = apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=90)

    assert result["rows_written"] == 1
    assert result["preserved_rows_written"] == 0

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        assert ws.cell(30, 19).value == "facility building depreciation"
        for row in (31, 32, 33):
            assert ws.cell(row, 2).value is None
            assert ws.cell(row, 19).value is None
    finally:
        wb.close()


def test_complete_v1_writer_does_not_duplicate_source_order_marker_on_rerun(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(168, 6).value = 100
    ws.cell(168, 19).value = "facility building depreciation"
    ws.cell(168, 20).value = (
        "COPY_SOURCE_MONTH_SAMPLE | "
        f"source_file={CANONICAL_SOURCE_FILE_ORDER[0]}; original_row=36; source-order-complete-v1"
    )
    wb.save(path)
    wb.close()

    apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=190)
    apply_complete_v1_source_order_to_workbook(path, start_row=30, clear_until_row=190)

    wb = load_workbook(path)
    try:
        ws = wb[SHEET]
        # COPY_SOURCE_MONTH_SAMPLE rows are now treated as generated file-order
        # policy rows and are skipped (cleared) during source-order re-layout.
        # This prevents phantom admin consumable rows from being preserved
        # without account codes under incorrect parent accounts.
        note = ws.cell(30, 20).value
        assert note is None or "COPY_SOURCE_MONTH_SAMPLE" not in str(note)
    finally:
        wb.close()


def test_run_e2e_complete_v1_final_source_order_writer_runs_after_reference_fill():
    text = Path("scripts/run_e2e.py").read_text(encoding="utf-8")

    assert "source_file_order=_annual_complete_v1_source_order(run_context)" in text
    assert "phase=\"final\"" in text
    assert "dynamic_allocation_rows=complete_v1_dynamic_allocation_rows" in text
    assert "dynamic_allocation_rows=complete_v1_dynamic_rows" in text
    assert "fiscal_periods=complete_v1_fiscal_periods" in text
    assert "fiscal_periods=complete_v1_periods" in text
    assert text.index("phase=\"pre-reference\"") < text.index("apply_mp_saisan_complete_v1(")
    assert text.index("apply_mp_saisan_complete_v1(") < text.index("phase=\"final\"")
