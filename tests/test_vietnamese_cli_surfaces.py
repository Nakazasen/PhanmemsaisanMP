from __future__ import annotations

import io
import json
import sys
from zipfile import BadZipFile

import pytest

from scripts import generate_quality_baseline as quality_baseline
from scripts import verify_fixed_assets_handover as handover
from src.utils.cli import VietnameseArgumentParser
from tools import compare_primary_reference as comparison
from tools import create_facility_file_order_preview as facility_preview
from tools import safe_workbook_inventory as inventory


ENGLISH_HELP_PHRASES = (
    "usage:",
    "options:",
    "positional arguments:",
    "show this help message and exit",
)


def test_vietnamese_argument_parser_localizes_help_and_keeps_flags():
    parser = VietnameseArgumentParser(prog="kiem-tra", description="Kiểm tra dữ liệu")
    parser.add_argument("--input", required=True, help="Tệp đầu vào")
    help_text = parser.format_help()

    assert "cách dùng:" in help_text
    assert "tùy chọn:" in help_text
    assert "hiển thị trợ giúp này và thoát" in help_text
    assert "--input" in help_text
    assert not any(phrase in help_text for phrase in ENGLISH_HELP_PHRASES)


def test_vietnamese_argument_parser_localizes_error_and_keeps_exit_code(capsys):
    parser = VietnameseArgumentParser(prog="kiem-tra")
    parser.add_argument("--input", required=True)

    with pytest.raises(SystemExit) as raised:
        parser.parse_args([])

    stderr = capsys.readouterr().err
    assert raised.value.code == 2
    assert "cách dùng:" in stderr
    assert "lỗi:" in stderr
    assert "bắt buộc phải có các đối số: --input" in stderr
    assert "the following arguments are required" not in stderr


def test_vietnamese_help_does_not_crash_on_cp932_console(monkeypatch):
    stdout_bytes = io.BytesIO()
    stderr_bytes = io.BytesIO()
    stdout = io.TextIOWrapper(stdout_bytes, encoding="cp932")
    stderr = io.TextIOWrapper(stderr_bytes, encoding="cp932")
    monkeypatch.setattr(sys, "stdout", stdout)
    monkeypatch.setattr(sys, "stderr", stderr)

    parser = VietnameseArgumentParser(prog="kiem-tra", description="Kiểm tra tiếng Việt")
    parser.add_argument("--input", help="Đường dẫn đầu vào")
    parser.print_help()
    stdout.flush()

    assert stdout.encoding.lower().replace("-", "") == "utf8"
    assert "Kiểm tra tiếng Việt" in stdout_bytes.getvalue().decode("utf-8")


def test_inventory_child_does_not_leak_raw_exception(monkeypatch, capsys, tmp_path):
    secret = "English library exception must stay private"
    workbook = tmp_path / "broken.xlsx"
    monkeypatch.setattr(sys, "argv", ["inventory", "--child-scan", str(workbook)])
    monkeypatch.setattr(inventory, "_scan_one", lambda _path: (_ for _ in ()).throw(RuntimeError(secret)))

    assert inventory.child_main() == 0
    payload = json.loads(capsys.readouterr().out)
    assert payload["read_status"] == "READ_ERROR:WORKBOOK_UNREADABLE"
    assert secret not in json.dumps(payload, ensure_ascii=False)


def test_facility_preview_does_not_leak_raw_exception(monkeypatch, capsys, tmp_path):
    secret = "English writer exception must stay private"
    monkeypatch.setattr(
        facility_preview,
        "write_facility_file_order_preview_workbook",
        lambda **_kwargs: (_ for _ in ()).throw(RuntimeError(secret)),
    )

    exit_code = facility_preview.main([
        "--template", str(tmp_path / "FORM.xlsx"),
        "--facility-source", str(tmp_path / "facility.xlsx"),
        "--output", str(tmp_path / "preview.xlsx"),
    ])

    stderr = capsys.readouterr().err
    assert exit_code == 1
    assert "Không thể tạo workbook xem trước Facility" in stderr
    assert secret not in stderr


def test_comparison_cli_does_not_leak_raw_exception(monkeypatch, capsys, tmp_path):
    secret = "English comparison exception must stay private"
    monkeypatch.setattr(
        comparison,
        "compare_workbooks",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError(secret)),
    )

    exit_code = comparison.main([
        "--generated", str(tmp_path / "generated.xlsx"),
        "--reference", str(tmp_path / "reference.xlsx"),
        "--out-dir", str(tmp_path / "reports"),
    ])

    stderr = capsys.readouterr().err
    assert exit_code == 1
    assert "Không thể tạo báo cáo so sánh" in stderr
    assert secret not in stderr


def test_handover_bad_zip_does_not_leak_raw_exception(monkeypatch, capsys, tmp_path):
    secret = "English zip exception must stay private"
    archive_path = tmp_path / "FY2027.zip"
    archive_path.write_bytes(b"not-a-zip")
    empty_fy2026 = tmp_path / "FY2026"
    empty_fy2026.mkdir()

    class BrokenZip:
        def __init__(self, *_args, **_kwargs):
            raise BadZipFile(secret)

    monkeypatch.setattr(handover, "REQUIRED_HASHES", {})
    monkeypatch.setattr(handover, "FY2026_DIR", empty_fy2026)
    monkeypatch.setattr(handover, "EXPECTED_FY2026_XLSX", 0)
    monkeypatch.setattr(handover, "FY2027_ZIP", archive_path)
    monkeypatch.setattr(handover, "ZipFile", BrokenZip)

    assert handover.verify(False) == 1
    stderr = capsys.readouterr().err
    assert "không hợp lệ hoặc chứa thành phần không an toàn" in stderr
    assert secret not in stderr


def test_quality_markdown_is_vietnamese_but_schema_keys_stay_stable():
    report = {
        "generated_at": "2026-01-01T00:00:00+00:00",
        "summary": {
            "python_files": 1,
            "python_lines": 2,
            "test_count": 3,
            "findings_by_severity": {"critical": 0, "high": 0, "medium": 0, "low": 0},
        },
        "findings": [],
    }
    text = quality_baseline.markdown(report)

    assert "# Kiểm toán kỹ thuật hiện tại" in text
    assert "## Tóm tắt" in text
    assert "## Tạo lại báo cáo" in text
    assert "Summary" not in text
    assert "Packaging decision" not in text
    assert set(report) == {"generated_at", "summary", "findings"}


def test_comparison_summary_keeps_schema_but_localizes_terminal_and_excel(monkeypatch, capsys, tmp_path):
    from openpyxl import load_workbook

    summary = {
        "generated_path": "ket-qua.xlsx",
        "reference_path": "tham-chieu.xlsx",
        "compare_mode": "strict_exact",
        "differences": 2,
    }
    original_keys = set(summary)
    display = comparison._display_summary(summary)

    assert set(summary) == original_keys
    assert "generated_path" not in display
    assert display["Đường dẫn tệp kết quả"] == "ket-qua.xlsx"
    assert display["Chế độ so sánh"] == "so sánh chính xác"

    report_path = tmp_path / "comparison.xlsx"
    comparison._write_excel_report(report_path, summary, [], [], [])
    workbook = load_workbook(report_path, data_only=False)
    try:
        sheet = workbook["Tóm_tắt"]
        labels = {sheet.cell(row=row, column=1).value for row in range(1, sheet.max_row + 1)}
        assert "Đường dẫn tệp kết quả" in labels
        assert "generated_path" not in labels
    finally:
        workbook.close()

    monkeypatch.setattr(
        comparison,
        "compare_workbooks",
        lambda *_args, **_kwargs: {
            "summary": summary,
            "xlsx_path": str(report_path),
            "json_path": str(tmp_path / "comparison.json"),
        },
    )
    assert comparison.main([
        "--generated", "ket-qua.xlsx",
        "--reference", "tham-chieu.xlsx",
        "--out-dir", str(tmp_path),
    ]) == 0
    output = capsys.readouterr().out
    assert "Đường dẫn tệp kết quả" in output
    assert '"generated_path"' not in output
