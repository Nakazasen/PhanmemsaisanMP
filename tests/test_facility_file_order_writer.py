from pathlib import Path

from openpyxl import load_workbook

from src.engine.facility_file_order_writer import write_facility_file_order_preview_workbook
from src.utils import excel_helpers as helpers

TEMPLATE = Path("FORM.xlsx")
FACILITY_SOURCE = Path("raw/施設課　MPFY2027.xlsx")


def _hub_sheet(workbook):
    return workbook[helpers.find_hub_sheet_name(workbook)]


def test_writer_creates_preview_workbook_without_touching_template(tmp_path):
    output_path = tmp_path / "facility_preview.xlsx"
    before_template_mtime = TEMPLATE.stat().st_mtime_ns
    before_source_mtime = FACILITY_SOURCE.stat().st_mtime_ns

    result = write_facility_file_order_preview_workbook(TEMPLATE, FACILITY_SOURCE, output_path)

    assert result == output_path
    assert output_path.exists()
    assert TEMPLATE.stat().st_mtime_ns == before_template_mtime
    assert FACILITY_SOURCE.stat().st_mtime_ns == before_source_mtime


def test_writer_writes_facility_rows_200_to_205(tmp_path):
    output_path = tmp_path / "facility_preview.xlsx"
    write_facility_file_order_preview_workbook(TEMPLATE, FACILITY_SOURCE, output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = _hub_sheet(workbook)
        assert [sheet.cell(row=row, column=5).value for row in range(200, 206)] == [
            "building_depreciation",
            "land_depreciation",
            "building_interest",
            "land_interest",
            "electricity",
            "water",
        ]
        expected_months_by_row = {
            200: ["=ROUND(293.02*$B$2,0)", "=ROUND(294.07*$B$2,0)", "=ROUND(296.68*$B$2,0)", "=ROUND(293.7*$B$2,0)", "=ROUND(287.07*$B$2,0)", "=ROUND(289.74*$B$2,0)", "=ROUND(294.29*$B$2,0)", "=ROUND(292.36*$B$2,0)", "=ROUND(291.13*$B$2,0)", "=ROUND(285.32*$B$2,0)", "=ROUND(273.97*$B$2,0)", "=ROUND(289.03*$B$2,0)"],
            201: ["=ROUND(19.78*$B$2,0)", "=ROUND(19.85*$B$2,0)", "=ROUND(20.03*$B$2,0)", "=ROUND(19.82*$B$2,0)", "=ROUND(19.38*$B$2,0)", "=ROUND(19.56*$B$2,0)", "=ROUND(19.86*$B$2,0)", "=ROUND(19.73*$B$2,0)", "=ROUND(19.65*$B$2,0)", "=ROUND(19.26*$B$2,0)", "=ROUND(18.49*$B$2,0)", "=ROUND(19.51*$B$2,0)"],
            202: ["=ROUND(188.35*$B$2,0)", "=ROUND(178.43*$B$2,0)", "=ROUND(180.02*$B$2,0)", "=ROUND(178.21*$B$2,0)", "=ROUND(174.19*$B$2,0)", "=ROUND(175.81*$B$2,0)", "=ROUND(178.57*$B$2,0)", "=ROUND(177.4*$B$2,0)", "=ROUND(176.65*$B$2,0)", "=ROUND(173.13*$B$2,0)", "=ROUND(166.24*$B$2,0)", "=ROUND(175.38*$B$2,0)"],
            203: ["=ROUND(23.88*$B$2,0)", "=ROUND(23.25*$B$2,0)", "=ROUND(23.46*$B$2,0)", "=ROUND(23.22*$B$2,0)", "=ROUND(22.7*$B$2,0)", "=ROUND(22.91*$B$2,0)", "=ROUND(23.27*$B$2,0)", "=ROUND(23.12*$B$2,0)", "=ROUND(23.02*$B$2,0)", "=ROUND(22.56*$B$2,0)", "=ROUND(21.66*$B$2,0)", "=ROUND(22.85*$B$2,0)"],
            204: [1577160, 2053287, 2791574, 3290266, 3419370, 3270194, 2959172, 2460327, 1758044, 1450135, 1247404, 1203521],
            205: [529671, 577177, 682783, 1015039, 859650, 874634, 861265, 646887, 635220, 566871, 438313, 637935],
        }
        for row_index, expected_months in expected_months_by_row.items():
            assert [sheet.cell(row=row_index, column=column_index).value for column_index in range(6, 18)] == expected_months
    finally:
        workbook.close()


def test_writer_leaves_blank_row_206(tmp_path):
    output_path = tmp_path / "facility_preview.xlsx"
    write_facility_file_order_preview_workbook(TEMPLATE, FACILITY_SOURCE, output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = _hub_sheet(workbook)
        assert sheet.cell(row=206, column=5).value is None
        assert sheet.cell(row=206, column=6).value is None
        assert sheet.cell(row=206, column=17).value is None
        assert sheet.cell(row=206, column=19).value is None
    finally:
        workbook.close()


def test_writer_contains_six_facility_items(tmp_path):
    output_path = tmp_path / "facility_preview.xlsx"
    write_facility_file_order_preview_workbook(TEMPLATE, FACILITY_SOURCE, output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = _hub_sheet(workbook)
        descriptions = [sheet.cell(row=row, column=19).value for row in range(200, 206)]
        assert descriptions == ["Khấu hao nhà", "Khấu hao đất", "Lãi nhà", "Lãi đất", "Điện", "Nước"]
        assert [sheet.cell(row=row, column=20).value for row in range(200, 206)] == [
            "ROUND_USD_BY_B2",
            "ROUND_USD_BY_B2",
            "ROUND_USD_BY_B2",
            "ROUND_USD_BY_B2",
            "COPY_VND_MONTHLY",
            "COPY_VND_MONTHLY",
        ]
    finally:
        workbook.close()


def test_default_hub_builder_export_flow_not_changed():
    hub_builder = Path("src/engine/hub_builder.py").read_text(encoding="utf-8")

    assert "write_facility_file_order_preview_workbook" not in hub_builder
    assert "facility_file_order_writer" not in hub_builder
    assert "self._write_fixed_rows(worksheet, target_cc)" in hub_builder


def test_writer_requires_explicit_output_path(tmp_path):
    try:
        write_facility_file_order_preview_workbook(TEMPLATE, FACILITY_SOURCE, "")
    except ValueError as exc:
        assert "output_path is required" in str(exc)
    else:
        raise AssertionError("Expected output_path validation failure")

    try:
        write_facility_file_order_preview_workbook(TEMPLATE, FACILITY_SOURCE, TEMPLATE)
    except ValueError as exc:
        assert "must not overwrite the template" in str(exc)
    else:
        raise AssertionError("Expected template overwrite validation failure")
