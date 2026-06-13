import csv
import sqlite3
import unittest
import warnings
from pathlib import Path
import shutil
import uuid

import openpyxl

from src.audit.pipeline_audit import write_pipeline_audit_report
from src.db.loader import _apply_mp2026_reference_unit_price, _parse_unit_price, load_allocation_rules
from src.db.schema import create_schema, init_sys_params
from src.engine.allocator import AllocationEngine
from src.engine.complete_v1_source_order_writer import apply_complete_v1_source_order_to_workbook
from src.engine.hub_builder import ExportIntegrityError, HubBuilder
from src.parsers.birthday import parse_birthday_workbook
from src.parsers.fixed_assets import parse_fixed_assets
from src.parsers.ga import parse_ga
from src.parsers.manual_event_drivers import parse_manual_event_drivers
from src.parsers.manual_headcount import (
    ensure_manual_headcount_template,
    get_required_headcount_periods,
    LEGACY_HEADCOUNT_SOURCE_WARNING,
    parse_manual_headcount,
    quarantine_manual_headcount_rows,
    resolve_manual_headcount_source_dir,
)
from src.parsers.manual_special_costs import parse_manual_special_costs
from src.parsers.nnn_paperwork import parse_nnn_paperwork
from src.utils.excel_helpers import find_hub_sheet_name, get_fy_months


TEST_TMP_ROOT = Path.cwd() / ".tmp_test_artifacts"
TEST_TMP_ROOT.mkdir(exist_ok=True)


def _mk_tmpdir() -> Path:
    path = TEST_TMP_ROOT / f"case_{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    return path


def _mk_conn():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    create_schema(conn)
    init_sys_params(conn, exchange_rate=26273, fiscal_year=2027)
    return conn


def _seed_cc(conn, code=1412000004, cost_type="一般"):
    conn.execute(
        """
        INSERT INTO dim_cost_centers
        (code, name_jp, seq_no, saisan_type, cost_type, staff_count, worker_count)
        VALUES (?, 'TEST_CC', 1, '製造', ?, 0, 0)
        """,
        (code, cost_type),
    )
    conn.commit()
    return code


def _find_system_cost_rows(ws):
    rows = []
    for row_index in range(1, ws.max_row + 1):
        row_text = " ".join(
            str(ws.cell(row_index, column_index).value).lower()
            for column_index in range(1, 21)
            if ws.cell(row_index, column_index).value is not None
        )
        if "system cost" in row_text:
            rows.append(row_index)
    return rows


class TestExportIntegrityGuard(unittest.TestCase):
    def test_malformed_one_cell_template_fails_without_overwriting_output(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            malformed_template = tmpdir / "FORM_1x1.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "内訳ﾘｽﾄ(4～3月)"
            workbook.save(malformed_template)
            workbook.close()

            conn.execute(
                """
                INSERT INTO fact_input_data
                (source, period, amount_vnd, cc_code, account_code, description)
                VALUES ('manual_special', '202604', 1000, ?, 5005136291, 'manual cost')
                """,
                (cc_code,),
            )
            conn.commit()

            output_path = tmpdir / "MP_CC_bad.xlsx"
            output_path.write_text("existing output", encoding="utf-8")

            with self.assertRaises(ExportIntegrityError):
                HubBuilder(conn, fiscal_year=2027).export_to_template(
                    str(malformed_template),
                    str(output_path),
                    cc_code=cc_code,
                )

            self.assertEqual(output_path.read_text(encoding="utf-8"), "existing output")
            self.assertFalse(output_path.with_name(f"{output_path.stem}.tmp_export{output_path.suffix}").exists())
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_missing_template_fails_without_creating_output(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            conn.execute(
                """
                INSERT INTO fact_input_data
                (source, period, amount_vnd, cc_code, account_code, description)
                VALUES ('manual_special', '202604', 1000, ?, 5005136291, 'manual cost')
                """,
                (cc_code,),
            )
            conn.commit()

            output_path = tmpdir / "MP_CC_missing_template.xlsx"
            with self.assertRaises(FileNotFoundError):
                HubBuilder(conn, fiscal_year=2027).export_to_template(
                    str(tmpdir / "missing_FORM.xlsx"),
                    str(output_path),
                    cc_code=cc_code,
                )

            self.assertFalse(output_path.exists())
            self.assertFalse(output_path.with_name(f"{output_path.stem}.tmp_export{output_path.suffix}").exists())
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_successful_export_has_business_rows(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            conn.execute(
                """
                INSERT INTO fact_input_data
                (source, period, amount_vnd, cc_code, account_code, description)
                VALUES ('manual_special', '202604', 1000, ?, 5005136291, 'manual cost')
                """,
                (cc_code,),
            )
            conn.commit()

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "MP_CC_good.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(
                str(template_path),
                str(output_path),
                cc_code=cc_code,
            )

            self.assertTrue(ok)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                worksheet = workbook[find_hub_sheet_name(workbook)]
                business_rows = HubBuilder(conn, fiscal_year=2027)._business_row_count(worksheet)
                self.assertGreater(business_rows, 0)
                self.assertGreaterEqual(worksheet.max_row, 30)
                self.assertGreaterEqual(worksheet.max_column, 19)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)


class TestManualHeadcountGenderSplit(unittest.TestCase):
    def _write_empty_headcount_csv(self, tmpdir: Path):
        csv_path = tmpdir / "headcount_manual.csv"
        with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "cc_code",
                    "period",
                    "headcount_staff",
                    "headcount_worker",
                    "headcount_male",
                    "headcount_female",
                    "description",
                ],
            )
            writer.writeheader()
        return csv_path

    def test_manual_headcount_template_is_header_only(self):
        tmpdir = _mk_tmpdir()
        try:
            csv_path = Path(ensure_manual_headcount_template(str(tmpdir), 2027))
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            self.assertEqual(
                reader.fieldnames,
                [
                    "cc_code",
                    "period",
                    "headcount_staff",
                    "headcount_worker",
                    "headcount_male",
                    "headcount_female",
                    "description",
                ],
            )
            self.assertEqual(rows, [])
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_parser_resolves_project_docs_source_to_raw_and_ignores_docs_csv(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            project_root = tmpdir / "project"
            docs_dir = project_root / "docs" / "MP2027"
            raw_dir = project_root / "raw"
            docs_dir.mkdir(parents=True)
            raw_dir.mkdir(parents=True)

            fieldnames = [
                "cc_code",
                "period",
                "headcount_staff",
                "headcount_worker",
                "headcount_male",
                "headcount_female",
                "description",
            ]
            for path, staff, worker, description in [
                (docs_dir / "headcount_manual.csv", "99", "0", "legacy docs row"),
                (raw_dir / "headcount_manual.csv", "7", "1", "canonical raw row"),
            ]:
                with path.open("w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerow(
                        {
                            "cc_code": str(cc_code),
                            "period": "202701",
                            "headcount_staff": staff,
                            "headcount_worker": worker,
                            "headcount_male": "",
                            "headcount_female": "",
                            "description": description,
                        }
                    )

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                resolved = resolve_manual_headcount_source_dir(str(docs_dir), base_dir=str(project_root))
            self.assertEqual(Path(resolved).resolve(), raw_dir.resolve())
            self.assertTrue(any(LEGACY_HEADCOUNT_SOURCE_WARNING in str(w.message) for w in caught))

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                result = parse_manual_headcount(conn, source_dir=str(docs_dir), base_dir=str(project_root))
            self.assertTrue(any(LEGACY_HEADCOUNT_SOURCE_WARNING in str(w.message) for w in caught))

            self.assertEqual(Path(str(result["template_path"])).resolve(), (raw_dir / "headcount_manual.csv").resolve())
            self.assertEqual(result["inserted"], 1)
            row = conn.execute(
                """
                SELECT headcount_staff, headcount_worker, description
                FROM fact_monthly_headcount
                WHERE cc_code=? AND period='202701' AND source='manual'
                """,
                (cc_code,),
            ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(float(row["headcount_staff"]), 7.0)
            self.assertEqual(float(row["headcount_worker"]), 1.0)
            self.assertEqual(row["description"], "canonical raw row")
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_explicit_docs_legacy_file_path_warns_and_does_not_import_rows(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            project_root = tmpdir / "project"
            docs_dir = project_root / "docs" / "MP2027"
            raw_dir = project_root / "raw"
            docs_dir.mkdir(parents=True)
            raw_dir.mkdir(parents=True)
            fieldnames = [
                "cc_code",
                "period",
                "headcount_staff",
                "headcount_worker",
                "headcount_male",
                "headcount_female",
                "description",
            ]
            docs_path = docs_dir / "headcount_manual.csv"
            with docs_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": "202701",
                        "headcount_staff": "99",
                        "headcount_worker": "0",
                        "headcount_male": "",
                        "headcount_female": "",
                        "description": "legacy file path must not import",
                    }
                )
            with (raw_dir / "headcount_manual.csv").open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                result = parse_manual_headcount(conn, source_dir=str(docs_path), base_dir=str(project_root))

            self.assertTrue(any(LEGACY_HEADCOUNT_SOURCE_WARNING in str(w.message) for w in caught))
            self.assertEqual(Path(str(result["template_path"])).resolve(), (raw_dir / "headcount_manual.csv").resolve())
            self.assertEqual(result["inserted"], 0)
            count = conn.execute("SELECT COUNT(*) FROM fact_monthly_headcount").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_docs_source_creates_new_headcount_template_only_in_raw(self):
        conn = _mk_conn()
        _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            project_root = tmpdir / "project"
            docs_dir = project_root / "docs" / "MP2027"
            docs_dir.mkdir(parents=True)

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                result = parse_manual_headcount(conn, source_dir=str(docs_dir), base_dir=str(project_root))

            self.assertTrue(any(LEGACY_HEADCOUNT_SOURCE_WARNING in str(w.message) for w in caught))
            raw_path = project_root / "raw" / "headcount_manual.csv"
            self.assertEqual(Path(str(result["template_path"])).resolve(), raw_path.resolve())
            self.assertTrue(raw_path.exists())
            self.assertFalse((docs_dir / "headcount_manual.csv").exists())
            self.assertEqual(result["inserted"], 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_quarantine_unconfirmed_headcount_removes_exact_keys_without_bus_or_other_cc_loss(self):
        conn = _mk_conn()
        target_cc = _seed_cc(conn, code=1412000006)
        other_cc = _seed_cc(conn, code=1412000004)
        tmpdir = _mk_tmpdir()
        periods = ["202701", "202702", "202703"]
        try:
            fieldnames = [
                "cc_code",
                "period",
                "headcount_staff",
                "headcount_worker",
                "headcount_male",
                "headcount_female",
                "description",
            ]
            csv_path = tmpdir / "headcount_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for period in periods:
                    writer.writerow(
                        {
                            "cc_code": str(target_cc),
                            "period": period,
                            "headcount_staff": "28",
                            "headcount_worker": "0",
                            "headcount_male": "",
                            "headcount_female": "",
                            "description": "",
                        }
                    )
                writer.writerow(
                    {
                        "cc_code": str(other_cc),
                        "period": "202701",
                        "headcount_staff": "3",
                        "headcount_worker": "2",
                        "headcount_male": "",
                        "headcount_female": "",
                        "description": "other cc must remain",
                    }
                )

            bus_path = tmpdir / "bus_headcount_manual.csv"
            with bus_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "bus_expat_count", "bus_vietnamese_count", "description"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(target_cc),
                        "bus_expat_count": "4",
                        "bus_vietnamese_count": "5",
                        "description": "bus survives quarantine",
                    }
                )

            conn.executemany(
                """
                INSERT INTO fact_monthly_headcount
                (period, cc_code, headcount_all, headcount_staff, headcount_worker, source, description)
                VALUES (?, ?, ?, ?, ?, 'manual', ?)
                """,
                [(period, target_cc, 28, 28, 0, "") for period in periods]
                + [("202701", other_cc, 5, 3, 2, "other cc must remain")],
            )
            conn.execute(
                """
                INSERT INTO fact_bus_headcount_drivers
                (cc_code, bus_expat_count, bus_vietnamese_count, source, description)
                VALUES (?, 4, 5, 'manual', 'bus survives quarantine')
                """,
                (str(target_cc),),
            )
            conn.commit()

            quarantine_path = tmpdir / "quarantine" / "headcount_1412000006_unconfirmed_quarantine.csv"
            result = quarantine_manual_headcount_rows(
                conn,
                source_dir=str(tmpdir),
                cc_code=str(target_cc),
                periods=periods,
                quarantine_path=str(quarantine_path),
                reason="not confirmed by user",
                quarantined_at="2026-06-13T00:00:00+00:00",
            )

            self.assertEqual(result["csv_rows_removed"], 3)
            self.assertEqual(result["logical_values_quarantined"], 6)
            self.assertEqual(result["db_rows_deleted"], 3)

            with quarantine_path.open("r", encoding="utf-8-sig", newline="") as f:
                quarantine_rows = list(csv.DictReader(f))
            self.assertEqual(len(quarantine_rows), 6)
            self.assertEqual({row["driver_type"] for row in quarantine_rows}, {"headcount_staff", "headcount_worker"})
            self.assertEqual({row["value"] for row in quarantine_rows if row["driver_type"] == "headcount_worker"}, {"0"})

            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                rows_after = list(csv.DictReader(f))
            self.assertFalse(
                any(str(row["cc_code"]) == str(target_cc) and row["period"] in periods for row in rows_after)
            )
            self.assertTrue(any(str(row["cc_code"]) == str(other_cc) for row in rows_after))

            target_count = conn.execute(
                """
                SELECT COUNT(*) FROM fact_monthly_headcount
                WHERE CAST(cc_code AS TEXT)=? AND period IN ('202701','202702','202703') AND source='manual'
                """,
                (str(target_cc),),
            ).fetchone()[0]
            self.assertEqual(target_count, 0)
            other_count = conn.execute(
                "SELECT COUNT(*) FROM fact_monthly_headcount WHERE CAST(cc_code AS TEXT)=? AND source='manual'",
                (str(other_cc),),
            ).fetchone()[0]
            self.assertEqual(other_count, 1)
            bus_count = conn.execute(
                "SELECT COUNT(*) FROM fact_bus_headcount_drivers WHERE CAST(cc_code AS TEXT)=?",
                (str(target_cc),),
            ).fetchone()[0]
            self.assertEqual(bus_count, 1)

            second = quarantine_manual_headcount_rows(
                conn,
                source_dir=str(tmpdir),
                cc_code=str(target_cc),
                periods=periods,
                quarantine_path=str(quarantine_path),
                reason="not confirmed by user",
                quarantined_at="2026-06-13T00:00:00+00:00",
            )
            self.assertEqual(second["csv_rows_removed"], 0)
            self.assertEqual(second["logical_values_quarantined"], 0)
            self.assertEqual(second["db_rows_deleted"], 0)

            parsed = parse_manual_headcount(conn, source_dir=str(tmpdir))
            self.assertEqual(parsed["errors"], 0)
            self.assertEqual(parsed["inserted"], 1)
            target_count_after_parse = conn.execute(
                """
                SELECT COUNT(*) FROM fact_monthly_headcount
                WHERE CAST(cc_code AS TEXT)=? AND period IN ('202701','202702','202703') AND source='manual'
                """,
                (str(target_cc),),
            ).fetchone()[0]
            self.assertEqual(target_count_after_parse, 0)
            other_count_after_parse = conn.execute(
                "SELECT COUNT(*) FROM fact_monthly_headcount WHERE CAST(cc_code AS TEXT)=? AND source='manual'",
                (str(other_cc),),
            ).fetchone()[0]
            self.assertEqual(other_count_after_parse, 1)
            bus_row = conn.execute(
                """
                SELECT bus_expat_count, bus_vietnamese_count
                FROM fact_bus_headcount_drivers
                WHERE CAST(cc_code AS TEXT)=?
                """,
                (str(target_cc),),
            ).fetchone()
            self.assertIsNotNone(bus_row)
            self.assertEqual(float(bus_row["bus_expat_count"]), 4.0)
            self.assertEqual(float(bus_row["bus_vietnamese_count"]), 5.0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_parser_accepts_required_baseline_and_is_idempotent(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_required_headcount_periods(2027)

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "headcount_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "headcount_staff",
                        "headcount_worker",
                        "headcount_male",
                        "headcount_female",
                        "description",
                    ],
                )
                writer.writeheader()
                for idx, period in enumerate(periods):
                    writer.writerow(
                        {
                            "cc_code": str(cc_code),
                            "period": period,
                            "headcount_staff": "27" if idx < 10 else "28",
                            "headcount_worker": "0",
                            "headcount_male": "",
                            "headcount_female": "",
                            "description": "full series",
                        }
                    )

            first = parse_manual_headcount(conn, source_dir=str(tmpdir))
            second = parse_manual_headcount(conn, source_dir=str(tmpdir))
            self.assertEqual(first["inserted"], 13)
            self.assertEqual(second["inserted"], 13)
            self.assertEqual(first["errors"], 0)
            self.assertEqual(second["errors"], 0)

            count = conn.execute(
                "SELECT COUNT(*) FROM fact_monthly_headcount WHERE cc_code=? AND source='manual'",
                (cc_code,),
            ).fetchone()[0]
            self.assertEqual(count, 13)
            baseline = conn.execute(
                """
                SELECT headcount_staff, headcount_worker, source, description
                FROM fact_monthly_headcount
                WHERE cc_code=? AND period='202603'
                """,
                (cc_code,),
            ).fetchone()
            self.assertIsNotNone(baseline)
            self.assertEqual(float(baseline["headcount_staff"]), 27.0)
            self.assertEqual(float(baseline["headcount_worker"]), 0.0)
            self.assertEqual(baseline["source"], "manual")
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_parser_reads_scalar_bus_headcount_drivers_idempotently(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            self._write_empty_headcount_csv(tmpdir)
            bus_path = tmpdir / "bus_headcount_manual.csv"
            with bus_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "bus_expat_count", "bus_vietnamese_count", "description"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "bus_expat_count": "3",
                        "bus_vietnamese_count": "7",
                        "description": "scalar bus counts",
                    }
                )

            first = parse_manual_headcount(conn, source_dir=str(tmpdir))
            second = parse_manual_headcount(conn, source_dir=str(tmpdir))

            self.assertEqual(first["bus_inserted"], 1)
            self.assertEqual(second["bus_inserted"], 1)
            self.assertEqual(first["bus_errors"], 0)
            self.assertEqual(second["bus_errors"], 0)

            rows = conn.execute("SELECT * FROM fact_bus_headcount_drivers WHERE cc_code=?", (str(cc_code),)).fetchall()
            self.assertEqual(len(rows), 1)
            self.assertEqual(float(rows[0]["bus_expat_count"]), 3.0)
            self.assertEqual(float(rows[0]["bus_vietnamese_count"]), 7.0)

            headcount_rows = conn.execute("SELECT COUNT(*) FROM fact_monthly_headcount").fetchone()[0]
            self.assertEqual(headcount_rows, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_parser_rejects_non_integer_bus_driver_counts(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            self._write_empty_headcount_csv(tmpdir)
            bus_path = tmpdir / "bus_headcount_manual.csv"
            with bus_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "bus_expat_count", "bus_vietnamese_count", "description"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "bus_expat_count": "1.5",
                        "bus_vietnamese_count": "2",
                        "description": "bad count",
                    }
                )

            result = parse_manual_headcount(conn, source_dir=str(tmpdir))
            self.assertEqual(result["bus_inserted"], 0)
            self.assertEqual(result["bus_errors"], 1)
            count = conn.execute("SELECT COUNT(*) FROM fact_bus_headcount_drivers").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_parser_requires_staff_and_worker_category_values(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "headcount_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "headcount_staff",
                        "headcount_worker",
                        "description",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": "202603",
                        "headcount_staff": "27",
                        "headcount_worker": "",
                        "description": "missing worker category",
                    }
                )

            result = parse_manual_headcount(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 1)
            count = conn.execute("SELECT COUNT(*) FROM fact_monthly_headcount").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_parser_reads_optional_gender_columns(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "headcount_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "headcount_staff",
                        "headcount_worker",
                        "headcount_male",
                        "headcount_female",
                        "description",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": periods[8],
                        "headcount_staff": "5",
                        "headcount_worker": "10",
                        "headcount_male": "6",
                        "headcount_female": "7",
                        "description": "health check split",
                    }
                )

            result = parse_manual_headcount(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)

            row = conn.execute(
                """
                SELECT headcount_all, headcount_staff, headcount_worker, headcount_male, headcount_female
                FROM fact_monthly_headcount
                WHERE cc_code=? AND period=? AND source='manual'
                """,
                (cc_code, periods[8]),
            ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(float(row["headcount_all"]), 15.0)
            self.assertEqual(float(row["headcount_male"]), 6.0)
            self.assertEqual(float(row["headcount_female"]), 7.0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)


class TestHeadcountMissingMatrix(unittest.TestCase):
    def test_pipeline_audit_reports_period_category_matrix_for_missing_target_series(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000040)
        tmpdir = _mk_tmpdir()
        try:
            result = write_pipeline_audit_report(
                conn=conn,
                output_dir=str(tmpdir),
                source_dir=str(tmpdir),
                fiscal_year=2027,
                target_cc=cc_code,
                parser_results={},
            )
            rows = []
            with open(result["missing_csv_path"], "r", encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    if row["area"] == "headcount_series":
                        rows.append(row)

            expected = [
                (period, category)
                for period in get_required_headcount_periods(2027)
                for category in ("headcount_staff", "headcount_worker")
            ]
            observed = [
                (
                    row["period"],
                    "headcount_staff" if "category=headcount_staff" in row["message"] else "headcount_worker",
                )
                for row in rows
            ]
            self.assertEqual(observed, expected)
            self.assertTrue(all(row["cc_code"] == str(cc_code) for row in rows))
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)


class TestHealthCheckAllocation(unittest.TestCase):
    def test_health_check_rules_use_gender_specific_counts(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)

        conn.execute(
            """
            INSERT INTO fact_monthly_headcount
            (
                period, cc_code, headcount_all, headcount_staff, headcount_worker,
                headcount_male, headcount_female, source, description
            )
            VALUES (?, ?, 20, 5, 15, 6, 9, 'manual', 'december split')
            """,
            (periods[8], cc_code),
        )

        conn.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES
            ('QA', 'Khám sức khỏe (cho CNV nam)', 'HC', 500001, 600001, 700001, '12月', 100, '/unit', 'headcount_worker', '12月'),
            ('QA', 'Khám sức khỏe (cho CNV nữ)', 'HC', 500002, 600002, 700002, '12月', 200, '/unit', 'headcount_worker', '12月')
            """
        )
        conn.commit()

        AllocationEngine(conn)._process_allocation_rules()

        male_amount = conn.execute(
            "SELECT amount_vnd FROM fact_input_data WHERE description='Alloc: Khám sức khỏe (cho CNV nam)'"
        ).fetchone()
        female_amount = conn.execute(
            "SELECT amount_vnd FROM fact_input_data WHERE description='Alloc: Khám sức khỏe (cho CNV nữ)'"
        ).fetchone()

        self.assertIsNotNone(male_amount)
        self.assertIsNotNone(female_amount)
        self.assertEqual(float(male_amount["amount_vnd"]), 600.0)
        self.assertEqual(float(female_amount["amount_vnd"]), 1800.0)
        conn.close()


class TestPostingMonthOverride(unittest.TestCase):
    def test_override_uses_months_from_business_sheet(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)

        conn.execute(
            """
            INSERT INTO fact_monthly_headcount
            (
                period, cc_code, headcount_all, headcount_staff, headcount_worker,
                headcount_male, headcount_female, source, description
            )
            VALUES
            (?, ?, 3, 0, 3, 0, 0, 'manual', 'may worker'),
            (?, ?, 4, 0, 4, 0, 0, 'manual', 'feb worker')
            """,
            (periods[1], cc_code, periods[10], cc_code),
        )
        conn.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES
            ('GA', 'Tiệc khuấy động năm tài chính / 決起コンパ（対象：全部門）', 'GA', 5004086291, 6004086651, 6004086551, '-', 300000, '/unit', 'headcount_worker', '-'),
            ('GA', '忘年会補助金 Hỗ trợ tiệc tất niên', 'GA', 5004086291, 6004086651, 6004086551, '-', 200000, '/unit', 'headcount_worker', '-')
            """
        )
        conn.commit()

        AllocationEngine(conn)._process_allocation_rules()

        kickoff = conn.execute(
            """
            SELECT period, amount_vnd
            FROM fact_input_data
            WHERE description LIKE 'Alloc: Tiệc khuấy động năm tài chính%'
            """
        ).fetchall()
        bonenkai = conn.execute(
            """
            SELECT period, amount_vnd
            FROM fact_input_data
            WHERE description = 'Alloc: 忘年会補助金 Hỗ trợ tiệc tất niên'
            """
        ).fetchall()

        self.assertEqual([(row["period"], float(row["amount_vnd"])) for row in kickoff], [(periods[1], 900000.0)])
        self.assertEqual([(row["period"], float(row["amount_vnd"])) for row in bonenkai], [(periods[10], 800000.0)])
        conn.close()


class TestEventDeltaHeadcountFailClosed(unittest.TestCase):
    def _insert_rule(
        self,
        conn,
        *,
        posting_month="\u5165\u793e\u6708",
        item_name="new hire stationery",
        unit_price=9100,
        driver_type="headcount_all",
    ):
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES ('GA', ?, 'Stationery', 5005246288, 6005246288, 7005246288, ?, ?, '/person', ?, ?)
            """,
            (item_name, posting_month, float(unit_price), driver_type, posting_month),
        )
        conn.commit()
        return cursor.lastrowid

    def _insert_headcount(self, conn, cc_code, period_values):
        conn.executemany(
            """
            INSERT INTO fact_monthly_headcount
            (period, cc_code, headcount_all, headcount_staff, headcount_worker, source, description)
            VALUES (?, ?, ?, ?, 0, 'manual', 'event delta test')
            """,
            [(period, cc_code, value, value) for period, value in period_values],
        )
        conn.commit()

    def _alloc_rows(self, conn, rule_id):
        return conn.execute(
            """
            SELECT period, amount_vnd
            FROM fact_input_data
            WHERE source = ?
            ORDER BY period
            """,
            (f"alloc_{rule_id}",),
        ).fetchall()

    def _missing_rows(self, conn):
        return conn.execute(
            """
            SELECT period, message
            FROM fact_missing_inputs
            WHERE source = 'allocator'
            ORDER BY id
            """
        ).fetchall()

    def test_complete_series_uses_real_month_delta_not_total_headcount(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        self._insert_headcount(
            conn,
            cc_code,
            [("202603", 27), *[(period, 27) for period in periods[:9]], *[(period, 28) for period in periods[9:]]],
        )
        rule_id = self._insert_rule(conn)

        AllocationEngine(conn)._process_allocation_rules()

        rows = self._alloc_rows(conn, rule_id)
        self.assertEqual([(row["period"], float(row["amount_vnd"])) for row in rows], [(periods[9], 9100.0)])
        self.assertNotIn(28 * 9100.0, {float(row["amount_vnd"]) for row in rows})
        self.assertEqual(self._missing_rows(conn), [])
        conn.close()

    def test_missing_previous_month_fails_closed_without_fake_first_delta(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        self._insert_headcount(conn, cc_code, [(periods[9], 28), (periods[10], 28), (periods[11], 28)])
        rule_id = self._insert_rule(conn)

        AllocationEngine(conn)._process_allocation_rules()

        rows = self._alloc_rows(conn, rule_id)
        self.assertEqual(rows, [])
        missing_messages = [row["message"] for row in self._missing_rows(conn)]
        self.assertTrue(any(f"cc={cc_code}, month={periods[9]}, previous_month={periods[8]}" in msg for msg in missing_messages))
        self.assertTrue(any("category=headcount_all" in msg for msg in missing_messages))
        self.assertFalse(any("amount_vnd" in str(row) and str(28 * 9100) in str(row) for row in rows))
        conn.close()

    def test_missing_middle_month_does_not_bridge_delta(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        self._insert_headcount(
            conn,
            cc_code,
            [
                ("202603", 27),
                (periods[0], 27),
                # periods[1] intentionally missing; Jun must not bridge Apr -> Jun.
                (periods[2], 28),
            ],
        )
        rule_id = self._insert_rule(conn)

        AllocationEngine(conn)._process_allocation_rules()

        self.assertEqual(self._alloc_rows(conn, rule_id), [])
        missing_messages = [row["message"] for row in self._missing_rows(conn)]
        self.assertTrue(any(f"month={periods[2]}, previous_month={periods[1]}" in msg for msg in missing_messages))
        self.assertTrue(any("missing=previous" in msg for msg in missing_messages))
        conn.close()

    def test_missing_all_monthly_headcount_records_missing_input_without_amount(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000040)
        rule_id = self._insert_rule(conn)

        AllocationEngine(conn)._process_allocation_rules()

        self.assertEqual(self._alloc_rows(conn, rule_id), [])
        missing_messages = [row["message"] for row in self._missing_rows(conn)]
        self.assertTrue(any(f"cc={cc_code}" in msg for msg in missing_messages))
        self.assertTrue(any("Missing complete monthly headcount driver" in msg for msg in missing_messages))
        conn.close()

    def test_non_event_headcount_allocation_still_uses_master_fallback(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            "UPDATE dim_cost_centers SET staff_count = 2, worker_count = 3 WHERE code = ?",
            (str(cc_code),),
        )
        conn.commit()
        periods = get_fy_months(2027)
        rule_id = self._insert_rule(
            conn,
            posting_month="every month",
            item_name="recurring non-event allocation",
            unit_price=100,
        )

        AllocationEngine(conn)._process_allocation_rules()

        rows = self._alloc_rows(conn, rule_id)
        self.assertEqual(len(rows), 12)
        self.assertEqual([(row["period"], float(row["amount_vnd"])) for row in rows], [(period, 500.0) for period in periods])
        self.assertEqual(self._missing_rows(conn), [])
        conn.close()


class TestBusHeadcountAllocation(unittest.TestCase):
    def _insert_bus_rule(self, conn, *, item_name: str, unit_price: float, account: int = 5004086291) -> int:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES ('GA', ?, '福利厚生費', ?, ?, ?, '毎月', ?, '/person', 'headcount_all', '人数')
            """,
            (item_name, account, account, account, float(unit_price)),
        )
        conn.commit()
        return cursor.lastrowid

    def _insert_bus_counts(self, conn, cc_code, *, expat: int, vietnamese: int):
        conn.execute(
            """
            INSERT INTO fact_bus_headcount_drivers
            (cc_code, bus_expat_count, bus_vietnamese_count, source, description)
            VALUES (?, ?, ?, 'manual', 'bus allocation test')
            """,
            (str(cc_code), float(expat), float(vietnamese)),
        )
        conn.commit()

    def _insert_bus_unit_price_series(self, conn, *, item_name: str, prices: list[int]):
        periods = get_fy_months(2027)
        conn.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, scenario_id, description)
            VALUES ('ga_unit_price', ?, ?, 0, 0, 'base', ?)
            """,
            [
                (period, float(price), f"{item_name}|headcount_per_person")
                for period, price in zip(periods, prices)
            ],
        )
        conn.commit()

    def test_bus_drivers_allocate_same_counts_to_all_12_months_and_export_formulas(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        expat_rule_id = self._insert_bus_rule(
            conn,
            item_name="出向者通勤送迎費 Xe đưa đón cho người Nhật",
            unit_price=1000,
        )
        vn_rule_id = self._insert_bus_rule(
            conn,
            item_name="ローカル通勤送迎費 Xe đưa đón cho người Việt",
            unit_price=2000,
        )
        self._insert_bus_counts(conn, cc_code, expat=3, vietnamese=4)
        conn.execute(
            """
            INSERT INTO fact_monthly_headcount
            (period, cc_code, headcount_all, headcount_staff, headcount_worker, source, description)
            VALUES ('202604', ?, 999, 888, 111, 'manual', 'must not drive bus')
            """,
            (cc_code,),
        )
        conn.commit()

        AllocationEngine(conn).run_allocation()

        expat_rows = conn.execute(
            "SELECT period, amount_vnd, form_row, description FROM fact_input_data WHERE source=? ORDER BY period",
            (f"alloc_{expat_rule_id}",),
        ).fetchall()
        vn_rows = conn.execute(
            "SELECT period, amount_vnd, form_row, description FROM fact_input_data WHERE source=? ORDER BY period",
            (f"alloc_{vn_rule_id}",),
        ).fetchall()

        self.assertEqual([row["period"] for row in expat_rows], periods)
        self.assertEqual([row["period"] for row in vn_rows], periods)
        self.assertTrue(all(float(row["amount_vnd"]) == 3000.0 for row in expat_rows))
        self.assertTrue(all(float(row["amount_vnd"]) == 8000.0 for row in vn_rows))
        self.assertTrue(all(int(row["form_row"]) == 53 for row in expat_rows))
        self.assertTrue(all(int(row["form_row"]) == 54 for row in vn_rows))
        self.assertTrue(all("driver_type=bus_expat_count" in row["description"] for row in expat_rows))
        self.assertTrue(all("driver_type=bus_vietnamese_count" in row["description"] for row in vn_rows))

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_bus_headcount_driver.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                for column in range(6, 18):
                    self.assertEqual(ws.cell(53, column).value, "=3*1000")
                    self.assertEqual(ws.cell(54, column).value, "=4*2000")
                self.assertEqual(ws["R53"].value, "=SUM(F53:Q53)")
                self.assertEqual(ws["R54"].value, "=SUM(F54:Q54)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_bus_drivers_use_monthly_ga_unit_price_source_before_footnote_rule(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        expat_prices = [856107, 894022, 841125, 843756, 840449, 826939, 807491, 851412, 890880, 868440, 823640, 810900]
        vn_prices = [1031546, 1666547, 1572350, 1523385, 1411084, 1510555, 1530445, 1556432, 1469160, 1461107, 1162654, 1116875]
        expat_rule_id = self._insert_bus_rule(
            conn,
            item_name="出向者通勤送迎費 Xe đưa đón cho người Nhật",
            unit_price=0,
        )
        vn_rule_id = self._insert_bus_rule(
            conn,
            item_name="ローカル通勤送迎費 Xe đưa đón cho người Việt",
            unit_price=0,
        )
        self._insert_bus_unit_price_series(
            conn,
            item_name="出向者送迎費 Xe đưa đón người Nhật",
            prices=expat_prices,
        )
        self._insert_bus_unit_price_series(
            conn,
            item_name="ローカル社員送迎費 Xe đưa đón người Việt",
            prices=vn_prices,
        )
        self._insert_bus_counts(conn, cc_code, expat=3, vietnamese=4)

        AllocationEngine(conn).run_allocation()

        expat_rows = conn.execute(
            "SELECT period, amount_vnd, description FROM fact_input_data WHERE source=? ORDER BY period",
            (f"alloc_{expat_rule_id}",),
        ).fetchall()
        vn_rows = conn.execute(
            "SELECT period, amount_vnd, description FROM fact_input_data WHERE source=? ORDER BY period",
            (f"alloc_{vn_rule_id}",),
        ).fetchall()

        self.assertEqual([row["period"] for row in expat_rows], periods)
        self.assertEqual([row["period"] for row in vn_rows], periods)
        self.assertEqual([float(row["amount_vnd"]) for row in expat_rows], [3.0 * price for price in expat_prices])
        self.assertEqual([float(row["amount_vnd"]) for row in vn_rows], [4.0 * price for price in vn_prices])
        self.assertTrue(all("unit_price_source=ga_unit_price" in row["description"] for row in expat_rows))
        self.assertTrue(all("source_cells=B9:M9" in row["description"] for row in expat_rows))
        self.assertTrue(all("source_cells=B10:M10" in row["description"] for row in vn_rows))
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM fact_missing_inputs WHERE area='bus_headcount_driver'").fetchone()[0],
            0,
        )

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_bus_monthly_unit_prices.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                for offset, price in enumerate(expat_prices):
                    self.assertEqual(ws.cell(53, 6 + offset).value, f"=3*{price}")
                for offset, price in enumerate(vn_prices):
                    self.assertEqual(ws.cell(54, 6 + offset).value, f"=4*{price}")
                self.assertEqual(ws["R53"].value, "=SUM(F53:Q53)")
                self.assertEqual(ws["R54"].value, "=SUM(F54:Q54)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_bus_unit_price_missing_fails_closed(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        expat_rule_id = self._insert_bus_rule(
            conn,
            item_name="出向者通勤送迎費 Xe đưa đón cho người Nhật",
            unit_price=0,
        )
        self._insert_bus_counts(conn, cc_code, expat=3, vietnamese=0)

        AllocationEngine(conn).run_allocation()

        fact_count = conn.execute(
            "SELECT COUNT(*) FROM fact_input_data WHERE source=?",
            (f"alloc_{expat_rule_id}",),
        ).fetchone()[0]
        self.assertEqual(fact_count, 0)

        missing = conn.execute(
            """
            SELECT message
            FROM fact_missing_inputs
            WHERE area='bus_headcount_driver'
            ORDER BY id
            """
        ).fetchall()
        self.assertTrue(any("missing=expat bus unit_price" in row["message"] for row in missing))
        conn.close()


class TestNewHireAllocationIdentityDedupe(unittest.TestCase):
    def _insert_notebook_rule(self, conn, *, unit_price=9100, driver_type="headcount_staff"):
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES ('GA', 'notebook source', 'Office supplies',
                    5005246288, 5005246288, 5005246288,
                    ?, ?, '/person', ?, ?)
            """,
            ("\u5165\u793e\u6708", float(unit_price), driver_type, "\u5165\u793e\u6708"),
        )
        conn.commit()
        return cursor.lastrowid

    def _insert_full_headcount_series(self, conn, cc_code, *, staff_values, worker_values):
        rows = []
        for period in ["202603", *get_fy_months(2027)]:
            staff = float(staff_values.get(period, 0.0))
            worker = float(worker_values.get(period, 0.0))
            rows.append((period, cc_code, staff + worker, staff, worker))
        conn.executemany(
            """
            INSERT INTO fact_monthly_headcount
            (period, cc_code, headcount_all, headcount_staff, headcount_worker, source, description)
            VALUES (?, ?, ?, ?, ?, 'manual', 'new hire identity test')
            """,
            rows,
        )
        conn.commit()

    def _export(self, conn, cc_code, output_name):
        tmpdir = _mk_tmpdir()
        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        output_path = tmpdir / output_name
        ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
        self.assertTrue(ok)
        return tmpdir, output_path

    def test_staff_notebook_allocation_maps_only_to_staff_notebook_row(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        jan_col = 6 + periods.index("202701")
        staff_values = {period: 27 for period in ["202603", *periods]}
        staff_values.update({period: 28 for period in periods[9:]})
        self._insert_full_headcount_series(conn, cc_code, staff_values=staff_values, worker_values={})
        rule_id = self._insert_notebook_rule(conn, unit_price=9100, driver_type="headcount_staff")

        AllocationEngine(conn)._process_allocation_rules()

        rows = conn.execute(
            "SELECT period, amount_vnd FROM fact_input_data WHERE source = ? ORDER BY period",
            (f"alloc_{rule_id}",),
        ).fetchall()
        self.assertEqual([(row["period"], float(row["amount_vnd"])) for row in rows], [("202701", 9100.0)])

        tmpdir, output_path = self._export(conn, cc_code, "out_staff_notebook_identity.xlsx")
        try:
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws.cell(97, jan_col).value, "=1*9100")
                self.assertIsNone(ws.cell(98, jan_col).value)
                self.assertIsNone(ws.cell(90, jan_col).value)
            finally:
                workbook.close()

            apply_complete_v1_source_order_to_workbook(output_path, start_row=168, clear_until_row=212)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                matching_rows = [
                    row_index
                    for row_index in range(30, 213)
                    if ws.cell(row_index, jan_col).value == "=1*9100"
                ]
                self.assertEqual(len(matching_rows), 1)
                self.assertNotIn(90, matching_rows)
                self.assertNotIn(97, matching_rows)
                self.assertNotIn(98, matching_rows)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_worker_notebook_allocation_maps_only_to_worker_notebook_row(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        jan_col = 6 + periods.index("202701")
        worker_values = {period: 8 for period in ["202603", *periods]}
        worker_values.update({period: 9 for period in periods[9:]})
        self._insert_full_headcount_series(conn, cc_code, staff_values={}, worker_values=worker_values)
        self._insert_notebook_rule(conn, unit_price=4000, driver_type="headcount_worker")

        AllocationEngine(conn)._process_allocation_rules()

        tmpdir, output_path = self._export(conn, cc_code, "out_worker_notebook_identity.xlsx")
        try:
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws.cell(98, jan_col).value, "=1*4000")
                self.assertIsNone(ws.cell(97, jan_col).value)
                self.assertIsNone(ws.cell(90, jan_col).value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_same_amount_staff_and_worker_notebooks_are_not_over_deduped(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        jan_col = 6 + periods.index("202701")
        staff_values = {period: 10 for period in ["202603", *periods]}
        staff_values.update({period: 11 for period in periods[9:]})
        worker_values = {period: 20 for period in ["202603", *periods]}
        worker_values.update({period: 21 for period in periods[9:]})
        self._insert_full_headcount_series(conn, cc_code, staff_values=staff_values, worker_values=worker_values)
        self._insert_notebook_rule(conn, unit_price=9100, driver_type="headcount_staff")
        self._insert_notebook_rule(conn, unit_price=9100, driver_type="headcount_worker")

        AllocationEngine(conn)._process_allocation_rules()

        tmpdir, output_path = self._export(conn, cc_code, "out_staff_worker_same_amount.xlsx")
        try:
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws.cell(97, jan_col).value, "=1*9100")
                self.assertEqual(ws.cell(98, jan_col).value, "=1*9100")
                self.assertIsNone(ws.cell(90, jan_col).value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_duplicate_same_source_identity_is_written_once(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        jan_col = 6 + periods.index("202701")
        rule_id = self._insert_notebook_rule(conn, unit_price=9100, driver_type="headcount_staff")
        conn.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, form_row, description)
            VALUES (?, '202701', 9100, 0, ?, 5005246288, NULL, 'Alloc: notebook source')
            """,
            [(f"alloc_{rule_id}", cc_code), (f"alloc_{rule_id}", cc_code)],
        )
        conn.commit()

        tmpdir, output_path = self._export(conn, cc_code, "out_duplicate_source_identity.xlsx")
        try:
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws.cell(97, jan_col).value, "=1*9100")
                self.assertNotEqual(ws.cell(97, jan_col).value, "=1*9100+1*9100")
                self.assertIsNone(ws.cell(90, jan_col).value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)


class TestRuleLoaderAndManualEventSafeguard(unittest.TestCase):
    def test_parse_unit_price_supports_currency_suffix(self):
        self.assertEqual(_parse_unit_price("145$"), 145.0)
        self.assertEqual(_parse_unit_price("1,259,500"), 1259500.0)
        self.assertEqual(_parse_unit_price(3000000), 3000000.0)
        self.assertIsNone(_parse_unit_price("※1"))
        self.assertIsNone(_parse_unit_price("abc"))

    def test_mp2026_reference_unit_price_fills_zero_rule_prices(self):
        self.assertEqual(_apply_mp2026_reference_unit_price("月餅 Bánh Trung Thu", 0), 56000.0)
        self.assertEqual(_apply_mp2026_reference_unit_price("運動会 Đại hội thể thao", 0), 107000.0)
        self.assertEqual(_apply_mp2026_reference_unit_price("運動会 Đại hội thể thao", 123), 123.0)

    def test_allocation_loader_keeps_footnote_unit_price_rules_as_missing_price_metadata(self):
        conn = _mk_conn()
        tmpdir = _mk_tmpdir()
        try:
            source_path = tmpdir / "FY2027_allocation.xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "allocation"
            ws.append(["dept", "item", "account", "mfg", "ga", "sales", "posting", "unit_price", "unit", "driver"])
            ws.append(
                [
                    "GA",
                    "出向者通勤送迎費 Xe đưa đón cho người Nhật",
                    "福利厚生費",
                    5004086291,
                    6004086651,
                    6004086551,
                    "毎月",
                    "※1",
                    "/person",
                    "人数",
                ]
            )
            workbook.save(source_path)
            workbook.close()

            loaded = load_allocation_rules(conn, alloc_path=str(source_path), fiscal_year=2027)
            self.assertEqual(loaded, 1)
            row = conn.execute("SELECT item_name, unit_price FROM map_allocation_rules").fetchone()
            self.assertIn("出向者通勤送迎費", row["item_name"])
            self.assertEqual(float(row["unit_price"]), 0.0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_allocation_loader_keeps_new_hire_notebook_staff_continuation_row(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        tmpdir = _mk_tmpdir()
        try:
            source_path = tmpdir / "FY2027配賦額一覧 (2025.12.29).xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "配賦額一覧"
            ws.append(["配布元", "内容", "科目名称", "製造コード", "間接コード", "販売コード", "計上月", "単価", "単位", "計上基準"])
            ws.append(
                [
                    "人事課",
                    "ノート Sổ",
                    "事務用品費",
                    5005246288,
                    6005126423,
                    6005246544,
                    "入社月",
                    4000,
                    "/冊",
                    "G7社員の配属人数で入社月に振替 Phân bổ theo số công nhân vào trong tháng",
                ]
            )
            ws.append(
                [
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    9100,
                    "/冊",
                    "スタッフの配属人数で入社月に振替 Phân bổ theo số nhân viên vào trong tháng",
                ]
            )
            workbook.save(source_path)
            workbook.close()

            loaded = load_allocation_rules(conn, alloc_path=str(source_path), fiscal_year=2027)
            self.assertEqual(loaded, 2)
            rules = conn.execute(
                """
                SELECT item_name, posting_month, unit_price, driver_type, mfg_account, ga_account, sales_account
                FROM map_allocation_rules
                ORDER BY unit_price
                """
            ).fetchall()
            self.assertEqual([float(row["unit_price"]) for row in rules], [4000.0, 9100.0])
            self.assertEqual([row["driver_type"] for row in rules], ["headcount_worker", "headcount_staff"])
            self.assertTrue(all(row["item_name"] == "ノート Sổ" for row in rules))
            self.assertTrue(all(row["posting_month"] == "入社月" for row in rules))
            self.assertTrue(all(int(row["mfg_account"]) == 5005246288 for row in rules))

            conn.executemany(
                """
                INSERT INTO fact_monthly_headcount
                (period, cc_code, headcount_all, headcount_staff, headcount_worker, source, description)
                VALUES (?, ?, ?, ?, ?, 'manual', 'new hire delta test')
                """,
                [
                    (periods[0], cc_code, 22, 22, 0),
                    (periods[1], cc_code, 22, 22, 0),
                    (periods[2], cc_code, 29, 26, 3),
                    (periods[3], cc_code, 30, 27, 3),
                ],
            )
            conn.commit()

            AllocationEngine(conn)._process_allocation_rules()
            rows = conn.execute(
                """
                SELECT period, amount_vnd, description
                FROM fact_input_data
                WHERE description LIKE 'Alloc: ノート Sổ%'
                ORDER BY period, amount_vnd
                """
            ).fetchall()
            observed = [(row["period"], float(row["amount_vnd"])) for row in rows]
            self.assertEqual(
                observed,
                [
                    (periods[2], 12000.0),
                    (periods[2], 36400.0),
                    (periods[3], 9100.0),
                ],
            )
            self.assertNotIn(periods[8], {row["period"] for row in rows})
            self.assertNotIn(26 * 9100.0, {float(row["amount_vnd"]) for row in rows})
            self.assertNotIn(3 * 9100.0, {float(row["amount_vnd"]) for row in rows})
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_acquisition_month_rules_are_skipped_until_manual_event_data_exists(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        period = get_fy_months(2027)[0]
        conn.execute(
            """
            INSERT INTO fact_monthly_headcount
            (
                period, cc_code, headcount_all, headcount_staff, headcount_worker,
                headcount_male, headcount_female, source, description
            )
            VALUES (?, ?, 20, 5, 15, 0, 0, 'manual', 'seed')
            """,
            (period, cc_code),
        )
        conn.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES
            ('GA', '旅券申請費用 Chi phi lam ho chieu', 'GA', 5005246286, 6005246673, 6005246542, '取得月', 1900000, '/nguoi', 'headcount_all', 'manual event')
            """
        )
        conn.commit()

        AllocationEngine(conn)._process_allocation_rules()
        count = conn.execute("SELECT COUNT(*) FROM fact_input_data WHERE source LIKE 'alloc_%'").fetchone()[0]
        self.assertEqual(count, 0)
        conn.close()


class TestManualSpecialCosts(unittest.TestCase):
    def test_manual_event_driver_parser_exports_formula_to_explicit_row(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()
        period = get_fy_months(2027)[1]

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "event_name",
                        "count",
                        "unit_price",
                        "amount_vnd",
                        "account_code",
                        "form_row",
                        "description",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": period,
                        "event_name": "no_trip_gift",
                        "count": "3",
                        "unit_price": "4000",
                        "amount_vnd": "",
                        "account_code": "5004086291",
                        "form_row": "137",
                        "description": "Manual event count",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_event.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B137"].value, 5004086291)
                self.assertEqual(ws["G137"].value, "=3*4000")
                self.assertEqual(ws["S137"].value, "出向者の書類申請費/Chi phí làm giấy tờ cho người biệt phái")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_travel_event_driver_posts_may_to_row_66(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089)
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()
        period = "202705"

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_code",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": period,
                        "event_name": "社員旅行 Du lịch công ty",
                        "event_type": "month_specific_driver",
                        "count": "111",
                        "unit_price": "1874000",
                        "account_code": "5004086291",
                        "row": "66",
                        "note": "Company trip manual driver for May",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT source, period, cc_code, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(fact_row["period"], "202705")
            self.assertEqual(str(fact_row["cc_code"]), "1412000089")
            self.assertEqual(int(fact_row["account_code"]), 5004086291)
            self.assertEqual(int(fact_row["form_row"]), 66)
            self.assertEqual(float(fact_row["amount_vnd"]), 208014000.0)
            self.assertIn("formula_expr=111*1874000", fact_row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_travel_event.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B66"].value, 5004086291)
                self.assertEqual(ws["G66"].value, "=111*1874000")
                month_cells = ["F66", "H66", "I66", "J66", "K66", "L66", "M66", "N66", "O66", "P66", "Q66"]
                self.assertNotIn("=111*1874000", [ws[cell].value for cell in month_cells])
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_staff_notebook_event_driver_posts_to_row_97(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089, cost_type="製造")
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5005246288, '事務用消耗品費', 'Office supplies')"
        )
        conn.commit()
        period = "202705"

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_code",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": period,
                        "event_name": "新入社員：ノート（スタッフ用） Người mới: Sổ tay nhân viên",
                        "event_type": "manual_count_unit_price",
                        "count": "3",
                        "unit_price": "9100",
                        "account_code": "5005246288",
                        "row": "97",
                        "note": "Staff notebook row 97 test",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT period, cc_code, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(fact_row["period"], "202705")
            self.assertEqual(str(fact_row["cc_code"]), "1412000089")
            self.assertEqual(int(fact_row["account_code"]), 5005246288)
            self.assertEqual(int(fact_row["form_row"]), 97)
            self.assertEqual(float(fact_row["amount_vnd"]), 27300.0)
            self.assertIn("formula_expr=3*9100", fact_row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_staff_notebook.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B97"].value, 5005246288)
                self.assertEqual(ws["G97"].value, "=3*9100")
                self.assertEqual(ws["R97"].value, "=SUM(F97:Q97)")
                other_month_cells = ["F97", "H97", "I97", "J97", "K97", "L97", "M97", "N97", "O97", "P97", "Q97"]
                self.assertTrue(all(ws[cell].value is None for cell in other_month_cells))
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_worker_notebook_event_driver_posts_to_row_98(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089, cost_type="製造")
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5005246288, '事務用消耗品費', 'Office supplies')"
        )
        conn.commit()
        period = "202705"

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_code",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": period,
                        "event_name": "新入社員：ノート (G7社員用） Người mới: Sổ tay công nhân",
                        "event_type": "manual_count_unit_price",
                        "count": "4",
                        "unit_price": "4000",
                        "account_code": "5005246288",
                        "row": "98",
                        "note": "Worker notebook row 98 test",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT period, cc_code, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(fact_row["period"], "202705")
            self.assertEqual(str(fact_row["cc_code"]), "1412000089")
            self.assertEqual(int(fact_row["account_code"]), 5005246288)
            self.assertEqual(int(fact_row["form_row"]), 98)
            self.assertEqual(float(fact_row["amount_vnd"]), 16000.0)
            self.assertIn("formula_expr=4*4000", fact_row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_worker_notebook.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B98"].value, 5005246288)
                self.assertEqual(ws["G98"].value, "=4*4000")
                self.assertEqual(ws["R98"].value, "=SUM(F98:Q98)")
                other_month_cells = ["F98", "H98", "I98", "J98", "K98", "L98", "M98", "N98", "O98", "P98", "Q98"]
                self.assertTrue(all(ws[cell].value is None for cell in other_month_cells))
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_recruitment_health_event_driver_source_month_posts_next_month_to_row_58(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089, cost_type="製造")
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "source_month",
                        "posting_rule",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_code",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "source_month": "202706",
                        "posting_rule": "next_month",
                        "event_name": "採用の健康診断費 Chi phí khám sức khỏe tuyển dụng",
                        "event_type": "manual_count_unit_price",
                        "count": "2",
                        "unit_price": "123000",
                        "account_code": "5004086291",
                        "row": "58",
                        "note": "Recruitment health source month next month test",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT period, cc_code, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(fact_row["period"], "202707")
            self.assertEqual(str(fact_row["cc_code"]), "1412000089")
            self.assertEqual(int(fact_row["account_code"]), 5004086291)
            self.assertEqual(int(fact_row["form_row"]), 58)
            self.assertEqual(float(fact_row["amount_vnd"]), 246000.0)
            self.assertIn("formula_expr=2*123000", fact_row["description"])
            self.assertIn("source_month=202706", fact_row["description"])
            self.assertIn("posting_rule=next_month", fact_row["description"])
            self.assertIn("shifted_to=202707", fact_row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_recruitment_health.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B58"].value, 5004086291)
                self.assertIsNone(ws["H58"].value)
                self.assertEqual(ws["I58"].value, "=2*123000")
                self.assertEqual(ws["R58"].value, "=SUM(F58:Q58)")
                self.assertIsNone(ws["I57"].value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_recruitment_health_source_month_outside_fy_fails_closed(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089, cost_type="製造")
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "source_month",
                        "posting_rule",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_code",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "source_month": "202803",
                        "posting_rule": "next_month",
                        "event_name": "採用の健康診断費 Chi phí khám sức khỏe tuyển dụng",
                        "event_type": "manual_count_unit_price",
                        "count": "1",
                        "unit_price": "123000",
                        "account_code": "5004086291",
                        "row": "58",
                        "note": "Recruitment health outside FY test",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 1)
            self.assertIn("outside FY", str(result.get("error_message", "")))
            count = conn.execute(
                "SELECT COUNT(*) FROM fact_input_data WHERE source = 'manual_event_driver' AND form_row = 58"
            ).fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_resolves_account_jp_name_for_travel_manufacturing(self):
        conn = _mk_conn()
        cc_code = 1412000089
        conn.execute(
            """
            INSERT INTO dim_cost_centers
            (code, name_jp, seq_no, saisan_type, cost_type, staff_count, worker_count)
            VALUES (?, 'MFG_CC', 1, '一般', '製造', 0, 0)
            """,
            (str(cc_code),),
        )
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, mfg_code, ga_code, sales_code)
            VALUES (5004086291, '福利厚生費', 'Welfare', 5004086291, 6004086651, 6004086551)
            """
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_jp_name",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "event_type": "month_specific_driver",
                        "count": "111",
                        "unit_price": "1874000",
                        "account_jp_name": "福利厚生費",
                        "row": "66",
                        "note": "Company trip manual driver for May",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT period, cc_code, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(fact_row["period"], "202705")
            self.assertEqual(str(fact_row["cc_code"]), "1412000089")
            self.assertEqual(int(fact_row["account_code"]), 5004086291)
            self.assertEqual(int(fact_row["form_row"]), 66)
            self.assertEqual(float(fact_row["amount_vnd"]), 208014000.0)
            self.assertIn("formula_expr=111*1874000", fact_row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_travel_event_resolved.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B66"].value, 5004086291)
                self.assertEqual(ws["G66"].value, "=111*1874000")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_resolves_account_jp_name_for_sales(self):
        conn = _mk_conn()
        cc_code = 1412000072
        conn.execute(
            """
            INSERT INTO dim_cost_centers
            (code, name_jp, seq_no, saisan_type, cost_type, staff_count, worker_count)
            VALUES (?, 'SALES_CC', 1, '一般', '販売', 0, 0)
            """,
            (str(cc_code),),
        )
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, mfg_code, ga_code, sales_code)
            VALUES (6004086551, '福利厚生費', 'Welfare', 5004086291, 6004086651, 6004086551)
            """
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_jp_name",
                        "row",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "event_type": "month_specific_driver",
                        "count": "1",
                        "unit_price": "100",
                        "account_jp_name": "福利厚生費",
                        "row": "66",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)
            account_code = conn.execute(
                "SELECT account_code FROM fact_input_data WHERE source = 'manual_event_driver'"
            ).fetchone()["account_code"]
            self.assertEqual(int(account_code), 6004086551)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_requires_account_code_or_account_jp_name(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089)
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "target_month", "event_name", "count", "unit_price", "row"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "count": "1",
                        "unit_price": "100",
                        "row": "66",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 1)
            count = conn.execute("SELECT COUNT(*) FROM fact_input_data").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_rejects_unknown_account_jp_name(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000089)
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "count",
                        "unit_price",
                        "account_jp_name",
                        "row",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "count": "1",
                        "unit_price": "100",
                        "account_jp_name": "DOES_NOT_EXIST",
                        "row": "66",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 1)
            count = conn.execute("SELECT COUNT(*) FROM fact_input_data").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def _seed_travel_unit_price_lookup(self, conn):
        conn.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES ('GA', '社員旅行', '福利厚生費', 5004086291, 6004086651, 6004086551,
                    '5月', 1874000, '/person', 'headcount_all', 'manual event')
            """
        )

    def _seed_travel_resolver_master(self, conn):
        cc_code = 1412000089
        conn.execute(
            """
            INSERT INTO dim_cost_centers
            (code, name_jp, seq_no, saisan_type, cost_type, staff_count, worker_count)
            VALUES (?, 'MFG_CC', 1, '一般', '製造', 0, 0)
            """,
            (str(cc_code),),
        )
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, mfg_code, ga_code, sales_code)
            VALUES (5004086291, '福利厚生費', 'Welfare', 5004086291, 6004086651, 6004086551)
            """
        )
        return cc_code

    def test_manual_event_driver_resolves_unit_price_key_for_travel(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        self._seed_travel_unit_price_lookup(conn)
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price_key",
                        "account_jp_name",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "event_type": "month_specific_driver",
                        "count": "111",
                        "unit_price_key": "社員旅行",
                        "account_jp_name": "福利厚生費",
                        "row": "66",
                        "note": "Company trip manual driver for May",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT account_code, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(int(fact_row["account_code"]), 5004086291)
            self.assertEqual(float(fact_row["amount_vnd"]), 208014000.0)
            self.assertIn("formula_expr=111*1874000", fact_row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_travel_unit_price_key.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["G66"].value, "=111*1874000")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_travel_absent_gift_event_driver_parses_june_driver(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        event_name = "社員旅行不参加対象者へのギフト贈呈 Quà tặng cho CNV không thể tham gia du lịch"
        unit_price_key = "社員旅行不参加対象者へのギフト贈呈 Quà tặng cho CNV không thể tham gia du lịch."
        conn.execute(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES ('GA', ?, '福利厚生費', 5004086291, 6004086651, 6004086551,
                    '6月', 1312500, '/person', 'working_days', 'manual event')
            """,
            (unit_price_key,),
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "unit_price_key",
                        "account_jp_name",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202606",
                        "event_name": event_name,
                        "event_type": "month_specific_driver",
                        "count": "7",
                        "unit_price": "1312500",
                        "unit_price_key": unit_price_key,
                        "account_jp_name": "福利厚生費",
                        "note": "Sample: travel absent gift June driver",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            fact_row = conn.execute(
                """
                SELECT source, period, cc_code, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(fact_row)
            self.assertEqual(fact_row["source"], "manual_event_driver")
            self.assertEqual(fact_row["period"], "202606")
            self.assertEqual(str(fact_row["cc_code"]), str(cc_code))
            self.assertEqual(int(fact_row["account_code"]), 5004086291)
            self.assertIsNone(fact_row["form_row"])
            self.assertEqual(float(fact_row["amount_vnd"]), 9187500.0)
            self.assertIn(event_name, fact_row["description"])
            self.assertIn("formula_expr=7*1312500", fact_row["description"])
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_bus_jp_event_driver_all_months_posts_to_row_53(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        periods = get_fy_months(2027)
        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_jp_name",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "all",
                        "event_name": "出向者BUS送迎費 Chi phí xe bus người JP",
                        "event_type": "manual_count_unit_price",
                        "count": "10",
                        "unit_price": "12345",
                        "account_jp_name": "福利厚生費",
                        "row": "53",
                        "note": "Bus JP all months test",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 12)
            self.assertEqual(result["errors"], 0)

            fact_rows = conn.execute(
                """
                SELECT period, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                ORDER BY period
                """
            ).fetchall()
            self.assertEqual([row["period"] for row in fact_rows], sorted(periods))
            for row in fact_rows:
                self.assertEqual(int(row["account_code"]), 5004086291)
                self.assertEqual(int(row["form_row"]), 53)
                self.assertEqual(float(row["amount_vnd"]), 123450.0)
                self.assertIn("formula_expr=10*12345", row["description"])
                self.assertIn("repeat=all_months", row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_bus_jp_all_months.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B53"].value, 5004086291)
                for column in range(6, 18):
                    self.assertEqual(ws.cell(53, column).value, "=10*12345")
                self.assertEqual(ws["R53"].value, "=SUM(F53:Q53)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_bus_vn_event_driver_all_months_posts_to_row_54(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        periods = get_fy_months(2027)
        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_jp_name",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "all",
                        "event_name": "ローカル社BUS送迎費 Chi phí xe bus người VN",
                        "event_type": "manual_count_unit_price",
                        "count": "20",
                        "unit_price": "23456",
                        "account_jp_name": "福利厚生費",
                        "row": "54",
                        "note": "Bus VN all months test",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 12)
            self.assertEqual(result["errors"], 0)

            fact_rows = conn.execute(
                """
                SELECT period, account_code, form_row, amount_vnd, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                ORDER BY period
                """
            ).fetchall()
            self.assertEqual([row["period"] for row in fact_rows], sorted(periods))
            for row in fact_rows:
                self.assertEqual(int(row["account_code"]), 5004086291)
                self.assertEqual(int(row["form_row"]), 54)
                self.assertEqual(float(row["amount_vnd"]), 469120.0)
                self.assertIn("formula_expr=20*23456", row["description"])
                self.assertIn("repeat=all_months", row["description"])

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_bus_vn_all_months.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B54"].value, 5004086291)
                for column in range(6, 18):
                    self.assertEqual(ws.cell(54, column).value, "=20*23456")
                self.assertEqual(ws["R54"].value, "=SUM(F54:Q54)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_explicit_unit_price_wins_over_key(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        self._seed_travel_unit_price_lookup(conn)
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "target_month", "event_name", "count", "unit_price", "unit_price_key", "account_jp_name", "row"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "count": "111",
                        "unit_price": "123",
                        "unit_price_key": "社員旅行",
                        "account_jp_name": "福利厚生費",
                        "row": "66",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)
            row = conn.execute("SELECT amount_vnd, description FROM fact_input_data").fetchone()
            self.assertEqual(float(row["amount_vnd"]), 13653.0)
            self.assertIn("formula_expr=111*123", row["description"])
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_rejects_unknown_unit_price_key(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "target_month", "event_name", "count", "unit_price_key", "account_jp_name", "row"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "count": "111",
                        "unit_price_key": "DOES_NOT_EXIST",
                        "account_jp_name": "福利厚生費",
                        "row": "66",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 1)
            count = conn.execute("SELECT COUNT(*) FROM fact_input_data").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_rejects_duplicate_unit_price_key(self):
        conn = _mk_conn()
        cc_code = self._seed_travel_resolver_master(conn)
        self._seed_travel_unit_price_lookup(conn)
        self._seed_travel_unit_price_lookup(conn)
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "target_month", "event_name", "count", "unit_price_key", "account_jp_name", "row"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": "202705",
                        "event_name": "社員旅行 Du lịch công ty",
                        "count": "111",
                        "unit_price_key": "社員旅行",
                        "account_jp_name": "福利厚生費",
                        "row": "66",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 1)
            count = conn.execute("SELECT COUNT(*) FROM fact_input_data").fetchone()[0]
            self.assertEqual(count, 0)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_without_form_row_appends_with_formula(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()
        period = get_fy_months(2027)[0]

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "event_name",
                        "count",
                        "unit_price",
                        "amount_vnd",
                        "account_code",
                        "form_row",
                        "description",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": period,
                        "event_name": "other",
                        "count": "2",
                        "unit_price": "5000",
                        "amount_vnd": "",
                        "account_code": "5004086291",
                        "form_row": "",
                        "description": "Append manual event",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_event_append.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                appended_row = next(
                    row_index
                    for row_index in range(168, 260)
                    if ws.cell(row_index, 2).value == 5004086291
                )
                self.assertEqual(ws.cell(appended_row, 6).value, "=2*5000")
                self.assertEqual(ws.cell(appended_row, 19).value, "other: Append manual event")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_template_includes_friendly_schema_columns(self):
        from src.parsers.manual_event_drivers import ensure_manual_event_drivers_template

        tmpdir = _mk_tmpdir()
        try:
            csv_path = Path(ensure_manual_event_drivers_template(str(tmpdir), 2027))
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                header = next(csv.reader(f))

            self.assertEqual(
                header,
                [
                    "cc_code",
                    "period",
                    "target_month",
                    "source_month",
                    "posting_rule",
                    "event_name",
                    "event_type",
                    "count",
                    "unit_price",
                    "unit_price_key",
                    "allocation_content",
                    "amount_vnd",
                    "account_code",
                    "account_jp_name",
                    "account_name",
                    "account_group",
                    "form_row",
                    "row",
                    "headcount_basis",
                    "description",
                    "note",
                ],
            )
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_accepts_new_alias_columns(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()
        period = get_fy_months(2027)[2]

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "target_month",
                        "event_name",
                        "event_type",
                        "count",
                        "unit_price",
                        "account_code",
                        "row",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "target_month": period,
                        "event_name": "month_specific_driver",
                        "event_type": "manual_count_unit_price",
                        "count": "4",
                        "unit_price": "2500",
                        "account_code": "5004086291",
                        "row": "137",
                        "note": "Alias driven event",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 0)

            row = conn.execute(
                """
                SELECT period, amount_vnd, form_row, description
                FROM fact_input_data
                WHERE source = 'manual_event_driver'
                """
            ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(row["period"], period)
            self.assertEqual(float(row["amount_vnd"]), 10000.0)
            self.assertEqual(int(row["form_row"]), 137)
            self.assertEqual(row["description"], "month_specific_driver: Alias driven event|formula_expr=4*2500")
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_rejects_conflicting_alias_values(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()
        periods = get_fy_months(2027)

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "target_month",
                        "event_name",
                        "count",
                        "unit_price",
                        "account_code",
                        "form_row",
                        "row",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": periods[0],
                        "target_month": periods[1],
                        "event_name": "period conflict",
                        "count": "1",
                        "unit_price": "100",
                        "account_code": "5004086291",
                        "form_row": "46",
                        "row": "46",
                    }
                )
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": periods[0],
                        "target_month": periods[0],
                        "event_name": "row conflict",
                        "count": "1",
                        "unit_price": "100",
                        "account_code": "5004086291",
                        "form_row": "46",
                        "row": "47",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 0)
            self.assertEqual(result["errors"], 2)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_driver_validates_event_type(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            "INSERT INTO dim_accounts (code, name_jp, name_vn) VALUES (5004086291, '福利厚生費', 'Welfare')"
        )
        conn.commit()
        period = get_fy_months(2027)[0]

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "cc_code",
                        "period",
                        "event_name",
                        "event_type",
                        "amount_vnd",
                        "account_code",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": period,
                        "event_name": "manual amount",
                        "event_type": "manual_amount",
                        "amount_vnd": "1234",
                        "account_code": "5004086291",
                    }
                )
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": period,
                        "event_name": "bad type",
                        "event_type": "unknown_event",
                        "amount_vnd": "5678",
                        "account_code": "5004086291",
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["errors"], 1)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_nnn_paperwork_workbook_parser_exports_to_row_137(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)

        tmpdir = _mk_tmpdir()
        try:
            workbook_path = tmpdir / "Dự tính chi phí làm giấy tờ cho NNN FY2027.xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "FY2027"
            ws.cell(row=2, column=4, value="Mã costcenter")
            ws.cell(row=2, column=5, value="Mã tài khoản")
            ws.cell(row=2, column=6, value="2026/04")
            ws.cell(row=2, column=7, value="2026/05")
            ws.cell(row=3, column=2, value="VN000001")
            ws.cell(row=3, column=3, value="TEST USER")
            ws.cell(row=3, column=4, value=cc_code)
            ws.cell(row=3, column=5, value=5005246286)
            ws.cell(row=3, column=6, value=1000)
            ws.cell(row=3, column=7, value=2000)
            workbook.save(workbook_path)
            workbook.close()

            result = parse_nnn_paperwork(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 2)
            self.assertEqual(result["errors"], 0)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_nnn_workbook.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B137"].value, 5005246286)
                self.assertEqual(ws["F137"].value, "=1000")
                self.assertEqual(ws["G137"].value, "=2000")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_birthday_workbook_parser_exports_to_row_59(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code)
            VALUES (5004086291, '福利厚生費', 'Welfare', NULL, NULL, 5004086291, 6004086551, 7004086777)
            """
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            workbook_path = tmpdir / "Sinh nhật MP FY2027.xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Sinh nhật MP FY2027"
            ws.cell(row=3, column=1, value="CostCenter")
            ws.cell(row=4, column=1, value=cc_code)
            ws.cell(row=4, column=2, value="TEST_CC")
            ws.cell(row=4, column=3, value=3)
            ws.cell(row=4, column=4, value=4)
            workbook.save(workbook_path)
            workbook.close()

            result = parse_birthday_workbook(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 2)
            self.assertEqual(result["errors"], 0)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_birthday_workbook.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B59"].value, 6004086551)
                self.assertEqual(ws["F59"].value, "=3*152000")
                self.assertEqual(ws["G59"].value, "=4*152000")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_special_cost_parser_and_export_to_explicit_form_row(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "special_costs_manual.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "period", "form_row", "account_code", "amount_vnd", "description"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "cc_code": str(cc_code),
                        "period": periods[0],
                        "form_row": "113",
                        "account_code": "5005246286",
                        "amount_vnd": "500",
                        "description": "manual row 113",
                    }
                )

            result = parse_manual_special_costs(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], 1)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_special_cost.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["F113"].value, "=500")
                self.assertIsNone(ws["G137"].value)
                self.assertIsNone(ws["O138"].value)
                self.assertIsNone(ws["S200"].value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_recurring_admin_rows_use_previous_month_headcount_formulas(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, description)
            VALUES
            ('ga_unit_price', ?, 100, 0, 0, 'gas|headcount_per_person'),
            ('ga_unit_price', ?, 110, 0, 0, 'gas|headcount_per_person'),
            ('ga_unit_price', ?, 120, 0, 0, 'gas|headcount_per_person'),
            ('ga_unit_price', ?, 200, 0, 0, '手洗い洗剤|headcount_per_person'),
            ('ga_unit_price', ?, 300, 0, 0, 'giay ve sinh|headcount_per_person'),
            ('ga_unit_price', ?, 400, 0, 0, 'cleaning|headcount_per_person'),
            ('manual', ?, 1, ?, 500001, 'seed')
            """,
            (
                periods[0],
                periods[1],
                periods[2],
                periods[1],
                periods[1],
                periods[2],
                periods[0],
                cc_code,
            ),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_prev_month_formula.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                # Gas: April uses April headcount; May uses April, not May; June uses May.
                self.assertEqual(ws["F46"].value, "=SUM(F$24:F$25)*100")
                self.assertEqual(ws["G46"].value, "=SUM(F$24:F$25)*110")
                self.assertNotIn("G$24:G$25", ws["G46"].value)
                self.assertEqual(ws["H46"].value, "=SUM(G$24:G$25)*120")

                # Other recurring administrative allocations follow the same previous-month pattern.
                self.assertEqual(ws["G48"].value, "=SUM(F$24:F$25)*200")
                self.assertEqual(ws["G49"].value, "=SUM(F$24:F$25)*300")
                self.assertEqual(ws["H51"].value, "=SUM(G$24:G$25)*400")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)


class TestHubBuilderExport(unittest.TestCase):
    def test_export_preserves_form_layout_and_appends_unmapped_rows(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code)
            VALUES (500001, 'TEST_ACC', 'Test Account', 'G', 'GV', 500001, 600001, 700001)
            """
        )
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, description)
            VALUES
            ('manual', ?, 123, ?, 500001, 'desc'),
            ('manual', ?, 456, ?, 500001, 'desc')
            """,
            (periods[0], cc_code, periods[1], cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out.xlsx"
            builder = HubBuilder(conn, fiscal_year=2027)
            ok = builder.export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B5"].value, cc_code)
                appended_row = next(
                    row_index
                    for row_index in range(168, 260)
                    if ws.cell(row_index, 2).value == 500001
                )
                self.assertEqual(ws.cell(appended_row, 19).value, "desc")
                self.assertEqual(ws.cell(appended_row, 6).value, "=123")
                self.assertEqual(ws.cell(appended_row, 7).value, "=456")
                self.assertEqual(ws.cell(appended_row, 18).value, f"=SUM(F{appended_row}:Q{appended_row})")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_export_routes_fixed_items_to_form_rows_and_clears_sample_rows(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        conn.executemany(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code)
            VALUES (?, 'TEST_ACC', 'Test Account', 'G', 'GV', ?, ?, ?)
            """,
            [
                (5004086291, 5004086291, 6004086651, 6004086551),
                (5005246281, 5005246281, 6005146627, 6005146541),
            ],
        )
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, description)
            VALUES
            ('alloc_238', ?, 111, ?, 5004086291, 'Alloc: 社員旅行 Du lịch công ty'),
            ('alloc_240', ?, 222, ?, 5004086291, 'Alloc: 京セラフェスティバルLễ hội Kyocera'),
            ('alloc_201', ?, 333, ?, 5005246281, 'Alloc: 社員証（新入社員用・再発行時、写真含む）\nThẻ từ chấm công + ảnh'),
            ('manual', ?, 444, ?, 500001, 'desc')
            """,
            (periods[1], cc_code, periods[5], cc_code, periods[3], cc_code, periods[0], cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_fixed_rows.xlsx"
            builder = HubBuilder(conn, fiscal_year=2027)
            ok = builder.export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["G66"].value, "=111")
                self.assertEqual(ws["K66"].value, "=222")
                self.assertEqual(ws["I79"].value, "=333")
                self.assertIsNone(ws["F67"].value)
                appended_row = next(
                    row_index
                    for row_index in range(168, 260)
                    if ws.cell(row_index, 19).value == "desc"
                )
                self.assertEqual(ws.cell(appended_row, 19).value, "desc")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_fixed_rows_follow_mp2027_form_layout(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        period = periods[0]
        conn.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, form_row, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("facility", period, 0, 1.5, cc_code, 0, None, "depreciation_building"),
                ("facility", period, 0, 2.5, cc_code, 0, None, "depreciation_land"),
                ("fixed_assets", period, 0, 3.5, cc_code, 0, None, "fixed_assets_depr|asset"),
                ("facility", period, 0, 4.5, cc_code, 0, None, "interest_building"),
                ("facility", period, 0, 5.5, cc_code, 0, None, "interest_land"),
                ("fixed_assets", period, 0, 6.5, cc_code, 0, None, "fixed_assets_interest|asset"),
                ("facility", period, 100, 0, cc_code, 0, None, "electric"),
                ("facility", period, 200, 0, cc_code, 0, None, "water"),
                ("ga_unit_price", period, 300, 0, 0, 0, None, "gas|headcount_per_person"),
                ("ga_unit_price", period, 400, 0, 0, 0, None, "nuoc rua tay|headcount_per_person"),
                ("ga_unit_price", period, 500, 0, 0, 0, None, "giay ve sinh|headcount_per_person"),
                ("ga_unit_price", period, 600, 0, 0, 0, None, "cleaning|headcount_per_person"),
                ("alloc_health", period, 700, 0, cc_code, 5004086291, None, "Alloc: kham suc khoe (cho cnv nam)"),
                ("alloc_new_hire_health", period, 800, 0, cc_code, 5004086291, None, "Alloc: kham suc khoe khi tuyen dung"),
                ("alloc_birthday", period, 900, 0, cc_code, 5004086291, None, "Alloc: sinh nhat"),
                ("it_sim", period, 0, 30, cc_code, 5005246282, None, "it_sim|component_term|vpn|qty=10|unit_usd=3"),
                ("alloc_note_staff", period, 1000, 0, cc_code, 5005246288, None, "Alloc: note staff"),
                ("alloc_note_worker", period, 1100, 0, cc_code, 5005246288, None, "Alloc: note worker"),
                ("manual_special_cost", period, 1200, 0, cc_code, 5005246286, 137, "manual NNN"),
            ],
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_fixed_form_layout.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                expected_accounts = {
                    36: 5006016260,
                    37: 5006016261,
                    38: 5006016244,
                    40: 9114120007,
                    41: 9114120007,
                    42: 9114120007,
                    44: 5005066281,
                    45: 5005066282,
                    46: 5005056281,
                    48: 5005016372,
                    49: 5005016372,
                    51: 5005246286,
                    57: 5004086291,
                    58: 5004086291,
                    59: 5004086291,
                    97: 5005246288,
                    98: 5005246288,
                    137: 5005246286,
                }
                for row_index, account_code in expected_accounts.items():
                    self.assertEqual(ws.cell(row_index, 2).value, account_code, f"B{row_index}")

                self.assertIn("Khấu hao (Nhà)", ws["S36"].value)
                self.assertIn("Lãi (Nhà)", ws["S40"].value)
                self.assertIn("Tiền điện", ws["S44"].value)
                self.assertIn("Hand wash", ws["S48"].value)
                self.assertIn("Chi phí khám sức khỏe hàng năm", ws["S57"].value)
                self.assertIn("Tiền sinh nhật", ws["S59"].value)
                system_rows = _find_system_cost_rows(ws)
                self.assertEqual(system_rows, [75])
                system_row = system_rows[0]
                self.assertIn("System Cost", ws.cell(system_row, 19).value)
                self.assertIn("Sổ tay", ws["S97"].value)
                self.assertIn("Chi phí làm giấy tờ", ws["S137"].value)

                self.assertEqual(ws["F36"].value, "=ROUND(1.5*$B$2,0)")
                self.assertEqual(ws["F37"].value, "=ROUND(2.5*$B$2,0)")
                self.assertEqual(ws["F38"].value, "=ROUND(3.5*$B$2,0)")
                self.assertEqual(ws["F40"].value, "=ROUND(4.5*$B$2,0)")
                self.assertEqual(ws["F41"].value, "=ROUND(5.5*$B$2,0)")
                self.assertEqual(ws["F42"].value, "=ROUND(6.5*$B$2,0)")
                self.assertEqual(ws["F44"].value, "=100")
                self.assertEqual(ws["F45"].value, "=200")
                self.assertEqual(ws["F46"].value, "=SUM(F$24:F$25)*300")
                self.assertEqual(ws["F48"].value, "=SUM(F$24:F$25)*400")
                self.assertEqual(ws["F49"].value, "=SUM(F$24:F$25)*500")
                self.assertEqual(ws["F51"].value, "=SUM(F$24:F$25)*600")
                self.assertEqual(ws["F57"].value, "=700")
                self.assertEqual(ws["F58"].value, "=800")
                self.assertEqual(ws["F59"].value, "=900")
                self.assertEqual(ws.cell(system_row, 6).value, "=ROUND((10*3)*$B$2,0)")
                self.assertEqual(ws.cell(system_row, 18).value, f"=SUM(F{system_row}:Q{system_row})")
                self.assertEqual(ws["F97"].value, "=1000")
                self.assertEqual(ws["F98"].value, "=1100")
                self.assertEqual(ws["F137"].value, "=1200")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_manual_event_reference_items_export_to_expected_rows_and_months(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        other_cc = _seed_cc(conn, code=1412999999)
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code)
            VALUES (5004086291, '福利厚生費', 'Welfare', NULL, NULL, 5004086291, 5004086291, 5004086291)
            """
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            csv_path = tmpdir / "event_drivers_manual.csv"
            fieldnames = [
                "cc_code",
                "period",
                "target_month",
                "source_month",
                "posting_rule",
                "event_name",
                "event_type",
                "count",
                "unit_price",
                "unit_price_key",
                "allocation_content",
                "amount_vnd",
                "account_code",
                "account_jp_name",
                "account_name",
                "account_group",
                "form_row",
                "row",
                "headcount_basis",
                "description",
                "note",
            ]
            cases = [
                ("FY2027部門方針発表会後の決起コンパ", 2, 1000, "F54", "=2*1000"),
                ("Tiệc khuấy động năm tài chính決起コンパ", 3, 1000, "G54", "=3*1000"),
                ("マイエピソード ～フィロソフィの実践～参加賞", 5, 1000, "I81", "=5*1000"),
                ("京セラフェスティバル", 6, 1000, "K66", "=6*1000"),
                ("月餅 Bánh Trung Thu", 7, 1000, "K71", "=7*1000"),
                ("10年勤続記念コンパ", 8, 1000, "L64", "=8*1000"),
                ("10年勤続記念品", 9, 1000, "L65", "=9*1000"),
                ("会社設立記念 感謝イベント", 10, 1000, "L68", "=10*1000"),
                ("ポケットカレンダー", 11, 1000, "M82", "=11*1000"),
                ("運動会 大 hội thể thao", 12, 1000, "M67", "=12*1000"),
                ("忘年会補助金", 13, 1000, "P70", "=13*1000"),
                ("お年玉 Tiền lì xì", 14, 1000, "P63", "=14*1000"),
            ]
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                for event_name, count, unit_price, _cell, _formula in cases:
                    writer.writerow(
                        {
                            "cc_code": cc_code,
                            "event_name": event_name,
                            "count": count,
                            "unit_price": unit_price,
                            "account_code": 5004086291,
                        }
                    )
                writer.writerow(
                    {
                        "cc_code": cc_code,
                        "event_name": "社員旅行不参加対象者へのギフト贈呈",
                        "count": 4,
                        "unit_price": 1000,
                        "account_code": 5004086291,
                    }
                )
                writer.writerow(
                    {
                        "cc_code": other_cc,
                        "event_name": "月餅 Bánh Trung Thu",
                        "count": 99,
                        "unit_price": 1000,
                        "account_code": 5004086291,
                    }
                )

            result = parse_manual_event_drivers(conn, source_dir=str(tmpdir))
            self.assertEqual(result["inserted"], len(cases) + 2)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_manual_events.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                for _event_name, _count, _unit_price, cell, formula in cases:
                    self.assertEqual(ws[cell].value, formula)
                self.assertIsNone(ws["H66"].value)
                appended_gift_rows = [
                    row_index
                    for row_index in range(168, ws.max_row + 1)
                    if "社員旅行不参加対象者へのギフト贈呈" in str(ws.cell(row=row_index, column=19).value or "")
                ]
                self.assertEqual(len(appended_gift_rows), 1)
                self.assertEqual(ws.cell(row=appended_gift_rows[0], column=8).value, "=4*1000")
                for row_index in (54, 63, 64, 65, 66, 67, 68, 70, 71, 81, 82):
                    self.assertEqual(ws.cell(row=row_index, column=18).value, f"=SUM(F{row_index}:Q{row_index})")
                self.assertNotEqual(ws["K71"].value, "=99*1000")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_ga_parser_loads_bus_unit_price_rows_from_admin_main_sheet(self):
        conn = _mk_conn()
        expat_prices = [856107, 894022, 841125, 843756, 840449, 826939, 807491, 851412, 890880, 868440, 823640, 810900]
        vn_prices = [1031546, 1666547, 1572350, 1523385, 1411084, 1510555, 1530445, 1556432, 1469160, 1461107, 1162654, 1116875]
        tmpdir = _mk_tmpdir()
        try:
            source_path = tmpdir / "総務課 FY2027 MP 振替予定.xlsx"
            months = get_fy_months(2027)
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "FY2027予定"
            ws.append(["Item FY2027 VND account", *months, None, "Basis VND Yotei", "account"])
            ws.append(["出向者送迎費\nXe đưa đón người Nhật", *expat_prices, None, "一人当たり", 5004086291])
            ws.append(["ローカル社員送迎費\nXe đưa đón người Việt", *vn_prices, None, "一人当たり", 5004086291])
            for _ in range(5):
                ws.append(["padding", *([0] * 12), None, "headcount", None])
            workbook.create_sheet("Cách tính phân bổ 振替計算")
            workbook.save(source_path)
            workbook.close()

            result = parse_ga(conn, source_dir=str(tmpdir))
            self.assertEqual(result["total"], 24)

            expat_rows = conn.execute(
                """
                SELECT period, amount_vnd, description
                FROM fact_input_data
                WHERE source='ga_unit_price'
                  AND description LIKE '%出向者送迎費%'
                ORDER BY period
                """
            ).fetchall()
            vn_rows = conn.execute(
                """
                SELECT period, amount_vnd, description
                FROM fact_input_data
                WHERE source='ga_unit_price'
                  AND description LIKE '%ローカル社員送迎費%'
                ORDER BY period
                """
            ).fetchall()

            self.assertEqual([row["period"] for row in expat_rows], months)
            self.assertEqual([float(row["amount_vnd"]) for row in expat_rows], [float(price) for price in expat_prices])
            self.assertTrue(all(row["description"].endswith("|headcount_per_person") for row in expat_rows))
            self.assertEqual([row["period"] for row in vn_rows], months)
            self.assertEqual([float(row["amount_vnd"]) for row in vn_rows], [float(price) for price in vn_prices])
            self.assertTrue(all(row["description"].endswith("|headcount_per_person") for row in vn_rows))
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_ga_admin_allocation_native_parser_exports_requirement_rows(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        other_cc = _seed_cc(conn, code=1412999999)

        tmpdir = _mk_tmpdir()
        try:
            source_path = tmpdir / "FY2027_MP_hanh_chinh.xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Cach tinh phan bo"
            months = get_fy_months(2027)

            # Main sheet signature for GA workbook discovery.
            ws.append(["Item FY2027", *months, None, None, "Basis VND Yotei 費"])
            ws.append(["Dummy", *([0] * 12), None, None, "headcount"])
            ws.append([])
            ws.append(["cc_code", "form_row", "account_code", "description", *months])
            ws.append([cc_code, 46, 5005056281, "Gas 食堂燃料", *range(101, 113)])
            ws.append([cc_code, 48, 5005016372, "Nước rửa tay", *range(201, 213)])
            ws.append([cc_code, 51, 5005246286, "Cleaning làm sạch", *range(301, 313)])
            ws.append([cc_code, 54, 5004086291, "Tiệc khuấy động", *range(401, 413)])
            ws.append([cc_code, 56, 5004086291, "Company sale", *range(501, 513)])
            ws.append([other_cc, 46, 5005056281, "Other CC must not leak", *([999999] * 12)])
            ws.row_dimensions[8].hidden = True
            ws.append([])
            ws.append(["cc_code", "form_row", "account_code", "description", "period", "quantity", "unit_price", "amount"])
            ws.append([cc_code, 97, 5005246288, "Sổ tay nhân viên mới", "2026/05", 3, 130000, None])
            ws.append([cc_code, 98, 5005246288, "Sổ tay công nhân mới", "2026/05", 4, 40000, None])
            ws.append([cc_code, 58, 5004086291, "Khám sức khỏe khi tuyển dụng", "2026/05", 2, 5000, None])
            ws.append([cc_code, 58, 5004086291, "Khám sức khỏe khi tuyển dụng extra", "2026/05", 1, 7000, None])
            ws.append([other_cc, 97, 5005246288, "Other new staff", "2026/05", 99, 130000, None])
            workbook.save(source_path)
            workbook.close()

            result = parse_ga(conn, source_dir=str(tmpdir))
            self.assertEqual(result["admin_allocation"], 77)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_admin_allocation.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                for row_index, start_value in ((46, 101), (48, 201), (51, 301), (54, 401), (56, 501)):
                    self.assertEqual(
                        [ws.cell(row=row_index, column=column_index).value for column_index in range(6, 18)],
                        [f"={value}" for value in range(start_value, start_value + 12)],
                    )
                    self.assertEqual(ws.cell(row=row_index, column=18).value, f"=SUM(F{row_index}:Q{row_index})")
                self.assertEqual(ws["G97"].value, "=3*130000")
                self.assertEqual(ws["G98"].value, "=4*40000")
                self.assertIn(ws["G58"].value, ("=2*5000+1*7000", "=1*7000+2*5000"))
                self.assertEqual(ws["R97"].value, "=SUM(F97:Q97)")
                self.assertEqual(ws["R98"].value, "=SUM(F98:Q98)")
                self.assertEqual(ws["R58"].value, "=SUM(F58:Q58)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_fixed_assets_source_driven_rows_38_42_handle_last_depreciation_months(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        other_cc = _seed_cc(conn, code=1412999999)

        tmpdir = _mk_tmpdir()
        try:
            source_path = tmpdir / "固定資産情報_Fixed_Assets_Information_2025.11 - Nov.xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "2025.11"

            def write_asset(row_index, cc, monthly_depr, last_month, last_month_depr, apr_interest, may_interest):
                ws.cell(row=row_index, column=3, value=f"A{row_index}")
                ws.cell(row=row_index, column=4, value=f"Asset {row_index}")
                ws.cell(row=row_index, column=8, value=cc)
                ws.cell(row=row_index, column=12, value=monthly_depr)
                ws.cell(row=row_index, column=16, value=last_month)
                ws.cell(row=row_index, column=17, value=last_month_depr)
                ws.cell(row=row_index, column=22, value=apr_interest)
                ws.cell(row=row_index, column=23, value=may_interest)

            write_asset(5, cc_code, 10.0, "2027/11/30", 9.0, 1.0, 2.0)
            write_asset(6, cc_code, 20.0, "2026/05/31", 7.0, 3.0, 4.0)
            write_asset(7, cc_code, 30.0, "2026/11/30", 8.0, 5.0, 6.0)
            write_asset(8, other_cc, 1000.0, "2027/11/30", 900.0, 100.0, 200.0)
            ws.row_dimensions[7].hidden = True
            workbook.save(source_path)
            workbook.close()

            result = parse_fixed_assets(conn, fa_path=str(source_path))
            self.assertEqual(result["total"], 68)

            template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
            output_path = tmpdir / "out_fixed_assets_source.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(
                    [ws.cell(row=38, column=column_index).value for column_index in range(6, 18)],
                    [
                        "=ROUND(60*$B$2,0)", "=ROUND(47*$B$2,0)", "=ROUND(40*$B$2,0)",
                        "=ROUND(40*$B$2,0)", "=ROUND(40*$B$2,0)", "=ROUND(40*$B$2,0)",
                        "=ROUND(40*$B$2,0)", "=ROUND(18*$B$2,0)", "=ROUND(10*$B$2,0)",
                        "=ROUND(10*$B$2,0)", "=ROUND(10*$B$2,0)", "=ROUND(10*$B$2,0)",
                    ],
                )
                self.assertEqual(
                    [ws.cell(row=42, column=column_index).value for column_index in range(6, 18)],
                    [
                        "=ROUND(9*$B$2,0)", "=ROUND(12*$B$2,0)", "=ROUND(8*$B$2,0)",
                        "=ROUND(8*$B$2,0)", "=ROUND(8*$B$2,0)", "=ROUND(8*$B$2,0)",
                        "=ROUND(8*$B$2,0)", "=ROUND(8*$B$2,0)", "=ROUND(2*$B$2,0)",
                        "=ROUND(2*$B$2,0)", "=ROUND(2*$B$2,0)", "=ROUND(2*$B$2,0)",
                    ],
                )
                self.assertEqual(ws["R38"].value, "=SUM(F38:Q38)")
                self.assertEqual(ws["R42"].value, "=SUM(F42:Q42)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_row_discovery_does_not_hardcode_row_75(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, cost_type="製造")
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES ('it_sim', ?, 0, 31.9, ?, 0, 'it_sim|component_term|vpn|qty=10|unit_usd=3.19')
            """,
            (periods[0], cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            moved_template = tmpdir / "form_system_row_moved.xlsx"
            shutil.copy2(template_path, moved_template)
            workbook = openpyxl.load_workbook(moved_template)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                ws["B75"].value = None
                ws["S75"].value = None
                ws["B76"].value = 5005246282
                ws["S76"].value = "System Cost moved template row"
                workbook.save(moved_template)
            finally:
                workbook.close()

            output_path = tmpdir / "out_it_moved_row.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(moved_template), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_rows = _find_system_cost_rows(ws)
                self.assertEqual(system_rows, [76])
                self.assertEqual(ws["B76"].value, 5005246282)
                self.assertIsNone(ws["F75"].value)
                self.assertEqual(ws["F76"].value, "=ROUND((10*3.19)*$B$2,0)")
                self.assertEqual(ws["R76"].value, "=SUM(F76:Q76)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_export_clears_unused_template_accounts_below_row_30(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000006, cost_type="陬ｽ騾")
        period = get_fy_months(2027)[0]
        conn.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("facility", period, 100, 0, cc_code, 0, "electric"),
                ("alloc_birthday", period, 900, 0, cc_code, 5004086291, "Alloc: sinh nhat"),
                ("it_sim", period, 0, 31.9, cc_code, 5005246282, "it_sim|component_term|vpn|qty=10|unit_usd=3.19"),
            ],
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_cleared_template_accounts.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_row = _find_system_cost_rows(ws)[0]
                self.assertEqual(ws["B44"].value, 5005066281)
                self.assertEqual(ws["B59"].value, 5004086291)
                self.assertEqual(ws.cell(system_row, 2).value, 5005246282)
                self.assertIsNone(ws["B31"].value)
                self.assertIsNone(ws["B66"].value)
                self.assertIsNone(ws["B130"].value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_health_check_export_deduplicates_periodic_and_hiring_business_items(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        period = get_fy_months(2027)[8]
        conn.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, form_row, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    "alloc_health_male",
                    period,
                    700,
                    0,
                    cc_code,
                    5004086291,
                    None,
                    "Alloc: Khám sức khỏe (cho CNV nam)",
                ),
                (
                    "alloc_health_hiring",
                    period,
                    800,
                    0,
                    cc_code,
                    5004086291,
                    None,
                    "Alloc: Khám sức khỏe khi tuyển dụng",
                ),
            ],
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_health_check_dedup.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["N57"].value, "=700")
                self.assertEqual(ws["N58"].value, "=800")
                self.assertTrue(
                    all(ws.cell(row=69, column=column_index).value is None for column_index in range(6, 18))
                )
                self.assertEqual(ws["R57"].value, "=SUM(F57:Q57)")
                self.assertEqual(ws["R58"].value, "=SUM(F58:Q58)")
                self.assertEqual(ws["R69"].value, "=SUM(F69:Q69)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_ga_admin_allocation_default_account_resolves_by_cc_cost_type(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        conn.execute(
            """
            INSERT INTO dim_accounts
            (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code)
            VALUES (5004086291, '福利厚生費', 'Welfare', NULL, NULL, 5004086291, 6004086551, 7004086777)
            """
        )
        conn.commit()

        tmpdir = _mk_tmpdir()
        try:
            source_path = tmpdir / "FY2027_MP_hanh_chinh.xlsx"
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Cach tinh phan bo"
            months = get_fy_months(2027)
            ws.append(["Item FY2027", *months, None, None, "Basis VND Yotei 雑費"])
            ws.append(["Dummy", *([0] * 12), None, None, "headcount"])
            ws.append([])
            ws.append(["cc_code", "form_row", "account_code", "description", *months])
            ws.append([cc_code, 54, None, "Tiệc khuấy động", *range(401, 413)])
            ws.append([cc_code, 54, None, "padding", *([0] * 12)])
            ws.append([cc_code, 54, None, "padding", *([0] * 12)])
            ws.append([cc_code, 54, None, "padding", *([0] * 12)])
            workbook.save(source_path)
            workbook.close()

            result = parse_ga(conn, source_dir=str(tmpdir))
            self.assertEqual(result["admin_allocation"], 12)

            row = conn.execute(
                """
                SELECT account_code, form_row
                FROM fact_input_data
                WHERE source = 'ga_admin_allocation'
                ORDER BY period
                LIMIT 1
                """
            ).fetchone()
            self.assertEqual(int(row["account_code"]), 6004086551)
            self.assertEqual(int(row["form_row"]), 54)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_export_rebuilds_accounts_when_template_column_b_is_blank_from_row_30(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, code=1412000006, cost_type="陬ｽ騾")
        period = get_fy_months(2027)[0]
        conn.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("facility", period, 100, 0, cc_code, 0, "electric"),
                ("it_sim", period, 0, 31.9, cc_code, 5005246282, "it_sim|component_term|vpn|qty=10|unit_usd=3.19"),
            ],
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            blank_b_template = tmpdir / "form_blank_b_from_30.xlsx"
            shutil.copy2(template_path, blank_b_template)
            workbook = openpyxl.load_workbook(blank_b_template)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                for row_index in range(30, ws.max_row + 1):
                    ws.cell(row=row_index, column=2).value = None
                workbook.save(blank_b_template)
            finally:
                workbook.close()

            output_path = tmpdir / "out_blank_b_from_30.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(blank_b_template), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_rows = _find_system_cost_rows(ws)
                self.assertEqual(system_rows, [75])
                self.assertEqual(ws["B44"].value, 5005066281)
                self.assertEqual(ws["B75"].value, 5005246282)
                self.assertEqual(ws["F75"].value, "=ROUND((10*3.19)*$B$2,0)")
                self.assertIsNone(ws["B66"].value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_account_resolves_from_cost_type_when_fact_account_is_zero(self):
        cases = [
            (1412000006, "製造", 5005246282),
            (1412000086, "一般", 6005146628),
            (1412000030, "販売", 6005146542),
        ]
        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        for cc_code, cost_type, expected_account in cases:
            with self.subTest(cost_type=cost_type):
                conn = _mk_conn()
                _seed_cc(conn, code=cc_code, cost_type=cost_type)
                period = get_fy_months(2027)[0]
                conn.execute(
                    """
                    INSERT INTO fact_input_data
                    (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
                    VALUES ('it_sim', ?, 0, 31.9, ?, 0, 'it_sim|component_term|vpn|qty=10|unit_usd=3.19')
                    """,
                    (period, cc_code),
                )
                conn.commit()

                tmpdir = _mk_tmpdir()
                try:
                    output_path = tmpdir / f"out_it_kdc_{cost_type}.xlsx"
                    ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
                    self.assertTrue(ok)

                    workbook = openpyxl.load_workbook(output_path, data_only=False)
                    try:
                        ws = workbook[find_hub_sheet_name(workbook)]
                        system_rows = _find_system_cost_rows(ws)
                        self.assertEqual(system_rows, [75])
                        system_row = system_rows[0]
                        self.assertEqual(ws.cell(system_row, 2).value, expected_account)
                        self.assertEqual(ws.cell(system_row, 6).value, "=ROUND((10*3.19)*$B$2,0)")
                    finally:
                        workbook.close()
                finally:
                    conn.close()
                    shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_account_prefers_valid_fact_kdc_account(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn, cost_type="製造")
        period = get_fy_months(2027)[0]
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES ('it_sim', ?, 0, 31.9, ?, 6005146628, 'it_sim|component_term|vpn|qty=10|unit_usd=3.19')
            """,
            (period, cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_it_fact_account.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_row = _find_system_cost_rows(ws)[0]
                self.assertEqual(ws.cell(system_row, 2).value, 6005146628)
                self.assertEqual(ws.cell(system_row, 6).value, "=ROUND((10*3.19)*$B$2,0)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_account_missing_cost_center_fails_closed(self):
        conn = _mk_conn()
        cc_code = 1412000999
        period = get_fy_months(2027)[0]
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES ('it_sim', ?, 0, 31.9, ?, 0, 'it_sim|component_term|vpn|qty=10|unit_usd=3.19')
            """,
            (period, cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_it_missing_cc.xlsx"
            with self.assertRaisesRegex(RuntimeError, "Unable to resolve KDC System Cost account"):
                HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_row_prefers_detail_formula_terms(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES
            ('it_sim', ?, 0, 31.9, ?, 5005246282, 'it_sim|component_term|vpn|qty=10|unit_usd=3.19'),
            ('it_sim', ?, 0, 230.2, ?, 5005246282, 'it_sim|component_term|mail|qty=20|unit_usd=11.51'),
            ('it_sim', ?, 0, 60.04, ?, 5005246282, 'it_sim|component_term|mes|qty=2|unit_usd=30.02'),
            ('it_sim', ?, 0, 45, ?, 5005246282, 'it_sim|component_term|vps|qty=20|unit_usd=2.25'),
            ('it_sim', ?, 9645870, 0, ?, 5005246282, 'it_sim|system_usage_total')
            """,
            (periods[0], cc_code, periods[0], cc_code, periods[0], cc_code, periods[0], cc_code, periods[0], cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_it_formula.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_rows = _find_system_cost_rows(ws)
                self.assertEqual(system_rows, [75])
                system_row = system_rows[0]
                self.assertEqual(ws.cell(system_row, 2).value, 5005246282)
                self.assertEqual(
                    ws.cell(system_row, 6).value,
                    "=ROUND((10*3.19+20*11.51+2*30.02+20*2.25)*$B$2,0)",
                )
                self.assertEqual(ws.cell(system_row, 18).value, f"=SUM(F{system_row}:Q{system_row})")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_summary_fallback_accepts_metadata_description(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        period = get_fy_months(2027)[0]
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES (
                'it_sim', ?, 9645870, 0, ?, 5005246282,
                'it_sim|system_usage_total|source_file=abc.xls|source_sheet=summary|source_filter=cc:1412000006|audit_status=OK|audit_diff_vnd=0'
            )
            """,
            (period, cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_it_summary_metadata.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_row = _find_system_cost_rows(ws)[0]
                self.assertEqual(ws.cell(system_row, 6).value, 9645870)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_it_system_component_fallback_accepts_metadata_description(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        period = get_fy_months(2027)[0]
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description)
            VALUES (
                'it_sim', ?, 0, 31.9, ?, 5005246282,
                'it_sim|component|vpn|source_file=abc.xls|source_sheet=VPN|source_filter=cc:1412000006'
            )
            """,
            (period, cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_it_component_metadata.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                system_row = _find_system_cost_rows(ws)[0]
                self.assertEqual(ws.cell(system_row, 6).value, "=ROUND((31.9)*$B$2,0)")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_nnn_paperwork_exports_to_row_137_f_to_q(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
            VALUES
            ('nnn_paperwork', ?, 1000000, ?, 5005246286, 137, 'base', 'NNN paperwork: VN0001 Worker A|formula_expr=1000000'),
            ('nnn_paperwork', ?, 2000000, ?, 5005246286, 137, 'base', 'NNN paperwork: VN0001 Worker A|formula_expr=2000000'),
            ('nnn_paperwork', ?, 12000000, ?, 5005246286, 137, 'base', 'NNN paperwork: VN0001 Worker A|formula_expr=12000000')
            """,
            (periods[0], cc_code, periods[1], cc_code, periods[11], cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_nnn_f_to_q.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B137"].value, 5005246286)
                self.assertEqual(ws["F137"].value, "=1000000")
                self.assertEqual(ws["G137"].value, "=2000000")
                self.assertEqual(ws["Q137"].value, "=12000000")
                self.assertEqual(ws["R137"].value, "=SUM(F137:Q137)")
                # Other months should be cleared/empty
                for month_col in ("H", "I", "J", "K", "L", "M", "N", "O", "P"):
                    self.assertIsNone(ws[f"{month_col}137"].value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_nnn_paperwork_sums_multiple_records_on_row_137(self):
        conn = _mk_conn()
        cc_code = _seed_cc(conn)
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
            VALUES
            ('nnn_paperwork', ?, 1000000, ?, 5005246286, 137, 'base', 'NNN paperwork: VN0001 Worker A|formula_expr=1000000'),
            ('nnn_paperwork', ?, 6000000, ?, 5005246286, 137, 'base', 'NNN paperwork: VN0002 Worker B|formula_expr=2000*3')
            """,
            (periods[0], cc_code, periods[0], cc_code),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_nnn_sum.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertEqual(ws["B137"].value, 5005246286)
                self.assertEqual(ws["F137"].value, "=1000000+2000*3")
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_nnn_paperwork_does_not_export_unknown_cost_center(self):
        conn = _mk_conn()
        cc_code_target = _seed_cc(conn, code=1412000004)
        cc_code_other = _seed_cc(conn, code=1412000018)
        periods = get_fy_months(2027)
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
            VALUES
            ('nnn_paperwork', ?, 1000000, ?, 5005246286, 137, 'base', 'NNN paperwork: VN0001 Worker A|formula_expr=1000000')
            """,
            (periods[0], cc_code_other),
        )
        conn.commit()

        # Seed at least one dummy record for the target CC so that HubBuilder knows the CC has fact input data
        conn.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
            VALUES
            ('dummy', ?, 1, ?, 500001, 200, 'base', 'dummy')
            """,
            (periods[0], cc_code_target),
        )
        conn.commit()

        template_path = Path(__file__).resolve().parents[1] / "docs" / "MP2027" / "FORM.xlsx"
        tmpdir = _mk_tmpdir()
        try:
            output_path = tmpdir / "out_nnn_other_cc.xlsx"
            ok = HubBuilder(conn, fiscal_year=2027).export_to_template(str(template_path), str(output_path), cc_code=cc_code_target)
            self.assertTrue(ok)

            workbook = openpyxl.load_workbook(output_path, data_only=False)
            try:
                ws = workbook[find_hub_sheet_name(workbook)]
                self.assertIsNone(ws["F137"].value)
            finally:
                workbook.close()
        finally:
            conn.close()
            shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
