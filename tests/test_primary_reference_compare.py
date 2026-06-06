import json
from pathlib import Path

import openpyxl

from tools.compare_primary_reference import compare_workbooks


def _make_workbook(path: Path, overrides=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "内訳ﾘｽﾄ(4～3月)"
    for row in (38, 42, 53, 54, 58, 66, 75, 97, 98, 137):
        ws[f"B{row}"] = 5000000000 + row
        for col in range(6, 18):
            ws.cell(row=row, column=col).value = f"=SUM({row},{col})"
        ws[f"R{row}"] = f"=SUM(F{row}:Q{row})"
        ws[f"S{row}"] = f"description {row}"
    ws["S38"] = "減価償却費（設備）/Khấu hao (Thiết bị)"
    ws["S42"] = "固定資産金利（設備）/Lãi (Thiết bị)"
    ws["S75"] = "System Cost (Mail,VPN,R3, MES,PLM,VPS)"
    for coord, value in (overrides or {}).items():
        ws[coord] = value
    wb.save(path)
    wb.close()


def _payload(result):
    return json.loads(Path(result["json_path"]).read_text(encoding="utf-8"))


def test_compare_primary_reference_identical_workbooks(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(reference)
    _make_workbook(generated)

    result = compare_workbooks(generated, reference, out_dir)

    assert result["summary"]["differences"] == 0
    assert result["summary"]["fixed_rows_compared"] == 1
    assert result["summary"]["identity_rows_checked"] == 9
    assert Path(result["xlsx_path"]).is_file()
    assert Path(result["json_path"]).is_file()


def test_compare_primary_reference_detects_fixed_row_b_account_high_severity(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(reference)
    _make_workbook(generated, {"B66": 9999999999})

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    assert result["summary"]["fixed_row_differences"] > 0
    b66 = [
        row for row in payload["important_row_diff"]
        if row["row"] == 66 and row["column"] == "B" and row["compare_mode"] == "fixed_row"
    ][0]
    assert b66["match"] is False
    assert b66["severity"] == "High"


def test_compare_primary_reference_detects_fixed_row_fq_medium_severity(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(reference)
    _make_workbook(generated, {"F66": "=999"})

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    f66 = [
        row for row in payload["important_row_diff"]
        if row["row"] == 66 and row["column"] == "F" and row["compare_mode"] == "fixed_row"
    ][0]
    assert result["summary"]["fixed_row_differences"] > 0
    assert f66["match"] is False
    assert f66["severity"] == "Medium"


def test_compare_primary_reference_aligns_system_cost_identity_row(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated)
    _make_workbook(reference, {
        "B75": 5005026371,
        "S75": "Temperature Sensor",
        "B80": 7770000075,
        "F80": "=SUM(80,6)",
        "R80": "=SUM(F80:Q80)",
        "S80": "System Cost Mail VPN R3 MES PLM VPS",
    })

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 75][0]
    assert alignment["reference_matched_row"] == 80
    assert alignment["match_method"] == "token_overlap"
    assert alignment["confidence"] == "Medium"
    assert not any(
        row for row in payload["important_row_diff"]
        if row["row"] == 75 and row["reference_row"] == 75 and row["compare_mode"] == "fixed_row"
    )


def test_compare_primary_reference_aligns_depreciation_identity_row(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated)
    _make_workbook(reference, {
        "B38": 5006016260,
        "S38": "建物",
        "B40": 7770000038,
        "R40": "=SUM(F40:Q40)",
        "S40": "減価償却 設備 depreciation",
    })

    result = compare_workbooks(generated, reference, out_dir)

    alignment = [row for row in _payload(result)["identity_row_alignment"] if row["generated_row"] == 38][0]
    assert alignment["reference_matched_row"] == 40
    assert alignment["match_method"] == "token_overlap"


def test_bus_rows_are_identity_aligned_not_fixed(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated, {
        "B54": 5004086291,
        "S54": "Local BUS transport xe bus nguoi VN",
    })
    _make_workbook(reference, {
        "B54": 5005016371,
        "S54": "Cooling spray Nabakem SF-1013",
        "B287": 5004086291,
        "F287": "=5*1031546",
        "R287": "=SUM(F287:Q287)",
        "S287": "local bus transport",
    })

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 54][0]
    assert alignment["reference_matched_row"] == 287
    assert alignment["match_method"] == "same_account"
    assert not any(
        row for row in payload["important_row_diff"]
        if row["row"] == 54 and row["compare_mode"] == "fixed_row"
    )


def test_bus_jp_identity_prefers_expats_transport(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated, {
        "B53": 5004086291,
        "S53": "日本人出向者バス transport",
    })
    _make_workbook(reference, {
        "B286": 5004086291,
        "S286": "出向者 社員送迎費",
        "B287": 5004086291,
        "S287": "ローカル社員送迎費",
    })

    result = compare_workbooks(generated, reference, out_dir)

    alignment = [row for row in _payload(result)["identity_row_alignment"] if row["generated_row"] == 53][0]
    assert alignment["reference_matched_row"] == 286
    assert alignment["match_method"] == "same_account"


def test_recruitment_health_row_is_identity_aligned_not_fixed(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated, {
        "B58": 5005246288,
        "S58": "Recruitment health checkup kham suc khoe tuyen dung",
    })
    _make_workbook(reference, {
        "B58": 5005016371,
        "S58": "Business trip meal unrelated item",
        "B210": 5005246288,
        "R210": "=SUM(F210:Q210)",
        "S210": "Recruitment health checkup 健康診断 入社",
    })

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 58][0]
    assert alignment["reference_matched_row"] == 210
    assert alignment["match_method"] == "same_account"
    assert not any(
        row for row in payload["important_row_diff"]
        if row["row"] == 58 and row["compare_mode"] == "fixed_row"
    )


def test_staff_notebook_row_is_identity_aligned_not_fixed(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated, {
        "B97": 5005246288,
        "S97": "Staff notebook So tay nhan vien",
    })
    _make_workbook(reference, {
        "B97": 5005016371,
        "S97": "Battery purchase unrelated item",
        "B210": 5005246288,
        "R210": "=SUM(F210:Q210)",
        "S210": "Staff notebook ノート nhan vien",
    })

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 97][0]
    assert alignment["reference_matched_row"] == 210
    assert alignment["match_method"] == "same_account"
    assert not any(
        row for row in payload["important_row_diff"]
        if row["row"] == 97 and row["compare_mode"] == "fixed_row"
    )


def test_worker_notebook_row_is_identity_aligned_not_fixed(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated, {
        "B98": 5005246288,
        "S98": "Worker G7 notebook So tay cong nhan",
    })
    _make_workbook(reference, {
        "B98": 5005016371,
        "S98": "SSD purchase unrelated item",
        "B211": 5005246288,
        "R211": "=SUM(F211:Q211)",
        "S211": "Worker G7 notebook ノート cong nhan",
    })

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 98][0]
    assert alignment["reference_matched_row"] == 211
    assert alignment["match_method"] == "same_account"
    assert not any(
        row for row in payload["important_row_diff"]
        if row["row"] == 98 and row["compare_mode"] == "fixed_row"
    )


def test_nnn_row_is_identity_aligned_not_fixed(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated, {
        "B137": 5004086291,
        "S137": "NNN paperwork giay to document",
    })
    _make_workbook(reference, {
        "B137": 5005016371,
        "S137": "Business trip meal",
        "B210": 5004086291,
        "R210": "=SUM(F210:Q210)",
        "S210": "NNN paperwork document",
    })

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 137][0]
    assert alignment["reference_matched_row"] == 210
    assert alignment["match_method"] == "same_account"
    assert not any(
        row for row in payload["important_row_diff"]
        if row["row"] == 137 and row["compare_mode"] == "fixed_row"
    )


def test_compare_primary_reference_records_identity_not_found(tmp_path):
    reference = tmp_path / "reference.xlsx"
    generated = tmp_path / "generated.xlsx"
    out_dir = tmp_path / "reports"
    _make_workbook(generated)
    _make_workbook(reference, {
        "B75": 5005026371,
        "S75": "Temperature Sensor",
    })
    for row in range(1, 120):
        if row != 75:
            wb = None
    # Remove all strong system tokens from reference while keeping workbook valid.
    ref_wb = openpyxl.load_workbook(reference)
    ws = ref_wb["内訳ﾘｽﾄ(4～3月)"]
    for row in range(1, ws.max_row + 1):
        if row != 75 and ws[f"S{row}"].value and "System" in str(ws[f"S{row}"].value):
            ws[f"S{row}"] = "unrelated detail"
            ws[f"B{row}"] = 1234567890
    ref_wb.save(reference)
    ref_wb.close()

    result = compare_workbooks(generated, reference, out_dir)

    payload = _payload(result)
    assert result["summary"]["identity_rows_not_found"] > 0
    alignment = [row for row in payload["identity_row_alignment"] if row["generated_row"] == 75][0]
    assert alignment["match_method"] == "not_found"
    assert alignment["reference_matched_row"] is None
