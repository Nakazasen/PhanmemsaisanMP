from pathlib import Path
from shutil import copy2
from openpyxl import load_workbook
from src.engine.system_cost_writer import apply_system_cost_to_workbook
from src.utils import excel_helpers as helpers
TEMPLATE=Path('FORM.xlsx')
SYSTEM=[Path('raw/システム課金金額(Simulation)_FY2027_Apr.2026 ~ June.2026.xls'),Path('raw/システム課金金額(Simulation)_FY2027_July.2026 ~ Dec.2026(Change AMS & PLM price).xls'),Path('raw/システム課金金額(Simulation)_FY2027_Jan.2027 ~ March.2027(Change SAP price).xls')]
def test_system_cost_writer_writes_row_211(tmp_path):
 out=tmp_path/'out.xlsx'; copy2(TEMPLATE,out); apply_system_cost_to_workbook(out,SYSTEM)
 wb=load_workbook(out); ws=wb[helpers.find_hub_sheet_name(wb)]
 try:
  assert ws.cell(211,5).value=='system_cost_combined'
  assert ws.cell(211,19).value=='System Cost / システム課金'
  assert ws.cell(211,6).value==6111363.0
  assert ws.cell(211,17).value==6111363.0
 finally: wb.close()
def test_system_cost_writer_blank_row_212(tmp_path):
 out=tmp_path/'out.xlsx'; copy2(TEMPLATE,out); apply_system_cost_to_workbook(out,SYSTEM)
 wb=load_workbook(out); ws=wb[helpers.find_hub_sheet_name(wb)]
 try:
  assert ws.cell(212,5).value is None
  assert ws.cell(212,6).value is None
  assert ws.cell(212,17).value is None
  assert ws.cell(212,19).value is None
 finally: wb.close()
def test_system_cost_writer_does_not_modify_sources_or_template(tmp_path):
 out=tmp_path/'out.xlsx'; copy2(TEMPLATE,out); mt=TEMPLATE.stat().st_mtime_ns; ms=[p.stat().st_mtime_ns for p in SYSTEM]
 apply_system_cost_to_workbook(out,SYSTEM)
 assert TEMPLATE.stat().st_mtime_ns==mt
 assert [p.stat().st_mtime_ns for p in SYSTEM]==ms
