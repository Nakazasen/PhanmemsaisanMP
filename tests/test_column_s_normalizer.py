from openpyxl import Workbook

from src.engine.column_s_normalizer import (
    cell_has_month_cost,
    normalize_output_description_column_s,
    worksheet_row_has_cost,
)


def _worksheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "内訳ﾘｽﾄ(4～3月)"
    return workbook, worksheet


def test_formula_cost_blank_s_gets_description_from_row_label():
    workbook, worksheet = _worksheet()
    try:
        worksheet.cell(30, 2).value = "new hire notebook"
        worksheet.cell(30, 6).value = "=1*9100"

        stats = normalize_output_description_column_s(worksheet)

        assert worksheet.cell(30, 19).value == "new hire notebook"
        assert stats["cost_rows_filled_description"] == 1
    finally:
        workbook.close()


def test_numeric_cost_blank_s_gets_description_from_row_label():
    workbook, worksheet = _worksheet()
    try:
        worksheet.cell(213, 2).value = "expat regular hours"
        worksheet.cell(213, 6).value = 153

        normalize_output_description_column_s(worksheet)

        assert worksheet.cell(213, 19).value == "expat regular hours"
    finally:
        workbook.close()


def test_no_cost_nonblank_s_is_cleared():
    workbook, worksheet = _worksheet()
    try:
        worksheet.cell(31, 19).value = "stale template text"
        worksheet.cell(31, 6).value = None
        worksheet.cell(31, 7).value = 0

        stats = normalize_output_description_column_s(worksheet)

        assert worksheet.cell(31, 19).value is None
        assert stats["no_cost_rows_cleared_description"] == 1
    finally:
        workbook.close()


def test_cost_row_with_existing_s_is_preserved():
    workbook, worksheet = _worksheet()
    try:
        worksheet.cell(40, 2).value = "fallback label"
        worksheet.cell(40, 6).value = 100
        worksheet.cell(40, 19).value = "valid source description"

        normalize_output_description_column_s(worksheet)

        assert worksheet.cell(40, 19).value == "valid source description"
    finally:
        workbook.close()


def test_formula_cost_detection_does_not_need_excel_recalculation():
    workbook, worksheet = _worksheet()
    try:
        worksheet.cell(50, 6).value = "=28*9100"
        worksheet.cell(51, 6).value = "=0"

        assert cell_has_month_cost("=28*9100") is True
        assert worksheet_row_has_cost(worksheet, 50) is True
        assert worksheet_row_has_cost(worksheet, 51) is False
    finally:
        workbook.close()


def test_reference_filled_cost_row_uses_b_label_when_s_blank():
    workbook, worksheet = _worksheet()
    try:
        worksheet.cell(219, 2).value = "expat people"
        worksheet.cell(219, 6).value = 1
        worksheet.cell(219, 20).value = "REFERENCE_FILLED_FROM_PRIMARY; primary_row=24; not source-derived"

        normalize_output_description_column_s(worksheet)

        assert worksheet.cell(219, 19).value == "expat people"
        assert "REFERENCE_FILLED_FROM_PRIMARY" in worksheet.cell(219, 20).value
    finally:
        workbook.close()
