import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from src.parsers.extracted_headcount_time_plan import parse_extracted_headcount_time_plan
from src.services.reference_staffing_extractor import (
    ExtractedDepartment,
    ExtractionError,
    _run_render_worker,
    evaluate_cell,
    extract_reference_staffing_sources,
)

ROOT = Path(__file__).resolve().parents[1]
REFERENCE = ROOT / "reference_outputs" / "secondary" / "FY2027"
OFFICIAL = ROOT / "raw" / "10.07.2026"


class FormulaEvaluatorTests(unittest.TestCase):
    def test_arithmetic_and_cell_references(self):
        wb=Workbook(); ws=wb.active; ws["A1"]=4; ws["B1"]="=A1*8.5-(2+3)"
        self.assertEqual(evaluate_cell(ws,"B1"),29)

    def test_rejects_excel_functions(self):
        wb=Workbook(); ws=wb.active; ws["A1"]="=SUM(B1:B2)"
        with self.assertRaises(ValueError): evaluate_cell(ws,"A1")


class ReferenceExtractionTests(unittest.TestCase):
    def test_real_65_file_extraction_is_complete_and_traceable(self):
        with tempfile.TemporaryDirectory() as tmp:
            result = extract_reference_staffing_sources(REFERENCE, tmp, 2027, OFFICIAL)
            self.assertEqual(result["files"], 65)
            self.assertEqual(len(result["errors"]), 0, [x.errors for x in result["errors"]])
            self.assertEqual(result["ready"] + result["split_required"], 65)
            self.assertGreater(result["ready"], 0)
            self.assertGreater(result["split_required"], 0)
            self.assertEqual(len({x.cc_code for x in result["results"]}), 65)
            self.assertEqual(len({x.template_source for x in result["results"]}), 1)
            self.assertTrue(Path(result["manifest_path"]).exists())

            ready = [item for item in result["results"] if item.status == "READY"]
            split_required = [
                item for item in result["results"] if item.status == "SPLIT_REQUIRED"
            ]
            self.assertTrue(
                all(item.split_source and not item.mismatch_details for item in ready)
            )
            self.assertTrue(
                all(
                    not item.split_source
                    and all(
                        row[metric] is None
                        for row in item.rows
                        for metric in (
                            "headcount_staff",
                            "headcount_worker",
                            "fixed_hours_staff",
                            "fixed_hours_worker",
                            "overtime_hours_staff",
                            "overtime_hours_worker",
                        )
                    )
                    for item in split_required
                )
            )

            cc14 = next(x for x in result["results"] if x.cc_code == "1412000006")
            self.assertEqual(cc14.status, "READY")
            self.assertEqual(cc14.rows[3]["headcount_staff"], 27)
            self.assertEqual(cc14.rows[3]["overtime_hours_local"], 385)
            parsed_ready = parse_extracted_headcount_time_plan(cc14.output_path, 2027)
            self.assertEqual(parsed_ready.status, "valid", parsed_ready.errors)
            self.assertEqual(len(parsed_ready.rows), 12)

            cc02 = next(x for x in result["results"] if x.cc_code == "1412000005")
            self.assertEqual(cc02.status, "SPLIT_REQUIRED")
            self.assertTrue(cc02.candidate_split_source)
            self.assertTrue(cc02.mismatch_details)
            self.assertTrue(
                any("fixed_hours_local" in detail for detail in cc02.mismatch_details)
            )
            parsed_cc02 = parse_extracted_headcount_time_plan(cc02.output_path, 2027)
            self.assertEqual(parsed_cc02.status, "valid", parsed_cc02.errors)
            self.assertIsNone(parsed_cc02.rows[0]["headcount_staff"])
            self.assertIsNone(parsed_cc02.rows[0]["fixed_hours_staff"])
            self.assertIsNone(parsed_cc02.rows[0]["overtime_hours_staff"])

            manifest = load_workbook(result["manifest_path"], read_only=True, data_only=True)
            try:
                headers = [cell.value for cell in manifest["EXTRACTION_MANIFEST"][1]]
                self.assertIn("candidate_split_source", headers)
                self.assertIn("mismatch_details", headers)
            finally:
                manifest.close()

class WorkerIsolationTests(unittest.TestCase):
    def _assert_worker_fails_closed(self, command_factory, timeout_seconds=5.0):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            stage_dir = root / "stage"
            output_dir = root / "output"
            stage_dir.mkdir()
            output_dir.mkdir()
            item = ExtractedDepartment(
                source_path=str(root / "source.xlsx"),
                cc_code="1412000001",
                department_name="Test Department",
                status="READY",
                output_path=str(output_dir / "01.Test_FY2027_staffing_truth.xlsx"),
            )
            with patch(
                "src.services.reference_staffing_extractor._render_worker_command",
                side_effect=command_factory,
            ):
                with self.assertRaises(ExtractionError):
                    _run_render_worker(
                        [item],
                        root / "template.xls",
                        stage_dir,
                        output_dir,
                        2027,
                        None,
                        timeout_seconds=timeout_seconds,
                    )
            self.assertTrue((output_dir / "_staffing_render_diagnostic.log").is_file())
            self.assertFalse(Path(item.output_path).exists())

    def test_worker_nonzero_exit_without_response_fails_closed(self):
        self._assert_worker_fails_closed(
            lambda request, response, status: [sys.executable, "-c", "raise SystemExit(7)"]
        )

    def test_worker_corrupt_response_fails_closed(self):
        def command(request, response, status):
            script = (
                "from pathlib import Path; "
                f"Path({str(response)!r}).write_text('{{', encoding='utf-8')"
            )
            return [sys.executable, "-c", script]

        self._assert_worker_fails_closed(command)

    def test_worker_success_response_without_output_fails_closed(self):
        def command(request, response, status):
            payload = json.dumps(
                {
                    "protocol_version": 1,
                    "success": True,
                    "rendered_files": ["01.Test_FY2027_staffing_truth.xlsx"],
                }
            )
            script = (
                "from pathlib import Path; "
                f"Path({str(response)!r}).write_text({payload!r}, encoding='utf-8')"
            )
            return [sys.executable, "-c", script]

        self._assert_worker_fails_closed(command)

    def test_worker_timeout_fails_closed(self):
        self._assert_worker_fails_closed(
            lambda request, response, status: [
                sys.executable,
                "-c",
                "import time; time.sleep(30)",
            ],
            timeout_seconds=0.05,
        )


if __name__ == "__main__": unittest.main()
