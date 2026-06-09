from pathlib import Path

from openpyxl import load_workbook

from src.engine.facility_file_order_preview import preview_facility_file_order

SOURCE = Path("raw/施設課　MPFY2027.xlsx")


def test_facility_dry_run_plans_six_items():
    preview = preview_facility_file_order(SOURCE)

    assert len(preview.items) == 6
    assert [item.item_id for item in preview.items] == [
        "building_depreciation",
        "land_depreciation",
        "building_interest",
        "land_interest",
        "electricity",
        "water",
    ]


def test_facility_dry_run_uses_planner_rows_200_to_205():
    preview = preview_facility_file_order(SOURCE)

    assert preview.planned_start_row == 200
    assert preview.planned_end_row == 205
    assert [item.planned_row for item in preview.items] == list(range(200, 206))


def test_facility_dry_run_has_blank_row_206():
    preview = preview_facility_file_order(SOURCE)

    assert preview.blank_row_after == 206


def test_facility_dry_run_does_not_write_workbook(tmp_path):
    before_mtime = SOURCE.stat().st_mtime_ns

    preview = preview_facility_file_order(SOURCE)

    assert preview.source_path == str(SOURCE)
    assert SOURCE.stat().st_mtime_ns == before_mtime
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "減価償却費（Depreciation）",
            "固定資産金利（Interest）",
            "水道光熱費（Electric & Water）",
            "B&L",
            "E&W",
        ]
    finally:
        workbook.close()
    assert not list(tmp_path.iterdir())


def test_facility_dry_run_contains_building_and_land_interest_separately():
    preview = preview_facility_file_order(SOURCE)
    by_id = {item.item_id: item for item in preview.items}

    assert by_id["building_interest"].source_sheet == "固定資産金利（Interest）"
    assert by_id["land_interest"].source_sheet == "固定資産金利（Interest）"
    assert by_id["building_interest"].source_row != by_id["land_interest"].source_row
    assert by_id["building_interest"].formula_policy == "ROUND_USD_BY_B2"
    assert by_id["land_interest"].formula_policy == "ROUND_USD_BY_B2"


def test_facility_dry_run_preserves_mixed_cost_center_row_pairs():
    preview = preview_facility_file_order(SOURCE, cost_center="1412000040")
    by_id = {item.item_id: item for item in preview.items}

    assert by_id["building_depreciation"].source_row == by_id["land_depreciation"].source_row - 1
    assert by_id["building_interest"].source_row == by_id["land_interest"].source_row - 1
    assert by_id["electricity"].source_row == by_id["water"].source_row - 1

    assert by_id["building_depreciation"].monthly_values != by_id["land_depreciation"].monthly_values
    assert by_id["building_interest"].monthly_values != by_id["land_interest"].monthly_values
    assert by_id["electricity"].monthly_values != by_id["water"].monthly_values

    assert by_id["electricity"].monthly_values[0] == 1577160
    assert by_id["water"].monthly_values[0] == 529671
