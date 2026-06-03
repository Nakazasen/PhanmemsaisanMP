import sqlite3
import unittest
from pathlib import Path
import shutil
import uuid
import openpyxl

from src.db.schema import create_schema, init_sys_params
from src.parsers.nnn_paperwork import parse_nnn_paperwork
from src.utils.excel_helpers import get_fy_months

TEST_TMP_ROOT = Path.cwd() / ".tmp_test_artifacts"
TEST_TMP_ROOT.mkdir(exist_ok=True)


def _mk_tmpdir() -> Path:
    path = TEST_TMP_ROOT / f"case_{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    return path


def _mk_conn():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    create_schema(conn)
    init_sys_params(conn, exchange_rate=26273, fiscal_year=2027)
    return conn


def _seed_cc(conn, code):
    conn.execute(
        """
        INSERT INTO dim_cost_centers
        (code, name_jp, seq_no, saisan_type, cost_type, staff_count, worker_count)
        VALUES (?, 'TEST_CC', 1, '製造', '一般', 0, 0)
        """,
        (str(code),),
    )
    conn.commit()


class TestNNNPaperworkParser(unittest.TestCase):
    def setUp(self):
        self.tmpdirs = []

    def tearDown(self):
        for d in self.tmpdirs:
            shutil.rmtree(d, ignore_errors=True)

    def _create_temp_dir(self) -> Path:
        d = _mk_tmpdir()
        self.tmpdirs.append(d)
        return d

    def test_nnn_parser_reads_standard_header_layout(self):
        conn = _mk_conn()
        cc_code = "1412000018"
        _seed_cc(conn, cc_code)
        periods = get_fy_months(2027)

        tmpdir = self._create_temp_dir()
        workbook_path = tmpdir / "nnn_giay_to_standard.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FY2027"

        # Row 1: Tiêu đề sheet
        ws.append(["Dự toán chi phí NNN"])
        # Row 2: Header (ở đây ta dùng 2026/04 hoặc 2027/04 đều được map đúng về tháng 4, tức periods[0])
        ws.append([
            "STT",
            "Mã nhân viên",
            "Tên nhân viên",
            "Mã Cost Center",
            "Mã tài khoản",
            "2026/04",
            "2026/05",
            "2026/06"
        ])
        # Row 3: Data row
        ws.append([
            1,
            "VN0001",
            "Worker A",
            cc_code,
            6005186429,
            1000.0,
            2000.0,
            0.0
        ])

        wb.save(workbook_path)
        wb.close()

        result = parse_nnn_paperwork(conn, workbook_path=str(workbook_path))
        self.assertEqual(result["errors"], 0)
        self.assertEqual(result["inserted"], 2)  # periods[0] và periods[1]

        # Verify database contents
        rows = conn.execute(
            """
            SELECT period, amount_vnd, cc_code, account_code, form_row, description
            FROM fact_input_data
            WHERE source = 'nnn_paperwork'
            ORDER BY period
            """
        ).fetchall()

        self.assertEqual(len(rows), 2)

        # Record 1
        self.assertEqual(rows[0]["period"], periods[0])
        self.assertEqual(float(rows[0]["amount_vnd"]), 1000.0)
        self.assertEqual(str(rows[0]["cc_code"]), str(cc_code))
        self.assertEqual(rows[0]["account_code"], 6005186429)
        self.assertEqual(rows[0]["form_row"], 137)
        self.assertIn("Worker A", rows[0]["description"])

        # Record 2
        self.assertEqual(rows[1]["period"], periods[1])
        self.assertEqual(float(rows[1]["amount_vnd"]), 2000.0)
        self.assertEqual(str(rows[1]["cc_code"]), str(cc_code))
        self.assertEqual(rows[1]["account_code"], 6005186429)
        self.assertEqual(rows[1]["form_row"], 137)
        self.assertIn("Worker A", rows[1]["description"])

        conn.close()

    def test_nnn_parser_reads_reordered_columns_by_header(self):
        conn = _mk_conn()
        cc_code = "1412000018"
        _seed_cc(conn, cc_code)
        periods = get_fy_months(2027)

        tmpdir = self._create_temp_dir()
        workbook_path = tmpdir / "nnn_reordered.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FY2027"

        # Trộn thứ tự cột: Account Code trước Cost Center, đổi thứ tự tháng
        ws.append(["Mã tài khoản", "Mã Cost Center", "Tên nhân viên", "2026/05", "2026/04", "Mã nhân viên"])
        ws.append([6005186429, cc_code, "Worker B", 2500.0, 1500.0, "VN0002"])

        wb.save(workbook_path)
        wb.close()

        result = parse_nnn_paperwork(conn, workbook_path=str(workbook_path))
        self.assertEqual(result["errors"], 0)
        self.assertEqual(result["inserted"], 2)

        # Verify DB
        rows = conn.execute(
            """
            SELECT period, amount_vnd, cc_code, account_code, form_row
            FROM fact_input_data
            WHERE source = 'nnn_paperwork'
            ORDER BY period
            """
        ).fetchall()

        self.assertEqual(len(rows), 2)

        # periods[0] amount 1500.0
        self.assertEqual(rows[0]["period"], periods[0])
        self.assertEqual(float(rows[0]["amount_vnd"]), 1500.0)
        self.assertEqual(str(rows[0]["cc_code"]), str(cc_code))
        self.assertEqual(rows[0]["account_code"], 6005186429)
        self.assertEqual(rows[0]["form_row"], 137)

        # periods[1] amount 2500.0
        self.assertEqual(rows[1]["period"], periods[1])
        self.assertEqual(float(rows[1]["amount_vnd"]), 2500.0)
        self.assertEqual(str(rows[1]["cc_code"]), str(cc_code))
        self.assertEqual(rows[1]["account_code"], 6005186429)
        self.assertEqual(rows[1]["form_row"], 137)

        conn.close()

    def test_nnn_parser_reads_multiple_people_same_cost_center(self):
        conn = _mk_conn()
        cc_code = "1412000018"
        _seed_cc(conn, cc_code)
        periods = get_fy_months(2027)

        tmpdir = self._create_temp_dir()
        workbook_path = tmpdir / "nnn_multiple_people.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FY2027"

        ws.append(["Mã Cost Center", "Mã tài khoản", "Tên nhân viên", "Mã nhân viên", "2026/04"])
        # Người A: 1000
        ws.append([cc_code, 6005186429, "Worker A", "VN0001", 1000.0])
        # Người B cùng Cost Center: 3000
        ws.append([cc_code, 6005186429, "Worker B", "VN0002", 3000.0])

        wb.save(workbook_path)
        wb.close()

        result = parse_nnn_paperwork(conn, workbook_path=str(workbook_path))
        self.assertEqual(result["errors"], 0)
        self.assertEqual(result["inserted"], 2)

        rows = conn.execute(
            """
            SELECT period, amount_vnd, cc_code, account_code, form_row, description
            FROM fact_input_data
            WHERE source = 'nnn_paperwork'
            ORDER BY description
            """
        ).fetchall()

        self.assertEqual(len(rows), 2)
        # Record của Worker A
        self.assertEqual(rows[0]["period"], periods[0])
        self.assertEqual(float(rows[0]["amount_vnd"]), 1000.0)
        self.assertEqual(str(rows[0]["cc_code"]), str(cc_code))
        self.assertIn("Worker A", rows[0]["description"])

        # Record của Worker B
        self.assertEqual(rows[1]["period"], periods[0])
        self.assertEqual(float(rows[1]["amount_vnd"]), 3000.0)
        self.assertEqual(str(rows[1]["cc_code"]), str(cc_code))
        self.assertIn("Worker B", rows[1]["description"])

        conn.close()

    def test_nnn_parser_rejects_unknown_cost_center(self):
        conn = _mk_conn()
        cc_code_valid = "1412000018"
        _seed_cc(conn, cc_code_valid)

        tmpdir = self._create_temp_dir()
        workbook_path = tmpdir / "nnn_unknown_cc.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FY2027"

        ws.append(["Mã Cost Center", "Mã tài khoản", "Tên nhân viên", "Mã nhân viên", "2026/04"])
        # CC hợp lệ
        ws.append([cc_code_valid, 6005186429, "Worker Valid", "VN0001", 1000.0])
        # CC không hợp lệ
        ws.append(["9999999999", 6005186429, "Worker Invalid", "VN9999", 2000.0])

        wb.save(workbook_path)
        wb.close()

        result = parse_nnn_paperwork(conn, workbook_path=str(workbook_path))
        self.assertEqual(result["errors"], 1)  # 1 CC không hợp lệ
        self.assertEqual(result["inserted"], 1)  # Chỉ insert dòng 1

        rows = conn.execute("SELECT * FROM fact_input_data WHERE source = 'nnn_paperwork'").fetchall()
        self.assertEqual(len(rows), 1)
        self.assertEqual(str(rows[0]["cc_code"]), str(cc_code_valid))

        conn.close()

    def test_nnn_parser_skips_blank_and_zero_amounts(self):
        conn = _mk_conn()
        cc_code = "1412000018"
        _seed_cc(conn, cc_code)
        periods = get_fy_months(2027)

        tmpdir = self._create_temp_dir()
        workbook_path = tmpdir / "nnn_blank_zero.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FY2027"

        ws.append(["Mã Cost Center", "Mã tài khoản", "Tên nhân viên", "Mã nhân viên", "2026/04", "2026/05", "2026/06"])
        # Month 1: 0.0, Month 2: None, Month 3: 1500
        ws.append([cc_code, 6005186429, "Worker C", "VN0003", 0.0, None, 1500.0])

        wb.save(workbook_path)
        wb.close()

        result = parse_nnn_paperwork(conn, workbook_path=str(workbook_path))
        self.assertEqual(result["errors"], 0)
        self.assertEqual(result["inserted"], 1)  # Chỉ periods[2] (tương ứng tháng 6)

        rows = conn.execute("SELECT * FROM fact_input_data WHERE source = 'nnn_paperwork'").fetchall()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["period"], periods[2])
        self.assertEqual(float(rows[0]["amount_vnd"]), 1500.0)

        conn.close()

    def test_nnn_parser_raises_value_error_if_no_valid_header_row(self):
        conn = _mk_conn()
        tmpdir = self._create_temp_dir()
        workbook_path = tmpdir / "nnn_no_header.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FY2027"

        # Chỉ có dữ liệu rác
        for _ in range(12):
            ws.append(["Rác", "Không có header đúng", 123, 456])

        wb.save(workbook_path)
        wb.close()

        with self.assertRaises(ValueError):
            parse_nnn_paperwork(conn, workbook_path=str(workbook_path))

        conn.close()


if __name__ == "__main__":
    unittest.main()
