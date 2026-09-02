from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from src.engine.manual_special_cost_sections import (
    LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET,
    MANUAL_SPECIAL_COST_METADATA_SHEET,
    ManualSpecialCostSectionError,
    preserve_manual_special_cost_section,
)
from scripts.run_e2e import _restore_manual_special_cost_section, _stage_manual_special_cost_source


def _make_output(path: Path, *, common_end_row: int) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Chi tiết MP"
    for row in range(37, common_end_row + 1):
        sheet.cell(row, 2).value = 5000000000 + row
        sheet.cell(row, 19).value = f"common-{row}"
        sheet.cell(row, 6).value = row
    workbook.save(path)
    workbook.close()


def _add_legacy_manual_rows(path: Path, start_row: int) -> None:
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Chi tiết MP"]
    sheet.cell(start_row, 2).value = 5005246286
    sheet.cell(start_row, 19).value = "Chi phí riêng A"
    sheet.cell(start_row, 6).value = 1250000
    sheet.cell(start_row, 18).value = f"=SUM(F{start_row}:Q{start_row})"
    sheet.cell(start_row, 19).fill = PatternFill("solid", fgColor="00FF00")
    sheet.cell(start_row + 1, 2).value = 5005246288
    sheet.cell(start_row + 1, 19).value = "Chi phí riêng B"
    sheet.cell(start_row + 1, 6).value = 2500000
    workbook.save(path)
    workbook.close()


def _metadata_row(path: Path) -> tuple[object, ...]:
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        sheet = workbook[MANUAL_SPECIAL_COST_METADATA_SHEET]
        return tuple(sheet.cell(2, column).value for column in range(1, 8))
    finally:
        workbook.close()


def test_legacy_manual_rows_move_below_new_dynamic_common_block_without_mutating_source(tmp_path):
    source = tmp_path / "MP_CC_1412000030_FY2026.xlsx"
    generated = tmp_path / "MP_CC_1412000030_FY2027.xlsx"
    _make_output(source, common_end_row=86)
    _add_legacy_manual_rows(source, 87)
    before = source.read_bytes()
    _make_output(generated, common_end_row=90)

    result = preserve_manual_special_cost_section(
        generated,
        cc_code="1412000030",
        source_workbook_path=source,
        legacy_start_row=87,
    )

    assert result["manual_rows_preserved"] == 2
    assert result["manual_start_row"] == 92
    assert source.read_bytes() == before
    workbook = openpyxl.load_workbook(generated, data_only=False)
    try:
        sheet = workbook.active
        assert sheet["S91"].value == "CHI PHÍ RIÊNG - NHẬP THỦ CÔNG"
        assert sheet["B92"].value == 5005246286
        assert sheet["S92"].value == "Chi phí riêng A"
        assert sheet["F92"].value == 1250000
        assert sheet["R92"].value == "=SUM(F92:Q92)"
        assert sheet["S92"].fill.fgColor.rgb.endswith("00FF00")
        assert sheet["B93"].value == 5005246288
        assert MANUAL_SPECIAL_COST_METADATA_SHEET in workbook.sheetnames
        assert workbook[MANUAL_SPECIAL_COST_METADATA_SHEET].sheet_state == "veryHidden"
    finally:
        workbook.close()
    assert _metadata_row(generated)[2:] == ("1412000030", 90, 92, 93, 1)


def test_metadata_snapshot_survives_rerun_when_common_block_grows(tmp_path):
    first = tmp_path / "MP_CC_1412000030_first.xlsx"
    rerun = tmp_path / "MP_CC_1412000030_rerun.xlsx"
    _make_output(first, common_end_row=86)
    _add_legacy_manual_rows(first, 87)
    preserve_manual_special_cost_section(first, "1412000030", source_workbook_path=first, legacy_start_row=87)
    _make_output(rerun, common_end_row=94)

    result = preserve_manual_special_cost_section(rerun, "1412000030", source_workbook_path=first)

    assert result["manual_start_row"] == 96
    workbook = openpyxl.load_workbook(rerun, data_only=False)
    try:
        sheet = workbook["Chi tiết MP"]
        assert sheet["B96"].value == 5005246286
        assert sheet["S96"].value == "Chi phí riêng A"
        assert sheet["B97"].value == 5005246288
    finally:
        workbook.close()


def test_manual_special_metadata_sheet_name_is_valid_for_excel(tmp_path):
    output = tmp_path / "output.xlsx"
    _make_output(output, common_end_row=86)

    preserve_manual_special_cost_section(output, "1412000030")

    workbook = openpyxl.load_workbook(output, data_only=False)
    try:
        assert len(MANUAL_SPECIAL_COST_METADATA_SHEET) <= 31
        assert MANUAL_SPECIAL_COST_METADATA_SHEET in workbook.sheetnames
    finally:
        workbook.close()


def test_legacy_metadata_name_is_read_and_migrated_to_excel_safe_name(tmp_path):
    source = tmp_path / "legacy.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_output(source, common_end_row=86)
    _add_legacy_manual_rows(source, 87)
    workbook = openpyxl.load_workbook(source)
    try:
        sheet_name = workbook.active.title
        metadata = workbook.create_sheet(LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET)
        metadata.sheet_state = "veryHidden"
        metadata.append(("sheet_name", "fiscal_year", "cc_code", "common_end_row", "manual_start_row", "manual_end_row", "schema_version"))
        metadata.append((sheet_name, 2026, "1412000030", 86, 87, 88, 1))
        workbook.save(source)
    finally:
        workbook.close()
    _make_output(generated, common_end_row=90)

    preserve_manual_special_cost_section(generated, "1412000030", source_workbook_path=source)

    workbook = openpyxl.load_workbook(generated, data_only=False)
    try:
        assert MANUAL_SPECIAL_COST_METADATA_SHEET in workbook.sheetnames
        assert LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET not in workbook.sheetnames
        assert workbook.active["B92"].value == 5005246286
    finally:
        workbook.close()


def test_unmarked_source_requires_explicit_start_row(tmp_path):
    source = tmp_path / "old.xlsx"
    generated = tmp_path / "new.xlsx"
    _make_output(source, common_end_row=86)
    _add_legacy_manual_rows(source, 87)
    _make_output(generated, common_end_row=90)

    with pytest.raises(ManualSpecialCostSectionError, match="chưa có dấu mốc"):
        preserve_manual_special_cost_section(generated, "1412000030", source_workbook_path=source)


def test_pipeline_treats_unmarked_current_output_as_first_run(tmp_path):
    output_dir = tmp_path / "OUTPUT_FY2027"
    workspace = tmp_path / "workspace"
    output_dir.mkdir()
    workspace.mkdir()
    current = output_dir / "MP_CC_1412000030.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_output(current, common_end_row=86)
    _make_output(generated, common_end_row=90)
    before = current.read_bytes()
    context = SimpleNamespace(
        output_dir=str(output_dir),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=None,
        manual_special_legacy_starts={},
    )

    source, kind = _stage_manual_special_cost_source(context, "1412000030")
    result = _restore_manual_special_cost_section(
        str(generated),
        run_context=context,
        cc_code="1412000030",
        source_path=source,
        source_kind=kind,
        log_callback=lambda _message: None,
    )

    assert (source, kind) == (None, "new_fiscal_year")
    assert result["manual_rows_preserved"] == 0
    assert current.read_bytes() == before


def test_pipeline_stages_unmarked_current_output_when_legacy_start_is_configured(tmp_path):
    output_dir = tmp_path / "OUTPUT_FY2027"
    workspace = tmp_path / "workspace"
    output_dir.mkdir()
    workspace.mkdir()
    current = output_dir / "MP_CC_1412000030.xlsx"
    _make_output(current, common_end_row=86)
    _add_legacy_manual_rows(current, 87)
    context = SimpleNamespace(
        output_dir=str(output_dir),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=None,
        manual_special_legacy_starts={"1412000030": 87},
    )

    source, kind = _stage_manual_special_cost_source(context, "1412000030")

    assert kind == "current_fiscal_year"
    assert source is not None


def test_pipeline_stages_current_output_with_saved_special_cost_metadata(tmp_path):
    output_dir = tmp_path / "OUTPUT_FY2027"
    workspace = tmp_path / "workspace"
    output_dir.mkdir()
    workspace.mkdir()
    current = output_dir / "MP_CC_1412000030.xlsx"
    _make_output(current, common_end_row=86)
    preserve_manual_special_cost_section(current, "1412000030")
    context = SimpleNamespace(
        output_dir=str(output_dir),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=None,
        manual_special_legacy_starts={},
    )

    source, kind = _stage_manual_special_cost_source(context, "1412000030")

    assert kind == "current_fiscal_year"
    assert source is not None


def test_pipeline_stages_saved_layout_from_previous_fiscal_year(tmp_path):
    previous_dir = tmp_path / "OUTPUT_FY2027"
    workspace = tmp_path / "workspace"
    previous_dir.mkdir()
    workspace.mkdir()
    inherited = previous_dir / "MP_CC_1412000030.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_output(inherited, common_end_row=86)
    _add_legacy_manual_rows(inherited, 87)
    preserve_manual_special_cost_section(
        inherited,
        "1412000030",
        source_workbook_path=inherited,
        legacy_start_row=87,
    )
    _make_output(generated, common_end_row=90)
    context = SimpleNamespace(
        output_dir=str(tmp_path / "OUTPUT_FY2028"),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=str(previous_dir),
        manual_special_legacy_starts={},
    )

    source, kind = _stage_manual_special_cost_source(context, "1412000030")
    result = _restore_manual_special_cost_section(
        str(generated),
        run_context=context,
        cc_code="1412000030",
        source_path=source,
        source_kind=kind,
        log_callback=lambda _message: None,
    )

    assert (source, kind) == (str(inherited), "previous_fiscal_year")
    assert result["manual_rows_preserved"] == 2
    workbook = openpyxl.load_workbook(generated, data_only=False)
    try:
        assert workbook["Chi tiết MP"]["B92"].value == 5005246286
        assert workbook["Chi tiết MP"]["S92"].value == "Chi phí riêng A"
        assert workbook["Chi tiết MP"]["F92"].value is None
        assert workbook["Chi tiết MP"]["R92"].value is None
    finally:
        workbook.close()


def test_pipeline_rejects_malformed_special_cost_metadata_instead_of_treating_it_as_first_run(tmp_path):
    output_dir = tmp_path / "OUTPUT_FY2027"
    workspace = tmp_path / "workspace"
    output_dir.mkdir()
    workspace.mkdir()
    current = output_dir / "MP_CC_1412000030.xlsx"
    _make_output(current, common_end_row=86)
    workbook = openpyxl.load_workbook(current)
    metadata = workbook.create_sheet(MANUAL_SPECIAL_COST_METADATA_SHEET)
    metadata.sheet_state = "veryHidden"
    metadata.append(("sheet_name", "fiscal_year", "cc_code", "common_end_row", "manual_start_row", "manual_end_row", "schema_version"))
    metadata.append(("Chi tiết MP", 2027, "1412000030", 86, 86, 86, 1))
    workbook.save(current)
    workbook.close()
    context = SimpleNamespace(
        output_dir=str(output_dir),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=None,
        manual_special_legacy_starts={},
    )

    with pytest.raises(ManualSpecialCostSectionError, match="không hợp lệ"):
        _stage_manual_special_cost_source(context, "1412000030")


def test_pipeline_prefers_current_fy_snapshot_over_prior_fy_inheritance(tmp_path):
    current_dir = tmp_path / "OUTPUT_FY2027"
    inherited_dir = tmp_path / "OUTPUT_FY2026"
    workspace = tmp_path / "workspace"
    current_dir.mkdir()
    inherited_dir.mkdir()
    workspace.mkdir()
    current = current_dir / "MP_CC_1412000030.xlsx"
    inherited = inherited_dir / "MP_CC_1412000030.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_output(current, common_end_row=86)
    _add_legacy_manual_rows(current, 87)
    _make_output(inherited, common_end_row=86)
    _add_legacy_manual_rows(inherited, 87)
    old = openpyxl.load_workbook(inherited)
    old["Chi tiết MP"]["S87"] = "Không được dùng"
    old.save(inherited)
    old.close()
    _make_output(generated, common_end_row=90)
    context = SimpleNamespace(
        output_dir=str(current_dir),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=str(inherited_dir),
        manual_special_legacy_starts={"1412000030": 87},
    )

    source, kind = _stage_manual_special_cost_source(context, "1412000030")
    result = _restore_manual_special_cost_section(
        str(generated),
        run_context=context,
        cc_code="1412000030",
        source_path=source,
        source_kind=kind,
        log_callback=lambda _message: None,
    )

    assert kind == "current_fiscal_year"
    assert result["manual_rows_preserved"] == 2
    workbook = openpyxl.load_workbook(generated, data_only=False)
    try:
        assert workbook["Chi tiết MP"]["S92"].value == "Chi phí riêng A"
    finally:
        workbook.close()


def test_previous_fy_legacy_inheritance_keeps_code_description_but_clears_money(tmp_path):
    previous_dir = tmp_path / "OUTPUT_FY2026"
    workspace = tmp_path / "workspace"
    previous_dir.mkdir()
    workspace.mkdir()
    source = previous_dir / "MP_CC_1412000030.xlsx"
    generated = tmp_path / "generated.xlsx"
    _make_output(source, common_end_row=86)
    _add_legacy_manual_rows(source, 87)
    _make_output(generated, common_end_row=90)
    messages = []
    context = SimpleNamespace(
        output_dir=str(tmp_path / "OUTPUT_FY2027"),
        workspace_dir=str(workspace),
        manual_special_inheritance_dir=str(previous_dir),
        manual_special_legacy_starts={"1412000030": 87},
    )

    _restore_manual_special_cost_section(
        str(generated),
        run_context=context,
        cc_code="1412000030",
        source_path=str(source),
        source_kind="previous_fiscal_year",
        log_callback=messages.append,
    )

    workbook = openpyxl.load_workbook(generated, data_only=False)
    source_workbook = openpyxl.load_workbook(source, data_only=False)
    try:
        sheet = workbook.active
        assert sheet["B92"].value == 5005246286
        assert sheet["S92"].value == source_workbook.active["S87"].value
        assert sheet["F92"].value is None
        assert sheet["R92"].value is None
    finally:
        workbook.close()
        source_workbook.close()
    assert len(messages) == 1
    assert "1412000030" in messages[0]
