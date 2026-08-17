import sqlite3
from pathlib import Path

import openpyxl
import pytest

from src.db.schema import create_schema, init_sys_params
from src.parsers.fixed_assets import (
    _asset_tag,
    _round_vnd,
    expand_depreciation_schedule,
    expand_interest_schedule,
    find_fixed_assets_file,
    inspect_fixed_assets_workbook,
    parse_fixed_assets,
)
from src.audit.fixed_assets_coverage import build_fixed_assets_coverage_report
from src.utils.fiscal_periods import fiscal_periods


def _mk_conn():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    create_schema(conn)
    init_sys_params(conn, exchange_rate=100.0, fiscal_year=2027)
    return conn


def _asset_row(category, asset_no, text, control_cc, depreciation_cc, monthly, last_month, last_amount, apr, may):
    row = [None] * 23
    row[1], row[2], row[3] = category, asset_no, text
    row[7], row[9] = control_cc, depreciation_cc
    row[11], row[15], row[16] = monthly, last_month, last_amount
    row[21], row[22] = apr, may
    return row


def _write_fixed_assets_fixture(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2025.11"
    for _ in range(4):
        ws.append([None] * 23)
    ws.append(_asset_row("MFG)MACHINERY AND EQUIPMENT", "A-001", "Machine A", "9999999999", "1412000040", 10, "2026-05-31", 4, 1, 2))
    ws.append(_asset_row("MFG)VEHICLES", "A-002", "Vehicle B", "9999999999", "1412000040", 20, "2027-11-30", 5, 3, 4))
    ws.append(_asset_row("MFG)TOOLS FURNITURE AND FIXTURES", "B-001", "Tool C", "1412000040", "1412000018", 7, None, None, 0, 0))
    ws.append(_asset_row("MFG)SOFTWARE", "OUT-001", "Software", "1412000040", "1412000040", 9, None, None, 0, 0))
    ws.append(_asset_row("MFG)MOLD", "NOCC", "Skipped", "1412000040", None, 99, None, None, 0, 0))

    history = wb.create_sheet("2025.10")
    for _ in range(4):
        history.append([None] * 23)
    history.append(_asset_row("MFG)MACHINERY AND EQUIPMENT", "OLD-001", "Historical", "1412000040", "1412000040", 999, None, None, 0, 0))
    wb.save(path)


def test_parse_fixed_assets_uses_current_sheet_depreciation_cc_and_is_idempotent(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    conn = _mk_conn()

    result = parse_fixed_assets(conn, fa_path=str(workbook))

    assert result["source_rows"] == 3
    assert result["parsed_assets"] == 3
    assert result["selected_sheets"] == ["2025.11"]
    assert result["by_cc"]["1412000040"] == 2
    assert result["by_cc"]["1412000018"] == 1
    assert result["skipped_reasons"]["missing_depreciation_cc"] == 1
    assert result["skipped_reasons"]["out_of_scope_category"] == 1

    audit_rows = conn.execute(
        """
        SELECT inclusion_status, exclusion_reason, source_sheet, source_row,
               monthly_depr_usd, terminal_period, terminal_depr_usd,
               formula_cache_status
        FROM audit_fixed_asset_import_rows
        WHERE fiscal_year=2027
        ORDER BY source_row
        """
    ).fetchall()
    assert len(audit_rows) == 5
    assert sum(row[0] == "INCLUDED" for row in audit_rows) == 3
    assert {(row[0], row[1]) for row in audit_rows if row[0] == "EXCLUDED"} == {
        ("EXCLUDED", "out_of_scope_category"),
        ("EXCLUDED", "missing_depreciation_cc"),
    }
    assert audit_rows[0][2:8] == ("2025.11", 5, 10.0, "202605", 4.0, "NO_FORMULA")

    row_count = conn.execute("SELECT COUNT(*) FROM fact_input_data WHERE source='fixed_assets'").fetchone()[0]
    assert row_count > 0
    parse_fixed_assets(conn, fa_path=str(workbook))
    assert conn.execute("SELECT COUNT(*) FROM fact_input_data WHERE source='fixed_assets'").fetchone()[0] == row_count

    depr_may = conn.execute(
        """
        SELECT amount_usd, account_code FROM fact_input_data
        WHERE cc_code='1412000040' AND period='202605'
          AND description LIKE 'fixed_assets_depr|machinery_equipment|A-001%'
        """
    ).fetchone()
    assert float(depr_may[0]) == 4.0
    assert int(depr_may[1]) == 5006016242

    after_last = conn.execute(
        """
        SELECT COUNT(*) FROM fact_input_data
        WHERE cc_code='1412000040' AND period='202606'
          AND description LIKE 'fixed_assets_depr|machinery_equipment|A-001%'
        """
    ).fetchone()[0]
    assert after_last == 0
    conn.close()


def test_fixed_assets_coverage_report_is_non_sensitive_and_flags_missing_parse(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    conn = _mk_conn()

    source_only = inspect_fixed_assets_workbook(workbook)
    assert source_only["source_rows"] == 3
    assert "1412000040" in source_only["by_cc"]

    report_before = build_fixed_assets_coverage_report(conn, workbook)
    assert "1412000040" in report_before["mismatches"]

    parse_fixed_assets(conn, fa_path=str(workbook))
    report_after = build_fixed_assets_coverage_report(conn, workbook)
    assert report_after["mismatches"] == {}
    assert report_after["db"]["period_rows_by_cc"]["1412000040"] > 0
    conn.close()


def test_fixed_assets_raw_workbook_coverage_if_present():
    workbook = Path("docs/MP2027/固定資産情報_Fixed_Assets_Information_2025.11 - Nov.xlsx")
    if not workbook.exists():
        workbook = Path("raw/固定資産情報_Fixed_Assets_Information_2025.11 - Nov.xlsx")
    if not workbook.exists():
        return
    conn = _mk_conn()
    result = parse_fixed_assets(conn, fa_path=str(workbook))
    report = build_fixed_assets_coverage_report(conn, workbook)
    assert result["source_rows"] >= result["parsed_assets"]
    assert isinstance(report["source"].get("by_cc", {}), dict)
    conn.close()


def test_parse_fixed_assets_fails_closed_without_authoritative_fy_or_fx(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    conn = _mk_conn()
    try:
        conn.execute("DELETE FROM sys_params WHERE key='exchange_rate_usd_vnd'")
        with pytest.raises(ValueError, match="exchange_rate_usd_vnd"):
            parse_fixed_assets(conn, fa_path=str(workbook))

        init_sys_params(conn, exchange_rate=100.0, fiscal_year=2027)
        conn.execute("DELETE FROM sys_params WHERE key='fiscal_year'")
        with pytest.raises(ValueError, match="fiscal_year"):
            parse_fixed_assets(conn, fa_path=str(workbook))
    finally:
        conn.close()


def test_parse_fixed_assets_does_not_substitute_l_when_terminal_q_is_missing(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    wb = openpyxl.load_workbook(workbook)
    # The first asset ends in FY2027, but Q is deliberately blank.
    wb["2025.11"].cell(row=5, column=17).value = None
    wb.save(workbook)
    conn = _mk_conn()
    try:
        with pytest.raises(ValueError, match="Thiếu số tiền khấu hao tháng cuối"):
            parse_fixed_assets(conn, fa_path=str(workbook))
    finally:
        conn.close()


def test_parse_fixed_assets_preserves_explicit_zero_q_and_negative_amounts(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    wb = openpyxl.load_workbook(workbook)
    sheet = wb["2025.11"]
    sheet.cell(row=5, column=17, value=0)
    sheet.cell(row=6, column=12, value=-20)
    wb.save(workbook)
    conn = _mk_conn()
    try:
        parse_fixed_assets(conn, fa_path=str(workbook))
        terminal = conn.execute(
            """
            SELECT amount_usd FROM fact_input_data
            WHERE description LIKE 'fixed_assets_depr|machinery_equipment|A-001%'
              AND period='202605'
            """
        ).fetchone()
        assert terminal is not None
        assert float(terminal[0]) == 0.0
        negative = conn.execute(
            """
            SELECT amount_usd FROM fact_input_data
            WHERE description LIKE 'fixed_assets_depr|vehicles|A-002%'
              AND period='202604'
            """
        ).fetchone()
        assert negative is not None
        assert float(negative[0]) == -20.0
    finally:
        conn.close()


def test_parse_fixed_assets_reimport_replaces_only_the_same_fiscal_year(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    conn = _mk_conn()
    try:
        init_sys_params(conn, exchange_rate=100.0, fiscal_year=2026)
        parse_fixed_assets(conn, fa_path=str(workbook))
        fy2026_rows = conn.execute(
            "SELECT COUNT(*) FROM fact_input_data WHERE source='fixed_assets' AND fiscal_year=2026"
        ).fetchone()[0]
        assert fy2026_rows > 0

        init_sys_params(conn, exchange_rate=100.0, fiscal_year=2027)
        parse_fixed_assets(conn, fa_path=str(workbook))
        fy2027_rows = conn.execute(
            "SELECT COUNT(*) FROM fact_input_data WHERE source='fixed_assets' AND fiscal_year=2027"
        ).fetchone()[0]
        assert fy2027_rows > 0
        assert conn.execute(
            "SELECT COUNT(*) FROM fact_input_data WHERE source='fixed_assets' AND fiscal_year=2026"
        ).fetchone()[0] == fy2026_rows
        assert conn.execute(
            "SELECT COUNT(*) FROM audit_fixed_asset_import_rows WHERE fiscal_year=2026"
        ).fetchone()[0] > 0
    finally:
        conn.close()


def test_parse_fixed_assets_fails_closed_for_formula_without_cached_value(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    wb = openpyxl.load_workbook(workbook)
    # openpyxl writes the formula but cannot create an Excel cached result.
    wb["2025.11"].cell(row=5, column=12, value="=10+1")
    wb.save(workbook)
    conn = _mk_conn()
    try:
        with pytest.raises(ValueError, match="Thiếu giá trị công thức đã tính sẵn"):
            parse_fixed_assets(conn, fa_path=str(workbook))
    finally:
        conn.close()


def test_parse_fixed_assets_supports_a_future_fiscal_year_from_sys_params(tmp_path):
    workbook = tmp_path / "Fixed_Assets_Information_fixture.xlsx"
    _write_fixed_assets_fixture(workbook)
    conn = _mk_conn()
    try:
        init_sys_params(conn, exchange_rate=100.0, fiscal_year=2028)
        result = parse_fixed_assets(conn, fa_path=str(workbook))
        periods = {
            row[0]
            for row in conn.execute(
                "SELECT DISTINCT period FROM fact_input_data WHERE source='fixed_assets' AND fiscal_year=2028"
            )
        }
        assert result["fiscal_year"] == 2028
        assert "202704" in periods
        assert "202803" in periods
    finally:
        conn.close()


def test_fixed_assets_file_and_sheet_are_selected_by_content_not_name_or_period(tmp_path):
    workbook = tmp_path / "opaque-input.xlsx"
    _write_fixed_assets_fixture(workbook)
    wb = openpyxl.load_workbook(workbook)
    wb["2025.11"].title = "asset-ledger"
    wb["2025.10"].title = "archive"
    wb.save(workbook)
    conn = _mk_conn()
    try:
        assert Path(find_fixed_assets_file(str(tmp_path))) == workbook
        result = parse_fixed_assets(conn, source_dir=str(tmp_path))
        assert result["selected_sheets"] == ["asset-ledger"]
        assert result["source_rows"] == 3
    finally:
        conn.close()


def test_asset_identity_keeps_source_coordinate_when_asset_labels_repeat():
    first = _asset_tag("", "Repeated asset", "asset-ledger", 8)
    second = _asset_tag("", "Repeated asset", "asset-ledger", 9)
    assert first != second
    assert first.endswith("@asset-ledger!8")
    assert second.endswith("@asset-ledger!9")


def test_asset_schedules_cover_terminal_before_within_after_and_no_terminal():
    months = ["202604", "202605", "202606", "202607"]
    assert expand_depreciation_schedule(10, "202603", 2, months) == {}
    assert expand_depreciation_schedule(10, "202606", 4, months) == {
        "202604": 10,
        "202605": 10,
        "202606": 4,
    }
    assert expand_depreciation_schedule(10, "202804", 2, months) == {
        period: 10 for period in months
    }
    assert expand_depreciation_schedule(10, None, None, months) == {
        period: 10 for period in months
    }


def test_interest_schedule_uses_v_in_april_w_from_may_and_keeps_terminal_month():
    months = ["202604", "202605", "202606", "202607"]
    assert expand_interest_schedule(1, 2, "202606", months) == {
        "202604": 1,
        "202605": 2,
        "202606": 2,
    }


def test_vnd_rounding_matches_excel_policy_for_positive_and_negative_half_values():
    assert _round_vnd(0.5, 1) == 1
    assert _round_vnd(-0.5, 1) == -1
