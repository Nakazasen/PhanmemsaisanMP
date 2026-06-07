from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

from src.engine.admin_consumables_writer import apply_admin_consumables_to_workbook
from src.utils import excel_helpers as helpers

TEMPLATE = Path("FORM.xlsx")
ADMIN_SOURCE = Path("raw/総務課 FY2027 MP 振替予定.xlsx")
ALLOC_SOURCE = Path("raw/FY2027配賦額一覧 (2025.12.29).xlsx")


def test_admin_consumables_writer_writes_rows_207_to_209(tmp_path):
    workbook_path = tmp_path / "output.xlsx"
    copy2(TEMPLATE, workbook_path)
    apply_admin_consumables_to_workbook(workbook_path, ADMIN_SOURCE, ALLOC_SOURCE)
    workbook = load_workbook(workbook_path)
    try:
        sheet = workbook[helpers.find_hub_sheet_name(workbook)]
        assert [sheet.cell(row=row, column=5).value for row in range(207, 210)] == [
            "toilet_paper",
            "hand_soap",
            "alcohol_disinfectant",
        ]
        assert [sheet.cell(row=row, column=19).value for row in range(207, 210)] == [
            "トイレットペーパー",
            "手洗い洗剤",
            "アルコール消毒",
        ]
    finally:
        workbook.close()


def test_admin_consumables_writer_blank_row_210(tmp_path):
    workbook_path = tmp_path / "output.xlsx"
    copy2(TEMPLATE, workbook_path)
    apply_admin_consumables_to_workbook(workbook_path, ADMIN_SOURCE, ALLOC_SOURCE)
    workbook = load_workbook(workbook_path)
    try:
        sheet = workbook[helpers.find_hub_sheet_name(workbook)]
        assert sheet.cell(row=210, column=5).value is None
        assert sheet.cell(row=210, column=6).value is None
        assert sheet.cell(row=210, column=17).value is None
        assert sheet.cell(row=210, column=19).value is None
        assert sheet.cell(row=210, column=20).value is None
    finally:
        workbook.close()


def test_admin_consumables_writer_does_not_modify_source_or_template(tmp_path):
    workbook_path = tmp_path / "output.xlsx"
    copy2(TEMPLATE, workbook_path)
    before_template_mtime = TEMPLATE.stat().st_mtime_ns
    before_admin_mtime = ADMIN_SOURCE.stat().st_mtime_ns
    before_alloc_mtime = ALLOC_SOURCE.stat().st_mtime_ns
    apply_admin_consumables_to_workbook(workbook_path, ADMIN_SOURCE, ALLOC_SOURCE)
    assert TEMPLATE.stat().st_mtime_ns == before_template_mtime
    assert ADMIN_SOURCE.stat().st_mtime_ns == before_admin_mtime
    assert ALLOC_SOURCE.stat().st_mtime_ns == before_alloc_mtime
