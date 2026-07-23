from pathlib import Path

import openpyxl

from src.services.fiscal_run import (
    create_fiscal_run_context,
    preflight_fiscal_run,
)
from src.universal_app import (
    MPManagerApp,
    _friendly_error_message,
    _is_missing_baseline_error,
    _pipeline_failure_summary,
    _validate_selected_template,
)
from src.utils import excel_helpers


ROOT = Path(__file__).resolve().parents[1]


def test_form_without_system_cost_account_is_valid():
    path = ROOT / "docs" / "MP2027" / "FORM.xlsx"
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook["内訳ﾘｽﾄ(4～3月)"]
        assert worksheet["B75"].value is None
    finally:
        workbook.close()

    assert _validate_selected_template(str(path)) is None


def test_form_validation_does_not_require_system_cost_text_or_account(tmp_path):
    source = ROOT / "docs" / "MP2027" / "FORM.xlsx"
    path = tmp_path / "FORM_neutral.xlsx"
    workbook = openpyxl.load_workbook(source)
    worksheet = workbook["内訳ﾘｽﾄ(4～3月)"]
    for row in worksheet.iter_rows(min_col=2, max_col=20):
        for cell in row:
            text = str(cell.value or "").lower()
            if "system cost" in text or "kdc" in text or "システム" in text:
                cell.value = None
    workbook.save(path)
    workbook.close()

    assert _validate_selected_template(str(path)) is None


def test_invalid_template_b2_clears_old_ui_rate():
    path = ROOT / "docs" / "MP2027" / "FORM 1.xlsx"

    class Variable:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

        def set(self, value):
            self.value = value

    app = object.__new__(MPManagerApp)
    app.template_path = Variable(str(path))
    app.exchange_rate = Variable("26273")
    messages = []
    app.log = messages.append

    assert app._reload_exchange_rate_from_template() is False
    assert app.exchange_rate.get() == ""
    assert "B2" in messages[-1]


def test_missing_march_headcount_has_actionable_user_message():
    raw_error = (
        "Kiểm tra nguồn nhân sự & thời gian FY2027 không đạt.\n"
        "Chương trình dừng trước khi xuất FORM để tránh kết quả sai.\n"
        "- Phòng 1412000004 – 機器製造1課: chưa có Tổng số người tháng 03/2026. "
        "Dữ liệu này cần để tính chi phí tháng 04/2026. "
        "Hãy chọn “Nhập nhân sự thủ công”, nhập dữ liệu tháng này và lưu lại"
    )

    message = _friendly_error_message(ValueError(raw_error))

    assert message.startswith("Thiếu dữ liệu nhân sự để xuất báo cáo.")
    assert "Phòng 1412000004 – 機器製造1課" in message
    assert "Tổng số người tháng 03/2026" in message
    assert "tính chi phí tháng 04/2026" in message
    assert "Nhập nhân sự thủ công" in message
    assert "CHẠY TÍNH TOÁN" in message
    assert "kiểm tra lại Tệp mẫu FORM" not in message
    assert _is_missing_baseline_error(raw_error)


def test_non_baseline_staffing_error_does_not_open_baseline_recovery():
    assert not _is_missing_baseline_error(
        "Kiểm tra nguồn nhân sự & thời gian FY2027 không đạt.\n"
        "- Phòng 1412000004: thiếu thời gian làm việc các tháng: 202604"
    )


def test_pipeline_failure_summary_excludes_import_log_but_keeps_error_block():
    summary = _pipeline_failure_summary(
        [
            "ĐÃ NẠP: file 1",
            "ĐÃ NẠP: file 2",
            "LỖI: Kiểm tra nguồn nhân sự không đạt.",
            "- Phòng 1412000004: chưa có Tổng số người tháng 03/2026.",
            "Chi tiết kỹ thuật đã được ẩn. Nếu cần điều tra sâu...",
        ],
        1,
    )
    assert "ĐÃ NẠP" not in summary
    assert summary.startswith("Kiểm tra nguồn nhân sự không đạt.")
    assert "chưa có Tổng số người tháng 03/2026" in summary


def test_refresh_cost_centers_seeds_empty_database_from_selected_form(tmp_path, monkeypatch):
    from src import universal_app
    from src.db.schema import get_connection

    class Variable:
        def __init__(self, value=""):
            self.value = value

        def get(self):
            return self.value

        def set(self, value):
            self.value = value

    class Widget:
        def __init__(self):
            self.values = []
            self.states = []

        def __setitem__(self, key, value):
            assert key == "values"
            self.values = list(value)

        def configure(self, **kwargs):
            if "state" in kwargs:
                self.states.append(kwargs["state"])

    monkeypatch.setattr(universal_app, "BASE_DIR", str(tmp_path))
    notices = []
    monkeypatch.setattr(
        universal_app.messagebox,
        "showinfo",
        lambda title, message: notices.append((title, message)),
    )
    monkeypatch.setattr(
        universal_app.messagebox,
        "showerror",
        lambda title, message: notices.append((title, message)),
    )
    app = object.__new__(MPManagerApp)
    app.project = type("Project", (), {"operational_database": str(tmp_path / "mp2027.db")})()
    app.template_path = Variable(str(ROOT / "docs" / "MP2027" / "FORM.xlsx"))
    app.cc_code_filter = Variable()
    app._available_cc_choices = []
    app._selected_cc_values = []
    app.refresh_btn = Widget()
    app.messages = []
    app.log = app.messages.append

    app.refresh_cost_centers_from_form()

    conn = get_connection(str(tmp_path / "mp2027.db"))
    try:
        assert conn.execute("SELECT COUNT(*) FROM dim_cost_centers").fetchone()[0] == 65
        assert conn.execute(
            "SELECT name_jp FROM dim_cost_centers WHERE code='1412000004'"
        ).fetchone()[0] == "機器製造1課"
    finally:
        conn.close()
    assert len(app._available_cc_choices) == 65
    assert "1412000004 - 機器製造1課" in app._available_cc_choices
    assert app.refresh_btn.states == [universal_app.tk.DISABLED, universal_app.tk.NORMAL]
    assert notices and notices[0][0] == "Nạp Trung tâm chi phí thành công"
    assert any("Đã nạp 65" in message for message in app.messages)


def test_refresh_cost_centers_uses_existing_database_without_reloading_form(tmp_path, monkeypatch):
    import sqlite3

    from src import universal_app
    from src.db.schema import create_schema

    class Variable:
        def __init__(self, value=""):
            self.value = value

        def get(self):
            return self.value

        def set(self, value):
            self.value = value

    class Widget:
        def __init__(self):
            self.values = []
            self.states = []

        def __setitem__(self, key, value):
            assert key == "values"
            self.values = list(value)

        def configure(self, **kwargs):
            if "state" in kwargs:
                self.states.append(kwargs["state"])

    monkeypatch.setattr(universal_app, "BASE_DIR", str(tmp_path))
    conn = sqlite3.connect(tmp_path / "mp2027.db")
    conn.row_factory = sqlite3.Row
    create_schema(conn)
    with conn:
        conn.execute(
            "INSERT INTO dim_cost_centers(code,name_jp,saisan_type,cost_type) VALUES(?,?,?,?)",
            ("1412000004", "機器製造1課", "製造", "製造"),
        )
    conn.close()

    def fail_if_reloaded(*_args, **_kwargs):
        raise AssertionError("Existing master must not reload FORM")

    monkeypatch.setattr(universal_app, "load_cost_centers", fail_if_reloaded)
    monkeypatch.setattr(universal_app.messagebox, "showinfo", fail_if_reloaded)
    app = object.__new__(MPManagerApp)
    app.project = type("Project", (), {"operational_database": str(tmp_path / "mp2027.db")})()
    app.template_path = Variable(str(ROOT / "docs" / "MP2027" / "FORM.xlsx"))
    app.cc_code_filter = Variable("1412000004 - 機器製造1課")
    app._available_cc_choices = ["1412000004 - 機器製造1課"]
    app._selected_cc_values = ["1412000004 - 機器製造1課"]
    app.refresh_btn = Widget()
    app.messages = []
    app.log = app.messages.append

    app.refresh_cost_centers_from_form()

    assert app._available_cc_choices == ["1412000004 - 機器製造1課"]
    assert app._selected_cc_values == ["1412000004 - 機器製造1課"]
    assert app.cc_code_filter.get() == "Tất cả Trung tâm chi phí (1)"
    assert app.refresh_btn.states == [universal_app.tk.DISABLED, universal_app.tk.NORMAL]
    assert any("Đã làm mới danh sách 1" in message for message in app.messages)


def _write_hygiene_test_form(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "内訳ﾘｽﾄ(4～3月)"
    worksheet["B2"] = 26273
    worksheet["S4"] = "structural label"
    worksheet["F8"] = "=SUM(1, 1)"
    worksheet["S30"] = "=IFERROR(VLOOKUP(B30,A:H,2,0),\"\")"
    worksheet["BC30"] = "preserve workbook dimensions"
    workbook.save(path)
    workbook.close()


def test_form_hygiene_ignores_structure_and_formulas_but_detects_payload(tmp_path):
    path = tmp_path / "FORM.xlsx"
    _write_hygiene_test_form(path)
    workbook = openpyxl.load_workbook(path, data_only=False)
    worksheet = workbook["内訳ﾘｽﾄ(4～3月)"]
    try:
        assert excel_helpers.find_form_template_hygiene_issues(workbook) == ()
        worksheet["B5"] = 1412000006
        worksheet["G9"] = 42
        worksheet["T38"] = "WBS-OLD"
        assert excel_helpers.find_form_template_hygiene_issues(workbook) == (
            "B5",
            "G9",
            "T38",
        )
    finally:
        workbook.close()


def test_fiscal_preflight_blocks_form_with_old_department_payload(tmp_path):
    template_path = tmp_path / "docs" / "MP2027" / "FORM.xlsx"
    template_path.parent.mkdir(parents=True)
    _write_hygiene_test_form(template_path)
    workbook = openpyxl.load_workbook(template_path)
    workbook["内訳ﾘｽﾄ(4～3月)"]["B5"] = 1412000006
    workbook.save(template_path)
    workbook.close()

    context = create_fiscal_run_context(2027, template_path=template_path, base_dir=tmp_path)
    report = preflight_fiscal_run(context)
    template_issues = [issue for issue in report.issues if issue.category == "template"]

    assert len(template_issues) == 1
    assert template_issues[0].code == "FORM_TEMPLATE_NOT_CLEAN"
    assert "B5" in template_issues[0].reason
    assert "FORM sạch" in template_issues[0].action
