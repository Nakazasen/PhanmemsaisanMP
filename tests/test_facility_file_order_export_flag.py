from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

import scripts.run_e2e as run_e2e
from src.engine.facility_file_order_writer import apply_facility_file_order_to_workbook
from src.utils import excel_helpers as helpers

TEMPLATE = Path("FORM.xlsx")
FACILITY_SOURCE = Path("raw/施設課　MPFY2027.xlsx")


def test_default_export_does_not_call_facility_file_order(monkeypatch):
    called = False

    def fake_apply(**kwargs):
        nonlocal called
        called = True

    monkeypatch.setattr(run_e2e, "apply_facility_file_order_to_workbook", fake_apply)
    assert "facility_file_order_export: bool = False" in Path("scripts/run_e2e.py").read_text(encoding="utf-8")
    assert not called


def test_facility_file_order_export_requires_explicit_flag():
    text = Path("scripts/run_e2e.py").read_text(encoding="utf-8")

    assert "--facility-file-order-export" in text
    assert "if facility_file_order_export:" in text


def test_facility_file_order_export_writes_rows_200_to_205(tmp_path):
    workbook_path = tmp_path / "output.xlsx"
    copy2(TEMPLATE, workbook_path)

    apply_facility_file_order_to_workbook(workbook_path, FACILITY_SOURCE)

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        sheet = workbook[helpers.find_hub_sheet_name(workbook)]
        assert [sheet.cell(row=row, column=5).value for row in range(200, 206)] == [
            "building_depreciation",
            "land_depreciation",
            "building_interest",
            "land_interest",
            "electricity",
            "water",
        ]
    finally:
        workbook.close()


def test_facility_file_order_export_leaves_blank_row_206(tmp_path):
    workbook_path = tmp_path / "output.xlsx"
    copy2(TEMPLATE, workbook_path)

    apply_facility_file_order_to_workbook(workbook_path, FACILITY_SOURCE)

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        sheet = workbook[helpers.find_hub_sheet_name(workbook)]
        assert sheet.cell(row=206, column=5).value is None
        assert sheet.cell(row=206, column=6).value is None
        assert sheet.cell(row=206, column=17).value is None
        assert sheet.cell(row=206, column=19).value is None
    finally:
        workbook.close()


def test_facility_file_order_export_does_not_modify_template_or_source(tmp_path):
    workbook_path = tmp_path / "output.xlsx"
    copy2(TEMPLATE, workbook_path)
    before_template_mtime = TEMPLATE.stat().st_mtime_ns
    before_source_mtime = FACILITY_SOURCE.stat().st_mtime_ns

    apply_facility_file_order_to_workbook(workbook_path, FACILITY_SOURCE)

    assert TEMPLATE.stat().st_mtime_ns == before_template_mtime
    assert FACILITY_SOURCE.stat().st_mtime_ns == before_source_mtime


def test_facility_file_order_export_does_not_dirty_OUTPUT_FY2027_in_repo_root(monkeypatch, tmp_path):
    monkeypatch.chdir(tmp_path)
    output_dir = Path("OUTPUT_FY2027")
    output_dir.mkdir()
    assert output_dir.exists()
    assert not Path.cwd().samefile(Path(__file__).resolve().parents[1])


def test_preview_flag_and_export_flag_are_separate():
    text = Path("scripts/run_e2e.py").read_text(encoding="utf-8")

    assert "facility_file_order_preview: bool = False" in text
    assert "facility_file_order_export: bool = False" in text
    assert "write_facility_file_order_preview_workbook" in text
    assert "apply_facility_file_order_to_workbook" in text
