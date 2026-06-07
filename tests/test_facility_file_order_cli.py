import subprocess
import sys

from openpyxl import load_workbook

from src.utils import excel_helpers as helpers

TEMPLATE = "FORM.xlsx"
FACILITY_SOURCE = "raw/施設課　MPFY2027.xlsx"
CLI = "tools/create_facility_file_order_preview.py"


def _run_cli(*args):
    return subprocess.run(
        [sys.executable, CLI, *args],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )


def test_cli_creates_preview_workbook_in_tmp_path(tmp_path):
    output = tmp_path / "facility_file_order_preview.xlsx"

    result = _run_cli(
        "--template", TEMPLATE,
        "--facility-source", FACILITY_SOURCE,
        "--output", str(output),
        "--cost-center", "1412000040",
        "--start-row", "200",
    )

    assert result.returncode == 0, result.stderr
    assert str(output) in result.stdout
    assert output.exists()


def test_cli_requires_explicit_output():
    result = _run_cli("--template", TEMPLATE, "--facility-source", FACILITY_SOURCE)

    assert result.returncode != 0
    assert "--output" in result.stderr


def test_cli_does_not_modify_template_or_source(tmp_path):
    output = tmp_path / "facility_file_order_preview.xlsx"
    before_template_mtime = __import__("pathlib").Path(TEMPLATE).stat().st_mtime_ns
    before_source_mtime = __import__("pathlib").Path(FACILITY_SOURCE).stat().st_mtime_ns

    result = _run_cli(
        "--template", TEMPLATE,
        "--facility-source", FACILITY_SOURCE,
        "--output", str(output),
    )

    assert result.returncode == 0, result.stderr
    assert __import__("pathlib").Path(TEMPLATE).stat().st_mtime_ns == before_template_mtime
    assert __import__("pathlib").Path(FACILITY_SOURCE).stat().st_mtime_ns == before_source_mtime


def test_cli_output_has_rows_200_to_205_and_blank_206(tmp_path):
    output = tmp_path / "facility_file_order_preview.xlsx"
    result = _run_cli("--template", TEMPLATE, "--facility-source", FACILITY_SOURCE, "--output", str(output))
    assert result.returncode == 0, result.stderr

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[helpers.find_hub_sheet_name(workbook)]
        assert [sheet.cell(row=row, column=5).value for row in range(200, 206)] == [
            "building_depreciation",
            "land_depreciation",
            "building_interest",
            "land_interest",
            "electricity",
            "water",
        ]
        assert sheet.cell(row=206, column=5).value is None
        assert sheet.cell(row=206, column=19).value is None
    finally:
        workbook.close()


def test_cli_help_or_argparse_available():
    result = _run_cli("--help")

    assert result.returncode == 0
    assert "--template" in result.stdout
    assert "--facility-source" in result.stdout
    assert "--output" in result.stdout
