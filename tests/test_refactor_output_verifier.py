from pathlib import Path

import openpyxl

from tools.verify_refactor_output import compare_output_directories, write_evidence

SHEET = "内訳ﾘｽﾄ(4～3月)"


def _write_output(folder: Path, cc: str, *, rows=(38, 39), locked=None) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET
    worksheet["B31"] = "=SUM(F31:Q31)"
    worksheet["R34"] = "=SUM(F34:Q34)"
    worksheet["S36"] = "Locked description"
    for row in rows:
        worksheet.cell(row=row, column=2).value = 5000000000 + row
        worksheet.cell(row=row, column=6).value = f"=SUM({row},1)"
    for coordinate, value in (locked or {}).items():
        worksheet[coordinate] = value
    path = folder / f"MP_CC_{cc}.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def test_identical_output_folders_pass_and_write_evidence(tmp_path):
    baseline, candidate, reports = tmp_path / "baseline", tmp_path / "candidate", tmp_path / "reports"
    _write_output(baseline, "1412000001")
    _write_output(candidate, "1412000001")

    evidence = compare_output_directories(baseline, candidate, baseline_ref="ca2bc52")
    json_path, xlsx_path = write_evidence(reports, evidence)

    assert evidence.passed
    result = evidence.results[0]
    assert result.baseline_data_rows == 2
    assert result.candidate_data_rows == 2
    assert result.row_count_passed
    assert result.locked_range_passed
    assert json_path.is_file()
    assert xlsx_path.is_file()
    report = openpyxl.load_workbook(xlsx_path, data_only=False)
    try:
        assert report["Tóm_tắt"]["B1"].value == "PASS"
    finally:
        report.close()


def test_changed_generated_row_count_fails_for_affected_cost_center(tmp_path):
    baseline, candidate = tmp_path / "baseline", tmp_path / "candidate"
    _write_output(baseline, "1412000001", rows=(38, 39))
    _write_output(candidate, "1412000001", rows=(38, 39, 40))

    evidence = compare_output_directories(baseline, candidate)

    result = evidence.results[0]
    assert not evidence.passed
    assert result.cost_center == "1412000001"
    assert (result.baseline_data_rows, result.candidate_data_rows) == (2, 3)
    assert not result.row_count_passed
    assert result.locked_range_passed


def test_locked_range_reports_exact_formula_change(tmp_path):
    baseline, candidate = tmp_path / "baseline", tmp_path / "candidate"
    _write_output(baseline, "1412000002")
    _write_output(candidate, "1412000002", locked={"R34": "=SUM(F34:P34)"})

    evidence = compare_output_directories(baseline, candidate)

    result = evidence.results[0]
    assert not evidence.passed
    assert result.row_count_passed
    assert not result.locked_range_passed
    assert [(item.coordinate, item.baseline, item.candidate) for item in result.locked_differences] == [
        ("R34", "=SUM(F34:Q34)", "=SUM(F34:P34)")
    ]


def test_missing_cost_center_output_fails(tmp_path):
    baseline, candidate = tmp_path / "baseline", tmp_path / "candidate"
    _write_output(baseline, "1412000001")
    _write_output(candidate, "1412000002")

    evidence = compare_output_directories(baseline, candidate)

    assert not evidence.passed
    assert [result.cost_center for result in evidence.results] == ["1412000001", "1412000002"]
    assert evidence.results[0].errors == ["Thiếu workbook ở bản refactor."]
    assert evidence.results[1].errors == ["Thiếu workbook ở baseline."]
