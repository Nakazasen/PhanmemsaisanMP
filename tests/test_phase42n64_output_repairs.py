from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.engine.complete_v1_source_order_writer import apply_complete_v1_source_order_to_workbook
from src.parsers.fixed_assets import HEADER_ALIASES, LEGACY_COLUMN_MAP

SHEET = "MP_DETAIL"


def _mk(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for r, desc in [(36, "facility building depreciation"), (37, "facility land depreciation"), (40, "facility building interest"), (41, "facility land interest"), (44, "facility electric"), (45, "facility water"), (57, "health check"), (97, "new employee notebook")]:
        ws.cell(r, 2).value = 5000000000 + r
        ws.cell(r, 3).value = f"=LOOKUP({r})"
        ws.cell(r, 4).value = f"=GROUP({r})"
        ws.cell(r, 6).value = f"={r}*100"
        ws.cell(r, 18).value = f"=SUM(F{r}:Q{r})"
        ws.cell(r, 19).value = desc
        ws.cell(r, 20).value = f"legacy_source_row={r}"
    ws.cell(52, 2).value = 5000000036
    ws.cell(52, 6).value = "=36*100"
    ws.cell(52, 19).value = "unrelated same amount"
    ws.row_dimensions[174].height = 33
    wb.save(path)
    wb.close()
    return path


def _business_values(ws, row):
    return [ws.cell(row, c).value for c in (2, 5, *range(6, 18), 19, 20)]


def test_canonical_emission_clears_exact_legacy_duplicates_and_retains_canonical(tmp_path):
    workbook = _mk(tmp_path / "out.xlsx")
    apply_complete_v1_source_order_to_workbook(workbook, start_row=168, clear_until_row=190)
    wb = load_workbook(workbook, data_only=False)
    try:
        ws = wb[SHEET]
        for row in [36, 37, 40, 41, 44, 45, 57]:
            assert all(ws.cell(row, c).value is None for c in range(2, 21))
        assert ws.cell(168, 19).value == "facility building depreciation"
        assert ws.cell(172, 19).value == "facility electric"
        assert ws.cell(173, 19).value == "facility water"
        assert "health check" in [ws.cell(r, 19).value for r in range(168, 191)]
    finally:
        wb.close()


def test_unrelated_same_amount_row_is_not_cleared(tmp_path):
    workbook = _mk(tmp_path / "out.xlsx")
    apply_complete_v1_source_order_to_workbook(workbook, start_row=168, clear_until_row=190)
    wb = load_workbook(workbook, data_only=False)
    try:
        ws = wb[SHEET]
        assert ws.cell(52, 19).value == "unrelated same amount"
        assert ws.cell(52, 6).value == "=36*100"
    finally:
        wb.close()


def test_strict_separator_clears_b_to_t_and_preserves_row_height(tmp_path):
    workbook = _mk(tmp_path / "out.xlsx")
    apply_complete_v1_source_order_to_workbook(workbook, start_row=168, clear_until_row=190)
    wb = load_workbook(workbook, data_only=False)
    try:
        ws = wb[SHEET]
        assert all(ws.cell(174, c).value is None for c in range(2, 21))
        assert ws.row_dimensions[174].height == 33
    finally:
        wb.close()


def test_row_after_last_business_template_total_only_not_business(tmp_path):
    workbook = _mk(tmp_path / "out.xlsx")
    apply_complete_v1_source_order_to_workbook(workbook, start_row=168, clear_until_row=190)
    wb = load_workbook(workbook, data_only=False)
    try:
        ws = wb[SHEET]
        business_rows = [r for r in range(168, 191) if any(_business_values(ws, r))]
        assert business_rows
        tail = max(business_rows) + 1
        ws.cell(tail, 18).value = f"=SUM(F{tail}:Q{tail})"
        assert not any(_business_values(ws, tail))
    finally:
        wb.close()


def test_fixed_assets_parser_keeps_legacy_depreciation_cc_and_header_detection():
    assert "depreciation_cc" in HEADER_ALIASES
    assert LEGACY_COLUMN_MAP["control_cc"] == 7
    assert LEGACY_COLUMN_MAP["depreciation_cc"] == 9


def test_no_agent_authored_description_suffixes_in_hub_rules():
    text = Path("src/engine/hub_builder.py").read_text(encoding="utf-8")
    forbidden = ["company trip/festival allocation", "sports day allocation", "mooncake allocation", "pen allocation"]
    for phrase in forbidden:
        assert phrase not in text
