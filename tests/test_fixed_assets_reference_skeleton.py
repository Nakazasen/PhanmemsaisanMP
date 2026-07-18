import csv
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.engine.fixed_assets_reference_skeleton import (
    PROVENANCE_LABEL,
    SHEET_NAME,
    apply_fixed_assets_reference_skeleton_to_workbook,
    load_fixed_assets_skeleton_candidates,
)
from scripts.run_e2e import run_universal_pipeline

CSV_PATH = Path("docs/audits/phase42n2e_5005026371_secondary_skeleton_patterns.csv")


def _make_workbook(path: Path, existing_row: int | None = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.cell(1, 2).value = "header"
    if existing_row:
        ws.cell(existing_row, 2).value = "EXISTING"
        ws.cell(existing_row, 19).value = "do not overwrite"
    wb.save(path)
    wb.close()
    return path


def _make_csv(path: Path, rows: list[dict[str, str]]) -> Path:
    fieldnames = [
        "account",
        "description",
        "month_F_sample",
        "month_Q_sample",
        "pattern_signature",
        "classification",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return path


def test_loads_only_48_candidates_from_42n2e_csv():
    candidates = load_fixed_assets_skeleton_candidates(CSV_PATH)
    assert len(candidates) == 48
    assert {row["classification"] for row in candidates} == {"REFERENCE_ASSISTED_FILL_CANDIDATE"}


def test_writes_provenance_label_and_summary(tmp_path):
    workbook = _make_workbook(tmp_path / "target.xlsx")
    csvp = _make_csv(
        tmp_path / "candidates.csv",
        [
            {
                "account": "5005026371",
                "description": "Skeleton A",
                "month_F_sample": "=ROUND(1*$B$2,0)",
                "month_Q_sample": "123",
                "pattern_signature": "F|B|B",
                "classification": "REFERENCE_ASSISTED_FILL_CANDIDATE",
            }
        ],
    )
    summary = apply_fixed_assets_reference_skeleton_to_workbook(workbook, csvp, start_row=10)
    assert summary == {
        "selected": 1,
        "written": 1,
        "skipped_existing": 0,
        "skipped_incomplete": 0,
        "start_row": 10,
    }
    wb = load_workbook(workbook, data_only=False)
    ws = wb[SHEET_NAME]
    assert ws.cell(10, 2).value == "5005026371"
    assert ws.cell(10, 19).value == "Skeleton A"
    assert ws.cell(10, 6).value == "=ROUND(1*$B$2,0)"
    assert ws.cell(10, 17).value == "123"
    assert ws.cell(10, 20).value == PROVENANCE_LABEL
    wb.close()


def test_does_not_overwrite_existing_rows(tmp_path):
    workbook = _make_workbook(tmp_path / "target.xlsx", existing_row=5)
    csvp = _make_csv(
        tmp_path / "candidates.csv",
        [
            {
                "account": "5005026371",
                "description": "Skeleton B",
                "month_F_sample": "1",
                "month_Q_sample": "2",
                "pattern_signature": "V|V",
                "classification": "REFERENCE_ASSISTED_FILL_CANDIDATE",
            }
        ],
    )
    summary = apply_fixed_assets_reference_skeleton_to_workbook(workbook, csvp, start_row=5)
    assert summary["written"] == 1
    wb = load_workbook(workbook)
    ws = wb[SHEET_NAME]
    assert ws.cell(5, 2).value == "EXISTING"
    assert ws.cell(6, 2).value == "5005026371"
    wb.close()


def test_does_not_invent_missing_values(tmp_path):
    workbook = _make_workbook(tmp_path / "target.xlsx")
    csvp = _make_csv(
        tmp_path / "candidates.csv",
        [
            {
                "account": "5005026371",
                "description": "",
                "month_F_sample": "",
                "month_Q_sample": "",
                "pattern_signature": "B|B",
                "classification": "REFERENCE_ASSISTED_FILL_CANDIDATE",
            }
        ],
    )
    summary = apply_fixed_assets_reference_skeleton_to_workbook(workbook, csvp, start_row=3)
    assert summary["written"] == 0
    assert summary["skipped_incomplete"] == 1
    wb = load_workbook(workbook)
    ws = wb[SHEET_NAME]
    assert ws.cell(3, 2).value is None
    assert ws.cell(3, 19).value is None
    assert ws.cell(3, 6).value is None
    assert ws.cell(3, 17).value is None
    wb.close()


def test_start_row_none_appends_after_last_business_row(tmp_path):
    workbook = _make_workbook(tmp_path / "target.xlsx", existing_row=8)
    csvp = _make_csv(
        tmp_path / "candidates.csv",
        [
            {
                "account": "5005026371",
                "description": "Append",
                "month_F_sample": "1",
                "month_Q_sample": "",
                "pattern_signature": "B|B",
                "classification": "REFERENCE_ASSISTED_FILL_CANDIDATE",
            }
        ],
    )
    summary = apply_fixed_assets_reference_skeleton_to_workbook(workbook, csvp)
    assert summary["start_row"] == 9
    wb = load_workbook(workbook)
    assert wb[SHEET_NAME].cell(9, 2).value == "5005026371"
    wb.close()


def test_cli_flag_default_off():
    import inspect

    parameter = inspect.signature(run_universal_pipeline).parameters[
        "fixed_assets_reference_skeleton_export"
    ]
    assert parameter.default is False


def test_duplicate_guard_with_primary_reference_fill(tmp_path):
    workbook = _make_workbook(tmp_path / "target.xlsx")
    csvp = _make_csv(tmp_path / "candidates.csv", [])
    ok, message = run_universal_pipeline(
        2027,
        str(workbook),
        str(tmp_path),
        target_cc=1412000040,
        primary_reference_fill=True,
        fixed_assets_reference_skeleton_export=True,
        fixed_assets_skeleton_csv=str(csvp),
    )
    assert ok is False
    assert "Nguy cơ trùng dữ liệu" in message
