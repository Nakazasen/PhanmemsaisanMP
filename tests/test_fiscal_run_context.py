import json
import os
from pathlib import Path

import pytest

from src.parsers.manual_event_drivers import _default_period_for_fiscal_year
from src.parsers.nnn_paperwork import month_map_for_fiscal_year
from src.services.fiscal_run import (
    _filename_period_coverage,
    annual_default_paths,
    create_fiscal_run_context,
    inspect_fiscal_year_evidence,
    preflight_fiscal_run,
    resolve_uniform_policy_path,
    validate_fiscal_year_evidence,
)
from src.utils.fiscal_periods import (
    SystemSourcePeriodError,
    fiscal_baseline_period,
    fiscal_periods,
    map_system_source_periods,
)
from src.utils.source_manifest import read_source_manifest
from src.services.project_config import (
    ProjectConfig,
    discover_or_create_project,
    read_last_project,
    remember_last_project,
)
from scripts.run_e2e import _resolve_primary_reference_path


def test_fiscal_periods_are_unbounded_and_keep_april_to_march_boundary():
    assert fiscal_periods(2028) == [
        "202704", "202705", "202706", "202707", "202708", "202709",
        "202710", "202711", "202712", "202801", "202802", "202803",
    ]
    assert fiscal_baseline_period(2028) == "202703"
    assert fiscal_periods(2029)[0] == "202804"
    assert fiscal_periods(2029)[-1] == "202903"


def test_new_fy_defaults_use_only_its_own_directories(tmp_path):
    defaults = annual_default_paths(2028, tmp_path)
    assert defaults["template_path"] == str(tmp_path / "docs" / "MP2028" / "FORM.xlsx")
    assert defaults["source_dir"] == str(tmp_path / "docs" / "MP2028")
    assert defaults["headcount_source_dir"] == str(tmp_path / "raw" / "FY2028")
    assert defaults["output_dir"] == str(tmp_path / "OUTPUT_FY2028")


def test_preflight_fy2028_never_falls_back_to_fy2027_sources(tmp_path):
    legacy_dir = tmp_path / "raw"
    legacy_dir.mkdir(parents=True)
    (legacy_dir / "only_FY2027.xlsx").write_bytes(b"not an FY2028 policy")
    (tmp_path / "docs" / "MP2028").mkdir(parents=True)
    (tmp_path / "raw" / "FY2028").mkdir(parents=True)

    context = create_fiscal_run_context(2028, base_dir=tmp_path)
    report = preflight_fiscal_run(context)

    assert not report.ok
    assert context.uniform_policy_path is None
    assert not report.resolved_sources["uniform_policy"]
    assert any(issue.category == "uniform_policy" for issue in report.issues)
    assert all("FY2027" not in " ".join(paths) for paths in report.resolved_sources.values())
    checks = report.as_dict()["checks"]
    assert any(check["category"] == "uniform_policy" and check["status"] == "SKIPPED" for check in checks)
    assert "Kiểm tra nguồn trước khi chạy FY2028" in report.as_markdown()


def test_dynamic_event_and_nnn_month_mappings_use_selected_fiscal_year():
    assert _default_period_for_fiscal_year({"period": "202604"}, 2028) == "202704"
    assert _default_period_for_fiscal_year({"period": "202702"}, 2028) == "202802"

    mapping = month_map_for_fiscal_year(2028)
    assert mapping["202704"] == 0
    assert mapping["202803"] == 11
    assert "202604" not in mapping


def test_system_source_filename_ranges_cover_the_selected_fiscal_year():
    covered = set()
    for name in (
        "system_FY2028_Apr.2027 ~ June.2027.xls",
        "system_FY2028_July.2027 ~ Dec.2027.xls",
        "system_FY2028_Jan.2028 ~ March.2028.xls",
    ):
        covered.update(_filename_period_coverage(name, 2028))
    assert covered == set(fiscal_periods(2028))


@pytest.mark.parametrize(
    "paths",
    [
        ["system_FY2028_Apr.2027 ~ March.2028.xls"],
        [
            "system_FY2028_Apr.2027 ~ Sep.2027.xls",
            "system_FY2028_Oct.2027 ~ March.2028.xls",
        ],
        [
            "system_FY2028_Apr.2027 ~ June.2027.xls",
            "system_FY2028_July.2027 ~ Dec.2027.xls",
            "system_FY2028_Jan.2028 ~ March.2028.xls",
        ],
        [
            "system_FY2028_Apr.2027 ~ June.2027.xls",
            "system_FY2028_July.2027 ~ Sep.2027.xls",
            "system_FY2028_Oct.2027 ~ Dec.2027.xls",
            "system_FY2028_Jan.2028 ~ March.2028.xls",
        ],
    ],
)
def test_system_source_mapping_accepts_one_to_four_files(paths):
    assignments = map_system_source_periods(paths, 2028)
    covered = [period for assignment in assignments for period in assignment.periods]
    assert covered == fiscal_periods(2028)


def test_system_source_mapping_is_independent_of_input_order():
    paths = [
        "system_FY2028_Jan.2028 ~ March.2028.xls",
        "system_FY2028_Apr.2027 ~ June.2027.xls",
        "system_FY2028_July.2027 ~ Dec.2027.xls",
    ]
    assignments = map_system_source_periods(paths, 2028)
    assert [Path(item.path).name for item in assignments] == [
        "system_FY2028_Apr.2027 ~ June.2027.xls",
        "system_FY2028_July.2027 ~ Dec.2027.xls",
        "system_FY2028_Jan.2028 ~ March.2028.xls",
    ]


@pytest.mark.parametrize(
    ("paths", "expected_code"),
    [
        (["system_FY2028_Apr.2027 ~ Feb.2028.xls"], "SYSTEM_PERIOD_MISSING"),
        (
            [
                "system_FY2028_Apr.2027 ~ Dec.2027.xls",
                "system_FY2028_Dec.2027 ~ March.2028.xls",
            ],
            "SYSTEM_PERIOD_OVERLAP",
        ),
        (["system_FY2028_unknown.xls"], "SYSTEM_PERIOD_UNRECOGNIZED"),
        (["system_FY2028_March.2028 ~ Apr.2027.xls"], "SYSTEM_PERIOD_REVERSED"),
        (
            ["system_FY2028_Apr.2027 ~ Sep.2027 ~ March.2028.xls"],
            "SYSTEM_PERIOD_AMBIGUOUS",
        ),
        (["system_FY2028_March.2027 ~ March.2028.xls"], "SYSTEM_PERIOD_OUTSIDE_FY"),
    ],
)
def test_system_source_mapping_fails_closed(paths, expected_code):
    with pytest.raises(SystemSourcePeriodError) as exc_info:
        map_system_source_periods(paths, 2028)
    assert exc_info.value.code == expected_code


def test_manifest_cannot_escape_selected_annual_source_directory(tmp_path):
    source_dir = tmp_path / "docs" / "MP2028"
    source_dir.mkdir(parents=True)
    (tmp_path / "docs" / "MP2027.xlsx").write_bytes(b"legacy")
    (source_dir / "source_file_order.csv").write_text(
        "order,category,filename,enabled\n1,facility,../MP2027.xlsx,1\n",
        encoding="utf-8",
    )
    entries = read_source_manifest(str(source_dir), include_missing=True)
    assert entries[0]["_invalid_path"] == "1"


def test_future_manual_input_store_rejects_rows_from_another_fy(tmp_path):
    from src.db.schema import create_schema, get_connection

    annual_dir = tmp_path / "raw" / "FY2028"
    annual_dir.mkdir(parents=True)
    store = annual_dir / "manual_inputs.db"
    conn = get_connection(str(store))
    create_schema(conn)
    conn.execute(
        "INSERT INTO fact_bus_headcount_drivers(cc_code,fiscal_year,bus_expat_count) VALUES('CC',2027,1)"
    )
    conn.commit()
    conn.close()
    context = create_fiscal_run_context(
        2028,
        headcount_source_dir=annual_dir,
        manual_input_store=str(store),
        base_dir=tmp_path,
    )

    report = preflight_fiscal_run(context)

    assert any(issue.category == "manual_inputs" for issue in report.issues)


def test_future_fy_reference_must_be_explicit_and_same_year(tmp_path):
    legacy = tmp_path / "reference FY2027.xlsx"
    legacy.write_bytes(b"not a workbook")
    try:
        _resolve_primary_reference_path("CC", str(legacy), fiscal_year=2028)
    except ValueError as exc:
        assert "FY2028" in str(exc)
    else:
        raise AssertionError("cross-year reference must be rejected")

    try:
        _resolve_primary_reference_path("CC", None, fiscal_year=2028)
    except ValueError as exc:
        assert "FY từ 2028" in str(exc)
    else:
        raise AssertionError("future FY reference must be explicit")

    unknown = tmp_path / "reference.xlsx"
    unknown.write_bytes(b"not a workbook")
    try:
        _resolve_primary_reference_path("CC", str(unknown), fiscal_year=2028)
    except ValueError as exc:
        assert "FY2028" in str(exc)
    else:
        raise AssertionError("future FY reference with no annual evidence must be rejected")


def test_fy_filename_cannot_mask_a_conflicting_business_sheet(tmp_path):
    from openpyxl import Workbook

    workbook_path = tmp_path / "reference_FY2028.xlsx"
    workbook = Workbook()
    workbook.active.title = "FY2027"
    workbook.active["A1"] = "FY2027 approved result"
    workbook.save(workbook_path)

    evidence = inspect_fiscal_year_evidence(workbook_path)
    assert evidence.conflict
    assert validate_fiscal_year_evidence(evidence, 2028)
    try:
        _resolve_primary_reference_path("CC", str(workbook_path), fiscal_year=2028)
    except ValueError as exc:
        assert "FY2028" in str(exc)
    else:
        raise AssertionError("conflicting evidence must be rejected")


def test_project_paths_are_relative_and_portable_when_project_moves(tmp_path):
    original = tmp_path / "original"
    original.mkdir()
    project = ProjectConfig.create_legacy_compatible(str(original), 2027)
    project.save()

    payload = json.loads((original / "project.json").read_text(encoding="utf-8"))
    assert payload["operational_database"] == "mp2027.db"
    assert payload["fiscal_years"]["2027"]["manual_input_store"] == "raw/FY2027/manual_inputs.db"

    moved = tmp_path / "moved"
    original.rename(moved)
    reloaded = ProjectConfig.load(str(moved / "project.json"))
    paths = reloaded.fiscal_paths(2027)

    assert reloaded.operational_database == os.path.abspath(moved / "mp2027.db")
    assert paths.template_path == os.path.abspath(moved / "docs" / "MP2027" / "FORM.xlsx")
    assert paths.manual_input_store == os.path.abspath(moved / "raw" / "FY2027" / "manual_inputs.db")


def test_manual_input_store_is_physically_isolated_by_fiscal_year(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)
    project.ensure_fiscal_year(2028)

    fy2027 = project.fiscal_paths(2027)
    fy2028 = project.fiscal_paths(2028)

    assert fy2027.manual_input_store != fy2028.manual_input_store
    assert fy2027.manual_input_store.endswith(os.path.join("raw", "FY2027", "manual_inputs.db"))
    assert fy2028.manual_input_store.endswith(os.path.join("raw", "FY2028", "manual_inputs.db"))


def test_recent_project_is_loaded_automatically_from_local_app_data(tmp_path):
    app_dir = tmp_path / "application"
    project_dir = tmp_path / "business-data"
    local_app_data = tmp_path / "local-app-data"
    app_dir.mkdir()
    project_dir.mkdir()

    project = ProjectConfig.create_legacy_compatible(str(project_dir), 2027)
    project.save()
    remember_last_project(project.config_path, local_app_data=str(local_app_data))

    assert read_last_project(local_app_data=str(local_app_data)) == project.config_path
    discovered, created = discover_or_create_project(
        str(app_dir), 2027, local_app_data=str(local_app_data)
    )

    assert created is False
    assert discovered.config_path == project.config_path
    assert not (app_dir / "project.json").exists()


def test_project_rejects_shared_manual_store_without_partial_mutation(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)
    project.ensure_fiscal_year(2028)
    original = json.dumps(project.data, sort_keys=True)
    shared_store = project.fiscal_paths(2027).manual_input_store

    try:
        project.update_fiscal_paths(2028, manual_input_store=shared_store)
    except ValueError as exc:
        assert "không được dùng chung kho nhập tay" in str(exc)
    else:
        raise AssertionError("two fiscal years must not share one manual input store")

    assert json.dumps(project.data, sort_keys=True) == original


def test_project_rejects_operational_database_that_is_a_manual_store(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)
    original = json.dumps(project.data, sort_keys=True)
    manual_store = project.fiscal_paths(2027).manual_input_store

    try:
        project.set_operational_database(manual_store)
    except ValueError as exc:
        assert "không được trùng kho nhập tay FY2027" in str(exc)
    else:
        raise AssertionError("operational database must not also be a manual input store")

    assert json.dumps(project.data, sort_keys=True) == original


def test_project_load_rejects_manual_store_reused_by_two_fiscal_years(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)
    project.ensure_fiscal_year(2028)
    project.data["fiscal_years"]["2028"]["manual_input_store"] = (
        project.data["fiscal_years"]["2027"]["manual_input_store"]
    )
    project.save()

    try:
        ProjectConfig.load(project.config_path)
    except ValueError as exc:
        assert "không được dùng chung kho nhập tay" in str(exc)
    else:
        raise AssertionError("invalid manually edited project.json must be rejected at load time")


def _write_minimal_uniform_policy(path: Path) -> None:
    from openpyxl import Workbook
    from src.engine.uniform_cup_rules import UNIFORM_ITEM_SPECS

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原価センタ"
    sheet.cell(1, 1, "原価センタ")
    for column, spec in enumerate(UNIFORM_ITEM_SPECS, start=2):
        sheet.cell(1, column, spec.header)
    workbook.save(path)
    workbook.close()


def test_new_project_leaves_uniform_policy_unset_for_annual_discovery(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)

    assert project.data["fiscal_years"]["2027"]["uniform_policy"] == ""
    assert project.fiscal_paths(2027).uniform_policy_path is None


def test_missing_generated_uniform_placeholder_is_migrated_as_unset(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)
    project.data["fiscal_years"]["2027"]["uniform_policy"] = (
        "docs/MP2027/uniform_eligibility.xlsx"
    )

    assert project.fiscal_paths(2027).uniform_policy_path is None


def test_missing_custom_uniform_policy_remains_explicit_and_fail_closed(tmp_path):
    project = ProjectConfig.create_legacy_compatible(str(tmp_path), 2027)
    project.data["fiscal_years"]["2027"]["uniform_policy"] = "custom/missing.xlsx"

    selected = project.fiscal_paths(2027).uniform_policy_path
    context = create_fiscal_run_context(2027, uniform_policy_path=selected, base_dir=tmp_path)
    report = preflight_fiscal_run(context)
    uniform_issues = [issue for issue in report.issues if issue.category == "uniform_policy"]

    assert selected == os.path.abspath(tmp_path / "custom" / "missing.xlsx")
    assert context.uniform_policy_path == selected
    assert len(uniform_issues) == 1
    assert "không tồn tại" in uniform_issues[0].reason
    assert not any(issue.category == "form_uniform_master" for issue in report.issues)


def test_uniform_discovery_follows_moved_project_root(tmp_path):
    original = tmp_path / "original"
    original.mkdir()
    policy = original / "raw" / "canonical.xlsx"
    _write_minimal_uniform_policy(policy)
    project = ProjectConfig.create_legacy_compatible(str(original), 2027)
    project.save()

    moved = tmp_path / "moved"
    original.rename(moved)
    reloaded = ProjectConfig.load(str(moved / "project.json"))
    paths = reloaded.fiscal_paths(2027)
    context = create_fiscal_run_context(
        2027,
        uniform_policy_path=paths.uniform_policy_path,
        base_dir=reloaded.root_dir,
    )

    assert context.uniform_policy_path == os.path.abspath(moved / "raw" / "canonical.xlsx")


def test_future_fy_uniform_discovery_is_annual_and_rejects_ambiguity(tmp_path):
    legacy = tmp_path / "raw" / "legacy_FY2027.xlsx"
    annual_a = tmp_path / "raw" / "FY2028" / "policy_a_FY2028.xlsx"
    annual_b = tmp_path / "raw" / "FY2028" / "policy_b_FY2028.xlsx"
    _write_minimal_uniform_policy(legacy)
    _write_minimal_uniform_policy(annual_a)

    assert resolve_uniform_policy_path(2028, base_dir=tmp_path) == str(annual_a.resolve())

    annual_a.unlink()
    assert resolve_uniform_policy_path(2028, base_dir=tmp_path) is None

    _write_minimal_uniform_policy(annual_a)
    _write_minimal_uniform_policy(annual_b)
    try:
        resolve_uniform_policy_path(2028, base_dir=tmp_path)
    except ValueError as exc:
        assert "nhiều file" in str(exc)
    else:
        raise AssertionError("ambiguous annual uniform policies must be rejected")


def _preflight_cache_context(tmp_path):
    from src.services.fiscal_run import create_fiscal_run_context

    source_dir = tmp_path / "sources"
    headcount_dir = tmp_path / "headcount"
    source_dir.mkdir()
    headcount_dir.mkdir()
    return create_fiscal_run_context(
        2028,
        template_path=str(tmp_path / "FORM.xlsx"),
        source_dir=str(source_dir),
        headcount_source_dir=str(headcount_dir),
        manual_input_store=str(tmp_path / "manual.db"),
        base_dir=tmp_path,
    )


def test_preflight_cache_hit_avoids_second_deep_check(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = str(tmp_path / "preflight_cache.json")
    calls = []

    def checker(_context):
        calls.append("deep")
        return RunPreflightReport(2028)

    first, first_hit = cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)
    second, second_hit = cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)

    assert first.ok and second.ok
    assert not first_hit
    assert second_hit
    assert calls == ["deep"]


def test_preflight_cache_misses_when_source_changes(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = str(tmp_path / "preflight_cache.json")
    calls = []

    def checker(_context):
        calls.append("deep")
        return RunPreflightReport(2028)

    cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)
    (tmp_path / "sources" / "facility.xlsx").write_bytes(b"changed")
    _, cache_hit = cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)

    assert not cache_hit
    assert calls == ["deep", "deep"]


def test_preflight_cache_force_refresh_bypasses_valid_entry(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = str(tmp_path / "preflight_cache.json")
    calls = []

    def checker(_context):
        calls.append("deep")
        return RunPreflightReport(2028)

    cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)
    _, cache_hit = cached_preflight_fiscal_run(
        context, cache_path=cache_path, force_refresh=True, checker=checker
    )

    assert not cache_hit
    assert calls == ["deep", "deep"]


def test_preflight_cache_corruption_falls_back_to_deep_check(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = tmp_path / "preflight_cache.json"
    cache_path.write_text("not-json", encoding="utf-8")
    calls = []

    def checker(_context):
        calls.append("deep")
        return RunPreflightReport(2028)

    report, cache_hit = cached_preflight_fiscal_run(
        context, cache_path=str(cache_path), checker=checker
    )

    assert report.ok
    assert not cache_hit
    assert calls == ["deep"]


def test_preflight_skips_unknown_source_without_making_it_a_global_blocker(tmp_path):
    import openpyxl

    source_dir = tmp_path / "sources"
    headcount_dir = tmp_path / "headcount"
    source_dir.mkdir()
    headcount_dir.mkdir()
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "unrecognized business source"
    workbook.save(source_dir / "new_cost_source.xlsx")
    context = create_fiscal_run_context(
        2028,
        template_path=str(tmp_path / "FORM.xlsx"),
        source_dir=str(source_dir),
        headcount_source_dir=str(headcount_dir),
        manual_input_store=str(tmp_path / "manual.db"),
        base_dir=tmp_path,
    )

    report = preflight_fiscal_run(context)

    review_issues = [issue for issue in report.issues if issue.code == "SOURCE_NEEDS_REVIEW"]
    assert review_issues
    assert review_issues[0].severity == "SOURCE_SKIPPED"
    assert "new_cost_source.xlsx" in review_issues[0].path
    assert review_issues[0] not in report.blocking_issues


def test_preflight_cache_ignores_nested_annual_source_files(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = str(tmp_path / "preflight_cache.json")
    calls = []

    def checker(_context):
        calls.append("deep")
        return RunPreflightReport(2028)

    cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)
    nested = tmp_path / "sources" / "archive"
    nested.mkdir()
    (nested / "old.xlsx").write_bytes(b"not consumed by top-level scanner")
    _, cache_hit = cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)

    assert cache_hit
    assert calls == ["deep"]


def test_preflight_cache_tracks_nested_headcount_files(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = str(tmp_path / "preflight_cache.json")
    calls = []

    def checker(_context):
        calls.append("deep")
        return RunPreflightReport(2028)

    cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)
    nested = tmp_path / "headcount" / "monthly"
    nested.mkdir()
    (nested / "apr.xlsx").write_bytes(b"tracked recursively")
    _, cache_hit = cached_preflight_fiscal_run(context, cache_path=cache_path, checker=checker)

    assert not cache_hit
    assert calls == ["deep", "deep"]


def test_preflight_cache_saves_report_against_post_check_state(tmp_path):
    from src.services.fiscal_run import RunPreflightReport
    from src.services.preflight_cache import cached_preflight_fiscal_run

    context = _preflight_cache_context(tmp_path)
    cache_path = str(tmp_path / "preflight_cache.json")
    calls = []

    def checker(_context):
        calls.append("deep")
        (tmp_path / "manual.db").write_bytes(b"initialized during preflight")
        return RunPreflightReport(2028)

    _, first_hit = cached_preflight_fiscal_run(
        context, cache_path=cache_path, checker=checker
    )
    _, second_hit = cached_preflight_fiscal_run(
        context, cache_path=cache_path, checker=checker
    )

    assert not first_hit
    assert second_hit
    assert calls == ["deep"]
