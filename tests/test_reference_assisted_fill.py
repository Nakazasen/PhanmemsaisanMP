import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.engine.reference_assisted_fill import (
    PROVENANCE_LABEL,
    apply_reference_assisted_fill_to_workbook,
    count_business_rows,
    identify_reference_assisted_rows,
    plan_reference_assisted_rows,
)

SHEET = "内訳ﾘｽﾄ(4～3月)"


def _make_workbook(path: Path, rows: list[int]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for row in rows:
        ws.cell(row, 2).value = "5005026371"
        ws.cell(row, 19).value = f"desc {row}"
        for col in range(6, 18):
            ws.cell(row, col).value = row * 100 + col
    wb.save(path)


def _make_invariant(path: Path) -> None:
    path.write_text(
        "primary_row,account_code,primary_description,old_source_status,new_classification,"
        "generated_match_row,reference_fill_scope,reference_fill_reason,nontech_reason\n"
        "10,500,already,SOURCE,ALREADY_GENERATED_FALSE_GAP,,,,\n"
        "11,501,matched,SOURCE,REAL_MISSING_SOURCE_UNCONFIRMED,205,,,\n"
        "12,502,unscoped,SOURCE,REAL_MISSING_SOURCE_UNCONFIRMED,,,,needs source mapping\n"
        "13,503,scoped,SOURCE,SCOPED_REFERENCE_FILL_ALLOWED,,,"
        "explicit scoped mapping for this CC,\n",
        encoding="utf-8",
    )


def test_identify_skips_false_gap_generated_match_and_unscoped_rows(tmp_path):
    primary = tmp_path / "primary.xlsx"
    generated = tmp_path / "generated.xlsx"
    invariant = tmp_path / "inv.csv"
    _make_workbook(primary, [10, 11, 12, 13])
    _make_workbook(generated, [200])
    _make_invariant(invariant)

    rows = identify_reference_assisted_rows(primary, generated, invariant)

    assert [row["primary_row"] for row in rows] == [13]


def test_plan_quarantines_unscoped_reference_rows(tmp_path):
    primary = tmp_path / "primary.xlsx"
    generated = tmp_path / "generated.xlsx"
    invariant = tmp_path / "inv.csv"
    _make_workbook(primary, [10, 11, 12, 13])
    _make_workbook(generated, [200])
    _make_invariant(invariant)

    plan = plan_reference_assisted_rows(primary, generated, invariant)

    assert [row["primary_row"] for row in plan["selected"]] == [13]
    assert [row["primary_row"] for row in plan["quarantined"]] == ["12"]
    assert plan["quarantined"][0]["quarantine_reason"] == "unscoped reference-assisted fill is not allowed"
    assert plan["skipped_false_gap"] == 1
    assert plan["skipped_generated_match"] == 1


def test_apply_fill_does_not_overwrite_200_212_and_writes_scoped_provenance(tmp_path):
    primary = tmp_path / "primary.xlsx"
    generated = tmp_path / "generated.xlsx"
    invariant = tmp_path / "inv.csv"
    _make_workbook(primary, [10, 11, 12, 13])
    _make_workbook(generated, [200, 211, 212])
    _make_invariant(invariant)

    result = apply_reference_assisted_fill_to_workbook(generated, primary, invariant, start_row=200)

    wb = load_workbook(generated, data_only=False)
    ws = wb[SHEET]
    assert result["written"] == 1
    assert result["quarantined"] == 1
    assert ws.cell(200, 19).value == "desc 200"
    assert ws.cell(211, 19).value == "desc 211"
    assert ws.cell(212, 19).value == "desc 212"
    assert ws.cell(213, 19).value == "desc 13"
    assert ws.cell(213, 6).value == 1306
    assert ws.cell(214, 19).value is None
    assert PROVENANCE_LABEL in ws.cell(213, 20).value
    assert "scoped-reference-fill" in ws.cell(213, 20).value
    assert "explicit scoped mapping for this CC" in ws.cell(213, 20).value
    assert "not source-derived" in ws.cell(213, 20).value
    wb.close()

    with open(result["quarantine_csv"], newline="", encoding="utf-8-sig") as handle:
        quarantine_rows = list(csv.DictReader(handle))
    assert [row["primary_row"] for row in quarantine_rows] == ["12"]
    assert quarantine_rows[0]["quarantine_reason"] == "unscoped reference-assisted fill is not allowed"


def test_count_business_rows(tmp_path):
    workbook = tmp_path / "book.xlsx"
    _make_workbook(workbook, [1, 3, 5])
    assert count_business_rows(workbook) == 3
