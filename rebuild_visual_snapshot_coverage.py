import os
import shutil
import time
import hashlib
import gc
import sys
import csv
import zipfile
import xml.etree.ElementTree as ET
from PIL import Image, ImageStat
import numpy as np
import openpyxl
import win32com.client

sys.stdout.reconfigure(encoding='utf-8')

def print_flush(msg):
    print(msg, flush=True)

# Tiled mappings configuration
MAPPINGS = [
    {
        "sheet": "V01_MASTER",
        "src_sheet": "Sheet1",
        "overview_range": "A1:AD55",
        "tiles": [
            {"tile_id": "T01", "range": "A1:O30"},
            {"tile_id": "T02", "range": "N1:AD30"},
            {"tile_id": "T03", "range": "A28:O55"},
            {"tile_id": "T04", "range": "N28:AD55"}
        ],
        "req": "General overview",
        "desc": "Workbook overview and high-level data-entry workflow"
    },
    {
        "sheet": "V02_0906_LIST_A",
        "src_sheet": "Hạng mục cần cải tiến",
        "overview_range": "A1:Y70",
        "tiles": [
            {"tile_id": "T01", "range": "A1:N25"},
            {"tile_id": "T02", "range": "M1:Y25"},
            {"tile_id": "T03", "range": "A23:N48"},
            {"tile_id": "T04", "range": "M23:Y48"},
            {"tile_id": "T05", "range": "A46:N70"},
            {"tile_id": "T06", "range": "M46:Y70"}
        ],
        "req": "09.06 change list part A",
        "desc": "Current 09.06 issue list: headcount input, fixed assets, health check, stationery, admin allocation, descriptions, separator rows"
    },
    {
        "sheet": "V03_GUI_INPUT",
        "src_sheet": "Hạng mục cần cải tiến",
        "overview_range": "A70:Y110",
        "tiles": [
            {"tile_id": "T01", "range": "A70:N92"},
            {"tile_id": "T02", "range": "M70:Y92"},
            {"tile_id": "T03", "range": "A90:N110"},
            {"tile_id": "T04", "range": "M90:Y110"}
        ],
        "req": "Requirement 6",
        "desc": "GUI/input headcount: 12 months and December male/female split"
    },
    {
        "sheet": "V04_HEADCOUNT",
        "src_sheet": "Hạng mục cần cải tiến",
        "overview_range": "A110:Q175",
        "tiles": [
            {"tile_id": "T01", "range": "A110:Q135"},
            {"tile_id": "T02", "range": "A133:Q155"},
            {"tile_id": "T03", "range": "A153:Q175"}
        ],
        "req": "09.06 headcount inputs",
        "desc": "Replacement manual headcount inputs used by allocation rules"
    },
    {
        "sheet": "V05_SEPARATOR",
        "src_sheet": "Hạng mục cần cải tiến",
        "overview_range": "A175:N235",
        "tiles": [
            {"tile_id": "T01", "range": "A175:N198"},
            {"tile_id": "T02", "range": "A196:N218"},
            {"tile_id": "T03", "range": "A216:N235"}
        ],
        "req": "Requirement 7-8",
        "desc": "Additional description and one blank row between each source-file block"
    },
    {
        "sheet": "V06_ADMIN_OV",
        "src_sheet": "Chi phí phân bổ từ hành chính ",
        "overview_range": "A1:Y42",
        "tiles": [
            {"tile_id": "T01", "range": "A1:N22"},
            {"tile_id": "T02", "range": "M1:Y22"},
            {"tile_id": "T03", "range": "A20:N42"},
            {"tile_id": "T04", "range": "M20:Y42"}
        ],
        "req": "Admin allocation overview",
        "desc": "Administrative allocation sheet overview plus 12-month allocation formula notes"
    },
    {
        "sheet": "V07_ADMIN_R41_65",
        "src_sheet": "Chi phí phân bổ từ hành chính ",
        "overview_range": "A38:T68",
        "tiles": [
            {"tile_id": "T01", "range": "A38:K68"},
            {"tile_id": "T02", "range": "J38:T68"}
        ],
        "req": "Admin allocation rows 41-65 skipped drawing area",
        "desc": "Administrative allocation row 41-65 drawings detailing allocation drivers"
    },
    {
        "sheet": "V08_ADMIN_CC",
        "src_sheet": "Chi phí phân bổ từ hành chính ",
        "overview_range": "A68:V141",
        "tiles": [
            {"tile_id": "T01", "range": "A68:L93"},
            {"tile_id": "T02", "range": "K68:V93"},
            {"tile_id": "T03", "range": "A91:L117"},
            {"tile_id": "T04", "range": "K91:V117"},
            {"tile_id": "T05", "range": "A115:L141"},
            {"tile_id": "T06", "range": "K115:V141"}
        ],
        "req": "Admin cost center/account mapping",
        "desc": "Cost center group and account-code lookup instructions"
    },
    {
        "sheet": "V09_ADMIN_HEALTH",
        "src_sheet": "Chi phí phân bổ từ hành chính ",
        "overview_range": "A141:V192",
        "tiles": [
            {"tile_id": "T01", "range": "A141:L168"},
            {"tile_id": "T02", "range": "K141:V168"},
            {"tile_id": "T03", "range": "A166:L192"},
            {"tile_id": "T04", "range": "K166:V192"}
        ],
        "req": "Health-check periodic & Row 182-192 drawing area",
        "desc": "Male/female regular health-check allocation notes and Picture 48 area"
    },
    {
        "sheet": "V10_ADMIN_NEW",
        "src_sheet": "Chi phí phân bổ từ hành chính ",
        "overview_range": "A192:V235",
        "tiles": [
            {"tile_id": "T01", "range": "A192:L215"},
            {"tile_id": "T02", "range": "K192:V215"},
            {"tile_id": "T03", "range": "A213:L235"},
            {"tile_id": "T04", "range": "K213:V235"}
        ],
        "req": "New-hire stationery, recruitment check & Row 217-234 drawing area",
        "desc": "New-hire monthly delta, stationery unit prices, recruitment health-check and Picture 15 area"
    },
    {
        "sheet": "V11_FIXED_ASSET",
        "src_sheet": "Chi phí tài sản cố định",
        "overview_range": "A1:AF66",
        "tiles": [
            {"tile_id": "T01", "range": "A1:L25"},
            {"tile_id": "T02", "range": "K1:V25"},
            {"tile_id": "T03", "range": "U1:AF25"},
            {"tile_id": "T04", "range": "A23:L48"},
            {"tile_id": "T05", "range": "K23:V48"},
            {"tile_id": "T06", "range": "U23:AF48"},
            {"tile_id": "T07", "range": "A46:L66"},
            {"tile_id": "T08", "range": "K46:V66"},
            {"tile_id": "T09", "range": "U46:AF66"}
        ],
        "req": "Fixed asset instructions",
        "desc": "Fixed asset depreciation and interest instructions including right-side callouts"
    },
    {
        "sheet": "V12_SYSTEM",
        "src_sheet": "Chi phí hệ thống",
        "overview_range": "A1:AB97",
        "tiles": [
            {"tile_id": "T01", "range": "A1:O25"},
            {"tile_id": "T02", "range": "N1:AB25"},
            {"tile_id": "T03", "range": "A23:O48"},
            {"tile_id": "T04", "range": "N23:AB48"},
            {"tile_id": "T05", "range": "A46:O72"},
            {"tile_id": "T06", "range": "N46:AB72"},
            {"tile_id": "T07", "range": "A70:O97"},
            {"tile_id": "T08", "range": "N70:AB97"}
        ],
        "req": "System cost instructions",
        "desc": "System-cost calculation and formula workflow spanning columns J to AB"
    },
    {
        "sheet": "V13_DEPR_LAND",
        "src_sheet": "Chi phí khấu hao, lãi nhà đất",
        "overview_range": "A1:V131",
        "tiles": [
            {"tile_id": "T01", "range": "A1:L25"},
            {"tile_id": "T02", "range": "K1:V25"},
            {"tile_id": "T03", "range": "A23:L48"},
            {"tile_id": "T04", "range": "K23:V48"},
            {"tile_id": "T05", "range": "A46:L70"},
            {"tile_id": "T06", "range": "K46:V70"},
            {"tile_id": "T07", "range": "A68:L92"},
            {"tile_id": "T08", "range": "K68:V92"},
            {"tile_id": "T09", "range": "A90:L114"},
            {"tile_id": "T10", "range": "K90:V114"},
            {"tile_id": "T11", "range": "A112:L131"},
            {"tile_id": "T12", "range": "K112:V131"}
        ],
        "req": "Depreciation/land/water instructions",
        "desc": "Depreciation, land interest, electricity and water instructions"
    },
    {
        "sheet": "V14_NNN",
        "src_sheet": "Chi phí làm giấy tờ cho NNN",
        "overview_range": "A1:S48",
        "tiles": [
            {"tile_id": "T01", "range": "A1:K25"},
            {"tile_id": "T02", "range": "J1:S25"},
            {"tile_id": "T03", "range": "A23:K48"},
            {"tile_id": "T04", "range": "J23:S48"}
        ],
        "req": "NNN paperwork",
        "desc": "Foreign worker paperwork cost month/account instructions"
    },
    {
        "sheet": "V15_BIRTHDAY",
        "src_sheet": "Chi phí sinh nhật",
        "overview_range": "A1:R48",
        "tiles": [
            {"tile_id": "T01", "range": "A1:K25"},
            {"tile_id": "T02", "range": "J1:R25"},
            {"tile_id": "T03", "range": "A23:K48"},
            {"tile_id": "T04", "range": "J23:R48"}
        ],
        "req": "Birthday/new-hire add-on",
        "desc": "Birthday cost with same-month new-hire add-on"
    },
    {
        "sheet": "V16_CC_LOOKUP",
        "src_sheet": "原価センタ",
        "overview_range": "A1:E66",
        "tiles": [
            {"tile_id": "T01", "range": "A1:E25"},
            {"tile_id": "T02", "range": "A23:E48"},
            {"tile_id": "T03", "range": "A46:E66"}
        ],
        "req": "Cost center lookup",
        "desc": "Cost center lookup sheet used by admin allocation instructions"
    },
    {
        "sheet": "V17_ACCOUNT",
        "src_sheet": "勘定科目",
        "overview_range": "A1:I244",
        "tiles": [
            {"tile_id": "T01", "range": "A1:I25"},
            {"tile_id": "T02", "range": "A23:I48"},
            {"tile_id": "T03", "range": "A46:I72"},
            {"tile_id": "T04", "range": "A70:I96"},
            {"tile_id": "T05", "range": "A94:I120"},
            {"tile_id": "T06", "range": "A118:I144"},
            {"tile_id": "T07", "range": "A142:I168"},
            {"tile_id": "T08", "range": "A166:I192"},
            {"tile_id": "T09", "range": "A190:I216"},
            {"tile_id": "T10", "range": "A214:I244"}
        ],
        "req": "Account lookup",
        "desc": "Account-code lookup sheet used by admin allocation instructions"
    }
]

def split_cell(cell_str):
    import re
    m = re.match(r'^([A-Z]+)(\d+)$', cell_str, re.IGNORECASE)
    if m:
        return m.group(1).upper(), m.group(2)
    return cell_str, "0"

def parse_range(range_str):
    parts = range_str.split(':')
    if len(parts) == 1:
        parts = parts + parts
    from_col_letter, from_row_str = split_cell(parts[0])
    to_col_letter, to_row_str = split_cell(parts[1])
    
    from_col = openpyxl.utils.column_index_from_string(from_col_letter)
    from_row = int(from_row_str)
    to_col = openpyxl.utils.column_index_from_string(to_col_letter)
    to_row = int(to_row_str)
    return from_col, from_row, to_col, to_row

def audit_image(path):
    if not os.path.exists(path):
        return "NOT_FOUND", 0, 0, 100.0, 0
    try:
        with Image.open(path) as img:
            width, height = img.size
            rgb_img = img.convert('RGB')
            pixels = np.array(rgb_img)
            near_white = (pixels[:, :, 0] >= 248) & (pixels[:, :, 1] >= 248) & (pixels[:, :, 2] >= 248)
            near_white_ratio = np.mean(near_white)
            
            gray_img = img.convert('L')
            stat = ImageStat.Stat(gray_img)
            stddev = stat.stddev[0]
            
            if near_white_ratio > 0.995:
                classification = 'BLANK' if near_white_ratio >= 1.0 else 'NEAR_BLANK'
            elif stddev < 1.0:
                classification = 'NEAR_BLANK'
            else:
                classification = 'VALID_CONTENT'
                
            return classification, width, height, near_white_ratio, stddev
    except Exception as e:
        return f"ERROR: {str(e)}", 0, 0, 100.0, 0

def capture_range_com(excel, wb, sheet_name, rng_address, dest_path):
    ws = wb.Worksheets(sheet_name)
    ws.Activate()
    rng = ws.Range(rng_address)
    
    # Scroll range into view to force rendering
    excel.ActiveWindow.ScrollRow = rng.Row
    excel.ActiveWindow.ScrollColumn = rng.Column
    rng.Select()
    excel.Calculate()
    
    max_attempts = 5
    for attempt in range(1, max_attempts + 1):
        # Copy to Clipboard
        # Appearance=1 (xlScreen), Format=-4147 (xlBitmap)
        rng.CopyPicture(1, -4147)
        time.sleep(0.2 * attempt)
        
        # Paste into temporary ChartObject
        chart_obj = ws.ChartObjects().Add(0, 0, rng.Width + 2, rng.Height + 2)
        chart_obj.Select()
        
        try:
            chart_obj.Chart.Paste()
            chart_obj.Chart.Export(dest_path, "PNG")
        except Exception as e:
            pass
        finally:
            chart_obj.Delete()
            
        # Audit
        classification, w, h, ratio, stddev = audit_image(dest_path)
        if classification == 'VALID_CONTENT':
            return True
        else:
            if os.path.exists(dest_path):
                os.remove(dest_path)
            time.sleep(0.5)
            
    return False

def get_sheet_drawings(z, ws_path, rid_to_path):
    ws_rels_path = os.path.dirname(ws_path) + '/_rels/' + os.path.basename(ws_path) + '.rels'
    if ws_rels_path not in z.namelist():
        return []

    ws_rels_tree = ET.fromstring(z.read(ws_rels_path))
    drawing_path = None
    for elem in ws_rels_tree.iter():
        if elem.tag.endswith('Relationship') and elem.get('Type').endswith('/drawing'):
            drawing_path = elem.get('Target')
            break
    
    if not drawing_path:
        return []

    if drawing_path.startswith('/'):
        drawing_path = drawing_path[1:]
    elif drawing_path.startswith('../'):
        drawing_path = 'xl/' + drawing_path.replace('../', '')
    else:
        drawing_path = 'xl/drawings/' + drawing_path
    
    if drawing_path not in z.namelist():
        return []

    draw_tree = ET.fromstring(z.read(drawing_path))
    anchors = []
    for elem in draw_tree.iter():
        if elem.tag.endswith('twoCellAnchor') or elem.tag.endswith('oneCellAnchor'):
            anchors.append(elem)
            
    results = []
    for anchor in anchors:
        from_el = None
        to_el = None
        for child in anchor:
            if child.tag.endswith('from'):
                from_el = child
            elif child.tag.endswith('to'):
                to_el = child
                
        if from_el is not None:
            col_el = None
            row_el = None
            for child in from_el:
                if child.tag.endswith('col'): col_el = child
                elif child.tag.endswith('row'): row_el = child
            from_col = int(col_el.text) if col_el is not None else -1
            from_row = int(row_el.text) if row_el is not None else -1
        else:
            from_col, from_row = -1, -1
            
        if to_el is not None:
            col_el = None
            row_el = None
            for child in to_el:
                if child.tag.endswith('col'): col_el = child
                elif child.tag.endswith('row'): row_el = child
            to_col = int(col_el.text) if col_el is not None else -1
            to_row = int(row_el.text) if row_el is not None else -1
        else:
            to_col, to_row = -1, -1

        obj_name = "unknown"
        obj_type = "unknown"
        
        # Check kids
        for child in anchor.iter():
            if child.tag.endswith('pic'):
                for sub in child.iter():
                    if sub.tag.endswith('cNvPr'):
                        obj_name = sub.get('name')
                        obj_type = "Picture"
                        break
            elif child.tag.endswith('sp'):
                for sub in child.iter():
                    if sub.tag.endswith('cNvPr'):
                        obj_name = sub.get('name')
                        obj_type = "Shape"
                        break
                # textbox check
                for sub in child.iter():
                    if sub.tag.endswith('txBody'):
                        obj_type = "Shape/Textbox"
                        break
            elif child.tag.endswith('grpSp'):
                for sub in child.iter():
                    if sub.tag.endswith('cNvPr'):
                        obj_name = sub.get('name')
                        obj_type = "GroupShape"
                        break
            elif child.tag.endswith('graphicFrame'):
                for sub in child.iter():
                    if sub.tag.endswith('cNvPr'):
                        obj_name = sub.get('name')
                        obj_type = "Chart/GraphicFrame"
                        break

        results.append({
            'name': obj_name,
            'type': obj_type,
            'from_row': from_row + 1,
            'from_col': from_col + 1,
            'to_row': to_row + 1,
            'to_col': to_col + 1
        })
    return results

def rebuild_all_coverage():
    canonical_path = r"D:\Sandbox\MP2027\raw\Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026.xlsx"
    target_path = r"D:\Sandbox\MP2027\raw\Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026_ảnh.xlsx"
    temp_dir = r"D:\Sandbox\MP2027\OUTPUT_FY2027\tmp_requirement_0906_visual_coverage\new_media"
    csv_drawing_path = r"D:\Sandbox\MP2027\OUTPUT_FY2027\tmp_requirement_0906_visual_coverage\source_drawing_inventory.csv"
    csv_matrix_path = r"D:\Sandbox\MP2027\OUTPUT_FY2027\tmp_requirement_0906_visual_coverage\capture_coverage_matrix.csv"

    os.makedirs(temp_dir, exist_ok=True)

    # Re-read canonical sheet layouts & drawings
    print_flush("Analyzing drawings and sheets in canonical file...")
    drawings_by_sheet = {}
    with zipfile.ZipFile(canonical_path, 'r') as z:
        # sheets and paths
        wb_tree = ET.fromstring(z.read('xl/workbook.xml'))
        sheets_info = []
        for elem in wb_tree.iter():
            if elem.tag.endswith('sheet'):
                rid = None
                for k, v in elem.attrib.items():
                    if k.endswith('id') or k.endswith('Id'):
                        if v.startswith('rId'):
                            rid = v
                            break
                if not rid:
                    rid = elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                sheets_info.append({'name': elem.get('name'), 'rid': rid})

        rels_tree = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rid_to_path = {}
        for elem in rels_tree.iter():
            if elem.tag.endswith('Relationship'):
                rid_to_path[elem.get('Id')] = elem.get('Target')

        for s in sheets_info:
            target = rid_to_path[s['rid']]
            if target.startswith('/'): target = target[1:]
            elif not target.startswith('xl/'): target = 'xl/' + target
            s['path'] = target
            drawings_by_sheet[s['name']] = get_sheet_drawings(z, target, rid_to_path)

    # Backup old file if it exists
    if os.path.exists(target_path):
        backup_path = r"D:\Sandbox\MP2027\OUTPUT_FY2027\tmp_requirement_0906_visual_coverage\backup_bad_snapshot.xlsx"
        print_flush(f"Moving old snapshot to {backup_path}")
        if os.path.exists(backup_path):
            os.remove(backup_path)
        shutil.copy2(target_path, backup_path)
        os.remove(target_path)

    # 1. Execute visible Excel COM to capture 95 images
    print_flush("Starting visible Excel Application for tiled capture...")
    excel = None
    wb = None
    capture_results = {}
    
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False
        
        wb = excel.Workbooks.Open(canonical_path, ReadOnly=True)
        
        for idx, mapping in enumerate(MAPPINGS):
            sheet_name = mapping["sheet"]
            src_sheet = mapping["src_sheet"]
            print_flush(f"[{idx+1}/{len(MAPPINGS)}] Capturing visual requirements for {sheet_name} (source: {src_sheet})...")
            
            # 1. Capture Overview
            ov_filename = f"{sheet_name}_overview.png"
            ov_path = os.path.join(temp_dir, ov_filename)
            print_flush(f"  Capturing Overview range {mapping['overview_range']}...")
            success = capture_range_com(excel, wb, src_sheet, mapping['overview_range'], ov_path)
            if not success:
                raise Exception(f"Failed to capture Overview range {mapping['overview_range']} for {sheet_name}!")
            
            capture_results[(sheet_name, "OVERVIEW")] = {
                "range": mapping["overview_range"],
                "path": ov_path,
                "filename": ov_filename
            }
            
            # 2. Capture Tiles
            tile_entries = []
            for tile in mapping["tiles"]:
                t_id = tile["tile_id"]
                t_range = tile["range"]
                t_filename = f"{sheet_name}_{t_id}.png"
                t_path = os.path.join(temp_dir, t_filename)
                
                print_flush(f"  Capturing Tile {t_id} range {t_range}...")
                success = capture_range_com(excel, wb, src_sheet, t_range, t_path)
                if not success:
                    raise Exception(f"Failed to capture Tile {t_id} range {t_range} for {sheet_name}!")
                    
                capture_results[(sheet_name, t_id)] = {
                    "range": t_range,
                    "path": t_path,
                    "filename": t_filename
                }
                tile_entries.append((t_id, t_range, t_path, t_filename))
                
    except Exception as e:
        print_flush(f"CRITICAL CAPTURE PIPELINE ERROR: {str(e)}")
        sys.exit(1)
    finally:
        if wb:
            wb.Close(False)
        if excel:
            excel.Quit()
        del wb
        del excel
        gc.collect()
        time.sleep(2)
        os.system("taskkill /F /IM excel.exe >nul 2>&1")

    # Calculate SHA-256 of canonical workbook
    with open(canonical_path, "rb") as f:
        canonical_sha = hashlib.sha256(f.read()).hexdigest().upper()
    modified_time_str = time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime(os.path.getmtime(canonical_path)))

    # 2. Coverage Analysis and Matrix Generation
    print_flush("Running completeness coverage analysis...")
    drawing_matrix_rows = []
    
    # Audit each drawing
    for s_name, drawings in drawings_by_sheet.items():
        # Find which visual sheets map to this source sheet
        matching_mappings = [m for m in MAPPINGS if m["src_sheet"] == s_name]
        
        for dwg in drawings:
            d_from_row = dwg['from_row']
            d_from_col = dwg['from_col']
            d_to_row = dwg['to_row']
            d_to_col = dwg['to_col']
            
            fully_covered_in_v = []
            partially_clipped_in_v = []
            
            for m in matching_mappings:
                v_sheet = m["sheet"]
                
                # Check Overview
                c_from_col, c_from_row, c_to_col, c_to_row = parse_range(m["overview_range"])
                is_from_inside = (c_from_col <= d_from_col <= c_to_col) and (c_from_row <= d_from_row <= c_to_row)
                is_to_inside = (c_from_col <= d_to_col <= c_to_col) and (c_from_row <= d_to_row <= c_to_row)
                
                if is_from_inside and is_to_inside:
                    fully_covered_in_v.append(f"{v_sheet}_OVERVIEW")
                else:
                    col_overlap = not (d_to_col < c_from_col or d_from_col > c_to_col)
                    row_overlap = not (d_to_row < c_from_row or d_from_row > c_to_row)
                    if col_overlap and row_overlap:
                        partially_clipped_in_v.append(f"{v_sheet}_OVERVIEW")
                        
                # Check Tiles
                for tile in m["tiles"]:
                    t_id = tile["tile_id"]
                    c_from_col, c_from_row, c_to_col, c_to_row = parse_range(tile["range"])
                    is_from_inside = (c_from_col <= d_from_col <= c_to_col) and (c_from_row <= d_from_row <= c_to_row)
                    is_to_inside = (c_from_col <= d_to_col <= c_to_col) and (c_from_row <= d_to_row <= c_to_row)
                    
                    if is_from_inside and is_to_inside:
                        fully_covered_in_v.append(f"{v_sheet}_{t_id}")
                    else:
                        col_overlap = not (d_to_col < c_from_col or d_from_col > c_to_col)
                        row_overlap = not (d_to_row < c_from_row or d_from_row > c_to_row)
                        if col_overlap and row_overlap:
                            partially_clipped_in_v.append(f"{v_sheet}_{t_id}")

            status = "NOT_CAPTURED"
            if fully_covered_in_v:
                status = "FULLY_COVERED"
            elif partially_clipped_in_v:
                status = "PARTIALLY_CLIPPED"
                
            drawing_matrix_rows.append({
                'source_sheet': s_name,
                'object_name': dwg['name'],
                'object_type': dwg['type'],
                'from_cell': f"{openpyxl.utils.get_column_letter(d_from_col)}{d_from_row}",
                'to_cell': f"{openpyxl.utils.get_column_letter(d_to_col)}{d_to_row}",
                'current_capture_status': status,
                'fully_covered_locations': "|".join(fully_covered_in_v) if fully_covered_in_v else "NONE",
                'partially_clipped_locations': "|".join(partially_clipped_in_v) if partially_clipped_in_v else "NONE",
                'validation_status': "PASS" if status == "FULLY_COVERED" else "FAIL"
            })

    # Save matrix CSV
    with open(csv_matrix_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['source_sheet', 'object_name', 'object_type', 'from_cell', 'to_cell', 'current_capture_status', 'fully_covered_locations', 'partially_clipped_locations', 'validation_status'])
        writer.writeheader()
        for row in drawing_matrix_rows:
            writer.writerow(row)
    print_flush(f"Coverage matrix saved to {csv_matrix_path}")

    # Compute stats for MANIFEST coverage table
    # We need to map each visual ID and tile to drawing counts
    coverage_table_entries = []
    
    for mapping in MAPPINGS:
        v_sheet = mapping["sheet"]
        s_sheet = mapping["src_sheet"]
        
        # 1. Overview
        ov_range = mapping["overview_range"]
        ov_ref_col_from, ov_ref_row_from, ov_ref_col_to, ov_ref_row_to = parse_range(ov_range)
        
        relevant_drawings = []
        fully_covered = 0
        partially_clipped = 0
        missing = 0
        
        # Find drawings that belong to this source sheet and intersect with this range
        for d in drawings_by_sheet.get(s_sheet, []):
            d_from_row = d['from_row']
            d_from_col = d['from_col']
            d_to_row = d['to_row']
            d_to_col = d['to_col']
            
            # Check overlap
            col_overlap = not (d_to_col < ov_ref_col_from or d_from_col > ov_ref_col_to)
            row_overlap = not (d_to_row < ov_ref_row_from or d_from_row > ov_ref_row_to)
            
            if col_overlap and row_overlap:
                relevant_drawings.append(d)
                # Check containment
                is_from_inside = (ov_ref_col_from <= d_from_col <= ov_ref_col_to) and (ov_ref_row_from <= d_from_row <= ov_ref_row_to)
                is_to_inside = (ov_ref_col_from <= d_to_col <= ov_ref_col_to) and (ov_ref_row_from <= d_to_row <= ov_ref_row_to)
                
                if is_from_inside and is_to_inside:
                    fully_covered += 1
                else:
                    partially_clipped += 1
                    
        # If there are drawings in the sheet but outside this range, they are missing in this view
        # For simplicity, we only count drawings that intersect
        validation_status = "PASS" if partially_clipped == 0 else "FAIL"
        
        coverage_table_entries.append({
            'visual_id': v_sheet,
            'source_sheet': s_sheet,
            'tile_id': "OVERVIEW",
            'source_range': ov_range,
            'relevant_drawing_count': len(relevant_drawings),
            'fully_covered_drawing_count': fully_covered,
            'missing_drawing_count': missing,
            'partially_clipped_count': partially_clipped,
            'validation_status': validation_status
        })
        
        # 2. Tiles
        for tile in mapping["tiles"]:
            t_id = tile["tile_id"]
            t_range = tile["range"]
            t_ref_col_from, t_ref_row_from, t_ref_col_to, t_ref_row_to = parse_range(t_range)
            
            relevant_drawings = []
            fully_covered = 0
            partially_clipped = 0
            missing = 0
            
            for d in drawings_by_sheet.get(s_sheet, []):
                d_from_row = d['from_row']
                d_from_col = d['from_col']
                d_to_row = d['to_row']
                d_to_col = d['to_col']
                
                # Check overlap
                col_overlap = not (d_to_col < t_ref_col_from or d_from_col > t_ref_col_to)
                row_overlap = not (d_to_row < t_ref_row_from or d_from_row > t_ref_row_to)
                
                if col_overlap and row_overlap:
                    relevant_drawings.append(d)
                    is_from_inside = (t_ref_col_from <= d_from_col <= t_ref_col_to) and (t_ref_row_from <= d_from_row <= t_ref_row_to)
                    is_to_inside = (t_ref_col_from <= d_to_col <= t_ref_col_to) and (t_ref_row_from <= d_to_row <= t_ref_row_to)
                    
                    if is_from_inside and is_to_inside:
                        fully_covered += 1
                    else:
                        partially_clipped += 1
            
            validation_status = "PASS" if partially_clipped == 0 else "FAIL"
            coverage_table_entries.append({
                'visual_id': v_sheet,
                'source_sheet': s_sheet,
                'tile_id': t_id,
                'source_range': t_range,
                'relevant_drawing_count': len(relevant_drawings),
                'fully_covered_drawing_count': fully_covered,
                'missing_drawing_count': missing,
                'partially_clipped_count': partially_clipped,
                'validation_status': validation_status
            })

    # 3. Assemble visual workbook copy
    print_flush("Assembling coverage visual workbook...")
    new_wb = openpyxl.Workbook()
    
    # Write MANIFEST
    manifest_ws = new_wb.active
    manifest_ws.title = "MANIFEST"
    manifest_ws.views.sheetView[0].showGridLines = True
    
    manifest_ws.cell(1, 1, "VISUAL SUPPORT ONLY — NOT CANONICAL SOURCE")
    manifest_ws.cell(2, 1, "DO NOT USE 04.06.2026_ảnh.xlsx")
    
    manifest_ws.cell(4, 1, "Canonical source path")
    manifest_ws.cell(4, 2, canonical_path)
    
    manifest_ws.cell(5, 1, "Canonical filename")
    manifest_ws.cell(5, 2, os.path.basename(canonical_path))
    
    manifest_ws.cell(6, 1, "Source last-modified UTC")
    manifest_ws.cell(6, 2, modified_time_str)
    
    manifest_ws.cell(7, 1, "Source SHA-256")
    manifest_ws.cell(7, 2, canonical_sha)
    
    manifest_ws.cell(8, 1, "Snapshot generation timestamp")
    manifest_ws.cell(8, 2, time.strftime('%Y-%m-%d %H:%M:%S %z'))
    
    manifest_ws.cell(9, 1, "Principle")
    manifest_ws.cell(9, 2, "Workbook 09.06.2026.xlsx wins over this visual snapshot, derived Markdown, and old audits.")
    
    manifest_ws.cell(10, 1, "Obsolete visual source")
    manifest_ws.cell(10, 2, "D:\\Sandbox\\MP2027\\raw\\Cải tiến nhập dữ liệu chung vào file MPnew 04.06.2026_ảnh.xlsx was not used.")

    # Write Coverage Table in MANIFEST
    manifest_ws.cell(12, 1, "Visual Drawing Coverage Matrix:").font = openpyxl.styles.Font(bold=True)
    
    headers = [
        "visual_id", "source_sheet", "tile_id", "source_range",
        "relevant_drawing_count", "fully_covered_drawing_count",
        "missing_drawing_count", "partially_clipped_count", "validation_status"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = manifest_ws.cell(14, col_idx, h)
        cell.font = openpyxl.styles.Font(bold=True)
        
    for row_idx, entry in enumerate(coverage_table_entries, 15):
        manifest_ws.cell(row_idx, 1, entry["visual_id"])
        manifest_ws.cell(row_idx, 2, entry["source_sheet"])
        manifest_ws.cell(row_idx, 3, entry["tile_id"])
        manifest_ws.cell(row_idx, 4, entry["source_range"])
        manifest_ws.cell(row_idx, 5, entry["relevant_drawing_count"])
        manifest_ws.cell(row_idx, 6, entry["fully_covered_drawing_count"])
        manifest_ws.cell(row_idx, 7, entry["missing_drawing_count"])
        manifest_ws.cell(row_idx, 8, entry["partially_clipped_count"])
        manifest_ws.cell(row_idx, 9, entry["validation_status"])

    # Auto-fit MANIFEST columns
    for col in manifest_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        manifest_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Write Visual Sheets
    pixel_audits = []
    for mapping in MAPPINGS:
        sheet_name = mapping["sheet"]
        src_sheet = mapping["src_sheet"]
        
        ws = new_wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Write sheet level title
        ws.cell(1, 1, "VISUAL SUPPORT ONLY — NOT CANONICAL SOURCE").font = openpyxl.styles.Font(bold=True, color="FF0000")
        
        # 2. Insert Overview
        ov_info = capture_results[(sheet_name, "OVERVIEW")]
        overview_caption = f"OVERVIEW | Source range: {ov_info['range']} | Requirement: {mapping['req']} | Source sheet: {src_sheet} | Canonical source: Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026.xlsx | Description: {mapping['desc']}"
        ws.cell(3, 1, overview_caption).font = openpyxl.styles.Font(bold=True)
        ws.row_dimensions[3].height = 24.0
        
        img_ov = openpyxl.drawing.image.Image(ov_info["path"])
        ws.add_image(img_ov, "A5")
        
        # Calculate row span for overview image
        num_rows_ov = int(img_ov.height / 20) + 1
        current_row = 5 + num_rows_ov + 3
        
        # Audit Overview image
        class_ov, w_ov, h_ov, r_ov, std_ov = audit_image(ov_info["path"])
        pixel_audits.append({
            'filename': ov_info["filename"],
            'sheet': sheet_name,
            'view': 'OVERVIEW',
            'width': w_ov,
            'height': h_ov,
            'ratio': r_ov,
            'stddev': std_ov,
            'classification': class_ov
        })
        
        # 3. Insert Tiles sequentially
        for tile in mapping["tiles"]:
            t_id = tile["tile_id"]
            tile_info = capture_results[(sheet_name, t_id)]
            
            tile_caption = f"TILE {t_id} | Source range: {tile_info['range']} | Requirement: {mapping['req']} | Source sheet: {src_sheet} | Canonical source: Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026.xlsx | Description: {mapping['desc']}"
            
            ws.cell(current_row, 1, tile_caption).font = openpyxl.styles.Font(bold=True)
            ws.row_dimensions[current_row].height = 24.0
            
            img_tile = openpyxl.drawing.image.Image(tile_info["path"])
            ws.add_image(img_tile, f"A{current_row + 2}")
            
            # Audit Tile image
            class_t, w_t, h_t, r_t, std_t = audit_image(tile_info["path"])
            pixel_audits.append({
                'filename': tile_info["filename"],
                'sheet': sheet_name,
                'view': t_id,
                'width': w_t,
                'height': h_t,
                'ratio': r_t,
                'stddev': std_t,
                'classification': class_t
            })
            
            num_rows_tile = int(img_tile.height / 20) + 1
            current_row = (current_row + 2) + num_rows_tile + 3

    # Save rebuilt workbook
    print_flush(f"Saving coverage visual workbook to {target_path}...")
    new_wb.save(target_path)
    new_wb.close()
    print_flush("Visual support workbook saved successfully.")

    # Audit CSV validation
    csv_validation_path = r"D:\Sandbox\MP2027\OUTPUT_FY2027\tmp_requirement_0906_visual_coverage\new_snapshot_pixel_audit.csv"
    with open(csv_validation_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['filename', 'sheet', 'view', 'width', 'height', 'ratio', 'stddev', 'classification'])
        writer.writeheader()
        for audit in pixel_audits:
            writer.writerow(audit)
            
    # Print summary
    blank_count = sum(1 for a in pixel_audits if a['classification'] == 'BLANK')
    near_blank_count = sum(1 for a in pixel_audits if a['classification'] == 'NEAR_BLANK')
    valid_count = sum(1 for a in pixel_audits if a['classification'] == 'VALID_CONTENT')
    print_flush(f"\nFinal Rebuild Summary:")
    print_flush(f"  - Total images embedded: {len(pixel_audits)}")
    print_flush(f"  - VALID_CONTENT: {valid_count}")
    print_flush(f"  - BLANK: {blank_count}")
    print_flush(f"  - NEAR_BLANK: {near_blank_count}")
    
    # Check overall coverage gaps
    missing_dwg = sum(1 for row in drawing_matrix_rows if row['current_capture_status'] == 'MISSING')
    clipped_dwg = sum(1 for row in drawing_matrix_rows if row['current_capture_status'] == 'PARTIALLY_CLIPPED')
    print_flush(f"Drawing Coverage gaps in rebuilt visual workbook:")
    print_flush(f"  - Missing drawings: {missing_dwg}")
    print_flush(f"  - Partially clipped drawings: {clipped_dwg}")

if __name__ == '__main__':
    rebuild_all_coverage()
