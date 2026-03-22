import sqlite3
import openpyxl
import os
import re

db_path = 'data/mp2027.db'
conn = sqlite3.connect(db_path)
print("--- Dim Cost Centers (First 20) ---")
for row in conn.execute("SELECT code, name_jp FROM dim_cost_centers LIMIT 20"):
    print(row)

# Find GA file
fn = [f for f in os.listdir('.') if '総務課' in f][0]
print(f"\n--- GA File: {fn} ---")
wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
sn = [s for s in wb.sheetnames if '計算' in s][0]
ws = wb[sn]
print(f"Sheet: {sn}")

for i, row in enumerate(ws.iter_rows(max_row=60, max_col=15)):
    vals = [c.value for c in row]
    if any(vals):
        print(f"Row {i}: {vals}")

# Test extract_cc_code logic
def extract_cc_code_test(val):
    if val is None: return None
    s = str(val).strip()
    try:
        n = int(float(s))
        if 1000 <= n <= 9999: return n
    except: pass
    match = re.search(r'(1\d{3}|2\d{3}|3\d{3}|4\d{3}|5\d{3}|6\d{3}|7\d{3}|8\d{3}|9\d{3})', s)
    if match: return int(match.group(1))
    return None

print("\n--- CC Extract Test ---")
test_vals = ["1412_HANOI", "1412", " (1412) ", "CC1412", "1412 - Dept", "Hanoi1412"]
for v in test_vals:
    print(f"'{v}' -> {extract_cc_code_test(v)}")
