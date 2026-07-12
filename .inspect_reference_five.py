from pathlib import Path
import openpyxl, xlrd, json
base=Path('reference_outputs/secondary/FY2027')
for prefix in ('14.','15.','16.','72.','73.'):
 p=next(base.glob(prefix+'*.xlsx'))
 wb=openpyxl.load_workbook(p,data_only=False,read_only=True)
 print('\n###',p.name,'sheets',wb.sheetnames)
 for sn in wb.sheetnames[:4]:
  ws=wb[sn]; print('SHEET',sn,ws.max_row,ws.max_column)
  for r in range(1,min(ws.max_row,25)+1):
   vals=[ws.cell(r,c).value for c in range(1,min(ws.max_column,20)+1)]
   non=[(openpyxl.utils.get_column_letter(c),v) for c,v in enumerate(vals,1) if v not in (None,'')]
   if non: print(r,non[:12])
 wb.close()
