"""Set fiscal period metadata on one source-manifest entry."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def set_periods(path: Path, order: str, period_start: str, period_end: str) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in worksheet[1]
    }
    for name in ("order", "period_start", "period_end"):
        if name not in headers:
            column = worksheet.max_column + 1
            worksheet.cell(1, column, name)
            headers[name] = column
    matching = [
        row
        for row in range(2, worksheet.max_row + 1)
        if str(worksheet.cell(row, headers["order"]).value or "").strip() == order
    ]
    if len(matching) != 1:
        workbook.close()
        raise ValueError(f"Manifest order {order!r} must identify exactly one row; found {len(matching)}")
    row = matching[0]
    worksheet.cell(row, headers["period_start"], period_start)
    worksheet.cell(row, headers["period_end"], period_end)
    workbook.save(path)
    workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("manifest", type=Path)
    parser.add_argument("order")
    parser.add_argument("period_start")
    parser.add_argument("period_end")
    args = parser.parse_args()
    set_periods(args.manifest, args.order, args.period_start, args.period_end)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
