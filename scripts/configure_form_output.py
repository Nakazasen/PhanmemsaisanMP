"""Attach dynamic output metadata to a repository FORM workbook.

The layout is discovered from headers and formulas.  No worksheet coordinate is
accepted on the command line or stored in Python source.
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.engine.dynamic_source_order_export import resolve_form_layout


OUTPUT_AREA_NAME = "MP_OUTPUT_AREA"
ROW_TEMPLATE_NAME = "MP_OUTPUT_ROW_TEMPLATE"


def _replace_defined_name(workbook, name: str, reference: str) -> None:
    if name in workbook.defined_names:
        del workbook.defined_names[name]
    workbook.defined_names.add(DefinedName(name, attr_text=reference))


def configure_form(path: str | Path) -> dict[str, object]:
    form_path = Path(path).resolve()
    workbook = load_workbook(form_path, data_only=False)
    layout = resolve_form_layout(workbook)
    worksheet = workbook[layout.sheet_name]

    # The three externally supplied identity columns must be clean in the
    # repository template.  Generated workbooks are cleared again before use.
    for row in range(layout.output_start_row, layout.output_end_row + 1):
        for column in (layout.account_col, layout.description_col, layout.wbs_col):
            worksheet.cell(row, column).value = None

    first_column = get_column_letter(layout.account_col)
    last_column = get_column_letter(layout.wbs_col)
    sheet = quote_sheetname(layout.sheet_name)
    output_reference = (
        f"{sheet}!${first_column}${layout.output_start_row}:"
        f"${last_column}${layout.output_end_row}"
    )
    template_reference = (
        f"{sheet}!${first_column}${layout.template_row}:"
        f"${last_column}${layout.template_row}"
    )
    _replace_defined_name(workbook, OUTPUT_AREA_NAME, output_reference)
    _replace_defined_name(workbook, ROW_TEMPLATE_NAME, template_reference)

    temporary = form_path.with_name(f".{form_path.name}.tmp")
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, form_path)
    return {
        "sheet": layout.sheet_name,
        "output_area": output_reference,
        "row_template": template_reference,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("form", type=Path)
    args = parser.parse_args()
    print(configure_form(args.form))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
