"""Đối chiếu chỉ đọc tài sản cố định FY2026/FY2027.

Kiểm tra so sánh sổ nguồn tài sản cố định của công ty với kết quả tham chiếu
theo phòng ban. Nó chủ động tính cả cách làm tròn chuẩn theo từng tài sản và
cách làm tròn theo nhóm của bộ ghi hiện tại để thấy rõ chênh lệch.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import sys
from typing import Any, Iterable

import openpyxl
import xlrd


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.utils.cli import VietnameseArgumentParser

AUDIT_DATE = "2026-07-16"
INTEREST_ACCOUNT = 9114120007
CATEGORY_SPECS = {
    "machinery_equipment": {
        "aliases": {"mfg)machinery and equipment", "machinery and equipment"},
        "account": 5006016242,
    },
    "vehicles": {"aliases": {"mfg)vehicles", "vehicles"}, "account": 5006016243},
    "tools_furniture_fixtures": {
        "aliases": {"mfg)tools furniture and fixtures", "tools furniture and fixtures"},
        "account": 5006016244,
    },
    "other_tangible_fixed_assets": {
        "aliases": {"mfg)other tangible fixed assets", "other tangible fixed assets"},
        "account": 5006016247,
    },
    "mold": {"aliases": {"mfg)mold", "mold"}, "account": 5005036246},
}
DEPRECIATION_ACCOUNTS = {int(spec["account"]) for spec in CATEGORY_SPECS.values()}
FIXED_ASSET_ACCOUNTS = DEPRECIATION_ACCOUNTS | {INTEREST_ACCOUNT}
OUT_OF_SCOPE_MARKERS = ("sga)", "software", "buildings", "structures", "land use rights")
CC_RE = re.compile(r"1412\d{6}")


@dataclass(frozen=True)
class FYConfig:
    name: str
    fiscal_year: int
    source_dir: Path
    reference_dir: Path
    raw_dir: Path

    @property
    def periods(self) -> tuple[str, ...]:
        return tuple(
            [f"{self.fiscal_year - 1}{month:02d}" for month in range(4, 13)]
            + [f"{self.fiscal_year}{month:02d}" for month in range(1, 4)]
        )


CONFIGS = (
    FYConfig(
        "FY2026",
        2026,
        ROOT / "docs" / "MP2026",
        ROOT / "reference_outputs" / "secondary" / "FY2026",
        ROOT / "raw" / "FY2026",
    ),
    FYConfig(
        "FY2027",
        2027,
        ROOT / "docs" / "MP2027",
        ROOT / "reference_outputs" / "secondary" / "FY2027",
        ROOT / "raw" / "FY2027",
    ),
)


def norm_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u3000", " ").strip().lower().split())


def normalize_cc(value: Any) -> str:
    text = str(value or "").strip()
    match = CC_RE.search(text.replace(".0", ""))
    return match.group(0) if match else ""


def normalize_period(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}{value.month:02d}"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        text = str(int(value))
        if len(text) == 6:
            return text
    text = str(value or "").strip()
    numbers = re.findall(r"\d+", text)
    if len(numbers) >= 2:
        year, month = int(numbers[0]), int(numbers[1])
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return f"{year:04d}{month:02d}"
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 6:
        return digits[:6]
    return ""


def number(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def excel_round(value: float, digits: int = 0) -> int | float:
    quantizer = Decimal("1").scaleb(-digits)
    rounded = Decimal(str(value)).quantize(quantizer, rounding=ROUND_HALF_UP)
    return int(rounded) if digits == 0 else float(rounded)


def category_key(value: Any) -> tuple[str, str]:
    text = norm_text(value)
    for key, spec in CATEGORY_SPECS.items():
        if text in spec["aliases"]:
            return key, "supported"
    if text and any(marker in text for marker in OUT_OF_SCOPE_MARKERS):
        return "", "out_of_scope"
    return "", "unknown"


def terminal_relation(period: str, periods: tuple[str, ...]) -> str:
    if not period:
        return "missing"
    if period < periods[0]:
        return "before_fy"
    if period > periods[-1]:
        return "after_fy"
    return "within_fy"


def schedule(
    *,
    monthly: float | None,
    terminal: str,
    terminal_amount: float | None,
    first_interest: float | None,
    later_interest: float | None,
    periods: tuple[str, ...],
    full_fy_when_terminal_before: bool,
) -> tuple[dict[str, float], dict[str, float]]:
    dep: dict[str, float] = {}
    interest: dict[str, float] = {}
    relation = terminal_relation(terminal, periods)
    for idx, period in enumerate(periods):
        if relation == "before_fy" and not full_fy_when_terminal_before:
            continue
        if terminal and terminal >= periods[0] and period > terminal:
            continue
        dep_amount = monthly
        if terminal and period == terminal:
            dep_amount = terminal_amount
        if dep_amount is not None:
            dep[period] = dep_amount
        interest_amount = first_interest if idx == 0 else later_interest
        if interest_amount is not None:
            interest[period] = interest_amount
    return dep, interest


def find_source_workbook(config: FYConfig) -> Path:
    matches = sorted(config.source_dir.glob("*Fixed_Assets_Information*.xlsx"))
    if len(matches) != 1:
        raise RuntimeError(f"{config.name}: expected one fixed-assets source, found {len(matches)}")
    return matches[0]


def load_asset_ledger(config: FYConfig) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    source = find_source_workbook(config)
    formulas = openpyxl.load_workbook(source, read_only=True, data_only=False)
    values = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        source_sheet = formulas[formulas.sheetnames[0]]
        value_sheet = values[values.sheetnames[0]]
        ledger: list[dict[str, Any]] = []
        counts: Counter[str] = Counter()
        for excel_row, (formula_row, value_row) in enumerate(
            zip(
                source_sheet.iter_rows(min_row=5, values_only=True),
                value_sheet.iter_rows(min_row=5, values_only=True),
            ),
            start=5,
        ):
            if not any(value not in (None, "") for value in value_row):
                continue
            category, status = category_key(value_row[1] if len(value_row) > 1 else None)
            cc = normalize_cc(value_row[9] if len(value_row) > 9 else None)
            monthly = number(value_row[11] if len(value_row) > 11 else None)
            terminal = normalize_period(value_row[15] if len(value_row) > 15 else None)
            terminal_amount = number(value_row[16] if len(value_row) > 16 else None)
            first_interest = number(value_row[21] if len(value_row) > 21 else None)
            later_interest = number(value_row[22] if len(value_row) > 22 else None)
            critical_formula_cells = {
                "L": formula_row[11] if len(formula_row) > 11 else None,
                "P": formula_row[15] if len(formula_row) > 15 else None,
                "Q": formula_row[16] if len(formula_row) > 16 else None,
                "V": formula_row[21] if len(formula_row) > 21 else None,
                "W": formula_row[22] if len(formula_row) > 22 else None,
            }
            critical_value_cells = {
                "L": value_row[11] if len(value_row) > 11 else None,
                "P": value_row[15] if len(value_row) > 15 else None,
                "Q": value_row[16] if len(value_row) > 16 else None,
                "V": value_row[21] if len(value_row) > 21 else None,
                "W": value_row[22] if len(value_row) > 22 else None,
            }
            cache_gaps = [
                col
                for col, formula in critical_formula_cells.items()
                if isinstance(formula, str)
                and formula.startswith("=")
                and critical_value_cells[col] in (None, "")
            ]
            relation = terminal_relation(terminal, config.periods)
            dep, interest = schedule(
                monthly=monthly,
                terminal=terminal,
                terminal_amount=terminal_amount,
                first_interest=first_interest,
                later_interest=later_interest,
                periods=config.periods,
                full_fy_when_terminal_before=False,
            )
            dep_alt, interest_alt = schedule(
                monthly=monthly,
                terminal=terminal,
                terminal_amount=terminal_amount,
                first_interest=first_interest,
                later_interest=later_interest,
                periods=config.periods,
                full_fy_when_terminal_before=True,
            )
            if status != "supported":
                counts[status] += 1
            elif not cc:
                counts["missing_cc"] += 1
            else:
                counts["supported_rows"] += 1
            if cache_gaps:
                counts["formula_cache_gap_rows"] += 1
            if any((amount or 0) < 0 for amount in (monthly, terminal_amount, first_interest, later_interest)):
                counts["negative_input_rows"] += 1
            if relation == "within_fy" and terminal_amount is None and (monthly or 0) != 0:
                counts["terminal_q_missing_rows"] += 1
            counts[f"terminal_{relation}"] += 1
            ledger.append(
                {
                    "fy": config.name,
                    "source_file": str(source.relative_to(ROOT)),
                    "source_sheet": source_sheet.title,
                    "source_row": excel_row,
                    "category_raw": value_row[1] if len(value_row) > 1 else None,
                    "category_key": category,
                    "category_status": status,
                    "asset_no": value_row[2] if len(value_row) > 2 else None,
                    "asset_text": value_row[3] if len(value_row) > 3 else None,
                    "control_cc": normalize_cc(value_row[7] if len(value_row) > 7 else None),
                    "depreciation_cc": cc,
                    "monthly_depr_usd": monthly,
                    "terminal_period": terminal,
                    "terminal_relation": relation,
                    "terminal_depr_usd": terminal_amount,
                    "apr_interest_usd": first_interest,
                    "may_interest_usd": later_interest,
                    "critical_formula_cache_gaps": ";".join(cache_gaps),
                    "depreciation_schedule": dep,
                    "interest_schedule": interest,
                    "alt_full_fy_depreciation_schedule": dep_alt,
                    "alt_full_fy_interest_schedule": interest_alt,
                }
            )
        return ledger, {
            "source_file": str(source.relative_to(ROOT)),
            "source_sheet": source_sheet.title,
            "rows": len(ledger),
            **dict(counts),
        }
    finally:
        formulas.close()
        values.close()


def formula_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(getattr(value, "text", value))


def is_facility_interest(description: str) -> bool:
    text = norm_text(description)
    return any(token in text for token in ("建物", "土地", "building", "land"))


def load_reference_rows(config: FYConfig) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    files = sorted(config.reference_dir.glob("*.xlsx"))
    rates: Counter[int] = Counter()
    cc_files: Counter[str] = Counter()
    failures: list[str] = []
    missing_cc_files: list[str] = []
    missing_rate_files: list[str] = []
    for index, path in enumerate(files, start=1):
        print(f"{config.name} reference {index}/{len(files)}: {path.name}", flush=True)
        try:
            formulas = openpyxl.load_workbook(path, read_only=True, data_only=False)
            values = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - evidence path
            failures.append(f"{path.name}: {type(exc).__name__}: {exc}")
            continue
        try:
            sheet_name = next((name for name in formulas.sheetnames if "内訳" in name), formulas.sheetnames[2])
            formula_sheet = formulas[sheet_name]
            value_sheet = values[sheet_name]
            formula_iter = formula_sheet.iter_rows(values_only=True)
            value_iter = value_sheet.iter_rows(values_only=True)
            pairs = list(zip(formula_iter, value_iter))
            cc = normalize_cc(pairs[4][1][1] if len(pairs) > 4 and len(pairs[4][1]) > 1 else None)
            rate_num = number(pairs[1][1][1] if len(pairs) > 1 and len(pairs[1][1]) > 1 else None)
            rate = int(rate_num) if rate_num is not None else 0
            if cc:
                cc_files[cc] += 1
            else:
                missing_cc_files.append(path.name)
            if rate:
                rates[rate] += 1
            else:
                missing_rate_files.append(path.name)
            for excel_row, (formula_row, value_row) in enumerate(pairs, start=1):
                account_num = number(value_row[1] if len(value_row) > 1 else None)
                account = int(account_num) if account_num is not None else 0
                if account not in FIXED_ASSET_ACCOUNTS:
                    continue
                description = str(value_row[18] if len(value_row) > 18 and value_row[18] not in (None, "") else "")
                kind = "depreciation" if account in DEPRECIATION_ACCOUNTS else "interest"
                scope = "facility" if kind == "interest" and is_facility_interest(description) else "fixed_assets"
                monthly_values: dict[str, float | None] = {}
                monthly_formulas: dict[str, str] = {}
                for offset, period in enumerate(config.periods, start=5):
                    raw_value = value_row[offset] if len(value_row) > offset else None
                    monthly_values[period] = number(raw_value)
                    raw_formula = formula_row[offset] if len(formula_row) > offset else None
                    monthly_formulas[period] = formula_text(raw_formula)
                rows.append(
                    {
                        "fy": config.name,
                        "reference_file": str(path.relative_to(ROOT)),
                        "sheet": sheet_name,
                        "row": excel_row,
                        "cc": cc,
                        "fx_rate": rate,
                        "account": account,
                        "kind": kind,
                        "scope": scope,
                        "description": description,
                        "monthly_values": monthly_values,
                        "monthly_formulas": monthly_formulas,
                        "source_candidate_status": "UNCLASSIFIED",
                        "selection_l1_delta_vnd": None,
                    }
                )
        finally:
            formulas.close()
            values.close()
    return rows, {
        "top_level_workbooks": len(files),
        "rates": dict(sorted(rates.items())),
        "cc_count": len(cc_files),
        "duplicate_cc_files": {cc: count for cc, count in cc_files.items() if count > 1},
        "missing_cc_files": missing_cc_files,
        "missing_rate_files": missing_rate_files,
        "failures": failures,
    }


def scan_raw_truth(config: FYConfig) -> dict[str, Any]:
    workbook_paths = sorted(
        path for path in config.raw_dir.iterdir() if path.is_file() and path.suffix.lower() in {".xls", ".xlsx"}
    )
    keyword_hits = 0
    account_hits = 0
    failed: list[str] = []
    keywords = ("固定資産", "減価償却", "depreciation", "fixed asset", "khấu hao")
    for path in workbook_paths:
        try:
            if path.suffix.lower() == ".xls":
                book = xlrd.open_workbook(path, on_demand=True)
                try:
                    for sheet in book.sheets():
                        for row_idx in range(sheet.nrows):
                            for col_idx in range(sheet.ncols):
                                value = sheet.cell_value(row_idx, col_idx)
                                text = norm_text(value)
                                if any(keyword in text for keyword in keywords):
                                    keyword_hits += 1
                                value_num = number(value)
                                if value_num is not None and int(value_num) in FIXED_ASSET_ACCOUNTS:
                                    account_hits += 1
                finally:
                    book.release_resources()
            else:
                book = openpyxl.load_workbook(path, read_only=True, data_only=True)
                try:
                    for sheet in book.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            for value in row:
                                text = norm_text(value)
                                if any(keyword in text for keyword in keywords):
                                    keyword_hits += 1
                                value_num = number(value)
                                if value_num is not None and int(value_num) in FIXED_ASSET_ACCOUNTS:
                                    account_hits += 1
                finally:
                    book.close()
        except Exception as exc:  # pragma: no cover - evidence path
            failed.append(f"{path.name}: {type(exc).__name__}: {exc}")
    return {
        "workbooks": len(workbook_paths),
        "fixed_asset_keyword_cells": keyword_hits,
        "fixed_asset_account_cells": account_hits,
        "failures": failed,
    }


def build_expected(
    config: FYConfig,
    ledger: list[dict[str, Any]],
    reference_rows: list[dict[str, Any]],
) -> tuple[dict[tuple[str, int, str], int], dict[tuple[str, int, str], int], dict[tuple[str, int, str], int]]:
    observed_rates = [int(row["fx_rate"]) for row in reference_rows if int(row["fx_rate"] or 0) > 0]
    modal_rate = Counter(observed_rates).most_common(1)[0][0] if observed_rates else 0
    rates_by_cc = {
        str(row["cc"]): int(row["fx_rate"])
        for row in reference_rows
        if row["cc"] and int(row["fx_rate"] or 0) > 0
    }
    per_asset: defaultdict[tuple[str, int, str], int] = defaultdict(int)
    alt_before: defaultdict[tuple[str, int, str], int] = defaultdict(int)
    writer_usd: defaultdict[tuple[str, str, str, str], float] = defaultdict(float)
    for asset in ledger:
        if asset["category_status"] != "supported" or not asset["depreciation_cc"]:
            continue
        cc = str(asset["depreciation_cc"])
        rate = rates_by_cc.get(cc, modal_rate)
        if not rate:
            continue
        category = str(asset["category_key"])
        dep_account = int(CATEGORY_SPECS[category]["account"])
        for period, amount in asset["depreciation_schedule"].items():
            per_asset[(cc, dep_account, period)] += int(excel_round(float(amount) * rate))
            writer_usd[(cc, "depreciation", category, period)] += float(amount)
        for period, amount in asset["interest_schedule"].items():
            per_asset[(cc, INTEREST_ACCOUNT, period)] += int(excel_round(float(amount) * rate))
            writer_usd[(cc, "interest", category, period)] += float(amount)
        for period, amount in asset["alt_full_fy_depreciation_schedule"].items():
            alt_before[(cc, dep_account, period)] += int(excel_round(float(amount) * rate))
        for period, amount in asset["alt_full_fy_interest_schedule"].items():
            alt_before[(cc, INTEREST_ACCOUNT, period)] += int(excel_round(float(amount) * rate))
    writer: defaultdict[tuple[str, int, str], int] = defaultdict(int)
    for (cc, kind, category, period), amount in writer_usd.items():
        rate = rates_by_cc.get(cc, modal_rate)
        if not rate:
            continue
        account = int(CATEGORY_SPECS[category]["account"] if kind == "depreciation" else INTEREST_ACCOUNT)
        writer[(cc, account, period)] += int(excel_round(amount * rate))
    return dict(per_asset), dict(writer), dict(alt_before)


def _row_vector(row: dict[str, Any], periods: tuple[str, ...]) -> tuple[int, ...]:
    return tuple(
        0 if row["monthly_values"].get(period) is None else int(excel_round(row["monthly_values"][period]))
        for period in periods
    )


def _best_subset(candidates: list[dict[str, Any]], target: tuple[int, ...], periods: tuple[str, ...]) -> tuple[set[int], int]:
    """Select the reference-row subset closest to the source-derived target."""
    vectors = [_row_vector(row, periods) for row in candidates]
    best_indices: set[int] = set()
    best_score = sum(abs(value) for value in target)
    if len(candidates) <= 18:
        masks: Iterable[int] = range(1, 1 << len(candidates))
        for mask in masks:
            combined = [0] * len(periods)
            selected_count = 0
            for index, vector in enumerate(vectors):
                if mask & (1 << index):
                    selected_count += 1
                    for month, value in enumerate(vector):
                        combined[month] += value
            score = sum(abs(combined[month] - target[month]) for month in range(len(periods)))
            if score < best_score or (score == best_score and selected_count < len(best_indices)):
                best_score = score
                best_indices = {index for index in range(len(candidates)) if mask & (1 << index)}
    else:
        # Conservative greedy fallback for an unexpectedly large manual-layer group.
        combined = [0] * len(periods)
        remaining = set(range(len(candidates)))
        while remaining:
            current_score = sum(abs(combined[month] - target[month]) for month in range(len(periods)))
            choices = []
            for index in remaining:
                trial = [combined[month] + vectors[index][month] for month in range(len(periods))]
                score = sum(abs(trial[month] - target[month]) for month in range(len(periods)))
                choices.append((score, index, trial))
            score, index, trial = min(choices)
            if score >= current_score:
                break
            best_indices.add(index)
            remaining.remove(index)
            combined = trial
            best_score = score
    return best_indices, best_score


def classify_reference_rows(
    config: FYConfig,
    reference_rows: list[dict[str, Any]],
    expected: dict[tuple[str, int, str], int],
) -> None:
    grouped: defaultdict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in reference_rows:
        if row["scope"] == "facility":
            row["source_candidate_status"] = "FACILITY_EXCLUDED"
            continue
        if not row["cc"] or not any(value is not None for value in row["monthly_values"].values()):
            row["source_candidate_status"] = "NO_MONTHLY_VALUE"
            continue
        grouped[(str(row["cc"]), int(row["account"]))].append(row)
    for (cc, account), candidates in grouped.items():
        target = tuple(int(expected.get((cc, account, period), 0)) for period in config.periods)
        if not any(target):
            for row in candidates:
                row["source_candidate_status"] = "NO_SOURCE_EXPECTATION"
            continue
        selected, score = _best_subset(candidates, target, config.periods)
        for index, row in enumerate(candidates):
            row["selection_l1_delta_vnd"] = score
            row["source_candidate_status"] = (
                "SELECTED_SOURCE_DERIVED_CANDIDATE"
                if index in selected
                else "EXCLUDED_MANUAL_OR_OTHER_LAYER"
            )


def build_actual(reference_rows: list[dict[str, Any]]) -> dict[tuple[str, int, str], int]:
    actual: defaultdict[tuple[str, int, str], int] = defaultdict(int)
    for row in reference_rows:
        if row["source_candidate_status"] != "SELECTED_SOURCE_DERIVED_CANDIDATE" or not row["cc"]:
            continue
        for period, amount in row["monthly_values"].items():
            if amount is not None:
                actual[(str(row["cc"]), int(row["account"]), period)] += int(excel_round(amount))
    return dict(actual)


def compare_months(
    config: FYConfig,
    expected: dict[tuple[str, int, str], int],
    writer: dict[tuple[str, int, str], int],
    alternative: dict[tuple[str, int, str], int],
    actual: dict[tuple[str, int, str], int],
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    keys = sorted(set(expected) | set(writer) | set(alternative) | set(actual))
    for cc, account, period in keys:
        if period not in config.periods:
            continue
        exp = expected.get((cc, account, period))
        act = actual.get((cc, account, period))
        wrt = writer.get((cc, account, period))
        alt = alternative.get((cc, account, period))
        if exp == act:
            classification = "EXACT_MATCH"
        elif act is None:
            classification = "MISSING_REFERENCE_OUTPUT"
        elif exp is None:
            classification = "EXTRA_REFERENCE_OUTPUT"
        elif act == wrt and wrt != exp:
            classification = "ROUNDING_ORDER"
        elif act == alt and alt != exp:
            classification = "TERMINAL_BEFORE_FY_FULL_YEAR"
        else:
            classification = "TRUE_AMOUNT_MISMATCH"
        result.append(
            {
                "fy": config.name,
                "cc": cc,
                "account": account,
                "period": period,
                "expected_per_asset_round_vnd": exp,
                "current_writer_category_round_vnd": wrt,
                "alternative_terminal_before_full_fy_vnd": alt,
                "reference_actual_vnd": act,
                "delta_reference_minus_expected": None if exp is None or act is None else act - exp,
                "classification": classification,
            }
        )
    return result


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            serial = {
                key: json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, (dict, list, tuple)) else value
                for key, value in row.items()
            }
            writer.writerow(serial)


def report_markdown(
    summary: dict[str, Any],
    comparisons: list[dict[str, Any]],
    reference_rows: list[dict[str, Any]],
    ledger: list[dict[str, Any]],
) -> str:
    classes = Counter(str(row["classification"]) for row in comparisons)
    classes_by_fy = {
        fy: Counter(str(row["classification"]) for row in comparisons if row["fy"] == fy)
        for fy in ("FY2026", "FY2027")
    }
    mismatches = [row for row in comparisons if row["classification"] != "EXACT_MATCH"]
    true_mismatches = [row for row in comparisons if row["classification"] == "TRUE_AMOUNT_MISMATCH"]
    selected_rows = Counter(
        row["fy"]
        for row in reference_rows
        if row["source_candidate_status"] == "SELECTED_SOURCE_DERIVED_CANDIDATE"
    )
    excluded_rows = Counter(
        row["fy"]
        for row in reference_rows
        if row["source_candidate_status"] == "EXCLUDED_MANUAL_OR_OTHER_LAYER"
    )
    critical_cache_gaps = Counter(row["fy"] for row in ledger if row["critical_formula_cache_gaps"])
    negative_rows = Counter(
        row["fy"]
        for row in ledger
        if any((number(row[field]) or 0) < 0 for field in ("monthly_depr_usd", "terminal_depr_usd", "apr_interest_usd", "may_interest_usd"))
    )
    missing_q_rows = Counter(
        row["fy"]
        for row in ledger
        if row["terminal_relation"] == "within_fy"
        and row["terminal_depr_usd"] is None
        and (row["monthly_depr_usd"] or 0) != 0
    )
    rounding_stats: dict[str, tuple[int, int, int]] = {}
    true_stats: dict[str, tuple[int, int, int]] = {}
    for fy in ("FY2026", "FY2027"):
        rounding_deltas = [
            int(row["delta_reference_minus_expected"])
            for row in comparisons
            if row["fy"] == fy
            and row["classification"] == "ROUNDING_ORDER"
            and row["delta_reference_minus_expected"] is not None
        ]
        rounding_stats[fy] = (
            len(rounding_deltas),
            sum(rounding_deltas),
            sum(abs(value) for value in rounding_deltas),
        )
        true_deltas = [
            int(row["delta_reference_minus_expected"])
            for row in comparisons
            if row["fy"] == fy
            and row["classification"] == "TRUE_AMOUNT_MISMATCH"
            and row["delta_reference_minus_expected"] is not None
        ]
        true_stats[fy] = (
            len(true_deltas),
            sum(1 for value in true_deltas if abs(value) > 1_000_000),
            max((abs(value) for value in true_deltas), default=0),
        )
    lines = [
        "# Kiểm toán đối chiếu chéo tài sản cố định — FY2026/FY2027",
        "",
        f"**Ngày kiểm toán:** `{AUDIT_DATE}`  ",
        "**Chế độ:** chỉ đọc dữ liệu nguồn/tham khảo; không thay đổi code kế toán vận hành  ",
        "**Phân loại:** `NOT_ACCEPTED_FIXED_ASSETS_CALCULATION`; vòng đời vẫn là `OPEN_AUDIT`",
        "",
        "## Phạm vi và căn cứ",
        "",
        "1. Yêu cầu chuẩn: `raw/Cải tiến nhập dữ liệu chung vào file MPnew 10.07.2026.xlsx`, sheet `Chi phí tài sản cố định`.",
        "2. Nguồn tính toán của công ty: hai tệp `固定資産情報_Fixed_Assets_Information_*.xlsx` trong `docs/MP2026` và `docs/MP2027`.",
        "3. Thư mục dữ liệu đúng do phòng ban nộp: `raw/FY2026`, `raw/FY2027`.",
        "4. Tệp kết quả tham khảo đã nộp: các tệp `.xlsx` cấp cao nhất trong `reference_outputs/secondary/FY2026` và `FY2027`.",
        "5. Code hiện tại chỉ được dùng làm bằng chứng triển khai, không phải căn cứ nghiệp vụ.",
        "",
        "## Tóm tắt tập dữ liệu",
        "",
        "| FY | Dòng nguồn | Dòng nguồn được hỗ trợ | Tệp tham khảo | Số CC tham khảo | Tỷ giá | Tệp dữ liệu đúng | Ô tài khoản tài sản cố định |",
        "|---|---:|---:|---:|---:|---|---:|---:|",
    ]
    for fy in ("FY2026", "FY2027"):
        item = summary[fy]
        lines.append(
            f"| {fy} | {item['source']['rows']} | {item['source'].get('supported_rows', 0)} | "
            f"{item['reference']['top_level_workbooks']} | {item['reference']['cc_count']} | "
            f"`{json.dumps(item['reference']['rates'], ensure_ascii=False)}` | "
            f"{item['raw']['workbooks']} | {item['raw']['fixed_asset_account_cells']} |"
        )
    lines.extend(
        [
            "",
            "`raw/FY2026` và `raw/FY2027` là dữ liệu kế hoạch nhân sự/thời gian trong tập dữ liệu hiện tại. FY2026 được quét sạch và không có ô mã tài khoản tài sản cố định. FY2027 có cùng kết quả ở các tệp đọc được, nhưng một tệp `.xls` cũ không đọc được bằng `xlrd`, nên kết luận chỉ áp dụng cho 63/64 tệp FY2027. Số tiền đúng của tài sản cố định nằm trong các tệp tính toán của công ty và tệp kết quả tham khảo cuối đã nộp, không nằm trong các tệp nhân sự thô đọc được.",
            "",
            "## Tách biệt các lớp tham khảo",
            "",
            "Tệp tham khảo thường chứa dòng suy ra từ nguồn cùng với dòng chuyển tiếp nhập tay, dòng lũy kế, cơ sở vật chất và tài sản tương lai có chung tài khoản. Với mỗi FY/CC/tài khoản, bộ đối chiếu chọn nhóm dòng có tổng chênh lệch tuyệt đối 12 tháng nhỏ nhất so với mục tiêu suy ra từ nguồn. Các dòng bị loại vẫn nằm trong sổ dòng tham khảo và không bao giờ bị bỏ qua âm thầm.",
            "",
            "| FY | Ứng viên suy ra từ nguồn được chọn | Lớp nhập tay/khác bị loại | Thiếu cache L/P/Q/V/W trọng yếu | Đầu vào trọng yếu âm | Kết thúc trong FY nhưng thiếu Q |",
            "|---|---:|---:|---:|---:|---:|",
            f"| FY2026 | {selected_rows['FY2026']} | {excluded_rows['FY2026']} | {critical_cache_gaps['FY2026']} | {negative_rows['FY2026']} | {missing_q_rows['FY2026']} |",
            f"| FY2027 | {selected_rows['FY2027']} | {excluded_rows['FY2027']} | {critical_cache_gaps['FY2027']} | {negative_rows['FY2027']} | {missing_q_rows['FY2027']} |",
            "",
            "## Kết quả đối chiếu theo tháng",
            "",
            "| Phân loại | FY2026 | FY2027 | Tổng |",
            "|---|---:|---:|---:|",
        ]
    )
    for name, count in sorted(classes.items()):
        lines.append(
            f"| `{name}` | {classes_by_fy['FY2026'].get(name, 0)} | "
            f"{classes_by_fy['FY2027'].get(name, 0)} | {count} |"
        )
    lines.extend(
        [
            "",
            f"Số ô tháng CC/tài khoản đã đối chiếu: **{len(comparisons)}**. Số ô không khớp hoàn toàn: **{len(mismatches)}**. Số ô chênh lệch số tiền thực sau khi tách trường hợp làm tròn/chính sách kết thúc: **{len(true_mismatches)}**.",
            "",
            "## Quyết định",
            "",
            "Phép tính tài sản cố định **chưa được chấp nhận là chính xác**. Có các ô khớp hoàn toàn, nhưng cách triển khai hiện tại vi phạm quy tắc làm tròn theo từng tài sản và hàng trăm ô tháng giữa nguồn/tham khảo vẫn chênh lệch đáng kể sau khi tách lớp nhập tay.",
            "",
            "| FY | Ô sai thứ tự làm tròn | Chênh lệch VND ròng so với làm tròn theo tài sản | Chênh lệch VND tuyệt đối | Ô chênh lệch thực | Chênh lệch thực > 1 triệu VND | Chênh lệch thực lớn nhất |",
            "|---|---:|---:|---:|---:|---:|---:|",
            f"| FY2026 | {rounding_stats['FY2026'][0]} | {rounding_stats['FY2026'][1]} | {rounding_stats['FY2026'][2]} | {true_stats['FY2026'][0]} | {true_stats['FY2026'][1]} | {true_stats['FY2026'][2]} |",
            f"| FY2027 | {rounding_stats['FY2027'][0]} | {rounding_stats['FY2027'][1]} | {rounding_stats['FY2027'][2]} | {true_stats['FY2027'][0]} | {true_stats['FY2027'][1]} | {true_stats['FY2027'][2]} |",
            "",
            "## Quy tắc và phát hiện đã được chứng minh",
            "",
            "- Tệp kết quả tham khảo dùng tỷ giá của tệp Excel tại `B2`; các tỷ giá quan sát theo FY là bằng chứng, không phải hằng số vận hành.",
            "- Các ô `ROUNDING_ORDER` là trường hợp kết quả đã nộp khớp với cách làm tròn theo nhóm của bộ ghi hiện tại nhưng khác cách làm tròn theo từng tài sản bắt buộc. Chênh lệch tiền nhỏ nhưng thứ tự tính vẫn sai.",
            "- Trường hợp kết thúc trong FY có ví dụ lỗi trực tiếp: nguồn `docs/MP2026/固定資産情報_Fixed_Assets_Information_2024.12 - December.xlsx`, `2024.12!L42/P42/Q42`, kết thúc tại `202601`; tệp tham khảo `24.KDTVN 品質保証課_MP FY2026_各予定(Ver01).xlsx`, dòng chi tiết 123, vẫn tiếp tục cùng mức khấu hao tháng đến `202602` và `202603`.",
            "- Trường hợp kết thúc trước FY được xác định ở cấp số tiền là không có chi phí trong FY: nguồn FY2027 `2025.11!L1257/P1257/Q1257` kết thúc tại `202512`; tệp tham khảo CC `1412000081`, tài khoản `5006016247`, dòng 45 bằng 0 cho mọi tháng FY2027. Theo yêu cầu chuẩn, kết quả mới phải biểu diễn giai đoạn sau kết thúc bằng ô trống, không phải số 0.",
            "- `TRUE_AMOUNT_MISMATCH` vẫn chưa được giải thích sau cả hai phép tính theo chính sách và cần rà soát cấp dòng/công thức trong các tệp CSV bằng chứng.",
            "- Có thể xác định một số chênh lệch thực lớn là khác biệt giữa ảnh chụp nguồn và tài sản tương lai nhập tay, không chỉ là số học; không được chấp nhận hoặc ghi đè nếu chưa có nguồn gốc cấp dòng.",
            "",
            "## Đánh giá code hiện tại",
            "",
            "- `src/parsers/fixed_assets.py` vẫn có giá trị FY/tỷ giá dự phòng, ánh xạ nhóm/tài khoản gắn cứng, lọc chỉ lấy số dương, dự phòng Q sang L và xóa toàn bộ lịch sử tài sản cố định khi nhập.",
            "- `src/engine/hub_builder.py::_load_fixed_asset_source_order_rows()` cộng USD tài sản theo nhóm rồi xuất một công thức `ROUND(sum*$B$2,0)`. Cách này không tuân thủ quy tắc làm tròn theo từng tài sản; tệp bằng chứng theo tháng đã định lượng các dòng bị ảnh hưởng.",
            "- Bộ ghi theo thứ tự nguồn đặt lại đúng vị trí các dòng tài sản cố định động khi được cung cấp, nhưng không thể sửa việc mất dữ liệu hoặc nguồn gốc ở thượng nguồn.",
            "",
            "## Các tệp bằng chứng",
            "",
            f"- `docs/audits/fixed_assets_asset_ledger_{AUDIT_DATE}.csv`",
            f"- `docs/audits/fixed_assets_reference_rows_{AUDIT_DATE}.csv`",
            f"- `docs/audits/fixed_assets_monthly_comparison_{AUDIT_DATE}.csv`",
            f"- `docs/audits/fixed_assets_cross_trace_summary_{AUDIT_DATE}.json`",
            "",
            "## Trạng thái",
            "",
            "Giữ `FA-OPEN` ở trạng thái `OPEN_AUDIT`. Việc sửa thứ tự làm tròn đã có bằng chứng; xử lý các dòng ảnh chụp nguồn/nhập tay còn lại cần ma trận quyết định và phân loại nguồn gốc trước khi thay đổi logic kế toán.",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = VietnameseArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=ROOT / "docs" / "audits")
    args = parser.parse_args()
    output_dir = args.output_dir.resolve()

    all_ledger: list[dict[str, Any]] = []
    all_reference_rows: list[dict[str, Any]] = []
    all_comparisons: list[dict[str, Any]] = []
    summary: dict[str, Any] = {}
    for config in CONFIGS:
        print(f"Đang nạp sổ nguồn {config.name}", flush=True)
        ledger, source_summary = load_asset_ledger(config)
        print(f"Đang quét tệp kết quả tham khảo {config.name}", flush=True)
        reference_rows, reference_summary = load_reference_rows(config)
        print(f"Đang quét thư mục dữ liệu đúng {config.name}", flush=True)
        raw_summary = scan_raw_truth(config)
        expected, current_writer, alternative = build_expected(config, ledger, reference_rows)
        classify_reference_rows(config, reference_rows, expected)
        actual = build_actual(reference_rows)
        comparisons = compare_months(config, expected, current_writer, alternative, actual)
        all_ledger.extend(ledger)
        all_reference_rows.extend(reference_rows)
        all_comparisons.extend(comparisons)
        summary[config.name] = {
            "source": source_summary,
            "reference": reference_summary,
            "raw": raw_summary,
            "comparison_classifications": dict(Counter(row["classification"] for row in comparisons)),
        }

    ledger_fields = [
        "fy", "source_file", "source_sheet", "source_row", "category_raw", "category_key",
        "category_status", "asset_no", "asset_text", "control_cc", "depreciation_cc",
        "monthly_depr_usd", "terminal_period", "terminal_relation", "terminal_depr_usd",
        "apr_interest_usd", "may_interest_usd", "critical_formula_cache_gaps",
        "depreciation_schedule", "interest_schedule", "alt_full_fy_depreciation_schedule",
        "alt_full_fy_interest_schedule",
    ]
    reference_fields = [
        "fy", "reference_file", "sheet", "row", "cc", "fx_rate", "account", "kind",
        "scope", "description", "monthly_values", "monthly_formulas",
        "source_candidate_status", "selection_l1_delta_vnd",
    ]
    comparison_fields = [
        "fy", "cc", "account", "period", "expected_per_asset_round_vnd",
        "current_writer_category_round_vnd", "alternative_terminal_before_full_fy_vnd",
        "reference_actual_vnd", "delta_reference_minus_expected", "classification",
    ]
    write_csv(output_dir / f"fixed_assets_asset_ledger_{AUDIT_DATE}.csv", all_ledger, ledger_fields)
    write_csv(output_dir / f"fixed_assets_reference_rows_{AUDIT_DATE}.csv", all_reference_rows, reference_fields)
    write_csv(output_dir / f"fixed_assets_monthly_comparison_{AUDIT_DATE}.csv", all_comparisons, comparison_fields)
    summary_path = output_dir / f"fixed_assets_cross_trace_summary_{AUDIT_DATE}.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    report_path = output_dir / f"fixed_assets_cross_trace_audit_{AUDIT_DATE}.md"
    report_path.write_text(report_markdown(summary, all_comparisons, all_reference_rows, all_ledger), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    print(f"ĐÃ GHI {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
