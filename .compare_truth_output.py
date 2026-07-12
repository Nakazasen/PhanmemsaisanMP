from pathlib import Path
import openpyxl, xlrd
src=Path('raw/10.07.2026'); out=Path('reference_outputs/secondary/FY2027')
for prefix in ('14.','15.','16.','72.','73.'):
 sp=next(src.glob(prefix+'*人員・時間計画表*.xls')); op=next(out.glob(prefix+'*.xlsx'))
 x=xlrd.open_workbook(str(sp)).sheet_by_index(0); wbv=openpyxl.load_workbook(op,data_only=False,read_only=True); wbd=openpyxl.load_workbook(op,data_only=True,read_only=True); sn=next(n for n in wbv.sheetnames if '内訳' in n); v=wbv[sn]; d=wbd[sn]
 print('\n###',prefix,'sourceCC',x.cell_value(4,0),'outputCC',v['B5'].value)
 pairs=[('JP HC',10,24),('VN HC',13,25),('JP fixed',17,8),('VN fixed',20,9),('JP OT',24,16),('VN OT',27,17)]
 for name,sr,orr in pairs:
  sv=[x.cell_value(sr-1,c) for c in range(2,14)]; ov=[d.cell(orr,c).value for c in range(6,18)]
  print(name,'MATCH' if sv==ov else 'DIFF','src=',sv,'out=',ov)
 wbv.close();wbd.close()
