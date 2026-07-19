import os
import shutil
import time
import hashlib
import gc
import sys
from pathlib import Path
from PIL import Image, ImageStat
import numpy as np
import openpyxl
import win32com.client

sys.stdout.reconfigure(encoding='utf-8')

def print_flush(msg):
    print(msg, flush=True)

# Mappings of the 17 sheets
MAPPINGS = [
    { "sheet": "V01_MASTER", "src_sheet": "Sheet1", "overview_range": "A1:F55", "zoom_range": "A50:F55", "req": "General overview", "desc": "Workbook overview and high-level data-entry workflow" },
    { "sheet": "V02_0906_LIST", "src_sheet": "Hạng mục cần cải tiến", "overview_range": "A176:N233", "zoom_range": "A194:N218", "req": "09.06 change list", "desc": "Current 09.06 issue list: headcount input, fixed assets, health check, stationery, admin allocation, descriptions, separator rows" },
    { "sheet": "V03_GUI_INPUT", "src_sheet": "Hạng mục cần cải tiến", "overview_range": "A70:N110", "zoom_range": "A87:N93", "req": "Requirement 6", "desc": "GUI/input headcount: 12 months and December male/female split" },
    { "sheet": "V04_HEADCOUNT", "src_sheet": "Hạng mục cần cải tiến", "overview_range": "A143:N171", "zoom_range": "A156:N169", "req": "09.06 headcount inputs", "desc": "Replacement manual headcount inputs used by allocation rules" },
    { "sheet": "V05_SEPARATOR", "src_sheet": "Hạng mục cần cải tiến", "overview_range": "A176:N218", "zoom_range": "A216:N218", "req": "Requirement 7-8", "desc": "Additional description and one blank row between each source-file block" },
    { "sheet": "V06_ADMIN_OV", "src_sheet": "Chi phí phân bổ từ hành chính ", "overview_range": "A1:V40", "zoom_range": "A31:T40", "req": "Admin allocation overview", "desc": "Administrative allocation sheet overview plus 12-month allocation formula notes" },
    { "sheet": "V07_ADMIN_CC", "src_sheet": "Chi phí phân bổ từ hành chính ", "overview_range": "A68:V141", "zoom_range": "A95:V141", "req": "Admin cost center/account mapping", "desc": "Cost center group and account-code lookup instructions" },
    { "sheet": "V08_ADMIN_A144", "src_sheet": "Chi phí phân bổ từ hành chính ", "overview_range": "A141:V182", "zoom_range": "A144:V169", "req": "Admin A144:A169", "desc": "Administrative allocation items from A144 through A169 and related right-side columns" },
    { "sheet": "V09_ADMIN_HEALTH", "src_sheet": "Chi phí phân bổ từ hành chính ", "overview_range": "A171:V182", "zoom_range": "A174:V180", "req": "Health-check periodic", "desc": "Male/female regular health-check allocation notes" },
    { "sheet": "V10_ADMIN_NEW", "src_sheet": "Chi phí phân bổ từ hành chính ", "overview_range": "A192:V215", "zoom_range": "A195:V215", "req": "New-hire stationery and recruitment health check", "desc": "New-hire monthly delta, stationery unit prices, and recruitment health-check instructions" },
    { "sheet": "V11_FIXED_ASSET", "src_sheet": "Chi phí tài sản cố định", "overview_range": "A1:AF66", "zoom_range": "A41:AF66", "req": "Fixed asset instructions", "desc": "Fixed asset depreciation and interest instructions including right-side callouts" },
    { "sheet": "V12_SYSTEM", "src_sheet": "Chi phí hệ thống", "overview_range": "A1:I97", "zoom_range": "A86:I97", "req": "System cost instructions", "desc": "System-cost calculation and formula workflow" },
    { "sheet": "V13_DEPR_LAND", "src_sheet": "Chi phí khấu hao, lãi nhà đất", "overview_range": "A1:V131", "zoom_range": "A60:V104", "req": "Depreciation/land/water instructions", "desc": "Depreciation, land interest, electricity and water instructions" },
    { "sheet": "V14_NNN", "src_sheet": "Chi phí làm giấy tờ cho NNN", "overview_range": "A1:Q48", "zoom_range": "A16:Q24", "req": "NNN paperwork", "desc": "Foreign worker paperwork cost month/account instructions" },
    { "sheet": "V15_BIRTHDAY", "src_sheet": "Chi phí sinh nhật", "overview_range": "A1:P48", "zoom_range": "A43:P48", "req": "Birthday/new-hire add-on", "desc": "Birthday cost with same-month new-hire add-on" },
    { "sheet": "V16_CC_LOOKUP", "src_sheet": "原価センタ", "overview_range": "A1:E66", "zoom_range": "A1:E66", "req": "Cost center lookup", "desc": "Cost center lookup sheet used by admin allocation instructions" },
    { "sheet": "V17_ACCOUNT", "src_sheet": "勘定科目", "overview_range": "A1:I244", "zoom_range": "A1:I80", "req": "Account lookup", "desc": "Account-code lookup sheet used by admin allocation instructions" }
]

def audit_image(path):
    if not os.path.exists(path):
        return "NOT_FOUND", 0, 0, 100.0, 0
    try:
        with Image.open(path) as img:
            width, height = img.size
            rgb_img = img.convert('RGB')
            pixels = np.array(rgb_img)
            near_white = (pixels[:, :, 0] >= 248) & (pixels[:, :, 1] >= 248) & (pixels[:, :, 2] >= 248)
            near_white_ratio = np.mean(near_white)
            
            gray_img = img.convert('L')
            stat = ImageStat.Stat(gray_img)
            stddev = stat.stddev[0]
            
            if near_white_ratio > 0.995:
                classification = 'BLANK' if near_white_ratio >= 1.0 else 'NEAR_BLANK'
            elif stddev < 1.0:
                classification = 'NEAR_BLANK'
            else:
                classification = 'VALID_CONTENT'
                
            return classification, width, height, near_white_ratio, stddev
    except Exception as e:
        return f"ERROR: {str(e)}", 0, 0, 100.0, 0

def capture_range_com(excel, wb, sheet_name, rng_address, dest_path):
    ws = wb.Worksheets(sheet_name)
    ws.Activate()
    rng = ws.Range(rng_address)
    
    # Scroll range into view to force rendering
    excel.ActiveWindow.ScrollRow = rng.Row
    excel.ActiveWindow.ScrollColumn = rng.Column
    rng.Select()
    excel.Calculate()
    
    max_attempts = 5
    for attempt in range(1, max_attempts + 1):
        print_flush(f"  Capture attempt {attempt} for {sheet_name} {rng_address}...")
        
        # Copy to Clipboard
        # Appearance=1 (xlScreen), Format=-4147 (xlBitmap)
        rng.CopyPicture(1, -4147)
        time.sleep(0.2 * attempt)
        
        # Paste into temporary ChartObject
        # Padding width/height slightly to avoid cropping borders
        chart_obj = ws.ChartObjects().Add(0, 0, rng.Width + 2, rng.Height + 2)
        chart_obj.Select()
        
        try:
            chart_obj.Chart.Paste()
            chart_obj.Chart.Export(dest_path, "PNG")
        except Exception as e:
            print_flush(f"    Chart paste/export failed: {str(e)}")
        finally:
            chart_obj.Delete()
            
        # Audit the exported PNG
        classification, w, h, ratio, stddev = audit_image(dest_path)
        print_flush(f"    Result: {classification} ({w}x{h}, near_white={ratio*100:.2f}%, stddev={stddev:.2f})")
        
        if classification == 'VALID_CONTENT':
            return True
        else:
            if os.path.exists(dest_path):
                os.remove(dest_path)
            time.sleep(0.5)
            
    return False

def rebuild_all():
    project_root = Path(__file__).resolve().parent
    source_root = project_root / "raw"
    output_root = project_root / "OUTPUT_FY2027" / "tmp_requirement_0906_visual_repair"
    canonical_path = str(source_root / "Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026.xlsx")
    target_path = str(source_root / "Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026_ảnh.xlsx")
    temp_dir = str(output_root / "new_media")
    backup_dir = str(output_root)
    csv_path = str(output_root / "new_snapshot_pixel_audit.csv")

    os.makedirs(temp_dir, exist_ok=True)

    # 1. Backup old snapshot if it exists
    if os.path.exists(target_path):
        backup_path = os.path.join(backup_dir, "backup_bad_snapshot.xlsx")
        print_flush(f"Backing up bad snapshot to {backup_path}")
        shutil.copy2(target_path, backup_path)
        os.remove(target_path)

    # Calculate SHA-256 of canonical workbook
    with open(canonical_path, "rb") as f:
        canonical_sha = hashlib.sha256(f.read()).hexdigest().upper()
    modified_time_ts = os.path.getmtime(canonical_path)
    modified_time_str = time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime(modified_time_ts))

    # 2. Start Excel and capture images
    excel = None
    wb = None
    images_captured = {}
    try:
        print_flush("Starting visible Excel application...")
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False
        
        print_flush(f"Opening read-only canonical workbook: {canonical_path}")
        wb = excel.Workbooks.Open(canonical_path, ReadOnly=True)
        
        for mapping in MAPPINGS:
            sheet_name = mapping["sheet"]
            src_sheet = mapping["src_sheet"]
            print_flush(f"Processing sheet {sheet_name} (from {src_sheet})...")
            
            # Overview
            overview_png = os.path.join(temp_dir, f"{sheet_name}_overview.png")
            success = capture_range_com(excel, wb, src_sheet, mapping["overview_range"], overview_png)
            if not success:
                raise Exception(f"Failed to capture overview range {mapping['overview_range']} for sheet {sheet_name}!")
            
            # Zoom
            zoom_png = os.path.join(temp_dir, f"{sheet_name}_zoom.png")
            success = capture_range_com(excel, wb, src_sheet, mapping["zoom_range"], zoom_png)
            if not success:
                raise Exception(f"Failed to capture zoom range {mapping['zoom_range']} for sheet {sheet_name}!")
                
            images_captured[sheet_name] = (overview_png, zoom_png)
            
    except Exception as e:
        print_flush(f"CRITICAL CAPTURE ERROR: {str(e)}")
        sys.exit(1)
    finally:
        if wb:
            print_flush("Closing Excel workbook...")
            wb.Close(False)
        if excel:
            print_flush("Quitting Excel Application...")
            excel.Quit()
        # Clean COM references
        del wb
        del excel
        gc.collect()
        time.sleep(2)
        # Kill any remaining excel processes to be absolutely safe
        os.system("taskkill /F /IM excel.exe >nul 2>&1")

    # 3. Assemble visual workbook using openpyxl
    print_flush("Assembling new visual workbook copy...")
    new_wb = openpyxl.Workbook()
    
    # Write MANIFEST sheet
    manifest_ws = new_wb.active
    manifest_ws.title = "MANIFEST"
    manifest_ws.views.sheetView[0].showGridLines = True
    
    manifest_ws.cell(1, 1, "VISUAL SUPPORT ONLY — NOT CANONICAL SOURCE")
    manifest_ws.cell(2, 1, "DO NOT USE 04.06.2026_ảnh.xlsx")
    
    manifest_ws.cell(4, 1, "Canonical source path")
    manifest_ws.cell(4, 2, canonical_path)
    
    manifest_ws.cell(5, 1, "Canonical filename")
    manifest_ws.cell(5, 2, os.path.basename(canonical_path))
    
    manifest_ws.cell(6, 1, "Source last-modified UTC")
    manifest_ws.cell(6, 2, modified_time_str)
    
    manifest_ws.cell(7, 1, "Source SHA-256")
    manifest_ws.cell(7, 2, canonical_sha)
    
    generation_ts = time.strftime('%Y-%m-%d %H:%M:%S %z')
    manifest_ws.cell(8, 1, "Snapshot generation timestamp")
    manifest_ws.cell(8, 2, generation_ts)
    
    manifest_ws.cell(9, 1, "Principle")
    manifest_ws.cell(9, 2, "Workbook 09.06.2026.xlsx wins over this visual snapshot, derived Markdown, and old audits.")
    
    manifest_ws.cell(10, 1, "Obsolete visual source")
    manifest_ws.cell(10, 2, "D:\\Sandbox\\MP2027\\raw\\Cải tiến nhập dữ liệu chung vào file MPnew 04.06.2026_ảnh.xlsx was not used.")
    
    headers = ["Snapshot sheet", "Source sheet", "Overview range", "Zoom range", "Requirement", "Description"]
    for col_idx, h in enumerate(headers, 1):
        cell = manifest_ws.cell(12, col_idx, h)
        cell.font = openpyxl.styles.Font(bold=True)
        
    for row_idx, mapping in enumerate(MAPPINGS, 13):
        manifest_ws.cell(row_idx, 1, mapping["sheet"])
        manifest_ws.cell(row_idx, 2, mapping["src_sheet"])
        manifest_ws.cell(row_idx, 3, mapping["overview_range"])
        manifest_ws.cell(row_idx, 4, mapping["zoom_range"])
        manifest_ws.cell(row_idx, 5, mapping["req"])
        manifest_ws.cell(row_idx, 6, mapping["desc"])

    # Auto-fit columns of MANIFEST sheet
    for col in manifest_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        manifest_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Write visual sheets
    pixel_audits = []
    for mapping in MAPPINGS:
        sheet_name = mapping["sheet"]
        overview_path, zoom_path = images_captured[sheet_name]
        
        ws = new_wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Write warning
        ws.cell(1, 1, "VISUAL SUPPORT ONLY — NOT CANONICAL SOURCE").font = openpyxl.styles.Font(bold=True, color="FF0000")
        
        # Overview caption
        overview_caption = f"OVERVIEW | Source range: {mapping['overview_range']} | Requirement: {mapping['req']} | Source sheet: {mapping['src_sheet']} | Canonical source: Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026.xlsx | Description: {mapping['desc']}"
        ws.cell(3, 1, overview_caption).font = openpyxl.styles.Font(bold=True)
        
        # Add Overview Image
        img_ov = openpyxl.drawing.image.Image(overview_path)
        ws.add_image(img_ov, "A5")
        
        # Calculate rows occupied by overview image
        # Excel default row height is ~20 pixels.
        num_rows_ov = int(img_ov.height / 20) + 1
        
        # Placement of Zoom section (leaving 3 rows of gap)
        zoom_caption_row = 5 + num_rows_ov + 3
        zoom_img_row = zoom_caption_row + 2
        
        # Zoom caption
        zoom_caption = f"ZOOM | Source range: {mapping['zoom_range']} | Requirement: {mapping['req']} | Source sheet: {mapping['src_sheet']} | Canonical source: Cải tiến nhập dữ liệu chung vào file MPnew 09.06.2026.xlsx | Description: {mapping['desc']}"
        ws.cell(zoom_caption_row, 1, zoom_caption).font = openpyxl.styles.Font(bold=True)
        
        # Add Zoom Image
        img_zm = openpyxl.drawing.image.Image(zoom_path)
        ws.add_image(img_zm, f"A{zoom_img_row}")
        
        # Set large row heights for caption rows to wrap text and avoid clipping
        ws.row_dimensions[3].height = 24.0
        ws.row_dimensions[zoom_caption_row].height = 24.0
        
        # Add to pixel audits
        classification_ov, w_ov, h_ov, ratio_ov, stddev_ov = audit_image(overview_path)
        pixel_audits.append({
            'filename': f"{sheet_name}_overview.png",
            'sheet': sheet_name,
            'view': 'OVERVIEW',
            'width': w_ov,
            'height': h_ov,
            'ratio': ratio_ov,
            'stddev': stddev_ov,
            'classification': classification_ov
        })
        
        classification_zm, w_zm, h_zm, ratio_zm, stddev_zm = audit_image(zoom_path)
        pixel_audits.append({
            'filename': f"{sheet_name}_zoom.png",
            'sheet': sheet_name,
            'view': 'ZOOM',
            'width': w_zm,
            'height': h_zm,
            'ratio': ratio_zm,
            'stddev': stddev_zm,
            'classification': classification_zm
        })

    # Save rebuilt workbook
    print_flush(f"Saving rebuilt visual support workbook to {target_path}...")
    new_wb.save(target_path)
    new_wb.close()
    print_flush("Visual support workbook saved successfully.")

    # Write Pixel Audit CSV
    print_flush(f"Writing pixel audit CSV to {csv_path}...")
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['filename', 'sheet', 'view', 'width', 'height', 'ratio', 'stddev', 'classification'])
        writer.writeheader()
        for audit in pixel_audits:
            writer.writerow(audit)
    print_flush("Pixel audit CSV written successfully.")

    # 4. Report summaries
    blank_count = sum(1 for a in pixel_audits if a['classification'] == 'BLANK')
    near_blank_count = sum(1 for a in pixel_audits if a['classification'] == 'NEAR_BLANK')
    valid_count = sum(1 for a in pixel_audits if a['classification'] == 'VALID_CONTENT')
    print_flush(f"\nFinal Assembly Verification:")
    print_flush(f" - Valid Content Count: {valid_count}")
    print_flush(f" - Blank Count: {blank_count}")
    print_flush(f" - Near-Blank Count: {near_blank_count}")

if __name__ == '__main__':
    import csv
    rebuild_all()
