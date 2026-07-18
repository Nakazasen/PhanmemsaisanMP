"""Resolve MP2027 source workbooks from a configurable ordered manifest."""

from __future__ import annotations

import csv
from functools import lru_cache
from hashlib import sha256
import json
import os
from pathlib import Path
import re
import unicodedata

import openpyxl


MANIFEST_FILENAME = "source_file_order.csv"
MANIFEST_XLSX_FILENAME = "source_file_order.xlsx"
LEGACY_MANIFEST_COLUMNS = ("order", "category", "filename", "enabled", "description")
MANIFEST_COLUMNS = LEGACY_MANIFEST_COLUMNS + ("status", "detection_method", "signature", "reason")
SUPPORTED_SOURCE_SUFFIXES = {".xls", ".xlsx", ".xlsm"}
SYSTEM_FILENAMES = {"form.xlsx", MANIFEST_FILENAME.lower(), MANIFEST_XLSX_FILENAME.lower()}
SYSTEM_PREFIXES = ("~$", "mp_cc_")
SYSTEM_EXACT_OUTPUTS = {"mp2027_audit_report.md", "mp2027_missing_inputs.csv"}
VALID_STATUSES = {"recognized", "needs_review", "ignored"}


DEFAULT_DESCRIPTIONS = {
    "facility": "Nguồn cơ sở vật chất",
    "fixed_assets": "Nguồn tài sản cố định",
    "it_simulation": "Nguồn mô phỏng hệ thống",
    "ga": "Nguồn Tổng vụ",
    "birthday": "Nguồn sinh nhật",
    "allocation_rules": "Nguồn quy tắc phân bổ",
    "nnn_paperwork": "Nguồn giấy tờ NNN",
}

CATEGORY_DISPLAY_NAMES = {
    "facility": "Cơ sở vật chất",
    "fixed_assets": "Tài sản cố định",
    "it_simulation": "Mô phỏng hệ thống",
    "ga": "Tổng vụ",
    "birthday": "Sinh nhật",
    "allocation_rules": "Quy tắc phân bổ",
    "nnn_paperwork": "Giấy tờ NNN",
}

CATEGORY_ORDER = {
    "facility": 10,
    "fixed_assets": 20,
    "it_simulation": 30,
    "ga": 40,
    "birthday": 50,
    "allocation_rules": 60,
    "nnn_paperwork": 70,
}


def _source_dir_path(source_dir: str | None) -> Path | None:
    if not source_dir:
        return None
    path = Path(source_dir)
    return path if path.is_dir() else None


def _normalize_row(row: dict[str, object], base_dir: Path) -> dict[str, str] | None:
    enabled = str(row.get("enabled", "1")).strip().lower()
    filename = str(row.get("filename", "")).strip()
    category = str(row.get("category", "")).strip()
    raw_status = str(row.get("status", "")).strip().lower()
    is_disabled = enabled in {"0", "false", "no", "n"}
    if not filename:
        return None
    if not category and not is_disabled and raw_status != "ignored":
        return None

    normalized = {key: str(row.get(key, "")).strip() for key in MANIFEST_COLUMNS}
    normalized["enabled"] = "0" if is_disabled else "1"
    if raw_status in VALID_STATUSES:
        normalized["status"] = raw_status
    elif is_disabled and not category:
        normalized["status"] = "ignored"
    else:
        normalized["status"] = "recognized" if category else "needs_review"
    if not normalized["detection_method"]:
        normalized["detection_method"] = "manifest" if category else "manual"
    # A manifest is an annual source boundary, not a general file picker.  Do
    # not allow `../FY2027/...` (or an absolute path) to silently escape it.
    candidate = (base_dir / filename).resolve()
    try:
        candidate.relative_to(base_dir.resolve())
    except ValueError:
        normalized["_path"] = ""
        normalized["_invalid_path"] = "1"
        return normalized
    normalized["_path"] = str(candidate)
    return normalized


def _sort_entries(entries: list[dict[str, str]]) -> list[dict[str, str]]:
    def order_key(row: dict[str, str]) -> tuple[int, str]:
        try:
            order = int(str(row.get("order", "")).strip())
        except ValueError:
            order = 9999
        return order, str(row.get("filename", "")).lower()

    return sorted(entries, key=order_key)


def _read_csv_manifest(path: Path, base_dir: Path) -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            normalized = _normalize_row(row, base_dir)
            if normalized:
                entries.append(normalized)
    return _sort_entries(entries)


def _read_xlsx_manifest(path: Path, base_dir: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = [
            str(worksheet.cell(row=1, column=col).value or "").strip()
            for col in range(1, worksheet.max_column + 1)
        ]
        header_index = {name: idx + 1 for idx, name in enumerate(headers)}
        if not all(column in header_index for column in ("order", "category", "filename")):
            return []

        entries: list[dict[str, str]] = []
        for row_index in range(2, worksheet.max_row + 1):
            raw = {
                column: worksheet.cell(row=row_index, column=header_index.get(column, 0)).value
                for column in MANIFEST_COLUMNS
                if header_index.get(column)
            }
            normalized = _normalize_row(raw, base_dir)
            if normalized:
                entries.append(normalized)
        return _sort_entries(entries)
    finally:
        workbook.close()


def _existing_enabled_entries(entries: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        entry
        for entry in entries
        if str(entry.get("enabled", "1")).strip() != "0"
        and str(entry.get("status", "recognized")).strip().lower() == "recognized"
        and not entry.get("_invalid_path")
        and Path(str(entry.get("_path", ""))).is_file()
    ]


def _is_system_or_generated_file(path: Path) -> bool:
    lower = path.name.lower()
    if lower in SYSTEM_FILENAMES or lower in SYSTEM_EXACT_OUTPUTS:
        return True
    if any(lower.startswith(prefix) for prefix in SYSTEM_PREFIXES):
        return True
    if lower.endswith(".tmp_export.xlsx") or lower.endswith(".tmp.xlsx"):
        return True
    return False


def _source_candidates(base_dir: Path) -> list[Path]:
    return [
        path
        for path in base_dir.iterdir()
        if path.is_file()
        and path.suffix.lower() in SUPPORTED_SOURCE_SUFFIXES
        and not _is_system_or_generated_file(path)
    ]


def inventory_source_files(source_dir: str | None) -> list[Path]:
    """Return every business-workbook candidate; never classify by omission."""
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        return []
    return sorted(_source_candidates(base_dir), key=lambda path: path.name.casefold())


def _normalize_signature_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.replace("\n", " ").replace("\u3000", " ").casefold().split())


def _filename_category(path: Path) -> str | None:
    name = path.name
    lower = name.lower()
    if "fixed_assets_information" in lower or "固定資産情報" in name:
        return "fixed_assets"
    if "simulation" in lower and re.search(r"fy\s*20\d{2}", lower):
        return "it_simulation"
    if re.search(r"mpfy\s*20\d{2}", lower) and "施設" in name:
        return "facility"
    if re.search(r"fy\s*20\d{2}\s+mp", lower) and "振替" in name:
        return "ga"
    if "sinh" in lower and re.search(r"fy\s*20\d{2}", lower):
        return "birthday"
    if "配賦額一覧" in name or (re.search(r"fy\s*20\d{2}", lower) and "allocation" in lower):
        return "allocation_rules"
    if "nnn" in lower or "giấy tờ" in lower or "giay to" in lower:
        return "nnn_paperwork"
    return None


def _sheet_blob(worksheet, *, rows: int = 12, columns: int = 20) -> str:
    values: list[str] = []
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(rows, worksheet.max_row),
        max_col=min(columns, worksheet.max_column),
        values_only=True,
    ):
        values.extend(
            _normalize_signature_text(value)
            for value in row
            if isinstance(value, str) and value.strip()
        )
    return " | ".join(values)


def _matches_structure(workbook) -> tuple[str, ...]:
    matches: set[str] = set()
    sheet_names = list(workbook.sheetnames)
    normalized_names = [_normalize_signature_text(name) for name in sheet_names]
    blobs = [_sheet_blob(workbook[name]) for name in sheet_names]
    all_blob = " | ".join(blobs)

    facility_markers = ("減価償却費", "固定資産金利", "水道光熱費")
    if all(any(marker in name for name in sheet_names) for marker in facility_markers):
        matches.add("facility")

    if any(re.fullmatch(r"sinh nhat mp fy20\d{2}", name) for name in normalized_names):
        matches.add("birthday")

    has_calc_sheet = any(
        any(token in name for token in ("cach tinh", "振替", "計算"))
        for name in normalized_names
    )
    has_ga_main = any(
        "vnd" in blob
        and any(token in blob for token in ("yotei", "item", "項目"))
        and any(token in blob for token in ("費", "製造", "tai khoan", "account"))
        for blob in blobs
    )
    if has_calc_sheet and has_ga_main:
        matches.add("ga")

    first_blob = blobs[0] if blobs else ""
    first_sheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None
    if (
        first_sheet is not None
        and first_sheet.max_row >= 20
        and first_sheet.max_column >= 10
        and "vnd" in first_blob
        and any(token in first_blob for token in ("don gia", "単価"))
        and any(token in first_blob for token in ("ma tai khoan", "tai khoan", "計上月"))
    ):
        matches.add("allocation_rules")

    has_fy_sheet = any(re.fullmatch(r"fy20\d{2}", name) for name in normalized_names)
    has_nnn_headers = (
        any(token in all_blob for token in ("cost center", "costcenter", "原価センタ", "cc_code"))
        and any(token in all_blob for token in ("account code", "accountcode", "勘定科目", "ma tai khoan"))
        and (
            re.search(r"(?<!\d)20\d{2}(?!\d)", all_blob) is not None
            or any(token in all_blob for token in ("4月", "apr"))
        )
    )
    if has_fy_sheet and has_nnn_headers:
        matches.add("nnn_paperwork")

    fixed_asset_match = False
    try:
        # Local import avoids the parser -> manifest resolver import cycle.
        from src.parsers.fixed_assets import _sheet_content_score

        fixed_asset_match = any(
            _sheet_content_score(workbook[name]) >= 110 for name in sheet_names
        )
    except Exception:
        fixed_asset_tokens = (
            "asset category", "asset class", "資産クラス", "資産分類",
            "depreciation cost center", "管理原価センタ", "費用負担原価センタ",
            "monthly depreciation", "減価償却費", "償却費",
        )
        fixed_asset_match = (
            sum(token in all_blob for token in fixed_asset_tokens) >= 2
            and any(token in all_blob for token in ("asset no", "資産番号", "固定資産番号"))
        )
    if fixed_asset_match and "facility" not in matches:
        matches.add("fixed_assets")

    component_tokens = ("vpn", "メール", "r3", "mes", "plm", "qlik", "vps", "ams")
    has_it_sheet = sum(any(token in name for token in component_tokens) for name in normalized_names) >= 2
    has_it_headers = (
        any(token in all_blob for token in ("原価センター", "cost center"))
        and any(token in all_blob for token in ("課金金額", "amount vnd", "amount usd"))
    )
    if has_it_sheet and has_it_headers:
        matches.add("it_simulation")

    return tuple(sorted(matches))


@lru_cache(maxsize=256)
def _inspect_source_signature_cached(path_text: str, size: int, mtime_ns: int) -> tuple[str, tuple[str, ...], str]:
    path = Path(path_text)
    if path.suffix.lower() == ".xls":
        payload = {"suffix": ".xls", "size": size, "mtime_ns": mtime_ns}
        signature = sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()
        category = _filename_category(path)
        matches = (category,) if category == "it_simulation" else ()
        reason = (
            "Định dạng XLS được parser mô phỏng hệ thống hỗ trợ; xác nhận theo tên và định dạng."
            if matches
            else "Không thể đọc dấu vân tay XLS an toàn; cần người dùng xác nhận."
        )
        return signature, matches, reason

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        payload = {"suffix": path.suffix.lower(), "size": size, "mtime_ns": mtime_ns}
        signature = sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()
        return signature, (), f"Không đọc được cấu trúc workbook ({type(exc).__name__})."

    try:
        structure = []
        for worksheet in workbook.worksheets:
            structure.append(
                {
                    "sheet": _normalize_signature_text(worksheet.title),
                    "max_column": worksheet.max_column,
                    "text": _sheet_blob(worksheet),
                }
            )
        matches = set(_matches_structure(workbook))
        filename_years = {
            int(value) for value in re.findall(r"(?i)fy\s*(20\d{2})", path.name)
        }
        if len(filename_years) == 1:
            try:
                from src.parsers.ga import _find_ga_main_sheet, _normalize_text

                normalized_sheets = [_normalize_text(name) for name in workbook.sheetnames]
                has_calc_sheet = any(
                    "cach tinh" in name or "振替" in name or "計算" in name
                    for name in normalized_sheets
                )
                if has_calc_sheet and _find_ga_main_sheet(workbook, filename_years.pop()) is not None:
                    matches.discard("allocation_rules")
                    matches.add("ga")
            except Exception:
                pass
        matches = tuple(sorted(matches))
    finally:
        workbook.close()
    signature = sha256(
        json.dumps(structure, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return signature, matches, ""


def inspect_source_signature(path: str | Path) -> dict[str, object]:
    """Inspect bounded workbook structure, cached by path metadata."""
    candidate = Path(path).resolve()
    stat = candidate.stat()
    signature, matches, reason = _inspect_source_signature_cached(
        str(candidate), stat.st_size, stat.st_mtime_ns
    )
    return {"signature": signature, "matches": matches, "reason": reason}


def _looks_like_form_workbook(path: Path) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    required_sheet_markers = (
        "採算表(usd)", "採算表(vnd)", "勘定科目", "原価センタ", "稼働日",
    )
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        normalized = {_normalize_signature_text(name) for name in workbook.sheetnames}
        return all(marker in normalized for marker in required_sheet_markers)
    finally:
        workbook.close()


def classify_source_candidate(path: str | Path) -> dict[str, str]:
    """Classify one candidate without ever hiding unknown or ambiguous files."""
    candidate = Path(path)
    if _looks_like_form_workbook(candidate):
        return {
            "category": "",
            "status": "ignored",
            "confidence": "high",
            "detection_method": "system_structure",
            "signature": str(inspect_source_signature(candidate)["signature"]),
            "reason": "Workbook có cấu trúc FORM đầu ra/mẫu, không phải file nguồn chi phí.",
            "_matches": "",
        }
    inspected = inspect_source_signature(candidate)
    matches = tuple(inspected["matches"])
    filename_category = _filename_category(candidate)
    signature = str(inspected["signature"])
    if len(matches) == 1:
        category = matches[0]
        method = "filename+structure" if filename_category == category else "structure"
        reason = f"Cấu trúc workbook khớp duy nhất loại {CATEGORY_DISPLAY_NAMES[category]}."
        if candidate.suffix.lower() == ".xls":
            method = "filename+parser"
            reason = str(inspected["reason"])
        return {
            "category": category,
            "status": "recognized",
            "confidence": "high",
            "detection_method": method,
            "signature": signature,
            "reason": reason,
            "_matches": category,
        }
    if len(matches) > 1:
        names = ", ".join(CATEGORY_DISPLAY_NAMES.get(item, item) for item in matches)
        reason = f"Cấu trúc khớp nhiều loại ({names}); cần người dùng xác nhận."
    else:
        reason = str(inspected["reason"] or "Chưa khớp hợp đồng cấu trúc của parser nào.")
    return {
        "category": "",
        "status": "needs_review",
        "confidence": "none",
        "detection_method": "structure",
        "signature": signature,
        "reason": reason,
        "_matches": "|".join(matches),
    }


def _classify_source_file(path: Path) -> str | None:
    """Backward-compatible category helper for callers that only need a key."""
    result = classify_source_candidate(path)
    return result["category"] or None


def _it_order(path: Path) -> int:
    lower = path.name.lower()
    if "apr" in lower or "june" in lower:
        return 0
    if "july" in lower or "dec" in lower:
        return 1
    if "jan" in lower or "march" in lower:
        return 2
    return 9


def _detected_entry_sort_key(entry: dict[str, str]) -> tuple[int, int, str]:
    category = entry.get("category", "")
    return (
        CATEGORY_ORDER.get(category, 999),
        _it_order(Path(entry["filename"])),
        entry["filename"].casefold(),
    )


def detect_source_files(source_dir: str | None) -> list[dict[str, str]]:
    """Inventory and classify every candidate workbook in the selected folder."""
    entries: list[dict[str, str]] = []
    for path in inventory_source_files(source_dir):
        classification = classify_source_candidate(path)
        category = classification["category"]
        entries.append(
            {
                "order": "",
                "category": category,
                "filename": path.name,
                "enabled": "1" if classification["status"] == "recognized" else "0",
                "description": (
                    DEFAULT_DESCRIPTIONS.get(category, "Cần xác nhận loại nguồn")
                    if classification["status"] != "ignored" else
                    "File hệ thống/FORM — không dùng làm nguồn chi phí"
                ),
                "status": classification["status"],
                "detection_method": classification["detection_method"],
                "signature": classification["signature"],
                "reason": classification["reason"],
                "_matches": classification.get("_matches", ""),
                "_path": str(path.resolve()),
            }
        )
    entries.sort(key=_detected_entry_sort_key)
    for index, entry in enumerate(entries, start=1):
        entry["order"] = str(index)
    return entries


def _read_saved_manifest(source_dir: str | None, include_csv: bool = True) -> list[dict[str, str]]:
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        return []
    xlsx_path = base_dir / MANIFEST_XLSX_FILENAME
    if xlsx_path.is_file():
        return _read_xlsx_manifest(xlsx_path, base_dir)
    csv_path = base_dir / MANIFEST_FILENAME
    if include_csv and csv_path.is_file():
        return _read_csv_manifest(csv_path, base_dir)
    return []


def merge_manifest_with_detected(
    source_dir: str | None,
    saved_entries: list[dict[str, str]] | None = None,
) -> list[dict[str, str]]:
    """Merge user decisions with a complete current-folder inventory."""
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        return []

    saved = saved_entries if saved_entries is not None else _read_saved_manifest(source_dir)
    detected = detect_source_files(source_dir)
    detected_by_name = {entry["filename"].casefold(): entry for entry in detected}

    merged: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in _sort_entries(saved):
        filename = str(row.get("filename", "")).strip()
        key = filename.casefold()
        if not filename or key in seen:
            continue
        if row.get("_invalid_path"):
            merged.append(dict(row))
            seen.add(key)
            continue

        current = detected_by_name.get(key)
        if current is None:
            missing = dict(row)
            missing.update(
                {
                    "category": "",
                    "enabled": "0",
                    "status": "needs_review",
                    "detection_method": "manifest",
                    "reason": "File đã lưu trong manifest hiện không còn trong thư mục nguồn.",
                    "_path": str((base_dir / filename).resolve()),
                }
            )
            merged.append(missing)
            seen.add(key)
            continue

        merged_row = dict(current)
        saved_status = str(row.get("status", "")).strip().lower()
        saved_category = str(row.get("category", "")).strip()
        saved_method = str(row.get("detection_method", "")).strip().lower()
        saved_signature = str(row.get("signature", "")).strip()
        is_ignored = saved_status == "ignored" or (
            str(row.get("enabled", "1")).strip() == "0" and not saved_category
        )

        if is_ignored:
            merged_row.update(
                {
                    "category": saved_category,
                    "enabled": "0",
                    "status": "ignored",
                    "detection_method": "manual",
                    "description": row.get("description", "") or "Bỏ qua có chủ đích",
                    "reason": row.get("reason", "") or "Người dùng đã chọn bỏ qua file này.",
                }
            )
        elif saved_category:
            current_matches = {
                item for item in str(current.get("_matches", "")).split("|") if item
            }
            structure_still_matches = (
                current.get("category") == saved_category
                or saved_category in current_matches
            )
            if structure_still_matches:
                signature_note = (
                    " Cấu trúc vẫn tương thích dù nội dung workbook đã thay đổi."
                    if saved_signature and saved_signature != current.get("signature") else
                    ""
                )
                merged_row.update(
                    {
                        "category": saved_category,
                        "enabled": row.get("enabled", "1"),
                        "status": "recognized",
                        "detection_method": saved_method or "manifest",
                        "description": row.get("description", "") or DEFAULT_DESCRIPTIONS.get(saved_category, ""),
                        "reason": (row.get("reason", "") or "Giữ phân loại đã xác nhận trong manifest.") + signature_note,
                    }
                )
            else:
                merged_row.update(
                    {
                        "category": "",
                        "enabled": "0",
                        "status": "needs_review",
                        "detection_method": "structure",
                        "reason": "Cấu trúc workbook đã thay đổi; loại đã lưu không còn tương thích với workbook hiện tại.",
                    }
                )
        merged_row["order"] = row.get("order", merged_row["order"])
        merged_row["_path"] = str((base_dir / filename).resolve())
        merged.append(merged_row)
        seen.add(key)

    for row in detected:
        key = row["filename"].casefold()
        if key not in seen:
            merged.append(row)
            seen.add(key)

    for index, row in enumerate(_sort_entries(merged), start=1):
        row["order"] = str(index)
    return merged


def write_source_manifest_xlsx(source_dir: str | None, entries: list[dict[str, str]]) -> str:
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        raise FileNotFoundError(f"Không tìm thấy thư mục nguồn: {source_dir}")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "source_file_order"
    worksheet.append(list(MANIFEST_COLUMNS))
    for entry in _sort_entries(entries):
        worksheet.append([entry.get(column, "") for column in MANIFEST_COLUMNS])
    worksheet.freeze_panes = "A2"
    widths = {"A": 10, "B": 18, "C": 70, "D": 10, "E": 32, "F": 16, "G": 22, "H": 22, "I": 64}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    path = base_dir / MANIFEST_XLSX_FILENAME
    temporary_path = path.with_name(f".{path.name}.tmp.xlsx")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()
    _cached_manifest_snapshot.cache_clear()
    return str(path)


def ensure_source_manifest(source_dir: str | None, *, refresh: bool = True) -> str:
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        raise FileNotFoundError(f"Không tìm thấy thư mục nguồn: {source_dir}")

    entries = merge_manifest_with_detected(source_dir) if refresh else _read_saved_manifest(source_dir)
    if not entries:
        entries = detect_source_files(source_dir)
    write_source_manifest_xlsx(source_dir, entries)
    return str(base_dir / MANIFEST_XLSX_FILENAME)


def read_source_manifest_inventory_fast(source_dir: str | None) -> list[dict[str, str]]:
    """Read saved decisions plus current filenames without opening source workbooks."""
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        return []

    candidates = inventory_source_files(source_dir)
    candidates_by_name = {path.name.casefold(): path for path in candidates}
    entries: list[dict[str, str]] = []
    seen: set[str] = set()

    for saved_row in _sort_entries(_read_saved_manifest(source_dir)):
        row = dict(saved_row)
        filename = str(row.get("filename", "")).strip()
        key = filename.casefold()
        if not filename or key in seen:
            continue
        current = candidates_by_name.get(key)
        if current is None and not row.get("_invalid_path"):
            row.update(
                {
                    "category": "",
                    "enabled": "0",
                    "status": "needs_review",
                    "detection_method": "manifest",
                    "reason": "Tệp đã lưu hiện không còn trong thư mục nguồn.",
                    "_path": str((base_dir / filename).resolve()),
                }
            )
        elif current is not None:
            row["_path"] = str(current.resolve())
        entries.append(row)
        seen.add(key)

    for candidate in candidates:
        key = candidate.name.casefold()
        if key in seen:
            continue
        entries.append(
            {
                "order": "",
                "category": "",
                "filename": candidate.name,
                "enabled": "0",
                "description": "Cần xác nhận loại nguồn",
                "status": "needs_review",
                "detection_method": "inventory",
                "signature": "",
                "reason": "Tệp mới được phát hiện theo tên; chưa quét nội dung.",
                "_path": str(candidate.resolve()),
            }
        )
        seen.add(key)

    for index, row in enumerate(_sort_entries(entries), start=1):
        row["order"] = str(index)
    return entries


def _manifest_inventory_fingerprint(base_dir: Path) -> tuple[tuple[str, int, int], ...]:
    rows: list[tuple[str, int, int]] = []
    for path in inventory_source_files(str(base_dir)):
        try:
            stat = path.stat()
            rows.append((path.name.casefold(), stat.st_size, stat.st_mtime_ns))
        except OSError:
            rows.append((path.name.casefold(), -1, -1))
    for name in (MANIFEST_FILENAME, MANIFEST_XLSX_FILENAME):
        path = base_dir / name
        if path.is_file():
            stat = path.stat()
            rows.append((name.casefold(), stat.st_size, stat.st_mtime_ns))
    return tuple(sorted(rows))


@lru_cache(maxsize=32)
def _cached_manifest_snapshot(
    source_dir: str,
    inventory_fingerprint: tuple[tuple[str, int, int], ...],
) -> tuple[tuple[tuple[str, str], ...], ...]:
    del inventory_fingerprint  # It exists solely to invalidate stale snapshots.
    entries = merge_manifest_with_detected(source_dir)
    return tuple(tuple(sorted((str(key), str(value)) for key, value in entry.items())) for entry in entries)


def read_source_manifest(source_dir: str | None, include_missing: bool = False) -> list[dict[str, str]]:
    """Read source file entries without rewriting the manifest on disk."""
    base_dir = _source_dir_path(source_dir)
    if base_dir is None:
        return []

    snapshot = _cached_manifest_snapshot(
        str(base_dir.resolve()),
        _manifest_inventory_fingerprint(base_dir),
    )
    entries = [dict(items) for items in snapshot]
    if include_missing:
        return entries
    return _existing_enabled_entries(entries)


def resolve_manifest_file(source_dir: str | None, category: str) -> str | None:
    """Return the first existing enabled file for a category."""
    for entry in read_source_manifest(source_dir):
        if entry.get("category") != category:
            continue
        path = Path(str(entry.get("_path", "")))
        if path.is_file():
            return str(path)
    return None


def resolve_manifest_files(source_dir: str | None, category: str) -> list[str]:
    """Return existing enabled files for a category in configured order."""
    paths: list[str] = []
    for entry in read_source_manifest(source_dir):
        if entry.get("category") != category:
            continue
        path = Path(str(entry.get("_path", "")))
        if path.is_file():
            paths.append(str(path))
    return paths


def describe_manifest(source_dir: str | None) -> list[str]:
    lines: list[str] = []
    status_labels = {
        "recognized": "ĐÃ NHẬN DIỆN",
        "needs_review": "CẦN XÁC NHẬN",
        "ignored": "ĐÃ BỎ QUA",
    }
    for entry in read_source_manifest(source_dir, include_missing=True):
        path = Path(str(entry.get("_path", "")))
        existence = "ĐỦ" if path.is_file() else "THIẾU"
        enabled = "BẬT" if str(entry.get("enabled", "1")).strip() != "0" else "TẮT"
        category_key = str(entry.get("category", "")).strip()
        category = CATEGORY_DISPLAY_NAMES.get(category_key, category_key or "Chưa phân loại")
        workflow_status = status_labels.get(str(entry.get("status", "")), "CẦN XÁC NHẬN")
        lines.append(
            "{order}. {category}: {filename} [{workflow_status}; {existence}; {enabled}] — {reason}".format(
                order=str(entry.get("order", "")).strip() or "?",
                category=category,
                filename=str(entry.get("filename", "")).strip(),
                workflow_status=workflow_status,
                existence=existence,
                enabled=enabled,
                reason=str(entry.get("reason", "")).strip() or "Không có ghi chú",
            )
        )
    return lines
