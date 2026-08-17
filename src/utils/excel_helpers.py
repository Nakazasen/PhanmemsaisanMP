"""Các tiện ích Excel dùng chung của MP2027 Manager.

Cung cấp các hàm hỗ trợ đọc và chuẩn hóa sổ làm việc tài chính.
"""
import pandas as pd
import openpyxl
import math
from src.services.template_manifest import (
    layout_hub_sheet_name,
    layout_hub_sheet_name_for_output,
    layout_payload_bounds,
)
from src.utils.fiscal_periods import fiscal_month_order, fiscal_periods
from datetime import datetime
from typing import Optional, Any
from pathlib import Path

# The FORM hub name is resolved only by the approved layout registry.
HUB_SHEET_CANDIDATES: tuple[str, ...] = ()

def get_month_mapping(fiscal_year: int = 2027) -> dict:
    """Returns a mapping of month Index (0-11) to Period String (YYYYMM)."""
    return {index: period for index, period in enumerate(fiscal_periods(fiscal_year))}

def get_fy_months(fiscal_year: int = 2027) -> list:
    """Returns a list of 12 YYYYMM strings for the given fiscal year."""
    return fiscal_periods(fiscal_year)

def get_fy_month_labels(fiscal_year: int = 2027) -> list:
    """Returns a list of 12 numeric month labels (4, 5, ..., 12, 1, 2, 3)."""
    return fiscal_month_order()

def normalize_period(value: Any) -> Optional[str]:
    """Universal period normalizer to YYYYMM format."""
    if pd.isna(value) or value in ('', None): return None
    if isinstance(value, datetime): return value.strftime('%Y%m')
    if hasattr(value, 'year') and hasattr(value, 'month'):
        return f"{int(value.year)}{int(value.month):02d}"
    s = str(value).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%m/%Y', '%m-%Y', '%Y%m'):
        try: return datetime.strptime(s, fmt).strftime('%Y%m')
        except ValueError: continue
    return None

def find_hub_sheet_name(workbook: openpyxl.Workbook) -> str:
    """Return the canonical hub sheet for a template or its derived output.

    Template admission stays fail-closed in ``resolve_template_layout``. This
    compatibility helper also supports derived workbooks, whose payload and
    dynamic rows legitimately change dimensions and approved-cell values.
    """
    try:
        return layout_hub_sheet_name(workbook)
    except ValueError:
        return layout_hub_sheet_name_for_output(workbook)


FORM_TEMPLATE_INPUT_ROWS = (8, 9, 16, 17, 24, 25)
FORM_TEMPLATE_MONTH_COLUMNS = tuple(range(6, 18))  # F:Q
FORM_TEMPLATE_PAYLOAD_COLUMNS = (2, 19, 20)  # B, S, T
# FORM 2026-07-21 reserves rows 30:37 for QLLN-owned structural data.
# Generated shared-cost rows must start below that protected block.
FORM_SHARED_COST_START_ROW = 38
FORM_TEMPLATE_PAYLOAD_START_ROW = FORM_SHARED_COST_START_ROW

# Rows 31:36 are an output contract owned by QLLN.  Two FORM variants have
# circulated in production: one contains these values and one contains blank
# account cells plus lookup formulas.  Generated MP workbooks must be identical
# regardless of which filename or local copy was selected.
FORM_QLNN_PROTECTED_VALUES = (
    (31, 9114120018, "部内間接経費", "部内間接経費"),
    (32, 9114120029, "部外間接経費1", "部外間接経費1"),
    (33, 9114120030, "部外間接経費2", "部外間接経費2"),
    (34, 9114120021, "工場間接経費", "工場間接経費"),
    (35, "\u00a0", "\u00a0", "\u00a0"),
    (36, 9114120009, "社内金利（在庫）", "在庫金利"),
)
FORM_PROFIT_SHEET_NAME = "採算表(VND)"
FORM_PROFIT_LOOKUP_KEY_ALIASES = {
    "部外間接経費1": "部外間接1経費",
    "部外間接経費2": "部外間接2経費",
}


def normalize_form_output_contract(workbook: openpyxl.Workbook) -> dict[str, int | bool]:
    """Repair stable FORM-owned cells on a copied output workbook.

    This intentionally runs only for a real MP FORM with its supporting master
    sheets.  It never mutates the selected template itself; callers invoke it
    after copying the template to the temporary output path.
    """
    required_sheets = {FORM_PROFIT_SHEET_NAME, "勘定科目", "原価センタ"}
    if not required_sheets.issubset(workbook.sheetnames):
        return {
            "form_contract_detected": False,
            "qlnn_cells_repaired": 0,
            "profit_lookup_keys_repaired": 0,
        }

    worksheet = workbook[find_hub_sheet_name(workbook)]
    qlnn_cells_repaired = 0
    for row, account, name, group in FORM_QLNN_PROTECTED_VALUES:
        for column, expected in ((2, account), (3, name), (4, group)):
            cell = worksheet.cell(row=row, column=column)
            if cell.value != expected:
                cell.value = expected
                qlnn_cells_repaired += 1

    profit_sheet = workbook[FORM_PROFIT_SHEET_NAME]
    profit_lookup_keys_repaired = 0
    for row in range(1, int(profit_sheet.max_row or 0) + 1):
        cell = profit_sheet.cell(row=row, column=2)
        replacement = FORM_PROFIT_LOOKUP_KEY_ALIASES.get(cell.value)
        if replacement is not None:
            cell.value = replacement
            profit_lookup_keys_repaired += 1

    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True

    return {
        "form_contract_detected": True,
        "qlnn_cells_repaired": qlnn_cells_repaired,
        "profit_lookup_keys_repaired": profit_lookup_keys_repaired,
    }


def is_form_template_payload_value(cell: Any) -> bool:
    """Return whether a cell contains a concrete value rather than a formula."""
    value = cell.value
    if value is None or (isinstance(value, str) and not value.strip()):
        return False
    return getattr(cell, "data_type", None) != "f" and not (
        isinstance(value, str) and value.lstrip().startswith("=")
    )


def find_form_template_hygiene_issues(
    workbook: openpyxl.Workbook,
    *,
    limit: int = 12,
) -> tuple[str, ...]:
    """List concrete department data that must not ship in a reusable FORM.

    Structural labels in rows 1–29, formulas, styles and master sheets are not
    considered contamination. The bounded result is suitable for user-facing
    preflight messages without dumping workbook contents.
    """
    worksheet = workbook[find_hub_sheet_name(workbook)]
    layout_payload_start, layout_payload_columns = layout_payload_bounds(workbook)
    issue_cells: list[str] = []

    def record(cell: Any) -> bool:
        if is_form_template_payload_value(cell):
            issue_cells.append(cell.coordinate)
        return len(issue_cells) >= max(int(limit), 1)

    if record(worksheet["B5"]):
        return tuple(issue_cells)

    for row_index in FORM_TEMPLATE_INPUT_ROWS:
        for column_index in FORM_TEMPLATE_MONTH_COLUMNS:
            if record(worksheet.cell(row=row_index, column=column_index)):
                return tuple(issue_cells)

    # In read-only mode, repeated worksheet.cell() calls can seek through the
    # worksheet XML for every coordinate. Read the bounded B:S range once and
    # inspect only the non-contiguous payload columns B, S and T in memory.
    payload_start = layout_payload_start
    payload_end = max(payload_start - 1, worksheet.max_row)
    payload_column_offsets = {
        column_index: column_index - 2
        for column_index in layout_payload_columns
    }
    for row in worksheet.iter_rows(
        min_row=payload_start,
        max_row=payload_end,
        min_col=2,
        max_col=max(layout_payload_columns),
        values_only=False,
    ):
        for column_index, offset in payload_column_offsets.items():
            if record(row[offset]):
                return tuple(issue_cells)
    return tuple(issue_cells)


def validate_exchange_rate(value: Any) -> float:
    """Return a safe USD/VND rate or raise instead of silently substituting one."""
    if isinstance(value, str):
        value = value.strip().replace(",", "")
    try:
        rate = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("Tỷ giá USD/VND phải là một số hợp lệ.") from exc
    if not math.isfinite(rate) or rate <= 0 or rate > 1_000_000:
        raise ValueError("Tỷ giá USD/VND phải lớn hơn 0 và không vượt quá 1,000,000.")
    return rate


def read_exchange_rate_from_form(form_path: str) -> float:
    """Read the USD/VND rate from the selected FORM hub-sheet B2 cell."""
    path = Path(form_path)
    if not path.exists(): raise FileNotFoundError(f'Không tìm thấy FORM.xlsx tại {path}')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = find_hub_sheet_name(wb)
        return validate_exchange_rate(wb[sheet_name]['B2'].value)
    finally: wb.close()


def write_exchange_rate_to_form(form_path: str, exchange_rate: Any) -> float:
    """Set the effective rate in a copied output FORM without mutating its template."""
    rate = validate_exchange_rate(exchange_rate)
    path = Path(form_path)
    if not path.exists():
        raise FileNotFoundError(f'Không tìm thấy FORM.xlsx tại {path}')
    wb = openpyxl.load_workbook(path)
    try:
        sheet_name = find_hub_sheet_name(wb)
        wb[sheet_name]['B2'].value = rate
        wb.save(path)
    finally:
        wb.close()
    return rate

import re

CC_CODE_PATTERNS = (
    r"\b\d{4}[A-Za-z]\d{5,}\b",
    r"\b\d{4,10}\b",
)


def normalize_cc_code(val: Any) -> Optional[str]:
    """Normalize cost center code from raw Excel/csv values."""
    if pd.isna(val) or val is None:
        return None

    if isinstance(val, (int, float)):
        try:
            number = int(float(val))
            if number >= 1000:
                return str(number)
        except Exception:
            pass

    s = str(val).strip()
    if not s:
        return None

    compact = s.replace(" ", "")
    for pattern in CC_CODE_PATTERNS:
        direct_match = re.fullmatch(pattern, compact)
        if direct_match:
            return direct_match.group(0).upper()

    for pattern in CC_CODE_PATTERNS:
        match = re.search(pattern, s)
        if match:
            return match.group(0).upper()
    return None


def extract_cc_code(val: Any) -> Optional[str]:
    """Backward-compatible alias for normalized cost center extraction."""
    return normalize_cc_code(val)

def safe_float(val: Any) -> float:
    """Convert a value to float, returning 0.0 for invalid values."""
    if pd.isna(val) or val is None: return 0.0
    try: return float(val)
    except (ValueError, TypeError): return 0.0

def normalize_text(value: str) -> str:
    """Normalize text for consistent mapping."""
    text = str(value or '').replace('\n', ' ').replace('\u3000', ' ')
    return ' '.join(text.split()).strip().lower()

def item_key(value: str) -> str:
    """Generate a lookup key from item description."""
    return normalize_text(str(value or '').split('\n')[0].strip())
