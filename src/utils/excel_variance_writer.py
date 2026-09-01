# -*- coding: utf-8 -*-
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

from typing import List
from src.engine.variance_analyzer import VarianceReport
from src.services.i18n import get_current_language, t
from src.ui.variance_chart import build_variance_chart_rows, resolve_multilingual_font_path


def _localized_headers() -> list[str]:
    return [
        t("col_account_code"),
        t("col_item_name"),
        t("col_prev_year"),
        t("col_curr_year"),
        t("col_curr_year_diff"),
        t("col_pct_diff"),
        t("col_status"),
        t("variance_export_notes"),
    ]


def _localized_status(line) -> str:
    return t(f"variance_status_{line.status.name.lower()}")

def export_variance_report(report: VarianceReport, output_path: str) -> None:
    """
    Export VarianceReport to an Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = t("variance_export_sheet_name")[:31]

    # Headers
    headers = _localized_headers()
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row_idx, line in enumerate(report.lines, start=2):
        ws.cell(row=row_idx, column=1, value=line.account_code)
        ws.cell(row=row_idx, column=2, value=line.item_name)

        c3 = ws.cell(row=row_idx, column=3, value=line.base_value)
        c3.number_format = '#,##0'

        c4 = ws.cell(row=row_idx, column=4, value=line.current_value)
        c4.number_format = '#,##0'

        c5 = ws.cell(row=row_idx, column=5, value=line.variance_absolute)
        c5.number_format = '#,##0'

        c6 = ws.cell(row=row_idx, column=6, value=(line.variance_percent / 100.0) if line.variance_percent is not None else "")
        c6.number_format = '0.00%'

        ws.cell(row=row_idx, column=7, value=_localized_status(line))
        ws.cell(row=row_idx, column=8, value="") # Placeholder for user notes

        if line.is_alert:
            fill = red_fill if line.variance_absolute > 0 else yellow_fill
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_num).fill = fill

    # Auto-fit columns
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['H'].width = 30

    _add_variance_chart_image(ws, report)

    wb.save(output_path)


def _add_variance_chart_image(ws, report: VarianceReport) -> None:
    """Embed a readable chart image rather than Excel's ambiguous default chart."""
    language = get_current_language()
    chart_rows = build_variance_chart_rows(report.lines, language=language)
    if not chart_rows:
        return

    from matplotlib.figure import Figure
    from matplotlib.font_manager import FontProperties
    from matplotlib.ticker import FuncFormatter

    display_rows = list(reversed(chart_rows))
    labels = [row.label for row in display_rows]
    values = [row.amount for row in display_rows]
    font_path = resolve_multilingual_font_path(language=language)
    chart_font = FontProperties(fname=str(font_path)) if font_path else FontProperties()
    figure = Figure(figsize=(12.5, max(4.2, 0.65 * len(display_rows) + 2.4)), dpi=150)
    axes = figure.add_subplot(111)
    positions = list(range(len(display_rows)))
    bars = axes.barh(positions, values, color=[row.color for row in display_rows], height=0.62)
    axes.axvline(0, color="#455A64", linewidth=1)
    axes.set_yticks(positions, labels, fontproperties=chart_font)
    axes.set_title(t("variance_chart_title"), fontproperties=chart_font, fontweight="bold", pad=14)
    axes.set_xlabel(
        f"{t('variance_chart_axis_label')} — {t('variance_chart_legend')}",
        fontproperties=chart_font,
    )
    axes.xaxis.set_major_formatter(FuncFormatter(lambda value, _position: f"{value / 1_000_000:,.1f}M"))
    axes.grid(axis="x", linestyle="--", alpha=0.3)
    axes.set_axisbelow(True)
    for bar, value in zip(bars, values):
        axes.text(
            value / 2,
            bar.get_y() + bar.get_height() / 2,
            f"{value / 1_000_000:+,.2f}M",
            ha="center",
            va="center",
            color="white",
            fontweight="bold",
            fontsize=9,
        )
    figure.subplots_adjust(left=0.34, right=0.94, top=0.88, bottom=0.17)

    image_buffer = BytesIO()
    figure.savefig(image_buffer, format="png", dpi=150)
    image_buffer.seek(0)
    image = OpenpyxlImage(image_buffer)
    image.width = 1180
    image.height = max(400, int(60 * len(display_rows) + 150))
    ws.add_image(image, "J2")

def batch_export_variance_reports(reports: List[VarianceReport], output_path: str) -> None:
    """
    Export multiple VarianceReports to a single Excel file (one sheet per CC).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    headers = _localized_headers()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for report in reports:
        # Sheet name max 31 chars
        title = str(report.context.cost_center_code)[:31]
        if title in wb.sheetnames:
            title = title[:28] + "_dup"
        ws = wb.create_sheet(title=title)

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, line in enumerate(report.lines, start=2):
            ws.cell(row=row_idx, column=1, value=line.account_code)
            ws.cell(row=row_idx, column=2, value=line.item_name)

            c3 = ws.cell(row=row_idx, column=3, value=line.base_value)
            c3.number_format = '#,##0'

            c4 = ws.cell(row=row_idx, column=4, value=line.current_value)
            c4.number_format = '#,##0'

            c5 = ws.cell(row=row_idx, column=5, value=line.variance_absolute)
            c5.number_format = '#,##0'

            c6 = ws.cell(row=row_idx, column=6, value=(line.variance_percent / 100.0) if line.variance_percent is not None else "")
            c6.number_format = '0.00%'

            ws.cell(row=row_idx, column=7, value=_localized_status(line))
            ws.cell(row=row_idx, column=8, value="")

            if line.is_alert:
                fill = red_fill if line.variance_absolute > 0 else yellow_fill
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_num).fill = fill

        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['H'].width = 30

    wb.save(output_path)
