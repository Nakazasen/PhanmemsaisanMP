"""Pipeline audit report for user confidence and missing-input review."""

from __future__ import annotations

import csv
import os
import sqlite3
from pathlib import Path
from typing import Any

from src.parsers.manual_event_drivers import TEMPLATE_FILENAME as EVENT_DRIVER_FILENAME
from src.parsers.manual_headcount import get_required_headcount_periods
from src.utils.excel_helpers import get_fy_months
from src.audit.fixed_assets_coverage import build_fixed_assets_coverage_report
from src.parsers.fixed_assets import find_fixed_assets_file


def _count_rows(conn: sqlite3.Connection, source: str) -> int:
    row = conn.execute("SELECT COUNT(*) FROM fact_input_data WHERE source = ?", (source,)).fetchone()
    return int(row[0] or 0)


def _distinct_cc_count(conn: sqlite3.Connection, source: str) -> int:
    row = conn.execute("SELECT COUNT(DISTINCT cc_code) FROM fact_input_data WHERE source = ?", (source,)).fetchone()
    return int(row[0] or 0)


def _manual_event_rows(source_dir: str) -> int:
    path = Path(source_dir) / EVENT_DRIVER_FILENAME
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return sum(1 for row in reader if any(str(value or "").strip() for value in row.values()))


def _manual_headcount_ccs(conn: sqlite3.Connection) -> set[str]:
    return {
        str(row[0]).strip()
        for row in conn.execute("SELECT DISTINCT cc_code FROM fact_monthly_headcount WHERE source = 'manual'").fetchall()
        if row[0] is not None
    }


def _manual_health_check_gender_ccs(conn: sqlite3.Connection, fiscal_year: int) -> set[str]:
    december_period = f"{fiscal_year - 1}12"
    return {
        str(row[0]).strip()
        for row in conn.execute(
            """
            SELECT DISTINCT cc_code
            FROM fact_monthly_headcount
            WHERE source = 'manual'
              AND period = ?
              AND (headcount_male > 0 OR headcount_female > 0)
            """,
            (december_period,),
        ).fetchall()
        if row[0] is not None
    }


def _headcount_rows_by_period(conn: sqlite3.Connection, cc_code: str) -> set[str]:
    return {
        str(row[0]).strip()
        for row in conn.execute(
            "SELECT DISTINCT period FROM fact_monthly_headcount WHERE CAST(cc_code AS TEXT) = ?",
            (str(cc_code).strip(),),
        ).fetchall()
        if row[0] is not None
    }


def _missing_headcount_series_rows(
    conn: sqlite3.Connection,
    *,
    cc_code: str,
    fiscal_year: int,
) -> list[dict[str, str]]:
    existing_periods = _headcount_rows_by_period(conn, cc_code)
    missing_rows: list[dict[str, str]] = []
    for period in get_required_headcount_periods(fiscal_year):
        if period in existing_periods:
            continue
        for category, category_label in (
            ("headcount_staff", "nhân viên"),
            ("headcount_worker", "công nhân"),
        ):
            missing_rows.append(
                {
                    "severity": "action",
                    "cc_code": cc_code,
                    "period": period,
                    "area": "headcount_series",
                    "message": (
                        "Thiếu dữ liệu số người chuẩn theo tháng: "
                        f"trung tâm chi phí={cc_code}, kỳ={period}, nhóm={category_label}"
                    ),
                    "action": (
                        "Hãy nhập số liệu gốc tháng 03/2026 và số người từng tháng của năm tài chính "
                        "cho cả nhân viên và công nhân trong headcount_manual.csv hoặc trên giao diện."
                    ),
                }
            )
    return missing_rows


def _target_ccs(conn: sqlite3.Connection, target_cc: object | None) -> list[str]:
    if target_cc:
        return [str(target_cc).strip()]
    return [
        str(row[0]).strip()
        for row in conn.execute("SELECT code FROM dim_cost_centers ORDER BY code").fetchall()
        if row[0] is not None
    ]


def _recorded_missing_inputs(conn: sqlite3.Connection, target_cc: object | None = None) -> list[dict[str, str]]:
    conditions = ""
    params: tuple[object, ...] = ()
    if target_cc:
        conditions = "WHERE CAST(cc_code AS TEXT) = ?"
        params = (str(target_cc).strip(),)
    rows = conn.execute(
        f"""
        SELECT severity, cc_code, period, area, message, action
        FROM fact_missing_inputs
        {conditions}
        ORDER BY id
        """,
        params,
    ).fetchall()
    return [
        {
            "severity": str(row["severity"] or "action"),
            "cc_code": str(row["cc_code"] or ""),
            "period": str(row["period"] or ""),
            "area": str(row["area"] or ""),
            "message": str(row["message"] or ""),
            "action": str(row["action"] or ""),
        }
        for row in rows
    ]


def write_pipeline_audit_report(
    *,
    conn: sqlite3.Connection,
    output_dir: str,
    source_dir: str,
    fiscal_year: int,
    target_cc: object | None,
    parser_results: dict[str, dict[str, Any]],
) -> dict[str, str]:
    """Write business-facing Excel audit reports for the completed run.

    The returned ``missing_report_path`` points to ``DU_LIEU_CON_THIEU.xlsx``.
    ``missing_csv_path`` remains as a compatibility alias for older callers.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "BAO_CAO_LAN_CHAY.xlsx"
    missing_report_path = out_dir / "DU_LIEU_CON_THIEU.xlsx"

    fy_months = get_fy_months(fiscal_year)
    manual_hc_ccs = _manual_headcount_ccs(conn)
    health_gender_ccs = _manual_health_check_gender_ccs(conn, fiscal_year)
    cc_list = _target_ccs(conn, target_cc)
    manual_event_count = _manual_event_rows(source_dir)
    fixed_assets_workbook = find_fixed_assets_file(source_dir)
    fixed_assets_coverage = build_fixed_assets_coverage_report(
        conn,
        fixed_assets_workbook,
        source_inspection=parser_results.get("fixed_assets"),
    )

    missing_rows: list[dict[str, str]] = []
    for cc_code in cc_list:
        missing_rows.extend(
            _missing_headcount_series_rows(conn, cc_code=cc_code, fiscal_year=fiscal_year)
        )

        if cc_code not in manual_hc_ccs:
            missing_rows.append(
                {
                    "severity": "review",
                    "cc_code": cc_code,
                    "period": "",
                    "area": "headcount",
                    "message": "Chưa có số người nhập bổ sung cho trung tâm chi phí này; chương trình sẽ dùng dữ liệu số người chính hoặc dữ liệu thay thế nếu có.",
                    "action": "Nếu trung tâm chi phí này cần tính theo số người thực tế từng tháng, hãy nhập vào headcount_manual.csv.",
                }
            )

        if cc_code not in health_gender_ccs:
            missing_rows.append(
                {
                    "severity": "review",
                    "cc_code": cc_code,
                    "period": f"{fiscal_year - 1}12",
                    "area": "health_check_gender_split",
                    "message": "Chưa có dữ liệu nam/nữ tháng 12 cho dòng 57 của phần kiểm tra sức khỏe.",
                    "action": "Nếu trung tâm chi phí này cần tính khám sức khỏe theo nam/nữ, hãy nhập số nam và số nữ tháng 12 trong headcount_manual.csv.",
                }
            )

    if manual_event_count == 0:
        missing_rows.append(
            {
                "severity": "action",
                "cc_code": str(target_cc or "Tất cả"),
                "period": ",".join(fy_months),
                "area": "manual_event_driver",
                "message": "Chưa có dòng sự kiện nhập bổ sung nào cho các số liệu không thể suy luận.",
                "action": (
                    "Nếu có xe đưa đón người Nhật/người Việt, quà cho người không đi du lịch, "
                    "kỷ niệm 10 năm, lễ kỷ niệm thành lập công ty hoặc dòng thị thực/hộ chiếu "
                    "khác dòng 137, hãy nhập vào event_drivers_manual.csv."
                ),
            }
        )

    missing_rows.extend(_recorded_missing_inputs(conn, target_cc))

    source_summary = {
        "manual_event_driver": _count_rows(conn, "manual_event_driver"),
        "nnn_paperwork": _count_rows(conn, "nnn_paperwork"),
        "birthday_workbook": _count_rows(conn, "birthday_workbook"),
        "manual_special_cost": _count_rows(conn, "manual_special_cost"),
        "it_sim": _count_rows(conn, "it_sim"),
        "facility": _count_rows(conn, "facility"),
        "fixed_assets": _count_rows(conn, "fixed_assets"),
    }

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    area_names = {
        "manual_event_driver": "Khoản phát sinh cần nhập",
        "headcount": "Số người",
        "headcount_series": "Số người theo tháng",
        "headcount_event_delta": "Biến động số người",
        "health_check_gender_split": "Số Nam/Nữ tháng 12",
    }
    severity_names = {"action": "Cần bổ sung", "review": "Cần xem lại", "warning": "Cảnh báo"}

    # Gộp các cảnh báo trùng để báo cáo không trở thành ma trận hàng nghìn dòng.
    grouped: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for row in missing_rows:
        key = (row["cc_code"], row["area"], row["message"], row["action"])
        if key not in grouped:
            grouped[key] = dict(row)
        else:
            periods = {p for p in grouped[key]["period"].split(",") if p}
            periods.update(p for p in row["period"].split(",") if p)
            grouped[key]["period"] = ", ".join(sorted(periods))
    concise_rows = list(grouped.values())

    missing_book = Workbook()
    missing_sheet = missing_book.active
    missing_sheet.title = "Dữ liệu còn thiếu"
    missing_sheet.append(["DỮ LIỆU CẦN BỔ SUNG HOẶC XÁC NHẬN"])
    missing_sheet.append(["Năm tài chính", fiscal_year])
    missing_sheet.append(["Phạm vi", str(target_cc) if target_cc else "Tất cả Trung tâm chi phí"])
    missing_sheet.append(["Tổng số vấn đề", len(concise_rows)])
    missing_sheet.append([])
    missing_sheet.append(["Mức độ", "Trung tâm chi phí", "Kỳ liên quan", "Nội dung", "Việc cần làm"])
    for row in concise_rows:
        area_label = area_names.get(
            row["area"],
            str(row["area"] or "Dữ liệu chưa đầy đủ").replace("_", " "),
        )
        detail = str(row["message"] or "").strip()
        content = f"{area_label}: {detail}" if detail else area_label
        missing_sheet.append([
            severity_names.get(row["severity"], "Cần xem lại"), row["cc_code"], row["period"],
            content, row["action"],
        ])
    if not concise_rows:
        missing_sheet.append(["Không có", "", "", "Chưa phát hiện dữ liệu cần bổ sung", "Không cần xử lý"])

    report_book = Workbook()
    report_sheet = report_book.active
    report_sheet.title = "Báo cáo lần chạy"
    report_sheet.append(["BÁO CÁO TÓM TẮT LẦN CHẠY"])
    report_sheet.append(["Năm tài chính", fiscal_year])
    report_sheet.append(["Phạm vi", str(target_cc) if target_cc else "Tất cả Trung tâm chi phí"])
    report_sheet.append(["Thư mục nguồn", source_dir])
    report_sheet.append(["Số vấn đề cần xem", len(concise_rows)])
    report_sheet.append(["Kết luận", "Đã hoàn thành, cần xử lý các mục còn thiếu" if concise_rows else "Đã hoàn thành, chưa phát hiện dữ liệu thiếu"])
    report_sheet.append([])
    report_sheet.append(["Nguồn dữ liệu", "Số dòng đã đọc", "Số Trung tâm chi phí"])
    source_names = {
        "manual_event_driver": "Khoản phát sinh nhập bổ sung", "nnn_paperwork": "Giấy tờ người nước ngoài",
        "birthday_workbook": "Chi phí sinh nhật", "manual_special_cost": "Chi phí đặc biệt nhập bổ sung",
        "it_sim": "Chi phí hệ thống", "facility": "Nhà xưởng, điện và nước", "fixed_assets": "Tài sản cố định",
    }
    for source, count in source_summary.items():
        report_sheet.append([source_names[source], count, _distinct_cc_count(conn, source)])

    for sheet in (missing_sheet, report_sheet):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
        sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        sheet["A1"].alignment = Alignment(horizontal="center")
        header_row = 6 if sheet is missing_sheet else 8
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.freeze_panes = f"A{header_row + 1}"
    for column, width in zip("ABCDE", (18, 22, 30, 42, 70)):
        missing_sheet.column_dimensions[column].width = width
    for column, width in zip("ABC", (42, 20, 24)):
        report_sheet.column_dimensions[column].width = width
    missing_book.save(missing_report_path)
    report_book.save(report_path)
    missing_book.close()
    report_book.close()
    missing_path = str(missing_report_path)
    return {
        "report_path": str(report_path),
        "missing_report_path": missing_path,
        "missing_csv_path": missing_path,
    }
