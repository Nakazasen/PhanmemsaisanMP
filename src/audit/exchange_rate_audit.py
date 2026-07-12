"""Guards for USD/VND formulas written into MP output workbooks."""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.utils.excel_helpers import find_hub_sheet_name, validate_exchange_rate

MONTH_START_COLUMN = 6  # F
MONTH_END_COLUMN = 17  # Q
FIRST_BUSINESS_ROW = 26
KNOWN_LEGACY_RATES = (25450, 26273)

# Only an unqualified B2 in the output hub sheet is the currency-rate cell.
_UNQUALIFIED_B2_RE = re.compile(
    r"(?<![A-Z0-9_!$])\$?B\$?2(?![A-Z0-9_])", re.IGNORECASE
)
_NUMBER_RE = re.compile(r"(?<![A-Z0-9_.])(\d+(?:\.0+)?)(?![A-Z0-9_.])")


def normalize_exchange_rate_formula(value: Any, exchange_rate: Any) -> Any:
    """Normalize copied output formulas to the hub-sheet `$B$2` rate cell.

    This is deliberately used only for formulas written into the output hub
    sheet. It converts legacy direct B2 references and known/historical
    hard-coded USD/VND rates; it never changes source workbooks.
    """
    if not isinstance(value, str) or not value.startswith("="):
        return value

    formula = _UNQUALIFIED_B2_RE.sub("$B$2", value)
    rates = {*KNOWN_LEGACY_RATES, int(round(validate_exchange_rate(exchange_rate)))}
    for rate in sorted(rates, reverse=True):
        formula = re.sub(
            rf"(?<![A-Z0-9_.]){re.escape(str(rate))}(?:\.0+)?(?![A-Z0-9_.])",
            "$B$2",
            formula,
        )
    return formula


def audit_exchange_rate_workbook(workbook_path: str | Path, exchange_rate: Any) -> dict[str, Any]:
    """Return all output-month formulas that use or hard-code an FX rate."""
    rate = validate_exchange_rate(exchange_rate)
    watched_rates = {*KNOWN_LEGACY_RATES, int(round(rate))}
    workbook_file = Path(workbook_path)
    workbook = load_workbook(workbook_file, read_only=True, data_only=False)
    try:
        sheet_name = find_hub_sheet_name(workbook)
        worksheet = workbook[sheet_name]
        try:
            output_b2_rate = validate_exchange_rate(worksheet["B2"].value)
            b2_matches_effective_rate = abs(output_b2_rate - rate) < 0.000001
        except ValueError:
            output_b2_rate = None
            b2_matches_effective_rate = False
        b2_formulas: list[str] = []
        hardcoded_rate_formulas: list[dict[str, str]] = []
        direct_b2_formulas: list[str] = []
        formula_cells = 0

        for row in worksheet.iter_rows(
            min_row=FIRST_BUSINESS_ROW,
            min_col=MONTH_START_COLUMN,
            max_col=MONTH_END_COLUMN,
        ):
            for cell in row:
                formula = cell.value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                formula_cells += 1
                if _UNQUALIFIED_B2_RE.search(formula):
                    b2_formulas.append(cell.coordinate)
                    if "$B$2" not in formula.upper():
                        direct_b2_formulas.append(cell.coordinate)
                matches = {
                    int(float(match.group(1)))
                    for match in _NUMBER_RE.finditer(formula)
                    if int(float(match.group(1))) in watched_rates
                }
                if matches:
                    hardcoded_rate_formulas.append(
                        {"cell": cell.coordinate, "formula": formula, "rates": ",".join(map(str, sorted(matches)))}
                    )

        return {
            "workbook": str(workbook_file),
            "hub_sheet": sheet_name,
            "effective_exchange_rate": rate,
            "output_b2_rate": output_b2_rate,
            "b2_matches_effective_rate": b2_matches_effective_rate,
            "formula_cells_checked": formula_cells,
            "b2_formula_cells": b2_formulas,
            "direct_b2_formula_cells": direct_b2_formulas,
            "hardcoded_rate_formulas": hardcoded_rate_formulas,
            "status": (
                "PASS"
                if b2_matches_effective_rate and not direct_b2_formulas and not hardcoded_rate_formulas
                else "FAIL"
            ),
        }
    finally:
        workbook.close()


def assert_exchange_rate_formulas_safe(workbook_path: str | Path, exchange_rate: Any) -> dict[str, Any]:
    result = audit_exchange_rate_workbook(workbook_path, exchange_rate)
    if result["status"] != "PASS":
        cells = [*result["direct_b2_formula_cells"], *(row["cell"] for row in result["hardcoded_rate_formulas"])]
        if not result["b2_matches_effective_rate"]:
            cells.insert(0, "B2")
        raise ValueError(
            "Phát hiện công thức tỷ giá không đồng nhất trong output: " + ", ".join(cells)
        )
    return result


def write_exchange_rate_audit_report(
    output_path: str | Path,
    exchange_rate: Any,
    exchange_rate_source: str,
    workbook_results: list[dict[str, Any]],
) -> Path:
    """Ghi báo cáo tỷ giá ngắn gọn, dễ mở bằng Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = Path(output_path).with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kiểm tra tỷ giá"
    sheet.append(["KIỂM TRA TỶ GIÁ CỦA TỆP KẾT QUẢ"])
    sheet.append(["Tỷ giá sử dụng", validate_exchange_rate(exchange_rate)])
    sheet.append(["Nguồn tỷ giá", exchange_rate_source])
    sheet.append([])
    sheet.append([
        "Tệp kết quả", "Tỷ giá đã ghi", "Số công thức đã kiểm tra",
        "Kết luận", "Chi tiết cần kiểm tra",
    ])
    overall_ok = True
    for result in workbook_results:
        ok = result.get("status") == "PASS"
        overall_ok = overall_ok and ok
        bad_cells = [*result.get("direct_b2_formula_cells", [])]
        bad_cells.extend(row.get("cell", "") for row in result.get("hardcoded_rate_formulas", []))
        if not result.get("b2_matches_effective_rate", False):
            bad_cells.insert(0, "B2")
        sheet.append([
            Path(str(result.get("workbook", ""))).name,
            result.get("output_b2_rate"),
            result.get("formula_cells_checked", 0),
            "Đạt" if ok else "Cần kiểm tra",
            ", ".join(cell for cell in bad_cells if cell),
        ])
    sheet.insert_rows(4)
    sheet["A4"] = "Kết luận chung"
    sheet["B4"] = "Đạt - tỷ giá được dùng thống nhất" if overall_ok else "Cần kiểm tra - có tỷ giá không thống nhất"
    sheet.merge_cells("A1:E1")
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")
    for cell in sheet[8]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for width, column in zip((34, 18, 24, 20, 42), "ABCDE"):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A9"
    workbook.save(path)
    workbook.close()
    return path
