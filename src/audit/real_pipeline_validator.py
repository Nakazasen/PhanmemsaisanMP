"""Read-only acceptance checks for one real fiscal pipeline run."""
from __future__ import annotations

from dataclasses import asdict, dataclass, field
from hashlib import sha256
import json
from pathlib import Path
import sqlite3
from typing import Any, Mapping

from openpyxl import load_workbook

from src.audit.exchange_rate_audit import assert_exchange_rate_formulas_safe
from src.utils.excel_helpers import find_hub_sheet_name
from src.utils.fiscal_periods import fiscal_baseline_period, fiscal_periods

REQUIRED_AUDIT_WORKBOOKS = (
    "BAO_CAO_LAN_CHAY.xlsx",
    "DU_LIEU_CON_THIEU.xlsx",
    "KIEM_TRA_TY_GIA.xlsx",
)
BASELINE_PROVENANCE = "USER_APPROVED_BASELINE_T3_FROM_T4"
REQUIRED_PIPELINE_STAGES = (
    "preflight",
    "initialize_database",
    "import_sources",
    "validate_staffing",
    "allocation",
    "export_workbooks",
    "audit_reports",
    "publication",
)


@dataclass(frozen=True)
class AcceptanceIssue:
    check: str
    message: str


@dataclass
class RealPipelineAcceptanceResult:
    run_id: str
    fiscal_year: int
    target_cc: str
    passed: bool = False
    checks: dict[str, Any] = field(default_factory=dict)
    issues: list[AcceptanceIssue] = field(default_factory=list)
    business_follow_up: dict[str, Any] = field(default_factory=dict)

    def fail(self, check: str, message: str) -> None:
        self.issues.append(AcceptanceIssue(check, message))

    def as_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["status"] = "PASS" if self.passed else "FAIL"
        return payload


def _readonly_connection(path: Path) -> sqlite3.Connection:
    if not path.is_file():
        raise FileNotFoundError(path)
    connection = sqlite3.connect(
        f"file:{path.resolve().as_posix()}?mode=ro", uri=True
    )
    connection.row_factory = sqlite3.Row
    return connection


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Cần một đối tượng JSON tại: {path}")
    return payload


def _validate_stage_evidence(path: Path, run_id: str) -> dict[str, Any]:
    payload = _load_json(path)
    if payload.get("schema_version") != 1:
        raise ValueError(f"Phiên bản lược đồ bằng chứng giai đoạn không được hỗ trợ: {payload.get('schema_version')!r}")
    if str(payload.get("run_id") or "") != str(run_id):
        raise ValueError(f"Mã lượt chạy trong bằng chứng giai đoạn không khớp: {payload.get('run_id')!r}")
    if payload.get("status") != "SUCCEEDED":
        raise ValueError(f"Bằng chứng giai đoạn chưa kết thúc ở trạng thái SUCCEEDED: {payload.get('status')!r}")
    if payload.get("current_stage") not in (None, ""):
        raise ValueError(f"Bằng chứng giai đoạn vẫn còn giai đoạn đang chạy: {payload.get('current_stage')!r}")
    stages = payload.get("stages")
    if not isinstance(stages, list):
        raise ValueError("Danh sách giai đoạn trong bằng chứng không hợp lệ")
    names = [stage.get("name") for stage in stages if isinstance(stage, dict)]
    if names != list(REQUIRED_PIPELINE_STAGES) or len(stages) != len(names):
        raise ValueError(
            f"Thứ tự giai đoạn không khớp: yêu cầu={list(REQUIRED_PIPELINE_STAGES)}, thực tế={names}"
        )
    for stage in stages:
        if stage.get("status") != "PASS":
            raise ValueError(f"Giai đoạn pipeline chưa đạt PASS: {stage.get('name')!r}")
        elapsed = stage.get("elapsed_seconds")
        if not isinstance(elapsed, (int, float)) or isinstance(elapsed, bool) or elapsed < 0:
            raise ValueError(f"Thời gian chạy giai đoạn không hợp lệ: {stage.get('name')!r}")
    total = payload.get("total_elapsed_seconds")
    if not isinstance(total, (int, float)) or isinstance(total, bool) or total < 0:
        raise ValueError("Tổng thời gian chạy các giai đoạn không hợp lệ")
    return {
        "path": str(path),
        "status": payload["status"],
        "stage_count": len(stages),
        "stages": stages,
        "total_elapsed_seconds": total,
    }


def _catalog_row(catalog: Path, run_id: str) -> dict[str, Any]:
    connection = _readonly_connection(catalog)
    try:
        row = connection.execute(
            "SELECT * FROM planning_runs WHERE run_id = ?", (run_id,)
        ).fetchone()
        if row is None:
            raise ValueError(f"Không có mã lượt chạy trong danh mục: {run_id}")
        return dict(row)
    finally:
        connection.close()


def _table_names(connection: sqlite3.Connection) -> set[str]:
    return {
        str(row[0])
        for row in connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )
    }


def _validate_staffing(
    database: Path,
    *,
    fiscal_year: int,
    target_cc: str,
    expected_provenance: str,
) -> dict[str, Any]:
    connection = _readonly_connection(database)
    try:
        required = {
            "fact_input_data",
            "fact_monthly_headcount",
            "fact_manual_headcount_baseline_override",
        }
        missing = sorted(required - _table_names(connection))
        if missing:
            raise ValueError("Cơ sở dữ liệu lượt chạy đang thiếu bảng: " + ", ".join(missing))

        baseline = fiscal_baseline_period(fiscal_year)
        baseline_rows = connection.execute(
            """SELECT period, description, headcount_all, headcount_expat,
                      headcount_staff, headcount_worker, headcount_local_total
               FROM fact_monthly_headcount
               WHERE period = ? AND CAST(cc_code AS TEXT) = ? AND source = 'manual'""",
            (baseline, target_cc),
        ).fetchall()
        if len(baseline_rows) != 1:
            raise ValueError(
                f"Cần đúng một baseline nhập tay cho {target_cc}/{baseline}; tìm thấy {len(baseline_rows)}"
            )
        baseline_row = baseline_rows[0]
        if str(baseline_row["description"] or "") != expected_provenance:
            raise ValueError(
                f"Nguồn gốc baseline nhập tay chưa được phê duyệt: {baseline_row['description']!r}"
            )
        override = connection.execute(
            """SELECT description FROM fact_manual_headcount_baseline_override
               WHERE fiscal_year = ? AND period = ? AND CAST(cc_code AS TEXT) = ?""",
            (fiscal_year, baseline, target_cc),
        ).fetchall()
        if len(override) != 1 or str(override[0]["description"] or "") != expected_provenance:
            raise ValueError("Không có bản ghi ghi đè baseline đã được phê duyệt trong run.db")

        def assert_balanced(row: sqlite3.Row, label: str) -> None:
            fields = (
                row["headcount_all"], row["headcount_expat"],
                row["headcount_staff"], row["headcount_worker"],
                row["headcount_local_total"],
            )
            if any(value is None for value in fields):
                raise ValueError(f"Dòng nhân sự đang thiếu thành phần: {label}")
            total, expat, staff, worker, local_total = map(float, fields)
            if any(value < 0 for value in fields):
                raise ValueError(f"Dòng nhân sự có thành phần âm: {label}")
            if abs(local_total - staff - worker) > 1e-6:
                raise ValueError(f"Tổng nhân sự địa phương không cân bằng: {label}")
            if abs(total - expat - local_total) > 1e-6:
                raise ValueError(f"Tổng nhân sự không cân bằng: {label}")

        assert_balanced(baseline_row, baseline)
        periods = fiscal_periods(fiscal_year)
        rows = connection.execute(
            """SELECT period, headcount_all, headcount_expat, headcount_staff,
                      headcount_worker, headcount_local_total
               FROM fact_monthly_headcount
               WHERE CAST(cc_code AS TEXT) = ? AND source = 'department_plan'""",
            (target_cc,),
        ).fetchall()
        by_period: dict[str, list[sqlite3.Row]] = {}
        for row in rows:
            by_period.setdefault(str(row["period"]), []).append(row)
        missing_periods = [period for period in periods if period not in by_period]
        duplicate_periods = [period for period in periods if len(by_period.get(period, [])) != 1]
        if missing_periods or duplicate_periods:
            raise ValueError(
                f"Phạm vi dữ liệu nhân sự phòng ban không hợp lệ; thiếu={missing_periods}, không_duy_nhất={duplicate_periods}"
            )
        for period in periods:
            assert_balanced(by_period[period][0], period)

        all_inputs = int(connection.execute("SELECT COUNT(*) FROM fact_input_data").fetchone()[0])
        target_inputs = int(connection.execute(
            "SELECT COUNT(*) FROM fact_input_data WHERE CAST(cc_code AS TEXT) = ?",
            (target_cc,),
        ).fetchone()[0])
        if all_inputs <= 0 or target_inputs <= 0:
            raise ValueError(f"Không có dữ liệu đầu vào thực tế: toàn_bộ={all_inputs}, mục_tiêu={target_inputs}")
        return {
            "baseline_period": baseline,
            "baseline_provenance": expected_provenance,
            "department_plan_periods": periods,
            "input_fact_rows": all_inputs,
            "target_input_fact_rows": target_inputs,
        }
    finally:
        connection.close()


def _validate_workbook(path: Path, exchange_rate: float) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        hub_name = find_hub_sheet_name(workbook)
        hub = workbook[hub_name]
        populated = 0
        formulas = 0
        for row in hub.iter_rows(min_row=26, min_col=6, max_col=17):
            for cell in row:
                if cell.value not in (None, ""):
                    populated += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
        if populated == 0:
            raise ValueError("Sheet trung tâm của tệp kết quả không có ô tháng nghiệp vụ nào chứa dữ liệu")
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    fx = assert_exchange_rate_formulas_safe(path, exchange_rate)
    return {
        "hub_sheet": hub_name,
        "sheet_names": sheet_names,
        "populated_month_cells": populated,
        "formula_cells": formulas,
        "exchange_rate_audit": fx,
        "sha256": _sha256_file(path),
    }


def _validate_audits(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    evidence: dict[str, Any] = {}
    follow_up: dict[str, Any] = {}
    for filename in REQUIRED_AUDIT_WORKBOOKS:
        workbook_path = path / filename
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            evidence[filename] = {
                "sheet_names": list(workbook.sheetnames),
                "sha256": _sha256_file(workbook_path),
            }
            if filename == "DU_LIEU_CON_THIEU.xlsx":
                sheet = workbook.active
                rows = max(int(sheet.max_row or 0) - 6, 0)
                follow_up = {
                    "declared_issue_count": sheet["B4"].value,
                    "detail_rows": rows,
                    "status": "REVIEW_REQUIRED" if rows else "NONE_REPORTED",
                }
        finally:
            workbook.close()
    return evidence, follow_up


def validate_real_pipeline_run(
    *,
    history_root: str | Path,
    fiscal_year: int,
    target_cc: object,
    run_id: str,
    expected_exchange_rate: float,
    public_output_dir: str | Path,
    preserved_public_workbooks: Mapping[str, str] | None = None,
    expected_baseline_provenance: str = BASELINE_PROVENANCE,
) -> RealPipelineAcceptanceResult:
    """Validate one completed real CLI run without mutating run artifacts."""
    fiscal_year = int(fiscal_year)
    target_cc = str(target_cc).strip()
    result = RealPipelineAcceptanceResult(str(run_id), fiscal_year, target_cc)
    history = Path(history_root).resolve()
    workspace = history / f"FY{fiscal_year}" / str(run_id)

    try:
        row = _catalog_row(history / "run_history.db", str(run_id))
        if int(row["fiscal_year"]) != fiscal_year:
            raise ValueError(f"Năm tài chính trong danh mục không khớp: {row['fiscal_year']}")
        if str(row["selected_cost_center"] or "") != target_cc:
            raise ValueError(f"Mã bộ phận mục tiêu trong danh mục không khớp: {row['selected_cost_center']!r}")
        if str(row["status"] or "") != "SUCCEEDED":
            raise ValueError(f"Trạng thái trong danh mục không phải SUCCEEDED: {row['status']!r}")
        if abs(float(row["exchange_rate"]) - float(expected_exchange_rate)) > 1e-6:
            raise ValueError(f"Tỷ giá trong danh mục không khớp: {row['exchange_rate']!r}")
        if row["error_summary"] not in (None, ""):
            raise ValueError(f"Lượt chạy thành công vẫn có tóm tắt lỗi: {row['error_summary']!r}")
        result.checks["catalog"] = row
    except Exception as exc:
        result.fail("catalog", str(exc))

    manifest_path = workspace / "run_manifest.json"
    try:
        manifest = _load_json(manifest_path)
        expected = {
            "run_id": str(run_id),
            "fiscal_year": fiscal_year,
            "baseline_period": fiscal_baseline_period(fiscal_year),
            "exchange_rate": float(expected_exchange_rate),
        }
        mismatches = {
            key: {"expected": value, "actual": manifest.get(key)}
            for key, value in expected.items() if manifest.get(key) != value
        }
        sources = manifest.get("source_checksums")
        if not isinstance(sources, dict) or not sources:
            mismatches["source_checksums"] = "cần ánh xạ có ít nhất một phần tử"
        if not manifest.get("template_checksum") or not manifest.get("manual_input_checksum"):
            mismatches["checksums"] = "cần checksum của tệp mẫu và dữ liệu nhập tay"
        if mismatches:
            raise ValueError(f"Tệp kê khai lượt chạy không khớp: {mismatches}")
        result.checks["manifest"] = {
            "path": str(manifest_path),
            "template_checksum": manifest["template_checksum"],
            "manual_input_checksum": manifest["manual_input_checksum"],
            "source_categories": sorted(sources),
        }
    except Exception as exc:
        result.fail("manifest", str(exc))

    preflight_path = workspace / "reports" / "preflight_report.json"
    try:
        preflight = _load_json(preflight_path)
        if preflight.get("ok") is not True or preflight.get("issues"):
            raise ValueError("Báo cáo kiểm tra trước khi chạy chưa sạch lỗi")
        if int(preflight.get("fiscal_year")) != fiscal_year:
            raise ValueError("Năm tài chính trong báo cáo kiểm tra trước khi chạy không khớp")
        result.checks["preflight"] = {"path": str(preflight_path), "status": "PASS"}
    except Exception as exc:
        result.fail("preflight", str(exc))

    stage_evidence_path = workspace / "reports" / "pipeline_stage_evidence.json"
    try:
        result.checks["pipeline_stages"] = _validate_stage_evidence(
            stage_evidence_path, str(run_id)
        )
    except Exception as exc:
        result.fail("pipeline_stages", str(exc))

    database = workspace / "run.db"
    try:
        result.checks["staffing_and_inputs"] = _validate_staffing(
            database, fiscal_year=fiscal_year, target_cc=target_cc,
            expected_provenance=expected_baseline_provenance,
        )
        result.checks["run_database"] = {"path": str(database), "sha256": _sha256_file(database)}
    except Exception as exc:
        result.fail("run_database", str(exc))

    staged = workspace / "outputs" / f"MP_CC_{target_cc}.xlsx"
    staged_hash: str | None = None
    try:
        evidence = _validate_workbook(staged, expected_exchange_rate)
        staged_hash = str(evidence["sha256"])
        result.checks["workbook"] = {"path": str(staged), **evidence}
    except Exception as exc:
        result.fail("workbook", str(exc))

    try:
        audits, follow_up = _validate_audits(workspace / "outputs" / "BAO_CAO_KIEM_TRA")
        result.checks["audit_workbooks"] = audits
        result.business_follow_up = follow_up
    except Exception as exc:
        result.fail("audit_workbooks", str(exc))

    failure_traceback = workspace / "reports" / "failure_traceback.txt"
    if failure_traceback.exists():
        result.fail("failure_evidence", f"Không gian làm việc thành công vẫn chứa tệp dấu vết lỗi: {failure_traceback}")
    else:
        result.checks["failure_evidence"] = "absent"

    try:
        public_dir = Path(public_output_dir).resolve()
        public_workbook = public_dir / f"MP_CC_{target_cc}.xlsx"
        public_hash = _sha256_file(public_workbook)
        if staged_hash is None or public_hash != staged_hash:
            raise ValueError("Tệp kết quả đã công bố không giống hoàn toàn tệp trong không gian lượt chạy")
        preserved: dict[str, str] = {}
        for filename, expected_hash in (preserved_public_workbooks or {}).items():
            actual_hash = _sha256_file(public_dir / filename)
            if actual_hash != expected_hash:
                raise ValueError(f"Tệp kết quả cần bảo toàn đã thay đổi hoặc bị thiếu: {filename}")
            preserved[filename] = actual_hash
        result.checks["publication"] = {
            "path": str(public_dir), "target_sha256": public_hash,
            "preserved_workbooks": preserved,
        }
    except Exception as exc:
        result.fail("publication", str(exc))

    result.passed = not result.issues
    return result


def write_acceptance_result(result: RealPipelineAcceptanceResult, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(result.as_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return target
