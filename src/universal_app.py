"""Ứng dụng giao diện chính của MP2027 Manager."""

import csv
import hashlib
import os
import queue
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import tkinter as tk
import unicodedata
from datetime import datetime
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl


def _default_fiscal_year(today: datetime | None = None) -> int:
    """Company FY ends in March: Apr-Dec belongs to the following FY."""
    current = today or datetime.now()
    return current.year + 1 if current.month >= 4 else current.year


def _filter_cost_center_choices(choices, query: str) -> list[str]:
    """Filter visible CC choices without mutating the source selection list."""
    normalized_query = "".join(
        character
        for character in unicodedata.normalize("NFKD", str(query or "").casefold())
        if not unicodedata.combining(character)
    ).strip()
    if not normalized_query:
        return list(choices)

    def searchable(choice: object) -> str:
        return "".join(
            character
            for character in unicodedata.normalize("NFKD", str(choice).casefold())
            if not unicodedata.combining(character)
        )

    return [str(choice) for choice in choices if normalized_query in searchable(choice)]

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Runtime business paths are rooted here. In source mode this is the repository;
# packaged startup may switch it to a writable per-user project directory.
BASE_DIR = APP_DIR

if APP_DIR not in sys.path:
    sys.path.append(APP_DIR)


def _copy_missing_tree(source_dir: str, target_dir: str) -> None:
    from src.services.runtime_health import copy_missing_tree

    copy_missing_tree(source_dir, target_dir)


def _directory_is_writable(path: str) -> bool:
    """Return whether *path* can host mutable portable project data."""
    from src.services.runtime_health import directory_is_writable

    return directory_is_writable(path)


def _packaged_project_root(app_dir: str, *, local_app_data: str | None = None) -> str:
    """Choose stable data storage independently from versioned application files."""
    from src.services.runtime_health import packaged_project_root

    return packaged_project_root(
        app_dir,
        local_app_data=local_app_data,
        writable_check=_directory_is_writable,
    )


def _ensure_external_runtime_data(*, local_app_data: str | None = None) -> str:
    """Seed editable bundled data into a stable writable project directory."""
    global BASE_DIR
    from src.services.runtime_health import ensure_external_runtime_data

    BASE_DIR = ensure_external_runtime_data(
        APP_DIR,
        resource_path("."),
        frozen=bool(getattr(sys, "frozen", False)),
        local_app_data=local_app_data,
        writable_check=_directory_is_writable,
    )
    return BASE_DIR


# Bundled data remains immutable seed material. Business data is resolved through
# project.json in BASE_DIR and is never written inside PyInstaller's _internal.

from src.db.loader import load_all, load_cost_centers
from src.db.migrations import CURRENT_SCHEMA_VERSION
from src.db.schema import create_schema, get_connection
from src.services.app_updates import (
    ApplicationUpdateError,
    application_install_root,
    install_runtime_application_update,
    launch_activated_update,
)
from src.services.update_delivery import (
    UpdateDeliveryError,
    check_update_source,
    current_release_version,
    discover_available_update,
    fetch_update_candidate,
    load_update_config,
    validate_update_source,
)
from src.services.content_packs import (
    ContentPackError,
    install_runtime_content_pack,
    load_runtime_content_rules,
)
from src.services.headcount_source_importer import (
    assess_headcount_time_source_coverage,
    cleanup_headcount_truth,
    count_headcount_truth_rows,
    import_headcount_time_sources,
    review_headcount_time_sources,
)
from src.services.manual_staffing_overrides import (
    copy_missing_baselines_from_april,
    find_missing_baseline_ccs,
    save_manual_baseline_override,
    save_manual_g6_to_g5_transitions,
    save_manual_time_overrides,
)
from src.services.fiscal_run import annual_default_paths, create_fiscal_run_context, preflight_fiscal_run
from src.services.preflight_cache import cached_preflight_fiscal_run, get_cached_preflight
from src.services.project_config import (
    ProjectConfig,
    discover_or_create_project,
    remember_last_project,
    remember_ui_language,
    read_ui_language,
)
from src.services.i18n import (
    t,
    translate_for_language,
    get_current_language,
    set_current_language,
    get_supported_languages,
    get_language_name,
    get_language_code,
    translation_key_for_text,
)
from src.services.user_guide_content import (
    get_user_guide_text,
    get_user_guide_search_suggestions,
    USER_GUIDE_VI,
)
from src.services.run_history import filter_runs
from src.services.operations_case_service import assemble_operational_case
from src.services.operations_ai_provider import CagentProviderPolicy, load_cagent_provider_policy_from_env
from src.ui.operations_assistant import OperationsAssistantDialog, OperationsBusinessChatDialog
from src.services.template_confirmation import inspect_form
from src.services.batch_publication import publish_selected_cc_batch
from src.parsers.manual_event_drivers import TEMPLATE_COLUMNS, ensure_manual_event_drivers_template
from src.parsers.manual_headcount import (
    BUS_DRIVER_COLUMNS,
    ensure_manual_bus_headcount_template,
    ensure_manual_headcount_template,
    get_required_headcount_periods,
    parse_manual_headcount,
    resolve_manual_headcount_source_dir,
    validate_manual_headcount_rows,
)
from src.utils.excel_helpers import (
    find_hub_sheet_name,
    get_fy_months,
    read_exchange_rate_from_form,
    validate_exchange_rate,
)
from src.utils.fiscal_periods import fiscal_month_labels
from src.utils.source_manifest import (
    CATEGORY_DISPLAY_NAMES,
    DEFAULT_DESCRIPTIONS,
    MANIFEST_COLUMNS,
    read_source_manifest_inventory_fast,
    write_source_manifest_xlsx,
)


def _headcount_save_error(period: str, field: str, raw_value: str, validation_rule: str, reason: str) -> dict:
    return {
        "period": period,
        "field": field,
        "raw_value": raw_value,
        "validation_rule": validation_rule,
        "reason": reason,
        "csv_row_written": False,
        "db_row_inserted": False,
    }


def _parse_blank_zero_save_int(period: str, field: str, raw_value: str, label: str) -> tuple[str | None, dict | None]:
    text = str(raw_value or "").strip()
    if text == "":
        return "0", None
    if not text.isdecimal():
        return None, _headcount_save_error(period, field, text, "INTEGER_GTE_0", t("hc_err_integer_gte_0", label=label.capitalize()))
    return str(int(text)), None


def _parse_optional_save_int(period: str, field: str, raw_value: str, label: str) -> tuple[str, dict | None]:
    text = str(raw_value or "").strip()
    if text == "":
        return "", None
    if not text.isdecimal():
        return "", _headcount_save_error(
            period,
            field,
            text,
            "INTEGER_GTE_0",
            t("hc_err_integer_gte_0", label=label.capitalize()),
        )
    return str(int(text)), None


def validate_headcount_save_period_rows(periods, month_values, label_by_period=None):
    """Validate GUI headcount inputs for an atomic full-series save."""
    label_by_period = label_by_period or {}
    periods = tuple(periods)
    baseline_period = periods[0] if periods else None
    rows = []
    errors = []
    for period in periods:
        values = month_values.get(period, {})
        label = label_by_period.get(period, period)
        row_error_count = len(errors)

        if period == baseline_period and not any(
            str(values.get(field, "") or "").strip()
            for field in ("expat", "staff", "worker")
        ):
            errors.append(
                _headcount_save_error(
                    period,
                    "baseline_t3",
                    "",
                    "REQUIRED",
                    t("hc_err_baseline_required"),
                )
            )
            continue

        expat, expat_error = _parse_blank_zero_save_int(
            period,
            "headcount_expat",
            values.get("expat", ""),
            t("hc_err_expat_at", label=str(label)),
        )
        staff, staff_error = _parse_blank_zero_save_int(
            period,
            "headcount_staff",
            values.get("staff", ""),
            t("hc_err_staff_at", label=str(label)),
        )
        worker, worker_error = _parse_blank_zero_save_int(
            period,
            "headcount_worker",
            values.get("worker", ""),
            t("hc_err_worker_at", label=str(label)),
        )
        if expat_error:
            errors.append(expat_error)
        if staff_error:
            errors.append(staff_error)
        if worker_error:
            errors.append(worker_error)

        male = ""
        female = ""
        if str(period).endswith("12"):
            male, male_error = _parse_optional_save_int(
                period,
                "headcount_male",
                values.get("male", ""),
                t("hc_err_male_at", label=str(label)),
            )
            female, female_error = _parse_optional_save_int(
                period,
                "headcount_female",
                values.get("female", ""),
                t("hc_err_female_at", label=str(label)),
            )
            if male_error:
                errors.append(male_error)
            if female_error:
                errors.append(female_error)

        if len(errors) != row_error_count:
            continue

        expat_int = int(expat or "0")
        staff_int = int(staff or "0")
        worker_int = int(worker or "0")
        male_int = int(male or "0")
        female_int = int(female or "0")
        if male_int + female_int > expat_int + staff_int + worker_int:
            errors.append(
                _headcount_save_error(
                    period,
                    "headcount_male/headcount_female",
                    f"{values.get('male', '')}/{values.get('female', '')}",
                    "SUM_LE_TOTAL",
                    t("hc_err_gender_sum_exceeded", label=label),
                )
            )
            continue

        rows.append(
            {
                "period": period,
                "headcount_staff": staff,
                "headcount_worker": worker,
                "headcount_male": male,
                "headcount_female": female,
                "description": str(values.get("description", "") or "").strip(),
            }
        )
    return rows, errors


def format_headcount_save_errors(errors) -> str:
    field_labels = {
        "baseline_t3": t("hc_field_baseline_t3"),
        "headcount_staff": t("hc_field_headcount_staff"),
        "headcount_worker": t("hc_field_headcount_worker"),
        "headcount_male": t("hc_field_headcount_male"),
        "headcount_female": t("hc_field_headcount_female"),
        "headcount_male/headcount_female": t("hc_field_gender_split"),
        "cc_code": t("hc_field_cc_code"),
        "bus_expat_count": t("hc_field_bus_expat"),
        "bus_vietnamese_count": t("hc_field_bus_vn"),
    }
    rule_labels = {
        "REQUIRED": t("hc_rule_required"),
        "INTEGER_GTE_0": t("hc_rule_integer_gte_0"),
        "SUM_LE_TOTAL": t("hc_rule_sum_le_total"),
        "VALID_CC": t("hc_rule_valid_cc"),
        "UNIQUE_CC": t("hc_rule_unique_cc"),
    }
    period_labels = {"bus": t("hc_period_bus")}
    lines = []
    for error in errors:
        raw_period = str(error.get("period", "") or "-")
        period = period_labels.get(raw_period, raw_period)
        field = field_labels.get(
            str(error.get("field", "") or "-"),
            str(error.get("field", "") or "-").replace("_", " "),
        )
        raw_value = str(error.get("raw_value", ""))
        raw_display = t("hc_val_empty") if raw_value == "" else raw_value
        raw_rule = str(error.get("validation_rule", "") or "-")
        rule = rule_labels.get(raw_rule, raw_rule.replace("_", " "))
        reason = str(error.get("reason", "") or "-")
        csv_written = t("hc_val_yes") if error.get("csv_row_written", False) else t("hc_val_no")
        db_inserted = t("hc_val_yes") if error.get("db_row_inserted", False) else t("hc_val_no")
        lines.append(
            f"{period} | {field} | {raw_display} | {rule} | {reason} | {csv_written} | {db_inserted}"
        )
    return "\n".join(lines)


def validate_bus_headcount_save_rows(rows, valid_cc_codes) -> list[dict]:
    errors = []
    seen_cc = set()
    for row_number, row in enumerate(rows, start=2):
        cc_code = str(row.get("cc_code", "") or "").strip()
        expat_count = str(row.get("bus_expat_count", "") or "").strip()
        vietnamese_count = str(row.get("bus_vietnamese_count", "") or "").strip()
        description = str(row.get("description", "") or "").strip()
        if not any([cc_code, expat_count, vietnamese_count, description]):
            continue
        if not cc_code or cc_code not in valid_cc_codes:
            error = _headcount_save_error("bus", "cc_code", cc_code, "VALID_CC", t("hc_err_bus_invalid_cc"))
            error["csv_row"] = row_number
            errors.append(error)
            continue
        if cc_code in seen_cc:
            error = _headcount_save_error("bus", "cc_code", cc_code, "UNIQUE_CC", t("hc_err_bus_duplicate_cc"))
            error["csv_row"] = row_number
            errors.append(error)
            continue
        if not expat_count.isdecimal():
            error = _headcount_save_error(
                "bus",
                "bus_expat_count",
                expat_count,
                "INTEGER_GTE_0",
                t("hc_err_bus_expat_int"),
            )
            error["csv_row"] = row_number
            errors.append(error)
            continue
        if not vietnamese_count.isdecimal():
            error = _headcount_save_error(
                "bus",
                "bus_vietnamese_count",
                vietnamese_count,
                "INTEGER_GTE_0",
                t("hc_err_bus_vn_int"),
            )
            error["csv_row"] = row_number
            errors.append(error)
            continue
        seen_cc.add(cc_code)
    return errors


def _annual_template_path(fiscal_year: int) -> str:
    return os.path.join(BASE_DIR, "docs", f"MP{int(fiscal_year)}", "FORM.xlsx")


def _annual_source_dir(fiscal_year: int) -> str:
    return os.path.join(BASE_DIR, "docs", f"MP{int(fiscal_year)}")


def _annual_headcount_source_dir(fiscal_year: int) -> str:
    annual_dir = os.path.join(BASE_DIR, "raw", f"FY{int(fiscal_year)}")
    if int(fiscal_year) == 2027:
        has_workbook = os.path.isdir(annual_dir) and any(
            name.lower().endswith((".xls", ".xlsx", ".xlsm"))
            for name in os.listdir(annual_dir)
        )
        if not has_workbook:
            return os.path.join(BASE_DIR, "raw")
    return annual_dir


def _annual_manual_input_store(fiscal_year: int) -> str:
    """Editable staffing overrides are isolated by FY, never shared mp2027.db."""
    path = annual_default_paths(int(fiscal_year), BASE_DIR)["manual_input_store"]
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path


def _format_manual_special_legacy_starts(starts: dict[str, int]) -> str:
    return ", ".join(f"{cc}:{row}" for cc, row in sorted(starts.items()))


def _parse_manual_special_legacy_starts(text: str) -> dict[str, int]:
    starts: dict[str, int] = {}
    for fragment in str(text or "").split(","):
        item = fragment.strip()
        if not item:
            continue
        if ":" not in item:
            raise ValueError(t("manual_special_legacy_start_invalid", value=item))
        cc_code, raw_row = (part.strip() for part in item.split(":", 1))
        try:
            row = int(raw_row)
        except ValueError as exc:
            raise ValueError(t("manual_special_legacy_start_invalid", value=item)) from exc
        if not cc_code or row < 1:
            raise ValueError(t("manual_special_legacy_start_invalid", value=item))
        starts[cc_code] = row
    return starts


def _default_template_path(fiscal_year: int = 2027) -> str:
    external_template = _annual_template_path(fiscal_year)
    if os.path.exists(external_template):
        return external_template
    if int(fiscal_year) == 2027:
        packaged_template = resource_path(os.path.join("docs", "MP2027", "FORM.xlsx"))
        if os.path.exists(packaged_template):
            return packaged_template
    return external_template


FORM_SYSTEM_ACCOUNT_CODES = {5005246282, 6005146628, 6005146542}


def _external_template_path(fiscal_year: int = 2027) -> str:
    return _annual_template_path(fiscal_year)


def _external_source_dir(fiscal_year: int = 2027) -> str:
    return _annual_source_dir(fiscal_year)


def _is_under_internal(path: str) -> bool:
    if not path:
        return False
    try:
        absolute_path = os.path.abspath(path)
        internal_dir = os.path.abspath(os.path.join(BASE_DIR, "_internal"))
        return os.path.commonpath([absolute_path, internal_dir]) == internal_dir
    except (OSError, ValueError):
        return False


def _validate_selected_template(path: str, fiscal_year: int = 2027) -> str | None:
    external_template = _external_template_path(fiscal_year)
    if not os.path.isfile(path):
        return t("err_template_not_found", template=external_template)
    if _is_legacy_root_template(path):
        return t("err_template_legacy_root", template=external_template)
    if _is_under_internal(path) and os.path.exists(external_template):
        return t("err_template_in_internal", template=external_template)
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return t("err_template_open_failed", template=external_template)
    try:
        try:
            sheet_name = find_hub_sheet_name(workbook)
        except Exception:
            return t("err_template_no_hub_sheet", template=external_template)
    finally:
        workbook.close()
    return None

def _validate_selected_source_dir(path: str, fiscal_year: int = 2027) -> str | None:
    external_source = _external_source_dir(fiscal_year)
    if not os.path.isdir(path):
        return t("err_source_dir_not_found", source=external_source)
    if _is_under_internal(path) and os.path.isdir(external_source):
        return t("err_source_dir_in_internal", source=external_source)
    return None


def _is_missing_baseline_error(error) -> bool:
    """Return True for both current preflight and legacy baseline error formats."""
    text = str(error or "").lower()
    return (
        "chưa có tổng số người tháng" in text
        or ("thiếu dữ liệu nhân sự/thời gian bắt buộc" in text and "baseline t3" in text)
    )


def _pipeline_failure_summary(output_lines: list[str], return_code: int) -> str:
    """Keep full subprocess output in the log while returning its final error block."""
    error_start = None
    for index in range(len(output_lines) - 1, -1, -1):
        if output_lines[index].strip().startswith("LỖI:"):
            error_start = index
            break
    if error_start is not None:
        error_lines = [
            line.strip()
            for line in output_lines[error_start:]
            if line.strip() and "chi tiết kỹ thuật đã được ẩn" not in line.lower()
        ]
        if error_lines:
            error_lines[0] = error_lines[0].removeprefix("LỖI:").strip()
            return "\n".join(error_lines)
    for line in reversed(output_lines):
        text = line.strip()
        if text and "chi tiết kỹ thuật đã được ẩn" not in text.lower():
            return text
    return t("pipeline_terminated_log", code=return_code)


def _headcount_coverage_error_message(fiscal_year: int, coverage: dict) -> str:
    """Make a selected-CC staffing coverage gap clear to a non-technical user."""
    missing = ", ".join(coverage.get("missing_cc_codes", ())) or t("hc_val_unknown")
    available_codes = tuple(coverage.get("available_cc_codes", ()))
    available = ", ".join(available_codes) if available_codes else t("hc_val_none")
    return t(
        "err_headcount_coverage",
        fiscal_year=fiscal_year,
        missing=missing,
        available=available,
    )


def _failed_run_database_from_output(
    output_lines: list[str], history_root: str, fiscal_year: int
) -> str | None:
    """Resolve the exact immutable run DB named by the subprocess traceback."""
    marker = "Chi tiết lỗi đã lưu:"
    expected_root = os.path.realpath(os.path.join(history_root, f"FY{int(fiscal_year)}"))
    for line in reversed(output_lines):
        if marker not in line:
            continue
        trace_path = line.split(marker, 1)[1].strip()
        candidate = os.path.realpath(
            os.path.join(os.path.dirname(os.path.dirname(trace_path)), "run.db")
        )
        try:
            inside_expected_root = os.path.commonpath([expected_root, candidate]) == expected_root
        except ValueError:
            inside_expected_root = False
        if inside_expected_root and os.path.isfile(candidate):
            return candidate
    return None


def _friendly_error_message(error) -> str:
    text = str(error or "").strip()
    lower_text = text.lower()
    external_template = _external_template_path()

    if "chưa có tổng số người tháng" in lower_text or "missing total headcount for month" in lower_text:
        issue_lines = [
            line[2:].strip()
            for line in text.splitlines()
            if line.strip().startswith("- ") and ("chưa có tổng số người tháng" in line.lower() or "missing total headcount" in line.lower())
        ]
        details = "\n".join(f"• {line}" for line in issue_lines)
        if not details:
            details = text
        return t("err_missing_headcount_for_report", details=details)

    if "unable to locate system cost row" in lower_text or "không tìm thấy dòng system cost" in lower_text or "chi phí hệ thống" in lower_text:
        return t("err_system_cost_row_not_found", template=external_template)

    if "unable to resolve kdc system cost account" in lower_text or "không xác định được tài khoản system cost" in lower_text:
        return t("err_system_cost_account_unresolved")

    if "form template not found" in lower_text or "không tìm thấy tệp mẫu form" in lower_text:
        return t("err_form_template_not_found", template=external_template)

    if "missing the mp detail sheet" in lower_text or "không có sheet chi tiết mp" in lower_text or "không có trang tính chi tiết mp" in lower_text:
        return t("err_missing_mp_detail_sheet", template=external_template)

    if "malformed or empty" in lower_text or "sai định dạng hoặc rỗng" in lower_text:
        return t("err_form_malformed_or_empty", template=external_template)

    if "append rows prepared" in lower_text or "dòng trống để ghi thêm" in lower_text or "không còn đủ dòng trống" in lower_text:
        return t("err_form_no_append_rows")

    if "permission denied" in lower_text or "access is denied" in lower_text or "đang được mở" in lower_text or "bị hạn chế quyền truy cập" in lower_text:
        return t("err_permission_denied")

    if "database is locked" in lower_text or "database locked" in lower_text or ("sqlite" in lower_text and "locked" in lower_text) or "đang bị khóa" in lower_text:
        return t("err_database_locked")

    if "fixed-assets" in lower_text or "tài sản cố định" in lower_text:
        if "missing authoritative exchange_rate" in lower_text or "tỷ giá" in lower_text or "thiếu tỷ giá" in lower_text:
            return t("err_fa_missing_rate")
        if "no recognizable current source sheet" in lower_text or "không chứa trang tính nguồn" in lower_text:
            return t("err_fa_no_sheet")

    if "must not overwrite" in lower_text or "không được ghi đè" in lower_text or "trùng với tệp dữ liệu nguồn" in lower_text:
        return t("err_must_not_overwrite")

    if "not found" in lower_text or "no such file" in lower_text or "không tìm thấy tệp hoặc thư mục" in lower_text:
        return t("err_file_or_dir_not_found")

    action_keywords = ("cách xử lý:", "hướng dẫn:", "対処方法:", "action:")
    if any(k in lower_text for k in action_keywords):
        return text

    if text:
        custom_action = t("err_custom_action")
        return f"{text}\n\n{custom_action}"

    return t("err_general_prefix", text=text) + "\n\n" + t("err_general_action")


def _uniform_policy_warning(issue) -> str | None:
    """Render the hidden uniform-policy dependency as a user action, not raw diagnostics."""
    category = str(getattr(issue, "category", ""))
    if category not in {"uniform_policy", "form_uniform_master"}:
        return None

    filename = os.path.basename(str(getattr(issue, "path", "") or "")) or t("preflight_no_file")
    reason = str(getattr(issue, "reason", "") or "")
    code = str(getattr(issue, "code", "") or "")
    normalized_reason = reason.casefold()

    if category == "form_uniform_master" and "cột bị trùng" in normalized_reason:
        duplicate_headers = reason.partition(":")[2].strip() or reason
        return t(
            "preflight_uniform_duplicate_warning",
            filename=filename,
            headers=duplicate_headers,
        )

    if code == "MISSING_SOURCE" or not str(getattr(issue, "path", "") or "").strip():
        return t("preflight_uniform_not_selected_warning")
    if "không tồn tại" in normalized_reason or "not found" in normalized_reason:
        return t("preflight_uniform_file_missing_warning", filename=filename)
    if "thiếu cột policy" in normalized_reason or "missing policy column" in normalized_reason:
        missing_columns = reason.partition(":")[2].strip() or reason
        return t(
            "preflight_uniform_missing_columns_warning",
            filename=filename,
            columns=missing_columns,
        )
    if "thiếu sheet" in normalized_reason or "missing sheet" in normalized_reason:
        return t("preflight_uniform_layout_warning", filename=filename)
    return t("preflight_uniform_incompatible_warning", filename=filename, details=reason)


def _localized_preflight_issue_warning(issue) -> str:
    """Return a fully localized, actionable warning for the incomplete-run dialog."""
    uniform_warning = _uniform_policy_warning(issue)
    if uniform_warning:
        return uniform_warning

    label = CATEGORY_DISPLAY_NAMES.get(issue.category, issue.category)
    filename = os.path.basename(issue.path) if issue.path else t("preflight_no_file")
    return t(
        "preflight_issue_impact_format",
        label=label,
        filename=filename,
        reason=issue.reason,
        impact=issue.impact,
    )


def _uniform_policy_signature(path: str | None) -> tuple[str, int | None, int | None]:
    """Detect a policy replacement even though that path is configured outside the main form."""
    if not path:
        return ("", None, None)
    normalized = os.path.abspath(path)
    try:
        stat = os.stat(normalized)
    except OSError:
        return (normalized, None, None)
    return (normalized, int(stat.st_mtime_ns), int(stat.st_size))


def _is_legacy_root_template(path: str) -> bool:
    selected_path = os.path.abspath(path)
    root_form = os.path.abspath(os.path.join(BASE_DIR, "FORM.xlsx"))
    if selected_path != root_form:
        return False

    canonical_candidates = [
        os.path.abspath(os.path.join(BASE_DIR, "docs", "MP2027", "FORM.xlsx")),
        os.path.abspath(resource_path(os.path.join("docs", "MP2027", "FORM.xlsx"))),
    ]
    canonical_form = next((candidate for candidate in canonical_candidates if os.path.exists(candidate)), None)
    if not os.path.exists(root_form):
        return True
    if canonical_form is None:
        return False

    def _sha256(file_path: str) -> str:
        digest = hashlib.sha256()
        with open(file_path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    return _sha256(root_form) != _sha256(canonical_form)


def _default_source_dir(fiscal_year: int = 2027) -> str:
    external_source = _annual_source_dir(fiscal_year)
    if os.path.isdir(external_source) or int(fiscal_year) != 2027:
        return external_source
    packaged_source = resource_path(os.path.join("docs", "MP2027"))
    return packaged_source if os.path.isdir(packaged_source) else external_source

USER_GUIDE_TEXT = USER_GUIDE_VI
USER_GUIDE_TEXT_LATEST = USER_GUIDE_VI


def _normalize_guide_search(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").casefold())
    return " ".join("".join(ch for ch in text if not unicodedata.combining(ch)).split())


def filter_user_guide_text(text: str, query: str) -> tuple[str, int]:
    """Return compact matching guide paragraphs using accent-insensitive terms."""
    terms = _normalize_guide_search(query).split()
    if not terms:
        return text, 0
    blocks = [block.strip() for block in re.split(r"\n\s*\n", text) if block.strip()]
    matches = [
        block for block in blocks
        if all(term in _normalize_guide_search(block) for term in terms)
    ]
    if not matches:
        examples = t("user_guide_search_examples")
        return (
            t("user_guide_search_not_found", examples=examples),
            0,
        )
    return t("user_guide_search_results_header") + "\n\n" + "\n\n".join(matches), len(matches)


class MPManagerApp:
    TITLE_OWNER = "Bùi Đức Vinh - Phòng Phát triển hệ thống Chế tạo"

    @staticmethod
    def _application_version() -> str:
        """Return the release version for a user-visible GUI label."""
        try:
            return current_release_version()
        except Exception:
            return t("app_version_unknown")

    @classmethod
    def _window_title(cls, fiscal_year: str, version: str) -> str:
        owner = t("owner_title")
        return t("window_title", fiscal_year=fiscal_year, version=version, owner=owner)

    @staticmethod
    def _initial_window_size(screen_width: int, screen_height: int) -> tuple[int, int]:
        """Choose a usable initial size without extending beyond smaller screens."""
        return (
            max(480, min(1180, int(screen_width) - 48)),
            max(360, min(800, int(screen_height) - 96)),
        )

    def __init__(self, root: tk.Tk):
        self.root = root
        try:
            from src.ui.operations_assistant import apply_modern_window_style
            apply_modern_window_style(self.root)
        except Exception:
            pass
        initial_lang = read_ui_language()
        set_current_language(initial_lang)
        self.language_var = tk.StringVar(value=get_language_name(initial_lang))
        width, height = self._initial_window_size(root.winfo_screenwidth(), root.winfo_screenheight())
        self.root.geometry(f"{width}x{height}")
        self.root.minsize(480, 360)
        self.application_version = self._application_version()
        initial_fiscal_year = _default_fiscal_year()
        self.project, project_created = discover_or_create_project(BASE_DIR, initial_fiscal_year)
        if self.project.ensure_fiscal_year(initial_fiscal_year):
            self.project.save()
        initial_paths = self.project.fiscal_paths(initial_fiscal_year)
        self.project_file = tk.StringVar(value=self.project.config_path)
        template_path = initial_paths.template_path
        self.fiscal_year = tk.StringVar(value=str(initial_fiscal_year))
        self.exchange_rate = tk.StringVar(value=self._initial_exchange_rate(template_path))
        self.cc_code_filter = tk.StringVar(value=t("all_cc_filter_label"))
        self._available_cc_choices: list[str] = []
        self._selected_cc_values: list[str] = []
        self.template_path = tk.StringVar(value=template_path)
        self.source_dir = tk.StringVar(value=initial_paths.source_dir)
        self.headcount_source_dir = tk.StringVar(value=initial_paths.headcount_source_dir)
        self._auto_path_fiscal_year = initial_fiscal_year
        # Keep semantic state for dynamic labels so a language change renders
        # the current state instead of retaining text from the old language.
        self._headcount_source_status_state = {"kind": "database"}
        self._preflight_status_state = {"kind": "key", "key": "preflight_untested_label", "params": {}}
        self.headcount_source_status = tk.StringVar(value=self._initial_headcount_source_status())
        self.preflight_status = tk.StringVar(value=t("preflight_untested_label"))
        self._preflight_token = 0
        self._approved_preflight_signature = None
        self._approved_preflight_report = None
        self._approved_uniform_policy_path = None
        self.last_excel_mtime = 0.0
        self.syncing_master = False
        self._pipeline_busy = False
        self._output_cost_row_order_editor = None
        self._legacy_headcount_editor = None
        self._event_driver_editor = None
        self._variance_editor = None
        self._user_guide_window = None
        self._run_history_window = None
        self.ui_thread_id = threading.get_ident()
        self.ui_queue = queue.Queue()
        # C-AGENT deployment policy: loaded from env vars at startup (fail-closed).
        # If env vars are absent or invalid, cagent_policy.enabled == False.
        # No network calls are made during loading.
        self.cagent_policy: CagentProviderPolicy = load_cagent_provider_policy_from_env()

        self.fiscal_year.trace_add("write", self._on_staffing_selection_changed)
        self.template_path.trace_add("write", self._on_source_selection_changed)
        self.source_dir.trace_add("write", self._on_source_selection_changed)
        self.headcount_source_dir.trace_add("write", self._on_source_selection_changed)
        self.setup_styles()
        self.setup_ui()
        self.preflight_status.trace_add("write", self._on_workflow_state_changed)
        self._refresh_fiscal_year_labels()
        self.set_icon()
        self.root.after(0, lambda: self.log(
             (t("project_profile_created", path=self.project.config_path) if project_created
              else t("project_profile_using", path=self.project.config_path))
        ))
        self.root.after(0, self._update_workflow_guide)
        self._startup_update_prompted = False
        self.root.after(50, self._drain_ui_queue)
        self.root.after(300, self.load_cc_list)
        self.root.after(500, self._mark_preflight_stale)
        self.root.after(1200, self._start_update_discovery)

    def _on_staffing_selection_changed(self, *_args):
        self._refresh_fiscal_year_labels()
        try:
            fiscal_year = int(self.fiscal_year.get())
        except ValueError:
            return
        created = self.project.ensure_fiscal_year(fiscal_year)
        paths = self.project.fiscal_paths(fiscal_year)
        if created:
            self.project.save()
        self.template_path.set(paths.template_path)
        self.source_dir.set(paths.source_dir)
        self.headcount_source_dir.set(paths.headcount_source_dir)
        self.exchange_rate.set(self._initial_exchange_rate(paths.template_path))
        self._auto_path_fiscal_year = fiscal_year
        if hasattr(self, "headcount_source_status"):
            self._set_headcount_source_status_from_database()
        self._mark_preflight_stale()

    def _on_source_selection_changed(self, *_args):
        self._mark_preflight_stale()

    def _mark_preflight_stale(self, force_refresh: bool = False):
        """Disable calculation immediately when the selected source changes."""
        if not hasattr(self, "start_btn"):
            return
        self._preflight_token += 1
        self._approved_preflight_signature = None
        self._approved_preflight_report = None
        self._approved_uniform_policy_path = None
        self._accepted_missing_categories = ()
        token = self._preflight_token
        self.start_btn.configure(state=tk.DISABLED)
        self._set_preflight_status(
            "preflight_refreshing_all" if force_refresh else "preflight_checking_selected"
        )
        self.root.after(350, lambda: self._start_preflight_check(token, force_refresh=force_refresh))

    def _start_preflight_check(self, token: int, *, force_refresh: bool = False):
        if token != self._preflight_token:
            return
        try:
            fiscal_year = int(self.fiscal_year.get())
            template = self.template_path.get().strip()
            source = self.source_dir.get().strip()
            headcount = self.headcount_source_dir.get().strip()
            exchange_rate = validate_exchange_rate(self.exchange_rate.get())
            selected_ccs = self._parse_selected_cc_codes()
        except Exception as exc:
            self._set_preflight_error_status(exc)
            return

        def worker():
            started_at = time.perf_counter()
            try:
                self._run_on_ui_thread(
                    self._set_preflight_status,
                    "preflight_scanning_form" if force_refresh else "preflight_checking_metadata",
                )
                paths = self._project_paths(fiscal_year)
                context = create_fiscal_run_context(
                    fiscal_year,
                    template_path=template,
                    source_dir=source,
                    headcount_source_dir=headcount,
                    uniform_policy_path=paths.uniform_policy_path,
                    output_dir=paths.output_dir,
                    exchange_rate=exchange_rate,
                    exchange_rate_source="FORM!B2 / người dùng xác nhận trên giao diện",
                    history_root=paths.history_root,
                    manual_input_store=paths.manual_input_store,
                    manual_special_inheritance_dir=paths.manual_special_inheritance_dir,
                    manual_special_legacy_starts=paths.manual_special_legacy_starts,
                    base_dir=self.project.root_dir,
                )
                checker = lambda active_context: preflight_fiscal_run(
                    active_context,
                    progress=lambda message: self._run_on_ui_thread(
                        self._update_preflight_progress,
                        token,
                        message,
                    ),
                )
                if force_refresh:
                    report, cache_hit = cached_preflight_fiscal_run(
                        context,
                        force_refresh=True,
                        extra_paths=(self.project.config_path,),
                        checker=checker,
                    )
                else:
                    report = get_cached_preflight(
                        context,
                        extra_paths=(self.project.config_path,),
                    )
                    cache_hit = report is not None
                    if report is None:
                        # A cache miss is not a source failure. Compute a fresh
                        # report instead of leaving the user with a disabled Start button.
                        report, cache_hit = cached_preflight_fiscal_run(
                            context,
                            force_refresh=True,
                            extra_paths=(self.project.config_path,),
                            checker=checker,
                        )
                coverage = self._selected_headcount_source_coverage(
                    fiscal_year,
                    headcount,
                    selected_ccs,
                )
                if coverage["missing_cc_codes"]:
                    summary = _headcount_coverage_error_message(fiscal_year, coverage)
                    self._run_on_ui_thread(
                        self._finish_preflight_check,
                        token,
                        False,
                        summary,
                        None,
                        None,
                        None,
                        False,
                        time.perf_counter() - started_at,
                    )
                    return
                summary = self._localized_preflight_summary(report)

                signature = (
                    fiscal_year,
                    os.path.abspath(template),
                    os.path.abspath(source),
                    os.path.abspath(headcount),
                    float(exchange_rate),
                    _uniform_policy_signature(context.uniform_policy_path),
                )
                self._run_on_ui_thread(
                    self._finish_preflight_check,
                    token,
                    report.ok,
                    summary,
                    signature,
                    report,
                    context,
                    cache_hit,
                    time.perf_counter() - started_at,
                )
            except Exception as exc:
                self._run_on_ui_thread(
                    self._finish_preflight_check,
                    token,
                    False,
                    _friendly_error_message(exc),
                    None,
                    None,
                    None,
                    False,
                    time.perf_counter() - started_at,
                )

        threading.Thread(target=worker, daemon=True).start()

    def _selected_headcount_source_coverage(
        self,
        fiscal_year: int,
        source_dir: str,
        selected_ccs: tuple[str, ...],
    ) -> dict:
        """Check the actual CC inside staffing files, not merely file count."""
        conn = None
        try:
            conn = get_connection(self._operational_database())
            create_schema(conn)
            return assess_headcount_time_source_coverage(
                conn,
                source_dir,
                fiscal_year,
                selected_ccs,
            )
        finally:
            if conn is not None:
                conn.close()

    def _update_preflight_progress(self, token: int, message: str) -> None:
        if token == self._preflight_token:
            # Progress from the worker is legacy free text. Keep the visible
            # status in the active language instead of retaining that text.
            self._set_preflight_status("preflight_checking_metadata")

    def _localized_preflight_summary(self, report) -> str:
        if report.ok:
            return t("preflight_summary_ready")
        if report.can_run:
            issue_lines = []
            for issue in report.skipped_issues:
                issue_lines.append(_localized_preflight_issue_warning(issue))
            return t("preflight_summary_incomplete") + " | ".join(issue_lines)
        issue_lines = []
        for issue in report.blocking_issues:
            label = CATEGORY_DISPLAY_NAMES.get(issue.category, issue.category)
            filename = os.path.basename(issue.path) if issue.path else t("preflight_no_file")
            issue_lines.append(
                t("preflight_issue_action_format", label=label, filename=filename, reason=issue.reason, action=issue.action)
            )
        return " | ".join(issue_lines) or t("preflight_summary_unmet")

    def _set_preflight_status(self, key: str, **params) -> None:
        self._preflight_status_state = {"kind": "key", "key": key, "params": dict(params)}
        self.preflight_status.set(t(key, **params))

    def _set_preflight_error_status(self, error) -> None:
        error_key = translation_key_for_text(
            str(error),
            keys=("err_select_at_least_one_cc",),
        )
        if error_key:
            self._preflight_status_state = {
                "kind": "error_key",
                "key": "preflight_cannot_check",
                "error_key": error_key,
            }
            self.preflight_status.set(t("preflight_cannot_check", error=t(error_key)))
            return
        self._set_preflight_status("preflight_cannot_check", error=_friendly_error_message(error))

    def _set_preflight_completed_status(self, ok: bool, summary: str, report, cache_hit: bool, elapsed_seconds: float) -> None:
        self._preflight_status_state = {
            "kind": "completed",
            "ok": ok,
            "summary": summary,
            "report": report,
            "cache_hit": cache_hit,
            "elapsed_seconds": elapsed_seconds,
        }
        self._refresh_preflight_status()

    def _manual_editor_selected_cc(self) -> str | None:
        """Return the single selected CC code for the manual editor, or None if multiple/invalid."""
        try:
            selected_ccs = tuple(self._parse_selected_cc_codes())
            if len(selected_ccs) == 1:
                return str(selected_ccs[0]).strip()
        except Exception:
            pass
        return None

    def _refresh_preflight_status(self) -> None:
        state = getattr(self, "_preflight_status_state", {"kind": "key", "key": "preflight_untested_label", "params": {}})
        if state["kind"] == "error_key":
            self.preflight_status.set(t(state["key"], error=t(state["error_key"])))
            return
        if state["kind"] == "completed":
            report = state.get("report")
            summary = self._localized_preflight_summary(report) if report is not None else state["summary"]
            if bool(getattr(report, "can_run", False)):
                pending_baselines = getattr(self, "_pending_baseline_ccs", ())
                if pending_baselines:
                    prefix = t("preflight_baseline_required", cc=", ".join(pending_baselines))
                elif getattr(report, "skipped_issues", ()):
                    prefix = t("preflight_warning_partial")
                else:
                    prefix = t("preflight_ready")
                mode = t("preflight_mode_cache") if state["cache_hit"] else t("preflight_mode_scan")
                self.preflight_status.set(t("preflight_done", prefix=prefix, mode=mode, elapsed=state["elapsed_seconds"]))
            else:
                self.preflight_status.set(t("preflight_blocked", summary=summary, elapsed=state["elapsed_seconds"]))
            return
        self.preflight_status.set(t(state["key"], **state.get("params", {})))

    def _finish_preflight_check(
        self,
        token: int,
        ok: bool,
        summary: str,
        signature=None,
        report=None,
        context=None,
        cache_hit: bool = False,
        elapsed_seconds: float = 0.0,
    ):
        if token != self._preflight_token:
            return
        reusable = bool(getattr(report, "can_run", False))
        if reusable:
            self._approved_preflight_signature = signature
            self._approved_preflight_report = report
            self._approved_uniform_policy_path = getattr(context, "uniform_policy_path", None)
            self._set_preflight_completed_status(ok, summary, report, cache_hit, elapsed_seconds)
            self.start_btn.configure(state=tk.NORMAL)
        else:
            self._set_preflight_completed_status(ok, summary, report, cache_hit, elapsed_seconds)
            self.start_btn.configure(state=tk.DISABLED)

    def _project_paths(self, fiscal_year: int | None = None):
        year = int(fiscal_year if fiscal_year is not None else self.fiscal_year.get())
        created = self.project.ensure_fiscal_year(year)
        paths = self.project.fiscal_paths(year)
        if created:
            self.project.save()
        return paths

    def _operational_database(self) -> str:
        return self.project.operational_database

    def _manual_input_store(self, fiscal_year: int | None = None) -> str:
        return self._project_paths(fiscal_year).manual_input_store

    def _initial_headcount_source_status(self) -> str:
        conn = None
        try:
            conn = get_connection(self._operational_database())
            create_schema(conn)
            values = {
                str(row[0]): str(row[1] or "")
                for row in conn.execute(
                    "SELECT key,value FROM sys_params WHERE key LIKE 'headcount_source_%'"
                )
            }
            updated = values.get("headcount_source_updated_at", "")
            if not updated:
                return t("db_status_unimported")
            imported = values.get("headcount_source_imported_files", "?")
            total = values.get("headcount_source_files", "?")
            skipped = values.get("headcount_source_skipped_files", "?")
            errors = values.get("headcount_source_error_files", "?")
            fy = values.get("headcount_source_fiscal_year", "?")
            return t("db_status_imported_summary", fy=fy, imported=imported, total=total, skipped=skipped, errors=errors, time=updated[:16])
        except Exception:
            return t("db_status_read_error")
        finally:
            if conn is not None:
                conn.close()

    def _set_headcount_source_status(self, key: str, **params) -> None:
        self._headcount_source_status_state = {"kind": "key", "key": key, "params": dict(params)}
        self.headcount_source_status.set(t(key, **params))

    def _set_headcount_source_status_from_database(self) -> None:
        self._headcount_source_status_state = {"kind": "database"}
        self.headcount_source_status.set(self._initial_headcount_source_status())

    def _refresh_headcount_source_status(self) -> None:
        state = getattr(self, "_headcount_source_status_state", {"kind": "database"})
        if state["kind"] == "database":
            self.headcount_source_status.set(self._initial_headcount_source_status())
            return
        self.headcount_source_status.set(t(state["key"], **state.get("params", {})))

    def _refresh_fiscal_year_labels(self, *_args):
        raw = self.fiscal_year.get().strip()
        label = raw if raw.isdigit() and len(raw) == 4 else "—"
        self.root.title(self._window_title(label, self.application_version))
        if hasattr(self, "main_heading"):
            self.main_heading.configure(text=t("main_heading", fiscal_year=label))

    @staticmethod
    def _initial_exchange_rate(template_path: str) -> str:
        try:
            return f"{read_exchange_rate_from_form(template_path):.0f}"
        except Exception:
            return ""

    def _activate_project(self, project: ProjectConfig) -> None:
        self._preflight_token += 1
        self.project = project
        remember_last_project(project.config_path)
        self.project_file.set(project.config_path)
        paths = self._project_paths(self._current_fiscal_year())
        self.template_path.set(paths.template_path)
        self.source_dir.set(paths.source_dir)
        self.headcount_source_dir.set(paths.headcount_source_dir)
        self.exchange_rate.set(self._initial_exchange_rate(paths.template_path))
        self._selected_cc_values = []
        self._available_cc_choices = []
        self._update_cc_selection_summary()
        self._set_headcount_source_status_from_database()
        self.load_cc_list()
        self._mark_preflight_stale()
        self.log(t("project_profile_using", path=project.config_path))

    def open_project(self) -> None:
        path = filedialog.askopenfilename(
            title=t("open_project_dialog_title"),
            initialdir=self.project.root_dir,
            filetypes=[(t("project_file_type"), "project.json"), (t("json_file_type"), "*.json")],
        )
        if not path:
            return
        try:
            self._activate_project(ProjectConfig.load(path))
        except Exception as exc:
            messagebox.showerror(t("cannot_open_project_title"), _friendly_error_message(exc))

    def create_project(self) -> None:
        root_dir = filedialog.askdirectory(title=t("choose_project_dir_title"))
        if not root_dir:
            return
        config_path = os.path.join(root_dir, "project.json")
        try:
            if os.path.isfile(config_path):
                project = ProjectConfig.load(config_path)
            else:
                project = ProjectConfig.create_legacy_compatible(
                    root_dir, self._current_fiscal_year(), config_path=config_path
                )
                project.save()
            self._activate_project(project)
        except Exception as exc:
            messagebox.showerror(t("cannot_create_project_title"), _friendly_error_message(exc))

    def configure_project_storage(self) -> None:
        """Edit shared and selected-FY storage paths without touching their data."""
        fiscal_year = self._current_fiscal_year()
        paths = self._project_paths(fiscal_year)
        dialog = tk.Toplevel(self.root)
        dialog.title(t("config_project_storage_title", fiscal_year=fiscal_year))
        dialog.geometry("820x430")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.columnconfigure(1, weight=1)

        fields = [
            (t("storage_op_db"), "operational_database", self.project.operational_database, "file"),
            (t("storage_uniform_policy"), "uniform_policy", paths.uniform_policy_path or "", "file"),
            (t("storage_manual_input"), "manual_input", paths.manual_input_store, "file"),
            (t("storage_manual_special_inheritance"), "manual_special_inheritance", paths.manual_special_inheritance_dir or "", "dir"),
            (t("storage_manual_special_legacy_starts"), "manual_special_legacy_starts", _format_manual_special_legacy_starts(paths.manual_special_legacy_starts), "text"),
            (t("storage_output_dir"), "output_dir", paths.output_dir, "dir"),
            (t("storage_history_dir"), "history_dir", paths.history_root, "dir"),
        ]
        variables = {}
        for row, (label, key_name, value, kind) in enumerate(fields):
            ttk.Label(dialog, text=label).grid(row=row, column=0, sticky="w", padx=12, pady=8)
            variable = tk.StringVar(value=value)
            variables[key_name] = variable
            ttk.Entry(dialog, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=8, pady=8)
            if kind == "dir":
                command = lambda var=variable, title=label: self._choose_project_directory(var, title)
            elif kind == "file":
                command = lambda var=variable, title=label: self._choose_project_file(var, title)
            else:
                command = None
            if command is not None:
                ttk.Button(dialog, text=t("btn_choose"), command=command).grid(row=row, column=2, padx=(0, 12), pady=8)

        ttk.Label(
            dialog,
            text=t("storage_paths_hint"),
            wraplength=760,
        ).grid(row=len(fields), column=0, columnspan=3, sticky="w", padx=12, pady=(8, 16))

        button_bar = ttk.Frame(dialog)
        button_bar.grid(row=len(fields) + 1, column=0, columnspan=3, sticky="e", padx=12, pady=12)
        ttk.Button(button_bar, text=t("btn_cancel"), command=dialog.destroy).pack(side="right")

        def save_configuration():
            try:
                self.project.update_storage_paths(
                    fiscal_year,
                    operational_database=variables["operational_database"].get().strip(),
                    uniform_policy_path=variables["uniform_policy"].get().strip(),
                    manual_input_store=variables["manual_input"].get().strip(),
                    manual_special_inheritance_dir=variables["manual_special_inheritance"].get().strip(),
                    manual_special_legacy_starts=_parse_manual_special_legacy_starts(
                        variables["manual_special_legacy_starts"].get()
                    ),
                    output_dir=variables["output_dir"].get().strip(),
                    history_root=variables["history_dir"].get().strip(),
                )
                self.project.save()
                refreshed = self._project_paths(fiscal_year)
                self.template_path.set(refreshed.template_path)
                self.source_dir.set(refreshed.source_dir)
                self.headcount_source_dir.set(refreshed.headcount_source_dir)
                self._mark_preflight_stale()
                self.log(t("project_profile_saved_storage", fy=fiscal_year))
                dialog.destroy()
            except Exception as exc:
                messagebox.showerror(
                    t("invalid_config_title"), _friendly_error_message(exc), parent=dialog
                )

        ttk.Button(button_bar, text=t("btn_save_config"), style="Primary.TButton", command=save_configuration).pack(
            side="right", padx=(0, 8)
        )

    @staticmethod
    def _choose_project_file(variable: tk.StringVar, title: str) -> None:
        path = filedialog.askopenfilename(title=t("choose_path_title", name=title))
        if path:
            variable.set(path)

    @staticmethod
    def _choose_project_directory(variable: tk.StringVar, title: str) -> None:
        path = filedialog.askdirectory(title=t("choose_path_title", name=title))
        if path:
            variable.set(path)

    def _save_path_preference(self, key: str, path: str, description: str) -> None:
        aliases = {
            "template_path": "template_path",
            "cost_source_dir": "source_dir",
            "headcount_source_dir": "headcount_source_dir",
        }
        argument = aliases.get(key)
        if argument is None:
            raise KeyError(t("project_profile_unsupported_path", key=key))
        self.project.update_fiscal_paths(self._current_fiscal_year(), **{argument: os.path.abspath(path)})
        self.project.save()

    def _reload_exchange_rate_from_template(self) -> bool:
        """Refresh the editable rate and never retain a stale value."""
        rate_text = self._initial_exchange_rate(self.template_path.get())
        if rate_text:
            self.exchange_rate.set(rate_text)
            self.log(t("fx_read_form_b2", rate=rate_text))
            return True

        self.exchange_rate.set("")
        self.log(t("fx_form_missing_b2"))
        return False

    def set_icon(self):
        icon_path = resource_path(os.path.join("assets", "app_icon.ico"))
        if os.path.exists(icon_path):
            try:
                # Windows specific icon loading
                self.root.iconbitmap(icon_path)
            except Exception as e:
                print(f"Lỗi khi nạp icon: {e}")


    def _on_workflow_state_changed(self, *_args):
        if hasattr(self, "workflow_cards"):
            self._update_workflow_guide()

    def _update_workflow_guide(self):
        if not hasattr(self, "workflow_cards"):
            return

        try:
            fiscal_year_ready = int(self.fiscal_year.get()) >= 2000
        except (TypeError, ValueError):
            fiscal_year_ready = False
        source_ready = all(
            (
                os.path.isfile(self.template_path.get()),
                os.path.isdir(self.source_dir.get()),
                os.path.isdir(self.headcount_source_dir.get()),
            )
        )
        report = self._approved_preflight_report
        preflight_ready = bool(report and getattr(report, "ok", False))
        preflight_warning = bool(report and getattr(report, "can_continue_incomplete", False))

        states = [
            "done" if fiscal_year_ready else "active",
            "done" if source_ready else ("active" if fiscal_year_ready else "pending"),
            "done" if preflight_ready else ("warning" if preflight_warning else ("active" if source_ready else "pending")),
            "optional" if (preflight_ready or preflight_warning) else "pending",
            "active" if (preflight_ready or preflight_warning) else "pending",
        ]
        palette = {
            "done": ("#dff4ea", "#176b4d", "✓"),
            "active": ("#e6efff", "#2457a6", "→"),
            "warning": ("#fff1cf", "#855b00", "!"),
            "optional": ("#eee9fb", "#5c3b8c", "○"),
            "pending": ("#f0f2f5", "#66707a", "·"),
        }
        for state, widgets in zip(states, self.workflow_cards):
            background, foreground, marker = palette[state]
            card, badge, title, detail = widgets
            card.configure(bg=background, highlightbackground=foreground)
            for widget in (badge, title, detail):
                widget.configure(bg=background, fg=foreground)
            badge.configure(text=marker)

        if not fiscal_year_ready:
            next_action = t("workflow_next_fiscal_year")
        elif not source_ready:
            next_action = t("workflow_next_sources")
        elif preflight_warning:
            next_action = t("workflow_next_warning")
        elif not preflight_ready:
            next_action = t("workflow_next_wait_check")
        else:
            next_action = t("workflow_next_run")
        self.workflow_next_action.configure(text=next_action)

    def setup_styles(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Header.TLabel", font=("Segoe UI", 15, "bold"), foreground="#0f172a")
        style.configure("WorkflowTitle.TLabel", font=("Segoe UI", 11, "bold"), foreground="#1e293b")
        style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=10)

    def setup_ui(self):
        shell = ttk.Frame(self.root, padding=12)
        shell.pack(fill=tk.BOTH, expand=True)
        shell.columnconfigure(0, weight=1)
        shell.rowconfigure(0, weight=1)

        content_shell = ttk.Frame(shell)
        content_shell.grid(row=0, column=0, sticky="nsew")
        content_shell.columnconfigure(0, weight=1)
        content_shell.rowconfigure(0, weight=1)
        content_canvas = tk.Canvas(content_shell, highlightthickness=0, borderwidth=0)
        vertical_scrollbar = ttk.Scrollbar(content_shell, orient=tk.VERTICAL, command=content_canvas.yview)
        horizontal_scrollbar = ttk.Scrollbar(content_shell, orient=tk.HORIZONTAL, command=content_canvas.xview)
        content_canvas.configure(
            yscrollcommand=vertical_scrollbar.set,
            xscrollcommand=horizontal_scrollbar.set,
        )
        content_canvas.grid(row=0, column=0, sticky="nsew")
        vertical_scrollbar.grid(row=0, column=1, sticky="ns")
        horizontal_scrollbar.grid(row=1, column=0, sticky="ew")

        container = ttk.Frame(content_canvas, padding=8)
        content_window = content_canvas.create_window((0, 0), window=container, anchor="nw")

        def refresh_scroll_region(_event=None):
            # Keep the scrollable content attached to the canvas origin.  Without
            # this normalization Tk can retain an offset while recomputing the
            # scroll region after a resize, exposing an empty white band above
            # the main screen.
            content_canvas.coords(content_window, 0, 0)
            content_canvas.configure(scrollregion=content_canvas.bbox("all"))

        def resize_content_width(event):
            content_canvas.coords(content_window, 0, 0)
            content_canvas.itemconfigure(
                content_window,
                width=max(event.width, container.winfo_reqwidth()),
            )

        container.bind("<Configure>", refresh_scroll_region)
        content_canvas.bind("<Configure>", resize_content_width)
        content_canvas.after_idle(lambda: content_canvas.yview_moveto(0.0))

        def scroll_content(event):
            content_canvas.yview_scroll(int(-event.delta / 120), "units")

        content_canvas.bind("<Enter>", lambda _event: content_canvas.bind_all("<MouseWheel>", scroll_content))
        content_canvas.bind("<Leave>", lambda _event: content_canvas.unbind_all("<MouseWheel>"))

        container.columnconfigure(1, weight=1)

        header_frame = ttk.Frame(container)
        header_frame.grid(row=0, column=0, sticky="w", pady=(0, 16))
        self.main_heading = ttk.Label(header_frame, text="", style="Header.TLabel")
        self.main_heading.pack(anchor="w")
        self.version_lbl = ttk.Label(
            header_frame,
            text=t("app_version", version=self.application_version),
            foreground="#4b5563",
        )
        self.version_lbl.pack(anchor="w", pady=(2, 0))
        project_bar = ttk.Frame(container)
        project_bar.grid(row=0, column=1, columnspan=2, sticky="e", pady=(0, 16))

        lang_frame = ttk.Frame(project_bar)
        lang_frame.pack(side="left", padx=(0, 12))
        self.lang_lbl = ttk.Label(lang_frame, text=t("language_label"))
        self.lang_lbl.pack(side="left", padx=(0, 4))
        self.lang_combo = ttk.Combobox(
            lang_frame,
            textvariable=self.language_var,
            values=[name for _, name in get_supported_languages()],
            state="readonly",
            width=11,
        )
        self.lang_combo.pack(side="left")
        self.lang_combo.bind("<<ComboboxSelected>>", self._on_language_selected)

        ttk.Label(project_bar, textvariable=self.project_file, width=40).pack(side="left", padx=(0, 6))
        self.open_proj_btn = ttk.Button(project_bar, text=t("open_project_btn"), command=self.open_project)
        self.open_proj_btn.pack(side="left")
        self.create_proj_btn = ttk.Button(project_bar, text=t("create_project_btn"), command=self.create_project)
        self.create_proj_btn.pack(side="left", padx=(6, 0))
        self.config_proj_btn = ttk.Button(project_bar, text=t("config_project_btn"), command=self.configure_project_storage)
        self.config_proj_btn.pack(
            side="left", padx=(6, 0)
        )

        self.fiscal_year_lbl = ttk.Label(container, text=t("fiscal_year_label"))
        self.fiscal_year_lbl.grid(row=1, column=0, sticky="w", pady=4)
        self.fiscal_year_entry = ttk.Entry(container, textvariable=self.fiscal_year, width=20)
        self.fiscal_year_entry.grid(row=1, column=1, sticky="w")

        # Tỷ giá (USD/VND)
        self.exchange_rate_lbl = ttk.Label(container, text=t("exchange_rate_label"))
        self.exchange_rate_lbl.grid(row=2, column=0, sticky="w", pady=4)
        self.exchange_rate_entry = ttk.Entry(container, textvariable=self.exchange_rate, width=20)
        self.exchange_rate_entry.grid(row=2, column=1, sticky="w")
        self.exchange_rate_hint_lbl = ttk.Label(container, text=t("exchange_rate_hint"))
        self.exchange_rate_hint_lbl.grid(
            row=2, column=2, sticky="w", padx=(12, 0)
        )

        self.template_lbl = ttk.Label(container, text=t("template_label"))
        self.template_lbl.grid(row=3, column=0, sticky="w", pady=(14, 4))
        self.template_path_entry = ttk.Entry(container, textvariable=self.template_path)
        self.template_path_entry.grid(
            row=3, column=1, columnspan=2, sticky="ew"
        )

        self.cost_source_lbl = ttk.Label(container, text=t("cost_source_dir_label"))
        self.cost_source_lbl.grid(row=4, column=0, sticky="w", pady=4)
        self.source_dir_entry = ttk.Entry(container, textvariable=self.source_dir)
        self.source_dir_entry.grid(
            row=4, column=1, columnspan=2, sticky="ew"
        )

        self.headcount_source_lbl = ttk.Label(container, text=t("headcount_source_dir_label"))
        self.headcount_source_lbl.grid(row=5, column=0, sticky="w", pady=4)
        self.headcount_source_dir_entry = ttk.Entry(container, textvariable=self.headcount_source_dir)
        self.headcount_source_dir_entry.grid(row=5, column=1, sticky="ew", padx=(0, 8))
        source_buttons = ttk.Frame(container)
        source_buttons.grid(row=5, column=2, sticky="w")
        self.update_db_btn = ttk.Button(
            source_buttons,
            text=t("update_db_btn"),
            command=self.update_headcount_database,
        )
        self.update_db_btn.pack(side="left")
        ttk.Label(
            container,
            textvariable=self.headcount_source_status,
            font=("Segoe UI", 9, "italic"),
        ).grid(row=6, column=1, columnspan=2, sticky="w", pady=(0, 8))

        self.cc_lbl = ttk.Label(container, text=t("cost_center_label"))
        self.cc_lbl.grid(row=7, column=0, sticky="w", pady=4)
        cc_frame = ttk.Frame(container)
        cc_frame.grid(row=7, column=1, sticky="ew")
        cc_frame.columnconfigure(0, weight=1)
        ttk.Entry(
            cc_frame,
            textvariable=self.cc_code_filter,
            state="readonly",
            width=42,
        ).grid(row=0, column=0, sticky="ew")
        self.cc_select_btn = ttk.Button(
            cc_frame,
            text=t("select_room_btn"),
            command=self._open_cc_selection_dialog,
        )
        self.cc_select_btn.grid(row=0, column=1, padx=(6, 0))
        self.refresh_btn = ttk.Button(
            cc_frame,
            text=t("reload_cc_btn"),
            command=self.refresh_cost_centers_from_form,
        )
        self.refresh_btn.grid(row=0, column=2, padx=(4, 0))
        self.cc_hint_lbl = ttk.Label(container, text=t("cc_hint"))
        self.cc_hint_lbl.grid(
            row=7, column=2, sticky="w", padx=(12, 0)
        )

        guide_panel = ttk.Frame(container, padding=(0, 6))
        guide_panel.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(4, 8))
        self.workflow_guide_lbl = ttk.Label(guide_panel, text=t("workflow_guide_title"), style="WorkflowTitle.TLabel")
        self.workflow_guide_lbl.pack(anchor="w", pady=(0, 6))
        workflow_row = ttk.Frame(guide_panel)
        workflow_row.pack(fill="x")
        workflow_steps = (
            ("1", "step1_title", "step1_desc"),
            ("2", "step2_title", "step2_desc"),
            ("3", "step3_title", "step3_desc"),
            ("4", "step4_title", "step4_desc"),
            ("5", "step5_title", "step5_desc"),
        )
        self.workflow_cards = []
        for column, (number, title_key, detail_key) in enumerate(workflow_steps):
            workflow_row.columnconfigure(column, weight=1, uniform="workflow")
            card = tk.Frame(workflow_row, bg="#f0f2f5", highlightthickness=1, highlightbackground="#66707a")
            card.grid(row=0, column=column, sticky="nsew", padx=(0 if column == 0 else 4, 0), ipadx=6, ipady=4)
            badge = tk.Label(card, text=number, font=("Segoe UI", 11, "bold"), bg="#f0f2f5", fg="#66707a")
            badge.pack(anchor="w")
            title = tk.Label(card, text=f"{number}. {t(title_key)}", font=("Segoe UI", 9, "bold"), bg="#f0f2f5", fg="#66707a")
            title.pack(anchor="w")
            detail = tk.Label(card, text=t(detail_key), font=("Segoe UI", 8), bg="#f0f2f5", fg="#66707a")
            detail.pack(anchor="w")
            self.workflow_cards.append((card, badge, title, detail))
        self.workflow_next_action = ttk.Label(guide_panel, text="", wraplength=900, font=("Segoe UI", 9, "bold"))
        self.workflow_next_action.pack(anchor="w", pady=(7, 0))

        actions = ttk.Frame(container)
        actions.grid(row=9, column=0, columnspan=3, sticky="w", pady=(4, 0))
        self.action_buttons = []
        for key, command in (
            ("manual_headcount_btn", self.open_headcount_editor_v2),
            ("event_driver_btn", self.open_event_driver_editor),
            ("source_order_btn", self.open_source_order_editor),
            ("output_cost_row_order_btn", self.open_output_cost_row_ordering),
            ("install_update_btn", self.install_application_update),
            ("variance_analysis_btn", self.open_variance_tab),
        ):
            btn = ttk.Button(actions, text=t(key), command=command)
            btn.pack(side="left", padx=(0, 8))
            self.action_buttons.append((btn, key))

        ttk.Separator(container, orient=tk.HORIZONTAL).grid(
            row=10, column=0, columnspan=3, sticky="ew", pady=12
        )

        check_actions = ttk.Frame(container)
        check_actions.grid(row=11, column=0, sticky="w")
        self.deep_scan_btn = ttk.Button(
            check_actions,
            text=t("deep_scan_btn"),
            command=lambda: self._mark_preflight_stale(force_refresh=True),
        )
        self.deep_scan_btn.pack(side="left")
        ttk.Label(
            container,
            textvariable=self.preflight_status,
            font=("Segoe UI", 9, "italic"),
            wraplength=700,
        ).grid(row=11, column=1, columnspan=2, sticky="w", padx=(8, 0))
        self.start_btn = ttk.Button(
            container,
            text=t("start_pipeline_btn"),
            style="Primary.TButton",
            command=self.start_pipeline,
        )
        self.start_btn.grid(row=12, column=0, columnspan=3, sticky="w", pady=(8, 0))

        log_frame = ttk.Frame(shell)
        log_frame.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        log_frame.columnconfigure(0, weight=1)
        self.log_title_lbl = ttk.Label(log_frame, text=t("process_log_title"))
        self.log_title_lbl.grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.log_widget = scrolledtext.ScrolledText(
            log_frame, height=5, state=tk.DISABLED, font=("Consolas", 9)
        )
        self.log_widget.grid(row=1, column=0, sticky="ew")

        self._setup_floating_mascot()

    def open_business_chat_assistant(self):
        """Mở cửa sổ Hỏi AI nội bộ (Business Chat Assistant)."""
        current_lang = get_current_language()
        history_root = None
        try:
            history_root = self._project_paths().history_root
        except Exception:
            history_root = None
        fiscal_year = None
        try:
            fiscal_year = self._current_fiscal_year()
        except Exception:
            fiscal_year = None
        return OperationsBusinessChatDialog.open(
            self.root,
            current_lang,
            open_history=self.open_run_history,
        )

    def _setup_floating_mascot(self):
        """Tạo mascot robot 3D có bóng đổ chân thực dưới chân, hiệu ứng bay và có thể kéo thả (Draggable)."""
        try:
            bg_color = "#f0f0f0"
            try:
                raw_bg = self.root.cget("bg")
                if raw_bg:
                    bg_color = raw_bg
            except Exception:
                bg_color = "#f0f0f0"

            canvas_w = 120
            canvas_h = 115
            self._mascot_canvas = tk.Canvas(
                self.root,
                width=canvas_w,
                height=canvas_h,
                bg=bg_color,
                highlightthickness=0,
                bd=0,
                cursor="hand2",
            )
            self._mascot_frame = self._mascot_canvas
            self._mascot_canvas.place(relx=1.0, rely=1.0, x=-16, y=-16, anchor="se")

            def _build_3d_mascot_image(hovering: bool = False):
                try:
                    img_path = resource_path("assets/operations_ai_mascot.png")
                    if not os.path.exists(img_path):
                        return None
                    from PIL import Image, ImageDraw, ImageFilter, ImageTk

                    bg_rgb = (240, 240, 240)
                    if isinstance(bg_color, str) and bg_color.startswith("#") and len(bg_color) == 7:
                        bg_rgb = (
                            int(bg_color[1:3], 16),
                            int(bg_color[3:5], 16),
                            int(bg_color[5:7], 16),
                        )

                    w, h = 90, 88
                    base = Image.new("RGBA", (w, h), (*bg_rgb, 255))

                    shadow = Image.new("RGBA", (w, h), (0, 0, 0, 0))
                    sdraw = ImageDraw.Draw(shadow)
                    sy = 75 if hovering else 73
                    sr_x = 24 if hovering else 21
                    sr_y = 6 if hovering else 7
                    s_alpha = 100 if hovering else 145
                    sdraw.ellipse(
                        [w // 2 - sr_x, sy - sr_y, w // 2 + sr_x, sy + sr_y],
                        fill=(71, 85, 105, s_alpha),
                    )
                    shadow = shadow.filter(ImageFilter.GaussianBlur(3))
                    base = Image.alpha_composite(base, shadow)

                    robot_y = 2 if hovering else 6
                    robot = Image.open(img_path).resize((64, 64), Image.Resampling.LANCZOS)
                    base.paste(robot, (w // 2 - 32, robot_y), robot)

                    return ImageTk.PhotoImage(base.convert("RGB"))
                except Exception:
                    return None

            self._mascot_img_normal = _build_3d_mascot_image(hovering=False)
            self._mascot_img_hover = _build_3d_mascot_image(hovering=True)
            self._mascot_photo = self._mascot_img_normal

            if self._mascot_img_normal:
                self._mascot_canvas_img_id = self._mascot_canvas.create_image(
                    canvas_w // 2, 68, image=self._mascot_img_normal, anchor="center"
                )

            def _draw_speech_bubble():
                self._mascot_canvas.delete("bubble")
                current_lang = get_current_language()
                title = translate_for_language("operations_business_chat_title", current_lang)
                bubble_text = f"✦ {title}"

                bx = canvas_w // 2
                bw = min(canvas_w - 4, max(68, len(bubble_text) * 7 + 16))
                x1 = bx - bw // 2
                y1 = 2
                x2 = bx + bw // 2
                y2 = 22

                self._mascot_canvas.create_rectangle(
                    x1 + 1, y1 + 1, x2 + 1, y2 + 1,
                    fill="#cbd5e1", outline="", tags="bubble"
                )
                self._mascot_canvas.create_rectangle(
                    x1, y1, x2, y2,
                    fill="#0284c7", outline="#0369a1", tags="bubble"
                )
                self._mascot_canvas.create_polygon(
                    bx - 4, y2, bx + 4, y2, bx, y2 + 4,
                    fill="#0284c7", outline="", tags="bubble"
                )
                self._mascot_bubble_text_id = self._mascot_canvas.create_text(
                    bx, (y1 + y2) // 2,
                    text=bubble_text,
                    fill="#ffffff",
                    font=("Segoe UI", 8, "bold"),
                    tags="bubble"
                )

            _draw_speech_bubble()
            self._mascot_draw_bubble = _draw_speech_bubble

            class _MascotTextProxy:
                def __init__(proxy_self, parent_app):
                    proxy_self._app = parent_app
                def configure(proxy_self, text=""):
                    if hasattr(proxy_self._app, "_mascot_draw_bubble"):
                        proxy_self._app._mascot_draw_bubble()
                def cget(proxy_self, key=""):
                    current_lang = get_current_language()
                    return f"✦ {translate_for_language('operations_business_chat_title', current_lang)}"

            self._mascot_text_lbl = _MascotTextProxy(self)

            self._mascot_drag = {"start_x": 0, "start_y": 0, "moved": False}

            def on_drag_start(event):
                self._mascot_drag["start_x"] = event.x_root
                self._mascot_drag["start_y"] = event.y_root
                self._mascot_drag["moved"] = False
                try:
                    tk.Misc.lift(self._mascot_canvas)
                except Exception:
                    pass

            def on_drag_motion(event):
                dx = event.x_root - self._mascot_drag["start_x"]
                dy = event.y_root - self._mascot_drag["start_y"]
                if abs(dx) > 3 or abs(dy) > 3:
                    self._mascot_drag["moved"] = True

                cur_x = self._mascot_canvas.winfo_x() + dx
                cur_y = self._mascot_canvas.winfo_y() + dy

                root_w = max(100, self.root.winfo_width())
                root_h = max(100, self.root.winfo_height())
                frame_w = self._mascot_canvas.winfo_width()
                frame_h = self._mascot_canvas.winfo_height()

                new_x = max(10, min(root_w - frame_w - 10, cur_x))
                new_y = max(10, min(root_h - frame_h - 10, cur_y))

                self._mascot_canvas.place(x=new_x, y=new_y, relx=0, rely=0, anchor="nw")
                self._mascot_drag["start_x"] = event.x_root
                self._mascot_drag["start_y"] = event.y_root

            def on_drag_release(event):
                if not self._mascot_drag["moved"]:
                    self.open_business_chat_assistant()

            def on_enter(_event):
                if hasattr(self, "_mascot_canvas_img_id") and self._mascot_img_hover:
                    self._mascot_canvas.itemconfig(self._mascot_canvas_img_id, image=self._mascot_img_hover)

            def on_leave(_event):
                if hasattr(self, "_mascot_canvas_img_id") and self._mascot_img_normal:
                    self._mascot_canvas.itemconfig(self._mascot_canvas_img_id, image=self._mascot_img_normal)

            self._mascot_canvas.bind("<Button-1>", on_drag_start)
            self._mascot_canvas.bind("<B1-Motion>", on_drag_motion)
            self._mascot_canvas.bind("<ButtonRelease-1>", on_drag_release)
            self._mascot_canvas.bind("<Enter>", on_enter)
            self._mascot_canvas.bind("<Leave>", on_leave)

        except Exception:
            pass

    def _on_language_selected(self, _event=None):
        selected_name = self.language_var.get()
        code = get_language_code(selected_name)
        set_current_language(code)
        remember_ui_language(code)
        self._refresh_localized_ui()
        self.log(t("lang_changed_log", lang_name=selected_name))

    def _refresh_localized_ui(self):
        self._refresh_fiscal_year_labels()
        if hasattr(self, "version_lbl"):
            self.version_lbl.configure(text=t("app_version", version=self.application_version))
        if hasattr(self, "lang_lbl"):
            self.lang_lbl.configure(text=t("language_label"))
        if hasattr(self, "open_proj_btn"):
            self.open_proj_btn.configure(text=t("open_project_btn"))
        if hasattr(self, "create_proj_btn"):
            self.create_proj_btn.configure(text=t("create_project_btn"))
        if hasattr(self, "config_proj_btn"):
            self.config_proj_btn.configure(text=t("config_project_btn"))
        if hasattr(self, "fiscal_year_lbl"):
            self.fiscal_year_lbl.configure(text=t("fiscal_year_label"))
        if hasattr(self, "exchange_rate_lbl"):
            self.exchange_rate_lbl.configure(text=t("exchange_rate_label"))
        if hasattr(self, "exchange_rate_hint_lbl"):
            self.exchange_rate_hint_lbl.configure(text=t("exchange_rate_hint"))
        if hasattr(self, "template_lbl"):
            self.template_lbl.configure(text=t("template_label"))
        if hasattr(self, "cost_source_lbl"):
            self.cost_source_lbl.configure(text=t("cost_source_dir_label"))
        if hasattr(self, "headcount_source_lbl"):
            self.headcount_source_lbl.configure(text=t("headcount_source_dir_label"))
        if hasattr(self, "update_db_btn"):
            self.update_db_btn.configure(text=t("update_db_btn"))
        if hasattr(self, "cc_lbl"):
            self.cc_lbl.configure(text=t("cost_center_label"))
        if hasattr(self, "cc_select_btn"):
            self.cc_select_btn.configure(text=t("select_room_btn"))
        if hasattr(self, "refresh_btn"):
            self.refresh_btn.configure(text=t("reload_cc_btn"))
        if hasattr(self, "cc_hint_lbl"):
            self.cc_hint_lbl.configure(text=t("cc_hint"))
        if hasattr(self, "workflow_guide_lbl"):
            self.workflow_guide_lbl.configure(text=t("workflow_guide_title"))
        if hasattr(self, "workflow_cards") and self.workflow_cards:
            step_keys = [
                ("1", "step1_title", "step1_desc"),
                ("2", "step2_title", "step2_desc"),
                ("3", "step3_title", "step3_desc"),
                ("4", "step4_title", "step4_desc"),
                ("5", "step5_title", "step5_desc"),
            ]
            for idx, (_card, _badge, title, detail) in enumerate(self.workflow_cards):
                if idx < len(step_keys):
                    num, t_key, d_key = step_keys[idx]
                    title.configure(text=f"{num}. {t(t_key)}")
                    detail.configure(text=t(d_key))
        if hasattr(self, "action_buttons"):
            for btn, key in self.action_buttons:
                btn.configure(text=t(key))
        if hasattr(self, "quick_check_btn"):
            self.quick_check_btn.configure(text=t("quick_check_btn"))
        if hasattr(self, "deep_scan_btn"):
            self.deep_scan_btn.configure(text=t("deep_scan_btn"))
        if hasattr(self, "start_btn"):
            self.start_btn.configure(text=t("start_pipeline_btn"))
        if hasattr(self, "log_title_lbl"):
            self.log_title_lbl.configure(text=t("process_log_title"))
        if hasattr(self, "headcount_source_status"):
            self._refresh_headcount_source_status()
        if hasattr(self, "preflight_status"):
            self._refresh_preflight_status()
        if hasattr(self, "_mascot_text_lbl") and self._mascot_text_lbl is not None:
            self._mascot_text_lbl.configure(text=f"✦ {t('operations_business_chat_title')}")
        self._update_workflow_guide()
        self._update_cc_selection_summary()


    def _run_on_ui_thread(self, callback, *args, **kwargs):
        if threading.get_ident() == self.ui_thread_id:
            callback(*args, **kwargs)
            return
        self.ui_queue.put((callback, args, kwargs))

    def _drain_ui_queue(self):
        while True:
            try:
                callback, args, kwargs = self.ui_queue.get_nowait()
            except queue.Empty:
                break
            callback(*args, **kwargs)
        self.root.after(50, self._drain_ui_queue)

    def log(self, message: str):
        if threading.get_ident() != self.ui_thread_id:
            self._run_on_ui_thread(self.log, message)
            return
        self.log_widget.configure(state=tk.NORMAL)
        self.log_widget.insert(tk.END, f"{datetime.now().strftime('[%H:%M:%S]')} {message}\n")
        self.log_widget.see(tk.END)
        self.log_widget.configure(state=tk.DISABLED)

    def install_content_package(self):
        try:
            fiscal_year = int(self.fiscal_year.get())
        except (TypeError, ValueError):
            messagebox.showerror(t("invalid_fiscal_year_title"), t("invalid_fiscal_year_msg"))
            return
        package_path = filedialog.askopenfilename(
            initialdir=BASE_DIR,
            title=t("choose_rule_pack_title"),
            filetypes=[(t("rule_pack_file_type"), "*.mpcontent")],
        )
        if not package_path:
            return
        try:
            installed_path = install_runtime_content_pack(
                package_path,
                BASE_DIR,
                fiscal_year=fiscal_year,
            )
        except ContentPackError as exc:
            message = _friendly_error_message(exc)
            self.log(t("update_pkg_install_failed", message=message))
            messagebox.showerror(t("invalid_rule_pack_title"), message)
            return
        except Exception as exc:
            message = _friendly_error_message(exc)
            self.log(t("update_pkg_install_failed", message=message))
            messagebox.showerror(t("cannot_install_rule_pack_title"), message)
            return
        self._mark_preflight_stale(force_refresh=True)
        self.log(t("update_pkg_activated", fy=fiscal_year, path=installed_path))
        messagebox.showinfo(
            t("installed_rule_pack_title"),
            t("installed_rule_pack_msg", fiscal_year=fiscal_year),
        )

    def _start_update_discovery(self):
        """Check configured sources after UI startup without blocking Tkinter."""
        if self._startup_update_prompted or getattr(self, "_application_update_running", False):
            return
        try:
            config = load_update_config(BASE_DIR)
            if not config["startup_check"] or not config["sources"]:
                return
            app_root = application_install_root(APP_DIR)
            current_version = current_release_version()
        except Exception as exc:
            self.log(t("update_startup_check_skipped", error=_friendly_error_message(exc)))
            return

        def worker():
            try:
                candidate = discover_available_update(
                    config["sources"],
                    current_version=current_version,
                    current_database_schema=CURRENT_SCHEMA_VERSION,
                )
                self._run_on_ui_thread(
                    self._offer_discovered_update, candidate, app_root, current_version
                )
            except Exception as exc:
                self._run_on_ui_thread(
                    self.log,
                    t("update_check_failed", error=_friendly_error_message(exc)),
                )

        threading.Thread(target=worker, daemon=True).start()

    def _offer_discovered_update(self, candidate, app_root, current_version):
        if candidate is None or self._startup_update_prompted or getattr(self, "_application_update_running", False):
            return
        self._startup_update_prompted = True
        notes = candidate.notes.strip() if candidate.notes else ""
        if notes:
            prompt = t("update_new_version_with_notes_prompt", version=candidate.version, current=current_version, notes=notes)
        else:
            prompt = t("update_new_version_prompt", version=candidate.version, current=current_version)
        if messagebox.askyesno(t("update_found_title"), prompt):
            self._install_discovered_update(candidate, app_root)
        else:
            self.log(t("update_deferred_log", version=candidate.version))

    def _install_discovered_update(self, candidate, app_root):
        if getattr(self, "_application_update_running", False):
            return
        self._application_update_running = True
        self.log(t("update_downloading_log", version=candidate.version))

        def worker():
            try:
                package_path = fetch_update_candidate(candidate, BASE_DIR)
                state = install_runtime_application_update(
                    package_path,
                    app_root,
                    BASE_DIR,
                    current_database_schema=CURRENT_SCHEMA_VERSION,
                )
                self._run_on_ui_thread(self._finish_application_update, state, None)
            except Exception as exc:
                self._run_on_ui_thread(self._finish_application_update, None, exc)

        threading.Thread(target=worker, daemon=True).start()

    def install_application_update(self):
        if getattr(self, "_application_update_running", False):
            return
        try:
            app_root = application_install_root(APP_DIR)
            config = load_update_config(BASE_DIR)
            current_version = current_release_version()
        except Exception as exc:
            messagebox.showerror(t("cannot_auto_update_title"), _friendly_error_message(exc))
            return
        if not config["sources"]:
            messagebox.showerror(
                t("no_update_source_title"),
                t("no_update_source_msg"),
            )
            return
        self._application_update_running = True
        self.log(t("update_scanning_log"))

        def worker():
            try:
                reachable_sources = []
                source_errors = []
                for source in config["sources"]:
                    if not source.get("enabled", True):
                        continue
                    try:
                        validated = validate_update_source(
                            str(source.get("type", "")),
                            str(source.get("location", "")),
                            enabled=True,
                        )
                        check_update_source(validated)
                        reachable_sources.append(source)
                    except Exception as exc:
                        source_errors.append(_friendly_error_message(exc))
                if not reachable_sources:
                    detail = source_errors[0] if source_errors else t("update_no_source_enabled")
                    raise UpdateDeliveryError(detail)
                candidate = discover_available_update(
                    reachable_sources,
                    current_version=current_version,
                    current_database_schema=CURRENT_SCHEMA_VERSION,
                )
            except Exception as exc:
                self._run_on_ui_thread(
                    self._finish_manual_update_discovery,
                    None,
                    app_root,
                    current_version,
                    exc,
                )
                return
            self._run_on_ui_thread(
                self._finish_manual_update_discovery,
                candidate,
                app_root,
                current_version,
                None,
            )

        threading.Thread(target=worker, daemon=True).start()

    def _finish_manual_update_discovery(self, candidate, app_root, current_version, error):
        self._application_update_running = False
        if error is not None:
            message = _friendly_error_message(error)
            self.log(t("update_scan_failed", message=message))
            messagebox.showerror(t("cannot_auto_update_title"), message)
            return
        if candidate is None:
            self.log(t("update_already_latest", current=current_version))
            messagebox.showinfo(
                t("latest_version_title"),
                t("latest_version_msg", version=current_version),
            )
            return
        notes = candidate.notes.strip() if candidate.notes else ""
        if notes:
            prompt = t("update_new_version_with_notes_prompt", version=candidate.version, current=current_version, notes=notes)
        else:
            prompt = t("update_new_version_prompt", version=candidate.version, current=current_version)
        if messagebox.askyesno(t("update_found_title"), prompt):
            self._install_discovered_update(candidate, app_root)
        else:
            self.log(t("update_deferred_log", version=candidate.version))

    def _finish_application_update(self, state, error):
        self._application_update_running = False
        if error is not None:
            self.log(t("update_app_failed", error=error))
            messagebox.showerror(
                t("update_failed_title"),
                t("update_failed_msg"),
            )
            return
        version = state.get("version", t("update_version_latest_word")) if isinstance(state, dict) else t("update_version_latest_word")
        self.log(t("update_app_activated", version=version))
        messagebox.showinfo(
            t("update_ready_title"),
            t("update_ready_msg", version=version),
        )
        try:
            app_root = application_install_root(APP_DIR)
            entrypoint = launch_activated_update(app_root, current_pid=os.getpid())
        except Exception as exc:
            self.log(t("update_restart_failed", error=exc))
            messagebox.showerror(
                t("cannot_auto_restart_title"),
                t("cannot_auto_restart_msg"),
            )
            return
        self.log(f"Đã lên lịch khởi động phiên bản mới: {entrypoint}")
        self.root.quit()
        self.root.destroy()

    def browse_template(self):
        current = self.template_path.get().strip()
        initial_dir = os.path.dirname(current) if os.path.isfile(current) else BASE_DIR
        path = filedialog.askopenfilename(initialdir=initial_dir, filetypes=[(t("excel_file_type"), "*.xlsx")])
        if path:
            validation_error = _validate_selected_template(path)
            if validation_error:
                messagebox.showerror(t("invalid_template_title"), validation_error)
                return
            self.template_path.set(path)
            self._save_path_preference("template_path", path, "Selected FORM template")
            self._reload_exchange_rate_from_template()

    def browse_source_dir(self):
        initial_dir = self.source_dir.get().strip()
        if not os.path.isdir(initial_dir):
            initial_dir = BASE_DIR
        path = filedialog.askdirectory(initialdir=initial_dir)
        if path:
            validation_error = _validate_selected_source_dir(path)
            if validation_error:
                messagebox.showerror(t("invalid_cost_source_title"), validation_error)
                return
            self.source_dir.set(path)
            self._save_path_preference("cost_source_dir", path, "Selected cost source folder")

    def browse_headcount_source_dir(self):
        initial_dir = self.headcount_source_dir.get().strip()
        if not os.path.isdir(initial_dir):
            initial_dir = BASE_DIR
        path = filedialog.askdirectory(initialdir=initial_dir)
        if path:
            self.headcount_source_dir.set(path)
            try:
                self._save_path_preference(
                    "headcount_source_dir", path, "Department plan source folder"
                )
                self._set_headcount_source_status("hc_source_needs_sync")
            except Exception as exc:
                self.log(f"Không lưu được thư mục nguồn nhân sự đã chọn: {exc}")

    def cleanup_headcount_database(self):
        try:
            fiscal_year = int(self.fiscal_year.get())
        except (TypeError, ValueError):
            messagebox.showerror(t("invalid_fiscal_year_title"), t("invalid_fiscal_year_msg"))
            return

        conn = None
        try:
            conn = get_connection(self._operational_database())
            create_schema(conn)
            counts = count_headcount_truth_rows(conn, fiscal_year)
            if counts["total_rows"] == 0:
                messagebox.showinfo(
                    t("no_data_cleanup_title"),
                    t("no_data_cleanup_msg", fiscal_year=fiscal_year, start=counts['periods'][0] if counts['periods'] else "", end=counts['periods'][-1] if counts['periods'] else ""),
                )
                return

            confirmed = messagebox.askyesno(
                t("cleanup_confirm_title"),
                t(
                    "cleanup_confirm_msg",
                    fiscal_year=fiscal_year,
                    start=counts['periods'][0],
                    end=counts['periods'][-1],
                    hc_rows=counts['monthly_headcount_rows'],
                    time_rows=counts['headcount_time_rows'],
                    total_rows=counts['total_rows'],
                ),
                icon="warning",
            )
            if not confirmed:
                self.log(t("cleanup_cancelled_log", fiscal_year=fiscal_year))
                return

            result = cleanup_headcount_truth(conn, fiscal_year)
        except Exception as exc:
            self.log(t("cleanup_failed_log", fiscal_year=fiscal_year, error=str(exc)))
            messagebox.showerror(
                t("cleanup_failed_title"),
                f"{exc}",
            )
            return
        finally:
            if conn is not None:
                conn.close()

        self._set_headcount_source_status("sync_cleaned_data", fy=fiscal_year)
        self.log(
            t(
                "sync_cleaned_truth",
                fy=fiscal_year,
                hc=result['monthly_headcount_rows'],
                time=result['headcount_time_rows'],
            )
        )
        messagebox.showinfo(
            t("cleanup_success_title"),
            t("cleanup_success_msg", total_rows=result['total_rows'], fiscal_year=fiscal_year),
        )

    def _confirm_headcount_source_exceptions(self, review):
        """Return approved/rejected file sets, or None when the user cancels."""
        unknown = list(review.get("unknown_cost_centers", []))
        mismatches = list(review.get("name_mismatches", []))
        if not unknown and not mismatches:
            return set(), set(), set()

        dialog = tk.Toplevel(self.root)
        dialog.title(t("headcount_source_exceptions_title"))
        dialog.geometry("940x640")
        dialog.minsize(760, 480)
        dialog.transient(self.root)
        dialog.grab_set()
        outcome = {"value": None}

        outer = ttk.Frame(dialog, padding=14)
        outer.pack(fill=tk.BOTH, expand=True)
        ttk.Label(
            outer,
            text=t("headcount_source_exceptions_guide"),
            wraplength=880,
            justify=tk.LEFT,
        ).pack(fill=tk.X, pady=(0, 10))

        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        content = ttk.Frame(canvas)
        content.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(window_id, width=event.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        variables = []

        def add_group(title, items, explanation, kind):
            if not items:
                return
            ttk.Label(content, text=title, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(8, 2))
            ttk.Label(content, text=explanation, wraplength=850, justify=tk.LEFT).pack(anchor="w", pady=(0, 6))
            for parsed in items:
                variable = tk.BooleanVar(value=False)
                expected = " / ".join(
                    value for value in (
                        getattr(parsed, "department_name_jp", ""),
                        getattr(parsed, "department_name_vn", ""),
                    ) if value
                )
                label = (
                    f"CC {parsed.cc_code} — B5: {parsed.department_name}\n"
                    f"Tệp: {os.path.basename(parsed.path)}"
                )
                if expected:
                    label += f"\nLookup dự kiến: {expected}"
                ttk.Checkbutton(content, text=label, variable=variable).pack(
                    anchor="w", fill=tk.X, padx=(12, 0), pady=5
                )
                variables.append((kind, os.path.abspath(parsed.path), variable))

        add_group(
            t("headcount_source_unknown_cc_group"),
            unknown,
            t("headcount_source_unknown_cc_desc"),
            "unknown",
        )
        add_group(
            t("headcount_source_name_mismatch_group"),
            mismatches,
            t("headcount_source_name_mismatch_desc"),
            "name",
        )

        buttons = ttk.Frame(dialog, padding=(14, 8, 14, 14))
        buttons.pack(fill=tk.X)

        def submit():
            approved_unknown = {path for kind, path, var in variables if kind == "unknown" and var.get()}
            rejected_unknown = {path for kind, path, var in variables if kind == "unknown" and not var.get()}
            approved_names = {path for kind, path, var in variables if kind == "name" and var.get()}
            outcome["value"] = (approved_unknown, rejected_unknown, approved_names)
            dialog.destroy()

        ttk.Button(buttons, text=t("btn_cancel"), command=dialog.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(buttons, text=t("btn_continue_selection"), command=submit).pack(side=tk.RIGHT)
        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)
        self.root.wait_window(dialog)
        return outcome["value"]

    def update_headcount_database(self):
        source_dir = self.headcount_source_dir.get().strip()
        if not os.path.isdir(source_dir):
            messagebox.showerror(t("invalid_source_title"), t("invalid_headcount_source_msg"))
            return
        conn = None
        try:
            fiscal_year = int(self.fiscal_year.get())
            conn = get_connection(self._operational_database())
            create_schema(conn)
            review = review_headcount_time_sources(conn, source_dir, fiscal_year)
            approvals = self._confirm_headcount_source_exceptions(review)
            if approvals is None:
                self.log(t("sync_cancelled"))
                return
            approved_unknown, rejected_unknown, approved_names = approvals
            result = import_headcount_time_sources(
                conn,
                source_dir,
                fiscal_year,
                approved_unknown_files=approved_unknown,
                rejected_unknown_files=rejected_unknown,
                approved_name_files=approved_names,
                scan_results=review["results"],
            )
        except Exception as exc:
            message = _friendly_error_message(exc)
            self.log(t("sync_failed", message=message))
            messagebox.showerror(t("update_db_failed_title"), message)
            return
        finally:
            if conn is not None:
                conn.close()
        for parsed, reason in result["skipped"]:
            self.log(t("sync_not_loaded_file", filename=os.path.basename(parsed.path), reason=reason))
        for parsed in result["errors"]:
            self.log(t("sync_error_file", filename=os.path.basename(parsed.path), error='; '.join(parsed.errors)))
        self._set_headcount_source_status(
            "sync_imported_summary",
            imported=result['imported_files'],
            total=result['files'],
            rows=result['imported_rows'],
            split=result.get('split_required_files', 0),
            skipped=len(result['skipped']),
            errors=len(result['errors']),
            time=f"{datetime.now():%H:%M}",
        )
        text = self.headcount_source_status.get()
        self.log(text)
        detail_lines = []
        for parsed, reason in result["skipped"]:
            detail_lines.append(
                t(
                    "sync_skip_detail",
                    filename=os.path.basename(parsed.path),
                    cc=parsed.cc_code or '?',
                    dept=parsed.department_name or '?',
                    reason=reason,
                )
            )
        for parsed in result["errors"]:
            detail_lines.append(
                t(
                    "sync_error_detail",
                    filename=os.path.basename(parsed.path),
                    errors='; '.join(parsed.errors) or '?',
                )
            )
        confirmed_unknown = result.get("confirmed_unknown_cost_centers", [])
        if confirmed_unknown:
            detail_lines.append(
                t("sync_master_recommendation")
                + "\n  ".join(
                    f"CC {parsed.cc_code} — {parsed.department_name}" for parsed in confirmed_unknown
                )
            )
        message = text
        if detail_lines:
            message += "\n\n" + t("sync_files_to_check") + "\n\n".join(detail_lines)
        messagebox.showinfo(t("update_db_success_title"), message)

    def open_output_cost_row_ordering(self):
        if self._focus_existing_editor("_output_cost_row_order_editor"):
            return
        try:
            fiscal_year = int(self.fiscal_year.get())
            output_dir = self._project_paths(fiscal_year).output_dir
        except (TypeError, ValueError, AttributeError):
            output_dir = BASE_DIR
        selected_path = filedialog.askopenfilename(
            initialdir=output_dir if os.path.isdir(output_dir) else BASE_DIR,
            title=t("output_cost_row_order_choose"),
            filetypes=[(t("excel_file_type"), "*.xlsx")],
        )
        if selected_path:
            self._open_cost_row_ordering_dialog(selected_path)

    def _open_cost_row_ordering_dialog(self, workbook_path: str):
        if self._focus_existing_editor("_output_cost_row_order_editor"):
            return
        from src.engine.output_cost_row_ordering import (
            OutputCostRowOrderError,
            read_cost_rows,
            save_cost_row_order,
        )

        match = re.search(r"MP_CC_([^\\/]+)\.xlsx$", os.path.basename(workbook_path), re.IGNORECASE)
        if not match:
            messagebox.showerror(
                t("output_cost_row_order_title"),
                t("output_cost_row_order_invalid_file"),
            )
            return
        cc_code = match.group(1).strip()
        try:
            rows = read_cost_rows(workbook_path, cc_code)
        except (OutputCostRowOrderError, OSError, ValueError) as exc:
            messagebox.showerror(
                t("output_cost_row_order_title"),
                t("output_cost_row_order_error", error=str(exc)),
            )
            return
        if not rows:
            messagebox.showwarning(
                t("output_cost_row_order_title"),
                t("output_cost_row_order_error", error="Không có dòng chi phí để sắp xếp."),
            )
            return

        editor = tk.Toplevel(self.root)
        editor.title(f"{t('output_cost_row_order_title')} - {cc_code}")
        editor.geometry("920x560")
        editor.transient(self.root)
        editor.lift()
        editor.focus_force()
        close_editor = self._register_singleton_editor("_output_cost_row_order_editor", editor)
        frame = ttk.Frame(editor, padding=12)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text=t("output_cost_row_order_hint"), wraplength=860).pack(anchor="w", pady=(0, 8))
        tree = ttk.Treeview(frame, columns=("no", "kind", "account", "description"), show="headings", height=18)
        for column, title, width in (
            ("no", "#", 50),
            ("kind", t("output_cost_row_order_common"), 150),
            ("account", t("col_account_code"), 170),
            ("description", t("col_item_name"), 500),
        ):
            tree.heading(column, text=title)
            tree.column(column, width=width, anchor="w")
        tree.pack(fill="both", expand=True)
        ordered_rows = list(rows)
        drag_row_id = {"value": None}

        def redraw(select_row_id: str | None = None):
            tree.delete(*tree.get_children())
            for index, row in enumerate(ordered_rows, start=1):
                kind_label = t("output_cost_row_order_manual") if row.row_kind == "manual" else t("output_cost_row_order_common")
                tree.insert("", "end", iid=row.row_id, values=(index, kind_label, row.account_code, row.description))
            if select_row_id and tree.exists(select_row_id):
                tree.selection_set(select_row_id)
                tree.focus(select_row_id)

        def move_selected(delta: int):
            selected = tree.selection()
            if not selected:
                return
            row_id = selected[0]
            index = next((i for i, row in enumerate(ordered_rows) if row.row_id == row_id), None)
            if index is None:
                return
            target = index + delta
            if not 0 <= target < len(ordered_rows):
                return
            ordered_rows[index], ordered_rows[target] = ordered_rows[target], ordered_rows[index]
            redraw(row_id)

        def start_drag(event):
            drag_row_id["value"] = tree.identify_row(event.y) or None

        def finish_drag(event):
            source_id = drag_row_id["value"]
            target_id = tree.identify_row(event.y)
            drag_row_id["value"] = None
            if not source_id or not target_id or source_id == target_id:
                return
            source_index = next(i for i, row in enumerate(ordered_rows) if row.row_id == source_id)
            target_index = next(i for i, row in enumerate(ordered_rows) if row.row_id == target_id)
            moved = ordered_rows.pop(source_index)
            ordered_rows.insert(target_index, moved)
            redraw(source_id)

        def save():
            try:
                save_cost_row_order(workbook_path, cc_code, [row.row_id for row in ordered_rows])
            except (OutputCostRowOrderError, OSError, ValueError) as exc:
                messagebox.showerror(
                    t("output_cost_row_order_title"),
                    t("output_cost_row_order_error", error=str(exc)),
                )
                return
            messagebox.showinfo(t("output_cost_row_order_title"), t("output_cost_row_order_saved", count=len(ordered_rows)))
            close_editor()

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x", pady=(8, 0))
        ttk.Button(buttons, text=t("btn_move_up"), command=lambda: move_selected(-1)).pack(side="left")
        ttk.Button(buttons, text=t("btn_move_down"), command=lambda: move_selected(1)).pack(side="left", padx=(6, 0))
        ttk.Button(buttons, text=t("btn_close"), command=close_editor).pack(side="right")
        ttk.Button(buttons, text=t("btn_save"), style="Primary.TButton", command=save).pack(side="right", padx=(0, 6))
        tree.bind("<ButtonPress-1>", start_drag)
        tree.bind("<ButtonRelease-1>", finish_drag)
        redraw()

    def open_source_order_editor(self):
        source_dir = self.source_dir.get() or BASE_DIR
        if not os.path.isdir(source_dir):
            messagebox.showerror(
                t("dir_not_found_title"),
                t("source_dir_not_found_msg", path=source_dir),
            )
            return

        editor = tk.Toplevel(self.root)
        editor.title(t("source_order_title"))
        editor.geometry("980x520")
        editor.transient(self.root)
        editor.grab_set()

        frame = ttk.Frame(editor, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            frame,
            text=t("source_order_instruction"),
            wraplength=900,
        ).grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 4))
        summary_var = tk.StringVar()
        ttk.Label(frame, textvariable=summary_var).grid(row=1, column=0, columnspan=6, sticky="w", pady=(0, 8))

        category_labels = {
            code: CATEGORY_DISPLAY_NAMES.get(code, code)
            for code in DEFAULT_DESCRIPTIONS
        }
        category_codes = {label: code for code, label in category_labels.items()}
        status_labels = {
            "recognized": t("source_status_recognized"),
            "needs_review": t("source_status_needs_review"),
            "ignored": t("source_status_ignored"),
        }
        status_codes = {label: code for code, label in status_labels.items()}
        detection_labels = {
            "manifest": t("source_detection_manifest"),
            "structure": t("source_detection_structure"),
            "system_structure": t("source_detection_system"),
            "manual": t("source_detection_manual"),
            "inventory": t("source_detection_inventory"),
        }
        detection_codes = {label: code for code, label in detection_labels.items()}

        def display_manifest_value(column: str, value: object) -> str:
            text = str(value or "")
            if column == "category":
                return category_labels.get(text, text)
            if column == "status":
                return status_labels.get(text, text)
            if column == "detection_method":
                return detection_labels.get(text, text)
            return text

        def internal_manifest_value(column: str, value: object) -> str:
            text = str(value or "")
            if column == "category":
                return category_codes.get(text, text)
            if column == "status":
                return status_codes.get(text, text)
            if column == "detection_method":
                return detection_codes.get(text, text)
            return text

        columns = MANIFEST_COLUMNS
        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=2, column=0, columnspan=6, sticky="nsew")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=13)
        headings = {
            "order": t("source_col_order"), "category": t("source_col_category"), "filename": t("source_col_filename"),
            "enabled": t("source_col_enabled"), "description": t("source_col_description"), "status": t("source_col_status"),
            "detection_method": t("source_col_detection"), "signature": t("source_col_signature"), "reason": t("source_col_reason"),
        }
        widths = {"order": 55, "category": 145, "filename": 330, "enabled": 50,
                  "description": 180, "status": 105, "detection_method": 165,
                  "signature": 90, "reason": 340}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], anchor="center" if column in {"order", "enabled", "status"} else "w")
        tree.grid(row=0, column=0, sticky="nsew")

        vertical_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        vertical_scrollbar.grid(row=0, column=1, sticky="ns")
        horizontal_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        horizontal_scrollbar.grid(row=1, column=0, sticky="ew")
        tree.configure(
            yscrollcommand=vertical_scrollbar.set,
            xscrollcommand=horizontal_scrollbar.set,
        )
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        form = ttk.Frame(frame)
        form.grid(row=3, column=0, columnspan=6, sticky="ew", pady=(10, 0))
        ttk.Label(form, text=t("source_form_category_label")).grid(row=0, column=0, sticky="w")
        category_var = tk.StringVar()
        category_combo = ttk.Combobox(
            form,
            textvariable=category_var,
            values=[""] + list(category_labels.values()),
            width=22,
            state="readonly",
        )
        category_combo.grid(row=1, column=0, sticky="w", padx=(0, 8))

        ttk.Label(form, text=t("source_form_filename_label")).grid(row=0, column=1, sticky="w")
        filename_var = tk.StringVar()
        ttk.Entry(form, textvariable=filename_var, width=70).grid(row=1, column=1, sticky="w")
        ttk.Button(
            form,
            text=t("btn_browse_file"),
            command=lambda: browse_manifest_file(),
        ).grid(row=1, column=2, sticky="w", padx=(6, 8))

        enabled_var = tk.IntVar(value=1)
        ttk.Checkbutton(form, text=t("source_use_row"), variable=enabled_var).grid(row=1, column=3, sticky="w")

        ttk.Label(form, text=t("source_col_description")).grid(row=2, column=0, sticky="w", pady=(8, 0))
        description_var = tk.StringVar()
        ttk.Entry(form, textvariable=description_var, width=96).grid(
            row=3, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        def rows_from_tree() -> list[dict[str, str]]:
            rows: list[dict[str, str]] = []
            for index, item_id in enumerate(tree.get_children(), start=1):
                values = tree.item(item_id, "values")
                rows.append({
                    column: internal_manifest_value(column, values[column_index])
                    for column_index, column in enumerate(columns)
                    if column_index < len(values)
                })
                rows[-1]["order"] = str(index)
            return rows

        def refresh_order_numbers() -> None:
            for index, item_id in enumerate(tree.get_children(), start=1):
                values = list(tree.item(item_id, "values"))
                values[0] = str(index)
                tree.item(item_id, values=values)

        def update_summary() -> None:
            counts = {status: 0 for status in ("recognized", "needs_review", "ignored")}
            for item_id in tree.get_children():
                displayed_status = str(tree.item(item_id, "values")[5]).strip()
                status = internal_manifest_value("status", displayed_status).lower()
                if status in counts:
                    counts[status] += 1
            summary_var.set(
                t("source_summary", total=sum(counts.values()), recognized=counts['recognized'],
                  needs_review=counts['needs_review'], ignored=counts['ignored'])
            )

        def load_rows() -> None:
            for item_id in tree.get_children():
                tree.delete(item_id)
            for row in read_source_manifest_inventory_fast(source_dir):
                tree.insert(
                    "",
                    tk.END,
                    values=tuple(
                        display_manifest_value(column, row.get(column, ""))
                        for column in columns
                    ),
                )
            refresh_order_numbers()
            update_summary()

        def selected_item() -> str | None:
            selection = tree.selection()
            return selection[0] if selection else None

        def fill_form_from_selection(_event=None) -> None:
            item_id = selected_item()
            if not item_id:
                return
            values = tree.item(item_id, "values")
            category_var.set(str(values[1]))
            filename_var.set(str(values[2]))
            enabled_var.set(1 if str(values[3]).strip() not in {"0", "False", "false"} else 0)
            description_var.set(str(values[4]))

        def browse_manifest_file() -> None:
            path = filedialog.askopenfilename(
                initialdir=source_dir,
                filetypes=[(t("excel_file_type"), "*.xlsx *.xls"), (t("all_files_type"), "*.*")],
            )
            if not path:
                return
            try:
                filename_var.set(os.path.relpath(path, source_dir))
            except ValueError:
                filename_var.set(os.path.basename(path))

        def add_or_update() -> None:
            category_label = category_var.get().strip()
            category = category_codes.get(category_label, category_label)
            filename = filename_var.get().strip()
            if not category or not filename:
                messagebox.showwarning(t("missing_data_title"), t("source_order_missing_data_msg"))
                return
            description = description_var.get().strip() or DEFAULT_DESCRIPTIONS.get(category, "")
            enabled = "1" if enabled_var.get() else "0"
            displayed_category = category_labels.get(category, category)
            item_id = selected_item()
            if item_id:
                values = list(tree.item(item_id, "values"))
                values[1:5] = [displayed_category, filename, enabled, description]
                values[5] = status_labels["recognized"]
                values[6] = detection_labels["manual"]
                values[8] = t("manifest_user_confirmed_reason")
                tree.item(item_id, values=values)
            else:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        "", displayed_category, filename, enabled, description,
                        status_labels["recognized"], detection_labels["manual"], "",
                        t("manifest_user_confirmed_reason"),
                    ),
                )
            refresh_order_numbers()
            update_summary()

        def ignore_selected() -> None:
            item_id = selected_item()
            if not item_id:
                messagebox.showwarning(t("no_file_selected_title"), t("source_order_no_file_selected_msg"))
                return
            values = list(tree.item(item_id, "values"))
            values[1] = ""
            values[3] = "0"
            values[5] = status_labels["ignored"]
            values[6] = detection_labels["manual"]
            values[8] = t("manifest_user_ignored_reason")
            tree.item(item_id, values=values)
            update_summary()

        def remove_selected() -> None:
            item_id = selected_item()
            if item_id:
                tree.delete(item_id)
                refresh_order_numbers()
                update_summary()

        def move_selected(delta: int) -> None:
            item_id = selected_item()
            if not item_id:
                return
            siblings = list(tree.get_children())
            current_index = siblings.index(item_id)
            target_index = current_index + delta
            if target_index < 0 or target_index >= len(siblings):
                return
            tree.move(item_id, "", target_index)
            tree.selection_set(item_id)
            refresh_order_numbers()

        def save_manifest() -> None:
            try:
                saved_path = write_source_manifest_xlsx(source_dir, rows_from_tree())
                self.log(t("manifest_saved_notification", path=saved_path))
                messagebox.showinfo(t("manifest_save_success_title"), t("manifest_save_success_msg", path=saved_path))
                self._mark_preflight_stale(force_refresh=True)
            except Exception as exc:
                messagebox.showerror(t("manifest_save_err_title"), t("manifest_save_err_msg", error=str(exc)))

        tree.bind("<<TreeviewSelect>>", fill_form_from_selection)
        load_rows()

        buttons = ttk.Frame(frame)
        buttons.grid(row=4, column=0, columnspan=6, sticky="w", pady=(10, 0))
        ttk.Button(buttons, text=t("btn_add_update"), command=add_or_update).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_confirm_type"), command=add_or_update).grid(row=0, column=1, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_ignore_file"), command=ignore_selected).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_delete_row"), command=remove_selected).grid(row=0, column=3, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_move_up"), command=lambda: move_selected(-1)).grid(row=0, column=4, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_move_down"), command=lambda: move_selected(1)).grid(row=0, column=5, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_save"), command=save_manifest).grid(row=0, column=6, padx=(0, 6))
        ttk.Button(buttons, text=t("btn_close"), command=editor.destroy).grid(row=0, column=7, padx=(0, 6))

        frame.rowconfigure(2, weight=1)
        frame.columnconfigure(2, weight=1)

    @property
    def ALL_COST_CENTERS_LABEL(self) -> str:
        return t("all_cost_centers_label")

    def _update_cc_selection_summary(self) -> None:
        selected = list(getattr(self, "_selected_cc_values", []))
        available = list(getattr(self, "_available_cc_choices", []))
        if not selected:
            summary = t("no_cc_selected")
        elif available and len(selected) == len(available):
            summary = t("all_cc_selected", count=len(available))
        elif len(selected) == 1:
            summary = selected[0]
        else:
            summary = t("selected_cc_count", count=len(selected))
        self.cc_code_filter.set(summary)

    def _set_cc_choices(self, choices) -> None:
        """Refresh available CCs while retaining every still-valid selection."""
        values = list(dict.fromkeys(str(choice).strip() for choice in choices if str(choice).strip()))
        current = list(getattr(self, "_selected_cc_values", []))
        self._available_cc_choices = values
        self._selected_cc_values = [choice for choice in current if choice in values]
        self._update_cc_selection_summary()

    def _open_cc_selection_dialog(self) -> None:
        choices = list(getattr(self, "_available_cc_choices", []))
        if not choices:
            messagebox.showwarning(
                t("no_cc_warning_title"),
                t("no_cc_warning_msg"),
            )
            return

        dialog = tk.Toplevel(self.root)
        dialog.title(t("cc_dialog_title"))
        dialog.geometry("660x600")
        dialog.minsize(520, 420)
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(
            dialog,
            text=t("cc_dialog_heading"),
            font=("Segoe UI", 14, "bold"),
        ).pack(anchor="w", padx=18, pady=(16, 4))
        ttk.Label(
            dialog,
            text=t("cc_dialog_desc"),
        ).pack(anchor="w", padx=18, pady=(0, 10))

        search_shell = ttk.Frame(dialog)
        search_shell.pack(fill="x", padx=18, pady=(0, 8))
        ttk.Label(search_shell, text=t("cc_search_label")).pack(side="left")
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_shell, textvariable=search_var, width=38)
        search_entry.pack(side="left", padx=(6, 6), fill="x", expand=True)
        search_result = ttk.Label(search_shell, text=t("cc_search_matches", count=len(choices)))
        search_result.pack(side="left", padx=(0, 6))

        list_shell = ttk.Frame(dialog)
        list_shell.pack(fill="both", expand=True, padx=18)
        canvas = tk.Canvas(list_shell, highlightthickness=1, highlightbackground="#cbd5e1")
        scrollbar = ttk.Scrollbar(list_shell, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        checklist = ttk.Frame(canvas, padding=8)
        checklist_window = canvas.create_window((0, 0), window=checklist, anchor="nw")
        checklist.bind(
            "<Configure>",
            lambda _event: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.bind(
            "<Configure>",
            lambda event: canvas.itemconfigure(checklist_window, width=event.width),
        )

        selected = set(getattr(self, "_selected_cc_values", []))
        variables = {}
        for choice in choices:
            variable = tk.BooleanVar(value=choice in selected)
            variables[choice] = variable
        checkbuttons = {}

        def render_filtered_choices(*_args) -> None:
            for checkbutton in checkbuttons.values():
                checkbutton.destroy()
            checkbuttons.clear()
            visible_choices = _filter_cost_center_choices(choices, search_var.get())
            for row, choice in enumerate(visible_choices):
                checkbutton = ttk.Checkbutton(checklist, text=choice, variable=variables[choice])
                checkbutton.grid(row=row, column=0, sticky="w", pady=2)
                checkbuttons[choice] = checkbutton
            search_result.configure(text=t("cc_search_matches", count=len(visible_choices)))
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.yview_moveto(0)

        def clear_search() -> None:
            search_var.set("")
            search_entry.focus_set()

        ttk.Button(search_shell, text=t("cc_search_clear"), command=clear_search).pack(side="right")
        search_var.trace_add("write", render_filtered_choices)
        render_filtered_choices()
        search_entry.focus_set()

        def set_all(value: bool) -> None:
            for variable in variables.values():
                variable.set(value)

        def apply_selection() -> None:
            self._selected_cc_values = [
                choice for choice in choices if variables[choice].get()
            ]
            self._update_cc_selection_summary()
            self._mark_preflight_stale()
            dialog.destroy()

        actions = ttk.Frame(dialog)
        actions.pack(fill="x", padx=18, pady=14)
        ttk.Button(actions, text=t("btn_select_all"), command=lambda: set_all(True)).pack(side="left")
        ttk.Button(actions, text=t("btn_deselect_all"), command=lambda: set_all(False)).pack(
            side="left", padx=(6, 0)
        )
        ttk.Button(actions, text=t("btn_cancel"), command=dialog.destroy).pack(side="right")
        ttk.Button(actions, text=t("btn_apply"), style="Primary.TButton", command=apply_selection).pack(
            side="right", padx=(0, 6)
        )
        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)

    def load_cc_list(self):
        db_path = self._operational_database()

        if not os.path.exists(db_path):
            self._set_cc_choices([])
            self.log(t("cc_no_base_data"))
            return

        conn = None
        try:
            conn = get_connection(db_path)
            rows = conn.execute("SELECT code, name_jp FROM dim_cost_centers ORDER BY code").fetchall()
            self._set_cc_choices([f"{row['code']} - {row['name_jp']}" for row in rows])
            if not rows:
                self.log(t("cc_empty_in_db"))
        except Exception as exc:
            self._set_cc_choices([])
            self.log(t("cc_load_error", error=exc))
        finally:
            if conn is not None:
                conn.close()

    def refresh_cost_centers_from_form(self):
        """Refresh existing CCs, or seed an empty master from the selected FORM."""
        db_path = self._operational_database()
        conn = None
        self.refresh_btn.configure(state=tk.DISABLED)
        try:
            conn = get_connection(db_path)
            create_schema(conn)
            current_count = int(conn.execute("SELECT COUNT(*) FROM dim_cost_centers").fetchone()[0])
            if current_count:
                rows = conn.execute(
                    "SELECT code, name_jp FROM dim_cost_centers ORDER BY code"
                ).fetchall()
                self._set_cc_choices([f"{row['code']} - {row['name_jp']}" for row in rows])
                self.log(t("cc_refreshed_from_db", count=current_count))
                return

            template = self.template_path.get().strip()
            template_error = _validate_selected_template(template, self._current_fiscal_year())
            if template_error:
                raise ValueError(template_error)

            loaded_count = load_cost_centers(conn, template)
            if loaded_count <= 0:
                raise RuntimeError(t("cc_form_no_valid_cc"))

            rows = conn.execute(
                "SELECT code, name_jp FROM dim_cost_centers ORDER BY code"
            ).fetchall()
            self._set_cc_choices([f"{row['code']} - {row['name_jp']}" for row in rows])
            self.log(t("cc_loaded_from_form", count=loaded_count))
            messagebox.showinfo(
                t("reload_cc_success_title"),
                t("reload_cc_success_msg", count=loaded_count),
            )
        except Exception as exc:
            message = _friendly_error_message(exc)
            self.log(t("cc_cannot_load_from_form", message=message))
            messagebox.showerror(t("reload_cc_failed_title"), message)
        finally:
            if conn is not None:
                conn.close()
            self.refresh_btn.configure(state=tk.NORMAL)

    def auto_init_master_data(self):
        """Automatically load master data if FORM.xlsx is available in current dir."""
        if self.syncing_master: return

        template = self.template_path.get()
        if not os.path.exists(template):
            template = _default_template_path()
            if not os.path.exists(template):
                return

        self.syncing_master = True
        self.log(t("auto_init_heading"))
        self.log(t("auto_init_template", path=template))
        self.log(t("auto_init_source", path=self.source_dir.get() or BASE_DIR))

        def run_sync():
            try:
                db_path = self._operational_database()
                fiscal_year = int(self.fiscal_year.get())
                content_rules = load_runtime_content_rules(BASE_DIR, fiscal_year=fiscal_year)
                load_all(
                    db_path=db_path,
                    template_path=template,
                    fiscal_year=fiscal_year,
                    content_rules=content_rules,
                )
                self.log(t("auto_init_success"))
                self._run_on_ui_thread(lambda: self.root.after(100, self.load_cc_list))
            except Exception as e:
                self.log(t("auto_init_failed", error=e))
            finally:
                self._run_on_ui_thread(setattr, self, "syncing_master", False)

        threading.Thread(target=run_sync, daemon=True).start()

    def _get_cc_choices(self):
        db_path = self._operational_database()
        if not os.path.exists(db_path):
            return []
        conn = get_connection(db_path)
        try:
            rows = conn.execute("SELECT code, name_jp FROM dim_cost_centers ORDER BY code").fetchall()
            return [f"{row['code']} - {row['name_jp']}" for row in rows]
        finally:
            conn.close()

    def _read_manual_headcount_rows(self, csv_path: str):
        if not os.path.exists(csv_path):
            return []
        with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return list(reader)

    def _write_manual_headcount_rows(self, csv_path: str, rows):
        fieldnames = [
            "cc_code",
            "period",
            "headcount_staff",
            "headcount_worker",
            "headcount_male",
            "headcount_female",
            "description",
        ]
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def _read_csv_rows(self, csv_path: str):
        if not os.path.exists(csv_path):
            return []
        with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return list(reader)

    def _write_csv_rows(self, csv_path: str, fieldnames, rows):
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def _read_manual_bus_headcount_rows(self, csv_path: str):
        rows = self._read_csv_rows(csv_path)
        return [
            {column: str(row.get(column, "") or "").strip() for column in BUS_DRIVER_COLUMNS}
            for row in rows
        ]

    def _write_manual_bus_headcount_rows(self, csv_path: str, rows):
        self._write_csv_rows(csv_path, BUS_DRIVER_COLUMNS, rows)

    def open_variance_tab(self):
        if self._focus_existing_editor("_variance_editor"):
            return
        try:
            from src.ui.tabs.variance_tab import VarianceTab
        except ImportError as e:
            from tkinter import messagebox
            messagebox.showerror(t("missing_lib_title"), t("missing_lib_msg"))
            return

        editor = tk.Toplevel(self.root)
        editor.title(t("variance_analysis_btn"))
        width, height = MPManagerApp._initial_window_size(self.root.winfo_screenwidth(), self.root.winfo_screenheight())
        editor.geometry(f"{width}x{height}")
        editor.minsize(800, 600)
        # Keep the YoY work window associated with its parent.  Without this,
        # Windows can put it behind the main MP window after a file dialog or
        # a message box takes focus.
        editor.transient(self.root)
        editor.lift()
        editor.focus_force()
        self._register_singleton_editor("_variance_editor", editor)
        VarianceTab(editor)

    def open_user_guide(self):
        if self._focus_existing_editor("_user_guide_window"):
            return
        guide = tk.Toplevel(self.root)
        guide.title(t("user_guide_title"))
        guide.geometry("920x700")
        guide.minsize(760, 560)
        guide.transient(self.root)
        guide.lift()
        guide.focus_force()
        close_guide = self._register_singleton_editor("_user_guide_window", guide)

        frame = ttk.Frame(guide, padding=14)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=t("user_guide_title"), style="Header.TLabel").pack(anchor="w")
        ttk.Label(
            frame,
            text=t("user_guide_subtitle"),
        ).pack(anchor="w", pady=(2, 10))

        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        quick = ttk.Frame(notebook, padding=14)
        details = ttk.Frame(notebook, padding=8)
        notebook.add(quick, text=t("user_guide_quick_tab"))
        notebook.add(details, text=t("user_guide_detail_tab"))

        diagram = tk.Canvas(quick, height=210, bg="#f7f9fc", highlightthickness=1, highlightbackground="#d7deea")
        diagram.pack(fill="x")
        steps = (
            ("1", t("user_guide_step1"), t("user_guide_step1_hint")),
            ("2", t("user_guide_step2"), t("user_guide_step2_hint")),
            ("3", t("user_guide_step3"), t("user_guide_step3_hint")),
            ("4", t("user_guide_step4"), t("user_guide_step4_hint")),
            ("5", t("user_guide_step5"), t("user_guide_step5_hint")),
        )

        def draw_diagram(_event=None):
            diagram.delete("all")
            width = max(diagram.winfo_width(), 760)
            margin = 70
            gap = (width - margin * 2) / 4
            centers = [margin + gap * index for index in range(5)]
            for index in range(4):
                diagram.create_line(centers[index] + 28, 70, centers[index + 1] - 28, 70, fill="#6b82a6", width=3, arrow=tk.LAST)
            colors = ("#2457a6", "#2457a6", "#176b4d", "#5c3b8c", "#b56a00")
            for center, color, (number, title, subtitle) in zip(centers, colors, steps):
                diagram.create_oval(center - 28, 42, center + 28, 98, fill=color, outline="")
                diagram.create_text(center, 70, text=number, fill="white", font=("Segoe UI", 16, "bold"))
                diagram.create_text(center, 125, text=title, fill="#1f344d", font=("Segoe UI", 9, "bold"))
                diagram.create_text(center, 148, text=subtitle, fill="#556273", font=("Segoe UI", 8))
            diagram.create_text(width / 2, 185, text=t("user_guide_diagram_tip"), fill="#33455d", font=("Segoe UI", 9, "italic"))

        diagram.bind("<Configure>", draw_diagram)

        ttk.Label(quick, text=t("user_guide_color_title"), style="WorkflowTitle.TLabel").pack(anchor="w", pady=(14, 8))
        legend = ttk.Frame(quick)
        legend.pack(fill="x")
        legend_items = (
            ("#dff4ea", "#176b4d", t("user_guide_green_title"), t("user_guide_green_desc")),
            ("#fff1cf", "#855b00", t("user_guide_yellow_title"), t("user_guide_yellow_desc")),
            ("#fde4e1", "#9b2c24", t("user_guide_red_title"), t("user_guide_red_desc")),
        )
        for column, (background, foreground, title, body) in enumerate(legend_items):
            legend.columnconfigure(column, weight=1, uniform="legend")
            card = tk.Frame(legend, bg=background, highlightthickness=1, highlightbackground=foreground)
            card.grid(row=0, column=column, sticky="nsew", padx=(0 if column == 0 else 6, 0), ipadx=8, ipady=8)
            tk.Label(card, text=title, bg=background, fg=foreground, font=("Segoe UI", 9, "bold"), wraplength=230, justify="left").pack(anchor="w")
            tk.Label(card, text=body, bg=background, fg=foreground, font=("Segoe UI", 8), wraplength=230, justify="left").pack(anchor="w", pady=(4, 0))

        tip = tk.Frame(quick, bg="#e8f1ff", highlightthickness=1, highlightbackground="#6b82a6")
        tip.pack(fill="x", pady=(14, 0), ipady=8)
        tk.Label(tip, text=t("user_guide_safety_tip_title"), bg="#e8f1ff", fg="#2457a6", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10)
        tk.Label(tip, text=t("user_guide_safety_tip_body"), bg="#e8f1ff", fg="#2457a6", font=("Segoe UI", 9)).pack(anchor="w", padx=10, pady=(2, 0))

        search_bar = ttk.Frame(details)
        search_bar.pack(fill="x", pady=(0, 8))
        ttk.Label(search_bar, text=t("user_guide_search_label")).pack(side="left")
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_bar, textvariable=search_var)
        search_entry.pack(side="left", fill="x", expand=True, padx=(8, 6))
        search_status = ttk.Label(search_bar, text=t("user_guide_search_hint"))
        search_status.pack(side="right", padx=(8, 0))

        suggestions = ttk.Frame(details)
        suggestions.pack(fill="x", pady=(0, 8))
        ttk.Label(suggestions, text=t("user_guide_suggestions_label")).pack(side="left")
        for keyword in get_user_guide_search_suggestions():
            ttk.Button(
                suggestions,
                text=keyword,
                command=lambda value=keyword: search_var.set(value),
            ).pack(side="left", padx=(6, 0))

        guide_text = scrolledtext.ScrolledText(details, wrap=tk.WORD, font=("Segoe UI", 10))
        guide_text.pack(fill=tk.BOTH, expand=True)

        def refresh_search(*_args):
            rendered, count = filter_user_guide_text(get_user_guide_text(), search_var.get())
            guide_text.configure(state=tk.NORMAL)
            guide_text.delete("1.0", tk.END)
            guide_text.insert("1.0", rendered)
            guide_text.configure(state=tk.DISABLED)
            guide_text.yview_moveto(0.0)
            if search_var.get().strip():
                search_status.configure(text=t("user_guide_search_match", count=count) if count else t("user_guide_search_no_match"))
            else:
                search_status.configure(text=t("user_guide_search_hint"))

        search_var.trace_add("write", refresh_search)
        refresh_search()

        ttk.Button(frame, text=t("user_guide_close_btn"), command=close_guide).pack(anchor="e", pady=(10, 0))

    def open_run_history(self):
        if self._focus_existing_editor("_run_history_window"):
            return
        history_root = self._project_paths().history_root
        dialog = tk.Toplevel(self.root)
        dialog.title(t("run_history_dialog_title"))
        dialog.geometry("1180x620")
        dialog.transient(self.root)
        dialog.lift()
        dialog.focus_force()
        self._register_singleton_editor("_run_history_window", dialog)
        frame = ttk.Frame(dialog, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)
        fiscal_var = tk.StringVar(value=str(self._current_fiscal_year()))
        status_labels = {
            "": t("history_status_all"),
            "PRECHECK_FAILED": t("history_status_precheck_failed"),
            "RUNNING": t("history_status_running"),
            "SUCCEEDED": t("history_status_succeeded"),
            "SUCCEEDED_INCOMPLETE": t("history_status_succeeded_incomplete"),
            "FAILED": t("history_status_failed"),
            "LEGACY_FY2027": t("history_status_legacy"),
        }
        status_values = {label: code for code, label in status_labels.items()}
        status_var = tk.StringVar(value=t("history_status_all"))
        cc_var = tk.StringVar()
        item_var = tk.StringVar()
        date_var = tk.StringVar()
        filters = ttk.Frame(frame)
        filters.pack(fill=tk.X, pady=(0, 8))
        for index, (label, variable, width) in enumerate((
            (t("history_filter_fiscal_year"), fiscal_var, 10),
            (t("history_filter_status"), status_var, 18),
            (t("history_filter_cc"), cc_var, 16),
            (t("history_filter_item"), item_var, 18),
            (t("history_filter_date"), date_var, 16),
        )):
            ttk.Label(filters, text=label).grid(row=0, column=index * 2, sticky="w", padx=(0 if index == 0 else 8, 3))
            if label == t("history_filter_status"):
                widget = ttk.Combobox(
                    filters,
                    textvariable=variable,
                    width=width,
                    state="readonly",
                    values=tuple(status_values),
                )
            else:
                widget = ttk.Entry(filters, textvariable=variable, width=width)
            widget.grid(row=0, column=index * 2 + 1, sticky="w")
        columns = ("run_id", "status", "started_at", "finished_at", "selected_cost_center", "output_path", "error_summary")
        table = ttk.Treeview(frame, columns=columns, show="headings")
        widths = {"run_id": 170, "status": 170, "started_at": 150, "finished_at": 150, "selected_cost_center": 120, "output_path": 210, "error_summary": 260}
        labels = {"run_id": t("history_col_run_id"), "status": t("history_col_status"), "started_at": t("history_col_started"), "finished_at": t("history_col_finished"), "selected_cost_center": t("history_col_cc"), "output_path": t("history_col_output"), "error_summary": t("history_col_error")}
        for column in columns:
            table.heading(column, text=labels[column])
            table.column(column, width=widths[column], stretch=True)
        table.pack(fill=tk.BOTH, expand=True)
        selected_rows: dict[str, dict[str, object]] = {}

        assistant_button = None

        def on_table_select(_event=None) -> None:
            if assistant_button is None or not dialog.winfo_exists():
                return
            row = selected()
            if not row:
                assistant_button.configure(state=tk.DISABLED)
                return
            status = str(row.get("status") or "").strip().upper()
            if status in ("FAILED", "PRECHECK_FAILED", "SUCCEEDED", "SUCCEEDED_INCOMPLETE", "LEGACY_FY2027"):
                assistant_button.configure(state=tk.NORMAL)
            else:
                assistant_button.configure(state=tk.DISABLED)

        table.bind("<<TreeviewSelect>>", on_table_select)

        def render_rows(rows: list[dict[str, object]]) -> None:
            if not dialog.winfo_exists():
                return
            for node in table.get_children():
                table.delete(node)
            selected_rows.clear()
            for row in rows:
                values = [
                    status_labels.get(str(row.get(column) or ""), str(row.get(column) or ""))
                    if column == "status" else str(row.get(column) or "")
                    for column in columns
                ]
                node = table.insert("", tk.END, values=values)
                selected_rows[node] = row
            on_table_select()

        filter_button = None
        filter_token = {"value": 0}

        def finish_refresh(token: int, rows=None, error: Exception | None = None) -> None:
            if not dialog.winfo_exists() or token != filter_token["value"]:
                return
            if error is None:
                render_rows(rows or [])
            else:
                messagebox.showerror(
                    t("history_read_failed_title"),
                    str(error),
                    parent=dialog,
                )
            if filter_button is not None:
                filter_button.configure(state=tk.NORMAL, text=t("btn_filter"))

        def refresh():
            nonlocal filter_button
            try:
                fiscal_year = int(fiscal_var.get()) if fiscal_var.get().strip() else None
            except ValueError:
                messagebox.showerror(t("invalid_filter_title"), t("invalid_fiscal_year_number_msg"), parent=dialog)
                return

            filter_token["value"] += 1
            token = filter_token["value"]
            criteria = {
                "status": status_values.get(status_var.get(), ""),
                "cost_center": cc_var.get().strip(),
                "item": item_var.get().strip(),
                "run_date": date_var.get().strip(),
            }
            if filter_button is not None:
                filter_button.configure(state=tk.DISABLED, text=t("btn_filtering"))

            def worker() -> None:
                try:
                    rows = filter_runs(history_root, fiscal_year, **criteria)
                except Exception as exc:
                    self._run_on_ui_thread(finish_refresh, token, None, exc)
                else:
                    self._run_on_ui_thread(finish_refresh, token, rows, None)

            threading.Thread(target=worker, daemon=True).start()

        def selected() -> dict[str, object] | None:
            nodes = table.selection()
            return selected_rows.get(nodes[0]) if nodes else None

        def open_path(path: str):
            if not path or not os.path.exists(path):
                messagebox.showerror(t("not_found_title"), t("run_file_not_found_msg"), parent=dialog)
                return
            os.startfile(path)  # type: ignore[attr-defined]

        def open_output():
            row = selected()
            if row:
                open_path(str(row.get("output_path") or ""))

        def open_run_file(relative: str):
            row = selected()
            if not row:
                return
            database_path = str(row.get("database_path") or "")
            workspace = os.path.dirname(database_path) if database_path else ""
            open_path(os.path.join(workspace, relative))

        def open_operations_assistant():
            row = selected()
            if not row:
                messagebox.showwarning(
                    t("operations_assistant_window_title"),
                    t("operations_assistant_no_run_selected"),
                    parent=dialog,
                )
                return

            status = str(row.get("status") or "").strip().upper()
            if status == "RUNNING" or status not in ("FAILED", "PRECHECK_FAILED", "SUCCEEDED", "SUCCEEDED_INCOMPLETE", "LEGACY_FY2027"):
                messagebox.showinfo(
                    t("operations_assistant_window_title"),
                    t("operations_assistant_run_not_finished"),
                    parent=dialog,
                )
                return

            run_id = str(row.get("run_id") or "").strip()
            if not run_id:
                messagebox.showerror(
                    t("operations_assistant_window_title"),
                    t("operations_assistant_unable_to_load_case"),
                    parent=dialog,
                )
                return

            try:
                current_lang = get_current_language()
                case = assemble_operational_case(history_root, run_id, current_lang)
                policy = getattr(self, "cagent_policy", None) or CagentProviderPolicy()
                OperationsBusinessChatDialog.open_with_case(
                    self.root,
                    current_lang,
                    case,
                    policy=policy,
                    history_root=history_root,
                    open_history=lambda: self.open_run_history(initial_cost_center=row.get("selected_cost_center")),
                )
            except Exception:
                messagebox.showerror(
                    t("operations_assistant_window_title"),
                    t("operations_assistant_unable_to_load_case"),
                    parent=dialog,
                )

        filter_button = ttk.Button(filters, text=t("btn_filter"), command=refresh)
        filter_button.grid(row=0, column=10, padx=(12, 0))
        actions = ttk.Frame(frame)
        actions.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(actions, text=t("btn_open_output"), command=open_output).pack(side=tk.LEFT)
        ttk.Button(actions, text=t("btn_open_preflight"), command=lambda: open_run_file(os.path.join("reports", "preflight_report.md"))).pack(side=tk.LEFT, padx=6)
        ttk.Button(actions, text=t("btn_open_uniform_log"), command=lambda: open_run_file("run.db")).pack(side=tk.LEFT)
        ttk.Button(actions, text=t("btn_open_asset_log"), command=lambda: open_run_file("run.db")).pack(side=tk.LEFT, padx=6)
        ttk.Button(actions, text=t("btn_open_run_db"), command=lambda: open_run_file("run.db")).pack(side=tk.LEFT)
        assistant_button = ttk.Button(
            actions,
            text=t("operations_assistant_btn"),
            command=open_operations_assistant,
            state=tk.DISABLED,
        )
        assistant_button.pack(side=tk.LEFT, padx=6)
        refresh()
        ttk.Label(frame, text=t("history_note")).pack(anchor="w", pady=(8, 0))

    def open_headcount_editor(self):
        if self._focus_existing_editor("_legacy_headcount_editor"):
            return
        try:
            fiscal_year = int(self.fiscal_year.get())
        except Exception:
            fiscal_year = 2027

        source_dir = resolve_manual_headcount_source_dir(self.source_dir.get() or BASE_DIR, base_dir=BASE_DIR)
        os.makedirs(source_dir, exist_ok=True)
        csv_path = ensure_manual_headcount_template(source_dir, fiscal_year)

        editor = tk.Toplevel(self.root)
        editor.title(t("manual_headcount_title"))
        editor.geometry("1020x600")
        editor.transient(self.root)
        editor.lift()
        editor.focus_force()
        close_editor = self._register_singleton_editor("_legacy_headcount_editor", editor)

        frame = ttk.Frame(editor, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=t("data_file_label", path=csv_path), font=("Segoe UI", 9, "italic")).grid(
            row=0, column=0, columnspan=6, sticky="w"
        )

        cc_var = tk.StringVar()
        period_var = tk.StringVar()
        staff_var = tk.StringVar()
        worker_var = tk.StringVar()
        desc_var = tk.StringVar()

        cc_choices = self._get_cc_choices()
        periods = get_required_headcount_periods(fiscal_year)

        ttk.Label(frame, text=t("event_label_cc")).grid(row=1, column=0, sticky="w", pady=5)
        cc_combo = ttk.Combobox(frame, textvariable=cc_var, values=cc_choices, width=34)
        cc_combo.grid(row=1, column=1, sticky="w")

        ttk.Label(frame, text=t("hc_time_header_period")).grid(row=1, column=2, sticky="w", pady=5, padx=(8, 0))
        period_combo = ttk.Combobox(frame, textvariable=period_var, values=periods, width=12)
        period_combo.grid(row=1, column=3, sticky="w")

        ttk.Label(frame, text=t("hc_header_staff")).grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=staff_var, width=14).grid(row=2, column=1, sticky="w")
        ttk.Label(frame, text=t("hc_header_worker")).grid(row=2, column=2, sticky="w", pady=5, padx=(8, 0))
        ttk.Entry(frame, textvariable=worker_var, width=14).grid(row=2, column=3, sticky="w")

        ttk.Label(frame, text=t("event_col_desc")).grid(row=3, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=desc_var, width=66).grid(row=3, column=1, columnspan=4, sticky="w")

        columns = ("cc_code", "period", "headcount_staff", "headcount_worker", "description")
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=18)
        for col, width, anchor, text in [
            ("cc_code", 130, "w", t("event_label_cc")),
            ("period", 100, "w", t("event_col_period")),
            ("headcount_staff", 130, "w", t("hc_header_staff")),
            ("headcount_worker", 130, "w", t("hc_header_worker")),
            ("description", 470, "w", t("event_col_desc")),
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor=anchor)
        tree.grid(row=5, column=0, columnspan=6, sticky="nsew", pady=(10, 0))

        scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.grid(row=5, column=6, sticky="ns", pady=(10, 0))

        frame.rowconfigure(5, weight=1)
        frame.columnconfigure(5, weight=1)

        def parse_cc_code(text: str) -> str:
            raw = (text or "").strip()
            if " - " in raw:
                raw = raw.split(" - ")[0].strip()
            return raw

        def clear_inputs():
            cc_var.set("")
            period_var.set("")
            staff_var.set("")
            worker_var.set("")
            desc_var.set("")

        def load_rows():
            for item in tree.get_children():
                tree.delete(item)
            if not os.path.exists(csv_path):
                return
            with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    tree.insert(
                        "",
                        tk.END,
                        values=(
                            str(row.get("cc_code", "")).strip(),
                            str(row.get("period", "")).strip(),
                            str(row.get("headcount_staff", "")).strip(),
                            str(row.get("headcount_worker", "")).strip(),
                            str(row.get("description", "")).strip(),
                        ),
                    )

        def add_or_update():
            cc_code = parse_cc_code(cc_var.get())
            period = period_var.get().strip()
            staff = staff_var.get().strip() or "0"
            worker = worker_var.get().strip() or "0"
            desc = desc_var.get().strip()
            if not cc_code or not period:
                messagebox.showerror(
                    t("missing_info_title"),
                    t("missing_info_msg")
                )
                return
            try:
                int(float(cc_code))
                float(staff)
                float(worker)
            except Exception:
                messagebox.showerror(
                    t("invalid_numeric_title"),
                    t("invalid_numeric_msg")
                )
                return
            selected = tree.selection()
            row_values = (cc_code, period, staff, worker, desc)
            if selected:
                tree.item(selected[0], values=row_values)
            else:
                tree.insert("", tk.END, values=row_values)
            clear_inputs()

        def remove_selected():
            selected = tree.selection()
            for item in selected:
                tree.delete(item)

        def on_select(_event):
            selected = tree.selection()
            if not selected:
                return
            values = tree.item(selected[0], "values")
            if not values:
                return
            cc_var.set(values[0])
            period_var.set(values[1])
            staff_var.set(values[2])
            worker_var.set(values[3])
            desc_var.set(values[4])

        def save_file():
            rows = []
            for item in tree.get_children():
                val = tree.item(item, "values")
                rows.append(
                    {
                        "cc_code": val[0],
                        "period": val[1],
                        "headcount_staff": val[2],
                        "headcount_worker": val[3],
                        "description": val[4],
                    }
                )
            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["cc_code", "period", "headcount_staff", "headcount_worker", "description"],
                )
                writer.writeheader()
                writer.writerows(rows)
            db_path = self._operational_database()
            conn = get_connection(db_path)
            try:
                create_schema(conn)
                result = parse_manual_headcount(conn, source_dir=source_dir)
            finally:
                conn.close()
            self.log(
                t(
                    "hc_saved_manual_summary",
                    rows=len(rows),
                    path=csv_path,
                    inserted=result.get("inserted", 0),
                    errors=result.get("errors", 0),
                )
            )
            messagebox.showinfo(
                t("save_success_title"),
                t("manual_hc_save_success_msg", rows=len(rows), inserted=result.get("inserted", 0), errors=result.get("errors", 0)),
            )

        btn = ttk.Frame(frame)
        btn.grid(row=4, column=0, columnspan=6, sticky="w", pady=(6, 0))
        ttk.Button(btn, text=t("btn_add_update"), command=add_or_update).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btn, text=t("btn_delete_selected"), command=remove_selected).grid(row=0, column=1, padx=(0, 6))
        ttk.Button(btn, text=t("btn_save_file"), command=save_file).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(btn, text=t("btn_close"), command=close_editor).grid(row=0, column=3, padx=(0, 6))

        tree.bind("<<TreeviewSelect>>", on_select)
        load_rows()

    def _focus_existing_editor(self, attribute_name: str) -> bool:
        """Raise a still-open editor instead of creating a conflicting copy."""
        editor = getattr(self, attribute_name, None)
        if editor is None:
            return False
        try:
            if not editor.winfo_exists():
                setattr(self, attribute_name, None)
                return False
            editor.deiconify()
            editor.transient(self.root)
            editor.lift()
            editor.focus_force()
            return True
        except tk.TclError:
            setattr(self, attribute_name, None)
            return False

    def _register_singleton_editor(self, attribute_name: str, editor: tk.Toplevel):
        """Remember an editor and release its slot whether it closes normally or externally."""
        setattr(self, attribute_name, editor)

        def close_editor():
            if getattr(self, attribute_name, None) is editor:
                setattr(self, attribute_name, None)
            if editor.winfo_exists():
                editor.destroy()

        def clear_reference(event):
            if event.widget is editor and getattr(self, attribute_name, None) is editor:
                setattr(self, attribute_name, None)

        editor.protocol("WM_DELETE_WINDOW", close_editor)
        editor.bind("<Destroy>", clear_reference, add="+")
        return close_editor

    def open_headcount_editor_v2(self, selected_cc=None):
        if self._focus_existing_editor("_headcount_editor_v2"):
            return
        try:
            fiscal_year = int(self.fiscal_year.get())
        except Exception:
            fiscal_year = 2027
        periods = get_required_headcount_periods(fiscal_year)
        fy_periods = set(get_fy_months(fiscal_year))
        editor = tk.Toplevel(self.root)
        editor.title(t("headcount_editor_title"))
        editor.geometry("1180x800")
        editor.transient(self.root)
        editor.lift()
        editor.focus_force()
        self._headcount_editor_v2 = editor

        def close_editor():
            if getattr(self, "_headcount_editor_v2", None) is editor:
                self._headcount_editor_v2 = None
            editor.destroy()

        def clear_editor_reference(event):
            if event.widget is editor and getattr(self, "_headcount_editor_v2", None) is editor:
                self._headcount_editor_v2 = None

        editor.protocol("WM_DELETE_WINDOW", close_editor)
        editor.bind("<Destroy>", clear_editor_reference, add="+")
        frame = ttk.Frame(editor, padding=10); frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=t("hc_v2_instruction"), font=("Segoe UI",9,"italic")).pack(anchor="w")
        top = ttk.Frame(frame); top.pack(fill="x", pady=8)
        ttk.Label(top,text=t("hc_v2_cc_label")).pack(side="left")
        cc_var=tk.StringVar(); cc_combo=ttk.Combobox(top,textvariable=cc_var,values=self._get_cc_choices(),width=42,state="readonly"); cc_combo.pack(side="left",padx=6)
        source_status=tk.StringVar(value=t("hc_v2_no_source")); ttk.Label(top,textvariable=source_status).pack(side="left",padx=8)
        bus_exp=tk.StringVar(value="0"); bus_vn=tk.StringVar(value="0"); bus_note=tk.StringVar()
        bus=ttk.LabelFrame(frame,text=t("hc_v2_bus_frame")); bus.pack(fill="x",pady=(0,8))
        for label,var in ((t("hc_v2_bus_expat"),bus_exp),(t("hc_v2_bus_vn"),bus_vn)):
            ttk.Label(bus,text=label).pack(side="left",padx=(8,4)); ttk.Entry(bus,textvariable=var,width=10).pack(side="left")
        ttk.Label(bus,text=t("hc_v2_note")).pack(side="left",padx=(12,4)); ttk.Entry(bus,textvariable=bus_note).pack(side="left",fill="x",expand=True,padx=(0,8))
        notebook=ttk.Notebook(frame); notebook.pack(fill="both",expand=True)
        people=ttk.Frame(notebook,padding=6); fixed=ttk.Frame(notebook,padding=6); overtime=ttk.Frame(notebook,padding=6)
        notebook.add(people, text=t("hc_v2_people_tab"))
        # Tạm ẩn hai tab nhập giờ; giữ nguyên frame, biến và logic để có thể bật lại khi cần.
        # notebook.add(fixed, text="Thời gian cố định")
        # notebook.add(overtime, text="Thời gian tăng ca")
        fields=("expat","staff","worker","g6_to_g5","male","female","total","note"); month_vars={p:{f:tk.StringVar() for f in fields} for p in periods}
        headers=(t("hc_time_header_period"),t("hc_header_expat"),t("hc_header_staff"),t("hc_header_worker"),t("hc_header_g6_to_g5"),"Nam (T12)","Nữ (T12)",t("hc_header_total"),t("hc_header_note"))
        for col,label in enumerate(headers): ttk.Label(people,text=label).grid(row=0,column=col,sticky="w",padx=3)
        def update_total(period):
            vals=month_vars[period]
            try: total=sum(float(vals[k].get() or 0) for k in ("expat","staff","worker")); vals["total"].set(f"{total:g}")
            except ValueError: vals["total"].set("?")
        for row,period in enumerate(periods,1):
            label=t("hc_month_label", month=int(period[-2:])) if period in fy_periods else t("hc_baseline_label", period=period)
            ttk.Label(people,text=label).grid(row=row,column=0,sticky="w",padx=3,pady=2)
            for col,key,width in ((1,"expat",9),(2,"staff",11),(3,"worker",11),(4,"g6_to_g5",10),(5,"male",10),(6,"female",10)):
                entry=ttk.Entry(people,textvariable=month_vars[period][key],width=width); entry.grid(row=row,column=col,padx=3,pady=2)
                if key in ("expat","staff","worker") and period in fy_periods: entry.state(["readonly"])
                if key == "g6_to_g5" and period not in fy_periods: entry.state(["disabled"])
                if key in ("male","female") and not period.endswith("12"): entry.state(["disabled"])
                if key in ("expat","staff","worker"): month_vars[period][key].trace_add("write",lambda *_args,p=period:update_total(p))
            total=ttk.Entry(people,textvariable=month_vars[period]["total"],width=11,state="readonly"); total.grid(row=row,column=7,padx=3)
            ttk.Entry(people,textvariable=month_vars[period]["note"],width=42).grid(row=row,column=8,sticky="ew",padx=3)
            if period not in fy_periods:
                ttk.Label(people,text=t("hc_baseline_hint"),foreground="#8A4B08").grid(row=row,column=9,sticky="w",padx=4)
        people.columnconfigure(8,weight=1)
        time_fields=("fixed_hours_expat","fixed_hours_local","overtime_hours_expat","overtime_hours_local")
        time_vars={p:{f:tk.StringVar() for f in time_fields} for p in get_fy_months(fiscal_year)}
        ttk.Label(fixed,text=t("hc_hours_optional"),font=("Segoe UI",9,"italic")).grid(row=0,column=0,columnspan=4,sticky="w",pady=(0,6))
        ttk.Label(overtime,text=t("hc_hours_optional"),font=("Segoe UI",9,"italic")).grid(row=0,column=0,columnspan=4,sticky="w",pady=(0,6))
        for tab in (fixed,overtime):
            for col,label in enumerate((t("hc_time_header_period"),t("hc_time_header_expat"),t("hc_time_header_local"),t("hc_time_header_total"))):
                ttk.Label(tab,text=label).grid(row=1,column=col,sticky="w",padx=4)
        def update_time_total(period, kind, total_var):
            jp_key=f"{kind}_hours_expat"; local_key=f"{kind}_hours_local"
            try: total_var.set(f"{float(time_vars[period][jp_key].get() or 0)+float(time_vars[period][local_key].get() or 0):g}")
            except ValueError: total_var.set("?")
        for row,period in enumerate(get_fy_months(fiscal_year),2):
            for tab,kind in ((fixed,"fixed"),(overtime,"overtime")):
                ttk.Label(tab,text=t("hc_time_month_label", month=int(period[-2:]), period=period)).grid(row=row,column=0,sticky="w",padx=4,pady=2)
                jp_key=f"{kind}_hours_expat"; local_key=f"{kind}_hours_local"; total_var=tk.StringVar(value="0")
                ttk.Entry(tab,textvariable=time_vars[period][jp_key],width=18).grid(row=row,column=1,padx=4,pady=2)
                ttk.Entry(tab,textvariable=time_vars[period][local_key],width=18).grid(row=row,column=2,padx=4,pady=2)
                ttk.Entry(tab,textvariable=total_var,width=18,state="readonly").grid(row=row,column=3,padx=4,pady=2)
                time_vars[period][jp_key].trace_add("write",lambda *_args,p=period,k=kind,t=total_var:update_time_total(p,k,t))
                time_vars[period][local_key].trace_add("write",lambda *_args,p=period,k=kind,t=total_var:update_time_total(p,k,t))
        for tab in (fixed,overtime): tab.columnconfigure(4,weight=1)
        def cc_code():
            raw=cc_var.get().strip(); return raw.split(" - ")[0] if " - " in raw else raw
        def clear():
            for vals in month_vars.values():
                for var in vals.values(): var.set("")
            for vals in time_vars.values():
                for var in vals.values(): var.set("")
            bus_exp.set("0"); bus_vn.set("0"); bus_note.set("")
        load_request = {"id": 0}

        def load_cc(*_):
            """Load one CC without blocking Tkinter while SQLite is busy."""
            load_request["id"] += 1
            request_id = load_request["id"]
            clear()
            cc = cc_code()
            if not cc:
                source_status.set(t("hc_no_cc_selected"))
                return

            source_status.set(t("hc_loading_cc", cc=cc))
            cc_combo.state(["disabled"])
            load_button.state(["disabled"])

            def read_cc_data():
                source_conn = None
                manual_conn = None
                try:
                    source_conn = get_connection(self._operational_database())
                    manual_conn = get_connection(self._manual_input_store(fiscal_year))
                    create_schema(source_conn)
                    create_schema(manual_conn)
                    fy_period_list = get_fy_months(fiscal_year)
                    period_placeholders = ",".join("?" for _ in fy_period_list)
                    source_rows = [
                        dict(row) for row in source_conn.execute(
                            f"""SELECT * FROM fact_monthly_headcount
                            WHERE CAST(cc_code AS TEXT)=? AND source='department_plan'
                            AND period IN ({period_placeholders}) ORDER BY period""",
                            (cc, *fy_period_list),
                        ).fetchall()
                    ]
                    manual_periods = periods
                    manual_placeholders = ",".join("?" for _ in manual_periods)
                    manual = {
                        row["period"]: dict(row)
                        for row in manual_conn.execute(
                            f"""SELECT * FROM fact_monthly_headcount
                            WHERE CAST(cc_code AS TEXT)=? AND source='manual'
                            AND period IN ({manual_placeholders})""",
                            (cc, *manual_periods),
                        ).fetchall()
                    }
                    busrow = manual_conn.execute(
                        "SELECT * FROM fact_bus_headcount_drivers WHERE cc_code=? AND fiscal_year=?",
                        (cc, fiscal_year),
                    ).fetchone()
                    timerows = [
                        dict(row) for row in source_conn.execute(
                            f"""SELECT * FROM fact_headcount_time_source
                            WHERE CAST(cc_code AS TEXT)=? AND period IN ({period_placeholders}) ORDER BY period""",
                            (cc, *fy_period_list),
                        ).fetchall()
                    ]
                    time_overrides = {
                        row["period"]: dict(row)
                        for row in manual_conn.execute(
                            f"""SELECT * FROM fact_manual_headcount_time_override
                            WHERE fiscal_year=? AND CAST(cc_code AS TEXT)=?
                            AND period IN ({period_placeholders}) ORDER BY period""",
                            (fiscal_year, cc, *fy_period_list),
                        ).fetchall()
                    }
                    transitions = {
                        row["period"]: dict(row)
                        for row in manual_conn.execute(
                            f"""SELECT * FROM fact_manual_g6_to_g5_transition
                            WHERE fiscal_year=? AND CAST(cc_code AS TEXT)=?
                            AND period IN ({period_placeholders}) ORDER BY period""",
                            (fiscal_year, cc, *fy_period_list),
                        ).fetchall()
                    }
                    return {
                        "source_rows": source_rows,
                        "manual": manual,
                        "busrow": dict(busrow) if busrow else None,
                        "timerows": timerows,
                        "time_overrides": time_overrides,
                        "transitions": transitions,
                    }
                finally:
                    if manual_conn is not None:
                        manual_conn.close()
                    if source_conn is not None:
                        source_conn.close()

            def apply_result(result=None, error=None):
                if request_id != load_request["id"] or not editor.winfo_exists():
                    return
                cc_combo.state(["!disabled"])
                load_button.state(["!disabled"])
                if error is not None:
                    source_status.set(t("hc_load_error", cc=cc, error=_friendly_error_message(error)))
                    return
                for row in result["source_rows"]:
                    period = row["period"]
                    if period in month_vars:
                        values = month_vars[period]
                        values["expat"].set(f"{float(row['headcount_expat'] or 0):g}")
                        values["staff"].set(f"{float(row['headcount_staff'] or 0):g}")
                        values["worker"].set(f"{float(row['headcount_worker'] or 0):g}")
                for period, row in result["manual"].items():
                    if period not in month_vars:
                        continue
                    values = month_vars[period]
                    if period not in fy_periods:
                        values["expat"].set(f"{float(row['headcount_expat'] or 0):g}")
                        values["staff"].set(f"{float(row['headcount_staff'] or 0):g}")
                        values["worker"].set(f"{float(row['headcount_worker'] or 0):g}")
                    values["male"].set(f"{float(row['headcount_male'] or 0):g}" if period.endswith("12") else "")
                    values["female"].set(f"{float(row['headcount_female'] or 0):g}" if period.endswith("12") else "")
                    values["note"].set(row["description"] or "")
                busrow = result["busrow"]
                if busrow:
                    bus_exp.set(f"{float(busrow['bus_expat_count'] or 0):g}")
                    bus_vn.set(f"{float(busrow['bus_vietnamese_count'] or 0):g}")
                    bus_note.set(busrow["description"] or "")
                for row in result["timerows"]:
                    period = row["period"]
                    if period in time_vars:
                        for key in time_fields:
                            time_vars[period][key].set(f"{float(row[key] or 0):g}")
                for period, row in result["time_overrides"].items():
                    if period in time_vars:
                        for key in time_fields:
                            time_vars[period][key].set(f"{float(row[key] or 0):g}")
                for period, row in result["transitions"].items():
                    if period in month_vars:
                        month_vars[period]["g6_to_g5"].set(f"{float(row['transition_count'] or 0):g}")
                source_status.set(
                    t("hc_has_source_data", count=len(result['source_rows']), fy=fiscal_year)
                    if result["source_rows"]
                    else t("hc_no_source_data", fy=fiscal_year)
                )

            def worker():
                try:
                    result = read_cc_data()
                except Exception as exc:
                    self._run_on_ui_thread(apply_result, None, exc)
                else:
                    self._run_on_ui_thread(apply_result, result)

            threading.Thread(target=worker, daemon=True).start()
        def nonneg(text,label):
            value = str(text or "").strip() or "0"
            if not value.isdecimal():
                raise ValueError(t("hc_err_non_negative_int", label=label))
            return float(value)
        def save():
            cc = cc_code()
            if not cc:
                return
            try:
                be = nonneg(bus_exp.get(), t("event_name_bus_jp"))
                bv = nonneg(bus_vn.get(), t("event_name_bus_vn"))
                month_values = {
                    period: {
                        "expat": values["expat"].get(),
                        "staff": values["staff"].get(),
                        "worker": values["worker"].get(),
                        "male": values["male"].get(),
                        "female": values["female"].get(),
                        "description": values["note"].get(),
                    }
                    for period, values in month_vars.items()
                }
                _, headcount_errors = validate_headcount_save_period_rows(
                    periods,
                    month_values,
                    {period: (t("hc_val_month_label", month=int(period[-2:])) if period in fy_periods else t("hc_val_baseline_label", period=period)) for period in periods},
                )
                if headcount_errors:
                    messagebox.showerror(t("invalid_data_title"), format_headcount_save_errors(headcount_errors))
                    return
            except ValueError as exc:
                messagebox.showerror(t("invalid_data_title"), _friendly_error_message(exc))
                return
            conn=get_connection(self._manual_input_store(fiscal_year)); create_schema(conn)
            try:
                with conn:
                    conn.execute("INSERT INTO fact_bus_headcount_drivers(cc_code,fiscal_year,bus_expat_count,bus_vietnamese_count,source,description) VALUES(?,?,?,?,'manual',?) ON CONFLICT(cc_code) DO UPDATE SET fiscal_year=excluded.fiscal_year,bus_expat_count=excluded.bus_expat_count,bus_vietnamese_count=excluded.bus_vietnamese_count,description=excluded.description",(cc,fiscal_year,be,bv,bus_note.get().strip()))
                    baseline_period = periods[0]
                    baseline = month_vars[baseline_period]
                    expat = nonneg(baseline["expat"].get(), f"{t('hc_field_expat_short')} {baseline_period}")
                    staff = nonneg(baseline["staff"].get(), f"{t('hc_field_staff_short')} {baseline_period}")
                    worker = nonneg(baseline["worker"].get(), f"{t('hc_field_worker_short')} {baseline_period}")
                    save_manual_baseline_override(conn,fiscal_year,cc,expat,staff,worker,baseline["note"].get().strip())
                    for period, v in month_vars.items():
                        if period not in fy_periods:
                            continue
                        male = nonneg(v["male"].get(), f"{t('hc_field_male_short')} {period}") if period.endswith("12") else 0
                        female = nonneg(v["female"].get(), f"{t('hc_field_female_short')} {period}") if period.endswith("12") else 0
                        note = v["note"].get().strip()
                        conn.execute("DELETE FROM fact_monthly_headcount WHERE cc_code=? AND period=? AND source='manual'",(cc,period))
                        if male or female or note:
                            conn.execute("INSERT INTO fact_monthly_headcount(period,cc_code,headcount_all,headcount_expat,headcount_staff,headcount_worker,headcount_male,headcount_female,source,description) VALUES(?,?,0,0,0,0,?,?,'manual',?)",(period,cc,male,female,note))
                    save_manual_time_overrides(conn,fiscal_year,cc,{p: {f: v[f].get() for f in time_fields} for p, v in time_vars.items()})
                    save_manual_g6_to_g5_transitions(conn,fiscal_year,cc,{p: v["g6_to_g5"].get() for p, v in month_vars.items() if p in fy_periods})
            except ValueError as exc:
                conn.rollback()
                messagebox.showerror(t("invalid_data_title"), _friendly_error_message(exc))
                return
            finally:
                conn.close()
            messagebox.showinfo(t("save_success_title"), t("headcount_save_success_msg"))
        buttons = ttk.Frame(frame)
        buttons.pack(fill="x", pady=(8, 0))
        load_button = ttk.Button(buttons, text=t("hc_v2_load_cc"), command=load_cc)
        load_button.pack(side="left")
        ttk.Button(
            buttons, text=t("hc_v2_save_btn"), style="Primary.TButton", command=save
        ).pack(side="left", padx=6)
        ttk.Button(buttons, text=t("btn_close"), command=close_editor).pack(side="left")
        cc_combo.bind("<<ComboboxSelected>>",load_cc)
        if cc_combo["values"]:
            initial=selected_cc if selected_cc in cc_combo["values"] else cc_combo["values"][0]
            cc_var.set(initial); load_cc()

    def open_event_driver_editor(self):
        if self._focus_existing_editor("_event_driver_editor"):
            return
        try:
            fiscal_year = int(self.fiscal_year.get())
        except Exception:
            fiscal_year = 2027

        source_dir = self.source_dir.get() or BASE_DIR
        os.makedirs(source_dir, exist_ok=True)
        csv_path = ensure_manual_event_drivers_template(source_dir, fiscal_year)
        periods = get_fy_months(fiscal_year)

        editor = tk.Toplevel(self.root)
        editor.title(t("event_driver_title"))
        editor.geometry("1260x760")
        editor.transient(self.root)
        editor.lift()
        editor.focus_force()
        close_editor = self._register_singleton_editor("_event_driver_editor", editor)

        frame = ttk.Frame(editor, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=t("event_file_path_label", path=csv_path), font=("Segoe UI", 9, "italic")).grid(
            row=0, column=0, columnspan=8, sticky="w"
        )
        ttk.Label(
            frame,
            text=t("event_instruction"),
            wraplength=1180,
        ).grid(row=1, column=0, columnspan=8, sticky="w", pady=(4, 6))

        guide = ttk.LabelFrame(frame, text=t("event_guide_frame_title"))
        guide.grid(row=2, column=0, columnspan=8, sticky="ew", pady=(0, 10))
        guide.columnconfigure(0, weight=1)
        ttk.Label(
            guide,
            text=t("event_guide_text"),
            wraplength=1160,
        ).grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
        event_help_var = tk.StringVar(
            value=t("event_default_help")
        )
        ttk.Label(
            guide,
            textvariable=event_help_var,
            foreground="#7a3f00",
            wraplength=1160,
        ).grid(row=1, column=0, sticky="w", padx=8, pady=(0, 6))

        cc_var = tk.StringVar()
        period_var = tk.StringVar()
        event_var = tk.StringVar()
        event_type_var = tk.StringVar(value=t("event_type_month_specific"))
        count_var = tk.StringVar()
        unit_price_var = tk.StringVar()
        unit_price_key_var = tk.StringVar()
        amount_var = tk.StringVar()
        bus_expat_people_var = tk.StringVar(value="0")
        bus_vietnamese_people_var = tk.StringVar(value="0")
        account_var = tk.StringVar()
        account_jp_name_var = tk.StringVar()
        form_row_var = tk.StringVar()
        desc_var = tk.StringVar()

        cc_choices = self._get_cc_choices()
        event_display_to_value = {
            "Câu chuyện của tôi": "My Episode",
            t("event_name_my_episode"): "My Episode",
            t("event_name_bus_jp"): "Xe bus JP",
            t("event_name_bus_vn"): "Xe bus VN",
            t("event_name_visa"): "VISA/Passport dòng khác 137",
        }
        event_value_to_display = {value: key for key, value in event_display_to_value.items()}
        event_type_display_to_value = {
            "Theo số lượng và đơn giá": "manual_count_unit_price",
            "Theo số tiền trực tiếp": "manual_amount",
            "Theo tháng riêng": "month_specific_driver",
            t("event_type_count_unit"): "manual_count_unit_price",
            t("event_type_direct_amount"): "manual_amount",
            t("event_type_month_specific"): "month_specific_driver",
        }
        event_type_value_to_display = {value: key for key, value in event_type_display_to_value.items()}
        event_help = {
            t("event_name_cup"): t("event_help_cup"),
            t("event_name_travel"): t("event_help_travel"),
            t("event_name_no_travel_gift"): t("event_help_no_travel_gift"),
            t("event_name_my_episode"): t("event_help_my_episode"),
            t("event_name_10yr_party"): t("event_help_10yr_party"),
            t("event_name_10yr_gift"): t("event_help_10yr_gift"),
            t("event_name_anniversary"): t("event_help_anniversary"),
            t("event_name_bus_jp"): t("event_help_bus_jp"),
            t("event_name_bus_vn"): t("event_help_bus_vn"),
            t("event_name_mar_prev"): t("event_help_mar_prev"),
            t("event_name_apr_event"): t("event_help_apr_event"),
            t("event_name_visa"): t("event_help_visa"),
            t("event_name_other"): t("event_help_other"),
        }
        event_choices = [
            t("event_name_cup"),
            t("event_name_travel"),
            t("event_name_no_travel_gift"),
            t("event_name_my_episode"),
            t("event_name_10yr_party"),
            t("event_name_10yr_gift"),
            t("event_name_anniversary"),
            t("event_name_bus_jp"),
            t("event_name_bus_vn"),
            t("event_name_mar_prev"),
            t("event_name_apr_event"),
            t("event_name_visa"),
            t("event_name_other"),
        ]

        def add_label_entry(row, column, label, variable, width=18, values=None):
            ttk.Label(frame, text=label).grid(row=row, column=column, sticky="w", padx=(0, 4), pady=3)
            if values is None:
                widget = ttk.Entry(frame, textvariable=variable, width=width)
            else:
                widget = ttk.Combobox(frame, textvariable=variable, values=values, width=width, state="readonly")
            widget.grid(row=row, column=column + 1, sticky="w", padx=(0, 12), pady=3)
            return widget

        add_label_entry(3, 0, t("event_label_cc"), cc_var, width=38, values=cc_choices)
        add_label_entry(3, 2, t("event_label_period"), period_var, width=12, values=periods)
        event_combo = add_label_entry(3, 4, t("event_label_event"), event_var, width=28, values=event_choices)
        add_label_entry(
            3,
            6,
            t("event_label_input_method"),
            event_type_var,
            width=24,
            values=list(event_type_display_to_value),
        )
        add_label_entry(4, 0, t("event_label_count"), count_var, width=16)
        add_label_entry(4, 2, t("event_label_unit_price"), unit_price_var, width=16)
        add_label_entry(4, 4, t("event_label_unit_price_key"), unit_price_key_var, width=24)
        add_label_entry(4, 6, t("event_label_direct_amount"), amount_var, width=18)
        add_label_entry(5, 0, t("event_label_bus_expat"), bus_expat_people_var, width=16)
        add_label_entry(5, 2, t("event_label_bus_vn"), bus_vietnamese_people_var, width=16)
        add_label_entry(6, 0, t("event_label_account"), account_var, width=16)
        add_label_entry(6, 2, t("event_label_account_jp"), account_jp_name_var, width=18)
        add_label_entry(6, 4, t("event_label_form_row"), form_row_var, width=12)
        ttk.Label(frame, text=t("event_label_note")).grid(row=6, column=6, sticky="w", padx=(0, 4), pady=3)
        ttk.Entry(frame, textvariable=desc_var, width=32).grid(row=6, column=7, sticky="w", pady=3)

        columns = tuple(TEMPLATE_COLUMNS)
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=20)
        headings = [
            ("cc_code", 105, t("event_col_cc")),
            ("period", 80, t("event_col_period")),
            ("target_month", 90, t("event_col_month")),
            ("event_name", 170, t("event_col_event")),
            ("event_type", 150, t("event_col_input_method")),
            ("count", 70, t("event_col_count")),
            ("unit_price", 100, t("event_col_unit_price")),
            ("unit_price_key", 120, t("event_col_unit_price_key")),
            ("allocation_content", 130, t("event_col_allocation")),
            ("amount_vnd", 115, t("event_col_amount")),
            ("bus_expat_people", 115, t("event_col_bus_expat")),
            ("bus_vietnamese_people", 115, t("event_col_bus_vn")),
            ("account_code", 95, t("event_col_account_code")),
            ("account_jp_name", 120, t("event_col_account_jp")),
            ("account_name", 120, t("event_col_account_alt")),
            ("account_group", 100, t("event_col_account_group")),
            ("form_row", 75, t("event_col_form_row")),
            ("row", 65, t("event_col_row")),
            ("source_month", 100, t("event_col_source_month")),
            ("headcount_basis", 120, t("event_col_hc_basis")),
            ("description", 180, t("event_col_desc")),
            ("note", 220, t("event_col_note")),
        ]
        for col, width, text in headings:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="w")
        tree.grid(row=12, column=0, columnspan=8, sticky="nsew", pady=(12, 0))
        scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.grid(row=8, column=8, sticky="ns", pady=(12, 0))
        frame.rowconfigure(8, weight=1)
        frame.columnconfigure(7, weight=1)

        def parse_cc_code(text: str) -> str:
            raw = (text or "").strip()
            if " - " in raw:
                raw = raw.split(" - ")[0].strip()
            return raw

        def clear_inputs():
            for variable in (
                cc_var,
                period_var,
                event_var,
                event_type_var,
                count_var,
                unit_price_var,
                unit_price_key_var,
                amount_var,
                bus_expat_people_var,
                bus_vietnamese_people_var,
                account_var,
                account_jp_name_var,
                form_row_var,
                desc_var,
            ):
                variable.set("")
            event_type_var.set(t("event_type_month_specific"))
            bus_expat_people_var.set("0")
            bus_vietnamese_people_var.set("0")

        def load_rows():
            for item in tree.get_children():
                tree.delete(item)
            for row in self._read_csv_rows(csv_path):
                values = []
                for col in columns:
                    value = str(row.get(col, "") or "").strip()
                    if col == "target_month" and not value:
                        value = str(row.get("period", "") or "").strip()
                    elif col == "row" and not value:
                        value = str(row.get("form_row", "") or "").strip()
                    elif col == "note" and not value:
                        value = str(row.get("description", "") or "").strip()
                    if col == "event_name":
                        value = event_value_to_display.get(value, value)
                    elif col == "event_type":
                        value = event_type_value_to_display.get(value, value)
                    values.append(value)
                tree.insert("", tk.END, values=tuple(values))

        def validate_numeric(raw, label, required=False):
            text = str(raw or "").strip()
            if not text:
                if required:
                    raise ValueError(t("event_err_missing_field", label=label))
                return ""
            float(text)
            return text

        def validate_non_negative_int(raw, label):
            text = str(raw or "").strip()
            if not text:
                return "0"
            if not text.isdecimal():
                raise ValueError(t("hc_err_non_negative_int", label=label))
            return str(int(text))

        def add_or_update():
            cc_code = parse_cc_code(cc_var.get())
            period = period_var.get().strip()
            event_name = event_var.get().strip()
            try:
                if not cc_code or not period or not event_name:
                    raise ValueError(t("event_err_required_fields"))
                count = validate_numeric(count_var.get(), t("field_quantity"))
                if event_name in (t("event_name_cup"), "Cốc xếp định kỳ"):
                    count = validate_non_negative_int(count_var.get(), t("event_name_cup"))
                    if int(period[-2:]) not in {2, 8}:
                        raise ValueError(t("event_err_cup_month"))
                unit_price = validate_numeric(unit_price_var.get(), t("field_unit_price"))
                unit_price_key = unit_price_key_var.get().strip()
                amount_vnd = validate_numeric(amount_var.get(), t("field_amount"))
                bus_expat_people = validate_non_negative_int(
                    bus_expat_people_var.get(), t("field_bus_expat")  # Người biệt phái đi xe bus
                )
                bus_vietnamese_people = validate_non_negative_int(
                    bus_vietnamese_people_var.get(), t("field_bus_vn")  # Người Việt Nam đi xe bus
                )
                account_code = validate_numeric(account_var.get(), t("field_account_code"))
                account_jp_name = account_jp_name_var.get().strip()
                form_row = validate_numeric(form_row_var.get(), t("field_form_row"))
                if not account_code and not account_jp_name:
                    raise ValueError(t("event_err_account_required"))
                if not ((count and (unit_price or unit_price_key)) or amount_vnd):
                    raise ValueError(t("event_err_pricing_required"))
            except Exception as exc:
                messagebox.showerror(t("invalid_data_title"), _friendly_error_message(exc))
                return

            row_data = {col: "" for col in columns}
            row_data.update(
                {
                    "cc_code": cc_code,
                    "period": period,
                    "target_month": period,
                    "event_name": event_name,
                    "event_type": event_type_var.get().strip() or t("event_type_month_specific"),
                    "count": count,
                    "unit_price": unit_price,
                    "unit_price_key": unit_price_key,
                    "allocation_content": unit_price_key,
                    "amount_vnd": amount_vnd,
                    "bus_expat_people": bus_expat_people,
                    "bus_vietnamese_people": bus_vietnamese_people,
                    "account_code": account_code,
                    "account_jp_name": account_jp_name,
                    "account_name": account_jp_name,
                    "form_row": form_row,
                    "row": form_row,
                    "description": desc_var.get().strip(),
                    "note": desc_var.get().strip(),
                }
            )
            values = tuple(row_data[col] for col in columns)
            selected = tree.selection()
            if selected:
                tree.item(selected[0], values=values)
            else:
                tree.insert("", tk.END, values=values)
            clear_inputs()

        def remove_selected():
            for item in tree.selection():
                tree.delete(item)

        def on_select(_event):
            selected = tree.selection()
            if not selected:
                return
            values = tree.item(selected[0], "values")
            row_data = {col: str(values[index]) if index < len(values) else "" for index, col in enumerate(columns)}
            cc_var.set(row_data.get("cc_code", ""))
            period_var.set(row_data.get("target_month") or row_data.get("period", ""))
            event_var.set(row_data.get("event_name", ""))
            event_type_var.set(row_data.get("event_type", "") or t("event_type_month_specific"))
            count_var.set(row_data.get("count", ""))
            unit_price_var.set(row_data.get("unit_price", ""))
            unit_price_key_var.set(row_data.get("unit_price_key") or row_data.get("allocation_content", ""))
            amount_var.set(row_data.get("amount_vnd", ""))
            bus_expat_people_var.set(row_data.get("bus_expat_people", "") or "0")
            bus_vietnamese_people_var.set(row_data.get("bus_vietnamese_people", "") or "0")
            account_var.set(row_data.get("account_code", ""))
            account_jp_name_var.set(row_data.get("account_jp_name") or row_data.get("account_name", ""))
            form_row_var.set(row_data.get("row") or row_data.get("form_row", ""))
            desc_var.set(row_data.get("note") or row_data.get("description", ""))

        def save_file():
            rows = []
            for item in tree.get_children():
                values = tree.item(item, "values")
                row = {col: values[index] if index < len(values) else "" for index, col in enumerate(columns)}
                row["event_name"] = event_display_to_value.get(row["event_name"], row["event_name"])
                row["event_type"] = event_type_display_to_value.get(row["event_type"], row["event_type"])
                rows.append(row)
            self._write_csv_rows(csv_path, columns, rows)
            self.log(t("event_saved_log", count=len(rows), path=csv_path))
            messagebox.showinfo(t("save_success_title"), t("event_save_success_msg", count=len(rows)))

        button_row = ttk.Frame(frame)
        button_row.grid(row=9, column=0, columnspan=8, sticky="w", pady=(10, 0))
        ttk.Button(button_row, text=t("btn_add_update"), command=add_or_update).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(button_row, text=t("btn_delete_selected"), command=remove_selected).grid(row=0, column=1, padx=(0, 6))
        ttk.Button(button_row, text=t("btn_save_file"), command=save_file).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(button_row, text=t("btn_close"), command=close_editor).grid(row=0, column=3, padx=(0, 6))

        def refresh_event_help(*_args):
            selected = event_var.get().strip()
            event_help_var.set(event_help.get(selected, t("event_default_help")))
            if selected == t("event_name_cup"):
                event_type_var.set(t("event_type_count_unit"))
                unit_price_var.set("")
                unit_price_key_var.set("折りたたみコップ Cốc xếp")
                account_var.set("")
                account_jp_name_var.set("福利厚生費")
                amount_var.set("")

        event_combo.bind("<<ComboboxSelected>>", refresh_event_help)
        tree.bind("<<TreeviewSelect>>", on_select)
        load_rows()

    def _parse_selected_cc_codes(self) -> tuple[str, ...]:
        selected = list(getattr(self, "_selected_cc_values", []))
        available = set(getattr(self, "_available_cc_choices", []))
        selected = [choice for choice in selected if choice in available]
        if not selected:
            raise ValueError(t("err_select_at_least_one_cc"))
        return tuple(
            choice.split(" - ")[0].strip() if " - " in choice else choice.strip()
            for choice in selected
        )

    def _all_cost_centers_selected(self) -> bool:
        available = list(getattr(self, "_available_cc_choices", []))
        selected = list(getattr(self, "_selected_cc_values", []))
        return bool(available) and len(selected) == len(available) and set(selected) == set(available)

    def _audit_output_dir(self) -> str:
        return self._project_paths().output_dir

    def _current_fiscal_year(self) -> int:
        try:
            return int(self.fiscal_year.get())
        except Exception:
            return 2027

    def _read_missing_inputs(self) -> list[dict[str, str]]:
        missing_path = os.path.join(
            self._audit_output_dir(), "BAO_CAO_KIEM_TRA", "DU_LIEU_CON_THIEU.xlsx"
        )
        if not os.path.exists(missing_path):
            return []
        workbook = load_workbook(missing_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(min_row=7, values_only=True))
            return [
                {
                    "severity": str(row[0] or ""), "cc_code": str(row[1] or ""),
                    "period": str(row[2] or ""), "area": str(row[3] or ""),
                    "message": str(row[3] or ""), "action": str(row[4] or ""),
                }
                for row in rows if any(value not in (None, "") for value in row)
            ]
        finally:
            workbook.close()

    def _manual_event_ccs(self) -> set[str]:
        source_dir = self.source_dir.get() or BASE_DIR
        path = ensure_manual_event_drivers_template(source_dir, self._current_fiscal_year())
        return {
            str(row.get("cc_code", "")).strip()
            for row in self._read_csv_rows(path)
            if str(row.get("cc_code", "")).strip()
        }


    def _confirm_selected_form(self, template: str, fiscal_year: int) -> bool:
        """Allow a usable FORM silently and report only a FORM that cannot run."""
        inspection = inspect_form(template)
        if not inspection.is_valid:
            messagebox.showerror(
                t("invalid_form_structure_title"),
                t("invalid_form_structure_msg"),
            )
            return False
        return True

    def _set_pipeline_ui_busy(self, busy: bool) -> None:
        """Prevent edits or project changes from racing a running calculation."""
        self._pipeline_busy = busy
        state = tk.DISABLED if busy else tk.NORMAL
        for attribute_name in (
            "cc_select_btn",
            "refresh_btn",
            "update_db_btn",
            "deep_scan_btn",
            "open_proj_btn",
            "create_proj_btn",
            "config_proj_btn",
            "fiscal_year_entry",
            "exchange_rate_entry",
            "template_path_entry",
            "source_dir_entry",
            "headcount_source_dir_entry",
        ):
            widget = getattr(self, attribute_name, None)
            if widget is not None:
                widget.configure(state=state)
        for button, _key in getattr(self, "action_buttons", ()):
            button.configure(state=state)
        if hasattr(self, "start_btn"):
            self.start_btn.configure(state=tk.DISABLED if busy else tk.NORMAL)

    def start_pipeline(self):
        try:
            fiscal_year = int(self.fiscal_year.get())
            exchange_rate = validate_exchange_rate(self.exchange_rate.get())

            selected_ccs = self._parse_selected_cc_codes()
            all_cost_centers = self._all_cost_centers_selected()
            run_targets: tuple[str | None, ...] = (None,) if all_cost_centers else selected_ccs

            template = self.template_path.get()
            source = self.source_dir.get()
            template_error = _validate_selected_template(template, fiscal_year)
            if template_error:
                messagebox.showerror(t("invalid_template_title"), template_error)
                return
            if not self._confirm_selected_form(template, fiscal_year):
                return

            source_error = _validate_selected_source_dir(source, fiscal_year)
            if source_error:
                messagebox.showerror(t("invalid_cost_source_title"), source_error)
                return

            if self.syncing_master:
                messagebox.showinfo(
                    t("syncing_data_title"),
                    t("syncing_data_msg"),
                )
                return

            headcount_source = self.headcount_source_dir.get().strip()
            if not os.path.isdir(headcount_source):
                messagebox.showerror(t("error_title"), t("invalid_headcount_dir_msg"))
                return
            coverage = self._selected_headcount_source_coverage(
                fiscal_year,
                headcount_source,
                selected_ccs,
            )
            if coverage["missing_cc_codes"]:
                message = _headcount_coverage_error_message(fiscal_year, coverage)
                self.log(message)
                messagebox.showerror(t("missing_headcount_title"), message)
                self._mark_preflight_stale(force_refresh=True)
                return
            missing_baselines = self._missing_baseline_ccs_for_selection(
                fiscal_year,
                selected_ccs,
            )
            if missing_baselines:
                self.log(t("hc_baseline_missing_before_calc", cc=", ".join(missing_baselines)))
                target_cc = selected_ccs[0] if len(selected_ccs) == 1 else None
                self._open_baseline_recovery_dialog(
                    fiscal_year,
                    target_cc,
                    missing_baselines,
                )
                return
            signature = (
                fiscal_year,
                os.path.abspath(template),
                os.path.abspath(source),
                os.path.abspath(headcount_source),
                float(exchange_rate),
                _uniform_policy_signature(self._project_paths(fiscal_year).uniform_policy_path),
            )
            approved_report = self._approved_preflight_report
            if (
                signature != self._approved_preflight_signature
                or not getattr(approved_report, "can_run", False)
            ):
                messagebox.showerror(
                    t("source_not_verified_title"),
                    t("source_not_verified_msg"),
                )
                self._mark_preflight_stale()
                return

            if getattr(approved_report, "skipped_issues", ()):
                warning_lines = "\n".join(
                    f"• {_localized_preflight_issue_warning(issue)}"
                    for issue in approved_report.skipped_issues
                )
                proceed_incomplete = messagebox.askyesno(
                    t("confirm_incomplete_run_title"),
                    t("confirm_incomplete_run_msg", warnings=warning_lines),
                )
                if not proceed_incomplete:
                    return
            self._accepted_missing_categories = ()

            if all_cost_centers:
                proceed = messagebox.askokcancel(
                    t("export_all_cc_title"),
                    t("export_all_cc_msg", count=len(selected_ccs)),
                )
                if not proceed:
                    return

            self._set_pipeline_ui_busy(True)
            self.log(t("pipeline_start_heading"))
            self.log(t("pipeline_template_log", path=template))
            self.log(t("pipeline_source_log", path=source))
            self.log(t("pipeline_headcount_log", path=headcount_source))
            self.log(t("pipeline_fx_log", rate=exchange_rate))
            if all_cost_centers:
                self.log(t("pipeline_scope_all_log", count=len(selected_ccs)))
            else:
                self.log(t("pipeline_scope_count_log", count=len(selected_ccs)))
            threading.Thread(
                target=self.run_process,
                args=(
                    fiscal_year,
                    template,
                    source,
                    headcount_source,
                    exchange_rate,
                    run_targets,
                ),
                daemon=True,
            ).start()
        except Exception as exc:
            if getattr(self, "_pipeline_busy", False):
                self._set_pipeline_ui_busy(False)
                self._mark_preflight_stale()
            messagebox.showerror(t("input_error_title"), _friendly_error_message(exc))

    def _run_pipeline_process(
        self,
        fiscal_year: int,
        template: str,
        source: str,
        headcount_source: str,
        rate: float,
        target_cc: str | None,
    ) -> tuple[bool, object, str | None]:
        try:
            cmd = self._pipeline_subprocess_command(
                fiscal_year, template, source, headcount_source, rate, target_cc,
                defer_publication=bool(getattr(self, "_defer_batch_publication", False)),
            )
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            process = subprocess.Popen(
                cmd,
                cwd=BASE_DIR,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
            )
            assert process.stdout is not None
            output_lines = []
            for line in process.stdout:
                text = line.rstrip()
                if text:
                    output_lines.append(text)
                    self.log(text)
            return_code = process.wait()
            success = return_code == 0
            failed_database = (
                None
                if success
                else _failed_run_database_from_output(
                    output_lines,
                    self._project_paths(fiscal_year).history_root,
                    fiscal_year,
                )
            )
            deferred_output = next(
                (
                    line.partition("=")[2]
                    for line in reversed(output_lines)
                    if line.startswith("PIPELINE_OUTPUT=")
                ),
                None,
            )
            result = (
                deferred_output or self._project_paths(fiscal_year).output_dir
                if success
                else _pipeline_failure_summary(output_lines, return_code)
            )
            return success, result, failed_database
        except Exception as exc:
            return False, exc, None

    def run_process(
        self,
        fiscal_year: int,
        template: str,
        source: str,
        headcount_source: str,
        rate: float,
        target_ccs: tuple[str | None, ...],
    ):
        total = len(target_ccs)
        final_result = self._project_paths(fiscal_year).output_dir
        selected_batch = total > 0 and all(target_cc is not None for target_cc in target_ccs)
        staged_output_root = None
        if selected_batch:
            staged_output_root = tempfile.mkdtemp(prefix=f"mp{fiscal_year}_cc_batch_")
        try:
            for index, target_cc in enumerate(target_ccs, start=1):
                label = target_cc or self.ALL_COST_CENTERS_LABEL
                self.log(t("pipeline_scope_log", index=index, total=total, label=label))
                self._last_pipeline_args = (
                    fiscal_year,
                    template,
                    source,
                    headcount_source,
                    rate,
                    target_cc,
                )
                self._defer_batch_publication = selected_batch
                success, result, failed_database = self._run_pipeline_process(
                    fiscal_year, template, source, headcount_source, rate, target_cc,
                )
                self._last_failed_run_database = failed_database
                if not success:
                    failure = RuntimeError(t("pipeline_cc_failed", label=label, error=_friendly_error_message(result)))
                    self._run_on_ui_thread(self._finish_pipeline, False, failure)
                    return
                if selected_batch:
                    assert staged_output_root is not None
                    self._stage_completed_cost_center_output(result, staged_output_root, str(target_cc))
                final_result = result
                self.log(t("pipeline_scope_done", index=index, total=total, label=label))
            if selected_batch:
                assert staged_output_root is not None
                final_result = publish_selected_cc_batch(
                    self._project_paths(fiscal_year).output_dir,
                    staged_output_root,
                    tuple(str(target_cc) for target_cc in target_ccs),
                )
                self.log(t("pipeline_batch_published"))
            self._run_on_ui_thread(self._finish_pipeline, True, final_result)
        except Exception as exc:
            self._run_on_ui_thread(self._finish_pipeline, False, exc)
        finally:
            self._defer_batch_publication = False
            if staged_output_root:
                shutil.rmtree(staged_output_root, ignore_errors=True)

    def _pipeline_subprocess_command(
        self,
        fiscal_year: int,
        template: str,
        source: str,
        headcount_source: str,
        rate: float,
        target_cc: int | str | None,
        *,
        defer_publication: bool = False,
    ) -> list[str]:
        if getattr(sys, "frozen", False):
            cmd = [sys.executable]
        else:
            cmd = [sys.executable, os.path.join(BASE_DIR, "scripts", "run_e2e.py")]
        cmd.extend(
            [
                "--fy",
                str(fiscal_year),
                "--template",
                template,
                "--source",
                source,
                "--headcount-source",
                headcount_source,
                "--exchange-rate",
                str(rate),
                "--exchange-rate-source",
                "FORM!B2 / người dùng xác nhận trên giao diện",
            ]
        )
        approved_uniform = getattr(self, "_approved_uniform_policy_path", None)
        paths = self._project_paths(fiscal_year)
        manual_special_inheritance_dir = getattr(paths, "manual_special_inheritance_dir", None)
        manual_special_legacy_starts = getattr(paths, "manual_special_legacy_starts", {}) or {}
        cmd.extend([
            "--operational-db", self._operational_database(),
            "--manual-input-store", paths.manual_input_store,
            "--manual-special-inheritance-dir", manual_special_inheritance_dir or "",
            "--output-dir", paths.output_dir,
            "--run-history-root", paths.history_root,
            "--project-config", self.project.config_path,
        ])
        if not manual_special_inheritance_dir:
            inheritance_index = cmd.index("--manual-special-inheritance-dir")
            del cmd[inheritance_index:inheritance_index + 2]
        for cc_code, start_row in manual_special_legacy_starts.items():
            cmd.extend(["--manual-special-legacy-start", f"{cc_code}:{start_row}"])
        if approved_uniform:
            cmd.extend(["--uniform-policy", str(approved_uniform)])
        if target_cc:
            cmd.extend(["--target-cc", str(target_cc)])
        if defer_publication:
            cmd.append("--defer-publication")
        return cmd

    def _stage_completed_cost_center_output(
        self,
        source_output: str | os.PathLike[str],
        staged_output_root: str | os.PathLike[str],
        target_cc: str,
    ) -> None:
        """Copy one verified private run result into the not-yet-public batch stage."""
        source = os.fspath(source_output)
        workbook_name = f"MP_CC_{target_cc}.xlsx"
        workbook = os.path.join(source, workbook_name)
        if not os.path.isfile(workbook):
            raise FileNotFoundError(f"Không tìm thấy sổ làm việc đã hoàn tất cho CC {target_cc}: {workbook}")
        destination = os.fspath(staged_output_root)
        shutil.copy2(workbook, os.path.join(destination, workbook_name))
        reports = os.path.join(source, "BAO_CAO_KIEM_TRA")
        if os.path.isdir(reports):
            staged_reports = os.path.join(destination, "BAO_CAO_KIEM_TRA")
            if os.path.isdir(staged_reports):
                shutil.rmtree(staged_reports)
            shutil.copytree(reports, staged_reports)

    def _missing_baseline_context(self, result):
        if not _is_missing_baseline_error(result):
            return None
        args=getattr(self,"_last_pipeline_args",None)
        run_database=getattr(self,"_last_failed_run_database",None)
        if not args or not run_database:return None
        fiscal_year,_,_,_,_,target_cc=args
        conn=get_connection(self._manual_input_store(fiscal_year)); create_schema(conn)
        try: missing=find_missing_baseline_ccs(conn,fiscal_year,target_cc=target_cc)
        finally: conn.close()
        return (fiscal_year,target_cc,missing,run_database) if missing else None

    def _missing_baseline_ccs_for_selection(self, fiscal_year, selected_ccs):
        conn=get_connection(self._manual_input_store(fiscal_year)); create_schema(conn)
        try:
            return find_missing_baseline_ccs(
                conn,
                fiscal_year,
                scope_ccs=selected_ccs,
            )
        finally:
            conn.close()

    def _copy_missing_baselines_from_selected_headcount_source(self, fiscal_year, missing_ccs):
        source_conn=sqlite3.connect(":memory:")
        source_conn.row_factory=sqlite3.Row
        create_schema(source_conn)
        try:
            template=self.template_path.get().strip()
            if os.path.isfile(template):
                load_cost_centers(source_conn, template)
            result=import_headcount_time_sources(
                source_conn,
                self.headcount_source_dir.get().strip(),
                fiscal_year,
                required_cc_codes=tuple(missing_ccs),
            )
            unresolved=tuple(result.get("missing_required_cc_codes") or ())
            if unresolved:
                raise ValueError(t("hc_baseline_missing_no_april", cc=", ".join(unresolved)))
            conn=get_connection(self._manual_input_store(fiscal_year)); create_schema(conn)
            try:
                copied=[]
                with conn:
                    for cc in missing_ccs:
                        copied.extend(copy_missing_baselines_from_april(
                            conn,
                            fiscal_year,
                            target_cc=cc,
                            source_conn=source_conn,
                        ))
                return list(dict.fromkeys(copied))
            finally:
                conn.close()
        finally:
            source_conn.close()

    def _copy_missing_baselines_from_failed_run(self, fiscal_year, missing_ccs, run_database):
        source_uri="file:"+os.path.realpath(run_database).replace("\\","/")+"?mode=ro"
        source_conn=sqlite3.connect(source_uri,uri=True)
        conn=get_connection(self._manual_input_store(fiscal_year)); create_schema(conn)
        try:
            copied=[]
            with conn:
                for cc in missing_ccs:
                    copied.extend(copy_missing_baselines_from_april(
                        conn,
                        fiscal_year,
                        target_cc=cc,
                        source_conn=source_conn,
                    ))
            return list(dict.fromkeys(copied))
        finally:
            source_conn.close()
            conn.close()

    def _open_baseline_recovery_dialog(self, fiscal_year, target_cc, missing_ccs, run_database=None):
        dialog=tk.Toplevel(self.root); dialog.title(t("baseline_missing_title")); dialog.geometry("680x300"); dialog.transient(self.root); dialog.grab_set()
        ttk.Label(dialog,text=t("baseline_missing_title"),font=("Segoe UI",15,"bold")).pack(anchor="w",padx=18,pady=(18,6))
        preview=", ".join(missing_ccs[:12])+("…" if len(missing_ccs)>12 else "")
        ttk.Label(dialog,text=t("baseline_missing_desc", ccs=preview),wraplength=640,justify="left").pack(anchor="w",padx=18)
        def use_april():
            try:
                if run_database:
                    copied=self._copy_missing_baselines_from_failed_run(
                        fiscal_year,
                        missing_ccs,
                        run_database,
                    )
                else:
                    copied=self._copy_missing_baselines_from_selected_headcount_source(
                        fiscal_year,
                        missing_ccs,
                    )
            except Exception as exc:
                messagebox.showerror(t("baseline_cannot_use_t4_title"),_friendly_error_message(exc),parent=dialog); return
            unresolved=[cc for cc in missing_ccs if cc not in copied]
            if unresolved:
                messagebox.showerror(t("baseline_cannot_use_t4_title"),t("baseline_cannot_use_t4_msg", ccs=", ".join(unresolved)),parent=dialog); return
            dialog.destroy()
            self.log(t("t4_copy_approved_log", cc=", ".join(copied)))
            self.start_pipeline()
        def manual():
            dialog.destroy()
            if target_cc:
                choices=list(self._get_cc_choices())
                selected=next((item for item in choices if item.split(" - ")[0]==str(target_cc)),None)
                if selected:self.root.after(100,lambda:self.open_headcount_editor_v2(selected))
                else:self.open_headcount_editor_v2()
            else:self.open_headcount_editor_v2()
        buttons=ttk.Frame(dialog); buttons.pack(fill="x",padx=18,pady=22)
        ttk.Button(buttons,text=t("baseline_option_use_april"),command=use_april).pack(fill="x",pady=3)
        ttk.Button(buttons,text=t("baseline_option_manual"),command=manual).pack(fill="x",pady=3)
        ttk.Button(buttons,text=t("baseline_option_cancel"),command=dialog.destroy).pack(fill="x",pady=3)
        dialog.protocol("WM_DELETE_WINDOW",dialog.destroy)

    def _open_path(self, path: str) -> None:
        if not path or not os.path.exists(path):
            messagebox.showwarning(t("no_file_warning_title"), t("file_not_found_msg", path=path))
            return
        os.startfile(os.path.abspath(path))

    def _finish_pipeline(self, success: bool, result):
        if success:
            self.log(t("pipeline_success_log", result=result))
            self.root.after(100, self.load_cc_list)
            open_output = messagebox.askyesno(
                t("pipeline_complete_title"),
                t("pipeline_complete_msg", result=result),
            )
            if open_output:
                result_path = os.path.abspath(str(result))
                output_dir = result_path if os.path.isdir(result_path) else os.path.dirname(result_path)
                self._open_path(output_dir)
        else:
            recovery=self._missing_baseline_context(result)
            message = _friendly_error_message(result)
            self.log(t("pipeline_failed_log", message=message))
            if recovery:
                self._open_baseline_recovery_dialog(*recovery)
            else:
                messagebox.showerror(t("pipeline_failed_title"), message)
        self._set_pipeline_ui_busy(False)
        # Source paths may have changed while the subprocess was running; do
        # not re-enable calculation until the current selection is checked.
        self._mark_preflight_stale()


def main() -> int:
    _ensure_external_runtime_data()
    if "--health-check" in sys.argv[1:]:
        from src.services.runtime_health import print_health_report as _print_health_report

        return _print_health_report(BASE_DIR)
    if "--reference-staffing-render-worker" in sys.argv[1:]:
        from src.services.reference_staffing_render_worker import main as _render_worker_main

        worker_args = [
            arg for arg in sys.argv[1:] if arg != "--reference-staffing-render-worker"
        ]
        return _render_worker_main(worker_args)
    # Support headless export from packaged exe: child CLI invocations delegate
    # to the pipeline instead of opening a second GUI window.
    if len(sys.argv) > 1 and any(arg.startswith("--") for arg in sys.argv[1:]):
        from scripts.run_e2e import main as _cli_main

        return _cli_main()
    root = tk.Tk()
    app = MPManagerApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
