"""Trích xuất dữ liệu nhân sự phòng ban có thể truy nguyên từ mẫu kế hoạch tổng thể của công ty."""
from __future__ import annotations

import ast
from dataclasses import asdict, dataclass, field
from datetime import datetime
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
import time
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.parsers.extracted_headcount_time_plan import parse_extracted_headcount_time_plan
from src.parsers.headcount_time_plan import parse_headcount_time_plan
from src.utils.excel_helpers import get_fy_months, normalize_cc_code

# Retained only for compatibility with callers importing the former constant.
# This identifier is deliberately not written to the company-form workbook.
SCHEMA_ID = "MP2027_EXTRACTED_STAFFING_V1"
DETAIL_SHEET = "内訳ﾘｽﾄ(4～3月)"
OUTPUT_SHEET = "人員・時間計画"
INDEX_SHEET = "Sheet1"
PROVENANCE_SHEET = "PROVENANCE"
MANIFEST_SHEET = "EXTRACTION_MANIFEST"
MONTH_COLUMNS = range(6, 18)
WORKING_DAYS_ROW = 2
METRIC_ROWS = {
    "fixed_hours_expat": 8,
    "fixed_hours_local": 9,
    "overtime_hours_expat": 16,
    "overtime_hours_local": 17,
    "headcount_expat": 24,
    "headcount_local_total": 25,
}
_CELL_REF = re.compile(r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?(\d+)")


class ExtractionError(ValueError):
    """Raised when a reference workbook cannot be converted safely."""


@dataclass
class ExtractedDepartment:
    """Validated source values and their audit provenance for one department."""

    source_path: str
    cc_code: str = ""
    department_name: str = ""
    status: str = "ERROR"
    output_path: str = ""
    rows: list[dict[str, Any]] = field(default_factory=list)
    working_days: list[float] = field(default_factory=list)
    formula_count: int = 0
    split_source: str = ""
    candidate_split_source: str = ""
    template_source: str = ""
    render_strategy: str = ""
    mismatch_details: list[str] = field(default_factory=list)
    provenance: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _numeric(value: Any, label: str) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ExtractionError(f"{label}: giá trị không phải số: {value!r}")
    numeric = float(value)
    if numeric < 0:
        raise ExtractionError(f"{label}: giá trị âm {numeric:g}")
    return numeric


def _safe_arithmetic(expression: str) -> float:
    tree = ast.parse(expression, mode="eval")

    def walk(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return walk(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)) and not isinstance(node.value, bool):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = walk(node.operand)
            return value if isinstance(node.op, ast.UAdd) else -value
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            left, right = walk(node.left), walk(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if right == 0:
                raise ExtractionError("Công thức chia cho 0")
            return left / right
        raise ExtractionError(f"Công thức chứa thành phần không được hỗ trợ: {ast.dump(node, include_attributes=False)}")

    return float(walk(tree))


def evaluate_cell(ws: Worksheet, coordinate: str, stack: tuple[str, ...] = (), workbook=None) -> float:
    """Evaluate the narrow formula subset used by the FY reference reports."""
    coordinate = coordinate.replace("$", "").upper()
    if coordinate in stack:
        raise ExtractionError(f"Công thức vòng lặp: {' -> '.join((*stack, coordinate))}")
    raw = ws[coordinate].value
    if raw in (None, ""):
        return 0.0
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return float(raw)
    if not isinstance(raw, str) or not raw.startswith("="):
        raise ExtractionError(f"{coordinate}: không phải số/công thức hỗ trợ: {raw!r}")
    expression = raw[1:].strip()
    if not expression:
        raise ExtractionError(f"{coordinate}: công thức trống")
    if expression.startswith("OFFSET(稼働日!$B$2,MATCH("):
        if workbook is None or "稼働日" not in workbook.sheetnames:
            raise ExtractionError(f"{coordinate}: thiếu sheet 稼働日")
        month = int(evaluate_cell(ws, f"{coordinate[0]}4", (*stack, coordinate), workbook))
        working_days = {
            int(row[0]): _numeric(row[1], f"稼働日 tháng {row[0]}")
            for row in workbook["稼働日"].iter_rows(min_row=3, max_col=2, values_only=True)
            if row[0] not in (None, "")
        }
        if month not in working_days:
            raise ExtractionError(f"{coordinate}: không tìm thấy ngày làm việc tháng {month}")
        return working_days[month]

    def replace(match: re.Match[str]) -> str:
        reference = f"{match.group(1)}{match.group(2)}"
        return repr(evaluate_cell(ws, reference, (*stack, coordinate), workbook))

    substituted = _CELL_REF.sub(replace, expression)
    if re.search(r"[A-Za-z_$!:,]", substituted):
        raise ExtractionError(f"{coordinate}: công thức Excel không được hỗ trợ: {raw}")
    return _safe_arithmetic(substituted)


def _department_name(workbook, cc_code: str) -> str:
    ws = workbook["原価センタ"]
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        if normalize_cc_code(row[0]) == cc_code:
            return str(row[1] or "").strip()
    return ""


def _official_splits(official_source_dir: str | Path | None, fiscal_year: int) -> dict[tuple[str, str], Any]:
    official: dict[tuple[str, str], Any] = {}
    if not official_source_dir:
        return official
    for path in Path(official_source_dir).glob("*.xls"):
        parsed = parse_headcount_time_plan(str(path), fiscal_year)
        if parsed.status == "valid":
            official[(parsed.cc_code, parsed.department_name)] = parsed
    return official


def _output_filename(path: Path, department_name: str, fiscal_year: int) -> str:
    matched = re.match(r"\s*(\d+)\.", path.name)
    sequence = matched.group(1) if matched else "00"
    return f"{sequence}.{department_name}_FY{fiscal_year}_staffing_truth.xlsx"


def _record_source_value(
    item: ExtractedDepartment,
    metric: str,
    period: str,
    source_path: Path,
    source_cell: str,
    raw_value: Any,
    extracted_value: float,
) -> None:
    item.provenance.append(
        {
            "cc_code": item.cc_code,
            "department_name": item.department_name,
            "period": period,
            "metric": metric,
            "source_file": str(source_path.resolve()),
            "source_sheet": DETAIL_SHEET,
            "source_cell": source_cell,
            "source_formula_or_value": str(raw_value or ""),
            "extracted_value": extracted_value,
        }
    )


def extract_reference_workbook(
    path: str | Path,
    output_dir: str | Path,
    fiscal_year: int,
    official: dict[tuple[str, str], Any],
    overwrite: bool = False,
) -> ExtractedDepartment:
    """Extract one report workbook; publish only after the whole batch is valid."""
    path, output_dir = Path(path), Path(output_dir)
    item = ExtractedDepartment(source_path=str(path.resolve()))
    workbook = None
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
        if DETAIL_SHEET not in workbook.sheetnames or "原価センタ" not in workbook.sheetnames:
            raise ExtractionError("Thiếu sheet nội訳 hoặc 原価センタ")
        ws = workbook[DETAIL_SHEET]
        item.cc_code = normalize_cc_code(ws["B5"].value) or ""
        if not item.cc_code:
            raise ExtractionError("Không đọc được CC tại B5")
        item.department_name = _department_name(workbook, item.cc_code)
        if not item.department_name:
            raise ExtractionError(f"CC {item.cc_code} không có tên trong sheet 原価センタ")

        periods = get_fy_months(fiscal_year)
        expected_months = [int(period[-2:]) for period in periods]
        months = [int(evaluate_cell(ws, ws.cell(4, col).coordinate, workbook=workbook)) for col in MONTH_COLUMNS]
        if months != expected_months:
            raise ExtractionError(f"Thứ tự tháng không đúng FY{fiscal_year}: {months}")

        for period, column in zip(periods, MONTH_COLUMNS):
            cell = ws.cell(WORKING_DAYS_ROW, column)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                item.formula_count += 1
            value = evaluate_cell(ws, cell.coordinate, workbook=workbook)
            item.working_days.append(value)
            _record_source_value(item, "working_days", period, path, cell.coordinate, cell.value, value)

        series: dict[str, list[float]] = {}
        for metric, row_number in METRIC_ROWS.items():
            values: list[float] = []
            for period, column in zip(periods, MONTH_COLUMNS):
                cell = ws.cell(row_number, column)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    item.formula_count += 1
                value = evaluate_cell(ws, cell.coordinate, workbook=workbook)
                if value < 0:
                    raise ExtractionError(f"{cell.coordinate}: giá trị âm {value:g}")
                values.append(value)
                _record_source_value(item, metric, period, path, cell.coordinate, cell.value, value)
            series[metric] = values

        split = official.get((item.cc_code, item.department_name))
        split_rows = {row["period"]: row for row in split.rows} if split else {}
        accepted_split = bool(split)
        if split:
            item.candidate_split_source = str(Path(split.path).resolve())
            for index, period in enumerate(periods):
                official_row = split_rows[period]
                official_totals = {
                    "headcount_expat": float(official_row["headcount_expat"]),
                    "headcount_local_total": float(official_row["headcount_staff"] + official_row["headcount_worker"]),
                    "fixed_hours_expat": float(official_row["fixed_hours_expat"]),
                    "fixed_hours_local": float(official_row["fixed_hours_staff"] + official_row["fixed_hours_worker"]),
                    "overtime_hours_expat": float(official_row["overtime_hours_expat"]),
                    "overtime_hours_local": float(official_row["overtime_hours_staff"] + official_row["overtime_hours_worker"]),
                }
                for metric, official_value in official_totals.items():
                    reference_value = float(series[metric][index])
                    if abs(reference_value - official_value) > 0.01:
                        item.mismatch_details.append(
                            f"{period} {metric}: báo cáo tham chiếu={reference_value:g}, "
                            f"Master Plan={official_value:g}"
                        )
            accepted_split = not item.mismatch_details

        for index, period in enumerate(periods):
            source_row = split_rows[period] if accepted_split else None
            staff = float(source_row["headcount_staff"]) if source_row else None
            worker = float(source_row["headcount_worker"]) if source_row else None
            fixed_staff = float(source_row["fixed_hours_staff"]) if source_row else None
            fixed_worker = float(source_row["fixed_hours_worker"]) if source_row else None
            overtime_staff = float(source_row["overtime_hours_staff"]) if source_row else None
            overtime_worker = float(source_row["overtime_hours_worker"]) if source_row else None
            item.rows.append(
                {
                    "period": period,
                    "headcount_expat": series["headcount_expat"][index],
                    "headcount_local_total": series["headcount_local_total"][index],
                    "headcount_staff": staff,
                    "headcount_worker": worker,
                    "fixed_hours_expat": series["fixed_hours_expat"][index],
                    "fixed_hours_staff": fixed_staff,
                    "fixed_hours_worker": fixed_worker,
                    "fixed_hours_local": series["fixed_hours_local"][index],
                    "overtime_hours_expat": series["overtime_hours_expat"][index],
                    "overtime_hours_staff": overtime_staff,
                    "overtime_hours_worker": overtime_worker,
                    "overtime_hours_local": series["overtime_hours_local"][index],
                }
            )

        item.split_source = item.candidate_split_source if accepted_split else ""
        item.status = "READY" if accepted_split else "SPLIT_REQUIRED"
        item.render_strategy = "COMMON_TEMPLATE_WITH_OFFICIAL_SPLIT" if accepted_split else "COMMON_TEMPLATE_WITH_TOTALS_ONLY"
        item.output_path = str((output_dir / _output_filename(path, item.department_name, fiscal_year)).resolve())
        if Path(item.output_path).exists() and not overwrite:
            raise ExtractionError(f"File đầu ra đã tồn tại: {item.output_path}")
    except Exception as exc:
        item.errors.append(str(exc))
        item.status = "ERROR"
    finally:
        if workbook is not None:
            workbook.close()
    return item


def _excel_column(column: int) -> str:
    text = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        text = chr(65 + remainder) + text
    return text


def _empty_zero(value: float | int | None) -> float | int | None:
    """The official form represents user-entered zeroes as blank cells."""
    return None if value in (None, 0, 0.0) else value


class _CompanyFormRenderer:
    """Use desktop Excel to preserve the exact company workbook form."""

    XLSX_FILE_FORMAT = 51  # xlOpenXMLWorkbook
    AUTOMATIC_CALCULATION = -4105  # xlCalculationAutomatic

    def __init__(self) -> None:
        self._excel = None
        self._pythoncom = None
        self._com_initialized = False

    def __enter__(self) -> "_CompanyFormRenderer":
        try:
            import pythoncom  # type: ignore[import-not-found]
            import win32com.client  # type: ignore[import-not-found]
        except ImportError as exc:
            raise ExtractionError("Không có win32com để tạo file theo form Excel công ty") from exc
        self._pythoncom = pythoncom
        pythoncom.CoInitialize()
        self._com_initialized = True
        try:
            self._excel = win32com.client.DispatchEx("Excel.Application")
            self._excel.Visible = False
            self._excel.DisplayAlerts = False
            self._excel.AskToUpdateLinks = False
        except Exception:
            pythoncom.CoUninitialize()
            self._com_initialized = False
            self._pythoncom = None
            raise
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        import gc

        excel = self._excel
        self._excel = None
        try:
            # Release any cyclic child proxies while Excel's RPC server is still alive.
            gc.collect()
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    # Excel may already have terminated or disconnected its RPC server.
                    # Generation has finished at this point; cleanup must not mask its result.
                    pass
        finally:
            if self._com_initialized and self._pythoncom is not None:
                self._pythoncom.CoUninitialize()
            self._com_initialized = False
            self._pythoncom = None
            # Do not explicitly release the parent proxy after Quit: Excel has already
            # disconnected its RPC server. The Python wrapper falls out of scope here.
            excel = None

    @staticmethod
    def _fill_company_form(sheet: Worksheet, item: ExtractedDepartment) -> None:
        """Populate the protected company form without removing its protection metadata."""
        sheet.cell(5, 1).number_format = "@"
        sheet.cell(5, 1).value = item.cc_code
        sheet.cell(5, 2).value = item.department_name
        for top_row, bottom_row in ((10, 14), (17, 21), (24, 28)):
            for row_cells in sheet.iter_rows(
                min_row=top_row,
                max_row=bottom_row,
                min_col=3,
                max_col=14,
            ):
                for cell in row_cells:
                    cell.value = None

        for column, row in enumerate(item.rows, start=3):
            column_letter = _excel_column(column)
            sheet.cell(6, column).value = _empty_zero(item.working_days[column - 3])

            sheet.cell(10, column).value = _empty_zero(row["headcount_expat"])
            sheet.cell(11, column).value = _empty_zero(row["headcount_staff"])
            sheet.cell(12, column).value = _empty_zero(row["headcount_worker"])
            sheet.cell(13, column).value = _empty_zero(row["headcount_local_total"])
            sheet.cell(14, column).value = f"={column_letter}10+{column_letter}13"

            sheet.cell(17, column).value = _empty_zero(row["fixed_hours_expat"])
            sheet.cell(18, column).value = _empty_zero(row["fixed_hours_staff"])
            sheet.cell(19, column).value = _empty_zero(row["fixed_hours_worker"])
            sheet.cell(20, column).value = _empty_zero(row["fixed_hours_local"])
            sheet.cell(21, column).value = f"={column_letter}17+{column_letter}20"

            sheet.cell(24, column).value = _empty_zero(row["overtime_hours_expat"])
            sheet.cell(25, column).value = _empty_zero(row["overtime_hours_staff"])
            sheet.cell(26, column).value = _empty_zero(row["overtime_hours_worker"])
            sheet.cell(27, column).value = _empty_zero(row["overtime_hours_local"])
            sheet.cell(28, column).value = f"={column_letter}24+{column_letter}27"

    def _convert_to_xlsx(self, source_path: Path, output_path: Path) -> None:
        workbooks = self._excel.Workbooks
        workbook = workbooks.Open(
            str(source_path.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
        )
        worksheets = None
        try:
            worksheets = workbook.Worksheets
            if worksheets.Count < 2:
                raise ExtractionError("Template công ty không đủ hai sheet")
            workbook.SaveAs(str(output_path.resolve()), FileFormat=self.XLSX_FILE_FORMAT)
        finally:
            workbook.Close(SaveChanges=False)
            worksheets = None
            workbook = None
            workbooks = None

    def _refresh_formula_cache(self, workbook_path: Path) -> None:
        workbooks = self._excel.Workbooks
        workbook = workbooks.Open(
            str(workbook_path.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
        )
        worksheet = None
        try:
            worksheet = workbook.Worksheets(OUTPUT_SHEET)
            worksheet.Calculate()
            workbook.Save()
        finally:
            workbook.Close(SaveChanges=False)
            worksheet = None
            workbook = None
            workbooks = None

    @staticmethod
    def _validate_rendered_output(
        item: ExtractedDepartment,
        workbook_path: Path,
        fiscal_year: int,
    ) -> None:
        parsed = parse_extracted_headcount_time_plan(str(workbook_path), fiscal_year)
        if parsed.status != "valid":
            raise ExtractionError(
                "File đầu ra không vượt qua hậu kiểm: " + "; ".join(parsed.errors)
            )
        if parsed.cc_code != item.cc_code or parsed.department_name != item.department_name:
            raise ExtractionError("File đầu ra không giữ đúng mã hoặc tên bộ phận")
        if len(parsed.rows) != 12:
            raise ExtractionError("File đầu ra không đủ 12 tháng")

        metrics = (
            "headcount_expat",
            "headcount_local_total",
            "headcount_staff",
            "headcount_worker",
            "fixed_hours_expat",
            "fixed_hours_staff",
            "fixed_hours_worker",
            "fixed_hours_local",
            "overtime_hours_expat",
            "overtime_hours_staff",
            "overtime_hours_worker",
            "overtime_hours_local",
        )
        for expected, actual in zip(item.rows, parsed.rows, strict=True):
            if actual["period"] != expected["period"]:
                raise ExtractionError(
                    f"File đầu ra sai kỳ: {actual['period']} != {expected['period']}"
                )
            for metric in metrics:
                expected_value = 0.0 if expected[metric] is None else float(expected[metric])
                actual_value = 0.0 if actual[metric] is None else float(actual[metric])
                if abs(expected_value - actual_value) > 0.01:
                    raise ExtractionError(
                        f"File đầu ra sai {expected['period']} {metric}: "
                        f"{actual_value:g} != {expected_value:g}"
                    )

        if item.status == "SPLIT_REQUIRED":
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            try:
                sheet = workbook[OUTPUT_SHEET]
                for row_number in (11, 12, 18, 19, 25, 26):
                    for column in range(3, 15):
                        if sheet.cell(row_number, column).value not in (None, ""):
                            raise ExtractionError(
                                "File cần bổ sung phân tách nhưng ô staff/worker không trống: "
                                f"{sheet.cell(row_number, column).coordinate}"
                            )
            finally:
                workbook.close()

    def render(
        self,
        item: ExtractedDepartment,
        template_path: Path,
        staging_output: Path,
        fiscal_year: int,
    ) -> None:
        if self._excel is None:
            raise ExtractionError("Excel renderer chưa được khởi tạo")
        if not template_path.exists():
            raise ExtractionError(f"Không tìm thấy template Master Plan: {template_path}")
        if staging_output.exists():
            staging_output.unlink()

        self._convert_to_xlsx(template_path, staging_output)
        workbook = load_workbook(staging_output, data_only=False)
        try:
            if OUTPUT_SHEET not in workbook.sheetnames or INDEX_SHEET not in workbook.sheetnames:
                raise ExtractionError("Template công ty không đủ hai sheet chuẩn")
            sheet = workbook[OUTPUT_SHEET]
            protection_hash = sheet.protection.password
            self._fill_company_form(sheet, item)
            if sheet.protection.password != protection_hash or not sheet.protection.sheet:
                raise ExtractionError("Không bảo toàn được trạng thái bảo vệ của form công ty")
            workbook.save(staging_output)
        finally:
            workbook.close()
        self._refresh_formula_cache(staging_output)
        self._validate_rendered_output(item, staging_output, fiscal_year)


def _write_manifest(results: list[ExtractedDepartment], manifest_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = MANIFEST_SHEET
    summary_headers = [
        "source_file", "output_file", "template_file", "render_strategy", "cc_code",
        "department_name", "status", "months", "formula_count", "split_source",
        "candidate_split_source", "mismatch_details", "errors",
    ]
    summary.append(summary_headers)
    for item in results:
        summary.append(
            [
                item.source_path, item.output_path, item.template_source, item.render_strategy,
                item.cc_code, item.department_name, item.status, len(item.rows), item.formula_count,
                item.split_source, item.candidate_split_source, "; ".join(item.mismatch_details),
                "; ".join(item.errors),
            ]
        )

    provenance = workbook.create_sheet(PROVENANCE_SHEET)
    provenance_headers = [
        "cc_code", "department_name", "period", "metric", "source_file", "source_sheet",
        "source_cell", "source_formula_or_value", "extracted_value",
    ]
    provenance.append(provenance_headers)
    for item in results:
        for entry in item.provenance:
            provenance.append([entry[key] for key in provenance_headers])
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(manifest_path)
    workbook.close()


def _backup_existing_truth_files(output_dir: Path, backup_dir: Path) -> None:
    existing = list(output_dir.glob("*_staffing_truth.xlsx"))
    manifest = output_dir / "extraction_manifest.xlsx"
    if manifest.exists():
        existing.append(manifest)
    if not existing:
        return
    backup_dir.mkdir(parents=True, exist_ok=False)
    for path in existing:
        shutil.move(str(path), str(backup_dir / path.name))


def _restore_backup(output_dir: Path, backup_dir: Path, published_paths: list[Path]) -> None:
    for path in published_paths:
        if path.exists():
            path.unlink()
    if backup_dir.exists():
        for path in backup_dir.iterdir():
            shutil.move(str(path), str(output_dir / path.name))
        backup_dir.rmdir()


_RENDER_WORKER_PROTOCOL = 1
_RENDER_WORKER_TIMEOUT_SECONDS = 15 * 60


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, sort_keys=True)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary, path)


def _render_worker_command(request_path: Path, response_path: Path, status_path: Path) -> list[str]:
    arguments = [
        "--request",
        str(request_path),
        "--response",
        str(response_path),
        "--status",
        str(status_path),
    ]
    if getattr(sys, "frozen", False):
        return [sys.executable, "--reference-staffing-render-worker", *arguments]
    return [
        sys.executable,
        "-m",
        "src.services.reference_staffing_render_worker",
        *arguments,
    ]


def _terminate_worker_tree(process: subprocess.Popen[Any]) -> None:
    if process.poll() is not None:
        return
    if os.name == "nt":
        subprocess.run(
            ["taskkill", "/PID", str(process.pid), "/T", "/F"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
    else:
        process.kill()
    try:
        process.wait(timeout=10)
    except subprocess.TimeoutExpired:
        process.kill()
        process.wait()


def _read_worker_status(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        status = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return status if status.get("protocol_version") == _RENDER_WORKER_PROTOCOL else {}


def _preserve_worker_diagnostic(
    output_dir: Path,
    stdout_path: Path,
    stderr_path: Path,
    reason: str,
) -> None:
    sections = [f"reason: {reason}"]
    for label, path in (("stdout", stdout_path), ("stderr", stderr_path)):
        try:
            content = path.read_text(encoding="utf-8", errors="replace").strip()
        except OSError:
            content = ""
        if content:
            sections.extend((f"\n[{label}]", content))
    diagnostic_path = output_dir / "_staffing_render_diagnostic.log"
    diagnostic_path.write_text("\n".join(sections) + "\n", encoding="utf-8")


def _run_render_worker(
    results: list[ExtractedDepartment],
    template_path: Path,
    stage_dir: Path,
    output_dir: Path,
    fiscal_year: int,
    progress_callback: Callable[[int, int, str, str], None] | None,
    timeout_seconds: float = _RENDER_WORKER_TIMEOUT_SECONDS,
) -> None:
    request_path = stage_dir / "render_request.json"
    response_path = stage_dir / "render_response.json"
    status_path = stage_dir / "render_status.json"
    stdout_path = stage_dir / "render_worker.stdout.log"
    stderr_path = stage_dir / "render_worker.stderr.log"
    expected_files = [Path(item.output_path).name for item in results]
    _atomic_write_json(
        request_path,
        {
            "protocol_version": _RENDER_WORKER_PROTOCOL,
            "template_path": str(template_path.resolve()),
            "stage_dir": str(stage_dir.resolve()),
            "fiscal_year": fiscal_year,
            "items": [asdict(item) for item in results],
        },
    )

    project_root = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parents[2]
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
    process: subprocess.Popen[Any] | None = None
    reported_completed = 0
    last_processing = ""
    started = time.monotonic()
    failure_reason = ""
    try:
        with stdout_path.open("wb") as stdout_handle, stderr_path.open("wb") as stderr_handle:
            process = subprocess.Popen(
                _render_worker_command(request_path, response_path, status_path),
                cwd=str(project_root),
                stdin=subprocess.DEVNULL,
                stdout=stdout_handle,
                stderr=stderr_handle,
                creationflags=creationflags,
            )
            while process.poll() is None:
                status = _read_worker_status(status_path)
                completed = status.get("completed_files", [])
                while reported_completed < min(len(completed), len(expected_files)):
                    if completed[reported_completed] != expected_files[reported_completed]:
                        failure_reason = "Tiến độ worker không khớp với thứ tự tệp được yêu cầu"
                        _terminate_worker_tree(process)
                        break
                    reported_completed += 1
                    if progress_callback is not None:
                        progress_callback(
                            reported_completed,
                            len(expected_files),
                            expected_files[reported_completed - 1],
                            "completed",
                        )
                if failure_reason:
                    break
                current_file = str(status.get("current_file") or "")
                if current_file and current_file != last_processing:
                    if current_file not in expected_files:
                        failure_reason = "Worker báo cáo một tệp kết quả không mong đợi"
                        _terminate_worker_tree(process)
                        break
                    last_processing = current_file
                    if progress_callback is not None:
                        progress_callback(
                            reported_completed,
                            len(expected_files),
                            current_file,
                            "processing",
                        )
                if time.monotonic() - started > timeout_seconds:
                    failure_reason = "Worker đã hết thời gian chờ"
                    _terminate_worker_tree(process)
                    break
                time.sleep(0.1)
            return_code = process.wait()

        if not failure_reason and return_code != 0:
            failure_reason = f"Worker đã kết thúc với mã lỗi {return_code}"
        try:
            response = json.loads(response_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            response = {}
            if not failure_reason:
                failure_reason = "Worker không trả về phản hồi hợp lệ"
        if response.get("protocol_version") != _RENDER_WORKER_PROTOCOL:
            failure_reason = failure_reason or "Giao thức phản hồi của worker không hợp lệ"
        if response.get("success") is not True:
            failure_reason = failure_reason or "Worker báo cáo không thể tạo tệp Excel"
        if response.get("rendered_files") != expected_files:
            failure_reason = failure_reason or "Phản hồi của worker thiếu các tệp đã tạo"
        if failure_reason:
            raise ExtractionError("Không thể hoàn tất tạo file nhân sự trong Excel")

        for index, (item, filename) in enumerate(zip(results, expected_files, strict=True), start=1):
            staging_output = stage_dir / filename
            if not staging_output.is_file() or staging_output.stat().st_size == 0:
                raise ExtractionError("Excel không tạo đủ file nhân sự trong vùng tạm")
            _CompanyFormRenderer._validate_rendered_output(item, staging_output, fiscal_year)
            if progress_callback is not None and index > reported_completed:
                progress_callback(index, len(expected_files), filename, "completed")
    except Exception as exc:
        if process is not None:
            _terminate_worker_tree(process)
        _preserve_worker_diagnostic(output_dir, stdout_path, stderr_path, failure_reason or str(exc))
        raise

def extract_reference_staffing_sources(
    source_dir: str | Path,
    output_dir: str | Path,
    fiscal_year: int = 2027,
    official_source_dir: str | Path | None = None,
    overwrite: bool = False,
    progress_callback: Callable[[int, int, str, str], None] | None = None,
) -> dict[str, Any]:
    """Validate all report files, then atomically publish company-form workbooks."""
    source_dir, output_dir = Path(source_dir), Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    official = _official_splits(official_source_dir, fiscal_year)
    results = [
        extract_reference_workbook(path, output_dir, fiscal_year, official, overwrite)
        for path in sorted(source_dir.glob("*.xlsx"))
        if not path.name.startswith("~$") and not path.name.endswith("_staffing_truth.xlsx")
    ]

    if not official:
        for item in results:
            if item.status != "ERROR":
                item.status = "ERROR"
                item.errors.append("Không có file Master Plan chính thức để dùng làm template form công ty")
    errors = [item for item in results if item.status == "ERROR"]

    if not errors:
        template_path = Path(sorted(official.values(), key=lambda value: value.path)[0].path).resolve()
        for item in results:
            item.template_source = str(template_path)
        stage_dir = Path(tempfile.mkdtemp(prefix=".staffing_truth_stage_", dir=output_dir))
        try:
            try:
                _run_render_worker(
                    results,
                    template_path,
                    stage_dir,
                    output_dir,
                    fiscal_year,
                    progress_callback,
                )
            except Exception as exc:
                failed_item = results[0]
                failed_item.status = "ERROR"
                failed_item.errors.append(str(exc))
                errors.append(failed_item)
                if progress_callback is not None:
                    progress_callback(0, len(results), Path(failed_item.output_path).name, "error")

            if not errors:
                staging_manifest = stage_dir / "extraction_manifest.xlsx"
                _write_manifest(results, staging_manifest)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_dir = output_dir / f"_staffing_truth_backup_{timestamp}"
                published_paths: list[Path] = []
                try:
                    _backup_existing_truth_files(output_dir, backup_dir)
                    for item in results:
                        staging_path = stage_dir / Path(item.output_path).name
                        final_path = Path(item.output_path)
                        os.replace(staging_path, final_path)
                        published_paths.append(final_path)
                    manifest_path = output_dir / "extraction_manifest.xlsx"
                    os.replace(staging_manifest, manifest_path)
                    published_paths.append(manifest_path)
                except Exception:
                    _restore_backup(output_dir, backup_dir, published_paths)
                    raise
        finally:
            shutil.rmtree(stage_dir, ignore_errors=True)

    manifest_path = output_dir / "extraction_manifest.xlsx"
    return {
        "files": len(results),
        "ready": sum(item.status == "READY" for item in results),
        "split_required": sum(item.status == "SPLIT_REQUIRED" for item in results),
        "errors": [item for item in results if item.status == "ERROR"],
        "results": results,
        "manifest_path": str(manifest_path.resolve()),
    }
