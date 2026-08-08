"""Structural compatibility helpers for MP FORM workbooks.

Known FORM layouts are retained for compatibility metadata. New usable layouts
are admitted by the person running the project through checksum confirmation.
"""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl


@dataclass(frozen=True)
class TemplateLayout:
    """A known or runtime FORM layout with its business-sheet contract."""

    version: str
    hub_sheet_name: str
    expected_sheet_names: tuple[str, ...]
    hub_min_rows: int
    hub_max_rows: int
    hub_min_columns: int
    hub_max_columns: int
    required_cells: tuple[tuple[str, object], ...]
    payload_start_row: int = 38
    payload_columns: tuple[int, ...] = (2, 19, 20)

    def matches(self, workbook: openpyxl.Workbook) -> bool:
        if tuple(workbook.sheetnames) != self.expected_sheet_names:
            return False
        if self.hub_sheet_name not in workbook.sheetnames:
            return False
        worksheet = workbook[self.hub_sheet_name]
        if not self.hub_min_rows <= worksheet.max_row <= self.hub_max_rows:
            return False
        if not self.hub_min_columns <= worksheet.max_column <= self.hub_max_columns:
            return False
        return all(worksheet[cell].value == value for cell, value in self.required_cells)


# FORM variants supplied with this repository. Values use structural identity
# only; department-specific payload is separately rejected by hygiene checks.
_COMMON_SHEETS = (
    "\u63a1\u7b97\u8868(USD)",
    "\u63a1\u7b97\u8868(VND)",
    "\u5185\u8a33\uff98\uff7d\uff84(4\uff5e3\u6708)",
    "\u8a2d\u5099\u6295\u8cc7\u8a08\u753b",
    "\u52d8\u5b9a\u79d1\u76ee",
    "\u539f\u4fa1\u30bb\u30f3\u30bf",
    "\u7a3c\u50cd\u65e5",
)
_HUB = "\u5185\u8a33\uff98\uff7d\uff84(4\uff5e3\u6708)"

APPROVED_TEMPLATE_LAYOUTS: tuple[TemplateLayout, ...] = (
    TemplateLayout(
        version="MP2027-FORM-1015",
        hub_sheet_name=_HUB,
        expected_sheet_names=_COMMON_SHEETS,
        hub_min_rows=1001,
        hub_max_rows=1020,
        hub_min_columns=55,
        hub_max_columns=55,
        required_cells=(("B2", 26273),),
    ),
    TemplateLayout(
        version="MP2027-FORM-1000",
        hub_sheet_name=_HUB,
        expected_sheet_names=_COMMON_SHEETS,
        hub_min_rows=990,
        hub_max_rows=1000,
        hub_min_columns=55,
        hub_max_columns=55,
        required_cells=(),
    ),
)


class TemplateLayoutError(ValueError):
    """Raised when a workbook has no compatible MP detail-sheet layout."""


def approved_layout_versions() -> tuple[str, ...]:
    return tuple(layout.version for layout in APPROVED_TEMPLATE_LAYOUTS)


def resolve_template_layout(workbook: openpyxl.Workbook) -> TemplateLayout:
    """Return a compatible FORM layout for technical validation.

    Known layouts retain their version metadata. A new workbook that still has
    the required MP detail sheet receives a runtime layout; its use is then
    controlled by the user's checksum confirmation rather than source edits.
    """
    matches = [layout for layout in APPROVED_TEMPLATE_LAYOUTS if layout.matches(workbook)]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise TemplateLayoutError("FORM khớp nhiều layout; cần kiểm tra lại cấu trúc.")
    hub_candidates = [name for name in workbook.sheetnames if "内訳" in name]
    if len(hub_candidates) != 1:
        raise TemplateLayoutError("FORM không có sheet chi tiết MP đúng cấu trúc.")
    worksheet = workbook[hub_candidates[0]]
    return TemplateLayout(
        version="runtime",
        hub_sheet_name=hub_candidates[0],
        expected_sheet_names=tuple(workbook.sheetnames),
        hub_min_rows=1,
        hub_max_rows=max(worksheet.max_row, 1),
        hub_min_columns=1,
        hub_max_columns=max(worksheet.max_column, 1),
        required_cells=(),
    )


def layout_hub_sheet_name(workbook: openpyxl.Workbook) -> str:
    """Resolve the canonical hub sheet from the approved template layout."""
    return resolve_template_layout(workbook).hub_sheet_name


def layout_payload_bounds(workbook: openpyxl.Workbook) -> tuple[int, tuple[int, ...]]:
    """Return the payload area used for hygiene validation."""
    layout = resolve_template_layout(workbook)
    return layout.payload_start_row, layout.payload_columns


def layout_hub_sheet_name_for_output(workbook: openpyxl.Workbook) -> str:
    """Resolve the hub sheet in a derived workbook.

    Exported workbooks deliberately carry CC payload and can grow beyond their
    source template dimensions. Their writers still require one unambiguous MP
    detail sheet, regardless of whether the source FORM was previously known.
    """
    hubs = [name for name in workbook.sheetnames if "内訳" in name]
    if len(hubs) == 1:
        return hubs[0]
    raise TemplateLayoutError("Workbook kết quả không có sheet chi tiết MP đúng cấu trúc.")
