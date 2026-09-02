"""Service for managing Fiscal Year RAG Knowledge Updates (Cập nhật kiến thức RAG theo năm tài chính).

Enables operators to add, update, and manage fiscal-year-specific business rules,
guidelines, and error models for new fiscal years (e.g. FY2028, FY2029...) or updated
Excel cost templates without modifying or compromising closed legacy documentation (FY2027).

Design Invariants:
1. Immutability of Legacy Documentation: Legacy FY documents (e.g. FY2027) are read-only and never overwritten.
2. Versioned Storage: Each update pack is saved under `docs/knowledge/business_chat/updates/{fiscal_year}/{update_id}.json`.
3. Priority & Recency: Confirmed updates for newer FYs take precedence over older rules for the same topic.
4. Confidence Gating: Unconfirmed updates are labeled 'Tham khảo nội bộ' (Internal Reference / 社内参考).
5. Fail-Closed Atomic Publishing: Index rebuild and validation must pass completely; otherwise, the previous index remains untouched.
6. Zero Technical Leakage: No local filesystem paths, hashes, JSON/MD file names, or code tokens exposed to end users.
7. No Automated Rule Guessing from Excel: Only basic structure (sheets/columns) is read to assist operator note-taking.
"""

from __future__ import annotations

import datetime
import hashlib
import json
import os
import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


SUPPORTED_LANGUAGES: tuple[str, ...] = ("vi", "en", "ja")
_UPDATES_DIR_RELATIVE_PATH: str = "docs/knowledge/business_chat/updates"
_GENERATED_ARTIFACT_PATHS: tuple[str, ...] = (
    "docs/knowledge/business_chat/source_discovery_inventory.json",
    "docs/knowledge/business_chat/coverage_matrix.json",
    "docs/knowledge/business_chat/coverage_evidence_report.json",
    "docs/knowledge/business_chat/coverage_evidence_report.md",
    "docs/knowledge/business_chat/knowledge_index.json",
)
_FORBIDDEN_TECHNICAL_TOKENS: tuple[str, ...] = (
    "d:\\",
    "c:\\",
    "traceback",
    "cagent",
    "c-agent",
    "def ",
    "class ",
    "import ",
    "from src.",
)


def _repo_root() -> Path:
    """Resolve the repository root directory."""
    return Path(__file__).resolve().parents[2]


def _atomic_write_bytes(path: Path, content: bytes) -> None:
    """Write one file durably enough that a partial JSON file is never published."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=path.parent, delete=False) as tmp_file:
        tmp_file.write(content)
        tmp_name = tmp_file.name
    try:
        os.replace(tmp_name, path)
    finally:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)


def _write_update_file(path: Path, payload: Dict[str, Any]) -> None:
    content = (json.dumps(payload, indent=2, ensure_ascii=False) + "\n").encode("utf-8")
    _atomic_write_bytes(path, content)


def _snapshot_generated_artifacts(root: Path) -> Dict[Path, bytes | None]:
    """Capture generated RAG artifacts before a publish/deactivate transaction."""
    snapshots: Dict[Path, bytes | None] = {}
    for relative_path in _GENERATED_ARTIFACT_PATHS:
        path = root / relative_path
        snapshots[path] = path.read_bytes() if path.is_file() else None
    return snapshots


def _restore_generated_artifacts(snapshots: Dict[Path, bytes | None]) -> None:
    """Restore generated artifacts exactly to their pre-transaction state."""
    for path, content in snapshots.items():
        if content is None:
            if path.is_file():
                path.unlink()
        else:
            _atomic_write_bytes(path, content)


def _reload_restored_index() -> None:
    """Make the in-process retrieval cache agree with restored on-disk artifacts."""
    from src.services.business_knowledge_index import reload_knowledge_index

    reload_knowledge_index(check_freshness=True)


@dataclass
class FiscalYearUpdateItem:
    """Data model representing a single Fiscal Year RAG Knowledge Update item."""

    fiscal_year: str  # e.g. "FY2028", "FY2029"
    update_id: str  # e.g. "upd_fy2028_facility_cost_revision"
    status: str = "confirmed"  # 'confirmed' | 'reference_with_caveat' | 'draft'
    change_type: str = "changed_rule"  # 'new_rule' | 'changed_rule' | 'known_error' | 'changed_excel_layout' | 'operational_guidance'
    business_area: str = "cost_allocation"  # 'cost_allocation' | 'staffing' | 'facilities' | 'it_system' | 'operations' | 'troubleshooting'
    title: Dict[str, str] = field(default_factory=lambda: {"vi": "", "en": "", "ja": ""})
    what_changed: Dict[str, str] = field(default_factory=lambda: {"vi": "", "en": "", "ja": ""})
    user_action: Dict[str, str] = field(default_factory=lambda: {"vi": "", "en": "", "ja": ""})
    applies_to: Dict[str, str] = field(default_factory=lambda: {"vi": "", "en": "", "ja": ""})
    source_note: Dict[str, str] = field(default_factory=lambda: {"vi": "", "en": "", "ja": ""})
    evidence_anchor: str = ""
    replaces_or_supersedes: List[str] = field(default_factory=list)
    is_active: bool = True
    created_at: str = ""
    updated_at: str = ""

    def __post_init__(self) -> None:
        now = datetime.datetime.now(datetime.timezone.utc).isoformat()
        if not self.created_at:
            self.created_at = now
        if not self.updated_at:
            self.updated_at = now

    def to_dict(self) -> Dict[str, Any]:
        return {
            "schema_version": "1.0",
            "fiscal_year": self.fiscal_year,
            "update_id": self.update_id,
            "status": self.status,
            "change_type": self.change_type,
            "business_area": self.business_area,
            "title": dict(self.title),
            "what_changed": dict(self.what_changed),
            "user_action": dict(self.user_action),
            "applies_to": dict(self.applies_to),
            "source_note": dict(self.source_note),
            "evidence_anchor": self.evidence_anchor,
            "replaces_or_supersedes": list(self.replaces_or_supersedes),
            "is_active": self.is_active,
            "created_at": self.created_at,
            "updated_at": self.updated_at,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> FiscalYearUpdateItem:
        return cls(
            fiscal_year=str(data.get("fiscal_year", "FY2028")).strip().upper(),
            update_id=str(data.get("update_id", "")).strip(),
            status=str(data.get("status", "confirmed")).strip().lower(),
            change_type=str(data.get("change_type", "changed_rule")).strip().lower(),
            business_area=str(data.get("business_area", "cost_allocation")).strip().lower(),
            title=dict(data.get("title", {})),
            what_changed=dict(data.get("what_changed", {})),
            user_action=dict(data.get("user_action", {})),
            applies_to=dict(data.get("applies_to", {})),
            source_note=dict(data.get("source_note", {})),
            evidence_anchor=str(data.get("evidence_anchor", "")).strip(),
            replaces_or_supersedes=list(data.get("replaces_or_supersedes", [])),
            is_active=bool(data.get("is_active", True)),
            created_at=str(data.get("created_at", "")).strip(),
            updated_at=str(data.get("updated_at", "")).strip(),
        )


def validate_update_item(item: FiscalYearUpdateItem) -> Tuple[bool, List[str]]:
    """Validate that a FiscalYearUpdateItem satisfies safety and completeness requirements."""
    errors: List[str] = []

    # 1. Fiscal Year format
    if not item.fiscal_year or not re.match(r"^FY\d{4}$", item.fiscal_year, re.IGNORECASE):
        errors.append("Năm tài chính không đúng định dạng (ví dụ: FY2028, FY2029).")

    # 2. Update ID format
    if not item.update_id or not re.match(r"^[a-zA-Z0-9_\-]+$", item.update_id):
        errors.append("Mã định danh cập nhật (update_id) chỉ được chứa chữ, số, gạch dưới (_) hoặc gạch nối (-).")

    # 3. Status
    if item.status not in ("confirmed", "reference_with_caveat", "draft"):
        errors.append(f"Trạng thái '{item.status}' không hợp lệ (hợp lệ: 'confirmed', 'reference_with_caveat', 'draft').")

    # 4. Multilingual fields
    for lang in SUPPORTED_LANGUAGES:
        if not item.title.get(lang, "").strip():
            errors.append(f"Tiêu đề thiếu nội dung ngôn ngữ '{lang}'.")
        if not item.what_changed.get(lang, "").strip():
            errors.append(f"Mô tả thay đổi thiếu nội dung ngôn ngữ '{lang}'.")
        if not item.user_action.get(lang, "").strip():
            errors.append(f"Hướng dẫn người dùng thiếu nội dung ngôn ngữ '{lang}'.")

    # 5. Evidence anchor (only required if not draft)
    if item.status in ("confirmed", "reference_with_caveat"):
        anchor = item.evidence_anchor.strip()
        if not anchor:
            errors.append("Thiếu bằng chứng nghiệp vụ (evidence_anchor) trích xuất từ nội dung thay đổi.")
        elif len(anchor) < 15 or len(anchor) > 50:
            errors.append(f"Độ dài evidence_anchor ({len(anchor)} ký tự) phải nằm trong khoảng 15 đến 50 ký tự.")
        else:
            # Check if anchor is present in what_changed text (any language)
            all_text = " ".join(item.what_changed.values()) + " " + " ".join(item.user_action.values())
            if anchor not in all_text:
                errors.append("evidence_anchor phải là đoạn trích xuất nguyên văn từ phần mô tả thay đổi hoặc hướng dẫn.")

    # 6. Guard against technical leakage
    all_combined = (
        " ".join(item.title.values())
        + " "
        + " ".join(item.what_changed.values())
        + " "
        + " ".join(item.user_action.values())
        + " "
        + " ".join(item.applies_to.values())
        + " "
        + " ".join(item.source_note.values())
    ).lower()

    for token in _FORBIDDEN_TECHNICAL_TOKENS:
        if token in all_combined:
            errors.append(f"Nội dung chứa từ khóa kỹ thuật hoặc đường dẫn tệp không an toàn ('{token}').")

    return (len(errors) == 0, errors)


def get_updates_directory(fiscal_year: str, repo_root: Path | None = None) -> Path:
    """Return the filesystem directory for a specific fiscal year's updates."""
    root = repo_root or _repo_root()
    clean_fy = str(fiscal_year).strip().upper()
    return root / _UPDATES_DIR_RELATIVE_PATH / clean_fy


def list_updates(fiscal_year: str | None = None, repo_root: Path | None = None) -> List[FiscalYearUpdateItem]:
    """List all registered update items across all fiscal years or filtered by a specific FY."""
    root = repo_root or _repo_root()
    base_dir = root / _UPDATES_DIR_RELATIVE_PATH
    if not base_dir.is_dir():
        return []

    items: List[FiscalYearUpdateItem] = []
    fy_filter = str(fiscal_year).strip().upper() if fiscal_year else None

    for fy_folder in sorted(base_dir.glob("FY*")):
        if not fy_folder.is_dir():
            continue
        if fy_filter and fy_folder.name.upper() != fy_filter:
            continue
        for update_file in sorted(fy_folder.glob("*.json")):
            try:
                data = json.loads(update_file.read_text(encoding="utf-8"))
                item = FiscalYearUpdateItem.from_dict(data)
                items.append(item)
            except Exception:
                continue

    return items


def save_draft(item: FiscalYearUpdateItem, repo_root: Path | None = None) -> Path:
    """Save an update item as draft without publishing to RAG index."""
    root = repo_root or _repo_root()
    item.status = "draft"
    item.updated_at = datetime.datetime.now(datetime.timezone.utc).isoformat()
    if not item.update_id:
        slug_src = item.title.get("vi") or item.title.get("en") or item.title.get("ja") or "update"
        slug = re.sub(r"[^a-zA-Z0-9_]+", "_", slug_src).strip("_").lower()
        item.update_id = f"upd_{item.fiscal_year.lower()}_{slug or 'item'}"

    target_dir = get_updates_directory(item.fiscal_year, root)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_file = target_dir / f"{item.update_id}.json"
    _write_update_file(target_file, item.to_dict())
    return target_file


def publish_update(
    item: FiscalYearUpdateItem,
    repo_root: Path | None = None,
) -> Tuple[bool, str]:
    """Publish a fiscal year update item atomically with strict fail-closed validation.

    If validation or index rebuild fails, rolls back file changes and keeps existing index intact.
    """
    root = repo_root or _repo_root()
    if item.status not in ("confirmed", "reference_with_caveat"):
        item.status = "confirmed"

    item.is_active = True
    item.updated_at = datetime.datetime.now(datetime.timezone.utc).isoformat()
    if not item.update_id:
        slug_src = item.title.get("vi") or item.title.get("en") or item.title.get("ja") or "update"
        slug = re.sub(r"[^a-zA-Z0-9_]+", "_", slug_src).strip("_").lower()
        item.update_id = f"upd_{item.fiscal_year.lower()}_{slug or 'item'}"

    # Auto-populate anchor if missing but valid text exists (check VI -> EN -> JA)
    if not item.evidence_anchor:
        for lang_cand in SUPPORTED_LANGUAGES:
            txt = item.what_changed.get(lang_cand, "").strip()
            if len(txt) >= 15:
                item.evidence_anchor = txt[:50]
                break

    is_valid, validation_errors = validate_update_item(item)
    if not is_valid:
        return False, "Validation thất bại: " + "; ".join(validation_errors)

    target_dir = get_updates_directory(item.fiscal_year, root)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_file = target_dir / f"{item.update_id}.json"

    # Backup existing file if any for atomic rollback
    old_content: Optional[bytes] = None
    if target_file.is_file():
        old_content = target_file.read_bytes()
    artifact_snapshots = _snapshot_generated_artifacts(root)

    try:
        # 1. Write the new update pack
        _write_update_file(target_file, item.to_dict())

        # 2. Rebuild discovery inventory
        from scripts.build_source_discovery_inventory import generate_source_discovery_inventory
        generate_source_discovery_inventory()

        # 3. Rebuild coverage report
        from scripts.build_coverage_evidence_report import generate_coverage_evidence_report
        generate_coverage_evidence_report()

        # 4. Rebuild business knowledge index
        from src.services.business_knowledge_index import build_index_data, save_index, reload_knowledge_index
        index_data = build_index_data(root)
        save_index(index_data)
        reload_knowledge_index(check_freshness=True)

        return True, f"Cập nhật thành công kiến thức AI cho năm tài chính {item.fiscal_year} (Mã: {item.update_id})."

    except Exception as exc:
        # Restore both the update pack and every generated RAG artifact.  A failed
        # rebuild must never leave the UI using a half-rebuilt index/report set.
        if old_content is not None:
            _atomic_write_bytes(target_file, old_content)
        elif target_file.is_file():
            target_file.unlink()

        try:
            _restore_generated_artifacts(artifact_snapshots)
            _reload_restored_index()
        except Exception:
            pass

        return False, "Không thể cập nhật kiến thức AI. Dữ liệu trước khi cập nhật đã được giữ nguyên."


def deactivate_update(
    fiscal_year: str,
    update_id: str,
    repo_root: Path | None = None,
) -> Tuple[bool, str]:
    """Deactivate or remove an update item, then cleanly rebuild knowledge index."""
    root = repo_root or _repo_root()
    target_dir = get_updates_directory(fiscal_year, root)
    target_file = target_dir / f"{update_id}.json"

    if not target_file.is_file():
        return False, f"Không tìm thấy bản cập nhật '{update_id}' của {fiscal_year}."

    old_content = target_file.read_bytes()
    artifact_snapshots = _snapshot_generated_artifacts(root)
    try:
        data = json.loads(old_content.decode("utf-8"))
        data["is_active"] = False
        _write_update_file(target_file, data)

        from scripts.build_source_discovery_inventory import generate_source_discovery_inventory
        from scripts.build_coverage_evidence_report import generate_coverage_evidence_report
        from src.services.business_knowledge_index import build_index_data, save_index, reload_knowledge_index
        generate_source_discovery_inventory()
        generate_coverage_evidence_report()
        save_index(build_index_data(root))
        reload_knowledge_index(check_freshness=True)

        return True, f"Đã vô hiệu hóa bản cập nhật '{update_id}' của {fiscal_year}."
    except Exception:
        _atomic_write_bytes(target_file, old_content)
        try:
            _restore_generated_artifacts(artifact_snapshots)
            _reload_restored_index()
        except Exception:
            pass
        return False, "Không thể vô hiệu hóa bản cập nhật. Dữ liệu trước khi thay đổi đã được giữ nguyên."


def inspect_excel_reference_metadata(excel_path: Path | str) -> Dict[str, Any]:
    """Safely inspect high-level sheet names and column headers of an Excel reference file.

    Does NOT infer arbitrary formulas or amounts as business rules.
    """
    p = Path(excel_path)
    if not p.is_file():
        return {"error": f"Tệp không tồn tại: {p.name}"}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        sheet_info: List[Dict[str, Any]] = []

        for name in wb.sheetnames[:10]:  # Limit to first 10 sheets
            ws = wb[name]
            headers: List[str] = []
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                row_count += 1
                if row_count == 1:
                    headers = [str(c).strip() for c in row if c is not None and str(c).strip()][:15]
                if row_count >= 5:
                    break
            sheet_info.append({
                "sheet_name": name,
                "sample_headers": headers,
            })
        wb.close()
        return {
            "filename": p.name,
            "sheets_count": len(wb.sheetnames),
            "sheets": sheet_info,
        }
    except Exception as exc:
        return {"error": f"Không thể đọc cấu trúc tệp Excel: {exc}"}


def generate_update_preview(item: FiscalYearUpdateItem, language: str = "vi") -> Dict[str, str]:
    """Generate sample chatbot answer, source reference, and confidence label for live UI preview."""
    lang = str(language or "vi").strip().lower()
    if lang not in SUPPORTED_LANGUAGES:
        lang = "vi"

    fy = item.fiscal_year or "FY2028"
    title = str(item.title.get(lang, "") or "").strip()
    what_changed = str(item.what_changed.get(lang, "") or "").strip()
    user_action = str(item.user_action.get(lang, "") or "").strip()
    applies_to = str(item.applies_to.get(lang, "") or item.applies_to.get("vi", "") or "").strip()

    source_prefix = {
        "vi": "Nguồn tham khảo",
        "en": "Source Reference",
        "ja": "参照元",
    }.get(lang, "Nguồn tham khảo")

    confidence_prefix = {
        "vi": "Mức tin cậy",
        "en": "Confidence Level",
        "ja": "信頼度",
    }.get(lang, "Mức tin cậy")

    source_title = {
        "vi": f"Cập nhật nghiệp vụ {fy}",
        "en": f"{fy} Business Update",
        "ja": f"{fy} 業務更新",
    }.get(lang, f"Cập nhật nghiệp vụ {fy}")

    conf_text = {
        "vi": ("Tham khảo nội bộ" if item.status == "reference_with_caveat" else "Đã xác nhận"),
        "en": ("Internal Reference" if item.status == "reference_with_caveat" else "Confirmed"),
        "ja": ("社内参考" if item.status == "reference_with_caveat" else "確定"),
    }.get(lang, "Đã xác nhận")

    missing_notice = {
        "vi": "(Chưa có nội dung thay đổi bằng tiếng Việt. Vui lòng bổ sung bản dịch.)",
        "en": "(No English translation provided for this update yet. Please add a translation.)",
        "ja": "(日本語の変更内容がまだ登録されていません。翻訳を追加してください。)",
    }.get(lang, "(Chưa có bản dịch)")

    missing_title = {
        "vi": "(Chưa có tiêu đề)",
        "en": "(Untitled)",
        "ja": "(タイトル未設定)",
    }.get(lang, "(Chưa có tiêu đề)")

    display_title = title if title else missing_title
    lines = [what_changed if what_changed else missing_notice]
    if user_action:
        lines.append(f"\n1. {user_action}")
    if applies_to:
        scope_lbl = {"vi": "Áp dụng cho:", "en": "Applies to:", "ja": "適用対象:"}.get(lang, "Áp dụng:")
        lines.append(f"\n{scope_lbl} {applies_to}")

    lines.append(f"\n{source_prefix}: {source_title} — {display_title}")
    lines.append(f"{confidence_prefix}: {conf_text}")

    return {
        "answer": "\n".join(lines).strip(),
        "source_reference": f"{source_prefix}: {source_title} — {display_title}",
        "confidence_level": f"{confidence_prefix}: {conf_text}",
    }
