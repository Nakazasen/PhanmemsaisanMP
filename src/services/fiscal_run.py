"""Xác định nguồn theo năm tài chính và kiểm tra trước theo nguyên tắc đóng khi lỗi.

Mô-đun độc lập với bộ máy tính toán. Trước khi cho phép lần chạy thay đổi dữ
liệu tính, nó trả lời câu hỏi: các nguồn này có đúng là nguồn đã được duyệt cho
năm tài chính được chọn hay không?
"""

from __future__ import annotations

from dataclasses import dataclass, field, replace
from hashlib import sha256
from pathlib import Path
from functools import lru_cache
import os
import re
import subprocess
import uuid
from typing import Callable, Iterable

import openpyxl

from src.engine.uniform_cup_rules import UNIFORM_ITEM_SPECS, normalize_uniform_text
from src.utils import excel_helpers
from src.utils.fiscal_periods import (
    SystemSourcePeriodError,
    fiscal_baseline_period,
    fiscal_periods,
    map_system_source_periods,
    system_source_periods_for_file,
)
from src.utils.source_manifest import read_source_manifest, resolve_manifest_files


SOURCE_CATEGORIES = (
    "facility",
    "fixed_assets",
    "it_simulation",
    "ga",
    "birthday",
    "allocation_rules",
    "nnn_paperwork",
    "uniform_policy",
)
# Compatibility alias: this is a registry of supported source categories, not a
# policy that every category must exist for every fiscal year.
REQUIRED_SOURCE_CATEGORIES = SOURCE_CATEGORIES
ISSUE_BLOCKING = "BLOCKING"
ISSUE_SOURCE_SKIPPED = "SOURCE_SKIPPED"
ISSUE_CONTINUABLE_MISSING_SOURCE = ISSUE_SOURCE_SKIPPED
ISSUE_INFORMATION = "INFORMATION"
SOURCE_SCOPED_CATEGORIES = frozenset(SOURCE_CATEGORIES) | {
    "source_inventory",
    "form_uniform_master",
}
REFERENCE_POLICY_DISABLED = "DISABLED"
REFERENCE_POLICY_EXPLICIT_SAME_FY = "EXPLICIT_SAME_FY"
REFERENCE_POLICY_LEGACY_FY2027_MAP = "LEGACY_FY2027_MAP"

_FY_PATTERN = re.compile(r"(?<!\d)FY\s*(20\d{2})(?!\d)", re.IGNORECASE)
_MPFY_PATTERN = re.compile(r"MPFY\s*(20\d{2})(?!\d)", re.IGNORECASE)
_PERIOD_PATTERN = re.compile(r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(?!\d)")


@dataclass(frozen=True)
class FiscalYearEvidence:
    """All FY markers found in one source, without choosing one arbitrarily."""

    path: str
    filename_years: tuple[int, ...] = ()
    directory_years: tuple[int, ...] = ()
    sheet_years: tuple[int, ...] = ()
    content_years: tuple[int, ...] = ()
    inspection_error: str | None = None

    @property
    def years(self) -> tuple[int, ...]:
        return tuple(sorted(set((*self.filename_years, *self.directory_years, *self.sheet_years, *self.content_years))))

    @property
    def anchor_years(self) -> tuple[int, ...]:
        """FY markers that identify the selected annual source, not old values."""
        return tuple(sorted(set((*self.filename_years, *self.directory_years, *self.sheet_years))))

    @property
    def resolved_fiscal_year(self) -> int | None:
        if len(self.anchor_years) == 1:
            return self.anchor_years[0]
        return self.content_years[0] if not self.anchor_years and len(self.content_years) == 1 else None

    @property
    def conflict(self) -> bool:
        # A historic price/snapshot year in a cell must not override a clear
        # FY label in the filename, folder or sheet.  Only conflicting annual
        # anchors are a cross-year safety failure.
        return len(self.anchor_years) > 1 or (not self.anchor_years and len(self.content_years) > 1)

    def as_dict(self) -> dict[str, object]:
        return {
            "path": self.path,
            "filename_years": list(self.filename_years),
            "directory_years": list(self.directory_years),
            "sheet_years": list(self.sheet_years),
            "content_years": list(self.content_years),
            "all_years": list(self.years),
            "inspection_error": self.inspection_error,
        }


@dataclass(frozen=True)
class SourceIssue:
    category: str
    path: str
    detected_fiscal_year: int | None
    reason: str
    action: str
    status: str = "FAILED"
    sheet: str | None = None
    period_coverage: tuple[str, ...] = ()
    code: str = "SOURCE_VALIDATION_FAILED"
    severity: str = ISSUE_BLOCKING
    impact: str = "Không thể bảo đảm kết quả tính toán chính xác."

    @property
    def is_continuable(self) -> bool:
        return self.severity == ISSUE_SOURCE_SKIPPED

    @property
    def is_source_skipped(self) -> bool:
        return self.severity == ISSUE_SOURCE_SKIPPED

    def as_dict(self) -> dict[str, object]:
        candidate = Path(self.path) if self.path else None
        return {
            "category": self.category,
            "selected_path": self.path,
            "detected_fiscal_year": self.detected_fiscal_year,
            "expected_fiscal_year": None,
            "status": self.status,
            "code": self.code,
            "severity": self.severity,
            "impact": self.impact,
            "checksum": sha256_file(candidate) if candidate and candidate.is_file() else None,
            "sheet": self.sheet,
            "period_coverage": list(self.period_coverage),
            "reason": self.reason,
            "required_action": self.action,
        }

    @classmethod
    def from_dict(cls, payload: dict[str, object]) -> "SourceIssue":
        return cls(
            category=str(payload.get("category", "")),
            path=str(payload.get("selected_path", "")),
            detected_fiscal_year=(
                int(payload["detected_fiscal_year"])
                if payload.get("detected_fiscal_year") is not None else None
            ),
            reason=str(payload.get("reason", "")),
            action=str(payload.get("required_action", "")),
            status=str(payload.get("status", "FAILED")),
            sheet=str(payload["sheet"]) if payload.get("sheet") is not None else None,
            period_coverage=tuple(str(value) for value in payload.get("period_coverage", []) or []),
            code=str(payload.get("code", "SOURCE_VALIDATION_FAILED")),
            severity=str(payload.get("severity", ISSUE_BLOCKING)),
            impact=str(payload.get("impact", "Không thể bảo đảm kết quả tính toán chính xác.")),
        )

    def as_text(self) -> str:
        detected = f"FY{self.detected_fiscal_year}" if self.detected_fiscal_year else "không xác định"
        return (
            f"[{self.category}/{self.code}/{self.severity}] file={self.path or '(không có)'}; "
            f"năm phát hiện={detected}; lý do={self.reason}; cần làm={self.action}"
        )


@dataclass(frozen=True)
class SourceCheck:
    """One auditable preflight decision, including successful sources."""

    category: str
    selected_path: str
    detected_fiscal_year: int | None
    expected_fiscal_year: int
    status: str
    checksum: str | None
    sheet: str | None = None
    period_coverage: tuple[str, ...] = ()
    reason: str = ""
    required_action: str = ""

    def as_dict(self) -> dict[str, object]:
        return {
            "category": self.category,
            "selected_path": self.selected_path,
            "detected_fiscal_year": self.detected_fiscal_year,
            "expected_fiscal_year": self.expected_fiscal_year,
            "status": self.status,
            "checksum": self.checksum,
            "sheet": self.sheet,
            "period_coverage": list(self.period_coverage),
            "reason": self.reason,
            "required_action": self.required_action,
        }

    @classmethod
    def from_dict(cls, payload: dict[str, object]) -> "SourceCheck":
        return cls(
            category=str(payload.get("category", "")),
            selected_path=str(payload.get("selected_path", "")),
            detected_fiscal_year=(
                int(payload["detected_fiscal_year"])
                if payload.get("detected_fiscal_year") is not None else None
            ),
            expected_fiscal_year=int(payload.get("expected_fiscal_year", 0) or 0),
            status=str(payload.get("status", "FAILED")),
            checksum=str(payload["checksum"]) if payload.get("checksum") is not None else None,
            sheet=str(payload["sheet"]) if payload.get("sheet") is not None else None,
            period_coverage=tuple(str(value) for value in payload.get("period_coverage", []) or []),
            reason=str(payload.get("reason", "")),
            required_action=str(payload.get("required_action", "")),
        )


@dataclass(frozen=True)
class RunPreflightReport:
    fiscal_year: int
    issues: tuple[SourceIssue, ...] = ()
    resolved_sources: dict[str, tuple[str, ...]] = field(default_factory=dict)
    checks: tuple[SourceCheck, ...] = ()

    def as_dict(self) -> dict[str, object]:
        issues = []
        for issue in self.issues:
            row = issue.as_dict()
            row["expected_fiscal_year"] = self.fiscal_year
            issues.append(row)
        return {
            "fiscal_year": self.fiscal_year,
            "ok": self.ok,
            "can_continue_incomplete": self.can_continue_incomplete,
            "issues": issues,
            "checks": [check.as_dict() for check in self.checks],
            "resolved_sources": {key: list(value) for key, value in self.resolved_sources.items()},
        }

    @classmethod
    def from_dict(cls, payload: dict[str, object]) -> "RunPreflightReport":
        return cls(
            fiscal_year=int(payload.get("fiscal_year", 0) or 0),
            issues=tuple(SourceIssue.from_dict(row) for row in payload.get("issues", []) or []),
            resolved_sources={
                str(key): tuple(str(path) for path in paths)
                for key, paths in dict(payload.get("resolved_sources", {}) or {}).items()
            },
            checks=tuple(SourceCheck.from_dict(row) for row in payload.get("checks", []) or []),
        )

    def as_markdown(self) -> str:
        """Human-readable companion to the JSON provenance report."""
        lines = [
            f"# Kiểm tra nguồn trước khi chạy FY{self.fiscal_year}",
            "",
            "| Nhóm nguồn | Trạng thái | File đã chọn | Năm phát hiện | Trang | Kỳ đọc được | Lý do / việc cần làm |",
            "|---|---|---|---|---|---|---|",
        ]
        for check in self.checks:
            year = f"FY{check.detected_fiscal_year}" if check.detected_fiscal_year else "Không xác định"
            detail = check.reason or check.required_action or "Đạt"
            path = check.selected_path or "(chưa có)"
            lines.append(
                "| {category} | {status} | {path} | {year} | {sheet} | {periods} | {detail} |".format(
                    category=check.category,
                    status=check.status,
                    path=path.replace("|", "\\|"),
                    year=year,
                    sheet=check.sheet or "",
                    periods=", ".join(check.period_coverage),
                    detail=detail.replace("|", "\\|"),
                )
            )
        return "\n".join(lines) + "\n"

    @property
    def ok(self) -> bool:
        return not self.issues

    @property
    def blocking_issues(self) -> tuple[SourceIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == ISSUE_BLOCKING)

    @property
    def skipped_issues(self) -> tuple[SourceIssue, ...]:
        return tuple(issue for issue in self.issues if issue.is_source_skipped)

    @property
    def continuable_issues(self) -> tuple[SourceIssue, ...]:
        """Compatibility alias for callers built before subset-run support."""
        return self.skipped_issues

    @property
    def can_run(self) -> bool:
        """Whether shared run prerequisites are safe, regardless of source count."""
        return not self.blocking_issues

    @property
    def can_continue_incomplete(self) -> bool:
        return self.can_run and bool(self.skipped_issues)

    @property
    def usable_sources(self) -> dict[str, tuple[str, ...]]:
        """Only source paths that passed validation; never rediscover rejected files."""
        rejected_categories = {
            issue.category for issue in self.skipped_issues
            if issue.category in SOURCE_SCOPED_CATEGORIES
            and issue.category != "it_simulation"
        }
        # Inventory issues point at one unclassified workbook and must not hide
        # otherwise valid, explicitly classified categories.
        rejected_categories.discard("source_inventory")
        if "form_uniform_master" in rejected_categories:
            rejected_categories.add("uniform_policy")

        system_paths = tuple(self.resolved_sources.get("it_simulation", ()))
        rejected_system_paths: set[str] = set()
        reject_all_system = False
        for issue in self.skipped_issues:
            if issue.category != "it_simulation":
                continue
            if issue.path in system_paths:
                rejected_system_paths.add(issue.path)
            else:
                # Overlap/duplicate issues can involve more than one path. Do not
                # guess which file should win; isolate the category in that case.
                reject_all_system = True

        usable: dict[str, tuple[str, ...]] = {}
        for category, paths in self.resolved_sources.items():
            if category in rejected_categories or not paths:
                continue
            if category == "it_simulation":
                filtered = () if reject_all_system else tuple(
                    path for path in paths if path not in rejected_system_paths
                )
                if filtered:
                    usable[category] = filtered
            else:
                usable[category] = tuple(paths)
        return usable

    def accepted_missing_categories(self) -> tuple[str, ...]:
        """Compatibility metadata; subset runs no longer require acceptance flags."""
        return tuple(sorted({issue.category for issue in self.skipped_issues}))

    def unaccepted_issues(self, accepted_categories: Iterable[str] = ()) -> tuple[SourceIssue, ...]:
        # Source-scoped failures are always isolated. The argument remains for
        # compatibility with saved commands and older callers.
        return self.blocking_issues

    def raise_if_invalid(self, accepted_categories: Iterable[str] = ()) -> None:
        remaining = self.blocking_issues
        if not remaining:
            return
        detail = "\n".join(f"- {issue.as_text()}" for issue in remaining)
        raise ValueError(
            f"Kiểm tra điều kiện chạy FY{self.fiscal_year} không đạt. Chương trình chưa tạo dữ liệu tính toán.\n{detail}"
        )


@dataclass(frozen=True)
class FiscalRunContext:
    run_id: str
    fiscal_year: int
    fiscal_periods: tuple[str, ...]
    baseline_period: str
    template_path: str
    source_dir: str
    headcount_source_dir: str
    uniform_policy_path: str | None
    output_dir: str
    exchange_rate: float
    exchange_rate_source: str = "not declared"
    resolved_sources: dict[str, tuple[str, ...]] = field(default_factory=dict)
    history_root: str | None = None
    workspace_dir: str | None = None
    database_path: str | None = None
    manual_input_store: str | None = None
    ordered_sources: tuple[dict[str, object], ...] = ()
    reference_policy: str = REFERENCE_POLICY_DISABLED
    application_version: str = "unknown"

    def with_resolution(self, sources: dict[str, tuple[str, ...]]) -> "FiscalRunContext":
        return replace(self, resolved_sources=sources, ordered_sources=resolve_ordered_sources(self))


def _base_dir() -> Path:
    return Path(__file__).resolve().parents[2]


def annual_default_paths(fiscal_year: int, base_dir: str | os.PathLike[str] | None = None) -> dict[str, str]:
    root = Path(base_dir) if base_dir else _base_dir()
    return {
        "template_path": str(root / "docs" / f"MP{fiscal_year}" / "FORM.xlsx"),
        "source_dir": str(root / "docs" / f"MP{fiscal_year}"),
        "headcount_source_dir": str(root / "raw" / f"FY{fiscal_year}"),
        "output_dir": str(root / f"OUTPUT_FY{fiscal_year}"),
        "history_root": str(root / "RUN_HISTORY"),
        "manual_input_store": str(root / "raw" / f"FY{fiscal_year}" / "manual_inputs.db"),
    }


def _has_headcount_workbook(directory: Path) -> bool:
    return directory.is_dir() and any(path.suffix.lower() in {".xls", ".xlsx", ".xlsm"} for path in directory.iterdir())


def _legacy_uniform_policy_path(root: Path) -> Path | None:
    for path in sorted((root / "raw").glob("*.xlsx")):
        try:
            if _is_uniform_policy(path):
                return path
        except Exception:
            continue
    return None


def _file_cache_key(path: str | os.PathLike[str]) -> tuple[str, int, int]:
    """Cache key that changes whenever the source file changes."""
    candidate = Path(path).resolve()
    stat = candidate.stat()
    return str(candidate), int(stat.st_mtime_ns), int(stat.st_size)


@lru_cache(maxsize=128)
def _is_uniform_policy_cached(path: str, _mtime_ns: int, _size: int) -> bool:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "原価センタ" not in workbook.sheetnames:
            return False
        sheet = workbook["原価センタ"]
        headers = {
            normalize_uniform_text(sheet.cell(1, column).value)
            for column in range(1, sheet.max_column + 1)
        }
        return all(normalize_uniform_text(spec.header) in headers for spec in UNIFORM_ITEM_SPECS)
    finally:
        workbook.close()


def _is_uniform_policy(path: Path) -> bool:
    try:
        return _is_uniform_policy_cached(*_file_cache_key(path))
    except OSError:
        return False


def resolve_uniform_policy_path(
    fiscal_year: int,
    explicit_path: str | None = None,
    *,
    base_dir: str | os.PathLike[str] | None = None,
) -> str | None:
    if explicit_path:
        return str(Path(explicit_path).resolve())

    root = Path(base_dir) if base_dir else _base_dir()
    annual_dir = root / "raw" / f"FY{fiscal_year}"
    candidates = []
    if annual_dir.is_dir():
        candidates = [path for path in annual_dir.glob("*.xlsx") if _is_uniform_policy(path)]
    if len(candidates) == 1:
        return str(candidates[0].resolve())
    if len(candidates) > 1:
        raise ValueError(
            f"FY{fiscal_year} có nhiều file bảng dấu chọn đồng phục/cốc xếp: "
            + "; ".join(str(path) for path in candidates)
        )
    if fiscal_year == 2027:
        legacy = _legacy_uniform_policy_path(root)
        return str(legacy.resolve()) if legacy else None
    return None


def create_fiscal_run_context(
    fiscal_year: int,
    *,
    template_path: str | None = None,
    source_dir: str | None = None,
    headcount_source_dir: str | None = None,
    uniform_policy_path: str | None = None,
    output_dir: str | None = None,
    exchange_rate: float = 25450.0,
    exchange_rate_source: str = "explicit pipeline input",
    history_root: str | None = None,
    manual_input_store: str | None = None,
    reference_policy: str | None = None,
    base_dir: str | os.PathLike[str] | None = None,
    run_id: str | None = None,
) -> FiscalRunContext:
    year = int(fiscal_year)
    defaults = annual_default_paths(year, base_dir)
    root = Path(base_dir) if base_dir else _base_dir()
    resolved_headcount_dir = headcount_source_dir or defaults["headcount_source_dir"]
    # The FY2027 delivered data predates the annual raw/FY folder convention.
    # Preserve that compatibility only when its annual folder has no workbooks.
    if year == 2027 and not headcount_source_dir and not _has_headcount_workbook(Path(resolved_headcount_dir)):
        resolved_headcount_dir = str(root / "raw")
    resolved_uniform = resolve_uniform_policy_path(year, uniform_policy_path, base_dir=root)
    try:
        git_revision = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"], cwd=root, text=True, stderr=subprocess.DEVNULL
        ).strip()
    except Exception:
        git_revision = "unknown"
    release_version = os.environ.get("MP_MANAGER_VERSION", "unversioned")
    application_version = f"release={release_version};git={git_revision}"
    resolved_reference_policy = reference_policy or (
        REFERENCE_POLICY_LEGACY_FY2027_MAP
        if year == 2027 else REFERENCE_POLICY_DISABLED
    )
    if resolved_reference_policy not in {
        REFERENCE_POLICY_DISABLED,
        REFERENCE_POLICY_EXPLICIT_SAME_FY,
        REFERENCE_POLICY_LEGACY_FY2027_MAP,
    }:
        raise ValueError(f"Chính sách file tham khảo không hợp lệ: {resolved_reference_policy}")
    if year != 2027 and resolved_reference_policy == REFERENCE_POLICY_LEGACY_FY2027_MAP:
        raise ValueError("Chỉ FY2027 mới được dùng bảng tham khảo tương thích cũ.")
    return FiscalRunContext(
        run_id=run_id or uuid.uuid4().hex,
        fiscal_year=year,
        fiscal_periods=tuple(fiscal_periods(year)),
        baseline_period=fiscal_baseline_period(year),
        template_path=str(Path(template_path or defaults["template_path"]).resolve()),
        source_dir=str(Path(source_dir or defaults["source_dir"]).resolve()),
        headcount_source_dir=str(Path(resolved_headcount_dir).resolve()),
        uniform_policy_path=resolved_uniform,
        output_dir=str(Path(output_dir or defaults["output_dir"]).resolve()),
        exchange_rate=float(exchange_rate),
        exchange_rate_source=str(exchange_rate_source or "").strip(),
        history_root=str(Path(history_root or defaults["history_root"]).resolve()),
        manual_input_store=str(Path(manual_input_store or defaults["manual_input_store"]).resolve()),
        reference_policy=resolved_reference_policy,
        application_version=application_version,
    )


def _years_in_text(value: object, *, allow_periods: bool = True) -> set[int]:
    text = str(value or "")
    years = {
        int(match.group(1))
        for pattern in (_FY_PATTERN, _MPFY_PATTERN)
        for match in pattern.finditer(text)
    }
    if allow_periods:
        for match in _PERIOD_PATTERN.finditer(text):
            calendar_year, month = int(match.group(1)), int(match.group(2))
            years.add(calendar_year + 1 if month >= 4 else calendar_year)
    return years


def inspect_fiscal_year_evidence(path: str | os.PathLike[str]) -> FiscalYearEvidence:
    """Inspect filename, sheet names and visible workbook text for FY markers.

    A source is intentionally *not* accepted based on the first marker found.
    This prevents a FY2028 filename from masking a FY2027 business sheet.
    """
    candidate = Path(path).resolve()
    try:
        stat = candidate.stat()
    except OSError:
        stat = None
    return _inspect_fiscal_year_evidence_cached(
        str(candidate),
        int(stat.st_mtime_ns) if stat else -1,
        int(stat.st_size) if stat else -1,
    )


@lru_cache(maxsize=256)
def _inspect_fiscal_year_evidence_cached(
    resolved_path: str,
    _mtime_ns: int,
    _size: int,
) -> FiscalYearEvidence:
    candidate = Path(resolved_path)
    filename_years = _years_in_text(candidate.name)
    directory_years = _years_in_text(candidate.parent.name, allow_periods=False)
    sheet_years: set[int] = set()
    content_years: set[int] = set()
    suffix = candidate.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return FiscalYearEvidence(
            str(candidate), tuple(sorted(filename_years)), tuple(sorted(directory_years))
        )
    try:
        workbook = openpyxl.load_workbook(candidate, read_only=True, data_only=True)
    except Exception as exc:
        return FiscalYearEvidence(
            str(candidate), tuple(sorted(filename_years)), tuple(sorted(directory_years)), inspection_error=str(exc)
        )
    try:
        for name in workbook.sheetnames:
            sheet_years.update(_years_in_text(name))
            # A bounded scan is deliberate: business FY headings are normally
            # in the title area and scanning entire large workbooks is unsafe.
            sheet = workbook[name]
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True):
                for value in row[: min(sheet.max_column, 20)]:
                    content_years.update(_years_in_text(value, allow_periods=False))
    finally:
        workbook.close()
    return FiscalYearEvidence(
        str(candidate),
        tuple(sorted(filename_years)),
        tuple(sorted(directory_years)),
        tuple(sorted(sheet_years)),
        tuple(sorted(content_years)),
    )


def validate_fiscal_year_evidence(
    evidence: FiscalYearEvidence,
    expected_fiscal_year: int,
    *,
    required: bool = True,
) -> str | None:
    """Return a deterministic rejection reason, or ``None`` when valid."""
    expected = int(expected_fiscal_year)
    if evidence.conflict:
        if not required and not evidence.anchor_years:
            # Fixed-asset snapshots often contain comparative FY headers.  A
            # separate parser audit validates their month range; do not reject
            # the source merely for historical labels in its cells.
            return None
        found = ", ".join(f"FY{year}" for year in (evidence.anchor_years or evidence.content_years))
        return f"Nguồn chứa dấu hiệu nhiều năm tài chính khác nhau: {found}."
    if evidence.resolved_fiscal_year is None:
        if required:
            return "Không xác định được năm tài chính trong tên file, trang hoặc phần tiêu đề nguồn."
        return None
    if evidence.resolved_fiscal_year != expected:
        return f"Nguồn thuộc FY{evidence.resolved_fiscal_year}, không phải FY{expected}."
    return None


def detect_fiscal_year(path: str | os.PathLike[str]) -> int | None:
    """Compatibility helper that returns a year only when all evidence agrees."""
    return inspect_fiscal_year_evidence(path).resolved_fiscal_year


def _manifest_sources(source_dir: str, category: str) -> tuple[str, ...]:
    return tuple(str(Path(path).resolve()) for path in resolve_manifest_files(source_dir, category))


def resolve_annual_sources(context: FiscalRunContext) -> dict[str, tuple[str, ...]]:
    """Resolve every business source inside the selected annual source folder only."""
    sources: dict[str, tuple[str, ...]] = {}
    for category in REQUIRED_SOURCE_CATEGORIES:
        if category == "uniform_policy":
            sources[category] = (context.uniform_policy_path,) if context.uniform_policy_path else ()
        else:
            sources[category] = _manifest_sources(context.source_dir, category)
    return sources


def resolve_ordered_sources(context: FiscalRunContext) -> tuple[dict[str, object], ...]:
    """Return the ordered annual manifest as immutable run provenance."""
    entries = read_source_manifest(context.source_dir, include_missing=True)
    result: list[dict[str, object]] = []
    for entry in entries:
        path = str(entry.get("_path", ""))
        result.append({
            "category": str(entry.get("category", "")),
            "order": int(str(entry.get("order", "9999")) or 9999),
            "path": path,
            "filename": str(entry.get("filename", "")),
            "status": str(entry.get("status", "recognized")),
            "detection_method": str(entry.get("detection_method", "")),
            "reason": str(entry.get("reason", "")),
            "fiscal_year": detect_fiscal_year(path) if path else None,
            "checksum": sha256_file(path) if path and Path(path).is_file() else None,
            "invalid_path": bool(entry.get("_invalid_path")),
        })
    return tuple(sorted(result, key=lambda row: (int(row["order"]), str(row["filename"]).lower())))


@lru_cache(maxsize=128)
def _workbook_sheet_names_cached(path: str, _mtime_ns: int, _size: int) -> tuple[str, ...]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return tuple(str(name) for name in workbook.sheetnames)
    finally:
        workbook.close()


def _workbook_sheet_names(path: str) -> tuple[str, ...]:
    return _workbook_sheet_names_cached(*_file_cache_key(path))


def _allocation_has_exact_sheet(path: str, fiscal_year: int) -> bool:
    try:
        names = _workbook_sheet_names(path)
    except Exception:
        return False
    return any(f"FY{fiscal_year}" in name.upper() for name in names)


def _workbook_has_fiscal_sheet(path: str, fiscal_year: int) -> bool:
    """Require an explicit FY marker in a workbook that is yearly by design."""
    try:
        names = _workbook_sheet_names(path)
    except Exception:
        return False
    expected = f"FY{int(fiscal_year)}"
    return any(expected in name.upper().replace(" ", "") for name in names)


def _matching_fiscal_sheet(path: str, fiscal_year: int) -> str | None:
    """Return the FY-labelled sheet used as evidence, without guessing."""
    try:
        names = _workbook_sheet_names(path)
    except Exception:
        return None
    expected = f"FY{int(fiscal_year)}"
    for name in names:
        if expected in name.upper().replace(" ", ""):
            return name
    return None


def _workbook_period_coverage(path: str, fiscal_year: int) -> tuple[str, ...]:
    """Read explicit YYYYMM headings from a bounded workbook area for audit."""
    candidate = Path(path)
    if candidate.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return ()
    values: set[str] = set()
    expected = set(fiscal_periods(fiscal_year))
    try:
        workbook = openpyxl.load_workbook(candidate, read_only=True, data_only=True)
    except Exception:
        return ()
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), max_col=min(sheet.max_column, 80), values_only=True):
                for value in row:
                    for match in _PERIOD_PATTERN.finditer(str(value or "")):
                        period = f"{match.group(1)}{match.group(2)}"
                        if period in expected:
                            values.add(period)
    finally:
        workbook.close()
    return tuple(period for period in fiscal_periods(fiscal_year) if period in values)


def _filename_period_coverage(path: str, fiscal_year: int) -> tuple[str, ...]:
    """Return shared System Cost filename coverage for the audit table."""
    try:
        return system_source_periods_for_file(path, fiscal_year)
    except SystemSourcePeriodError:
        return ()


def _preflight_headcount_coverage(context: FiscalRunContext) -> tuple[tuple[str, ...], str | None]:
    """Validate the selected FY's staffing files without writing a run DB."""
    try:
        from src.services.headcount_source_importer import scan_headcount_time_sources
        results = scan_headcount_time_sources(context.headcount_source_dir, context.fiscal_year)
    except Exception as exc:
        return (), f"Không đọc được nguồn nhân sự/thời gian: {exc}"
    valid = [result for result in results if getattr(result, "status", "") == "valid"]
    if not valid:
        return (), "Không có file nhân sự/thời gian đúng cấu trúc và đúng FY trong thư mục đã chọn."
    expected = tuple(fiscal_periods(context.fiscal_year))
    missing_files = []
    for result in valid:
        covered = {str(row.get("period", "")) for row in getattr(result, "rows", ())}
        if not set(expected).issubset(covered):
            missing_files.append(Path(str(getattr(result, "path", ""))).name)
    if missing_files:
        return expected, "Nguồn nhân sự thiếu một hoặc nhiều tháng FY: " + ", ".join(missing_files[:8])
    return expected, None


def _validate_manual_input_store(context: FiscalRunContext) -> str | None:
    """Reject a future-FY manual store that is unscoped or contains another FY."""
    if context.fiscal_year == 2027:
        return None  # handled by the explicit one-time FY2027 migration
    store = Path(context.manual_input_store or "")
    if not store:
        return "Chưa xác định được kho dữ liệu nhập tay theo năm."
    try:
        store.resolve().relative_to(Path(context.headcount_source_dir).resolve())
    except ValueError:
        return "Kho dữ liệu nhập tay phải nằm trong raw/FY của năm đang chạy."
    if not store.exists():
        return None
    import sqlite3
    try:
        conn = sqlite3.connect(store)
        try:
            for table in (
                "fact_manual_headcount_time_override",
                "fact_manual_headcount_baseline_override",
                "fact_bus_headcount_drivers",
            ):
                exists = conn.execute(
                    "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
                ).fetchone()
                if not exists:
                    continue
                columns = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
                if "fiscal_year" not in columns:
                    return f"Kho nhập tay thiếu trường fiscal_year ở bảng {table}."
                foreign = conn.execute(
                    f"SELECT COUNT(*) FROM {table} WHERE fiscal_year<>?", (context.fiscal_year,)
                ).fetchone()[0]
                if foreign:
                    return f"Kho nhập tay có {foreign} dòng {table} không thuộc FY{context.fiscal_year}."
        finally:
            conn.close()
    except sqlite3.Error as exc:
        return f"Không đọc được kho dữ liệu nhập tay: {exc}"
    return None


def _preflight_checks(
    context: FiscalRunContext,
    sources: dict[str, tuple[str, ...]],
    issues: list[SourceIssue],
) -> tuple[SourceCheck, ...]:
    """Build a complete source table; failures are never hidden by JSON only."""
    candidates: list[tuple[str, str]] = [
        ("template", context.template_path),
        ("headcount", context.headcount_source_dir),
        ("manual_inputs", str(context.manual_input_store or "")),
        ("exchange_rate", context.exchange_rate_source),
    ]
    for category in REQUIRED_SOURCE_CATEGORIES:
        paths = sources.get(category, ())
        if paths:
            candidates.extend((category, str(path)) for path in paths)
        else:
            candidates.append((category, ""))
    if any(issue.category == "form_uniform_master" for issue in issues):
        candidates.append(("form_uniform_master", context.uniform_policy_path or context.template_path))
    checks: list[SourceCheck] = []
    for category, path in candidates:
        is_file = bool(path) and Path(path).is_file()
        evidence = inspect_fiscal_year_evidence(path) if is_file else FiscalYearEvidence(path)
        related = [
            issue for issue in issues
            if issue.category == category and (not issue.path or issue.path == path or path in issue.path)
        ]
        issue = related[0] if related else None
        checks.append(SourceCheck(
            category=category,
            selected_path=path,
            detected_fiscal_year=evidence.resolved_fiscal_year,
            expected_fiscal_year=context.fiscal_year,
            status=(
                "SKIPPED" if issue and issue.is_source_skipped else
                "FAILED" if issue else
                "OK"
            ),
            checksum=sha256_file(path) if is_file else None,
            sheet=(
                "原価センタ" if category == "uniform_policy" and is_file else
                _matching_fiscal_sheet(path, context.fiscal_year) if is_file else None
            ),
            period_coverage=(
                issue.period_coverage if issue and issue.period_coverage else
                _filename_period_coverage(path, context.fiscal_year) if category == "it_simulation" and is_file else
                _workbook_period_coverage(path, context.fiscal_year) if is_file else ()
            ),
            reason=(
                issue.reason if issue else
                f"Tỷ giá {context.exchange_rate:,.6g} từ {context.exchange_rate_source}."
                if category == "exchange_rate" else
                "Đã xác nhận nguồn phù hợp với lần chạy."
            ),
            required_action=issue.action if issue else "",
        ))
    return tuple(checks)


def _template_validation_error(path: str) -> str | None:
    candidate = Path(path)
    if candidate.name.lower() == "form_old.xlsx" or not candidate.is_file():
        return "FORM không tồn tại hoặc không đúng cấu trúc."
    try:
        workbook = openpyxl.load_workbook(candidate, read_only=True, data_only=False)
    except Exception:
        return "FORM không thể mở hoặc không đúng cấu trúc."
    try:
        if not any("内訳" in name or "採算" in name for name in workbook.sheetnames):
            return "FORM không có sheet chi tiết MP đúng cấu trúc."
        try:
            issue_cells = excel_helpers.find_form_template_hygiene_issues(workbook)
        except ValueError:
            return "FORM không có sheet chi tiết MP đúng cấu trúc."
        if issue_cells:
            return (
                "FORM còn dữ liệu của bộ phận cũ tại các ô: "
                f"{', '.join(issue_cells)}."
            )
        return None
    finally:
        workbook.close()


def _template_looks_valid(path: str) -> bool:
    return _template_validation_error(path) is None


def _validate_form_master_and_uniform(context: FiscalRunContext) -> str | None:
    """Use production readers against an in-memory DB before calculation starts."""
    if not Path(context.template_path).is_file() or not context.uniform_policy_path:
        return None
    import sqlite3
    from src.db.loader import load_accounts, load_cost_centers, load_uniform_entitlements
    from src.db.schema import create_schema

    conn = sqlite3.connect(":memory:")
    try:
        create_schema(conn)
        cc_count = load_cost_centers(conn, context.template_path)
        account_count = load_accounts(conn, context.template_path)
        if cc_count <= 0 or account_count <= 0:
            return "FORM không đọc được danh mục phòng ban hoặc tài khoản."
        load_uniform_entitlements(
            conn,
            context.uniform_policy_path,
            fiscal_year=context.fiscal_year,
        )
    except Exception as exc:
        return str(exc)
    finally:
        conn.close()
    return None


def preflight_fiscal_run(
    context: FiscalRunContext,
    *,
    progress: Callable[[str], None] | None = None,
) -> RunPreflightReport:
    """Validate annual isolation without writing to calculation or output storage."""
    notify = progress or (lambda _message: None)
    issues: list[SourceIssue] = []
    notify("Đang kiểm kê toàn bộ workbook nguồn chi phí...")
    sources = resolve_annual_sources(context)
    ordered = resolve_ordered_sources(context)
    for entry in ordered:
        if entry["invalid_path"]:
            issues.append(SourceIssue(
                str(entry["category"]), str(entry["filename"]), None,
                "Đường dẫn manifest đi ra ngoài thư mục nguồn của năm", "Sửa filename trong manifest; không dùng ../ hoặc đường dẫn tuyệt đối."
            ))
        if entry["status"] == "needs_review":
            issues.append(SourceIssue(
                "source_inventory",
                str(entry["filename"]),
                None,
                "Workbook chưa được chọn để tính: " + (str(entry["reason"]) or "chưa nhận diện được loại nguồn"),
                "Mở ‘Thứ tự file nguồn’ để xác nhận loại, hoặc giữ trạng thái Bỏ qua.",
                code="SOURCE_NEEDS_REVIEW",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Workbook này không được đưa vào kết quả; các nguồn đã xác nhận vẫn được chạy.",
            ))

    notify("Đang kiểm tra FORM sạch và đúng cấu trúc...")
    template_error = _template_validation_error(context.template_path)
    if template_error:
        issues.append(SourceIssue(
            "template",
            context.template_path,
            detect_fiscal_year(context.template_path),
            template_error,
            "Chọn FORM sạch đúng FY; không dùng FORM còn mã phòng, nhân sự hoặc chi phí cũ.",
            code="FORM_TEMPLATE_NOT_CLEAN",
        ))
    if not Path(context.source_dir).is_dir():
        issues.append(SourceIssue("source_dir", context.source_dir, None, "Thư mục nguồn chi phí không tồn tại", f"Tạo hoặc chọn docs/MP{context.fiscal_year}."))
    notify("Đang kiểm tra nguồn nhân sự và dữ liệu nhập tay...")
    if not Path(context.headcount_source_dir).is_dir():
        issues.append(SourceIssue("headcount", context.headcount_source_dir, None, "Thư mục nguồn nhân sự không tồn tại", f"Tạo hoặc chọn raw/FY{context.fiscal_year}."))
    elif not _has_headcount_workbook(Path(context.headcount_source_dir)):
        issues.append(SourceIssue("headcount", context.headcount_source_dir, None, "Không có file nguồn nhân sự/thời gian", "Bổ sung đủ nguồn nhân sự có tháng 3 mốc và 12 tháng FY."))
    elif context.fiscal_year >= 2028:
        coverage, coverage_error = _preflight_headcount_coverage(context)
        if coverage_error:
            issues.append(SourceIssue(
                "headcount", context.headcount_source_dir, context.fiscal_year,
                coverage_error,
                "Bổ sung đủ 12 tháng nhân sự/thời gian theo đúng mẫu FY; tháng 3 mốc phải được nhập tay có truy vết nếu không có trong file.",
                period_coverage=coverage,
            ))
    if context.exchange_rate <= 0 or not context.exchange_rate_source:
        issues.append(SourceIssue(
            "exchange_rate", "", None,
            "Tỷ giá không hợp lệ hoặc chưa có nguồn truy vết.",
            "Nhập tỷ giá dương và nêu rõ nguồn đã được duyệt (ví dụ FORM!B2).",
        ))
    manual_store_error = _validate_manual_input_store(context)
    if manual_store_error:
        issues.append(SourceIssue(
            "manual_inputs", str(context.manual_input_store or ""), context.fiscal_year,
            manual_store_error,
            f"Dùng kho raw/FY{context.fiscal_year}/manual_inputs.db với từng dòng mang fiscal_year={context.fiscal_year}.",
        ))

    notify("Đang kiểm tra FY và cấu trúc của từng nguồn chi phí...")
    for category in SOURCE_CATEGORIES:
        paths = sources.get(category, ())
        if not paths:
            issues.append(SourceIssue(
                category,
                "",
                None,
                "Không có nguồn được chọn cho category này",
                "Không cần bổ sung nếu category nằm ngoài phạm vi lần chạy.",
                code="MISSING_SOURCE",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Category này không được tính; các nguồn hợp lệ khác vẫn được chạy.",
            ))
            continue
        if category != "it_simulation" and len(paths) != 1:
            issues.append(SourceIssue(
                category, "; ".join(paths), None,
                "Có nhiều nguồn cùng loại nên không thể chọn an toàn",
                "Chọn rõ một nguồn trong manifest; category này sẽ được bỏ qua ở lần chạy hiện tại.",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Category này không được tính; các nguồn hợp lệ khác vẫn được chạy.",
            ))
        for path in paths:
            # Uniform policy has a stricter sequential pipeline below:
            # existence -> structure -> fiscal evidence -> FORM compatibility.
            # Running generic FY validation here would produce cascading issues.
            if category == "uniform_policy":
                continue
            evidence = inspect_fiscal_year_evidence(path)
            # Fixed-asset snapshot workbooks can legitimately use a prior
            # calendar snapshot label instead of an annual FY marker.
            reason = validate_fiscal_year_evidence(
                evidence,
                context.fiscal_year,
                required=category != "fixed_assets",
            )
            if reason:
                issues.append(SourceIssue(
                    category,
                    path,
                    evidence.resolved_fiscal_year,
                    reason,
                    f"Thay bằng nguồn xác nhận đúng FY{context.fiscal_year}; category này đang được bỏ qua.",
                    severity=ISSUE_SOURCE_SKIPPED,
                    impact="Nguồn này không được tính; các nguồn hợp lệ khác vẫn được chạy.",
                ))

    system_paths = sources.get("it_simulation", ())
    if system_paths:
        rejected_system_paths = {
            issue.path for issue in issues
            if issue.category == "it_simulation" and issue.path in system_paths
        }
        valid_system_paths: list[str] = []
        for path in system_paths:
            if path in rejected_system_paths:
                continue
            try:
                map_system_source_periods(
                    (path,),
                    context.fiscal_year,
                    require_complete=False,
                )
            except SystemSourcePeriodError as exc:
                issues.append(SourceIssue(
                    "it_simulation",
                    path,
                    context.fiscal_year,
                    str(exc),
                    "Đổi tên/chọn file thể hiện rõ khoảng kỳ; chỉ file này bị bỏ qua.",
                    period_coverage=tuple(exc.periods),
                    code=exc.code,
                    severity=ISSUE_SOURCE_SKIPPED,
                    impact="File System Cost này không được dùng; các file hợp lệ khác vẫn được tính.",
                ))
            else:
                valid_system_paths.append(path)

        if valid_system_paths:
            try:
                map_system_source_periods(
                    valid_system_paths,
                    context.fiscal_year,
                    require_complete=False,
                )
            except SystemSourcePeriodError as exc:
                issues.append(SourceIssue(
                    "it_simulation",
                    "; ".join(exc.paths or valid_system_paths),
                    context.fiscal_year,
                    str(exc),
                    "Loại file trùng kỳ hoặc sửa phạm vi kỳ; System Cost sẽ bị bỏ qua để tránh tự chọn sai.",
                    period_coverage=tuple(exc.periods),
                    code=exc.code,
                    severity=ISSUE_SOURCE_SKIPPED,
                    impact="System Cost không được ghi; các category hợp lệ khác vẫn được chạy.",
                ))

    fixed_asset_paths = sources.get("fixed_assets", ())
    if len(fixed_asset_paths) == 1:
        evidence = inspect_fiscal_year_evidence(fixed_asset_paths[0])
        if evidence.inspection_error:
            issues.append(SourceIssue(
                "fixed_assets", fixed_asset_paths[0], evidence.resolved_fiscal_year,
                f"Không đọc được snapshot tài sản cố định: {evidence.inspection_error}",
                "Cung cấp file snapshot mở được; category này đang được bỏ qua.",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Tài sản cố định không được tính; các category hợp lệ khác vẫn được chạy.",
            ))

    allocation_paths = sources.get("allocation_rules", ())
    if len(allocation_paths) == 1 and not _allocation_has_exact_sheet(allocation_paths[0], context.fiscal_year):
        issues.append(SourceIssue(
            "allocation_rules", allocation_paths[0], detect_fiscal_year(allocation_paths[0]),
            "Không có sheet quy tắc đúng FY", f"Bổ sung sheet FY{context.fiscal_year}; category này đang được bỏ qua.",
            severity=ISSUE_SOURCE_SKIPPED,
            impact="Không nạp quy tắc phân bổ từ file này; các nguồn hợp lệ khác vẫn được chạy.",
        ))

    for category in ("ga", "birthday", "nnn_paperwork"):
        paths = sources.get(category, ())
        if len(paths) == 1 and not _workbook_has_fiscal_sheet(paths[0], context.fiscal_year):
            issues.append(SourceIssue(
                category, paths[0], detect_fiscal_year(paths[0]),
                "Không có sheet xác nhận đúng năm tài chính",
                f"Bổ sung sheet FY{context.fiscal_year}; category này đang được bỏ qua.",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Category này không được tính; các nguồn hợp lệ khác vẫn được chạy.",
            ))

    notify("Đang kiểm tra policy và độ tương thích với FORM...")
    uniform_paths = sources.get("uniform_policy", ())
    uniform_ready = False
    if len(uniform_paths) == 1:
        uniform_path = Path(uniform_paths[0])
        if not uniform_path.is_file():
            issues.append(SourceIssue(
                "uniform_policy",
                str(uniform_path),
                None,
                "File policy đồng phục/cốc xếp không tồn tại",
                "Chọn file policy hợp lệ; category này đang được bỏ qua.",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Đồng phục/cốc xếp không được tính; các category khác vẫn được chạy.",
            ))
        elif not _is_uniform_policy(uniform_path):
            issues.append(SourceIssue(
                "uniform_policy",
                str(uniform_path),
                detect_fiscal_year(uniform_path),
                "Thiếu sheet 原価センタ hoặc các cột policy đồng phục/cốc xếp F:U",
                "Dùng bảng policy đúng cấu trúc; category này đang được bỏ qua.",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Đồng phục/cốc xếp không được tính; các category khác vẫn được chạy.",
            ))
        else:
            evidence = inspect_fiscal_year_evidence(uniform_path)
            fiscal_reason = validate_fiscal_year_evidence(
                evidence,
                context.fiscal_year,
                required=context.fiscal_year != 2027,
            )
            if fiscal_reason:
                issues.append(SourceIssue(
                    "uniform_policy",
                    str(uniform_path),
                    evidence.resolved_fiscal_year,
                    fiscal_reason,
                    f"Thay bằng policy FY{context.fiscal_year}; category này đang được bỏ qua.",
                    severity=ISSUE_SOURCE_SKIPPED,
                    impact="Đồng phục/cốc xếp không được tính; các category khác vẫn được chạy.",
                ))
            else:
                uniform_ready = True

    if uniform_ready:
        master_uniform_error = _validate_form_master_and_uniform(context)
        if master_uniform_error:
            issues.append(SourceIssue(
                "form_uniform_master",
                context.uniform_policy_path or context.template_path,
                detect_fiscal_year(context.uniform_policy_path) if context.uniform_policy_path else None,
                master_uniform_error,
                "Sửa policy hoặc FORM; policy đồng phục sẽ bị bỏ qua cho lần chạy hiện tại.",
                severity=ISSUE_SOURCE_SKIPPED,
                impact="Đồng phục/cốc xếp không được tính; các category khác vẫn được chạy.",
            ))

    notify("Đang tổng hợp kết quả kiểm tra an toàn...")
    return RunPreflightReport(
        context.fiscal_year,
        tuple(issues),
        sources,
        _preflight_checks(context, sources, issues),
    )


def sha256_file(path: str | os.PathLike[str]) -> str:
    digest = sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
