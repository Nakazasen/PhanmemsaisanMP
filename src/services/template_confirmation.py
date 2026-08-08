"""Inspect a user-selected FORM and track whether its exact file was confirmed."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
import hashlib
import os
from pathlib import Path
from typing import Any, Mapping

import openpyxl

from src.utils.excel_helpers import find_form_template_hygiene_issues, find_hub_sheet_name


@dataclass(frozen=True)
class FormInspection:
    """The minimum technical evidence needed before a user can use a FORM."""

    path: str
    checksum: str
    hub_sheet_name: str
    sheet_names: tuple[str, ...]
    issue_cells: tuple[str, ...] = ()
    error: str = ""

    @property
    def is_valid(self) -> bool:
        return not self.error and not self.issue_cells

    def as_confirmation(self) -> dict[str, Any]:
        return {
            **asdict(self),
            "confirmed_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        }


def sha256_file(path: str | Path) -> str:
    """Return a streaming SHA-256 without loading a workbook into memory."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_form(path: str | Path) -> FormInspection:
    """Check that a FORM is readable and structurally usable by the pipeline."""
    candidate = Path(path).resolve()
    if not candidate.is_file():
        return FormInspection(str(candidate), "", "", (), error="missing")
    try:
        checksum = sha256_file(candidate)
        workbook = openpyxl.load_workbook(candidate, read_only=True, data_only=False)
    except Exception:
        return FormInspection(str(candidate), "", "", (), error="unreadable")
    try:
        try:
            hub_sheet_name = find_hub_sheet_name(workbook)
            issue_cells = find_form_template_hygiene_issues(workbook)
        except (KeyError, ValueError):
            return FormInspection(
                str(candidate), checksum, "", tuple(workbook.sheetnames), error="invalid_layout"
            )
        return FormInspection(
            str(candidate), checksum, hub_sheet_name, tuple(workbook.sheetnames), issue_cells
        )
    finally:
        workbook.close()


def confirmation_status(
    confirmations: Mapping[str, Mapping[str, Any]] | None,
    inspection: FormInspection,
) -> str:
    """Classify a usable FORM as already known, new, or changed at its old path."""
    if not inspection.is_valid:
        return "invalid"
    records = confirmations or {}
    if inspection.checksum in records:
        return "known"
    selected_path = os.path.normcase(os.path.abspath(inspection.path))
    for record in records.values():
        previous_path = str(record.get("path", ""))
        if previous_path and os.path.normcase(os.path.abspath(previous_path)) == selected_path:
            return "changed"
    return "new"
