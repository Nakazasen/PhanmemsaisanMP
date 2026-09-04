"""Bộ nạp dữ liệu gốc của MP2027 Manager.

Đọc trung tâm chi phí, tài khoản và quy tắc phân bổ từ các tệp Excel nguồn.
"""
import csv
import sqlite3
import os
import re
import unicodedata
import pandas as pd
import openpyxl
from src.db.schema import get_connection, create_schema, init_sys_params
from src.utils.excel_helpers import normalize_cc_code, read_exchange_rate_from_form, validate_exchange_rate
from src.utils.source_manifest import resolve_manifest_file
from src.services.fiscal_run import resolve_uniform_policy_path

import sys

from src.engine.uniform_cup_rules import SOURCE_BACKED_UNIFORM_ITEM_SPECS, normalize_uniform_text

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _packaged_base_dir() -> str:
    return getattr(sys, "_MEIPASS", BASE_DIR)


def _resolve_mp2027_docs_dir() -> str:
    external_docs_dir = os.path.join(BASE_DIR, 'docs', 'MP2027')
    if os.path.isdir(external_docs_dir):
        return external_docs_dir

    packaged_docs_dir = os.path.join(_packaged_base_dir(), 'docs', 'MP2027')
    if os.path.isdir(packaged_docs_dir):
        return packaged_docs_dir

    raise FileNotFoundError(
        f"Không tìm thấy thư mục tài liệu nguồn bắt buộc: {external_docs_dir}. "
        "Yêu cầu có thư mục docs/MP2027 ở cạnh ứng dụng hoặc được đóng gói bên trong tệp thực thi."
    )


# Legacy defaults are resolved lazily.  Importing a shared loader must not
# require FY2027 to exist when a future FY is supplied explicitly.
MP2027_DOCS_DIR = os.path.join(BASE_DIR, "docs", "MP2027")
FORM_PATH = os.path.join(MP2027_DOCS_DIR, 'FORM.xlsx')
ALLOC_PATH = os.path.join(MP2027_DOCS_DIR, 'FY2027配賦額一覧 (2025.12.29).xlsx')
UNIFORM_REQUIREMENTS_FILENAME = 'Cải tiến nhập dữ liệu chung vào file MPnew 10.07.2026.xlsx'
UNIFORM_REQUIREMENTS_SHEET = '原価センタ'


def resolve_uniform_requirements_path(
    explicit_path: str | None = None,
    fiscal_year: int = 2027,
) -> str:
    if explicit_path:
        path = os.path.abspath(explicit_path)
        if os.path.isfile(path):
            return path
        candidate = os.path.join(BASE_DIR, "raw", os.path.basename(path))
        if os.path.isfile(candidate):
            return candidate
        raise FileNotFoundError(f"Không tìm thấy file yêu cầu đồng phục/cốc xếp: {path}")

    annual = resolve_uniform_policy_path(fiscal_year, base_dir=BASE_DIR)
    if annual and os.path.isfile(annual):
        return annual

    # Only FY2027 can use the packaged legacy policy workbook.  Future fiscal
    # years must provide a policy inside raw/FY<year>.
    candidates = ()
    if int(fiscal_year) == 2027:
        candidates = (os.path.join(_packaged_base_dir(), "raw", UNIFORM_REQUIREMENTS_FILENAME),)
    for path in candidates:
        if os.path.isfile(path):
            return path
    raise FileNotFoundError(
        "Không tìm thấy nguồn đối tượng đồng phục/cốc xếp chính thức: "
        + "; ".join(candidates)
    )


def load_uniform_entitlements(
    conn: sqlite3.Connection,
    requirements_path: str | None = None,
    fiscal_year: int = 2027,
) -> int:
    """Load all F:W entitlement decisions with cell-level provenance."""
    path = resolve_uniform_requirements_path(requirements_path, fiscal_year=fiscal_year)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if UNIFORM_REQUIREMENTS_SHEET not in workbook.sheetnames:
            raise ValueError(f"Thiếu sheet {UNIFORM_REQUIREMENTS_SHEET} trong {path}")
        worksheet = workbook[UNIFORM_REQUIREMENTS_SHEET]
        headers: dict[str, int] = {}
        duplicates: set[str] = set()
        for column in range(1, worksheet.max_column + 1):
            normalized = normalize_uniform_text(worksheet.cell(1, column).value)
            if not normalized:
                continue
            if normalized in headers:
                duplicates.add(normalized)
            headers[normalized] = column
        if duplicates:
            raise ValueError(f"Cột bị trùng trong sheet {UNIFORM_REQUIREMENTS_SHEET}: {sorted(duplicates)}")

        cc_header = normalize_uniform_text("原価センタ")
        if cc_header not in headers:
            raise ValueError(f"Thiếu cột mã phòng trong sheet {UNIFORM_REQUIREMENTS_SHEET}")
        missing = [
            spec.header
            for spec in SOURCE_BACKED_UNIFORM_ITEM_SPECS
            if normalize_uniform_text(spec.header) not in headers
        ]
        if missing:
            raise ValueError("Thiếu cột đối tượng đồng phục/cốc xếp: " + ", ".join(missing))

        rows: list[tuple[str, str, str, int, str, str, str]] = []
        seen_cc: set[str] = set()
        for row_number in range(2, worksheet.max_row + 1):
            cc_code = normalize_cc_code(worksheet.cell(row_number, headers[cc_header]).value)
            if not cc_code:
                continue
            if cc_code in seen_cc:
                raise ValueError(f"Mã phòng bị trùng trong sheet {UNIFORM_REQUIREMENTS_SHEET}: {cc_code}")
            seen_cc.add(cc_code)
            for spec in SOURCE_BACKED_UNIFORM_ITEM_SPECS:
                column = headers[normalize_uniform_text(spec.header)]
                raw_mark = str(worksheet.cell(row_number, column).value or "").strip()
                if raw_mark not in ("", "〇"):
                    raise ValueError(
                        f"Dấu chọn không hợp lệ tại {UNIFORM_REQUIREMENTS_SHEET}!"
                        f"{worksheet.cell(row_number, column).coordinate}: {raw_mark!r}"
                    )
                rows.append(
                    (
                        cc_code,
                        spec.key,
                        spec.header,
                        1 if raw_mark == "〇" else 0,
                        os.path.abspath(path),
                        UNIFORM_REQUIREMENTS_SHEET,
                            f"{openpyxl.utils.get_column_letter(column)}{row_number}",
                    )
                )

        known_cc = {str(row[0]).strip() for row in conn.execute("SELECT code FROM dim_cost_centers")}
        source_cc = {row[0] for row in rows}
        missing_cc = sorted(known_cc - source_cc)
        unknown_cc = sorted(source_cc - known_cc)
        if missing_cc or unknown_cc:
            raise ValueError(
                "Danh sách mã phòng của nguồn đồng phục không khớp danh mục hiện hành. "
                f"Thiếu={missing_cc}; không nhận diện={unknown_cc}"
            )

        cursor = conn.cursor()
        cursor.execute("DELETE FROM map_cost_center_uniform_items")
        cursor.executemany(
            """
            INSERT INTO map_cost_center_uniform_items
            (cc_code, item_key, item_name, eligible, source_file, source_sheet, source_cell)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
        return len(rows)
    finally:
        workbook.close()


def _normalize_text(value) -> str:
    text = str(value or "").replace("\n", " ").replace("\u3000", " ").strip().lower()
    return " ".join(text.split())


def _looks_like_allocation_rules_workbook(path: str) -> bool:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        if worksheet.max_row < 20 or worksheet.max_column < 10:
            return False

        header_cells: list[str] = []
        for row in worksheet.iter_rows(min_row=1, max_row=6, max_col=10, values_only=True):
            for value in row:
                if value is not None:
                    header_cells.append(_normalize_text(value))
        header_blob = " | ".join(header_cells)
        return (
            "vnd" in header_blob
            and ("don gia" in header_blob or "単価" in header_blob)
            and ("ma tai khoan" in header_blob or "tai khoan" in header_blob or "計上月" in header_blob)
        )
    finally:
        workbook.close()


def find_allocation_rules_file(search_dir: str | None = None, fiscal_year: int = 2027) -> str | None:
    manifest_path = resolve_manifest_file(search_dir, "allocation_rules")
    if manifest_path:
        return manifest_path

    candidates: list[tuple[int, str]] = []
    base_search_dir = search_dir or BASE_DIR
    if not os.path.isdir(base_search_dir):
        return None

    for name in os.listdir(base_search_dir):
        lower_name = name.lower()
        if not lower_name.endswith(".xlsx") or lower_name.startswith("~$"):
            continue
        path = os.path.join(base_search_dir, name)
        if not _looks_like_allocation_rules_workbook(path):
            continue

        score = 0
        if f"fy{fiscal_year}".lower() in lower_name:
            score += 3
        if "配賦" in name or "allocation" in lower_name:
            score += 2
        if "2025.12.29" in name:
            score += 1
        candidates.append((score, path))

    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], os.path.basename(item[1]).lower()))
    return candidates[0][1]


def _classify_driver(raw_text: str) -> str:
    """Classify allocation driver from raw Japanese/Vietnamese text."""
    if not raw_text or pd.isna(raw_text):
        return 'unknown'
    text = str(raw_text).strip()
    if 'G7社員' in text or '公nhân' in text.lower() or 'công nhân' in text.lower():
        return 'headcount_worker'
    elif 'スタッフ' in text or 'nhân viên' in text.lower():
        return 'headcount_staff'
    elif '配属人数' in text or '人数' in text or 'số người' in text.lower():
        return 'headcount_all'
    elif '稼働日数' in text or 'ngày' in text.lower():
        return 'working_days'
    elif '固定' in text or 'tỷ lệ' in text.lower():
        return 'fixed_ratio'
    else:
        return 'headcount_all'  # Default fallback


def _parse_unit_price(value) -> float | None:
    """Parse numeric unit prices such as `145$`, `1,259,500`, or plain floats."""
    if pd.isna(value):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if re.fullmatch(r"[※*＊]\s*\d+", text):
        return None

    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        return float(match.group(0))
    except (ValueError, TypeError):
        return None


def _is_footnote_unit_price(value) -> bool:
    if pd.isna(value):
        return False
    return re.fullmatch(r"[※*＊]\s*\d+", str(value).strip()) is not None


def _apply_approved_unit_price_override(item_name: str, unit_price: float, fiscal_year: int) -> float:
    """Apply an explicitly approved, year-scoped override only.

    A blank price in an annual allocation workbook must never inherit a price
    from another fiscal year.  FY2027 compatibility values are data in the
    approved override CSV, not a fallback rule for future years.
    """
    if float(unit_price or 0.0) > 0:
        return float(unit_price)
    path = os.path.join(BASE_DIR, "docs", "config", "approved_unit_price_overrides.csv")
    if not os.path.isfile(path):
        return float(unit_price or 0.0)
    normalized_item = _normalize_text(item_name)
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if str(row.get("fiscal_year", "")).strip() != str(int(fiscal_year)):
                continue
            token = _normalize_text(row.get("item_token", ""))
            if not token or token not in normalized_item:
                continue
            try:
                return float(str(row.get("approved_unit_price", "")).replace(",", ""))
            except (TypeError, ValueError):
                continue
    return float(unit_price or 0.0)


def load_cost_centers(conn: sqlite3.Connection, form_path: str = None) -> int:
    """Load cost centers from FORM.xlsx 原価センタ sheet."""
    path = form_path or FORM_PATH
    if not os.path.exists(path):
        print(f"Không tìm thấy FORM.xlsx tại {path}")
        return 0

    xl = pd.ExcelFile(path, engine='openpyxl')
    # Find the cost center sheet (原価センタ)
    cc_sheet = None
    for name in xl.sheet_names:
        if '原価' in name or 'センタ' in name or 'cost' in name.lower():
            cc_sheet = name
            break

    if not cc_sheet:
        # Fall back to known index from extract_samples.py (index 5)
        if len(xl.sheet_names) > 5:
            cc_sheet = xl.sheet_names[5]
            print(f"Thông tin: đang dùng trang tính theo vị trí: {cc_sheet}")
        else:
            print("Cảnh báo: không tìm thấy trang tính mã bộ phận")
            return 0

    print(f"Đang đọc mã bộ phận từ trang tính: {cc_sheet}")
    df = pd.read_excel(path, sheet_name=cc_sheet, engine='openpyxl')

    cursor = conn.cursor()
    # PRE-SYNC: Clear existing cost centers
    cursor.execute("DELETE FROM dim_cost_centers")
    count = 0
    for _, row in df.iterrows():
        raw_code = row.iloc[0]  # Unnamed: 0 = code
        if pd.isna(raw_code):
            continue
        code = normalize_cc_code(raw_code)
        if not code:
            continue

        name_jp = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ''
        seq_no = float(row.iloc[2]) if not pd.isna(row.iloc[2]) else None
        saisan = str(row.iloc[3]).strip() if len(row) > 3 and not pd.isna(row.iloc[3]) else ''
        cost_type = str(row.iloc[4]).strip() if len(row) > 4 and not pd.isna(row.iloc[4]) else ''

        if not name_jp:
            continue

        cursor.execute("""
            INSERT OR REPLACE INTO dim_cost_centers
            (code, name_jp, seq_no, saisan_type, cost_type)
            VALUES (?, ?, ?, ?, ?)
        """, (code, name_jp, seq_no, saisan, cost_type))
        count += 1

    conn.commit()
    print(f"Đã nạp {count} mã bộ phận.")
    return count


def load_accounts(conn: sqlite3.Connection, form_path: str = None) -> int:
    """Load accounts from FORM.xlsx 勘定科目 sheet."""
    path = form_path or FORM_PATH
    if not os.path.exists(path):
        return 0

    xl = pd.ExcelFile(path, engine='openpyxl')
    # Find the account sheet (勘定科目)
    acc_sheet = None
    for name in xl.sheet_names:
        if '勘定' in name or '科目' in name:
            acc_sheet = name
            break

    if not acc_sheet:
        if len(xl.sheet_names) > 4:
            acc_sheet = xl.sheet_names[4]
            print(f" Đang dùng trang tính theo vị trí: {acc_sheet}")
        else:
            print("Cảnh báo: không tìm thấy trang tính tài khoản")
            return 0

    print(f" Đang đọc tài khoản từ trang tính: {acc_sheet}")
    df = pd.read_excel(path, sheet_name=acc_sheet, engine='openpyxl')

    cursor = conn.cursor()
    # PRE-SYNC: Clear existing accounts
    cursor.execute("DELETE FROM dim_accounts")
    count = 0
    for _, row in df.iterrows():
        code = row.get('Account_Code', row.iloc[0] if len(row) > 0 else None)
        if pd.isna(code):
            continue
        try:
            code = int(float(code))
        except (ValueError, TypeError):
            continue

        name_jp = str(row.iloc[1]).strip() if len(row) > 1 and not pd.isna(row.iloc[1]) else ''
        name_vn = str(row.iloc[2]).strip() if len(row) > 2 and not pd.isna(row.iloc[2]) else None
        group_name = str(row.iloc[3]).strip() if len(row) > 3 and not pd.isna(row.iloc[3]) else None
        group_vn = str(row.iloc[4]).strip() if len(row) > 4 and not pd.isna(row.iloc[4]) else None

        # 製造/一般/販売 codes (columns 5, 6, 7)
        mfg_code = None
        ga_code = None
        sales_code = None
        if len(row) > 5 and not pd.isna(row.iloc[5]):
            try:
                mfg_code = int(float(row.iloc[5]))
            except (ValueError, TypeError):
                pass
        if len(row) > 6 and not pd.isna(row.iloc[6]):
            try:
                ga_code = int(float(row.iloc[6]))
            except (ValueError, TypeError):
                pass
        if len(row) > 7 and not pd.isna(row.iloc[7]):
            try:
                sales_code = int(float(row.iloc[7]))
            except (ValueError, TypeError):
                pass

        remark = str(row.iloc[8]).strip() if len(row) > 8 and not pd.isna(row.iloc[8]) else None

        if not name_jp:
            continue

        cursor.execute("""
            INSERT OR REPLACE INTO dim_accounts
            (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (code, name_jp, name_vn, group_name, group_vn, mfg_code, ga_code, sales_code, remark))
        count += 1

    conn.commit()
    print(f"Đã nạp {count} tài khoản.")
    return count

def _select_allocation_rules_sheet(sheet_names: list[str], fiscal_year: int) -> str:
    """Select the fiscal-year allocation sheet instead of blindly using the first sheet."""
    fiscal_token = f"FY{fiscal_year}"
    for sheet_name in sheet_names:
        if fiscal_token in sheet_name:
            return sheet_name

    for sheet_name in sheet_names:
        if "配賦額一覧" in sheet_name or "allocation" in sheet_name.lower():
            return sheet_name

    return sheet_names[0]


def _merge_verified_content_rules(
    conn: sqlite3.Connection,
    content_rules: list[dict] | None,
) -> int:
    """Insert verified data-only rules without overriding workbook identities."""
    if not content_rules:
        return 0
    from src.services.content_packs import CONTENT_SCHEMA, validate_rules

    rules = validate_rules({"schema": CONTENT_SCHEMA, "rules": content_rules})
    existing = {
        (str(row[0]).strip().casefold(), str(row[1]).strip().casefold())
        for row in conn.execute("SELECT source_dept, item_name FROM map_allocation_rules")
    }
    conflicts = [
        f"{rule['source_dept']}/{rule['item_name']}"
        for rule in rules
        if (str(rule["source_dept"]).strip().casefold(), str(rule["item_name"]).strip().casefold()) in existing
    ]
    if conflicts:
        raise ValueError(
            "Quy tắc content pack trùng với workbook; không quy tắc nào được ghi: "
            + ", ".join(conflicts)
        )
    conn.executemany(
        """
        INSERT INTO map_allocation_rules
        (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
         posting_month, unit_price, unit, driver_type, driver_raw)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                str(rule["source_dept"]).strip(),
                str(rule["item_name"]).strip(),
                rule.get("account_name"),
                rule.get("mfg_account"),
                rule.get("ga_account"),
                rule.get("sales_account"),
                rule.get("posting_month"),
                float(rule["unit_price"]),
                rule.get("unit"),
                rule["driver_type"],
                rule.get("driver_raw"),
            )
            for rule in rules
        ],
    )
    return len(rules)


def load_allocation_rules(
    conn: sqlite3.Connection,
    alloc_path: str = None,
    search_dir: str | None = None,
    fiscal_year: int = 2027,
    content_rules: list[dict] | None = None,
) -> int:
    """Load workbook rules and append already verified content-pack rules."""
    path = alloc_path or ALLOC_PATH
    if fiscal_year != 2027 and not alloc_path:
        path = os.path.join(search_dir or BASE_DIR, f"FY{fiscal_year}配賦額一覧.xlsx")
    if not path or not os.path.exists(path):
        discovered = find_allocation_rules_file(search_dir=search_dir, fiscal_year=fiscal_year)
        if discovered:
            path = discovered
    if not path or not os.path.exists(path):
        if not content_rules:
            print(f"Cảnh báo: không tìm thấy tệp quy tắc phân bổ tại {path}")
            return 0
        print("Không có workbook quy tắc; đang nạp content pack đã xác thực.")
        df = pd.DataFrame()
    else:
        print(f"Đang đọc quy tắc phân bổ từ: {os.path.basename(path)}")
        xl = pd.ExcelFile(path, engine='openpyxl')
        target_sheet = _select_allocation_rules_sheet(xl.sheet_names, fiscal_year)
        if f"FY{fiscal_year}" not in target_sheet.upper():
            raise ValueError(
                f"File quy tắc {path} không có sheet FY{fiscal_year}; không được dùng sheet của năm khác."
            )
        print(f"Đang đọc quy tắc phân bổ từ trang tính: {target_sheet}")
        df = pd.read_excel(path, sheet_name=target_sheet, engine='openpyxl')

    cursor = conn.cursor()
    workbook_rows: list[tuple] = []
    current_dept = None
    current_item = None
    current_account_name = None
    current_mfg_acc = None
    current_ga_acc = None
    current_sales_acc = None
    current_posting_month = None

    for _, row in df.iterrows():
        # Column mapping from master data:
        # 0: 配布元 (source dept) - may be NaN for continuation rows
        # 1: 内容 (item name)
        # 2: 科目名称 (account name)
        # 3: 製造コード
        # 4: 間接コード
        # 5: 販売コード
        # 6: 計上月 (posting month)
        # 7: 単価 (unit price)
        # 8: 単位 (unit)
        # 9: 計上基準 (driver/criteria)

        dept = row.iloc[0] if not pd.isna(row.iloc[0]) else current_dept
        raw_item = row.iloc[1] if len(row) > 1 and not pd.isna(row.iloc[1]) else None

        # Skip header rows
        item_str = str(raw_item or current_item or "")
        if not item_str:
            continue
        if '内　容' in item_str or 'Nội dung' in item_str:
            continue

        if not pd.isna(row.iloc[0]):
            current_dept = str(row.iloc[0]).strip()

        if not current_dept:
            continue

        # Keep footnote-only rows as metadata with unit_price=0 so downstream
        # allocation can fail closed and report the missing unit price.
        raw_unit_price = row.iloc[7] if len(row) > 7 else None
        raw_unit = row.iloc[8] if len(row) > 8 else None
        raw_driver = row.iloc[9] if len(row) > 9 else None
        if raw_item is None and pd.isna(raw_unit) and pd.isna(raw_driver):
            continue
        unit_price = _parse_unit_price(raw_unit_price)
        if unit_price is None:
            if _is_footnote_unit_price(raw_unit_price):
                unit_price = 0.0
            else:
                continue
        unit_price = _apply_approved_unit_price_override(item_str, unit_price, fiscal_year)

        def _safe_int(val):
            """Convert value to int, handling '-', empty strings, etc."""
            if pd.isna(val):
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        if raw_item is not None:
            current_item = item_str.strip()
            current_account_name = str(row.iloc[2]).strip() if len(row) > 2 and not pd.isna(row.iloc[2]) else None
            current_mfg_acc = _safe_int(row.iloc[3]) if len(row) > 3 else None
            current_ga_acc = _safe_int(row.iloc[4]) if len(row) > 4 else None
            current_sales_acc = _safe_int(row.iloc[5]) if len(row) > 5 else None
            current_posting_month = str(row.iloc[6]).strip() if len(row) > 6 and not pd.isna(row.iloc[6]) else None

        account_name = (
            str(row.iloc[2]).strip()
            if len(row) > 2 and not pd.isna(row.iloc[2])
            else current_account_name
        )
        mfg_acc = _safe_int(row.iloc[3]) if len(row) > 3 and not pd.isna(row.iloc[3]) else current_mfg_acc
        ga_acc = _safe_int(row.iloc[4]) if len(row) > 4 and not pd.isna(row.iloc[4]) else current_ga_acc
        sales_acc = _safe_int(row.iloc[5]) if len(row) > 5 and not pd.isna(row.iloc[5]) else current_sales_acc
        posting_month = (
            str(row.iloc[6]).strip()
            if len(row) > 6 and not pd.isna(row.iloc[6])
            else current_posting_month
        )
        unit = str(raw_unit).strip() if len(row) > 8 and not pd.isna(raw_unit) else None
        driver_raw = str(raw_driver).strip() if len(row) > 9 and not pd.isna(raw_driver) else None

        driver_type = _classify_driver(driver_raw)
        normalized_item = unicodedata.normalize("NFKD", item_str).lower()
        normalized_item = "".join(ch for ch in normalized_item if not unicodedata.combining(ch))
        recurring_total_tokens = (
            "食堂燃料", "gas", "トイレットペーパー", "giay ve sinh", "toilet paper",
            "手洗い洗剤", "nuoc rua tay", "hand wash", "清掃費", "phi lam sach", "cleaning",
        )
        if any(token in normalized_item for token in recurring_total_tokens):
            driver_type = "headcount_all"

        workbook_rows.append(
            (
                current_dept,
                item_str.strip(),
                account_name,
                mfg_acc,
                ga_acc,
                sales_acc,
                posting_month,
                unit_price,
                unit,
                driver_type,
                driver_raw,
            )
        )

    savepoint = "allocation_rules_rebuild"
    cursor.execute(f"SAVEPOINT {savepoint}")
    savepoint_active = True
    try:
        cursor.execute("DELETE FROM map_allocation_rules")
        cursor.executemany(
            """
            INSERT INTO map_allocation_rules
            (source_dept, item_name, account_name, mfg_account, ga_account, sales_account,
             posting_month, unit_price, unit, driver_type, driver_raw)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            workbook_rows,
        )
        count = len(workbook_rows) + _merge_verified_content_rules(conn, content_rules)
        cursor.execute(f"RELEASE SAVEPOINT {savepoint}")
        savepoint_active = False
        conn.commit()
    except Exception:
        if savepoint_active:
            # Cleanup must never replace the original validation/database error.
            try:
                cursor.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
            except Exception:
                pass
            try:
                cursor.execute(f"RELEASE SAVEPOINT {savepoint}")
            except Exception:
                pass
        else:
            # RELEASE succeeded, so a later commit failure has no savepoint to use.
            try:
                conn.rollback()
            except Exception:
                pass
        raise
    print(f"Đã nạp {count} quy tắc phân bổ.")
    return count


def load_all(db_path: str = None, template_path: str = None,
             rules_path: str = None, fiscal_year: int = 2027,
             exchange_rate: float | None = None, search_dir: str | None = None,
             exchange_rate_source: str = "explicit pipeline input",
             uniform_eligibility_path: str | None = None,
             include_allocation_rules: bool = True,
             include_uniform_entitlements: bool = True,
             content_rules: list[dict] | None = None) -> dict:
    """Load shared master data and only the explicitly enabled optional sources."""
    # Future FY must name its own FORM; no implicit FY2027 fallback is allowed.
    if template_path:
        t_path = template_path
    elif int(fiscal_year) == 2027:
        t_path = FORM_PATH
    else:
        t_path = os.path.join(BASE_DIR, "docs", f"MP{int(fiscal_year)}", "FORM.xlsx")
    if not os.path.isfile(t_path):
        raise FileNotFoundError(f"Thiếu FORM đúng FY{fiscal_year}: {t_path}")
    discovery_dir = search_dir or (os.path.dirname(os.path.abspath(t_path)) if t_path else BASE_DIR)
    r_path = rules_path or (
        ALLOC_PATH if int(fiscal_year) == 2027 else os.path.join(discovery_dir, f"FY{fiscal_year}配賦額一覧.xlsx")
    )

    if exchange_rate is None:
        exchange_rate = read_exchange_rate_from_form(t_path)
        exchange_rate_source = f"FORM B2 ({os.path.basename(t_path)})"
    else:
        exchange_rate = validate_exchange_rate(exchange_rate)
    print(f"Nguồn tỷ giá hiệu lực: {exchange_rate_source}: {exchange_rate:,.0f}")

    conn = get_connection(db_path)
    # Ensure Row factory for Row-based access in loaders if needed (schema.py usually sets this)
    conn.row_factory = sqlite3.Row
    create_schema(conn)

    # Initialize system params with SSOT rate
    init_sys_params(
        conn,
        exchange_rate=exchange_rate,
        fiscal_year=fiscal_year,
        exchange_rate_source=exchange_rate_source,
    )

    # Do not discover an allocation workbook when preflight excluded the
    # category from this run's approved source scope.
    if include_allocation_rules and not rules_path and not os.path.exists(r_path):
        discovered = find_allocation_rules_file(search_dir=discovery_dir, fiscal_year=fiscal_year)
        if discovered:
            r_path = discovered

    results = {
        'cost_centers': load_cost_centers(conn, t_path),
        'accounts': load_accounts(conn, t_path),
        'allocation_rules': (
            load_allocation_rules(
                conn,
                r_path,
                search_dir=discovery_dir,
                fiscal_year=fiscal_year,
                content_rules=content_rules,
            )
            if include_allocation_rules or content_rules else 0
        ),
    }
    results['uniform_entitlements'] = (
        load_uniform_entitlements(
            conn,
            uniform_eligibility_path,
            fiscal_year=fiscal_year,
        )
        if include_uniform_entitlements else 0
    )

    conn.close()
    return results


if __name__ == '__main__':
    results = load_all()
    print(f"\nTóm tắt: {results}")
