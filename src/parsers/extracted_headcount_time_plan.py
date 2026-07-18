"""Parser for extracted staffing truth workbooks in the company Master Plan form."""
from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.parsers.headcount_time_plan import PlanParseResult
from src.utils.excel_helpers import get_fy_months, normalize_cc_code

EXTRACTED_SOURCE = "extracted_department_plan"
OUTPUT_SHEET = "人員・時間計画"
INDEX_SHEET = "Sheet1"
_FORM_TITLE = "FY"


class ExtractedPlanParseError(ValueError):
    """Raised when an extracted file does not keep the required company form."""


def _number(value: Any, label: str, *, allow_blank: bool = False) -> float | None:
    if value in (None, ""):
        return None if allow_blank else 0.0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ExtractedPlanParseError(f"{label} không phải số")
    numeric = float(value)
    if numeric < 0:
        raise ExtractedPlanParseError(f"{label} không được âm")
    return numeric


def _cc_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def _require_company_layout(workbook, worksheet) -> None:
    if workbook.sheetnames != [OUTPUT_SHEET, INDEX_SHEET]:
        raise ExtractedPlanParseError(
            "Workbook nguồn sự thật phải chỉ có hai sheet theo form công ty: "
            f"{OUTPUT_SHEET}, {INDEX_SHEET}"
        )
    if worksheet.max_row < 28 or worksheet.max_column < 14:
        raise ExtractedPlanParseError(
            f"Cấu trúc form không đủ A1:N28: {worksheet.max_row} dòng, {worksheet.max_column} cột"
        )
    if _FORM_TITLE not in str(worksheet.cell(1, 1).value or ""):
        raise ExtractedPlanParseError("Không nhận diện được tiêu đề form Master Plan tại A1")


def _status_and_split(
    period: str,
    local_total: float,
    staff: float | None,
    worker: float | None,
    fixed_hours_local: float,
    fixed_hours_staff: float | None,
    fixed_hours_worker: float | None,
    overtime_hours_local: float,
    overtime_hours_staff: float | None,
    overtime_hours_worker: float | None,
) -> tuple[str, float | None, float | None]:
    """Fail closed when any non-zero local metric lacks its staff/worker split."""
    split_groups = (
        (local_total, staff, worker),
        (fixed_hours_local, fixed_hours_staff, fixed_hours_worker),
        (overtime_hours_local, overtime_hours_staff, overtime_hours_worker),
    )
    if any(total != 0 and detail_staff is None and detail_worker is None
           for total, detail_staff, detail_worker in split_groups):
        return "SPLIT_REQUIRED", None, None

    normalized_staff = 0.0 if staff is None else staff
    normalized_worker = 0.0 if worker is None else worker
    if abs(local_total - normalized_staff - normalized_worker) > 0.01:
        raise ExtractedPlanParseError(
            f"{period}: staff + worker không khớp local ({normalized_staff + normalized_worker:g} != {local_total:g})"
        )
    return "READY", normalized_staff, normalized_worker


def parse_extracted_headcount_time_plan(path: str, fiscal_year: int) -> PlanParseResult:
    """Parse a company-form `.xlsx` truth file without writing to the database."""
    resolved_path = str(Path(path).resolve())
    result = PlanParseResult(path=resolved_path, status="error", fiscal_year=fiscal_year)
    matched = re.match(r"\s*(\d+)\.", os.path.basename(path))
    result.department_no = matched.group(1) if matched else ""
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if OUTPUT_SHEET not in workbook.sheetnames:
            raise ExtractedPlanParseError(f"Thiếu sheet {OUTPUT_SHEET}")
        worksheet = workbook[OUTPUT_SHEET]
        _require_company_layout(workbook, worksheet)
        result.sheet_name = OUTPUT_SHEET
        raw_cc = _cc_text(worksheet.cell(5, 1).value)
        result.cc_code = normalize_cc_code(raw_cc) or raw_cc
        result.department_name = str(worksheet.cell(5, 2).value or "").strip()
        if not result.cc_code or not result.department_name:
            raise ExtractedPlanParseError("Thiếu mã bộ phận hoặc tên phòng tại hàng 5")

        periods = get_fy_months(fiscal_year)
        actual_months = [_number(worksheet.cell(8, column).value, f"tháng cột {column}") for column in range(3, 15)]
        expected_months = [float(int(period[-2:])) for period in periods]
        if actual_months != expected_months:
            raise ExtractedPlanParseError(f"Thứ tự tháng không đúng FY{fiscal_year}: {actual_months}")

        rows = {
            "headcount_expat": 10,
            "headcount_staff": 11,
            "headcount_worker": 12,
            "headcount_local_total": 13,
            "headcount_total": 14,
            "fixed_hours_expat": 17,
            "fixed_hours_staff": 18,
            "fixed_hours_worker": 19,
            "fixed_hours_local": 20,
            "fixed_hours_total": 21,
            "overtime_hours_expat": 24,
            "overtime_hours_staff": 25,
            "overtime_hours_worker": 26,
            "overtime_hours_local": 27,
            "overtime_hours_total": 28,
        }
        for index, period in enumerate(periods, start=3):
            values = {
                metric: _number(
                    worksheet.cell(row_number, index).value,
                    f"{period} {metric}",
                    allow_blank=metric.endswith(("staff", "worker")),
                )
                for metric, row_number in rows.items()
            }
            status, staff, worker = _status_and_split(
                period,
                float(values["headcount_local_total"] or 0),
                values["headcount_staff"],
                values["headcount_worker"],
                float(values["fixed_hours_local"] or 0),
                values["fixed_hours_staff"],
                values["fixed_hours_worker"],
                float(values["overtime_hours_local"] or 0),
                values["overtime_hours_staff"],
                values["overtime_hours_worker"],
            )
            checks = (
                (values["headcount_total"], values["headcount_expat"] + values["headcount_local_total"], "tổng số người"),
                (values["fixed_hours_total"], values["fixed_hours_expat"] + values["fixed_hours_local"], "tổng giờ cố định"),
                (values["overtime_hours_total"], values["overtime_hours_expat"] + values["overtime_hours_local"], "tổng tăng ca"),
            )
            for actual, expected, label in checks:
                if abs(float(actual or 0) - float(expected or 0)) > 0.01:
                    raise ExtractedPlanParseError(
                        f"{period}: {label} không khớp ({float(actual or 0):g} != {float(expected or 0):g})"
                    )
            if status == "READY":
                detailed_checks = (
                    (values["fixed_hours_local"], (values["fixed_hours_staff"] or 0) + (values["fixed_hours_worker"] or 0), "tổng giờ cố định local"),
                    (values["overtime_hours_local"], (values["overtime_hours_staff"] or 0) + (values["overtime_hours_worker"] or 0), "tổng tăng ca local"),
                )
                for actual, expected, label in detailed_checks:
                    if abs(float(actual or 0) - float(expected or 0)) > 0.01:
                        raise ExtractedPlanParseError(
                            f"{period}: {label} không khớp ({float(actual or 0):g} != {float(expected or 0):g})"
                        )

            result.rows.append(
                {
                    "period": period,
                    "headcount_expat": float(values["headcount_expat"] or 0),
                    "headcount_local_total": float(values["headcount_local_total"] or 0),
                    "headcount_staff": staff,
                    "headcount_worker": worker,
                    "split_status": status,
                    "fixed_hours_expat": float(values["fixed_hours_expat"] or 0),
                    "fixed_hours_staff": values["fixed_hours_staff"],
                    "fixed_hours_worker": values["fixed_hours_worker"],
                    "fixed_hours_local": float(values["fixed_hours_local"] or 0),
                    "overtime_hours_expat": float(values["overtime_hours_expat"] or 0),
                    "overtime_hours_staff": values["overtime_hours_staff"],
                    "overtime_hours_worker": values["overtime_hours_worker"],
                    "overtime_hours_local": float(values["overtime_hours_local"] or 0),
                    "source_cells": "Master Plan C:N; rows 10-28",
                }
            )
        result.status = "valid"
    except Exception as exc:
        result.errors.append(str(exc))
    finally:
        if workbook is not None:
            workbook.close()
    return result
