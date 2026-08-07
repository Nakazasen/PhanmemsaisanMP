"""Bộ đọc tài sản cố định của MP2027 Manager.

Xử lý lịch khấu hao và lãi tài sản theo quy tắc chốt cuối tháng.
"""
from __future__ import annotations

import sqlite3
from hashlib import sha256
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from src.utils import excel_helpers as helpers
from src.utils.source_manifest import resolve_manifest_file


HEADER_ALIASES = {
    "category": ("category", "asset category", "asset class", "資産クラス", "資産分類"),
    "asset_no": ("asset no", "asset_no", "資産番号", "固定資産番号", "no."),
    "asset_text": ("asset text", "asset_text", "資産テキスト", "資産名", "名称"),
    "control_cc": ("control cost center", "management cost center", "管理原価センタ"),
    "depreciation_cc": (
        "code phòng chịu chi phí", "depreciation cost center", "depr cost center",
        "khấu hao cost center", "償却原価センタ", "費用負担原価センタ",
    ),
    "monthly_depr": ("chi phí khấu hao", "monthly depr", "monthly depreciation", "減価償却費", "償却費"),
    "last_month": ("tháng khấu hao cuối cùng", "last month", "last depreciation month", "償却終了", "最終償却", "最終月"),
    "last_month_depr": ("chi phí khấu hao của tháng cuối cùng", "last month depr", "last month depreciation", "最終月償却", "最終償却額"),
    "apr_interest": ("chi phí lãi tháng 4", "apr interest", "april interest", "4月利息", "4月金利", "4月"),
    "may_interest": ("chi phí lãi", "may interest", "interest from may", "5月以降利息", "5月金利", "5月以降"),
}

LEGACY_COLUMN_MAP = {
    "category": 1,
    "asset_no": 2,
    "asset_text": 3,
    "control_cc": 7,
    "depreciation_cc": 9,
    "monthly_depr": 11,
    "last_month": 15,
    "last_month_depr": 16,
    "apr_interest": 21,
    "may_interest": 22,
}

CATEGORY_SPECS = {
    "machinery_equipment": {
        "aliases": ("mfg)machinery and equipment", "machinery and equipment"),
        "depreciation_account": 5006016242,
        "label": "Machinery and Equipment",
    },
    "vehicles": {
        "aliases": ("mfg)vehicles", "vehicles"),
        "depreciation_account": 5006016243,
        "label": "Vehicles",
    },
    "tools_furniture_fixtures": {
        "aliases": ("mfg)tools furniture and fixtures", "tools furniture and fixtures"),
        "depreciation_account": 5006016244,
        "label": "Tools Furniture and Fixtures",
    },
    "other_tangible_fixed_assets": {
        "aliases": ("mfg)other tangible fixed assets", "other tangible fixed assets"),
        "depreciation_account": 5006016247,
        "label": "Other Tangible Fixed Assets",
    },
    "mold": {
        "aliases": ("mfg)mold", "mold"),
        "depreciation_account": 5005036246,
        "label": "Mold",
    },
}
INTEREST_ACCOUNT = 9114120007
OUT_OF_SCOPE_CATEGORY_MARKERS = (
    "sga)", "software", "buildings", "structures", "land use rights",
)


def _category_status(value: Any) -> tuple[str | None, str]:
    key = _normalize_category(value)
    if key is not None:
        return key, "supported"
    text = " ".join(_norm(value).split())
    if text and any(marker in text for marker in OUT_OF_SCOPE_CATEGORY_MARKERS):
        return None, "out_of_scope"
    return None, "unknown"



def _norm(value: Any) -> str:
    return helpers.normalize_text(str(value or "")).replace("\u3000", " ")


def _header_score(row: tuple[Any, ...]) -> int:
    text = " | ".join(_norm(cell) for cell in row if cell not in (None, ""))
    score = 0
    for aliases in HEADER_ALIASES.values():
        if any(_norm(alias) in text for alias in aliases):
            score += 1
    return score


def _find_header_row(ws) -> tuple[int | None, dict[str, int]]:
    best_row = None
    best_score = 0
    best_values: tuple[Any, ...] = ()
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        score = _header_score(row)
        if score > best_score:
            best_score = score
            best_row = idx
            best_values = row
    if best_row is None or best_score < 2:
        return None, {}

    mapping: dict[str, int] = {}
    normalized_cells = [_norm(cell) for cell in best_values]
    for field, aliases in HEADER_ALIASES.items():
        for col_idx, cell_text in enumerate(normalized_cells):
            if cell_text and any(_norm(alias) in cell_text for alias in aliases):
                mapping[field] = col_idx
                break
    if "depreciation_cc" not in mapping:
        return None, {}
    return best_row, mapping


def _value(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> Any:
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _asset_tag(asset_no: Any, asset_text: Any, sheet_name: str, row_number: int) -> str:
    no = str(asset_no or "").strip()
    text = str(asset_text or "").strip()
    if no and text:
        base = f"{no}|{text}"
    elif no:
        base = no
    elif text:
        base = text
    else:
        base = "asset"
    # Asset number/text are not guaranteed unique in historical spreadsheets.
    # The source coordinate keeps staging, output provenance and terminal checks
    # one-to-one without using a FORM row as a business identity.
    return f"{base}@{sheet_name}!{row_number}"


def _sheet_plan(ws) -> tuple[int, dict[str, int], str]:
    header_row, mapping = _find_header_row(ws)
    if header_row is not None:
        return header_row + 1, mapping, "header"
    return 5, dict(LEGACY_COLUMN_MAP), "legacy"


def _normalize_category(value: Any) -> str | None:
    text = " ".join(_norm(value).split())
    for key, spec in CATEGORY_SPECS.items():
        if text in {_norm(alias) for alias in spec["aliases"]}:
            return key
    return None


def _sheet_content_score(ws) -> int:
    """Score a sheet by asset-ledger content, never by its visible name."""
    header_row, mapping = _find_header_row(ws)
    if header_row is not None:
        score = 100 + len(mapping)
        start_row = header_row + 1
        active_map = mapping
    else:
        score = 0
        start_row = 5
        active_map = LEGACY_COLUMN_MAP
    for row in ws.iter_rows(min_row=start_row, max_row=min(ws.max_row, start_row + 200), values_only=True):
        category_key, status = _category_status(_value(row, active_map, "category"))
        cc = helpers.extract_cc_code(_value(row, active_map, "depreciation_cc"))
        values = (
            helpers.safe_float(_value(row, active_map, "monthly_depr")),
            helpers.safe_float(_value(row, active_map, "apr_interest")),
            helpers.safe_float(_value(row, active_map, "may_interest")),
        )
        if cc and category_key and status == "supported":
            score += 10
        if any(value != 0 for value in values):
            score += 1
    return score


def _selected_worksheets(wb) -> list[Any]:
    """Select the strongest fixed-assets sheet from headers and row content."""
    scored = [(_sheet_content_score(ws), ws.max_row, ws) for ws in wb.worksheets]
    valid = [item for item in scored if item[0] > 0]
    if not valid:
        return []
    highest_score = max(item[0] for item in valid)
    candidates = [item for item in valid if item[0] == highest_score]
    highest_rows = max(item[1] for item in candidates)
    return [next(item[2] for item in candidates if item[1] == highest_rows)]



def find_fixed_assets_file(source_dir: str = None) -> str | None:
    manifest_path = resolve_manifest_file(source_dir, "fixed_assets")
    if manifest_path:
        return manifest_path

    search_dir = Path(source_dir or Path(__file__).resolve().parents[2])
    candidates: list[tuple[int, Path]] = []
    for path in sorted(search_dir.glob("*.xlsx")):
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                score = max((_sheet_content_score(ws) for ws in workbook.worksheets), default=0)
            finally:
                workbook.close()
            if score > 0:
                candidates.append((score, path))
        except Exception:
            continue
    if candidates:
        highest_score = max(score for score, _ in candidates)
        top = [path for score, path in candidates if score == highest_score]
        if len(top) == 1:
            return str(top[0])
        raise ValueError("Multiple source workbooks contain equally likely fixed-assets ledgers")
    return None


def _source_snapshot(path: Path) -> str:
    """Return a content identity for an import without relying on its filename."""
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _round_vnd(amount_usd: float, exchange_rate: float) -> int:
    """Match the audited Excel ROUND(..., 0) policy for one asset at a time."""
    value = float(amount_usd) * float(exchange_rate)
    return math.floor(value + 0.5) if value >= 0 else math.ceil(value - 0.5)


def expand_depreciation_schedule(
    monthly_depr: float,
    last_month: str | None,
    last_month_depr: float | None,
    fy_months: list[str],
) -> dict[str, float]:
    result = {}
    for period in fy_months:
        if last_month and period > last_month:
            continue
        if last_month and period == last_month:
            if last_month_depr is None:
                raise ValueError("Missing Q/last-month depreciation amount for an in-FY terminal period")
            amount = last_month_depr
        else:
            amount = monthly_depr
        result[period] = amount
    return result


def expand_interest_schedule(apr_interest: float, may_interest: float, last_month: str | None, fy_months: list[str]) -> dict[str, float]:
    result = {}
    for period in fy_months:
        if last_month and period > last_month:
            continue
        amount = apr_interest if period == fy_months[0] else may_interest
        result[period] = amount
    return result


def inspect_fixed_assets_workbook(fa_path: str | Path) -> dict:
    """Return non-sensitive source-row coverage counts by sheet, CC and Category."""
    path = Path(fa_path)
    if not path.is_file():
        return {"source_rows": 0, "by_cc": {}, "by_category": {}, "by_sheet": {}, "skipped_reasons": {"missing_file": 1}}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    by_cc: Counter[str] = Counter()
    by_category: Counter[str] = Counter()
    by_sheet: dict[str, dict[str, Any]] = {}
    skipped: Counter[str] = Counter()
    try:
        selected = _selected_worksheets(wb)
        if not selected:
            return {"source_rows": 0, "by_cc": {}, "by_category": {}, "by_sheet": {}, "skipped_reasons": {"missing_source_sheet": 1}}
        for ws in selected:
            start_row, mapping, mode = _sheet_plan(ws)
            sheet_rows = 0
            sheet_skipped: Counter[str] = Counter()
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not any(cell not in (None, "") for cell in row):
                    continue
                cc_code = helpers.extract_cc_code(_value(row, mapping, "depreciation_cc"))
                monthly_depr = helpers.safe_float(_value(row, mapping, "monthly_depr"))
                apr_interest = helpers.safe_float(_value(row, mapping, "apr_interest"))
                may_interest = helpers.safe_float(_value(row, mapping, "may_interest"))
                category_key, category_status = _category_status(_value(row, mapping, "category"))
                if not cc_code:
                    sheet_skipped["missing_depreciation_cc"] += 1
                    continue
                if monthly_depr <= 0 and apr_interest <= 0 and may_interest <= 0:
                    sheet_skipped["no_fixed_asset_amount"] += 1
                    continue
                if category_status == "out_of_scope":
                    sheet_skipped["out_of_scope_category"] += 1
                    continue
                if category_key is None:
                    sheet_skipped["missing_or_unknown_category"] += 1
                    continue
                by_cc[cc_code] += 1
                by_category[category_key] += 1
                sheet_rows += 1
            skipped.update(sheet_skipped)
            by_sheet[ws.title] = {"mode": mode, "selected": True, "source_rows": sheet_rows, "skipped_reasons": dict(sheet_skipped)}
        return {
            "source_rows": sum(by_cc.values()), "by_cc": dict(by_cc), "by_category": dict(by_category),
            "by_sheet": by_sheet, "selected_sheets": [ws.title for ws in selected], "skipped_reasons": dict(skipped),
        }
    finally:
        wb.close()


def parse_fixed_assets(conn: sqlite3.Connection, fa_path: str = None, source_dir: str = None) -> dict:
    fpath = fa_path or find_fixed_assets_file(source_dir)
    if not fpath or not Path(fpath).is_file():
        return {"total": 0, "source_rows": 0, "parsed_assets": 0, "skipped_reasons": {"missing_file": 1}}

    rate_row = conn.execute("SELECT value FROM sys_params WHERE key='exchange_rate_usd_vnd'").fetchone()
    if not rate_row or rate_row[0] in (None, ""):
        raise ValueError("Missing authoritative exchange_rate_usd_vnd in sys_params")
    rate = float(rate_row[0])
    fy_row = conn.execute("SELECT value FROM sys_params WHERE key='fiscal_year'").fetchone()
    if not fy_row or fy_row[0] in (None, ""):
        raise ValueError("Missing authoritative fiscal_year in sys_params")
    fiscal_year = int(str(fy_row[0]).replace("FY", ""))
    fy_months = helpers.get_fy_months(fiscal_year)
    source_snapshot = _source_snapshot(Path(fpath))

    wb = openpyxl.load_workbook(Path(fpath), read_only=True, data_only=True)
    formula_wb = openpyxl.load_workbook(Path(fpath), read_only=True, data_only=False)
    pending: list[tuple[str, str, float, float, str, int, str, int, str]] = []
    parsed_assets = 0
    source_rows = 0
    depr_rows = 0
    interest_rows = 0
    by_cc: Counter[str] = Counter()
    by_category: Counter[str] = Counter()
    by_sheet: dict[str, dict[str, Any]] = {}
    skipped: Counter[str] = Counter()
    warnings: Counter[str] = Counter()
    audit_pending: list[tuple[Any, ...]] = []

    try:
        selected = _selected_worksheets(wb)
        if not selected:
            raise ValueError("Fixed-assets workbook has no recognizable current source sheet")
        for ws in selected:
            formula_ws = formula_wb[ws.title]
            start_row, mapping, mode = _sheet_plan(ws)
            sheet_source_rows = 0
            sheet_parsed_assets = 0
            sheet_skipped: Counter[str] = Counter()
            for row_number, (row, formula_row) in enumerate(
                zip(
                    ws.iter_rows(min_row=start_row, values_only=True),
                    formula_ws.iter_rows(min_row=start_row, values_only=True),
                ),
                start=start_row,
            ):
                if not any(cell not in (None, "") for cell in row):
                    continue
                cache_gap_fields = [
                    field
                    for field in ("monthly_depr", "last_month", "last_month_depr", "apr_interest", "may_interest")
                    if isinstance(_value(formula_row, mapping, field), str)
                    and str(_value(formula_row, mapping, field)).startswith("=")
                    and _value(row, mapping, field) in (None, "")
                ]
                if cache_gap_fields:
                    raise ValueError(
                        f"Missing cached formula value at {ws.title}!{row_number} for "
                        + ", ".join(cache_gap_fields)
                    )
                formula_fields = [
                    field
                    for field in ("monthly_depr", "last_month", "last_month_depr", "apr_interest", "may_interest")
                    if isinstance(_value(formula_row, mapping, field), str)
                    and str(_value(formula_row, mapping, field)).startswith("=")
                ]
                formula_cache_status = (
                    "NO_FORMULA"
                    if not formula_fields
                    else "FORMULA_CACHE_PRESENT:" + ",".join(formula_fields)
                )
                cc_code = helpers.extract_cc_code(_value(row, mapping, "depreciation_cc"))
                monthly_depr = helpers.safe_float(_value(row, mapping, "monthly_depr"))
                last_month = helpers.normalize_period(_value(row, mapping, "last_month"))
                raw_last_month_depr = _value(row, mapping, "last_month_depr")
                last_month_depr = (
                    None
                    if raw_last_month_depr in (None, "")
                    else helpers.safe_float(raw_last_month_depr)
                )
                apr_interest = helpers.safe_float(_value(row, mapping, "apr_interest"))
                may_interest = helpers.safe_float(_value(row, mapping, "may_interest"))
                category_value = _value(row, mapping, "category")
                category_key, category_status = _category_status(category_value)
                if category_key is None and mode == "legacy" and category_value in (None, ""):
                    category_key = "tools_furniture_fixtures"
                    category_status = "supported"
                    warnings["legacy_missing_category_defaulted_to_tools"] += 1
                audit_base = (
                    fiscal_year,
                    source_snapshot,
                    str(Path(fpath)),
                    ws.title,
                    row_number,
                    str(_value(row, mapping, "asset_no") or ""),
                    str(_value(row, mapping, "asset_text") or ""),
                    str(category_value or ""),
                    category_key,
                    str(_value(row, mapping, "control_cc") or ""),
                    str(_value(row, mapping, "depreciation_cc") or ""),
                    monthly_depr,
                    last_month,
                    last_month_depr,
                    apr_interest,
                    may_interest,
                    formula_cache_status,
                )
                if not cc_code:
                    sheet_skipped["missing_depreciation_cc"] += 1
                    audit_pending.append((*audit_base, "EXCLUDED", "missing_depreciation_cc"))
                    continue
                if monthly_depr == 0 and apr_interest == 0 and may_interest == 0:
                    sheet_skipped["no_fixed_asset_amount"] += 1
                    audit_pending.append((*audit_base, "EXCLUDED", "no_fixed_asset_amount"))
                    continue
                if category_status == "out_of_scope":
                    sheet_skipped["out_of_scope_category"] += 1
                    audit_pending.append((*audit_base, "EXCLUDED", "out_of_scope_category"))
                    continue
                if category_key is None:
                    raw_category = str(category_value or "<blank>").strip()
                    raise ValueError(f"Unknown fixed-assets Category at {ws.title}!{row_number}: {raw_category}")

                asset_tag = _asset_tag(_value(row, mapping, "asset_no"), _value(row, mapping, "asset_text"), ws.title, row_number)
                source_rows += 1
                sheet_source_rows += 1
                parsed_assets += 1
                sheet_parsed_assets += 1
                by_cc[cc_code] += 1
                by_category[category_key] += 1
                if last_month in fy_months and last_month_depr is None:
                    raise ValueError(
                        f"Missing Q/last-month depreciation at {ws.title}!{row_number}; "
                        "the parser must not substitute the monthly L value"
                    )
                audit_pending.append((*audit_base, "INCLUDED", None))

                dep_desc = f"fixed_assets_depr|{category_key}|{asset_tag}"
                int_desc = f"fixed_assets_interest|{category_key}|{asset_tag}"
                dep_account = int(CATEGORY_SPECS[category_key]["depreciation_account"])
                for period, val in expand_depreciation_schedule(monthly_depr, last_month, last_month_depr, fy_months).items():
                    pending.append((
                        "fixed_assets", period, _round_vnd(val, rate), val, cc_code, dep_account, dep_desc,
                        fiscal_year, source_snapshot,
                    ))
                    depr_rows += 1
                for period, val in expand_interest_schedule(apr_interest, may_interest, last_month, fy_months).items():
                    pending.append((
                        "fixed_assets", period, _round_vnd(val, rate), val, cc_code, INTEREST_ACCOUNT, int_desc,
                        fiscal_year, source_snapshot,
                    ))
                    interest_rows += 1
            skipped.update(sheet_skipped)
            by_sheet[ws.title] = {
                "mode": mode, "selected": True, "source_rows": sheet_source_rows,
                "parsed_assets": sheet_parsed_assets, "skipped_reasons": dict(sheet_skipped),
            }
    finally:
        wb.close()
        formula_wb.close()

    cursor = conn.cursor()
    cursor.execute("SAVEPOINT fixed_assets_import")
    try:
        cursor.execute(
            "DELETE FROM fact_input_data WHERE source='fixed_assets' AND fiscal_year=?",
            (fiscal_year,),
        )
        cursor.execute(
            "DELETE FROM audit_fixed_asset_import_rows WHERE fiscal_year=?",
            (fiscal_year,),
        )
        cursor.executemany(
            """
            INSERT INTO audit_fixed_asset_import_rows
            (fiscal_year, source_snapshot, source_file, source_sheet, source_row, asset_no, asset_text,
             category_raw, category_key, control_cc, depreciation_cc, monthly_depr_usd, terminal_period,
             terminal_depr_usd, apr_interest_usd, may_interest_usd, formula_cache_status,
             inclusion_status, exclusion_reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            audit_pending,
        )
        cursor.executemany(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, amount_usd, cc_code, account_code, description, fiscal_year, source_snapshot)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            pending,
        )
        cursor.execute("RELEASE SAVEPOINT fixed_assets_import")
        conn.commit()
    except Exception:
        cursor.execute("ROLLBACK TO SAVEPOINT fixed_assets_import")
        cursor.execute("RELEASE SAVEPOINT fixed_assets_import")
        raise

    return {
        "total": len(pending), "source_rows": source_rows, "parsed_assets": parsed_assets,
        "depreciation_rows": depr_rows, "interest_rows": interest_rows,
        "fiscal_year": fiscal_year, "exchange_rate": rate, "source_snapshot": source_snapshot,
        "selected_sheets": [ws.title for ws in selected], "by_cc": dict(by_cc),
        "by_category": dict(by_category), "by_sheet": by_sheet,
        "skipped_reasons": dict(skipped), "warnings": dict(warnings),
    }
