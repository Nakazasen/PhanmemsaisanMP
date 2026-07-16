"""Parser for FY2027 foreign-worker paperwork costs."""

from __future__ import annotations

import ast
import datetime
import operator
import sqlite3
from pathlib import Path

import openpyxl

from src.utils.excel_helpers import get_fy_months, normalize_cc_code, safe_float
from src.utils.source_manifest import resolve_manifest_file

SOURCE_NAME = "nnn_paperwork"
FORM_ROW = 137


def _format_formula_number(value: float) -> str:
    return str(int(round(value))) if abs(value - round(value)) < 1e-9 else str(value)


def _normalize_header(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y%m")
    val_str = str(val).strip().lower()
    for char in ["/", "-", ".", " ", "_", "\n", "\r", "\t"]:
        val_str = val_str.replace(char, "")
    return val_str


def _matches_header(cell: str, aliases: set[str]) -> bool:
    return cell in aliases or any(alias and alias in cell for alias in aliases)


_FORMULA_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
}


def _evaluate_simple_formula(expr: str) -> float | None:
    """Evaluate simple numeric Excel formulas without cell refs/functions."""
    try:
        node = ast.parse(expr, mode="eval")
    except SyntaxError:
        return None

    def _eval(inner):
        if isinstance(inner, ast.Expression):
            return _eval(inner.body)
        if isinstance(inner, ast.Constant) and isinstance(inner.value, (int, float)):
            return float(inner.value)
        if isinstance(inner, ast.UnaryOp) and type(inner.op) in _FORMULA_OPERATORS:
            return _FORMULA_OPERATORS[type(inner.op)](_eval(inner.operand))
        if isinstance(inner, ast.BinOp) and type(inner.op) in _FORMULA_OPERATORS:
            return _FORMULA_OPERATORS[type(inner.op)](_eval(inner.left), _eval(inner.right))
        raise ValueError("unsupported formula")

    try:
        return float(_eval(node))
    except Exception:
        return None


CC_RAW = [
    "cost center", "costcenter", "code phòng chịu chi phí", "code phong chiu chi phi",
    "mã cost center", "ma cost center", "mã costcenter", "ma costcenter",
    "コードセンター", "原価センタ", "cc_code", "cost_center", "cc code", "cccode"
]
ACC_RAW = [
    "account code", "accountcode", "code tài khoản chịu chi phí", "code tai khan chiu chi phi",
    "code tài khoản", "code tai khoan", "mã tài khoản", "ma tai khoan", "コード",
    "勘定科目", "account"
]
EMP_CODE_RAW = [
    "mã nhân viên", "ma nhan vien", "employee code", "mã nv", "ma nv", "emp code", "社員番号", "code nhân viên"
]
EMP_NAME_RAW = [
    "tên nhân viên", "ten nhan vien", "employee name", "tên nv", "ten nv", "emp name", "氏名", "お名前", "name", "tên người nước ngoài"
]

CC_ALIASES = { _normalize_header(x) for x in CC_RAW }
ACC_ALIASES = { _normalize_header(x) for x in ACC_RAW }
EMP_CODE_ALIASES = { _normalize_header(x) for x in EMP_CODE_RAW }
EMP_NAME_ALIASES = { _normalize_header(x) for x in EMP_NAME_RAW }

def month_map_for_fiscal_year(fiscal_year: int) -> dict[str, int]:
    """Build accepted month headers for a selected FY without a year ceiling."""
    result: dict[str, int] = {}
    english = ("jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec")
    for index, period in enumerate(get_fy_months(fiscal_year)):
        year = int(period[:4])
        month = int(period[-2:])
        for label in (period, f"{year}{month:02d}", f"{english[month - 1]}{year}", f"{month}月", f"{year}年{month}月"):
            result[_normalize_header(label)] = index
    return result


def find_nnn_paperwork_file(source_dir: str | None = None) -> str | None:
    manifest_path = resolve_manifest_file(source_dir, "nnn_paperwork")
    if manifest_path:
        return manifest_path

    search_dir = Path(source_dir or Path.cwd())
    for path in search_dir.glob("*.xlsx"):
        name = path.name.lower()
        if "nnn" in name and "giay" in name:
            return str(path)
    for path in search_dir.glob("*.xlsx"):
        name = path.name.lower()
        if "nnn" in name:
            return str(path)
    return None


def parse_nnn_paperwork(conn: sqlite3.Connection, source_dir: str | None = None, workbook_path: str | None = None) -> dict[str, int | str]:
    """Load NNN/VISA/GPLD/Passport paperwork workbook into explicit FORM row 137."""
    fy_row = conn.execute("SELECT value FROM sys_params WHERE key='fiscal_year'").fetchone()
    if not fy_row:
        raise ValueError("Thiếu năm tài chính trong dữ liệu lần chạy; không được tự mặc định FY2027.")
    fiscal_year = int(str(fy_row[0]).upper().replace("FY", "").strip())
    fy_months = get_fy_months(fiscal_year)
    month_map = month_map_for_fiscal_year(fiscal_year)
    path = workbook_path or find_nnn_paperwork_file(source_dir)
    if not path or not Path(path).is_file():
        return {"inserted": 0, "skipped": 0, "errors": 0, "path": path or ""}

    valid_cc_codes = {
        str(row[0]).strip()
        for row in conn.execute("SELECT code FROM dim_cost_centers").fetchall()
        if row[0] is not None
    }
    cursor = conn.cursor()
    cursor.execute("DELETE FROM fact_input_data WHERE source = ?", (SOURCE_NAME,))

    inserted = 0
    skipped = 0
    errors = 0

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    formula_wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        expected_sheet = f"FY{fiscal_year}"
        if expected_sheet not in wb.sheetnames:
            raise ValueError(f"Thiếu sheet {expected_sheet} trong file giấy tờ người nước ngoài.")
        ws = wb[expected_sheet]

        # Scan up to 10 rows to locate header row
        header_row_idx = None
        header_cells = None
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            if not row:
                continue
            normalized_row = [_normalize_header(cell) for cell in row]
            has_cc = any(_matches_header(cell, CC_ALIASES) for cell in normalized_row)
            has_acc = any(_matches_header(cell, ACC_ALIASES) for cell in normalized_row)
            has_month = any(cell in month_map for cell in normalized_row)
            if has_cc and has_acc and has_month:
                header_row_idx = row_idx
                header_cells = normalized_row
                break

        if header_row_idx is None:
            raise ValueError(f"Không tìm thấy dòng header hợp lệ trong file: {path}")

        cc_col_idx = None
        acc_col_idx = None
        emp_code_col_idx = None
        emp_name_col_idx = None
        month_col_indices = {}

        for col_idx, cell in enumerate(header_cells):
            if _matches_header(cell, CC_ALIASES) and cc_col_idx is None:
                cc_col_idx = col_idx
            elif _matches_header(cell, ACC_ALIASES) and acc_col_idx is None:
                acc_col_idx = col_idx
            elif _matches_header(cell, EMP_CODE_ALIASES) and emp_code_col_idx is None:
                emp_code_col_idx = col_idx
            elif _matches_header(cell, EMP_NAME_ALIASES) and emp_name_col_idx is None:
                emp_name_col_idx = col_idx
            elif cell in month_map:
                period = month_map[cell]
                if period not in month_col_indices:
                    month_col_indices[period] = col_idx

        # Check required columns
        if cc_col_idx is None or acc_col_idx is None or not month_col_indices:
            raise ValueError(f"File thiếu các cột bắt buộc (Cost Center, Account Code hoặc Tháng) trong file: {path}")

        formula_ws = formula_wb[ws.title]
        formula_rows = formula_ws.iter_rows(min_row=header_row_idx + 1, values_only=True)
        for row, formula_row in zip(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), formula_rows):
            cc_code = normalize_cc_code(row[cc_col_idx] if len(row) > cc_col_idx else None)
            if not cc_code:
                skipped += 1
                continue
            if cc_code not in valid_cc_codes:
                errors += 1
                continue

            account_code = int(safe_float(row[acc_col_idx] if len(row) > acc_col_idx else 0))
            if account_code <= 0:
                errors += 1
                continue

            employee_code = ""
            if emp_code_col_idx is not None and len(row) > emp_code_col_idx:
                employee_code = str(row[emp_code_col_idx] or "").strip()

            employee_name = ""
            if emp_name_col_idx is not None and len(row) > emp_name_col_idx:
                employee_name = str(row[emp_name_col_idx] or "").strip()

            description = "NNN paperwork"
            if employee_code or employee_name:
                description = f"{description}: {employee_code} {employee_name}".strip()

            for offset, period in enumerate(fy_months):
                if offset not in month_col_indices:
                    continue
                col_idx = month_col_indices[offset]
                amount = safe_float(row[col_idx] if len(row) > col_idx else 0)
                raw_formula = formula_row[col_idx] if len(formula_row) > col_idx else None
                if isinstance(raw_formula, str) and raw_formula.startswith("="):
                    formula_expr = raw_formula[1:]
                    if amount <= 0:
                        evaluated = _evaluate_simple_formula(formula_expr)
                        if evaluated is not None:
                            amount = evaluated
                else:
                    formula_expr = _format_formula_number(amount)
                if amount <= 0:
                    continue
                cursor.execute(
                    """
                    INSERT INTO fact_input_data
                    (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
                    VALUES (?, ?, ?, ?, ?, ?, 'base', ?)
                    """,
                    (
                        SOURCE_NAME,
                        period,
                        amount,
                        cc_code,
                        account_code,
                        FORM_ROW,
                        f"{description}|formula_expr={formula_expr}",
                    ),
                )
                inserted += 1
    finally:
        wb.close()
        formula_wb.close()

    conn.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors, "path": str(path)}
