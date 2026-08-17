"""Trình phân tích cú pháp cho tệp kế hoạch nhân sự và thời gian dạng bảng."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import Any
from openpyxl import load_workbook
import xlrd
import unicodedata
from src.utils.excel_helpers import normalize_cc_code
from src.utils.fiscal_periods import fiscal_periods as get_fy_months

HEADCOUNT_SOURCE = "department_plan"


@dataclass
class PlanParseResult:
    path: str
    status: str
    fiscal_year: int
    department_no: str = ""
    cc_code: str = ""
    department_name: str = ""
    department_name_jp: str = ""
    department_name_vn: str = ""
    sheet_name: str = ""
    lookup_status: str = "missing"
    verification_method: str = "unverified"
    errors: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def _cc_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _resolve_lookup_identity(
    cc_code: str,
    department_name: str,
    lookup_rows: list[tuple[str, str, str]],
) -> tuple[str, str, str]:
    if not lookup_rows:
        return "missing", "", ""
    cc_text = _norm_text(cc_code)
    cc_norm = normalize_cc_code(cc_text) or cc_text
    dept_norm = _norm_text(department_name)
    exact_matches = [
        row for row in lookup_rows
        if (normalize_cc_code(_norm_text(row[0])) or _norm_text(row[0])) == cc_norm and (
            _norm_text(row[1]) == dept_norm or _norm_text(row[2]) == dept_norm
        )
    ]
    if len(exact_matches) == 1:
        return "matched", str(exact_matches[0][1] or "").strip(), str(exact_matches[0][2] or "").strip()
    if len(exact_matches) > 1:
        return "ambiguous", str(exact_matches[0][1] or "").strip(), str(exact_matches[0][2] or "").strip()
    cc_matches = [
        row for row in lookup_rows
        if (normalize_cc_code(_norm_text(row[0])) or _norm_text(row[0])) == cc_norm
    ]
    if cc_matches:
        return "mismatch", str(cc_matches[0][1] or "").strip(), str(cc_matches[0][2] or "").strip()
    return "missing", "", ""


def parse_headcount_time_plan(path: str, fiscal_year: int) -> PlanParseResult:
    """Parse one departmental plan layout without writing the database."""
    result = PlanParseResult(path=os.path.abspath(path), status="error", fiscal_year=fiscal_year)
    match = re.match(r"\s*(\d+)\.", os.path.basename(path))
    result.department_no = match.group(1) if match else ""
    workbook = None
    try:
        if path.lower().endswith(".xlsx"):
            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
            sheet = workbook.worksheets[0]
            cell_value = lambda row, col: sheet.cell(row + 1, col + 1).value
            nrows, ncols, sheet_name = sheet.max_row, sheet.max_column, sheet.title
            lookup_rows = []
            for candidate in workbook.worksheets[1:]:
                for row in candidate.iter_rows(min_col=1, max_col=3, values_only=True):
                    code = normalize_cc_code(_cc_text(row[0])) or _cc_text(row[0])
                    if code:
                        lookup_rows.append((code, _cc_text(row[1]), _cc_text(row[2])))
        else:
            workbook = xlrd.open_workbook(path)
            sheet = workbook.sheet_by_index(0)
            cell_value = sheet.cell_value
            nrows, ncols, sheet_name = sheet.nrows, sheet.ncols, sheet.name
            lookup_rows = []
            for candidate in workbook.sheets()[1:]:
                for row_index in range(candidate.nrows):
                    code = normalize_cc_code(_cc_text(candidate.cell_value(row_index, 0))) or _cc_text(candidate.cell_value(row_index, 0))
                    if code:
                        name_jp = _cc_text(candidate.cell_value(row_index, 1)) if candidate.ncols > 1 else ""
                        name_vn = _cc_text(candidate.cell_value(row_index, 2)) if candidate.ncols > 2 else ""
                        lookup_rows.append((code, name_jp, name_vn))
    except Exception as exc:
        result.errors.append(
            f"Không mở được file: {type(exc).__name__}. "
            "Nguyên nhân: Tệp Excel bị khóa, không đúng định dạng hoặc bị hỏng. "
            "Cách xử lý: Kiểm tra lại tệp Excel kế hoạch nhân sự và đóng các ứng dụng đang mở tệp này."
        )
        return result
    try:
        result.sheet_name = sheet_name
        if nrows < 28 or ncols < 14:
            result.errors.append(
                f"Cấu trúc không đủ: {nrows} dòng, {ncols} cột. "
                "Nguyên nhân: Bảng tính kế hoạch nhân sự cần tối thiểu 28 dòng và 14 cột theo định dạng chuẩn. "
                "Cách xử lý: Sử dụng đúng biểu mẫu kế hoạch nhân sự và thời gian quy định."
            )
            return result
        raw_cc = _cc_text(cell_value(4, 0))
        result.cc_code = normalize_cc_code(raw_cc) or raw_cc
        result.department_name = _cc_text(cell_value(4, 1))
        lookup_status, name_jp, name_vn = _resolve_lookup_identity(
            result.cc_code, result.department_name, lookup_rows
        )
        result.lookup_status = lookup_status
        result.department_name_jp = name_jp
        result.department_name_vn = name_vn
        if lookup_status == "matched":
            result.verification_method = "workbook_bilingual_lookup"
        elif lookup_status == "mismatch":
            result.verification_method = "name_confirmation_required"
        elif lookup_status == "ambiguous":
            result.errors.append(
                "Lookup nội bộ có nhiều cặp tên cùng khớp CC và B5. "
                "Nguyên nhân: Có nhiều hơn một phòng ban trùng khớp mã trung tâm chi phí trong sheet tra cứu. "
                "Cách xử lý: Kiểm tra và chuẩn hóa lại sheet tra cứu nội bộ của tệp."
            )
        months = [int(_number(cell_value(7, col)) or 0) for col in range(2, 14)]
        if months != [int(period[-2:]) for period in get_fy_months(fiscal_year)]:
            result.errors.append(
                f"Thứ tự tháng không đúng FY{fiscal_year}: {months}. "
                "Nguyên nhân: Các cột tháng phải theo thứ tự năm tài chính từ tháng 4 đến tháng 3. "
                "Cách xử lý: Hiệu chỉnh lại dòng tiêu đề tháng (dòng 8) theo đúng thứ tự 4, 5, ..., 12, 1, 2, 3."
            )
            return result
        row_map = {
            "headcount_expat": 9, "headcount_staff": 10, "headcount_worker": 11,
            "headcount_local_total": 12, "headcount_total": 13,
            "fixed_hours_expat": 16, "fixed_hours_staff": 17, "fixed_hours_worker": 18,
            "fixed_hours_local_total": 19, "fixed_hours_total": 20,
            "overtime_hours_expat": 23, "overtime_hours_staff": 24, "overtime_hours_worker": 25,
            "overtime_hours_local_total": 26, "overtime_hours_total": 27,
        }
        values: dict[str, list[float]] = {}
        for metric, row_index in row_map.items():
            parsed = [_number(cell_value(row_index, col)) for col in range(2, 14)]
            if any(value is None for value in parsed):
                result.errors.append(
                    f"Giá trị không phải số tại dòng {row_index + 1} ({metric}). "
                    "Nguyên nhân: Ô chứa ký tự hoặc giá trị trống không thể quy đổi ra số. "
                    f"Cách xử lý: Nhập đầy đủ số liệu dạng số tại dòng {row_index + 1} các cột tháng."
                )
            else:
                values[metric] = [float(value or 0) for value in parsed]
        if result.errors:
            return result
        for index, period in enumerate(get_fy_months(fiscal_year)):
            checks = (
                (values["headcount_local_total"][index], values["headcount_staff"][index] + values["headcount_worker"][index], "tổng người Việt"),
                (values["headcount_total"][index], values["headcount_expat"][index] + values["headcount_local_total"][index], "tổng số người"),
                (values["fixed_hours_local_total"][index], values["fixed_hours_staff"][index] + values["fixed_hours_worker"][index], "tổng giờ cố định người Việt"),
                (values["fixed_hours_total"][index], values["fixed_hours_expat"][index] + values["fixed_hours_local_total"][index], "tổng giờ cố định"),
                (values["overtime_hours_local_total"][index], values["overtime_hours_staff"][index] + values["overtime_hours_worker"][index], "tổng tăng ca người Việt"),
                (values["overtime_hours_total"][index], values["overtime_hours_expat"][index] + values["overtime_hours_local_total"][index], "tổng tăng ca"),
            )
            for actual, expected, label in checks:
                if abs(actual - expected) > 0.01:
                    result.errors.append(
                        f"{period}: {label} không khớp ({actual} != {expected}). "
                        "Nguyên nhân: Tổng số thực tế trên bảng tính không khớp với tổng các thành phần chi tiết. "
                        f"Cách xử lý: Kiểm tra lại công thức hoặc số liệu {label} tại kỳ {period}."
                    )
            result.rows.append({
                "period": period,
                "headcount_expat": values["headcount_expat"][index],
                "headcount_staff": values["headcount_staff"][index],
                "headcount_worker": values["headcount_worker"][index],
                "fixed_hours_expat": values["fixed_hours_expat"][index],
                "fixed_hours_staff": values["fixed_hours_staff"][index],
                "fixed_hours_worker": values["fixed_hours_worker"][index],
                "fixed_hours_local": values["fixed_hours_local_total"][index],
                "overtime_hours_expat": values["overtime_hours_expat"][index],
                "overtime_hours_staff": values["overtime_hours_staff"][index],
                "overtime_hours_worker": values["overtime_hours_worker"][index],
                "overtime_hours_local": values["overtime_hours_local_total"][index],
                "source_cells": "C:N; rows 10-28",
            })
        result.status = "valid" if not result.errors else "error"
        return result
    finally:
        if path.lower().endswith(".xlsx") and workbook is not None:
            workbook.close()
