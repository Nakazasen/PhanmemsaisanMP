"""Bộ đọc sổ làm việc chi phí sinh nhật FY2027."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl

from src.engine.account_resolver import resolve_account_code_for_source
from src.utils.excel_helpers import get_fy_months, normalize_cc_code, safe_float, normalize_text
from src.utils.source_manifest import resolve_manifest_file


SOURCE_NAME = "birthday_workbook"
FORM_ROW = 59


def _formula_number(value: float) -> str:
    return str(int(round(value))) if abs(value - round(value)) < 1e-9 else str(value)


def _birthday_unit_price(conn: sqlite3.Connection) -> float:
    """Resolve birthday price from the annual allocation-rules source."""
    matches = []
    for row in conn.execute("SELECT id, item_name, unit_price FROM map_allocation_rules").fetchall():
        name = normalize_text(row["item_name"] if hasattr(row, "keys") else row[1])
        if any(token in name for token in ("birthday", "sinh nhat", "誕生日")):
            matches.append(row)
    if len(matches) != 1:
        raise ValueError(
            "Không xác định duy nhất đơn giá sinh nhật trong file quy tắc phân bổ của năm đang chạy."
        )
    price = float(matches[0]["unit_price"] if hasattr(matches[0], "keys") else matches[0][2] or 0)
    if price <= 0:
        raise ValueError("Đơn giá sinh nhật của năm đang chạy phải lớn hơn 0.")
    return price


def find_birthday_file(source_dir: str | None = None) -> str | None:
    manifest_path = resolve_manifest_file(source_dir, "birthday")
    if manifest_path:
        return manifest_path

    search_dir = Path(source_dir or Path.cwd())
    for path in search_dir.glob("*.xlsx"):
        name = path.name.lower()
        if "sinh" in name and "nh" in name:
            return str(path)
    return None


def parse_birthday_workbook(conn: sqlite3.Connection, source_dir: str | None = None, workbook_path: str | None = None) -> dict[str, int | str]:
    """Load birthday workbook amounts into explicit FORM row 59."""
    fy_row = conn.execute("SELECT value FROM sys_params WHERE key='fiscal_year'").fetchone()
    if not fy_row:
        raise ValueError("Thiếu năm tài chính trong dữ liệu lần chạy; không được tự mặc định FY2027.")
    fiscal_year = int(str(fy_row[0]).upper().replace("FY", "").strip())
    fy_months = get_fy_months(fiscal_year)
    unit_price_vnd = _birthday_unit_price(conn)
    path = workbook_path or find_birthday_file(source_dir)
    if not path or not Path(path).is_file():
        return {"inserted": 0, "skipped": 0, "errors": 0, "path": path or ""}

    valid_cc_codes = {
        str(row[0]).strip()
        for row in conn.execute("SELECT code FROM dim_cost_centers").fetchall()
        if row[0] is not None
    }
    cursor = conn.cursor()
    cursor.execute("DELETE FROM fact_input_data WHERE source = ?", (SOURCE_NAME,))

    inserted = 0
    skipped = 0
    errors = 0

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheet = f"Sinh nhật MP FY{fiscal_year}"
        if expected_sheet not in wb.sheetnames:
            raise ValueError(f"Thiếu sheet {expected_sheet} trong file sinh nhật.")
        ws = wb[expected_sheet]
        for row in ws.iter_rows(min_row=4, values_only=True):
            cc_code = normalize_cc_code(row[0] if len(row) > 0 else None)
            if not cc_code:
                skipped += 1
                continue
            if cc_code not in valid_cc_codes:
                errors += 1
                continue

            cc_name = str(row[1] or "").strip() if len(row) > 1 else ""
            for offset, period in enumerate(fy_months):
                count = safe_float(row[2 + offset] if len(row) > 2 + offset else 0)
                if count <= 0:
                    continue
                amount = count * unit_price_vnd
                count_text = str(int(round(count))) if abs(count - round(count)) < 1e-9 else str(count)
                cursor.execute(
                    """
                    INSERT INTO fact_input_data
                    (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
                    VALUES (?, ?, ?, ?, ?, ?, 'base', ?)
                    """,
                    (
                        SOURCE_NAME,
                        period,
                        amount,
                        cc_code,
                        resolve_account_code_for_source(conn, SOURCE_NAME, cc_code),
                        FORM_ROW,
                        f"Birthday workbook: {cc_name}|formula_expr={count_text}*{_formula_number(unit_price_vnd)}",
                    ),
                )
                inserted += 1
    finally:
        wb.close()

    conn.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors, "path": str(path)}
