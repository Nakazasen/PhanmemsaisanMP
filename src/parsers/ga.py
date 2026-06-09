"""
GA parser.

Loads:
- per-unit GA costs into fact_input_data as source='ga_unit_price'
- working days into sys_params as working_days_YYYYMM
- monthly headcount by CC into fact_monthly_headcount as source='ga'
"""

import os
import re
import sqlite3
import unicodedata
from collections import defaultdict
from typing import Any

import openpyxl
import pandas as pd

from src.engine.account_resolver import resolve_account_code_for_source
from src.utils.excel_helpers import get_fy_months, normalize_cc_code, safe_float
from src.utils.source_manifest import resolve_manifest_file

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))

GA_FILE_KEYWORDS = ("fy", "mp")
CALC_SHEET_KEYWORDS = ("cach tinh", "振替", "計算", "tinh")
WORKING_DAY_KEYWORDS = ("稼働日", "ngay lam")
HEADCOUNT_KEYWORDS = ("人員数", "so nguoi", "headcount", "staff", "worker", "direct", "indirect")
STAFF_CATEGORY_KEYWORDS = ("indirect", "gian tiep", "staff", "nhan vien", "dinh phi")
WORKER_CATEGORY_KEYWORDS = ("direct", "truc tiep", "worker", "cong nhan", "bien phi", "g7")
EXCLUDE_HEADCOUNT_ROW_KEYWORDS = ("vnd", "tiền", "tien", "alloc", "rate", "ratio", "đơn giá", "don gia")


def _normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\n", " ").replace("\u3000", " ").strip().lower()
    return " ".join(text.split())


def _looks_like_ga_workbook(path: str, fiscal_year: int) -> bool:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        normalized_sheets = [_normalize_text(name) for name in workbook.sheetnames]
        if not any("cach tinh" in name or "振替" in name or "計算" in name for name in normalized_sheets):
            return False

        worksheet = workbook[workbook.sheetnames[0]]
        if worksheet.max_row < 8 or worksheet.max_column < 12:
            return False

        header_cells: list[str] = []
        for row in worksheet.iter_rows(min_row=1, max_row=6, max_col=18, values_only=True):
            for value in row:
                if value is not None:
                    header_cells.append(_normalize_text(value))
        header_blob = " | ".join(header_cells)
        return (
            str(fiscal_year) in header_blob
            and "vnd" in header_blob
            and ("yotei" in header_blob or "item" in header_blob or "項目" in header_blob)
            and ("費" in header_blob or "製造" in header_blob or "tai khoan" in header_blob or "account" in header_blob)
        )
    finally:
        workbook.close()


def _find_ga_file(source_dir: str | None, fiscal_year: int) -> str | None:
    manifest_path = resolve_manifest_file(source_dir, "ga")
    if manifest_path:
        return manifest_path

    search_dir = source_dir or BASE_DIR
    if not os.path.isdir(search_dir):
        return None

    candidates: list[tuple[int, str]] = []
    for name in os.listdir(search_dir):
        lower_name = name.lower()
        if not lower_name.endswith(".xlsx") or lower_name.startswith("~$"):
            continue

        path = os.path.join(search_dir, name)
        if not _looks_like_ga_workbook(path, fiscal_year):
            continue

        score = 0
        if f"fy{fiscal_year}" in lower_name:
            score += 2
        if "mp" in lower_name:
            score += 1
        if "総務課" in name or "tong vu" in lower_name or "hanh chinh" in lower_name:
            score += 2
        candidates.append((score, path))

    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], os.path.basename(item[1]).lower()))
    return candidates[0][1]


def _iter_calc_sheet_names(sheet_names: list[str]) -> list[str]:
    result: list[str] = []
    for sheet_name in sheet_names:
        normalized = _normalize_text(sheet_name)
        if any(token in normalized for token in ("cach tinh", "振替", "計算", "tinh")):
            result.append(sheet_name)
            continue
        if any(keyword in normalized for keyword in CALC_SHEET_KEYWORDS):
            result.append(sheet_name)
    return result


def _parse_ga_main_sheet(df: pd.DataFrame, fy_months: list[str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    header_row = None

    for i in range(min(25, len(df))):
        first_col = _normalize_text(df.iloc[i, 0] if len(df.columns) > 0 else "")
        if "item" in first_col or "内容" in first_col or "項目" in first_col:
            header_row = i
            break

    if header_row is None:
        return records

    for i in range(header_row + 1, len(df)):
        row = df.iloc[i]
        item_name = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
        if not item_name:
            continue

        lowered_name = _normalize_text(item_name)
        if "total" in lowered_name or "tổng" in lowered_name or "tong" in lowered_name or "合計" in item_name:
            continue

        basis = _normalize_text(row.iloc[14] if len(row) > 14 else "")
        driver = "working_days" if any(keyword in basis for keyword in WORKING_DAY_KEYWORDS) else "headcount_per_person"

        for month_index in range(min(12, len(fy_months))):
            value_index = month_index + 1
            if value_index >= len(row):
                continue
            amount = safe_float(row.iloc[value_index])
            if amount <= 0:
                continue
            records.append(
                {
                    "period": fy_months[month_index],
                    "amount": amount,
                    "item": item_name,
                    "driver": driver,
                }
            )

    return records


def _parse_working_days_sheet(df: pd.DataFrame, fy_months: list[str]) -> dict[str, float]:
    working_days: dict[str, float] = {}
    for i in range(len(df)):
        label = _normalize_text(df.iloc[i, 1] if len(df.columns) > 1 else "")
        if not any(keyword in label for keyword in WORKING_DAY_KEYWORDS):
            continue
        for month_index in range(min(12, len(fy_months))):
            col_index = month_index + 2
            if col_index >= len(df.columns):
                continue
            value = safe_float(df.iloc[i, col_index])
            if value > 0:
                working_days[fy_months[month_index]] = value
        break
    return working_days


def _detect_category(row_text: str, current_category: str) -> str:
    if any(keyword in row_text for keyword in WORKER_CATEGORY_KEYWORDS):
        return "worker"
    if any(keyword in row_text for keyword in STAFF_CATEGORY_KEYWORDS):
        return "staff"
    return current_category


def _extract_candidate_codes(cell_value: Any) -> list[str]:
    candidates: list[str] = []
    if cell_value is None:
        return candidates

    text = str(cell_value).strip()
    if not text:
        return candidates

    normalized = normalize_cc_code(text)
    if normalized:
        candidates.append(normalized)

    for token in re.findall(r"\d{4,10}", text):
        candidates.append(token)

    # Keep stable order and remove duplicates.
    seen: set[str] = set()
    ordered: list[str] = []
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        ordered.append(candidate)
    return ordered


def _extract_valid_cc_from_row(row: tuple[Any, ...], valid_cc_codes: set[str]) -> str | None:
    # Scan left-side cells first where CC usually appears.
    for cell in row[:20]:
        candidates = _extract_candidate_codes(cell)
        for candidate in candidates:
            if candidate in valid_cc_codes:
                return candidate
            # Suffix match for 4-8 digit codes (e.g. 1136 matches 1412001136)
            if candidate.isdigit() and 4 <= len(candidate) <= 8:
                str_cand = candidate
                for valid in valid_cc_codes:
                    if str(valid).endswith(str_cand):
                        return valid
    return None


def _parse_ga_headcount_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    fy_months: list[str],
    valid_cc_codes: set[str],
) -> dict[tuple[str, str], dict[str, float]]:
    try:
        worksheet = workbook[sheet_name]
    except Exception:
        return {}

    aggregated: dict[tuple[int, str], dict[str, float]] = defaultdict(
        lambda: {"headcount_all": 0.0, "headcount_staff": 0.0, "headcount_worker": 0.0}
    )

    active_cc: str | None = None
    current_category = "staff"

    try:
        for row in worksheet.iter_rows(values_only=True):
            if not row:
                continue

            cc_from_row = _extract_valid_cc_from_row(row, valid_cc_codes)
            if cc_from_row is not None:
                active_cc = cc_from_row
                current_category = "staff"

            if active_cc is None:
                continue

            row_text = " ".join(str(value) for value in row[:20] if value is not None).lower()
            current_category = _detect_category(row_text, current_category)

            if not any(keyword in row_text for keyword in HEADCOUNT_KEYWORDS):
                continue
            if any(keyword in row_text for keyword in EXCLUDE_HEADCOUNT_ROW_KEYWORDS):
                continue

            for start in range(1, 20):
                sequence = [safe_float(row[start + offset]) for offset in range(12) if start + offset < len(row)]
                if len(sequence) != 12:
                    continue
                if not any(value > 0 for value in sequence):
                    continue
                if not all(value < 5000 for value in sequence):
                    continue

                category = current_category
                if any(keyword in row_text for keyword in WORKER_CATEGORY_KEYWORDS):
                    category = "worker"
                elif any(keyword in row_text for keyword in STAFF_CATEGORY_KEYWORDS):
                    category = "staff"

                for month_index, value in enumerate(sequence):
                    if value <= 0:
                        continue
                    key = (active_cc, fy_months[month_index])
                    aggregated[key]["headcount_all"] += value
                    if category == "worker":
                        aggregated[key]["headcount_worker"] += value
                    else:
                        aggregated[key]["headcount_staff"] += value
                break
    except Exception:
        return {}

    return aggregated


ADMIN_SOURCE_NAME = "ga_admin_allocation"
ADMIN_FORM_ROW_BY_TOKEN = (
    (46, ("gas", "食堂燃料", "nhiên liệu", "nhien lieu"), ()),
    (48, ("手洗い洗剤", "nước rửa tay", "nuoc rua tay", "soap"), ()),
    (51, ("清掃", "làm sạch", "lam sach", "cleaning"), ()),
    (54, ("決起コンパ", "tiệc khuấy động", "tiec khuay dong"), ()),
    (56, ("社内販売", "company sale", "bán nội bộ", "ban noi bo"), ()),
    (58, ("khám sức khỏe khi tuyển dụng", "kham suc khoe khi tuyen dung", "người mới", "nguoi moi", "new hire health"), ()),
    (97, ("sổ tay nhân viên", "so tay nhan vien", "staff notebook", "nhân viên mới", "nhan vien moi"), ("worker", "cong nhan", "công nhân")),
    (98, ("sổ tay công nhân", "so tay cong nhan", "worker notebook", "công nhân mới", "cong nhan moi"), ("staff", "nhan vien", "nhân viên")),
)
ADMIN_BASE_ACCOUNT_BY_ROW = {
    46: 5005056281,
    48: 5005016372,
    51: 5005246286,
    54: 5004086291,
    56: 5004086291,
    58: 5004086291,
    97: 5005246288,
    98: 5005246288,
}
ADMIN_DEFAULT_UNIT_PRICE_BY_ROW = {97: 130000, 98: 40000}


def _admin_match_form_row(text: str) -> int | None:
    normalized = _normalize_text(text)
    for row_index, tokens, exclude_tokens in ADMIN_FORM_ROW_BY_TOKEN:
        if any(_normalize_text(token) in normalized for token in exclude_tokens):
            continue
        if any(_normalize_text(token) in normalized for token in tokens):
            return row_index
    return None


def _admin_header_key(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    if any(token in text for token in ("account", "tài khoản", "tai khoan", "勘定科目")):
        return "account_code"
    if any(token in text for token in ("cc", "costcenter", "cost center", "code phòng", "code phong", "部門")):
        return "cc_code"
    if any(token in text for token in ("form_row", "form row", "row", "dòng", "dong")):
        return "form_row"
    if any(token in text for token in ("period", "month", "tháng", "thang", "月")):
        return "period"
    if any(token in text for token in ("amount", "số tiền", "so tien", "金額")):
        return "amount"
    if any(token in text for token in ("formula", "công thức", "cong thuc")):
        return "formula"
    if any(token in text for token in ("quantity", "qty", "số người", "so nguoi", "人数")):
        return "quantity"
    if any(token in text for token in ("unit", "đơn giá", "don gia", "単価")):
        return "unit_price"
    if any(token in text for token in ("description", "nội dung", "noi dung", "内容")):
        return "description"
    return ""


def _admin_period_from_value(value: Any, fy_months: list[str]) -> str | None:
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month"):
        period = f"{int(value.year)}{int(value.month):02d}"
        return period if period in fy_months else None
    text = _normalize_text(value)
    if not text:
        return None
    compact = text.replace("/", "").replace("-", "")
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", compact)
    if match:
        period = f"{match.group(1)}{match.group(2)}"
        return period if period in fy_months else None
    month_match = re.search(r"(?:thang|tháng|月)\s*(\d{1,2})", text)
    if not month_match:
        month_match = re.fullmatch(r"(\d{1,2})", text)
    if month_match:
        month = int(month_match.group(1))
        for period in fy_months:
            if int(period[-2:]) == month:
                return period
    return None


def _admin_format_number(value: float) -> str:
    return str(int(round(value))) if abs(value - round(value)) < 1e-9 else str(value)


def _admin_insert_record(
    cursor: sqlite3.Cursor,
    *,
    period: str,
    cc_code: str,
    form_row: int,
    account_code: int | None,
    amount_vnd: float,
    description: str,
    formula_expr: str | None = None,
) -> int:
    if amount_vnd <= 0 and not formula_expr:
        return 0
    final_account = int(account_code or 0)
    if final_account <= 0:
        final_account = resolve_account_code_for_source(
            cursor.connection,
            ADMIN_SOURCE_NAME,
            cc_code,
            form_row=form_row,
        )
    if final_account <= 0:
        return 0
    final_description = f"Admin allocation: {description or form_row}"
    if formula_expr:
        final_description = f"{final_description}|formula_expr={formula_expr}"
    cursor.execute(
        """
        INSERT INTO fact_input_data
        (source, period, amount_vnd, cc_code, account_code, form_row, scenario_id, description)
        VALUES (?, ?, ?, ?, ?, ?, 'base', ?)
        """,
        (ADMIN_SOURCE_NAME, period, amount_vnd, cc_code, final_account, form_row, final_description),
    )
    return 1


def _parse_admin_allocation_tables(
    workbook: openpyxl.Workbook,
    cursor: sqlite3.Cursor,
    fy_months: list[str],
    valid_cc_codes: set[str],
) -> int:
    inserted = 0
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        active_header: dict[str, int] | None = None
        month_columns: dict[int, str] = {}
        for row in rows:
            if not row:
                continue
            keys = [_admin_header_key(value) for value in row]
            header_map = {key: index for index, key in enumerate(keys) if key}
            detected_months = {
                index: period
                for index, value in enumerate(row)
                if (period := _admin_period_from_value(value, fy_months)) is not None
            }
            if "cc_code" in header_map and ("form_row" in header_map or detected_months):
                active_header = header_map
                month_columns = detected_months
                continue
            if active_header is None:
                continue

            cc_index = active_header.get("cc_code")
            if cc_index is None or cc_index >= len(row):
                continue
            cc_code = normalize_cc_code(row[cc_index])
            if not cc_code or cc_code not in valid_cc_codes:
                continue

            row_text = " ".join(str(value) for value in row if value is not None)
            form_row = None
            form_row_index = active_header.get("form_row")
            if form_row_index is not None and form_row_index < len(row):
                maybe_row = safe_float(row[form_row_index])
                if maybe_row > 0:
                    form_row = int(maybe_row)
            if form_row is None:
                form_row = _admin_match_form_row(row_text)
            if form_row not in ADMIN_BASE_ACCOUNT_BY_ROW:
                continue

            account_code = None
            account_index = active_header.get("account_code")
            if account_index is not None and account_index < len(row):
                account = safe_float(row[account_index])
                account_code = int(account) if account > 0 else None

            desc_index = active_header.get("description")
            description = str(row[desc_index] or "").strip() if desc_index is not None and desc_index < len(row) else row_text

            if month_columns:
                for month_index, period in month_columns.items():
                    if month_index >= len(row):
                        continue
                    amount = safe_float(row[month_index])
                    inserted += _admin_insert_record(
                        cursor,
                        period=period,
                        cc_code=cc_code,
                        form_row=form_row,
                        account_code=account_code,
                        amount_vnd=amount,
                        description=description,
                    )
                continue

            period_index = active_header.get("period")
            period = _admin_period_from_value(row[period_index], fy_months) if period_index is not None and period_index < len(row) else None
            if period is None:
                continue
            formula_index = active_header.get("formula")
            formula_expr = str(row[formula_index] or "").strip() if formula_index is not None and formula_index < len(row) else ""
            if formula_expr.startswith("="):
                formula_expr = formula_expr[1:]
            amount_index = active_header.get("amount")
            amount = safe_float(row[amount_index]) if amount_index is not None and amount_index < len(row) else 0.0
            if not formula_expr:
                qty_index = active_header.get("quantity")
                unit_index = active_header.get("unit_price")
                qty = safe_float(row[qty_index]) if qty_index is not None and qty_index < len(row) else 0.0
                unit = safe_float(row[unit_index]) if unit_index is not None and unit_index < len(row) else 0.0
                if qty > 0:
                    if unit <= 0:
                        unit = float(ADMIN_DEFAULT_UNIT_PRICE_BY_ROW.get(form_row, 0))
                    if unit > 0:
                        formula_expr = f"{_admin_format_number(qty)}*{_admin_format_number(unit)}"
                        amount = qty * unit
            inserted += _admin_insert_record(
                cursor,
                period=period,
                cc_code=cc_code,
                form_row=form_row,
                account_code=account_code,
                amount_vnd=amount,
                description=description,
                formula_expr=formula_expr or None,
            )
    return inserted


def parse_ga(conn: sqlite3.Connection, source_dir: str | None = None) -> dict[str, int]:
    """Main entry for GA parsing."""
    fy_row = conn.execute("SELECT value FROM sys_params WHERE key='fiscal_year'").fetchone()
    fiscal_year = int(str(fy_row[0]).upper().replace("FY", "").strip()) if fy_row else 2027
    fy_months = get_fy_months(fiscal_year)

    path = _find_ga_file(source_dir, fiscal_year)
    if not path or not os.path.exists(path):
        return {"total": 0, "working_days": 0, "headcount": 0, "admin_allocation": 0}

    cursor = conn.cursor()
    cursor.execute("DELETE FROM fact_input_data WHERE source = 'ga_unit_price'")
    cursor.execute("DELETE FROM fact_input_data WHERE source = ?", (ADMIN_SOURCE_NAME,))
    cursor.execute("DELETE FROM fact_monthly_headcount WHERE source = 'ga'")

    excel_file = pd.ExcelFile(path, engine="openpyxl")

    main_sheet = excel_file.sheet_names[0]
    main_df = pd.read_excel(excel_file, sheet_name=main_sheet, header=None, engine="openpyxl")
    unit_price_records = _parse_ga_main_sheet(main_df, fy_months)
    for record in unit_price_records:
        cursor.execute(
            """
            INSERT INTO fact_input_data
            (source, period, amount_vnd, cc_code, account_code, scenario_id, description)
            VALUES ('ga_unit_price', ?, ?, 0, 0, 'base', ?)
            """,
            (record["period"], record["amount"], f"{record['item']}|{record['driver']}"),
        )

    working_days_written = 0
    calc_sheet_names = _iter_calc_sheet_names(excel_file.sheet_names)
    for sheet_name in calc_sheet_names:
        calc_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine="openpyxl")
        working_days = _parse_working_days_sheet(calc_df, fy_months)
        for period, value in working_days.items():
            cursor.execute(
                """
                INSERT OR REPLACE INTO sys_params (key, value, description)
                VALUES (?, ?, ?)
                """,
                (f"working_days_{period}", str(value), f"Working days for {period}"),
            )
            working_days_written += 1
        if working_days:
            break

    valid_cc_codes = {str(row[0]).strip() for row in conn.execute("SELECT code FROM dim_cost_centers").fetchall() if row[0] is not None}

    aggregated_headcount: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {"headcount_all": 0.0, "headcount_staff": 0.0, "headcount_worker": 0.0}
    )
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        admin_inserted = _parse_admin_allocation_tables(workbook, cursor, fy_months, valid_cc_codes)
        for sheet_name in calc_sheet_names:
            sheet_data = _parse_ga_headcount_sheet(workbook, sheet_name, fy_months, valid_cc_codes)
            for key, values in sheet_data.items():
                aggregated_headcount[key]["headcount_all"] += values["headcount_all"]
                aggregated_headcount[key]["headcount_staff"] += values["headcount_staff"]
                aggregated_headcount[key]["headcount_worker"] += values["headcount_worker"]
    finally:
        workbook.close()

    for (cc_code, period), values in aggregated_headcount.items():
        cursor.execute(
            """
            INSERT INTO fact_monthly_headcount
            (period, cc_code, headcount_all, headcount_staff, headcount_worker, source)
            VALUES (?, ?, ?, ?, ?, 'ga')
            ON CONFLICT(period, cc_code, source) DO UPDATE SET
                headcount_all = excluded.headcount_all,
                headcount_staff = excluded.headcount_staff,
                headcount_worker = excluded.headcount_worker
            """,
            (period, cc_code, values["headcount_all"], values["headcount_staff"], values["headcount_worker"]),
        )

    conn.commit()
    return {
        "total": len(unit_price_records),
        "working_days": working_days_written,
        "headcount": len(aggregated_headcount),
        "admin_allocation": admin_inserted,
    }
