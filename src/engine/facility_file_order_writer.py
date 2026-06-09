"""Explicit Facility file-order preview/export workbook writer.

This module is intentionally separate from HubBuilder. It only writes when callers
explicitly provide an output path, and it never changes the default export flow.
"""

from __future__ import annotations

from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook

from src.engine.facility_file_order_preview import FacilityPreviewItem, preview_facility_file_order
from src.utils import excel_helpers as helpers

VISIBLE_MONTH_START_COL = 6  # F
VISIBLE_MONTH_END_COL = 17  # Q
DESCRIPTION_COL = 19  # S
NOTE_COL = 20  # T
ITEM_ID_COL = 5  # E


def _preview_value(item: FacilityPreviewItem, month_column: int) -> Any:
    if item.confidence != "HIGH":
        return "UNKNOWN"
    month_index = month_column - VISIBLE_MONTH_START_COL
    if 0 <= month_index < len(item.monthly_values):
        source_value = item.monthly_values[month_index]
        if item.formula_policy == "ROUND_USD_BY_B2":
            return f"=ROUND({source_value}*$B$2,0)"
        return source_value
    return None


def _clear_preview_row(worksheet, row_index: int) -> None:
    worksheet.cell(row=row_index, column=ITEM_ID_COL).value = None
    worksheet.cell(row=row_index, column=DESCRIPTION_COL).value = None
    worksheet.cell(row=row_index, column=NOTE_COL).value = None
    for column_index in range(VISIBLE_MONTH_START_COL, VISIBLE_MONTH_END_COL + 1):
        worksheet.cell(row=row_index, column=column_index).value = None


def _write_facility_preview_rows(worksheet, facility_source_path: str | Path, cost_center: str | int, start_row: int) -> None:
    preview = preview_facility_file_order(facility_source_path, cost_center=cost_center, start_row=start_row)
    for item in preview.items:
        row_index = item.planned_row
        _clear_preview_row(worksheet, row_index)
        worksheet.cell(row=row_index, column=ITEM_ID_COL, value=item.item_id)
        worksheet.cell(row=row_index, column=DESCRIPTION_COL, value=item.display_name)
        worksheet.cell(row=row_index, column=NOTE_COL, value=item.note if item.confidence != "HIGH" else item.formula_policy)
        for column_index in range(VISIBLE_MONTH_START_COL, VISIBLE_MONTH_END_COL + 1):
            worksheet.cell(row=row_index, column=column_index, value=_preview_value(item, column_index))

    if preview.blank_row_after is not None:
        _clear_preview_row(worksheet, preview.blank_row_after)


def apply_facility_file_order_to_workbook(
    workbook_path: str | Path,
    facility_source_path: str | Path,
    cost_center: str | int = "1412000040",
    start_row: int = 200,
) -> Path:
    """Apply Facility file-order rows to an explicit existing workbook path."""
    if not workbook_path:
        raise ValueError("workbook_path is required for Facility file-order export")
    workbook_file = Path(workbook_path)
    facility_source = Path(facility_source_path)
    if workbook_file.resolve() == facility_source.resolve():
        raise ValueError("workbook_path must not overwrite the Facility source workbook")

    workbook = load_workbook(workbook_file)
    try:
        worksheet = workbook[helpers.find_hub_sheet_name(workbook)]
        _write_facility_preview_rows(worksheet, facility_source, cost_center, start_row)
        workbook.save(workbook_file)
    finally:
        workbook.close()
    return workbook_file


def write_facility_file_order_preview_workbook(
    template_path: str | Path,
    facility_source_path: str | Path,
    output_path: str | Path,
    cost_center: str | int = "1412000040",
    start_row: int = 200,
) -> Path:
    """Write an explicit Facility file-order preview workbook to ``output_path``."""
    if not output_path:
        raise ValueError("output_path is required for Facility file-order preview writing")

    template = Path(template_path)
    facility_source = Path(facility_source_path)
    output = Path(output_path)
    if template.resolve() == output.resolve():
        raise ValueError("output_path must not overwrite the template workbook")
    if facility_source.resolve() == output.resolve():
        raise ValueError("output_path must not overwrite the Facility source workbook")

    output.parent.mkdir(parents=True, exist_ok=True)
    copy2(template, output)
    return apply_facility_file_order_to_workbook(output, facility_source, cost_center=cost_center, start_row=start_row)
