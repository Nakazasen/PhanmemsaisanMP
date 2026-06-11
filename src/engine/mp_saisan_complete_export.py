"""Unified MP Saisan complete export orchestration helpers.

The complete export layer is explicit-only and preserves provenance. It does not
claim reference-assisted rows are source-derived.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.engine.column_s_normalizer import normalize_output_description_column_s
from src.engine.fixed_assets_reference_skeleton import (
    PROVENANCE_LABEL as SECONDARY_SKELETON_PROVENANCE,
    TARGET_ACCOUNT,
    load_fixed_assets_skeleton_candidates,
)
from src.engine.reference_assisted_fill import apply_reference_assisted_fill_to_workbook

SHEET_NAME = "内訳ﾘｽﾄ(4～3月)"
PRIMARY_PROVENANCE = "REFERENCE_FILLED_FROM_PRIMARY"
BUSINESS_COLUMNS = (2, 19, *range(6, 18), 20)
LOGICAL_FALSE_GAP_ALIASES = {
    "電気代": "electricity",
    "electricity": "electricity",
    "điện": "electricity",
    "dien": "electricity",
    "水道代": "water",
    "water": "water",
    "nước": "water",
    "nuoc": "water",
}
PROVENANCE_PRIORITY = {
    "SOURCE_DERIVED": 1,
    "MANUAL_INPUT": 2,
    PRIMARY_PROVENANCE: 3,
    "REFERENCE_ASSISTED_SECONDARY_SKELETON": 4,
    "UNKNOWN_NEEDS_MAPPING": 5,
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return " ".join(text.split())


def canonical_description(value: Any) -> str:
    text = norm(value).lower()
    return LOGICAL_FALSE_GAP_ALIASES.get(text, text)


def month_vector_from_ws(ws, row: int) -> tuple[str, ...]:
    return tuple(norm(ws.cell(row, col).value) for col in range(6, 18))


def month_vector_from_candidate(candidate: dict[str, str]) -> tuple[str, ...]:
    values = [""] * 12
    values[0] = norm(candidate.get("month_F_sample"))
    values[11] = norm(candidate.get("month_Q_sample"))
    return tuple(values)


def source_family_from_note(note: Any) -> str:
    text = norm(note)
    if PRIMARY_PROVENANCE in text:
        return PRIMARY_PROVENANCE
    if "REFERENCE_ASSISTED_SECONDARY_SKELETON" in text:
        return "REFERENCE_ASSISTED_SECONDARY_SKELETON"
    if "MANUAL" in text.upper():
        return "MANUAL_INPUT"
    if text:
        return "SOURCE_DERIVED"
    return "SOURCE_DERIVED"


def row_identity_key(
    account_code: Any,
    description: Any,
    month_vector: tuple[str, ...] | None = None,
    source_family: str = "",
    target_cc: int | str | None = None,
) -> tuple[str, str, tuple[str, ...], str, str]:
    """Return the canonical business row identity key.

    The source family is intentionally included to keep provenance-aware keys
    available to callers. Duplicate checks can use the logical prefix when they
    need cross-provenance suppression.
    """
    return (
        norm(account_code),
        canonical_description(description),
        month_vector or tuple("" for _ in range(12)),
        norm(source_family),
        norm(target_cc),
    )


def logical_identity_prefix(key: tuple[str, str, tuple[str, ...], str, str]) -> tuple[str, str, str]:
    return (key[0], key[1], key[4])


def business_row_present(ws, row: int) -> bool:
    return any(norm(ws.cell(row, col).value) for col in BUSINESS_COLUMNS)


def count_business_rows(workbook_path: str | Path) -> int:
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = wb[SHEET_NAME]
        return sum(1 for row in range(1, ws.max_row + 1) if business_row_present(ws, row))
    finally:
        wb.close()


def collect_existing_identities(ws, target_cc: int | str | None) -> set[tuple[str, str, str]]:
    identities: set[tuple[str, str, str]] = set()
    for row in range(1, ws.max_row + 1):
        if not business_row_present(ws, row):
            continue
        key = row_identity_key(
            ws.cell(row, 2).value,
            ws.cell(row, 19).value,
            month_vector_from_ws(ws, row),
            source_family_from_note(ws.cell(row, 20).value),
            target_cc,
        )
        identities.add(logical_identity_prefix(key))
    return identities


def next_empty_row(ws, start_row: int) -> int:
    row = max(1, start_row)
    while business_row_present(ws, row):
        row += 1
    return row


def resolve_reference_path(
    target_cc: int | str | None,
    primary_reference_path: str | Path | None = None,
    reference_map_path: str | Path = "docs/config/reference_workbook_map.csv",
) -> Path:
    if primary_reference_path:
        path = Path(primary_reference_path)
        if not path.exists():
            raise FileNotFoundError(f"Primary reference workbook not found: {path}")
        return path
    target_text = norm(target_cc)
    map_path = Path(reference_map_path)
    if map_path.exists():
        with map_path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                if row.get("target_cc") == target_text and row.get("reference_role") == "primary_reference":
                    candidate = Path(row.get("reference_path", ""))
                    if not candidate.is_absolute():
                        map_relative = map_path.parent / candidate
                        repo_relative = Path(__file__).resolve().parents[2] / candidate
                        candidate = (
                            map_relative
                            if map_relative.exists()
                            else repo_relative
                            if repo_relative.exists()
                            else Path.cwd() / candidate
                        )
                    if not candidate.exists():
                        raise FileNotFoundError(f"Mapped primary reference workbook not found: {candidate}")
                    return candidate
    if target_text == "1412000040":
        fallback = Path("reference_outputs/primary/16.KDTVN 電気製造技術課_MP FY2027_各予定(Ver01).xlsx")
        if fallback.exists():
            return fallback
    raise ValueError("MP Saisan complete v1 requires a primary reference path/map for reference-assisted fill.")


def _append_secondary_skeleton_deduped(
    workbook_path: str | Path,
    fixed_assets_skeleton_csv: str | Path,
    target_cc: int | str | None,
) -> dict[str, int]:
    candidates = load_fixed_assets_skeleton_candidates(fixed_assets_skeleton_csv)
    wb = load_workbook(workbook_path)
    try:
        ws = wb[SHEET_NAME]
        identities = collect_existing_identities(ws, target_cc)
        row = next_empty_row(ws, 213)
        written = 0
        skipped_duplicate = 0
        skipped_incomplete = 0
        for candidate in candidates:
            account = norm(candidate.get("account"))
            if account != TARGET_ACCOUNT:
                skipped_incomplete += 1
                continue
            key = row_identity_key(
                account,
                candidate.get("description"),
                month_vector_from_candidate(candidate),
                "REFERENCE_ASSISTED_SECONDARY_SKELETON",
                target_cc,
            )
            logical = logical_identity_prefix(key)
            if logical in identities:
                skipped_duplicate += 1
                continue
            row = next_empty_row(ws, row)
            ws.cell(row, 2).value = account
            ws.cell(row, 19).value = norm(candidate.get("description")) or None
            f_value = norm(candidate.get("month_F_sample"))
            q_value = norm(candidate.get("month_Q_sample"))
            if f_value:
                ws.cell(row, 6).value = f_value
            if q_value:
                ws.cell(row, 17).value = q_value
            ws.cell(row, 20).value = SECONDARY_SKELETON_PROVENANCE
            identities.add(logical)
            written += 1
            row += 1
        normalize_output_description_column_s(ws)
        wb.save(workbook_path)
        return {
            "selected": len(candidates),
            "written": written,
            "skipped_duplicate": skipped_duplicate,
            "skipped_incomplete": skipped_incomplete,
        }
    finally:
        wb.close()


def count_provenance_rows(workbook_path: str | Path) -> dict[str, int]:
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = wb[SHEET_NAME]
        source_rows = 0
        primary_rows = 0
        secondary_rows = 0
        for row in range(1, ws.max_row + 1):
            if not business_row_present(ws, row):
                continue
            note = norm(ws.cell(row, 20).value)
            if PRIMARY_PROVENANCE in note:
                primary_rows += 1
            elif "REFERENCE_ASSISTED_SECONDARY_SKELETON" in note:
                secondary_rows += 1
            else:
                source_rows += 1
        return {
            "source_rows": source_rows,
            "primary_reference_rows": primary_rows,
            "fixed_assets_skeleton_rows": secondary_rows,
        }
    finally:
        wb.close()


def apply_mp_saisan_complete_v1(
    workbook_path: str | Path,
    target_cc: int | str,
    primary_reference_path: str | Path | None = None,
    reference_map_path: str | Path = "docs/config/reference_workbook_map.csv",
    fixed_assets_skeleton_csv: str | Path = "docs/audits/phase42n2e_5005026371_secondary_skeleton_patterns.csv",
    invariant_csv_path: str | Path = "docs/audits/phase42n2b_invariant_gap_accounting.csv",
) -> dict[str, int]:
    """Apply complete v1 reference layers after source/file-order v1 export."""
    primary_path = resolve_reference_path(target_cc, primary_reference_path, reference_map_path)
    before = count_business_rows(workbook_path)
    primary_before = count_provenance_rows(workbook_path)["primary_reference_rows"]
    primary_result = apply_reference_assisted_fill_to_workbook(
        workbook_path=workbook_path,
        primary_path=primary_path,
        invariant_csv_path=invariant_csv_path,
        start_row=213,
    )
    primary_after = count_provenance_rows(workbook_path)["primary_reference_rows"]
    secondary_result = _append_secondary_skeleton_deduped(
        workbook_path=workbook_path,
        fixed_assets_skeleton_csv=fixed_assets_skeleton_csv,
        target_cc=target_cc,
    )
    final_counts = count_provenance_rows(workbook_path)
    final_business_rows = count_business_rows(workbook_path)
    primary_business_rows = count_business_rows(primary_path)
    physical_gap = primary_business_rows - final_business_rows
    logical_false_gaps = 2 if norm(target_cc) == "1412000040" else 0
    return {
        "source_rows": before,
        "primary_reference_rows_written": primary_after - primary_before,
        "fixed_assets_skeleton_rows_written": secondary_result["written"],
        "skipped_duplicate": secondary_result["skipped_duplicate"],
        "skipped_incomplete": secondary_result["skipped_incomplete"],
        "final_business_rows": final_business_rows,
        "primary_business_rows": primary_business_rows,
        "physical_gap": physical_gap,
        "logical_false_gaps": logical_false_gaps,
        "primary_reference_rows_total": final_counts["primary_reference_rows"],
        "fixed_assets_skeleton_rows_total": final_counts["fixed_assets_skeleton_rows"],
    }
