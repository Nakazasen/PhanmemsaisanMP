"""Explicit System Cost file-order writer."""
from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from src.engine.column_s_normalizer import normalize_output_description_column_s
from src.engine.system_cost_preview import preview_system_cost_file_order
from src.utils import excel_helpers as helpers

ITEM_ID_COL=5
MONTH_START_COL=6
DESCRIPTION_COL=19
NOTE_COL=20

def _clear_row(ws, row:int):
    ws.cell(row=row,column=ITEM_ID_COL).value=None
    ws.cell(row=row,column=DESCRIPTION_COL).value=None
    ws.cell(row=row,column=NOTE_COL).value=None
    for c in range(MONTH_START_COL, MONTH_START_COL+12):
        ws.cell(row=row,column=c).value=None

def apply_system_cost_to_workbook(workbook_path, system_source_paths, cost_center: str | int='1412000040', start_row:int=211):
    workbook_file=Path(workbook_path)
    for src in system_source_paths:
        if workbook_file.resolve()==Path(src).resolve():
            raise ValueError('workbook_path must not overwrite a System Cost source workbook')
    preview=preview_system_cost_file_order(system_source_paths, cost_center=cost_center, start_row=start_row)
    wb=load_workbook(workbook_file)
    try:
        ws=wb[helpers.find_hub_sheet_name(wb)]
        item=preview.items[0]
        _clear_row(ws,item.planned_row)
        ws.cell(row=item.planned_row,column=ITEM_ID_COL,value=item.item_id)
        ws.cell(row=item.planned_row,column=DESCRIPTION_COL,value=item.display_name)
        ws.cell(row=item.planned_row,column=NOTE_COL,value=item.formula_policy if item.confidence=='HIGH' else item.note)
        if item.confidence=='HIGH':
            for offset,value in enumerate(item.month_values):
                ws.cell(row=item.planned_row,column=MONTH_START_COL+offset,value=value)
        _clear_row(ws,preview.blank_row_after)
        normalize_output_description_column_s(ws)
        wb.save(workbook_file)
    finally:
        wb.close()
    return workbook_file
