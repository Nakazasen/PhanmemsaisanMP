"""Reference-assisted workbook fill helpers.

These helpers deliberately label copied rows as reference-assisted. They do not
claim source-derived provenance.
"""
from __future__ import annotations

from copy import copy
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.engine.column_s_normalizer import normalize_output_description_column_s

SHEET_NAME = "内訳ﾘｽﾄ(4～3月)"
PROVENANCE_LABEL = "REFERENCE_FILLED_FROM_PRIMARY"
FALSE_GAP_CLASS = "ALREADY_GENERATED_FALSE_GAP"
ALLOWED_REFERENCE_CLASSIFICATIONS = {
    "REFERENCE_FILL_ALLOWED",
    "SCOPED_REFERENCE_FILL_ALLOWED",
}
ALLOWED_SCOPE_VALUES = {
    "1",
    "allowed",
    "reference_fill_allowed",
    "scoped",
    "scoped_allowed",
    "true",
    "yes",
}
BUSINESS_COLUMNS = [2, 5, *range(6, 18), 19, 20]
QUARANTINE_FIELDS = [
    "primary_row",
    "account_code",
    "description",
    "classification",
    "quarantine_reason",
    "scope_reason",
    "nontech_reason",
]


def _norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _business_row_present(ws, row: int) -> bool:
    return any(_norm(ws.cell(row, col).value) for col in (2, 19, *range(6, 18)))


def _month_vector(ws, row: int) -> tuple[str, ...]:
    return tuple(_norm(ws.cell(row, col).value) for col in range(6, 18))


def _existing_identity_keys(ws) -> set[tuple[str, str, tuple[str, ...]]]:
    keys: set[tuple[str, str, tuple[str, ...]]] = set()
    for row in range(1, ws.max_row + 1):
        if not _business_row_present(ws, row):
            continue
        keys.add((_norm(ws.cell(row, 2).value), _norm(ws.cell(row, 19).value), _month_vector(ws, row)))
    return keys


def count_business_rows(workbook_path: str | Path) -> int:
    """Count non-empty business rows on the MP detail sheet."""
    wb = load_workbook(workbook_path, data_only=False)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return 0
        ws = wb[SHEET_NAME]
        return sum(1 for row in range(1, ws.max_row + 1) if _business_row_present(ws, row))
    finally:
        wb.close()


def _load_invariant_rows(invariant_csv_path: str | Path) -> list[dict[str, str]]:
    with open(invariant_csv_path, newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def _scope_flag(row: dict[str, str]) -> str:
    raw = row.get("reference_fill_scope") or row.get("scope_status") or row.get("scope") or ""
    return _norm(raw).lower().replace("-", "_").replace(" ", "_")


def _scope_reason(row: dict[str, str]) -> str:
    return (
        _norm(row.get("reference_fill_reason"))
        or _norm(row.get("scope_reason"))
        or _norm(row.get("nontech_reason"))
        or _norm(row.get("action"))
        or _norm(row.get("new_classification"))
        or "explicit reference-fill scope allow"
    )


def _reference_fill_allowed(row: dict[str, str]) -> bool:
    classification = _norm(row.get("new_classification")).upper()
    return classification in ALLOWED_REFERENCE_CLASSIFICATIONS or _scope_flag(row) in ALLOWED_SCOPE_VALUES


def _row_record(primary_ws, row: dict[str, str], primary_row: int, reason: str = "") -> dict[str, str]:
    return {
        "primary_row": str(primary_row),
        "classification": _norm(row.get("new_classification")),
        "account_code": _norm(primary_ws.cell(primary_row, 2).value) or _norm(row.get("account_code")),
        "description": _norm(primary_ws.cell(primary_row, 19).value) or _norm(row.get("primary_description")),
        "scope_reason": reason,
        "nontech_reason": _norm(row.get("nontech_reason")),
    }


def _default_quarantine_path(workbook_path: str | Path) -> Path:
    workbook_file = Path(workbook_path)
    return workbook_file.with_name(f"{workbook_file.stem}_REFERENCE_FILL_QUARANTINE.csv")


def _write_quarantine_csv(path: str | Path, rows: list[dict[str, str]]) -> None:
    quarantine_path = Path(path)
    quarantine_path.parent.mkdir(parents=True, exist_ok=True)
    with quarantine_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUARANTINE_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in QUARANTINE_FIELDS})


def plan_reference_assisted_rows(
    primary_path: str | Path,
    generated_path: str | Path,
    invariant_csv_path: str | Path,
) -> dict[str, Any]:
    """Classify primary reference rows into scoped fill or quarantine buckets."""
    primary_wb = load_workbook(primary_path, data_only=False)
    generated_wb = load_workbook(generated_path, data_only=False)
    try:
        primary_ws = primary_wb[SHEET_NAME]
        generated_ws = generated_wb[SHEET_NAME]
        existing_keys = _existing_identity_keys(generated_ws)
        selected: list[dict[str, Any]] = []
        quarantined: list[dict[str, str]] = []
        skipped_false_gap = 0
        skipped_generated_match = 0
        skipped_empty_primary = 0
        skipped_existing_identity = 0
        for row in _load_invariant_rows(invariant_csv_path):
            if row.get("new_classification") == FALSE_GAP_CLASS:
                skipped_false_gap += 1
                continue
            if _norm(row.get("generated_match_row")):
                skipped_generated_match += 1
                continue
            primary_row = int(row["primary_row"])
            if not _business_row_present(primary_ws, primary_row):
                skipped_empty_primary += 1
                continue
            identity = (
                _norm(primary_ws.cell(primary_row, 2).value),
                _norm(primary_ws.cell(primary_row, 19).value),
                _month_vector(primary_ws, primary_row),
            )
            if identity in existing_keys:
                skipped_existing_identity += 1
                continue
            if not _reference_fill_allowed(row):
                record = _row_record(primary_ws, row, primary_row)
                record["quarantine_reason"] = "unscoped reference-assisted fill is not allowed"
                quarantined.append(record)
                continue
            reason = _scope_reason(row)
            record = _row_record(primary_ws, row, primary_row, reason)
            selected.append(
                {
                    "primary_row": primary_row,
                    "classification": record["classification"],
                    "account_code": record["account_code"],
                    "description": record["description"],
                    "scope_reason": reason,
                }
            )
            existing_keys.add(identity)
        return {
            "selected": selected,
            "quarantined": quarantined,
            "skipped_false_gap": skipped_false_gap,
            "skipped_generated_match": skipped_generated_match,
            "skipped_empty_primary": skipped_empty_primary,
            "skipped_existing_identity": skipped_existing_identity,
        }
    finally:
        generated_wb.close()
        primary_wb.close()


def identify_reference_assisted_rows(
    primary_path: str | Path,
    generated_path: str | Path,
    invariant_csv_path: str | Path,
) -> list[dict[str, Any]]:
    """Return primary rows eligible for reference-assisted copying.

    Skips false gaps and rows with an invariant generated match. The generated
    workbook parameter is retained for API clarity and future safety checks.
    """
    return plan_reference_assisted_rows(primary_path, generated_path, invariant_csv_path)["selected"]


def _next_safe_row(ws, start_row: int) -> int:
    row = max(start_row, 213)
    while _business_row_present(ws, row):
        row += 1
    return row


def _copy_business_row(primary_ws, target_ws, primary_row: int, target_row: int, scope_reason: str) -> None:
    for col in BUSINESS_COLUMNS:
        source = primary_ws.cell(primary_row, col)
        dest = target_ws.cell(target_row, col)
        dest.value = source.value
        if source.has_style:
            dest._style = copy(source._style)
        if source.number_format:
            dest.number_format = source.number_format
    note_cell = target_ws.cell(target_row, 20)
    label = (
        f"{PROVENANCE_LABEL}; primary_row={primary_row}; scoped-reference-fill; "
        f"reason={scope_reason}; not source-derived"
    )
    if _norm(note_cell.value):
        note_cell.value = f"{note_cell.value} | {label}"
    else:
        note_cell.value = label


def apply_reference_assisted_fill_to_workbook(
    workbook_path: str | Path,
    primary_path: str | Path,
    invariant_csv_path: str | Path,
    start_row: int = 213,
) -> dict[str, int | str]:
    """Append eligible primary rows to workbook with provenance labels."""
    plan = plan_reference_assisted_rows(primary_path, workbook_path, invariant_csv_path)
    selected = plan["selected"]
    quarantined = plan["quarantined"]
    quarantine_path = _default_quarantine_path(workbook_path)
    target_wb = load_workbook(workbook_path)
    primary_wb = load_workbook(primary_path, data_only=False)
    try:
        target_ws = target_wb[SHEET_NAME]
        primary_ws = primary_wb[SHEET_NAME]
        target_row = _next_safe_row(target_ws, start_row)
        written = 0
        for item in selected:
            if target_row < 213:
                target_row = 213
            _copy_business_row(primary_ws, target_ws, item["primary_row"], target_row, item["scope_reason"])
            written += 1
            target_row += 1
        normalize_output_description_column_s(target_ws)
        target_wb.save(workbook_path)
        _write_quarantine_csv(quarantine_path, quarantined)
        return {
            "selected": len(selected),
            "written": written,
            "quarantined": len(quarantined),
            "skipped_false_gap": plan["skipped_false_gap"],
            "skipped_generated_match": plan["skipped_generated_match"],
            "skipped_empty_primary": plan["skipped_empty_primary"],
            "skipped_existing_identity": plan["skipped_existing_identity"],
            "start_row": start_row,
            "quarantine_csv": str(quarantine_path),
        }
    finally:
        primary_wb.close()
        target_wb.close()
