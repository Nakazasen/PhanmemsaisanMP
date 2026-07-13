"""Single-pass, source-ordered export for the MP detail worksheet.

The exporter deliberately does not use legacy destination-row metadata. Workbook
coordinates are discovered from FORM headers (or optional defined names), and
business rows are built from database provenance before anything is written.
"""

from __future__ import annotations

from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
import os
from pathlib import Path
import re
import shutil
import sqlite3
from typing import Iterable, Mapping
import unicodedata

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

from src.utils import excel_helpers as helpers


OUTPUT_AREA_NAME = "MP_OUTPUT_AREA"
OUTPUT_ROW_NAME = "MP_OUTPUT_ROW_TEMPLATE"
AUDIT_SHEET_NAME = "_mp2027_output_audit"
FORMULA_MARKER = "formula_expr="
MISSING_MARKERS = {"missing_separate_count=1", "explicit_zero_count=1"}
STAFFING_LABEL_ALIASES = {
    "fixed_hours_expat": ("定時間（日本社員）", "fixed hours expat"),
    "fixed_hours_local": ("定時間（ベトナム社員）", "fixed hours local"),
    "overtime_hours_expat": ("残業時間（日本社員）", "overtime hours expat"),
    "overtime_hours_local": ("残業時間（ベトナム社員）", "overtime hours local"),
    "headcount_expat": ("出向社員(人)", "expat headcount"),
    "headcount_local": ("ローカル社員(人)", "local headcount"),
}

DEFAULT_GROUP_ORDER = (
    "facility",
    "fixed_assets",
    "it_simulation",
    "ga",
    "birthday",
    "allocation_rules",
    "nnn_paperwork",
)


class DynamicExportError(RuntimeError):
    """Raised when FORM cannot be identified safely or output is ambiguous."""


@dataclass(frozen=True)
class FormLayout:
    sheet_name: str
    header_row: int
    period_row: int
    output_start_row: int
    output_end_row: int
    template_row: int
    account_col: int
    lookup_name_col: int
    lookup_group_col: int
    item_col: int
    month_cols: tuple[int, ...]
    total_col: int
    description_col: int
    wbs_col: int
    cost_center_row: int

    @property
    def managed_cols(self) -> tuple[int, ...]:
        first = min(self.account_col, self.lookup_name_col, self.lookup_group_col, self.item_col)
        last = max(self.wbs_col, self.total_col, *self.month_cols)
        return tuple(range(first, last + 1))


@dataclass
class OutputRecord:
    source_group: str
    item_key: str
    account_code: int
    description: str
    item_order: int
    source_files: tuple[str, ...] = ()
    amounts: dict[str, float] = field(default_factory=lambda: defaultdict(float))
    formula_terms: dict[str, list[str]] = field(default_factory=lambda: defaultdict(list))
    highlighted_periods: set[str] = field(default_factory=set)


def _norm(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return " ".join(text.replace("\n", " ").replace("\u3000", " ").strip().split()).lower()


def _defined_range(workbook, name: str, sheet_name: str) -> tuple[int, int, int, int] | None:
    defined = workbook.defined_names.get(name)
    if defined is None:
        return None
    try:
        destinations = list(defined.destinations)
    except Exception:
        return None
    matching = [coordinate for target_sheet, coordinate in destinations if target_sheet == sheet_name]
    if len(matching) != 1:
        return None
    return range_boundaries(matching[0])


def _find_unique_header(
    worksheet,
    aliases: Iterable[str],
    *,
    max_row: int | None = None,
    max_col: int | None = None,
) -> tuple[int, int]:
    wanted = {_norm(alias) for alias in aliases}
    matches: list[tuple[int, int]] = []
    last_row = min(int(worksheet.max_row or 1), int(max_row or worksheet.max_row or 1))
    for row in range(1, last_row + 1):
        last_col = min(int(worksheet.max_column or 1), int(max_col or worksheet.max_column or 1))
        for col in range(1, last_col + 1):
            if _norm(worksheet.cell(row, col).value) in wanted:
                matches.append((row, col))
    if len(matches) != 1:
        raise DynamicExportError(
            f"FORM phải có đúng một tiêu đề trong {sorted(wanted)}; tìm thấy {len(matches)}."
        )
    return matches[0]


def _find_period_columns(worksheet) -> tuple[int, tuple[int, ...], int]:
    expected = (4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3)
    candidates: list[tuple[int, tuple[int, ...], int]] = []
    for row_cells in worksheet.iter_rows():
        row = row_cells[0].row if row_cells else 0
        values: list[int | None] = []
        for cell in row_cells:
            value = cell.value
            try:
                values.append(int(float(value)) if value not in (None, "") else None)
            except (TypeError, ValueError):
                values.append(None)
        for start in range(0, max(0, len(values) - len(expected) + 1)):
            if tuple(values[start : start + len(expected)]) != expected:
                continue
            month_cols = tuple(range(start + 1, start + 1 + len(expected)))
            total_col = month_cols[-1] + 1
            if _norm(worksheet.cell(row, total_col).value) in {_norm("合計"), "total"}:
                candidates.append((row, month_cols, total_col))
    if len(candidates) != 1:
        raise DynamicExportError(
            f"Không xác định duy nhất được dãy tháng 4→3 và cột tổng; tìm thấy {len(candidates)}."
        )
    return candidates[0]


def _find_business_headers(worksheet, *, wbs_max_col: int) -> dict[str, tuple[int, int]]:
    aliases = {
        "lookup_name": {_norm(value) for value in ("勘定科目名称", "account name", "tên tài khoản")},
        "lookup_group": {_norm(value) for value in ("採算科目", "account group", "nhóm tài khoản")},
        "item": {_norm(value) for value in ("ﾃｰﾏｺｰﾄﾞ", "テーマコード", "item", "mã hạng mục")},
        "description": {_norm(value) for value in ("備考", "description", "mô tả")},
        "wbs": {_norm("WBS")},
    }
    matches: dict[str, list[tuple[int, int]]] = {key: [] for key in aliases}
    for row in worksheet.iter_rows():
        for cell in row:
            normalized = _norm(cell.value)
            if not normalized:
                continue
            for key, wanted in aliases.items():
                if key == "wbs" and cell.column > wbs_max_col:
                    continue
                if normalized in wanted:
                    matches[key].append((cell.row, cell.column))
    invalid = {key: len(value) for key, value in matches.items() if len(value) != 1}
    if invalid:
        raise DynamicExportError(f"Không xác định duy nhất được tiêu đề FORM: {invalid}.")
    return {key: value[0] for key, value in matches.items()}


def _formula_template_candidate(worksheet, row: int, cols: Iterable[int]) -> bool:
    formula_count = sum(
        1
        for col in cols
        if isinstance(worksheet.cell(row, col).value, str)
        and str(worksheet.cell(row, col).value).startswith("=")
    )
    return formula_count >= 2


def _detect_output_start(worksheet, *, after_row: int, account_col: int, formula_cols: tuple[int, ...]) -> int:
    candidates: list[int] = []
    for row in range(after_row + 1, int(worksheet.max_row or 0) + 1):
        if worksheet.cell(row, account_col).value not in (None, ""):
            continue
        if not _formula_template_candidate(worksheet, row, formula_cols):
            continue
        previous = worksheet.cell(row - 1, account_col).value if row > 1 else None
        if previous not in (None, ""):
            candidates.append(row)
    if not candidates:
        raise DynamicExportError("Không tìm thấy dòng mẫu output sạch sau các dòng tài khoản có sẵn.")
    return candidates[0]


def _detect_output_end(
    worksheet,
    start_row: int,
    formula_cols: tuple[int, ...],
    month_cols: tuple[int, ...],
) -> int:
    # FORM summary cells define the managed detail capacity (for example
    # SUM(F29:F1000)).  Reading that formula keeps the boundary in the
    # workbook contract instead of duplicating it in Python.
    referenced_ends: list[int] = []
    for row in range(1, start_row):
        for col in month_cols:
            value = worksheet.cell(row, col).value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            for match in re.finditer(r"\$?[A-Z]+\$?(\d+)\s*:\s*\$?[A-Z]+\$?(\d+)", value.upper()):
                range_start, range_end = int(match.group(1)), int(match.group(2))
                if range_start < start_row <= range_end:
                    referenced_ends.append(range_end)
    if referenced_ends:
        return max(referenced_ends)

    last = start_row - 1
    for row in range(start_row, int(worksheet.max_row or 0) + 1):
        if _formula_template_candidate(worksheet, row, formula_cols):
            last = row
            continue
        if last >= start_row:
            break
    if last < start_row:
        raise DynamicExportError("FORM không có vùng dòng mẫu liên tục cho output.")
    return last


def _audit_output_start(workbook, sheet_name: str) -> int | None:
    if AUDIT_SHEET_NAME not in workbook.sheetnames:
        return None
    audit = workbook[AUDIT_SHEET_NAME]
    rows = [
        int(row[1])
        for row in audit.iter_rows(min_row=2, values_only=True)
        if len(row) > 1 and row[0] == sheet_name and row[1] is not None
    ]
    return min(rows) if rows else None


def resolve_form_layout(workbook, sheet_name: str | None = None) -> FormLayout:
    resolved_sheet = sheet_name or helpers.find_hub_sheet_name(workbook)
    worksheet = workbook[resolved_sheet]
    period_row, month_cols, total_col = _find_period_columns(worksheet)
    headers = _find_business_headers(worksheet, wbs_max_col=total_col + 2)
    header_row, lookup_name_col = headers["lookup_name"]
    group_row, lookup_group_col = headers["lookup_group"]
    item_row, item_col = headers["item"]
    description_row, description_col = headers["description"]
    wbs_row, wbs_col = headers["wbs"]
    if len({header_row, group_row, item_row, description_row, wbs_row}) != 1:
        raise DynamicExportError("Các tiêu đề nghiệp vụ của FORM không nằm trên cùng một dòng.")
    account_col = lookup_name_col - 1
    if account_col <= 0:
        raise DynamicExportError("Không xác định được cột tài khoản đứng trước cột tên tài khoản.")

    area = _defined_range(workbook, OUTPUT_AREA_NAME, resolved_sheet)
    row_template = _defined_range(workbook, OUTPUT_ROW_NAME, resolved_sheet)
    formula_cols = (lookup_name_col, lookup_group_col, total_col)
    if area:
        _, output_start, _, output_end = area
    else:
        output_start = _audit_output_start(workbook, resolved_sheet) or _detect_output_start(
            worksheet,
            after_row=max(period_row, header_row),
            account_col=account_col,
            formula_cols=formula_cols,
        )
        output_end = _detect_output_end(worksheet, output_start, formula_cols, month_cols)
    template_row = row_template[1] if row_template else output_start

    cc_matches: list[int] = []
    for row in range(1, output_start):
        if _norm(worksheet.cell(row, account_col).value) in {_norm("原価センター"), _norm("原価センタ")}:
            cc_matches.append(row + 1)
    if len(cc_matches) != 1:
        raise DynamicExportError("Không xác định duy nhất được ô Cost Center của FORM.")

    return FormLayout(
        sheet_name=resolved_sheet,
        header_row=header_row,
        period_row=period_row,
        output_start_row=output_start,
        output_end_row=output_end,
        template_row=template_row,
        account_col=account_col,
        lookup_name_col=lookup_name_col,
        lookup_group_col=lookup_group_col,
        item_col=item_col,
        month_cols=month_cols,
        total_col=total_col,
        description_col=description_col,
        wbs_col=wbs_col,
        cost_center_row=cc_matches[0],
    )


def _source_group(source: str, description: str) -> str:
    source = str(source or "").strip()
    if source == "facility":
        return "facility"
    if source == "fixed_assets":
        return "fixed_assets"
    if source == "it_sim":
        return "it_simulation"
    if source in {"ga_admin_allocation", "ga_unit_price"}:
        return "ga"
    if source == "birthday_workbook":
        return "birthday"
    if source == "nnn_paperwork":
        return "nnn_paperwork"
    if source.startswith("alloc_") and "driver_type=bus_" in str(description or ""):
        return "ga"
    return "allocation_rules"


def _clean_description(description: str) -> str:
    visible = str(description or "").split("|", 1)[0].strip()
    if visible.lower().startswith("alloc:"):
        visible = visible.split(":", 1)[1].strip()
    return visible


def _item_identity(source_group: str, source: str, account_code: int, description: str) -> tuple[str, str]:
    clean = _clean_description(description)
    if source_group == "fixed_assets":
        parts = str(description or "").split("|")
        key = "|".join(parts[:2]) if len(parts) >= 2 else clean
        return f"{source_group}:{account_code}:{key}", key
    if source_group == "it_simulation":
        return f"{source_group}:{account_code}", clean
    if source_group in {"birthday", "nnn_paperwork"}:
        base = clean.split(":", 1)[0].strip()
        return f"{source_group}:{account_code}", base
    return f"{source_group}:{account_code}:{clean}", clean


def _formula_term(description: str) -> str | None:
    for part in str(description or "").split("|"):
        if part.startswith(FORMULA_MARKER):
            term = part[len(FORMULA_MARKER) :].strip()
            return term[1:] if term.startswith("=") else term
    return None


def _manifest_group_data(manifest_entries: Iterable[Mapping[str, object]]) -> tuple[tuple[str, ...], dict[str, tuple[str, ...]]]:
    ordered: list[str] = []
    files: dict[str, list[str]] = defaultdict(list)
    def order_key(entry: Mapping[str, object]) -> tuple[int, str]:
        try:
            order = int(str(entry.get("order", "")).strip())
        except ValueError:
            order = 2**31 - 1
        return order, str(entry.get("filename", "")).casefold()

    for entry in sorted(list(manifest_entries), key=order_key):
        if str(entry.get("enabled", "1")).strip().lower() in {"0", "false", "no", "n"}:
            continue
        category = str(entry.get("category", "")).strip()
        filename = str(entry.get("filename", "")).strip()
        if category and category not in ordered:
            ordered.append(category)
        if category and filename and filename not in files[category]:
            files[category].append(filename)
    for category in DEFAULT_GROUP_ORDER:
        if category not in ordered:
            ordered.append(category)
    return tuple(ordered), {key: tuple(value) for key, value in files.items()}


def backfill_output_metadata(conn: sqlite3.Connection, manifest_entries: Iterable[Mapping[str, object]]) -> int:
    """Populate provenance for legacy/parser rows without consulting FORM coordinates."""
    _, files_by_group = _manifest_group_data(manifest_entries)
    rows = conn.execute(
        """SELECT id, source, description, account_code, source_group, source_file,
                  source_row, item_key, item_order
           FROM fact_input_data ORDER BY id"""
    ).fetchall()
    updates: list[tuple[object, ...]] = []
    for sequence, row in enumerate(rows, start=1):
        group = str(row["source_group"] or "").strip() or _source_group(row["source"], row["description"])
        item_key, _ = _item_identity(group, str(row["source"] or ""), int(row["account_code"] or 0), str(row["description"] or ""))
        source_files = files_by_group.get(group, ())
        updates.append(
            (
                group,
                str(row["source_file"] or "").strip() or (source_files[0] if source_files else None),
                str(row["item_key"] or "").strip() or item_key,
                int(row["item_order"] or 0) or sequence,
                int(row["id"]),
            )
        )
    conn.executemany(
        """UPDATE fact_input_data
           SET source_group=?, source_file=?, item_key=?, item_order=?
           WHERE id=?""",
        updates,
    )
    conn.commit()
    return len(updates)


def load_output_records(
    conn: sqlite3.Connection,
    cc_code: object,
    manifest_entries: Iterable[Mapping[str, object]],
) -> tuple[tuple[str, ...], list[OutputRecord]]:
    group_order, files_by_group = _manifest_group_data(manifest_entries)
    backfill_output_metadata(conn, manifest_entries)
    rows = conn.execute(
        """SELECT id, source, period, amount_vnd, account_code, description,
                  source_group, source_file, source_row, item_key, item_order
           FROM fact_input_data
           WHERE CAST(cc_code AS TEXT)=? AND account_code > 0
           ORDER BY COALESCE(item_order, source_row, id), id""",
        (str(cc_code).strip(),),
    ).fetchall()
    grouped: dict[tuple[str, str, int], OutputRecord] = {}
    descriptions: dict[tuple[str, str, int], list[str]] = defaultdict(list)
    for row in rows:
        group = str(row["source_group"] or "").strip() or _source_group(row["source"], row["description"])
        account = int(row["account_code"] or 0)
        derived_key, display = _item_identity(group, str(row["source"] or ""), account, str(row["description"] or ""))
        item_key = str(row["item_key"] or "").strip() or derived_key
        if group == "it_simulation":
            item_key = derived_key
        identity = (group, item_key, account)
        source_file = str(row["source_file"] or "").strip()
        record = grouped.get(identity)
        if record is None:
            record = OutputRecord(
                source_group=group,
                item_key=item_key,
                account_code=account,
                description=display,
                item_order=int(row["item_order"] or row["source_row"] or row["id"]),
                source_files=tuple(file for file in (source_file,) if file),
            )
            grouped[identity] = record
        elif source_file and source_file not in record.source_files:
            record.source_files = (*record.source_files, source_file)
        if display and display not in descriptions[identity]:
            descriptions[identity].append(display)
        period = str(row["period"] or "").strip()
        description = str(row["description"] or "")
        term = _formula_term(description)
        if term:
            record.formula_terms[period].append(term)
        else:
            record.amounts[period] += float(row["amount_vnd"] or 0.0)
        if any(marker in description for marker in MISSING_MARKERS):
            record.highlighted_periods.add(period)

    records = list(grouped.values())
    for identity, record in grouped.items():
        if record.source_group == "it_simulation":
            labels = descriptions.get(identity, [])
            record.description = " / ".join(labels) if labels else record.item_key
        configured_files = files_by_group.get(record.source_group, ())
        record.source_files = tuple(dict.fromkeys((*record.source_files, *configured_files)))
    order_index = {group: index for index, group in enumerate(group_order)}
    records.sort(key=lambda row: (order_index.get(row.source_group, len(order_index)), row.item_order, row.item_key))
    return group_order, records


def _translate_formula(value: object, source_row: int, target_row: int, col: int) -> object:
    if not (isinstance(value, str) and value.startswith("=")):
        return value
    try:
        origin = f"{get_column_letter(col)}{source_row}"
        target = f"{get_column_letter(col)}{target_row}"
        return Translator(value, origin=origin).translate_formula(target)
    except Exception:
        return value


def _clear_output(worksheet, layout: FormLayout) -> None:
    for row in range(layout.output_start_row, layout.output_end_row + 1):
        for col in layout.managed_cols:
            worksheet.cell(row, col).value = None
            worksheet.cell(row, col).comment = None


def _prepare_row(worksheet, layout: FormLayout, target_row: int) -> None:
    for col in layout.managed_cols:
        source = worksheet.cell(layout.template_row, col)
        target = worksheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.value = _translate_formula(source.value, layout.template_row, target_row, col)
    for col in range(max(1, layout.account_col - 1), layout.lookup_group_col + 1):
        worksheet.cell(target_row, col).fill = PatternFill(fill_type=None)


def _write_record(worksheet, layout: FormLayout, row_number: int, record: OutputRecord, fiscal_periods: tuple[str, ...]) -> None:
    _prepare_row(worksheet, layout, row_number)
    worksheet.cell(row_number, layout.account_col).value = record.account_code
    worksheet.cell(row_number, layout.item_col).value = None
    worksheet.cell(row_number, layout.description_col).value = record.description
    worksheet.cell(row_number, layout.wbs_col).value = None
    for period, col in zip(fiscal_periods, layout.month_cols):
        terms = list(record.formula_terms.get(period, ()))
        numeric = float(record.amounts.get(period, 0.0) or 0.0)
        if terms and abs(numeric) > 1e-9:
            terms.append(str(int(numeric)) if numeric.is_integer() else str(numeric))
        if terms:
            worksheet.cell(row_number, col).value = "=" + "+".join(term.lstrip("=") for term in terms)
        elif abs(numeric) > 1e-9:
            worksheet.cell(row_number, col).value = numeric
        else:
            worksheet.cell(row_number, col).value = None
        if period in record.highlighted_periods:
            worksheet.cell(row_number, col).fill = PatternFill("solid", fgColor="FFC7CE")
    worksheet.cell(row_number, layout.total_col).value = (
        f"=SUM({get_column_letter(layout.month_cols[0])}{row_number}:"
        f"{get_column_letter(layout.month_cols[-1])}{row_number})"
    )
    provenance = ", ".join(record.source_files) or record.source_group
    worksheet.cell(row_number, layout.description_col).comment = Comment(
        f"source_group={record.source_group}; item_key={record.item_key}; source={provenance}",
        "MP2027",
    )


def _write_audit_sheet(workbook, layout: FormLayout, audit_rows: list[tuple[object, ...]]) -> None:
    if AUDIT_SHEET_NAME in workbook.sheetnames:
        del workbook[AUDIT_SHEET_NAME]
    audit = workbook.create_sheet(AUDIT_SHEET_NAME)
    audit.sheet_state = "veryHidden"
    audit.append(("sheet", "output_row", "source_group", "item_key", "account_code", "source_files"))
    for row in audit_rows:
        audit.append(row)


def _find_staffing_rows(worksheet, layout: FormLayout) -> dict[str, int]:
    wanted = {
        key: {_norm(alias) for alias in aliases}
        for key, aliases in STAFFING_LABEL_ALIASES.items()
    }
    matches: dict[str, list[int]] = {key: [] for key in wanted}
    for row in worksheet.iter_rows(min_row=1, max_row=layout.output_start_row - 1):
        row_values = {_norm(cell.value) for cell in row if cell.value not in (None, "")}
        for key, aliases in wanted.items():
            if row_values & aliases:
                matches[key].append(row[0].row)
    invalid = {key: rows for key, rows in matches.items() if len(rows) != 1}
    if invalid:
        raise DynamicExportError(f"Không xác định duy nhất được các dòng nhân sự/thời gian: {invalid}.")
    return {key: rows[0] for key, rows in matches.items()}


def _write_staffing_time(
    conn: sqlite3.Connection,
    worksheet,
    layout: FormLayout,
    cc_code: object,
    periods: tuple[str, ...],
) -> None:
    placeholders = ",".join("?" for _ in periods)
    headcount_rows = conn.execute(
        f"""SELECT period,headcount_expat,headcount_staff,headcount_worker,headcount_local_total
            FROM fact_monthly_headcount
            WHERE CAST(cc_code AS TEXT)=? AND source='department_plan'
              AND period IN ({placeholders})""",
        (str(cc_code), *periods),
    ).fetchall()
    time_rows = conn.execute(
        f"""SELECT period,fixed_hours_expat,fixed_hours_local,
                   overtime_hours_expat,overtime_hours_local
            FROM fact_headcount_time_source
            WHERE CAST(cc_code AS TEXT)=? AND period IN ({placeholders})""",
        (str(cc_code), *periods),
    ).fetchall()
    headcount = {str(row["period"]): row for row in headcount_rows}
    time = {str(row["period"]): row for row in time_rows}
    missing_headcount = [period for period in periods if period not in headcount]
    missing_time = [period for period in periods if period not in time]
    if missing_headcount or missing_time:
        raise DynamicExportError(
            f"CC {cc_code} thiếu nguồn nhân sự={missing_headcount}, thời gian={missing_time}."
        )
    target_rows = _find_staffing_rows(worksheet, layout)
    for period, month_col in zip(periods, layout.month_cols):
        headcount_row = headcount[period]
        time_row = time[period]
        local_total = headcount_row["headcount_local_total"]
        if local_total is None:
            local_total = float(headcount_row["headcount_staff"] or 0) + float(headcount_row["headcount_worker"] or 0)
        values = {
            "fixed_hours_expat": float(time_row["fixed_hours_expat"] or 0),
            "fixed_hours_local": float(time_row["fixed_hours_local"] or 0),
            "overtime_hours_expat": float(time_row["overtime_hours_expat"] or 0),
            "overtime_hours_local": float(time_row["overtime_hours_local"] or 0),
            "headcount_expat": float(headcount_row["headcount_expat"] or 0),
            "headcount_local": float(local_total or 0),
        }
        for metric, value in values.items():
            worksheet.cell(target_rows[metric], month_col).value = value


def export_dynamic_source_order(
    conn: sqlite3.Connection,
    *,
    fiscal_year: int,
    template_path: str | os.PathLike[str],
    output_path: str | os.PathLike[str],
    cc_code: object,
    manifest_entries: Iterable[Mapping[str, object]],
    sheet_name: str | None = None,
    require_staffing: bool = True,
) -> dict[str, int]:
    """Export one CC in a single pass with one blank row between non-empty groups."""
    template = Path(template_path)
    output = Path(output_path)
    if not template.is_file():
        raise FileNotFoundError(f"Không tìm thấy FORM template: {template}")
    output.parent.mkdir(parents=True, exist_ok=True)
    temp = output.with_name(f"{output.stem}.tmp_export{output.suffix or '.xlsx'}")
    if temp.exists():
        temp.unlink()
    shutil.copy2(template, temp)

    workbook = openpyxl.load_workbook(temp, data_only=False)
    try:
        layout = resolve_form_layout(workbook, sheet_name)
        worksheet = workbook[layout.sheet_name]
        worksheet.cell(layout.cost_center_row, layout.account_col).value = (
            int(str(cc_code)) if str(cc_code).isdigit() else str(cc_code)
        )
        periods = tuple(helpers.get_fy_months(fiscal_year))
        if require_staffing:
            _write_staffing_time(conn, worksheet, layout, cc_code, periods)
        group_order, records = load_output_records(conn, cc_code, manifest_entries)
        by_group: dict[str, list[OutputRecord]] = defaultdict(list)
        for record in records:
            if any(record.formula_terms.values()) or any(abs(value) > 1e-9 for value in record.amounts.values()):
                by_group[record.source_group].append(record)

        _clear_output(worksheet, layout)
        current = layout.output_start_row
        blocks = rows_written = separators = 0
        audit_rows: list[tuple[object, ...]] = []
        for group in group_order:
            group_rows = by_group.get(group, [])
            if not group_rows:
                continue
            if blocks:
                if current > layout.output_end_row:
                    raise DynamicExportError("FORM không đủ dòng cho dòng phân cách giữa các nhóm.")
                separators += 1
                current += 1
            for record in group_rows:
                if current > layout.output_end_row:
                    raise DynamicExportError("FORM không đủ vùng output cho toàn bộ chi phí.")
                _write_record(worksheet, layout, current, record, periods)
                audit_rows.append(
                    (
                        layout.sheet_name,
                        current,
                        record.source_group,
                        record.item_key,
                        record.account_code,
                        ", ".join(record.source_files),
                    )
                )
                current += 1
                rows_written += 1
            blocks += 1
        _write_audit_sheet(workbook, layout, audit_rows)
        workbook.save(temp)
    finally:
        workbook.close()
    os.replace(temp, output)
    return {
        "source_blocks_written": blocks,
        "rows_written": rows_written,
        "blank_rows_written": separators,
        "start_row": layout.output_start_row,
        "end_row": current - 1 if rows_written else layout.output_start_row - 1,
    }
