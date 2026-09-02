"""Preserve user-owned special-cost rows in one CC workbook.

The generated common-cost block is deliberately rebuilt on every run.  This
module keeps the Excel rows below that block out of the generator's ownership
without relying on a shared physical row number across cost centers.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator

from src.utils import excel_helpers as helpers
from src.engine.output_cost_row_ordering import has_saved_cost_row_order, restore_cost_layout


MANUAL_SPECIAL_COST_METADATA_SHEET = "_mp2027_manual_cost_meta"
LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET = "_mp2027_manual_special_cost_meta"
MANUAL_SPECIAL_COST_SCHEMA_VERSION = 1
COMMON_COST_START_ROW = helpers.FORM_SHARED_COST_START_ROW
MANUAL_SPECIAL_COST_SEPARATOR = "CHI PHÍ RIÊNG - NHẬP THỦ CÔNG"
FIRST_COST_COLUMN = 1
LAST_COST_COLUMN = 20  # A:T, including account, visible months and note.


class ManualSpecialCostSectionError(ValueError):
    """Raised when a user-owned Excel section cannot be identified safely."""


@dataclass(frozen=True)
class _CellSnapshot:
    value: object
    font: object
    fill: object
    border: object
    alignment: object
    protection: object
    number_format: str
    comment: object
    hyperlink: object


@dataclass(frozen=True)
class _RowSnapshot:
    source_row: int
    cells: tuple[_CellSnapshot, ...]
    height: float | None
    hidden: bool


@dataclass(frozen=True)
class _ManualSectionSnapshot:
    sheet_name: str
    cc_code: str
    rows: tuple[_RowSnapshot, ...]


def _normalise_cc_code(value: object) -> str:
    return str(value or "").strip()


def _find_hub_sheet(workbook):
    try:
        return workbook[helpers.find_hub_sheet_name(workbook)]
    except ValueError:
        return workbook.active


def _row_has_content(worksheet, row: int) -> bool:
    for column in range(FIRST_COST_COLUMN, LAST_COST_COLUMN + 1):
        cell = worksheet.cell(row, column)
        if cell.value not in (None, "") or cell.comment is not None:
            return True
    return False


def _last_content_row(worksheet, start_row: int) -> int:
    for row in range(int(worksheet.max_row or 0), int(start_row) - 1, -1):
        if _row_has_content(worksheet, row):
            return row
    return int(start_row) - 1


def _metadata_sheet(workbook, *, create: bool = False):
    if MANUAL_SPECIAL_COST_METADATA_SHEET in workbook.sheetnames:
        return workbook[MANUAL_SPECIAL_COST_METADATA_SHEET]
    if not create and LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET in workbook.sheetnames:
        return workbook[LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET]
    if not create:
        return None
    if LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET in workbook.sheetnames:
        legacy = workbook[LEGACY_MANUAL_SPECIAL_COST_METADATA_SHEET]
        sheet = workbook.create_sheet(MANUAL_SPECIAL_COST_METADATA_SHEET)
        sheet.sheet_state = "veryHidden"
        for row in legacy.iter_rows(values_only=True):
            sheet.append(row)
        workbook.remove(legacy)
        return sheet
    sheet = workbook.create_sheet(MANUAL_SPECIAL_COST_METADATA_SHEET)
    sheet.sheet_state = "veryHidden"
    sheet.append((
        "sheet_name",
        "fiscal_year",
        "cc_code",
        "common_end_row",
        "manual_start_row",
        "manual_end_row",
        "schema_version",
    ))
    return sheet


def _read_metadata(workbook, *, cc_code: str) -> tuple[str, int, int] | None:
    sheet = _metadata_sheet(workbook)
    if sheet is None:
        return None
    for row in range(2, int(sheet.max_row or 1) + 1):
        row_cc = _normalise_cc_code(sheet.cell(row, 3).value)
        if row_cc != cc_code:
            continue
        sheet_name = str(sheet.cell(row, 1).value or "").strip()
        common_end = int(sheet.cell(row, 4).value or 0)
        manual_start = int(sheet.cell(row, 5).value or 0)
        manual_end = int(sheet.cell(row, 6).value or 0)
        schema_version = int(sheet.cell(row, 7).value or 0)
        if schema_version != MANUAL_SPECIAL_COST_SCHEMA_VERSION:
            raise ManualSpecialCostSectionError(
                f"Dấu mốc chi phí riêng của CC {cc_code} có phiên bản không hỗ trợ."
            )
        if not sheet_name or manual_start <= common_end or manual_end < manual_start - 1:
            raise ManualSpecialCostSectionError(
                f"Dấu mốc chi phí riêng của CC {cc_code} không hợp lệ."
            )
        return sheet_name, manual_start, manual_end
    return None


def _write_metadata(
    workbook,
    *,
    sheet_name: str,
    cc_code: str,
    common_end_row: int,
    manual_start_row: int,
    manual_end_row: int,
) -> None:
    metadata = _metadata_sheet(workbook, create=True)
    metadata.sheet_state = "veryHidden"
    target_row = None
    for row in range(2, int(metadata.max_row or 1) + 1):
        if _normalise_cc_code(metadata.cell(row, 3).value) == cc_code:
            target_row = row
            break
    if target_row is None:
        target_row = int(metadata.max_row or 1) + 1
    values = (
        sheet_name,
        None,
        cc_code,
        int(common_end_row),
        int(manual_start_row),
        int(manual_end_row),
        MANUAL_SPECIAL_COST_SCHEMA_VERSION,
    )
    for column, value in enumerate(values, start=1):
        metadata.cell(target_row, column).value = value


def has_saved_manual_special_cost_layout(
    workbook_path: str | Path,
    cc_code: object,
) -> bool:
    """Return whether a workbook proves it owns special-cost layout for one CC.

    A file's presence alone is not proof: original forms and common-cost-only
    first-run outputs are unmarked.  Malformed metadata is deliberately still
    raised by ``_read_metadata`` instead of being downgraded to first-run.
    """
    normalized_cc = _normalise_cc_code(cc_code)
    if not normalized_cc:
        return False
    path = Path(workbook_path)
    if not path.is_file():
        return False
    if has_saved_cost_row_order(path, normalized_cc):
        return True
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        return _read_metadata(workbook, cc_code=normalized_cc) is not None
    finally:
        workbook.close()


def _snapshot_manual_section(
    source_workbook_path: Path,
    *,
    cc_code: str,
    legacy_start_row: int | None,
) -> _ManualSectionSnapshot:
    if not source_workbook_path.is_file():
        raise ManualSpecialCostSectionError(
            f"Không tìm thấy workbook nguồn chi phí riêng của CC {cc_code}: {source_workbook_path}"
        )
    workbook = openpyxl.load_workbook(source_workbook_path, data_only=False)
    try:
        record = _read_metadata(workbook, cc_code=cc_code)
        if record is None:
            if legacy_start_row is None:
                raise ManualSpecialCostSectionError(
                    f"Workbook nguồn của CC {cc_code} chưa có dấu mốc chi phí riêng. "
                    "Hãy xác nhận dòng bắt đầu chi phí riêng trước khi chạy."
                )
            worksheet = _find_hub_sheet(workbook)
            start_row = int(legacy_start_row)
            if start_row < COMMON_COST_START_ROW:
                raise ManualSpecialCostSectionError(
                    f"Dòng bắt đầu chi phí riêng của CC {cc_code} không hợp lệ: {start_row}."
                )
            end_row = _last_content_row(worksheet, start_row)
        else:
            sheet_name, start_row, end_row = record
            if sheet_name not in workbook.sheetnames:
                raise ManualSpecialCostSectionError(
                    f"Dấu mốc chi phí riêng của CC {cc_code} trỏ tới sheet không tồn tại: {sheet_name}."
                )
            worksheet = workbook[sheet_name]

        rows: list[_RowSnapshot] = []
        for row in range(start_row, end_row + 1):
            cells = tuple(
                _CellSnapshot(
                    value=worksheet.cell(row, column).value,
                    font=copy(worksheet.cell(row, column).font),
                    fill=copy(worksheet.cell(row, column).fill),
                    border=copy(worksheet.cell(row, column).border),
                    alignment=copy(worksheet.cell(row, column).alignment),
                    protection=copy(worksheet.cell(row, column).protection),
                    number_format=worksheet.cell(row, column).number_format,
                    comment=copy(worksheet.cell(row, column).comment),
                    hyperlink=copy(worksheet.cell(row, column).hyperlink),
                )
                for column in range(FIRST_COST_COLUMN, LAST_COST_COLUMN + 1)
            )
            dimension = worksheet.row_dimensions[row]
            rows.append(_RowSnapshot(row, cells, dimension.height, bool(dimension.hidden)))
        return _ManualSectionSnapshot(worksheet.title, cc_code, tuple(rows))
    finally:
        workbook.close()


def _translate_formula(value: object, *, source_row: int, target_row: int, column: int) -> object:
    if not isinstance(value, str) or not value.startswith("=") or source_row == target_row:
        return value
    origin = f"{openpyxl.utils.get_column_letter(column)}{source_row}"
    target = f"{openpyxl.utils.get_column_letter(column)}{target_row}"
    try:
        return Translator(value, origin=origin).translate_formula(target)
    except Exception:
        return value


def _common_end_row(worksheet) -> int:
    return _last_content_row(worksheet, COMMON_COST_START_ROW)


def _write_separator(worksheet, row: int) -> None:
    for column in range(FIRST_COST_COLUMN, LAST_COST_COLUMN + 1):
        cell = worksheet.cell(row, column)
        cell.value = None
        cell.comment = None
    worksheet.cell(row, 19).value = MANUAL_SPECIAL_COST_SEPARATOR


def preserve_manual_special_cost_section(
    generated_workbook_path: str | Path,
    cc_code: object,
    *,
    source_workbook_path: str | Path | None = None,
    legacy_start_row: int | None = None,
    source_kind: str = "current_fiscal_year",
) -> dict[str, int | str]:
    """Put the source CC's manual section below a newly generated common block.

    ``source_workbook_path`` is opened read-only from the caller's perspective:
    this function never saves it.  With no source, an empty, metadata-marked
    section is created for a first-year workbook.
    """
    output_path = Path(generated_workbook_path)
    normalized_cc = _normalise_cc_code(cc_code)
    if not normalized_cc:
        raise ManualSpecialCostSectionError("Thiếu mã Trung tâm chi phí cho vùng chi phí riêng.")
    if not output_path.is_file():
        raise ManualSpecialCostSectionError(f"Không tìm thấy workbook kết quả: {output_path}")

    if source_workbook_path and has_saved_cost_row_order(source_workbook_path, normalized_cc):
        result = restore_cost_layout(
            output_path,
            normalized_cc,
            source_workbook_path,
            source_kind=source_kind,
        )
        return {
            **result,
            "common_end_row": 0,
            "manual_start_row": 0,
            "manual_end_row": 0,
            "source_workbook": str(source_workbook_path),
        }

    snapshot = None
    if source_workbook_path:
        snapshot = _snapshot_manual_section(
            Path(source_workbook_path),
            cc_code=normalized_cc,
            legacy_start_row=legacy_start_row,
        )

    workbook = openpyxl.load_workbook(output_path, data_only=False)
    try:
        worksheet = _find_hub_sheet(workbook)
        common_end = _common_end_row(worksheet)
        separator_row = common_end + 1
        manual_start = separator_row + 1
        _write_separator(worksheet, separator_row)

        manual_rows = tuple(snapshot.rows) if snapshot is not None else ()
        for offset, source_row in enumerate(manual_rows):
            target_row = manual_start + offset
            for column, source_cell in enumerate(source_row.cells, start=FIRST_COST_COLUMN):
                target = worksheet.cell(target_row, column)
                target.value = (
                    None
                    if source_kind == "previous_fiscal_year" and 6 <= column <= 18
                    else _translate_formula(
                        source_cell.value,
                        source_row=source_row.source_row,
                        target_row=target_row,
                        column=column,
                    )
                )
                target.font = copy(source_cell.font)
                target.fill = copy(source_cell.fill)
                target.border = copy(source_cell.border)
                target.alignment = copy(source_cell.alignment)
                target.protection = copy(source_cell.protection)
                target.number_format = source_cell.number_format
                target.comment = copy(source_cell.comment)
                target._hyperlink = copy(source_cell.hyperlink)
            target_dimension = worksheet.row_dimensions[target_row]
            target_dimension.height = source_row.height
            target_dimension.hidden = source_row.hidden

        manual_end = manual_start + len(manual_rows) - 1
        _write_metadata(
            workbook,
            sheet_name=worksheet.title,
            cc_code=normalized_cc,
            common_end_row=common_end,
            manual_start_row=manual_start,
            manual_end_row=manual_end,
        )
        workbook.save(output_path)
    finally:
        workbook.close()

    return {
        "cc_code": normalized_cc,
        "common_end_row": common_end,
        "manual_start_row": manual_start,
        "manual_end_row": manual_end,
        "manual_rows_preserved": len(manual_rows),
        "new_common_rows": 0,
        "source_workbook": str(source_workbook_path or ""),
    }
