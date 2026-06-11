"""Explicit Admin consumables file-order writer."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.engine.admin_consumables_preview import AdminConsumablePreviewItem, preview_admin_consumables_file_order
from src.engine.column_s_normalizer import normalize_output_description_column_s
from src.utils import excel_helpers as helpers

VISIBLE_MONTH_START_COL = 6
VISIBLE_MONTH_END_COL = 17
ITEM_ID_COL = 5
DESCRIPTION_COL = 19
NOTE_COL = 20


def _clear_preview_row(worksheet, row_index: int) -> None:
    worksheet.cell(row=row_index, column=ITEM_ID_COL).value = None
    worksheet.cell(row=row_index, column=DESCRIPTION_COL).value = None
    worksheet.cell(row=row_index, column=NOTE_COL).value = None
    for column_index in range(VISIBLE_MONTH_START_COL, VISIBLE_MONTH_END_COL + 1):
        worksheet.cell(row=row_index, column=column_index).value = None


def _preview_value(item: AdminConsumablePreviewItem, month_column: int) -> Any:
    apr, mar = item.source_month_values_sample
    if month_column == VISIBLE_MONTH_START_COL:
        return apr if apr not in (None, "") else "UNKNOWN"
    if month_column == VISIBLE_MONTH_END_COL:
        return mar if mar not in (None, "") else "UNKNOWN"
    return None


def apply_admin_consumables_to_workbook(
    workbook_path: str | Path,
    admin_source_path: str | Path,
    allocation_source_path: str | Path | None = None,
    cost_center: str | int = "1412000040",
    start_row: int = 207,
) -> Path:
    """Apply Admin consumables rows to an explicit existing workbook path."""
    workbook_file = Path(workbook_path)
    admin_source = Path(admin_source_path)
    if workbook_file.resolve() == admin_source.resolve():
        raise ValueError("workbook_path must not overwrite the Admin source workbook")
    if allocation_source_path and workbook_file.resolve() == Path(allocation_source_path).resolve():
        raise ValueError("workbook_path must not overwrite the allocation source workbook")

    preview = preview_admin_consumables_file_order(
        admin_source_path=admin_source,
        allocation_source_path=allocation_source_path,
        cost_center=cost_center,
        start_row=start_row,
    )
    workbook = load_workbook(workbook_file)
    try:
        worksheet = workbook[helpers.find_hub_sheet_name(workbook)]
        for item in preview.items:
            _clear_preview_row(worksheet, item.planned_row)
            worksheet.cell(row=item.planned_row, column=ITEM_ID_COL, value=item.item_id)
            worksheet.cell(row=item.planned_row, column=DESCRIPTION_COL, value=item.display_name)
            worksheet.cell(row=item.planned_row, column=NOTE_COL, value=item.formula_policy if item.confidence == "HIGH" else item.note)
            worksheet.cell(row=item.planned_row, column=VISIBLE_MONTH_START_COL, value=_preview_value(item, VISIBLE_MONTH_START_COL))
            worksheet.cell(row=item.planned_row, column=VISIBLE_MONTH_END_COL, value=_preview_value(item, VISIBLE_MONTH_END_COL))
        _clear_preview_row(worksheet, preview.blank_row_after)
        normalize_output_description_column_s(worksheet)
        workbook.save(workbook_file)
    finally:
        workbook.close()
    return workbook_file
