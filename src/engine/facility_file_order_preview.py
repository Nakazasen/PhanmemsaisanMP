"""Read-only Facility file-order preview for MP2027.

The preview scans the Facility workbook and combines source evidence with the
pure output placement planner. It does not save or modify any workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.engine.output_mode import get_group_spec
from src.engine.output_placement import plan_output_groups

FACILITY_SOURCE_SHEETS = (
    "減価償却費（Depreciation）",
    "固定資産金利（Interest）",
    "水道光熱費（Electric & Water）",
    "B&L",
    "E&W",
)

FACILITY_ITEMS = (
    ("building_depreciation", "Khấu hao nhà", "減価償却費（Depreciation）", "ROUND_USD_BY_B2", -1),
    ("land_depreciation", "Khấu hao đất", "減価償却費（Depreciation）", "ROUND_USD_BY_B2", 0),
    ("building_interest", "Lãi nhà", "固定資産金利（Interest）", "ROUND_USD_BY_B2", -1),
    ("land_interest", "Lãi đất", "固定資産金利（Interest）", "ROUND_USD_BY_B2", 0),
    ("electricity", "Điện", "水道光熱費（Electric & Water）", "COPY_VND_MONTHLY", -1),
    ("water", "Nước", "水道光熱費（Electric & Water）", "COPY_VND_MONTHLY", 0),
)


@dataclass(frozen=True)
class FacilityPreviewItem:
    item_id: str
    display_name: str
    source_sheet: str
    source_row: int | None
    source_formula_evidence: str | None
    source_amount_sample_apr: Any
    source_amount_sample_mar: Any
    formula_policy: str
    planned_row: int
    confidence: str
    note: str


@dataclass(frozen=True)
class FacilityFileOrderPreview:
    cost_center: str
    source_path: str
    scanned_sheets: tuple[str, ...]
    planned_start_row: int
    planned_end_row: int
    blank_row_after: int | None
    items: tuple[FacilityPreviewItem, ...]


def _find_cost_center_row(worksheet, cost_center: str) -> int | None:
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if any(str(cell).strip() == cost_center for cell in row if cell is not None):
            return row_index
    return None


def _row_samples(value_ws, formula_ws, row_index: int | None) -> tuple[Any, Any, str | None]:
    if row_index is None or row_index < 1:
        return None, None, None
    apr = value_ws.cell(row=row_index, column=4).value
    mar = value_ws.cell(row=row_index, column=15).value
    formula = formula_ws.cell(row=row_index, column=4).value
    return apr, mar, str(formula) if formula is not None else None


def preview_facility_file_order(
    source_path: str | Path,
    cost_center: str | int = "1412000040",
    start_row: int = 200,
) -> FacilityFileOrderPreview:
    """Build a read-only Facility file-order preview for one cost center."""
    source = Path(source_path)
    cc_key = str(cost_center)
    specs = (get_group_spec("facility"),)
    planned = plan_output_groups(specs, start_row=start_row, item_counts={"facility": 6})[0]
    planned_rows = dict(planned.item_rows)

    value_wb = load_workbook(source, read_only=True, data_only=True)
    formula_wb = load_workbook(source, read_only=True, data_only=False)
    try:
        items: list[FacilityPreviewItem] = []
        for item_id, display_name, sheet_name, formula_policy, row_offset in FACILITY_ITEMS:
            value_ws = value_wb[sheet_name]
            formula_ws = formula_wb[sheet_name]
            cc_row = _find_cost_center_row(value_ws, cc_key)
            source_row = cc_row + row_offset if cc_row is not None else None
            apr, mar, formula = _row_samples(value_ws, formula_ws, source_row)
            confidence = "HIGH" if source_row and apr is not None and mar is not None else "UNKNOWN"
            note = "matched cost-center row with paired building/electric row above" if row_offset < 0 else "matched cost-center row"
            if confidence == "UNKNOWN":
                note = "NEED_SOURCE_HEADER_CLARIFICATION"
            items.append(
                FacilityPreviewItem(
                    item_id=item_id,
                    display_name=display_name,
                    source_sheet=sheet_name,
                    source_row=source_row,
                    source_formula_evidence=formula,
                    source_amount_sample_apr=apr,
                    source_amount_sample_mar=mar,
                    formula_policy=formula_policy,
                    planned_row=int(planned_rows[item_id]) if item_id in planned_rows else start_row + len(items),
                    confidence=confidence,
                    note=note,
                )
            )
        return FacilityFileOrderPreview(
            cost_center=cc_key,
            source_path=str(source),
            scanned_sheets=FACILITY_SOURCE_SHEETS,
            planned_start_row=int(planned.start_row or start_row),
            planned_end_row=int(planned.end_row or start_row),
            blank_row_after=planned.blank_row_after,
            items=tuple(items),
        )
    finally:
        value_wb.close()
        formula_wb.close()


def preview_to_markdown(preview: FacilityFileOrderPreview) -> str:
    """Render a preview as an audit-friendly markdown table."""
    lines = [
        "# Phase 42N1N-A - Facility File-Order Dry Run",
        "",
        "## Scope",
        "",
        "Read-only dry-run only. No Excel output was written and no writer flow was changed.",
        "",
        "## Planned placement",
        "",
        f"- Cost center: `{preview.cost_center}`",
        f"- Facility planned rows: `{preview.planned_start_row}`-`{preview.planned_end_row}`",
        f"- Blank row after Facility group: `{preview.blank_row_after}`",
        "",
        "## Scanned sheets",
        "",
    ]
    lines.extend(f"- `{sheet}`" for sheet in preview.scanned_sheets)
    lines.extend([
        "",
        "## Dry-run items",
        "",
        "| # | item_id | display_name | source_sheet | source_row | Apr sample | Mar sample | formula_policy | planned_row | confidence | note |",
        "|---|---|---|---|---:|---:|---:|---|---:|---|---|",
    ])
    for index, item in enumerate(preview.items, start=1):
        lines.append(
            "| {index} | `{item_id}` | {display_name} | `{source_sheet}` | {source_row} | {apr} | {mar} | `{policy}` | {planned_row} | {confidence} | {note} |".format(
                index=index,
                item_id=item.item_id,
                display_name=item.display_name,
                source_sheet=item.source_sheet,
                source_row=item.source_row or "",
                apr=item.source_amount_sample_apr if item.source_amount_sample_apr is not None else "",
                mar=item.source_amount_sample_mar if item.source_amount_sample_mar is not None else "",
                policy=item.formula_policy,
                planned_row=item.planned_row,
                confidence=item.confidence,
                note=item.note,
            )
        )
    return "\n".join(lines) + "\n"
