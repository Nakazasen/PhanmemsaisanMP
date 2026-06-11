"""Column S description rule for generated MP detail workbooks."""

from __future__ import annotations

from collections.abc import Iterable
from numbers import Number
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.utils import excel_helpers as helpers

DEFAULT_START_ROW = 30
DEFAULT_MONTH_COLUMNS = tuple(range(6, 18))  # F:Q
DEFAULT_DESCRIPTION_COLUMN = 19  # S
DEFAULT_LABEL_COLUMNS = (2, 3, 4, 5)  # B:E


def _text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _numeric_value(value: str) -> float | None:
    text = value.replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _is_obvious_blank_or_zero_formula(value: str) -> bool:
    expression = value[1:].strip() if value.startswith("=") else value.strip()
    if expression in {"", '""', "''"}:
        return True
    numeric = _numeric_value(expression)
    return numeric is not None and abs(numeric) <= 1e-9


def cell_has_month_cost(value: Any) -> bool:
    """Return True when a visible F:Q cell carries a cost signal.

    Formula strings count as cost-bearing because generated workbooks are not
    recalculated by Excel during automated tests or CLI runs.
    """
    if value is None:
        return False
    if isinstance(value, Number):
        return abs(float(value)) > 1e-9
    text = _text(value)
    if not text:
        return False
    if text.startswith("="):
        return not _is_obvious_blank_or_zero_formula(text)
    numeric = _numeric_value(text)
    if numeric is not None:
        return abs(numeric) > 1e-9
    return True


def worksheet_row_has_cost(
    worksheet,
    row_index: int,
    month_columns: Iterable[int] = DEFAULT_MONTH_COLUMNS,
) -> bool:
    return any(cell_has_month_cost(worksheet.cell(row_index, column).value) for column in month_columns)


def _label_from_value(value: Any) -> str:
    text = _text(value)
    if not text or text.startswith("="):
        return ""
    if _numeric_value(text) is not None:
        return ""
    return text


def row_description_fallback(
    worksheet,
    row_index: int,
    label_columns: Iterable[int] = DEFAULT_LABEL_COLUMNS,
) -> str:
    for column in label_columns:
        label = _label_from_value(worksheet.cell(row_index, column).value)
        if label:
            return label
    return ""


def normalize_output_description_column_s(
    worksheet,
    *,
    start_row: int = DEFAULT_START_ROW,
    month_columns: Iterable[int] = DEFAULT_MONTH_COLUMNS,
    description_column: int = DEFAULT_DESCRIPTION_COLUMN,
    label_columns: Iterable[int] = DEFAULT_LABEL_COLUMNS,
) -> dict[str, int]:
    """Clear stale S text on no-cost rows and fill S on cost rows when possible."""
    stats = {
        "cost_rows": 0,
        "cost_rows_preserved_description": 0,
        "cost_rows_filled_description": 0,
        "cost_rows_missing_description": 0,
        "no_cost_rows_cleared_description": 0,
    }
    month_columns = tuple(month_columns)
    label_columns = tuple(label_columns)
    for row_index in range(max(1, int(start_row)), int(worksheet.max_row or 0) + 1):
        description_cell = worksheet.cell(row=row_index, column=description_column)
        description = _text(description_cell.value)
        has_cost = worksheet_row_has_cost(worksheet, row_index, month_columns)
        if not has_cost:
            if description:
                description_cell.value = None
                stats["no_cost_rows_cleared_description"] += 1
            continue

        stats["cost_rows"] += 1
        if description:
            stats["cost_rows_preserved_description"] += 1
            continue
        fallback = row_description_fallback(worksheet, row_index, label_columns)
        if fallback:
            description_cell.value = fallback
            stats["cost_rows_filled_description"] += 1
        else:
            stats["cost_rows_missing_description"] += 1
    return stats


def apply_column_s_rule_to_workbook(
    workbook_path: str | Path,
    *,
    sheet_name: str | None = None,
    start_row: int = DEFAULT_START_ROW,
) -> dict[str, int]:
    workbook_file = Path(workbook_path)
    workbook = load_workbook(workbook_file, data_only=False)
    try:
        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            try:
                worksheet = workbook[helpers.find_hub_sheet_name(workbook)]
            except ValueError:
                worksheet = workbook.active
        stats = normalize_output_description_column_s(worksheet, start_row=start_row)
        workbook.save(workbook_file)
        return stats
    finally:
        workbook.close()
