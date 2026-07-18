"""
MP2027 Manager - Hub Builder.

Writes shared-cost data back into the MP detail sheet while preserving the
original FORM layout and formulas.
"""

from __future__ import annotations

from collections import defaultdict
from copy import copy
import os
import re
import shutil
import sqlite3
from typing import Optional
import unicodedata

import openpyxl
from openpyxl.utils import get_column_letter

from src.engine.account_resolver import AccountResolutionError, resolve_account_code_for_connection
from src.engine.column_s_normalizer import normalize_output_description_column_s
from src.engine.output_mode import OutputGroupSpec, get_default_output_group_specs
from src.parsers.fixed_assets import CATEGORY_SPECS, INTEREST_ACCOUNT
from src.services.headcount_source_policy import HeadcountSourceError, load_canonical_headcount
from src.utils import excel_helpers as helpers
from src.utils.fiscal_periods import fiscal_baseline_period


VISIBLE_MONTH_START_COL = 6   # F
TOTAL_COL = 18                # R
ACCOUNT_COL = 2               # B
DESCRIPTION_COL = 19          # S
WBS_COL = 20                  # T
LOOKUP_NAME_COL = 3           # C
LOOKUP_GROUP_COL = 4          # D
APPEND_TEMPLATE_ROW = 29
MIN_APPEND_START_ROW = 168
APPEND_START_ROW = MIN_APPEND_START_ROW
MIN_APPEND_LAST_ROW = 1000
TEMPLATE_ACCOUNT_CLEAR_START_ROW = 30
APPEND_LEFT_FILL = "CCFFFF"
APPEND_MONTH_FILL = "CCFFFF"
APPEND_NOTE_FILL = "CCFFFF"
MISSING_SEPARATE_COUNT_MARKER = "missing_separate_count=1"
EXPLICIT_ZERO_COUNT_MARKER = "explicit_zero_count=1"
MISSING_SEPARATE_COUNT_FILL = "FFC7CE"
FORM_SOURCE_DRIVER_ROWS = {
    "fixed_hours_expat": 8,
    "fixed_hours_local": 9,
    "overtime_hours_expat": 16,
    "overtime_hours_local": 17,
    "headcount_expat": 24,
    "headcount_local": 25,
}
IT_COMPONENT_ORDER = ("vpn", "mail", "r3", "mes", "plm", "qlik_sense", "vps", "ams")
IT_SYSTEM_ACCOUNT_CODES = {5005246282, 6005146628, 6005146542}
IT_SYSTEM_ACCOUNT_BY_COST_TYPE = {
    "製造": 5005246282,
    "一般": 6005146628,
    "販売": 6005146542,
}
IT_SYSTEM_ROW_TEXT_TOKENS = ("system cost", "kdc", "ｋｄｃ", "システム", "社内システム")
IT_SYSTEM_DEFAULT_ROW = 75
IT_SYSTEM_DESCRIPTION = "System Cost (Mail,VPN,R3, Mes,PLM,VPS,...)"
MONTHLY_HEADCOUNT_FIXED_ROWS = (46, 48, 49, 51)
FIXED_ROW_DESCRIPTIONS = {
    36: "減価償却費（建物）/Khấu hao (Nhà)",
    37: "減価償却費（土地）/Khấu hao (Đất)",
    38: "減価償却費（設備）/Khấu hao (Thiết bị)",
    40: "固定資産金利（建物）/Lãi (Nhà)",
    41: "固定資産金利（土地）/Lãi (Đất)",
    42: "固定資産金利（設備）/Lãi (Thiết bị)",
    44: "電気代/Tiền điện",
    45: "水道代/Tiền nước",
    46: "ガス代/Tiền gas",
    48: "Hand wash",
    49: "Toilet paper",
    51: "cleaning fee",
    57: "定年の健康診断費/Chi phí khám sức khỏe hàng năm",
    58: "採用の健康診断費/Chi phí khám sức khỏe tuyển dụng",
    59: "誕生日会/Tiền sinh nhật",
    97: "新入社員：ノート（スタッフ用）/Người mới: Sổ tay (Dùng cho nhân viên)",
    98: "新入社員：ノート (G7社員用）/Người mới: Sổ tay (Dùng cho công nhân)",
    137: "出向者の書類申請費/Chi phí làm giấy tờ cho người biệt phái",
}
FIXED_ALLOCATION_ROW_MATCHERS = {
    57: {
        "tokens": ("kham suc khoe (cho cnv nam)", "kham suc khoe (cho cnv nu)", "health check"),
        "exclude_tokens": ("tuyen dung", "gpld"),
    },
    59: {
        "tokens": ("sinh nhat", "birthday"),
        "exclude_tokens": (),
    },
    97: {
        "tokens": ("ノート", "note", "notebook"),
        "exclude_tokens": ("g7", "worker", "cong nhan"),
        "account_codes": (5005246288,),
        "driver_types": ("headcount_staff",),
    },
    98: {
        "tokens": ("ノート", "note", "notebook"),
        "exclude_tokens": ("staff", "nhan vien"),
        "account_codes": (5005246288,),
        "driver_types": ("headcount_worker",),
    },
    90: {
        "tokens": ("alloc: ノート", "ノート sổ"),
        "exclude_tokens": (),
        "driver_types": ("__disabled_row_90_notebook_legacy__",),
    },
}
MANAGED_FIXED_ROWS = tuple(
    sorted(
        set(range(38, 91))
        | set(range(93, 110))
        | set(range(111, 153))
    )
)


class ExportIntegrityError(RuntimeError):
    """Raised when an export would create a malformed or empty MP workbook."""


class HubBuilder:
    def __init__(self, conn: sqlite3.Connection, fiscal_year: int = 2027, source_file_by_category: dict[str, str] | None = None):
        self.conn = conn
        self.fiscal_year = fiscal_year
        self.source_file_by_category = dict(source_file_by_category or {})
        self.fy_months = helpers.get_fy_months(fiscal_year)
        self.rule_unit_price_by_source = self._load_rule_unit_price_by_source()
        self.rule_identity_by_source = self._load_rule_identity_by_source()
        try:
            self.canonical_headcount = load_canonical_headcount(conn, fiscal_year)
        except HeadcountSourceError as exc:
            raise ExportIntegrityError(str(exc)) from exc

    def _output_group_specs(self) -> tuple[OutputGroupSpec, ...]:
        """Return canonical output group specs for future row-placement planning."""
        return get_default_output_group_specs()

    def _normalize_text(self, value: object) -> str:
        text = unicodedata.normalize("NFKD", str(value or ""))
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.replace("\n", " ").replace("\u3000", " ").strip().lower()
        return " ".join(text.split())

    def _normalize_visible_fiscal_year_labels(self, workbook) -> int:
        """Replace template FY labels in visible header rows with the selected FY."""
        replacement = f"FY{int(self.fiscal_year)}"
        fiscal_label = re.compile(r"FY\s*20\d{2}(?!\d)", re.IGNORECASE)
        changed = 0
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            for row in worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row)):
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or value.startswith("="):
                        continue
                    normalized = fiscal_label.sub(replacement, value)
                    if normalized != value:
                        cell.value = normalized
                        changed += 1
        return changed

    def _format_number(self, value: float) -> str:
        number = float(value or 0.0)
        if abs(number - round(number)) < 1e-9:
            return str(int(round(number)))
        return f"{number:.6f}".rstrip("0").rstrip(".")

    def _as_int(self, value: object) -> int | None:
        try:
            if value is None or str(value).strip() == "":
                return None
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return None

    def _series_has_output(self, values: dict[str, float]) -> bool:
        return any(abs(float(amount or 0.0)) > 1e-9 for amount in values.values())

    def _formula_series_has_output(
        self,
        terms_by_period: dict[str, list[str]],
        numeric_values: dict[str, float] | None = None,
    ) -> bool:
        if any(terms for terms in terms_by_period.values()):
            return True
        return self._series_has_output(numeric_values or {})

    def _find_it_system_total_row(self, worksheet) -> int:
        candidates: list[tuple[int, int]] = []
        for row_index in range(1, worksheet.max_row + 1):
            account_code = self._as_int(worksheet.cell(row=row_index, column=ACCOUNT_COL).value)
            has_kdc_account = account_code in IT_SYSTEM_ACCOUNT_CODES
            row_text = " ".join(
                self._normalize_text(worksheet.cell(row=row_index, column=column_index).value)
                for column_index in range(1, WBS_COL + 1)
                if worksheet.cell(row=row_index, column=column_index).value is not None
            )
            has_system_text = any(token in row_text for token in IT_SYSTEM_ROW_TEXT_TOKENS)
            if has_kdc_account or has_system_text:
                score = (2 if has_kdc_account else 0) + (3 if has_system_text else 0)
                candidates.append((score, row_index))

        if not candidates:
            if worksheet.max_row >= IT_SYSTEM_DEFAULT_ROW:
                return IT_SYSTEM_DEFAULT_ROW
            raise RuntimeError(
                "Không tìm thấy dòng System Cost trong FORM template. "
                "Hãy dùng docs/MP2027/FORM.xlsx mới nhất hoặc giữ đủ vùng dòng output chuẩn."
            )
        candidates.sort(key=lambda item: (-item[0], item[1]))
        return candidates[0][1]

    def _find_recurring_admin_rows(self, worksheet) -> dict[str, int]:
        specs = {
            "gas": (46, 5005056281, ("ガス代", "食堂燃料", "tien gas", "gas")),
            "handwash": (48, 5005016372, ("手洗い", "nuoc rua tay", "hand wash")),
            "toilet_paper": (49, 5005016372, ("トイレット", "giay ve sinh", "toilet paper")),
            "cleaning": (51, 5005246286, ("清掃", "lam sach", "cleaning fee")),
        }
        resolved: dict[str, int] = {}
        for item_key, (canonical_row, account_code, tokens) in specs.items():
            normalized_tokens = tuple(self._normalize_text(token) for token in tokens)
            exact_matches: list[int] = []
            label_matches_with_blank_account: list[int] = []
            for row_index in range(1, worksheet.max_row + 1):
                row_account = self._as_int(worksheet.cell(row=row_index, column=ACCOUNT_COL).value)
                text = " ".join(
                    self._normalize_text(worksheet.cell(row=row_index, column=column).value)
                    for column in (DESCRIPTION_COL, WBS_COL)
                )
                if not any(token in text for token in normalized_tokens):
                    continue
                if row_account == account_code:
                    exact_matches.append(row_index)
                elif row_account is None:
                    label_matches_with_blank_account.append(row_index)

            matches = exact_matches or label_matches_with_blank_account
            if len(matches) == 1:
                resolved[item_key] = matches[0]
                continue
            if len(matches) > 1:
                raise ExportIntegrityError(
                    f"FORM không xác định duy nhất dòng {item_key} theo mã tài khoản và tên khoản chi phí: "
                    f"tìm thấy nhiều dòng {matches}."
                )

            canonical_identity = " ".join(
                self._normalize_text(worksheet.cell(row=canonical_row, column=column).value)
                for column in (ACCOUNT_COL, DESCRIPTION_COL, WBS_COL)
                if worksheet.cell(row=canonical_row, column=column).value is not None
            )
            if canonical_identity:
                raise ExportIntegrityError(
                    f"FORM không xác định duy nhất dòng {item_key} theo mã tài khoản và tên khoản chi phí: "
                    f"không tìm thấy; dòng chuẩn {canonical_row} đang chứa identity khác."
                )
            resolved[item_key] = canonical_row
        return resolved

    def _resolve_it_system_account_code(self, cc_code: int, fact_account_codes: set[int]) -> int | None:
        valid_fact_accounts = fact_account_codes & IT_SYSTEM_ACCOUNT_CODES
        if len(valid_fact_accounts) == 1:
            return next(iter(valid_fact_accounts))

        row = self.conn.execute(
            "SELECT cost_type FROM dim_cost_centers WHERE code = ? LIMIT 1",
            (str(cc_code),),
        ).fetchone()
        if row:
            cost_type = str(row["cost_type"] or "").strip()
            if cost_type in IT_SYSTEM_ACCOUNT_BY_COST_TYPE:
                return IT_SYSTEM_ACCOUNT_BY_COST_TYPE[cost_type]
        return None

    def _load_rule_unit_price_by_source(self) -> dict[str, float]:
        rows = self.conn.execute("SELECT id, unit_price FROM map_allocation_rules").fetchall()
        return {f"alloc_{int(row['id'])}": float(row["unit_price"] or 0.0) for row in rows}

    def _load_rule_identity_by_source(self) -> dict[str, dict[str, object]]:
        rows = self.conn.execute(
            """
            SELECT id, item_name, account_name, mfg_account, ga_account, sales_account,
                   posting_month, unit_price, unit, driver_type, driver_raw
            FROM map_allocation_rules
            """
        ).fetchall()
        return {
            f"alloc_{int(row['id'])}": {
                "item_name": row["item_name"],
                "account_name": row["account_name"],
                "mfg_account": row["mfg_account"],
                "ga_account": row["ga_account"],
                "sales_account": row["sales_account"],
                "posting_month": row["posting_month"],
                "unit_price": float(row["unit_price"] or 0.0),
                "unit": row["unit"],
                "driver_type": row["driver_type"],
                "driver_raw": row["driver_raw"],
            }
            for row in rows
        }

    def _copy_row_style(self, worksheet, source_row: int, target_row: int) -> None:
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column_index in range(1, WBS_COL + 1):
            source_cell = worksheet.cell(row=source_row, column=column_index)
            target_cell = worksheet.cell(row=target_row, column=column_index)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

    def _write_lookup_formulas(self, worksheet, row_index: int) -> None:
        worksheet.cell(
            row=row_index,
            column=LOOKUP_NAME_COL,
            value=(
                f'=IFERROR(IF(VLOOKUP($B{row_index},勘定科目!$A:$H,'
                f'HLOOKUP($E$5,勘定科目!$F$1:$H$2,2,0),0)="","",'
                f'VLOOKUP($B{row_index},勘定科目!$A:$E,2,0)),"")'
            ),
        )
        worksheet.cell(
            row=row_index,
            column=LOOKUP_GROUP_COL,
            value=f'=IF(C{row_index}="","",VLOOKUP($B{row_index},勘定科目!$A:$E,4,0))',
        )
        worksheet.cell(
            row=row_index,
            column=TOTAL_COL,
            value=f"=SUM(F{row_index}:Q{row_index})",
        )

    def _clear_visible_months(self, worksheet, row_index: int) -> None:
        for offset in range(len(self.fy_months)):
            worksheet.cell(row=row_index, column=VISIBLE_MONTH_START_COL + offset).value = None
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _clear_managed_fixed_row(self, worksheet, row_index: int) -> None:
        self._clear_visible_months(worksheet, row_index)
        worksheet.cell(row=row_index, column=DESCRIPTION_COL).value = None
        worksheet.cell(row=row_index, column=WBS_COL).value = None

    def _write_fixed_description(self, worksheet, row_index: int, description: str | None = None) -> None:
        text = description or FIXED_ROW_DESCRIPTIONS.get(row_index)
        if text:
            worksheet.cell(row=row_index, column=DESCRIPTION_COL, value=text)

    def _prepare_append_row(self, worksheet, row_index: int) -> None:
        self._copy_row_style(worksheet, APPEND_TEMPLATE_ROW, row_index)
        self._write_lookup_formulas(worksheet, row_index)
        for column_index in range(1, WBS_COL + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if VISIBLE_MONTH_START_COL <= column_index <= TOTAL_COL:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=APPEND_MONTH_FILL)
            elif column_index in (DESCRIPTION_COL, WBS_COL):
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=APPEND_NOTE_FILL)
            else:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=APPEND_LEFT_FILL)
        worksheet.cell(row=row_index, column=5).value = None
        worksheet.cell(row=row_index, column=ACCOUNT_COL).value = None
        worksheet.cell(row=row_index, column=DESCRIPTION_COL).value = None
        worksheet.cell(row=row_index, column=WBS_COL).value = None
        self._clear_visible_months(worksheet, row_index)

    def _clear_append_area(self, worksheet, start_row: int) -> None:
        meaningful_columns = (
            ACCOUNT_COL,
            5,
            *range(VISIBLE_MONTH_START_COL, TOTAL_COL + 1),
            DESCRIPTION_COL,
            WBS_COL,
        )
        for row_index in range(start_row, self._append_last_row(worksheet) + 1):
            if any(
                self._cell_has_user_visible_value(worksheet, row_index, column_index)
                for column_index in meaningful_columns
            ):
                for column_index in range(ACCOUNT_COL, WBS_COL + 1):
                    worksheet.cell(row=row_index, column=column_index).value = None

    def _append_last_row(self, worksheet) -> int:
        return max(int(worksheet.max_row or 0), MIN_APPEND_LAST_ROW)

    def _clear_template_business_payload(
        self,
        worksheet,
        start_row: int = TEMPLATE_ACCOUNT_CLEAR_START_ROW,
    ) -> None:
        for row_index in range(max(int(start_row or 1), 1), self._append_last_row(worksheet) + 1):
            for column_index in (ACCOUNT_COL, DESCRIPTION_COL, WBS_COL):
                cell = worksheet.cell(row=row_index, column=column_index)
                if helpers.is_form_template_payload_value(cell):
                    cell.value = None

    def _cell_has_user_visible_value(self, worksheet, row_index: int, column_index: int) -> bool:
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            return False
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return False
            # Template lookup/total formulas are pre-filled in many blank rows.
            # They should not make a row count as occupied.
            if text.startswith("="):
                return False
            return True
        if isinstance(value, (int, float)):
            return abs(float(value)) > 1e-9
        return True

    def _business_row_count(self, worksheet) -> int:
        business_rows = 0
        for row_index in range(TEMPLATE_ACCOUNT_CLEAR_START_ROW, worksheet.max_row + 1):
            account = worksheet.cell(row=row_index, column=ACCOUNT_COL).value
            months = [
                worksheet.cell(row=row_index, column=column_index).value
                for column_index in range(VISIBLE_MONTH_START_COL, TOTAL_COL)
            ]
            if self._cell_has_user_visible_value(worksheet, row_index, ACCOUNT_COL):
                business_rows += 1
                continue
            if any(value not in (None, "") for value in months):
                business_rows += 1
        return business_rows

    def _fact_count_for_cc(self, cc_code: object) -> int:
        row = self.conn.execute(
            "SELECT COUNT(*) AS count FROM fact_input_data WHERE cc_code = ?",
            (str(cc_code),),
        ).fetchone()
        return int(row["count"] if row else 0)

    def _validate_template_workbook(self, workbook, template_path: str) -> str:
        try:
            hub_sheet_name = helpers.find_hub_sheet_name(workbook)
        except ValueError as exc:
            raise ExportIntegrityError(
                f"FORM template không có sheet chi tiết MP đúng định dạng: {template_path}"
            ) from exc

        worksheet = workbook[hub_sheet_name]
        if worksheet.max_row < TEMPLATE_ACCOUNT_CLEAR_START_ROW or worksheet.max_column < DESCRIPTION_COL:
            raise ExportIntegrityError(
                "FORM template sai định dạng hoặc rỗng: "
                f"{template_path}; sheet={hub_sheet_name!r}; "
                f"số dòng={worksheet.max_row}; số cột={worksheet.max_column}"
            )
        return hub_sheet_name

    def _validate_exported_workbook(self, workbook, output_path: str, cc_code: object, fact_count: int) -> int:
        try:
            hub_sheet_name = self._validate_template_workbook(workbook, output_path)
        except ExportIntegrityError as exc:
            raise ExportIntegrityError(f"Exported workbook failed integrity check for CC {cc_code}: {exc}") from exc

        worksheet = workbook[hub_sheet_name]
        business_rows = self._business_row_count(worksheet)
        if fact_count > 0 and business_rows <= 0:
            raise ExportIntegrityError(
                "Exported workbook has no business rows although DB facts exist: "
                f"cc={cc_code}, facts={fact_count}, output={output_path}, "
                f"sheet={hub_sheet_name}, max_row={worksheet.max_row}, max_column={worksheet.max_column}"
            )
        return business_rows

    def _resolve_append_start_row(self, worksheet, requested_start_row: int) -> int:
        """Find the first safe append row for this CC's generated workbook.

        ``requested_start_row`` is a minimum, not a hard-coded destination.
        Different CCs can have different final occupied rows, so append output
        starts after the last row that contains visible business data.
        """
        minimum_start = max(int(requested_start_row or MIN_APPEND_START_ROW), MIN_APPEND_START_ROW)
        meaningful_columns = (
            ACCOUNT_COL,
            5,
            *range(VISIBLE_MONTH_START_COL, TOTAL_COL + 1),
            DESCRIPTION_COL,
            WBS_COL,
        )
        last_occupied = minimum_start - 1
        for row_index in range(1, self._append_last_row(worksheet) + 1):
            if any(
                self._cell_has_user_visible_value(worksheet, row_index, column_index)
                for column_index in meaningful_columns
            ):
                last_occupied = max(last_occupied, row_index)
        return max(minimum_start, last_occupied + 1)

    def _input_rows_for_cc(self, cc_code: object, source: str | None = None) -> list[sqlite3.Row]:
        cc_key = str(cc_code)
        if source is None:
            query = """
                SELECT source, period, description, account_code, amount_vnd, amount_usd
                FROM fact_input_data
                WHERE cc_code = ?
            """
            params = (cc_key,)
        else:
            query = """
                SELECT source, period, description, account_code, amount_vnd, amount_usd
                FROM fact_input_data
                WHERE cc_code = ? AND source = ?
            """
            params = (cc_key, source)
        return self.conn.execute(query, params).fetchall()

    def _month_series(
        self,
        cc_code: int,
        *,
        source: str | None = None,
        description: str | None = None,
        description_like: str | None = None,
        account_code: int | None = None,
        value_column: str = "amount_vnd",
    ) -> dict[str, float]:
        conditions = ["cc_code = ?"]
        params: list[object] = [str(cc_code)]
        if source is not None:
            conditions.append("source = ?")
            params.append(source)
        if description is not None:
            conditions.append("description = ?")
            params.append(description)
        if description_like is not None:
            conditions.append("description LIKE ?")
            params.append(description_like)
        if account_code is not None:
            conditions.append("account_code = ?")
            params.append(int(account_code))

        query = f"""
            SELECT period, SUM(COALESCE({value_column}, 0)) AS amount
            FROM fact_input_data
            WHERE {' AND '.join(conditions)}
            GROUP BY period
        """
        rows = self.conn.execute(query, params).fetchall()
        return {str(row["period"]): float(row["amount"] or 0.0) for row in rows}

    def _ga_unit_price_series(self, match_tokens: tuple[str, ...]) -> dict[str, float]:
        rows = self.conn.execute(
            """
            SELECT period, description, amount_vnd
            FROM fact_input_data
            WHERE source = 'ga_unit_price'
            """
        ).fetchall()
        result: dict[str, float] = {}
        normalized_tokens = tuple(self._normalize_text(token) for token in match_tokens)
        for row in rows:
            description = self._normalize_text(row["description"])
            if not any(token in description for token in normalized_tokens):
                continue
            result[str(row["period"])] = float(row["amount_vnd"] or 0.0)
        return result

    def _write_numeric_series(self, worksheet, row_index: int, values: dict[str, float]) -> None:
        self._clear_visible_months(worksheet, row_index)
        for offset, period in enumerate(self.fy_months):
            amount = float(values.get(period, 0.0))
            formula_value = f"={self._format_number(amount)}" if amount else None
            worksheet.cell(
                row=row_index,
                column=VISIBLE_MONTH_START_COL + offset,
                value=formula_value,
            )
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _write_formula_series(
        self,
        worksheet,
        row_index: int,
        terms_by_period: dict[str, list[str]],
        numeric_values: dict[str, float] | None = None,
        highlight_periods: set[str] | None = None,
    ) -> None:
        self._clear_visible_months(worksheet, row_index)
        numeric_values = numeric_values or {}
        highlight_periods = highlight_periods or set()
        missing_fill = openpyxl.styles.PatternFill("solid", fgColor=MISSING_SEPARATE_COUNT_FILL)
        for offset, period in enumerate(self.fy_months):
            terms = list(terms_by_period.get(period, []))
            numeric_amount = float(numeric_values.get(period, 0.0))
            if numeric_amount:
                terms.append(self._format_number(numeric_amount))
            if not terms:
                continue
            worksheet.cell(
                row=row_index,
                column=VISIBLE_MONTH_START_COL + offset,
                value=f"={' + '.join(terms)}".replace(" + ", "+"),
            )
            if period in highlight_periods:
                worksheet.cell(row=row_index, column=VISIBLE_MONTH_START_COL + offset).fill = missing_fill
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _write_fx_formula_series(self, worksheet, row_index: int, values_usd: dict[str, float]) -> None:
        self._clear_visible_months(worksheet, row_index)
        for offset, period in enumerate(self.fy_months):
            amount_usd = float(values_usd.get(period, 0.0))
            if amount_usd <= 0:
                continue
            worksheet.cell(
                row=row_index,
                column=VISIBLE_MONTH_START_COL + offset,
                value=f"=ROUND({self._format_number(amount_usd)}*$B$2,0)",
            )
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _write_headcount_formula_series(
        self,
        worksheet,
        row_index: int,
        unit_prices: dict[str, float],
        start_headcount_row: int = 24,
        end_headcount_row: int = 25,
    ) -> None:
        self._clear_visible_months(worksheet, row_index)
        for offset, period in enumerate(self.fy_months):
            unit_price = float(unit_prices.get(period, 0.0))
            if unit_price <= 0:
                continue
            column_letter = get_column_letter(VISIBLE_MONTH_START_COL + offset)
            worksheet.cell(
                row=row_index,
                column=VISIBLE_MONTH_START_COL + offset,
                value=(
                    f"=SUM({column_letter}${start_headcount_row}:{column_letter}${end_headcount_row})"
                    f"*{self._format_number(unit_price)}"
                ),
            )
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _write_prev_month_headcount_formula_series(
        self,
        worksheet,
        row_index: int,
        unit_prices: dict[str, float],
        cc_code: object,
        item_label: str = "Chi phí cần số người",
    ) -> None:
        """Write current-month price multiplied by canonical prior-month total."""
        self._clear_visible_months(worksheet, row_index)
        headcount_by_period = self._monthly_headcount_series(cc_code, "headcount_all")
        baseline = fiscal_baseline_period(self.fiscal_year)
        previous_periods = [baseline, *self.fy_months[:-1]]
        for offset, (period, source_period) in enumerate(zip(self.fy_months, previous_periods)):
            if period not in unit_prices:
                continue
            unit_price = float(unit_prices[period] or 0.0)
            headcount = headcount_by_period.get(source_period)
            if headcount is None:
                raise ExportIntegrityError(
                    f"Không thể xuất CC {cc_code}. {item_label} tháng {period[-2:]}/{period[:4]} "
                    f"cần Tổng người kỳ {source_period} nhưng dữ liệu này chưa có theo đúng nguồn quy định."
                )
            worksheet.cell(
                row=row_index,
                column=VISIBLE_MONTH_START_COL + offset,
                value=f"={self._format_number(headcount)}*{self._format_number(unit_price)}",
            )
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _monthly_headcount_series(self, cc_code: object, driver_type: str) -> dict[str, float]:
        result: dict[str, float] = {}
        normalized_cc = str(cc_code).strip()
        for (row_cc, period), row in self.canonical_headcount.items():
            if row_cc != normalized_cc:
                continue
            value = getattr(row, driver_type, row.headcount_all)
            result[period] = float(value)
        return result

    def _write_source_staffing_time_rows(self, worksheet, cc_code: object) -> None:
        """Write complete selected-FY staffing/time source series to FORM F:Q."""
        marks = ",".join("?" for _ in self.fy_months)
        headcount_rows = self.conn.execute(
            f"SELECT period,headcount_expat,headcount_staff,headcount_worker,headcount_local_total "
            f"FROM fact_monthly_headcount WHERE CAST(cc_code AS TEXT)=? "
            f"AND source='department_plan' AND period IN ({marks})",
            (str(cc_code), *self.fy_months),
        ).fetchall()
        time_rows = self.conn.execute(
            f"SELECT period,fixed_hours_expat,fixed_hours_local,overtime_hours_expat,overtime_hours_local FROM fact_headcount_time_source WHERE CAST(cc_code AS TEXT)=? AND period IN ({marks})",
            (str(cc_code), *self.fy_months),
        ).fetchall()
        headcount = {str(row["period"]): row for row in headcount_rows}
        time = {str(row["period"]): row for row in time_rows}
        missing_h = [p for p in self.fy_months if p not in headcount]
        missing_t = [p for p in self.fy_months if p not in time]
        if missing_h or missing_t:
            details = []
            if missing_h: details.append("số người: " + ", ".join(missing_h))
            if missing_t: details.append("thời gian: " + ", ".join(missing_t))
            raise ExportIntegrityError(f"CC {cc_code} thiếu nguồn sự thật FY{self.fiscal_year} ({'; '.join(details)}). Không xuất FORM để tránh dùng dữ liệu sai năm.")
        for offset, period in enumerate(self.fy_months):
            h, t = headcount[period], time[period]
            local_total = h["headcount_local_total"]
            if local_total is None:
                local_total = float(h["headcount_staff"] or 0) + float(h["headcount_worker"] or 0)
            values = {
                "headcount_expat": float(h["headcount_expat"] or 0),
                "headcount_local": float(local_total),
                "fixed_hours_expat": float(t["fixed_hours_expat"] or 0),
                "fixed_hours_local": float(t["fixed_hours_local"] or 0),
                "overtime_hours_expat": float(t["overtime_hours_expat"] or 0),
                "overtime_hours_local": float(t["overtime_hours_local"] or 0),
            }
            for metric, row_index in FORM_SOURCE_DRIVER_ROWS.items():
                worksheet.cell(row=row_index, column=VISIBLE_MONTH_START_COL + offset, value=values[metric])

    def _match_description(self, description: str, tokens: tuple[str, ...], exclude_tokens: tuple[str, ...]) -> bool:
        normalized_description = self._normalize_text(description)
        normalized_tokens = tuple(self._normalize_text(token) for token in tokens)
        normalized_excludes = tuple(self._normalize_text(token) for token in exclude_tokens)
        return any(token in normalized_description for token in normalized_tokens) and not any(
            token in normalized_description for token in normalized_excludes
        )

    def _infer_driver_type_from_description(self, description: str) -> str:
        normalized_description = self._normalize_text(description)
        if any(token in normalized_description for token in ("worker", "cong nhan", "g7")):
            return "headcount_worker"
        if any(token in normalized_description for token in ("staff", "nhan vien")):
            return "headcount_staff"
        return ""

    def _driver_type_for_input_row(self, row: sqlite3.Row) -> str:
        source = str(row["source"] or "")
        identity = self.rule_identity_by_source.get(source) or {}
        driver_type = str(identity.get("driver_type") or "").strip()
        if driver_type:
            return driver_type
        return self._infer_driver_type_from_description(str(row["description"] or ""))

    def _row_matches_allocation_matcher(
        self,
        row: sqlite3.Row,
        *,
        tokens: tuple[str, ...],
        exclude_tokens: tuple[str, ...] = (),
        account_codes: tuple[int, ...] = (),
        driver_types: tuple[str, ...] = (),
    ) -> bool:
        if not self._match_description(str(row["description"] or ""), tokens, exclude_tokens):
            return False

        if account_codes:
            row_account_code = int(row["account_code"] or 0)
            if row_account_code not in {int(code) for code in account_codes}:
                return False

        if driver_types:
            driver_type = self._driver_type_for_input_row(row)
            if driver_type not in set(driver_types):
                return False

        return True

    def _allocation_output_identity(self, row: sqlite3.Row) -> tuple[object, ...]:
        source = str(row["source"] or "")
        identity = self.rule_identity_by_source.get(source) or {}
        unit_price = float(identity.get("unit_price") or self.rule_unit_price_by_source.get(source, 0.0) or 0.0)
        return (
            source,
            str(row["period"]),
            int(row["account_code"] or 0),
            self._normalize_text(row["description"] or ""),
            self._driver_type_for_input_row(row),
            self._format_number(unit_price),
        )

    def _series_from_tokens(
        self,
        cc_code: int,
        *,
        tokens: tuple[str, ...],
        exclude_tokens: tuple[str, ...] = (),
        account_codes: tuple[int, ...] = (),
        driver_types: tuple[str, ...] = (),
        source_prefix: str = "alloc_",
        value_column: str = "amount_vnd",
    ) -> dict[str, float]:
        result: dict[str, float] = defaultdict(float)
        seen_identities: set[tuple[object, ...]] = set()
        for row in self._input_rows_for_cc(cc_code):
            source = str(row["source"] or "")
            if source_prefix and not source.startswith(source_prefix):
                continue
            if not self._row_matches_allocation_matcher(
                row,
                tokens=tokens,
                exclude_tokens=exclude_tokens,
                account_codes=account_codes,
                driver_types=driver_types,
            ):
                continue
            row_identity = self._allocation_output_identity(row)
            if row_identity in seen_identities:
                continue
            seen_identities.add(row_identity)
            result[str(row["period"])] += float(row[value_column] or 0.0)
        return dict(result)

    def _alloc_formula_term_from_row(self, row: sqlite3.Row) -> str | None:
        description = str(row["description"] or "")
        explicit_formula = self._explicit_formula_term_from_description(description)
        if explicit_formula and "business_identity=recruitment_health" in self._normalize_text(description):
            return explicit_formula
        source = str(row["source"] or "")
        if not source.startswith("alloc_"):
            return None
        unit_price = float(self.rule_unit_price_by_source.get(source, 0.0) or 0.0)
        keys = row.keys()
        raw_amount = row["amount_vnd"] if "amount_vnd" in keys else row["amount"]
        amount_vnd = float(raw_amount or 0.0)
        if unit_price <= 0 or amount_vnd <= 0:
            return None
        driver_value = amount_vnd / unit_price
        if abs(driver_value - round(driver_value)) < 1e-9:
            driver_value = round(driver_value)
        return f"{self._format_number(driver_value)}*{self._format_number(unit_price)}"

    def _alloc_formula_series_from_tokens(
        self,
        cc_code: object,
        *,
        tokens: tuple[str, ...],
        exclude_tokens: tuple[str, ...] = (),
        account_codes: tuple[int, ...] = (),
        driver_types: tuple[str, ...] = (),
        source_prefix: str = "alloc_",
    ) -> tuple[dict[str, list[str]], dict[str, float]]:
        terms_by_period: dict[str, list[str]] = defaultdict(list)
        numeric_values: dict[str, float] = defaultdict(float)
        seen_identities: set[tuple[object, ...]] = set()
        for row in self._input_rows_for_cc(cc_code):
            source = str(row["source"] or "")
            if source_prefix and not source.startswith(source_prefix):
                continue
            if not self._row_matches_allocation_matcher(
                row,
                tokens=tokens,
                exclude_tokens=exclude_tokens,
                account_codes=account_codes,
                driver_types=driver_types,
            ):
                continue
            row_identity = self._allocation_output_identity(row)
            if row_identity in seen_identities:
                continue
            seen_identities.add(row_identity)
            term = self._alloc_formula_term_from_row(row)
            if term:
                terms_by_period[str(row["period"])].append(term)
            else:
                numeric_values[str(row["period"])] += float(row["amount_vnd"] or 0.0)
        return dict(terms_by_period), dict(numeric_values)

    def _account_code_from_tokens(
        self,
        cc_code: int,
        *,
        tokens: tuple[str, ...],
        exclude_tokens: tuple[str, ...] = (),
        account_codes: tuple[int, ...] = (),
        driver_types: tuple[str, ...] = (),
        source_prefix: str = "alloc_",
    ) -> int | None:
        matched_account_codes: set[int] = set()
        for row in self._input_rows_for_cc(cc_code):
            source = str(row["source"] or "")
            if source_prefix and not source.startswith(source_prefix):
                continue
            if not self._row_matches_allocation_matcher(
                row,
                tokens=tokens,
                exclude_tokens=exclude_tokens,
                account_codes=account_codes,
                driver_types=driver_types,
            ):
                continue
            code = int(row["account_code"] or 0)
            if code > 0:
                matched_account_codes.add(code)
        if len(matched_account_codes) == 1:
            return next(iter(matched_account_codes))
        return None

    def _fixed_row_for_description(self, description: str) -> int | None:
        normalized = self._normalize_text(description)
        if "business_identity=recruitment_health" in normalized or self._match_description(
            description,
            ("採用の健康診断費", "採用時健診", "kham suc khoe tuyen dung", "kham suc khoe khi tuyen dung"),
            ("hang nam", "dinh ky", "cho cnv nam", "cho cnv nu"),
        ):
            return -1  # Semantically managed inside the health group; never append.
        for row_index, matcher in FIXED_ALLOCATION_ROW_MATCHERS.items():
            if self._match_description(description, matcher["tokens"], matcher["exclude_tokens"]):
                return row_index
        return None

    def _find_recruitment_health_row(self, worksheet) -> int:
        recruitment_tokens = (
            "採用の健康診断費", "採用時健診", "kham suc khoe tuyen dung", "kham suc khoe khi tuyen dung"
        )
        annual_tokens = ("定年の健康診断費", "kham suc khoe hang nam", "health check")
        recruitment_rows: list[int] = []
        annual_rows: list[int] = []
        for row_index in range(1, worksheet.max_row + 1):
            row_text = " ".join(
                self._normalize_text(worksheet.cell(row=row_index, column=column).value)
                for column in (DESCRIPTION_COL, WBS_COL)
                if worksheet.cell(row=row_index, column=column).value is not None
            )
            if any(self._normalize_text(token) in row_text for token in recruitment_tokens):
                recruitment_rows.append(row_index)
            if any(self._normalize_text(token) in row_text for token in annual_tokens) and "tuyen dung" not in row_text:
                annual_rows.append(row_index)
        if len(recruitment_rows) == 1:
            return recruitment_rows[0]
        if len(recruitment_rows) > 1:
            raise ExportIntegrityError("FORM có nhiều dòng khám sức khỏe tuyển dụng; không thể xác định vị trí an toàn.")
        if len(annual_rows) != 1:
            # The distributed FORM may intentionally leave fixed rows 57–58
            # unlabeled while retaining their lookup/total formulas. Use the
            # declared canonical recruitment slot only when both identity
            # cells are empty, so unrelated template content is never replaced.
            canonical_row = 58
            canonical_identity = " ".join(
                self._normalize_text(worksheet.cell(row=canonical_row, column=column).value)
                for column in (ACCOUNT_COL, DESCRIPTION_COL, WBS_COL)
                if worksheet.cell(row=canonical_row, column=column).value is not None
            )
            if not canonical_identity:
                return canonical_row
            raise ExportIntegrityError("Không xác định được nhóm chi phí sức khỏe trong FORM để đặt khám sức khỏe tuyển dụng.")
        target_row = annual_rows[0] + 1
        existing_text = " ".join(
            self._normalize_text(worksheet.cell(row=target_row, column=column).value)
            for column in (DESCRIPTION_COL, WBS_COL)
            if worksheet.cell(row=target_row, column=column).value is not None
        )
        if existing_text and not any(self._normalize_text(token) in existing_text for token in recruitment_tokens):
            worksheet.insert_rows(target_row, 1)
            self._copy_row_style(worksheet, annual_rows[0], target_row)
        return target_row

    def _write_recruitment_health_row(self, worksheet, cc_code: int) -> None:
        tokens = (
            "business_identity=recruitment_health",
            "採用の健康診断費",
            "採用時健診",
            "kham suc khoe tuyen dung",
            "kham suc khoe khi tuyen dung",
        )
        exclude_tokens = ("hang nam", "dinh ky", "cho cnv nam", "cho cnv nu")
        terms_by_period, numeric_values = self._alloc_formula_series_from_tokens(
            cc_code, tokens=tokens, exclude_tokens=exclude_tokens
        )
        if not self._formula_series_has_output(terms_by_period, numeric_values):
            return
        row_index = self._find_recruitment_health_row(worksheet)
        account_code = self._account_code_from_tokens(cc_code, tokens=tokens, exclude_tokens=())
        self._clear_visible_months(worksheet, row_index)
        if account_code:
            worksheet.cell(row=row_index, column=ACCOUNT_COL, value=account_code)
            self._write_lookup_formulas(worksheet, row_index)
        worksheet.cell(row=row_index, column=DESCRIPTION_COL, value="採用の健康診断費/Chi phí khám sức khỏe tuyển dụng")
        worksheet.cell(row=row_index, column=WBS_COL, value="business_identity=recruitment_health; placement=health_group")
        self._write_formula_series(worksheet, row_index, terms_by_period, numeric_values)

    def _load_explicit_form_rows(self, cc_code: int) -> list[dict[str, object]]:
        rows = self.conn.execute(
            """
            SELECT form_row, account_code, description, period, SUM(amount_vnd) AS amount
            FROM fact_input_data
            WHERE cc_code = ?
              AND account_code > 0
              AND form_row IS NOT NULL
            GROUP BY form_row, account_code, description, period
            ORDER BY form_row, account_code, description, period
            """,
            (str(cc_code),),
        ).fetchall()

        grouped: dict[int, dict[str, object]] = {}
        for row in rows:
            row_index = int(row["form_row"] or 0)
            if row_index <= 0:
                continue
            bucket = grouped.setdefault(
                row_index,
                {
                    "form_row": row_index,
                    "account_codes": set(),
                    "descriptions": set(),
                    "months": defaultdict(float),
                    "terms": defaultdict(list),
                    "numeric_values": defaultdict(float),
                    "highlight_periods": set(),
                },
            )
            account_code = int(row["account_code"] or 0)
            if account_code > 0:
                bucket["account_codes"].add(account_code)
            description = str(row["description"] or "").strip()
            if description:
                clean_description = self._strip_explicit_formula_metadata(description)
                if clean_description:
                    bucket["descriptions"].add(clean_description)
            period = str(row["period"])
            amount = float(row["amount"] or 0.0)
            term = self._explicit_formula_term_from_description(description)
            if MISSING_SEPARATE_COUNT_MARKER in description or EXPLICIT_ZERO_COUNT_MARKER in description:
                bucket["highlight_periods"].add(period)
            if term:
                bucket["terms"][period].append(term)
            else:
                bucket["numeric_values"][period] += amount
                bucket["months"][period] += amount

        result: list[dict[str, object]] = []
        for row_index in sorted(grouped):
            bucket = grouped[row_index]
            account_codes = bucket.pop("account_codes")
            descriptions = bucket.pop("descriptions")
            bucket["account_code"] = next(iter(account_codes)) if len(account_codes) == 1 else None
            bucket["description"] = next(iter(descriptions)) if len(descriptions) == 1 else None
            bucket["months"] = dict(bucket["months"])
            bucket["terms"] = dict(bucket["terms"])
            bucket["numeric_values"] = dict(bucket["numeric_values"])
            bucket["highlight_periods"] = set(bucket["highlight_periods"])
            result.append(bucket)
        return result

    def _explicit_formula_term_from_description(self, description: str) -> str | None:
        marker = "formula_expr="
        for part in str(description or "").split("|"):
            if part.startswith(marker):
                formula = part[len(marker):].strip()
                return formula[1:] if formula.startswith("=") else formula
        return None

    def _strip_explicit_formula_metadata(self, description: str) -> str:
        return "|".join(
            part
            for part in str(description or "").split("|")
            if not part.startswith("formula_expr=")
            and part != MISSING_SEPARATE_COUNT_MARKER
            and part != EXPLICIT_ZERO_COUNT_MARKER
        ).strip()

    def _append_output_description(self, grouped_description: str, original_description: str) -> str:
        parts = [part.strip() for part in str(grouped_description or "").split("|") if part.strip()]
        identity = ""
        visible_parts: list[str] = []
        for part in parts:
            if part.startswith("business_identity="):
                identity = part.split("=", 1)[1].strip()
            else:
                visible_parts.append(part)
        visible = "|".join(visible_parts).strip()
        if identity == "new_worker_cup":
            return f"{visible} - công nhân mới"
        if identity == "periodic_cup":
            if MISSING_SEPARATE_COUNT_MARKER in str(original_description or ""):
                return f"{visible} định kỳ - chưa nhập số lượng"
            return f"{visible} định kỳ"
        return grouped_description

    def _parse_it_component_term(self, description: str) -> tuple[str, float, float] | None:
        parts = description.split("|")
        if len(parts) < 5 or parts[0:2] != ["it_sim", "component_term"]:
            return None

        component_key = parts[2]
        quantity = 0.0
        unit_price_usd = 0.0
        for part in parts[3:]:
            if part.startswith("qty="):
                quantity = float(part.split("=", 1)[1] or 0.0)
            elif part.startswith("unit_usd="):
                unit_price_usd = float(part.split("=", 1)[1] or 0.0)
        if quantity <= 0 or unit_price_usd <= 0:
            return None
        return component_key, quantity, unit_price_usd

    def _write_explicit_form_rows(self, worksheet, cc_code: int) -> None:
        for row in self._load_explicit_form_rows(cc_code):
            row_index = int(row["form_row"])
            self._clear_visible_months(worksheet, row_index)
            account_code = row.get("account_code")
            if account_code:
                worksheet.cell(row=row_index, column=ACCOUNT_COL, value=int(account_code))
            existing_description = worksheet.cell(row=row_index, column=DESCRIPTION_COL).value
            if not existing_description:
                worksheet.cell(
                    row=row_index,
                    column=DESCRIPTION_COL,
                    value=FIXED_ROW_DESCRIPTIONS.get(row_index) or row.get("description"),
                )
            if row.get("terms"):
                self._write_formula_series(
                    worksheet,
                    row_index,
                    row["terms"],
                    row["numeric_values"],
                    row.get("highlight_periods"),
                )
            else:
                self._write_numeric_series(worksheet, row_index, row["months"])

    def _write_it_system_total_row(self, worksheet, cc_code: int) -> None:
        rows = self._input_rows_for_cc(cc_code, source="it_sim")
        total_vnd_by_period: dict[str, float] = {}
        component_usd_by_period: dict[str, dict[str, float]] = defaultdict(dict)
        component_terms_by_period: dict[str, dict[str, list[tuple[float, float]]]] = defaultdict(lambda: defaultdict(list))
        account_codes: set[int] = set()

        for row in rows:
            description = str(row["description"] or "")
            period = str(row["period"])
            account_code = int(row["account_code"] or 0)
            if account_code > 0:
                account_codes.add(account_code)

            if description.startswith("it_sim|system_usage_total"):
                total_vnd_by_period[period] = float(row["amount_vnd"] or 0.0)
                continue

            parsed_term = self._parse_it_component_term(description)
            if parsed_term is not None:
                component_key, quantity, unit_price_usd = parsed_term
                component_terms_by_period[period][component_key].append((quantity, unit_price_usd))
                continue

            if description.startswith("it_sim|component|"):
                parts = description.split("|")
                if len(parts) >= 3:
                    component_key = parts[2]
                    component_usd_by_period[period][component_key] = float(row["amount_usd"] or 0.0)

        row_index = getattr(self, "_resolved_it_system_row", None) or self._find_it_system_total_row(worksheet)
        account_code = self._resolve_it_system_account_code(cc_code, account_codes)
        if account_code is None:
            raise RuntimeError(
                f"Không xác định được tài khoản System Cost cho mã bộ phận {cc_code}. "
                "Hãy kiểm tra loại chi phí của mã bộ phận trong master CC."
            )
        worksheet.cell(row=row_index, column=ACCOUNT_COL, value=account_code)
        worksheet.cell(row=row_index, column=DESCRIPTION_COL, value=IT_SYSTEM_DESCRIPTION)

        self._clear_visible_months(worksheet, row_index)
        for offset, period in enumerate(self.fy_months):
            component_terms = component_terms_by_period.get(period, {})
            ordered_terms = []
            for key in IT_COMPONENT_ORDER:
                for quantity, unit_price_usd in component_terms.get(key, []):
                    ordered_terms.append(f"{self._format_number(quantity)}*{self._format_number(unit_price_usd)}")

            component_values = component_usd_by_period.get(period, {})
            ordered_values = [
                float(component_values[key])
                for key in IT_COMPONENT_ORDER
                if float(component_values.get(key, 0.0)) > 0
            ]
            cell = worksheet.cell(row=row_index, column=VISIBLE_MONTH_START_COL + offset)
            if ordered_terms:
                cell.value = f"=ROUND(({'+'.join(ordered_terms)})*$B$2,0)"
                continue
            if ordered_values:
                formula = "+".join(self._format_number(value) for value in ordered_values)
                cell.value = f"=ROUND(({formula})*$B$2,0)"
                continue

            total_amount = float(total_vnd_by_period.get(period, 0.0))
            cell.value = int(round(total_amount)) if total_amount else None
        worksheet.cell(row=row_index, column=TOTAL_COL, value=f"=SUM(F{row_index}:Q{row_index})")

    def _write_fixed_rows_legacy(self, worksheet, cc_code: int) -> None:
        fixed_account_codes = {
            36: 5006016260,
            37: 5006016261,
            38: 5006016244,
            40: 9114120007,
            41: 9114120007,
            42: 9114120007,
            44: 5005066281,
            45: 5005066282,
            46: 5005056281,
            48: 5005016372,
            49: 5005016372,
            51: 5005246286,
            57: 5004086291,
            58: 5004086291,
            59: 5004086291,
            97: 5005246288,
            98: 5005246288,
            137: 5005246286,
        }
        for row_index in MANAGED_FIXED_ROWS:
            self._clear_visible_months(worksheet, row_index)
        for row_index, account_code in fixed_account_codes.items():
            worksheet.cell(row=row_index, column=ACCOUNT_COL, value=account_code)

        self._write_numeric_series(
            worksheet,
            44,
            self._month_series(cc_code, source="facility", description="electric"),
        )
        self._write_numeric_series(
            worksheet,
            45,
            self._month_series(cc_code, source="facility", description="water"),
        )
        self._write_prev_month_headcount_formula_series(
            worksheet,
            46,
            self._ga_unit_price_series(("gas|headcount_per_person", "食堂燃料費")),
            cc_code,
        )
        self._write_prev_month_headcount_formula_series(
            worksheet,
            51,
            self._ga_unit_price_series(("清掃費", "chi phí làm sạch|headcount_per_person")),
            cc_code,
        )
        cleaning_series = self._ga_unit_price_series(("cleaning|headcount_per_person",))
        if cleaning_series:
            self._write_prev_month_headcount_formula_series(worksheet, 51, cleaning_series, cc_code)
        self._write_prev_month_headcount_formula_series(
            worksheet,
            48,
            self._ga_unit_price_series(("手洗い洗剤", "nuoc rua tay|headcount_per_person", "nước rửa tay|headcount_per_person")),
            cc_code,
        )
        self._write_prev_month_headcount_formula_series(
            worksheet,
            49,
            self._ga_unit_price_series(("トイレットペーパー", "giay ve sinh|headcount_per_person", "giấy vệ sinh|headcount_per_person")),
            cc_code,
        )
        self._write_fx_formula_series(
            worksheet,
            36,
            self._month_series(cc_code, source="facility", description="depreciation_building", value_column="amount_usd"),
        )
        self._write_fx_formula_series(
            worksheet,
            37,
            self._month_series(cc_code, source="facility", description="depreciation_land", value_column="amount_usd"),
        )
        self._write_fx_formula_series(
            worksheet,
            38,
            self._month_series(cc_code, source="fixed_assets", description_like="fixed_assets_depr|%", value_column="amount_usd"),
        )
        self._write_fx_formula_series(
            worksheet,
            40,
            self._month_series(cc_code, source="facility", description="interest_building", value_column="amount_usd"),
        )
        self._write_fx_formula_series(
            worksheet,
            41,
            self._month_series(cc_code, source="facility", description="interest_land", value_column="amount_usd"),
        )
        self._write_fx_formula_series(
            worksheet,
            42,
            self._month_series(cc_code, source="fixed_assets", description_like="fixed_assets_interest|%", value_column="amount_usd"),
        )
        self._write_it_system_total_row(worksheet, cc_code)

        for row_index, matcher in FIXED_ALLOCATION_ROW_MATCHERS.items():
            series = self._series_from_tokens(
                cc_code,
                tokens=matcher["tokens"],
                exclude_tokens=matcher["exclude_tokens"],
                account_codes=matcher.get("account_codes", ()),
                driver_types=matcher.get("driver_types", ()),
            )
            if not series:
                continue
            account_code = self._account_code_from_tokens(
                cc_code,
                tokens=matcher["tokens"],
                exclude_tokens=matcher["exclude_tokens"],
                account_codes=matcher.get("account_codes", ()),
                driver_types=matcher.get("driver_types", ()),
            )
            if account_code:
                worksheet.cell(row=row_index, column=ACCOUNT_COL, value=account_code)
            output_description = matcher.get("output_description")
            if output_description:
                worksheet.cell(row=row_index, column=DESCRIPTION_COL, value=output_description)
                worksheet.cell(row=row_index, column=WBS_COL, value=f"allocation_rule_row={row_index}; exact_identity={output_description}")
            terms_by_period, numeric_values = self._alloc_formula_series_from_tokens(
                cc_code,
                tokens=matcher["tokens"],
                exclude_tokens=matcher["exclude_tokens"],
                account_codes=matcher.get("account_codes", ()),
                driver_types=matcher.get("driver_types", ()),
            )
            if terms_by_period:
                self._write_formula_series(worksheet, row_index, terms_by_period, numeric_values)
            else:
                self._write_numeric_series(worksheet, row_index, series)

        self._write_explicit_form_rows(worksheet, cc_code)

    def _write_fixed_rows(self, worksheet, cc_code: int) -> None:
        fixed_account_codes = {
            36: 5006016260,
            37: 5006016261,
            38: 5006016244,
            40: 9114120007,
            41: 9114120007,
            42: 9114120007,
            44: 5005066281,
            45: 5005066282,
            57: 5004086291,
            58: 5004086291,
            59: 5004086291,
            97: 5005246288,
            98: 5005246288,
            137: 5005246286,
        }
        recurring_rows = getattr(self, "_resolved_recurring_rows", None) or self._find_recurring_admin_rows(worksheet)
        for item_key, account_code in {
            "gas": 5005056281,
            "handwash": 5005016372,
            "toilet_paper": 5005016372,
            "cleaning": 5005246286,
        }.items():
            fixed_account_codes[recurring_rows[item_key]] = account_code
        for row_index in MANAGED_FIXED_ROWS:
            self._clear_managed_fixed_row(worksheet, row_index)

        def _set_fixed_row(row_index: int, description: str | None = None) -> None:
            account_code = fixed_account_codes.get(row_index)
            if account_code:
                worksheet.cell(row=row_index, column=ACCOUNT_COL, value=account_code)
                self._write_lookup_formulas(worksheet, row_index)
            self._write_fixed_description(worksheet, row_index, description)

        electric_series = self._month_series(cc_code, source="facility", description="electric")
        if self._series_has_output(electric_series):
            _set_fixed_row(44)
            self._write_numeric_series(worksheet, 44, electric_series)

        water_series = self._month_series(cc_code, source="facility", description="water")
        if self._series_has_output(water_series):
            _set_fixed_row(45)
            self._write_numeric_series(worksheet, 45, water_series)

        gas_series = self._ga_unit_price_series(("gas|headcount_per_person", "食堂燃料費"))
        if self._series_has_output(gas_series):
            _set_fixed_row(recurring_rows["gas"], FIXED_ROW_DESCRIPTIONS[46])
            self._write_prev_month_headcount_formula_series(worksheet, recurring_rows["gas"], gas_series, cc_code, "Tiền gas")

        legacy_cleaning_series = self._ga_unit_price_series(("清掃費", "chi ph\u00ed l\u00e0m s\u1ea1ch|headcount_per_person"))
        if self._series_has_output(legacy_cleaning_series):
            _set_fixed_row(recurring_rows["cleaning"], FIXED_ROW_DESCRIPTIONS[51])
            self._write_prev_month_headcount_formula_series(worksheet, recurring_rows["cleaning"], legacy_cleaning_series, cc_code, "Chi phí làm sạch")

        cleaning_series = self._ga_unit_price_series(("cleaning|headcount_per_person",))
        if self._series_has_output(cleaning_series):
            _set_fixed_row(recurring_rows["cleaning"], FIXED_ROW_DESCRIPTIONS[51])
            self._write_prev_month_headcount_formula_series(worksheet, recurring_rows["cleaning"], cleaning_series, cc_code, "Chi phí làm sạch")

        handwash_series = self._ga_unit_price_series(
            ("手洗い洗剤", "nuoc rua tay|headcount_per_person", "nước rửa tay|headcount_per_person")
        )
        if self._series_has_output(handwash_series):
            _set_fixed_row(recurring_rows["handwash"], FIXED_ROW_DESCRIPTIONS[48])
            self._write_prev_month_headcount_formula_series(worksheet, recurring_rows["handwash"], handwash_series, cc_code, "Nước rửa tay")

        toilet_paper_series = self._ga_unit_price_series(
            ("トイレットペーパー", "giay ve sinh|headcount_per_person", "giấy vệ sinh|headcount_per_person")
        )
        if self._series_has_output(toilet_paper_series):
            _set_fixed_row(recurring_rows["toilet_paper"], FIXED_ROW_DESCRIPTIONS[49])
            self._write_prev_month_headcount_formula_series(worksheet, recurring_rows["toilet_paper"], toilet_paper_series, cc_code, "Giấy vệ sinh")

        building_depr_series = self._month_series(
            cc_code,
            source="facility",
            description="depreciation_building",
            value_column="amount_usd",
        )
        if self._series_has_output(building_depr_series):
            _set_fixed_row(36)
            self._write_fx_formula_series(worksheet, 36, building_depr_series)

        land_depr_series = self._month_series(
            cc_code,
            source="facility",
            description="depreciation_land",
            value_column="amount_usd",
        )
        if self._series_has_output(land_depr_series):
            _set_fixed_row(37)
            self._write_fx_formula_series(worksheet, 37, land_depr_series)

        equipment_depr_series = self._month_series(
            cc_code,
            source="fixed_assets",
            description_like="fixed_assets_depr|%",
            value_column="amount_usd",
        )
        if self._series_has_output(equipment_depr_series):
            _set_fixed_row(38)
            self._write_fx_formula_series(worksheet, 38, equipment_depr_series)

        building_interest_series = self._month_series(
            cc_code,
            source="facility",
            description="interest_building",
            value_column="amount_usd",
        )
        if self._series_has_output(building_interest_series):
            _set_fixed_row(40)
            self._write_fx_formula_series(worksheet, 40, building_interest_series)

        land_interest_series = self._month_series(
            cc_code,
            source="facility",
            description="interest_land",
            value_column="amount_usd",
        )
        if self._series_has_output(land_interest_series):
            _set_fixed_row(41)
            self._write_fx_formula_series(worksheet, 41, land_interest_series)

        equipment_interest_series = self._month_series(
            cc_code,
            source="fixed_assets",
            description_like="fixed_assets_interest|%",
            value_column="amount_usd",
        )
        if self._series_has_output(equipment_interest_series):
            _set_fixed_row(42)
            self._write_fx_formula_series(worksheet, 42, equipment_interest_series)

        self._write_it_system_total_row(worksheet, cc_code)

        for row_index, matcher in FIXED_ALLOCATION_ROW_MATCHERS.items():
            series = self._series_from_tokens(
                cc_code,
                tokens=matcher["tokens"],
                exclude_tokens=matcher["exclude_tokens"],
                account_codes=matcher.get("account_codes", ()),
                driver_types=matcher.get("driver_types", ()),
            )
            if not series:
                continue
            account_code = self._account_code_from_tokens(
                cc_code,
                tokens=matcher["tokens"],
                exclude_tokens=matcher["exclude_tokens"],
                account_codes=matcher.get("account_codes", ()),
                driver_types=matcher.get("driver_types", ()),
            )
            if account_code:
                worksheet.cell(row=row_index, column=ACCOUNT_COL, value=account_code)
                self._write_lookup_formulas(worksheet, row_index)
            self._write_fixed_description(worksheet, row_index)
            output_description = matcher.get("output_description")
            if output_description:
                worksheet.cell(row=row_index, column=DESCRIPTION_COL, value=output_description)
                worksheet.cell(row=row_index, column=WBS_COL, value=f"allocation_rule_row={row_index}; exact_identity={output_description}")
            terms_by_period, numeric_values = self._alloc_formula_series_from_tokens(
                cc_code,
                tokens=matcher["tokens"],
                exclude_tokens=matcher["exclude_tokens"],
                account_codes=matcher.get("account_codes", ()),
                driver_types=matcher.get("driver_types", ()),
            )
            if self._formula_series_has_output(terms_by_period, numeric_values):
                self._write_formula_series(worksheet, row_index, terms_by_period, numeric_values)
            else:
                self._write_numeric_series(worksheet, row_index, series)

        self._write_recruitment_health_row(worksheet, cc_code)
        self._write_explicit_form_rows(worksheet, cc_code)

    def _load_fixed_asset_source_order_rows(self, cc_code: int) -> list[dict[str, object]]:
        """Build source-order values from VND amounts rounded per individual asset.

        The parser applies the authoritative runtime rate and Excel rounding to
        each asset before it reaches this aggregation.  Writing the resulting
        VND value avoids a formula longer than Excel's 8,192-character limit
        for cost centers with hundreds of assets.
        """
        category_order = tuple(CATEGORY_SPECS)
        rows = self.conn.execute(
            """
            SELECT description, account_code, period, amount_vnd, amount_usd
            FROM fact_input_data
            WHERE CAST(cc_code AS TEXT) = ? AND source = 'fixed_assets'
              AND amount_usd IS NOT NULL
            ORDER BY description, period
            """,
            (str(cc_code),),
        ).fetchall()
        grouped: dict[tuple[str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
        explicit_zeroes: dict[tuple[str, str], set[str]] = defaultdict(set)
        for row in rows:
            description = str(row["description"] or "")
            parts = description.split("|", 2)
            if len(parts) < 3:
                continue
            kind_token, category_key = parts[0], parts[1]
            if category_key not in CATEGORY_SPECS:
                continue
            kind = "depreciation" if kind_token == "fixed_assets_depr" else "interest" if kind_token == "fixed_assets_interest" else ""
            if not kind:
                continue
            key = (kind, category_key)
            period = str(row["period"])
            rounded_vnd = int(round(float(row["amount_vnd"] or 0.0)))
            grouped[key][period] += rounded_vnd
            if float(row["amount_usd"] or 0.0) == 0.0:
                explicit_zeroes[key].add(period)

        payload: list[dict[str, object]] = []
        for kind in ("depreciation", "interest"):
            for category_key in category_order:
                months = dict(grouped.get((kind, category_key), {}))
                if not any(amount != 0 for amount in months.values()):
                    continue
                spec = CATEGORY_SPECS[category_key]
                account_code = int(spec["depreciation_account"] if kind == "depreciation" else INTEREST_ACCOUNT)
                label = str(spec["label"])
                description = (
                    f"Khấu hao tài sản cố định - {label}"
                    if kind == "depreciation"
                    else f"Lãi tài sản cố định - {label}"
                )
                numeric_months = {
                    period: amount
                    for period, amount in months.items()
                    if amount != 0
                }
                explicit_zero_periods = {
                    period for period in explicit_zeroes[(kind, category_key)] if months.get(period, 0) == 0
                }
                payload.append({
                    "source_file": self.source_file_by_category.get("fixed_assets", "fixed_assets"),
                    "account_code": account_code,
                    "description": description,
                    "months": {},
                    "terms": {},
                    "numeric_months": numeric_months,
                    "explicit_zero_periods": explicit_zero_periods,
                    "highlight_periods": set(),
                    "provenance": f"fixed_assets_accounting|{kind}|{category_key}",
                    "audit_trail": (
                        "fixed_assets_audit_table=audit_fixed_asset_import_rows; "
                        f"fiscal_year={self.fiscal_year}; depreciation_cc={cc_code}; "
                        f"category_key={category_key}; account_code={account_code}"
                    ),
                })
        return payload

    def _load_append_rows(self, cc_code: int) -> list[dict[str, object]]:
        rows = self.conn.execute(
            """
            SELECT source, account_code, description, period, SUM(amount_vnd) AS amount
            FROM fact_input_data
            WHERE cc_code = ?
              AND account_code > 0
              AND form_row IS NULL
              AND source NOT IN ('facility', 'fixed_assets', 'it_sim', 'ga_unit_price')
            GROUP BY source, account_code, description, period
            ORDER BY account_code, description, source, period
            """,
            (str(cc_code),),
        ).fetchall()

        grouped: dict[tuple[int, str], dict[str, object]] = {}
        for row in rows:
            description = str(row["description"] or "")
            if self._fixed_row_for_description(description) is not None:
                continue

            clean_description = self._strip_explicit_formula_metadata(description)
            key = (int(row["account_code"]), clean_description)
            bucket = grouped.setdefault(
                key,
                {
                    "account_code": int(row["account_code"]),
                    "description": self._append_output_description(clean_description, description),
                    "months": {},
                    "terms": defaultdict(list),
                    "numeric_months": defaultdict(float),
                    "highlight_periods": set(),
                },
            )
            period = str(row["period"])
            term = self._explicit_formula_term_from_description(description) or self._alloc_formula_term_from_row(row)
            if MISSING_SEPARATE_COUNT_MARKER in description or EXPLICIT_ZERO_COUNT_MARKER in description:
                bucket["highlight_periods"].add(period)
            if term:
                bucket["terms"][period].append(term)
            else:
                amount = float(row["amount"] or 0.0)
                bucket["numeric_months"][period] += amount
                bucket["months"][period] = bucket["numeric_months"][period]
        allocation_rows = list(grouped.values())
        return self._load_fixed_asset_source_order_rows(cc_code) + allocation_rows

    def _load_nnn_source_order_rows(self, cc_code: int) -> list[dict[str, object]]:
        rows = self.conn.execute(
            """
            SELECT account_code, description, period, SUM(amount_vnd) AS amount
            FROM fact_input_data
            WHERE cc_code = ?
              AND source = 'nnn_paperwork'
              AND account_code > 0
            GROUP BY account_code, description, period
            ORDER BY account_code, description, period
            """,
            (str(cc_code),),
        ).fetchall()

        grouped: dict[tuple[int, str], dict[str, object]] = {}
        for row in rows:
            description = str(row["description"] or "")
            clean_description = self._strip_explicit_formula_metadata(description)
            raw_account_code = int(row["account_code"])
            try:
                account_code = resolve_account_code_for_connection(self.conn, cc_code, raw_account_code)
            except AccountResolutionError as exc:
                raise ExportIntegrityError(
                    "Không thể chuẩn hóa account NNN theo cost type của Cost Center "
                    f"{cc_code}: raw_account={raw_account_code}. {exc}"
                ) from exc
            key = (account_code, clean_description)
            bucket = grouped.setdefault(
                key,
                {
                    "account_code": account_code,
                    "description": self._append_output_description(clean_description, description),
                    "months": {},
                    "terms": defaultdict(list),
                    "numeric_months": defaultdict(float),
                    "highlight_periods": set(),
                    "source_group_index": 6,
                    "audit_trail": (
                        f"nnn_account_resolved_by_cc; raw_account_code={raw_account_code}; "
                        f"resolved_account_code={account_code}; cc_code={cc_code}"
                    ),
                },
            )
            period = str(row["period"])
            term = self._explicit_formula_term_from_description(description) or self._alloc_formula_term_from_row(row)
            if term:
                bucket["terms"][period].append(term)
            else:
                amount = float(row["amount"] or 0.0)
                bucket["numeric_months"][period] += amount
                bucket["months"][period] = bucket["numeric_months"][period]
        return list(grouped.values())

    def export_to_template(
        self,
        template_path: str,
        output_path: str,
        cc_code: Optional[object] = None,
        sheet_name: Optional[str] = None,
        start_row: int = APPEND_START_ROW,
    ) -> bool:
        target_cc = str(cc_code).strip() if cc_code else None
        if target_cc is None:
            return False

        fact_count = self._fact_count_for_cc(target_cc)
        if fact_count <= 0:
            return False

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Không tìm thấy tệp FORM template: {template_path}")

        template_workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=False)
        try:
            self._validate_template_workbook(template_workbook, template_path)
        finally:
            template_workbook.close()

        output_dir = os.path.dirname(os.path.abspath(output_path))
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        root, extension = os.path.splitext(output_path)
        temp_output_path = f"{root}.tmp_export{extension or '.xlsx'}"
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
        shutil.copy2(template_path, temp_output_path)

        workbook = openpyxl.load_workbook(temp_output_path)
        try:
            try:
                hub_sheet_name = sheet_name if sheet_name and sheet_name in workbook.sheetnames else helpers.find_hub_sheet_name(workbook)
                worksheet = workbook[hub_sheet_name]

                worksheet.cell(
                    row=5,
                    column=ACCOUNT_COL,
                    value=int(target_cc) if target_cc.isdigit() else target_cc,
                )
                self._resolved_recurring_rows = self._find_recurring_admin_rows(worksheet)
                self._resolved_it_system_row = self._find_it_system_total_row(worksheet)
                self._clear_template_business_payload(worksheet)
                self._write_source_staffing_time_rows(worksheet, target_cc)
                self._write_fixed_rows(worksheet, target_cc)

                append_start_row = self._resolve_append_start_row(worksheet, start_row)
                self._clear_append_area(worksheet, append_start_row)
                max_data_row = self._append_last_row(worksheet)
                current_row = append_start_row
                for row in self._load_append_rows(target_cc):
                    if current_row > max_data_row:
                        raise ValueError(
                            "Sheet chi tiết MP trong FORM không còn đủ dòng trống để ghi thêm chi phí phát sinh."
                        )
                    self._prepare_append_row(worksheet, current_row)
                    worksheet.cell(row=current_row, column=ACCOUNT_COL, value=int(row["account_code"]))
                    worksheet.cell(row=current_row, column=DESCRIPTION_COL, value=row["description"])
                    if row["terms"]:
                        self._write_formula_series(
                            worksheet,
                            current_row,
                            dict(row["terms"]),
                            dict(row["numeric_months"]),
                            row.get("highlight_periods"),
                        )
                    else:
                        self._write_numeric_series(worksheet, current_row, row["months"])
                    current_row += 1

                normalize_output_description_column_s(worksheet)
                self._normalize_visible_fiscal_year_labels(workbook)
                workbook.save(temp_output_path)
                self._validate_exported_workbook(workbook, temp_output_path, target_cc, fact_count)
            finally:
                workbook.close()
            os.replace(temp_output_path, output_path)
            return True
        except Exception:
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
            raise
