"""Keep user-selected output cost-row order independent from row position.

The workbook stores ownership per cost row in a very-hidden sheet.  Therefore a
manual row may be displayed between common rows without becoming generator
owned on a later run.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator

from src.utils import excel_helpers as helpers


OUTPUT_COST_ROW_ORDER_METADATA_SHEET = "_mp2027_output_cost_row_order"
OUTPUT_COST_ROW_ORDER_SCHEMA_VERSION = 1
LEGACY_MANUAL_METADATA_SHEETS = (
    "_mp2027_manual_cost_meta",
    "_mp2027_manual_special_cost_meta",
)
MANUAL_METADATA_SHEET = LEGACY_MANUAL_METADATA_SHEETS[0]
INVALID_LEGACY_MANUAL_METADATA_SHEET = LEGACY_MANUAL_METADATA_SHEETS[1]
FIRST_COST_ROW = helpers.FORM_SHARED_COST_START_ROW
FIRST_COST_COLUMN = 1
LAST_COST_COLUMN = 20
ACCOUNT_COLUMN = 2
MONEY_FIRST_COLUMN = 6
MONEY_LAST_COLUMN = 18
DESCRIPTION_COLUMN = 19
ROW_ORDER_HEADERS = (
    "cc_code", "sheet_name", "row_id", "row_kind", "signature",
    "sort_order", "current_row", "schema_version",
)


class OutputCostRowOrderError(ValueError):
    """Raised when an output layout cannot be safely identified."""


@dataclass(frozen=True)
class CostRow:
    row_id: str
    row_kind: str
    description: str
    account_code: object
    current_row: int


@dataclass(frozen=True)
class _CellSnapshot:
    value: object
    font: object
    fill: object
    border: object
    alignment: object
    protection: object
    number_format: str
    comment: object
    hyperlink: object


@dataclass(frozen=True)
class _RowSnapshot:
    row_id: str
    row_kind: str
    signature: str
    source_row: int
    cells: tuple[_CellSnapshot, ...]
    height: float | None
    hidden: bool


def _normalise(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _cc(value: object) -> str:
    return str(value or "").strip()


def _hub_sheet(workbook):
    try:
        return workbook[helpers.find_hub_sheet_name(workbook)]
    except ValueError:
        return workbook.active


def _row_has_content(worksheet, row: int) -> bool:
    return any(
        worksheet.cell(row, column).value not in (None, "")
        or worksheet.cell(row, column).comment is not None
        for column in range(FIRST_COST_COLUMN, LAST_COST_COLUMN + 1)
    )


def _is_separator(worksheet, row: int) -> bool:
    return _normalise(worksheet.cell(row, DESCRIPTION_COLUMN).value).upper().startswith("CHI PHÍ RIÊNG")


def _last_cost_row(worksheet) -> int:
    for row in range(int(worksheet.max_row or 0), FIRST_COST_ROW - 1, -1):
        if _row_has_content(worksheet, row):
            return row
    return FIRST_COST_ROW - 1


def _business_rows(worksheet) -> list[int]:
    return [
        row
        for row in range(FIRST_COST_ROW, _last_cost_row(worksheet) + 1)
        if _row_has_content(worksheet, row) and not _is_separator(worksheet, row)
    ]


def _signature(worksheet, row: int, seen: dict[str, int]) -> str:
    base = "|".join((_normalise(worksheet.cell(row, ACCOUNT_COLUMN).value), _normalise(worksheet.cell(row, DESCRIPTION_COLUMN).value)))
    if base == "|":
        raise OutputCostRowOrderError(f"Dòng {row} không có mã hoặc mô tả chi phí để nhận dạng an toàn.")
    seen[base] = seen.get(base, 0) + 1
    return f"{base}|{seen[base]}"


def _legacy_manual_bounds(workbook, cc_code: str) -> tuple[str, int, int] | None:
    for metadata_sheet in LEGACY_MANUAL_METADATA_SHEETS:
        if metadata_sheet not in workbook.sheetnames:
            continue
        metadata = workbook[metadata_sheet]
        for row in range(2, int(metadata.max_row or 1) + 1):
            if _cc(metadata.cell(row, 3).value) != cc_code:
                continue
            sheet_name = str(metadata.cell(row, 1).value or "").strip()
            common_end = int(metadata.cell(row, 4).value or 0)
            manual_start = int(metadata.cell(row, 5).value or 0)
            manual_end = int(metadata.cell(row, 6).value or 0)
            schema_version = int(metadata.cell(row, 7).value or 0)
            if schema_version != 1 or not sheet_name or manual_start <= common_end or manual_end < manual_start - 1:
                raise OutputCostRowOrderError(f"Dấu mốc chi phí riêng của CC {cc_code} không hợp lệ.")
            return sheet_name, manual_start, manual_end
    return None


def _metadata_sheet(workbook, *, create: bool = False):
    if OUTPUT_COST_ROW_ORDER_METADATA_SHEET in workbook.sheetnames:
        return workbook[OUTPUT_COST_ROW_ORDER_METADATA_SHEET]
    if not create:
        return None
    sheet = workbook.create_sheet(OUTPUT_COST_ROW_ORDER_METADATA_SHEET)
    sheet.sheet_state = "veryHidden"
    sheet.append(ROW_ORDER_HEADERS)
    return sheet


def _migrate_legacy_manual_metadata(workbook) -> None:
    """Replace the pre-fix 32-character metadata title before saving."""
    if INVALID_LEGACY_MANUAL_METADATA_SHEET not in workbook.sheetnames:
        return
    legacy = workbook[INVALID_LEGACY_MANUAL_METADATA_SHEET]
    if MANUAL_METADATA_SHEET not in workbook.sheetnames:
        migrated = workbook.create_sheet(MANUAL_METADATA_SHEET)
        migrated.sheet_state = "veryHidden"
        for row in legacy.iter_rows(values_only=True):
            migrated.append(row)
    workbook.remove(legacy)


def _read_records(workbook, cc_code: str) -> list[dict[str, object]] | None:
    metadata = _metadata_sheet(workbook)
    if metadata is None:
        return None
    records: list[dict[str, object]] = []
    for row in range(2, int(metadata.max_row or 1) + 1):
        if _cc(metadata.cell(row, 1).value) != cc_code:
            continue
        record = {
            "cc_code": cc_code,
            "sheet_name": str(metadata.cell(row, 2).value or "").strip(),
            "row_id": str(metadata.cell(row, 3).value or "").strip(),
            "row_kind": str(metadata.cell(row, 4).value or "").strip(),
            "signature": str(metadata.cell(row, 5).value or "").strip(),
            "sort_order": int(metadata.cell(row, 6).value or 0),
            "current_row": int(metadata.cell(row, 7).value or 0),
            "schema_version": int(metadata.cell(row, 8).value or 0),
        }
        if (
            record["schema_version"] != OUTPUT_COST_ROW_ORDER_SCHEMA_VERSION
            or not record["sheet_name"]
            or not record["row_id"]
            or record["row_kind"] not in {"common", "manual"}
            or record["sort_order"] < 1
            or record["current_row"] < FIRST_COST_ROW
        ):
            raise OutputCostRowOrderError(f"Metadata thứ tự dòng của CC {cc_code} không hợp lệ.")
        records.append(record)
    if not records:
        return None
    if len({str(record["row_id"]) for record in records}) != len(records):
        raise OutputCostRowOrderError(f"Metadata thứ tự dòng của CC {cc_code} có mã dòng trùng.")
    if len({int(record["sort_order"]) for record in records}) != len(records):
        raise OutputCostRowOrderError(f"Metadata thứ tự dòng của CC {cc_code} có thứ tự trùng.")
    return sorted(records, key=lambda record: int(record["sort_order"]))


def _records_from_legacy(workbook, worksheet, cc_code: str) -> list[dict[str, object]]:
    bounds = _legacy_manual_bounds(workbook, cc_code)
    if bounds is None:
        raise OutputCostRowOrderError(
            f"Workbook CC {cc_code} chưa có dấu mốc chi phí riêng. Hãy xác nhận dòng bắt đầu chi phí riêng trước khi sắp xếp."
        )
    sheet_name, manual_start, manual_end = bounds
    if sheet_name != worksheet.title:
        raise OutputCostRowOrderError(f"Dấu mốc chi phí riêng của CC {cc_code} trỏ tới sheet không đúng.")
    seen: dict[str, int] = {}
    records: list[dict[str, object]] = []
    manual_index = 0
    for sort_order, row in enumerate(_business_rows(worksheet), start=1):
        if manual_start <= row <= manual_end:
            manual_index += 1
            row_kind = "manual"
            signature = f"manual|{manual_index}"
            row_id = f"manual:{manual_index}"
        else:
            row_kind = "common"
            signature = _signature(worksheet, row, seen)
            row_id = f"common:{signature}"
        records.append({
            "cc_code": cc_code, "sheet_name": worksheet.title, "row_id": row_id,
            "row_kind": row_kind, "signature": signature, "sort_order": sort_order,
            "current_row": row, "schema_version": OUTPUT_COST_ROW_ORDER_SCHEMA_VERSION,
        })
    return records


def _records_for_generated_common(worksheet, cc_code: str) -> list[dict[str, object]]:
    seen: dict[str, int] = {}
    return [
        {
            "cc_code": cc_code, "sheet_name": worksheet.title,
            "row_id": f"common:{signature}", "row_kind": "common", "signature": signature,
            "sort_order": index, "current_row": row,
            "schema_version": OUTPUT_COST_ROW_ORDER_SCHEMA_VERSION,
        }
        for index, (row, signature) in enumerate(
            ((row, _signature(worksheet, row, seen)) for row in _business_rows(worksheet)), start=1
        )
    ]


def _snapshot_row(worksheet, record: dict[str, object]) -> _RowSnapshot:
    row = int(record["current_row"])
    if not _row_has_content(worksheet, row):
        raise OutputCostRowOrderError(f"Dòng {row} trong metadata CC {record['cc_code']} không còn dữ liệu.")
    cells = tuple(
        _CellSnapshot(
            worksheet.cell(row, column).value,
            copy(worksheet.cell(row, column).font),
            copy(worksheet.cell(row, column).fill),
            copy(worksheet.cell(row, column).border),
            copy(worksheet.cell(row, column).alignment),
            copy(worksheet.cell(row, column).protection),
            worksheet.cell(row, column).number_format,
            copy(worksheet.cell(row, column).comment),
            copy(worksheet.cell(row, column).hyperlink),
        )
        for column in range(FIRST_COST_COLUMN, LAST_COST_COLUMN + 1)
    )
    dimension = worksheet.row_dimensions[row]
    return _RowSnapshot(
        str(record["row_id"]), str(record["row_kind"]), str(record["signature"]), row,
        cells, dimension.height, bool(dimension.hidden),
    )


def _translate_formula(value: object, source_row: int, target_row: int, column: int) -> object:
    if not isinstance(value, str) or not value.startswith("=") or source_row == target_row:
        return value
    origin = f"{openpyxl.utils.get_column_letter(column)}{source_row}"
    target = f"{openpyxl.utils.get_column_letter(column)}{target_row}"
    try:
        return Translator(value, origin=origin).translate_formula(target)
    except Exception:
        return value


def _clear_row(worksheet, row: int) -> None:
    for column in range(FIRST_COST_COLUMN, LAST_COST_COLUMN + 1):
        cell = worksheet.cell(row, column)
        cell.value = None
        cell.comment = None
        cell.hyperlink = None
        cell._style = None


def _write_snapshot(worksheet, snapshot: _RowSnapshot, target_row: int, *, clear_money: bool) -> None:
    for offset, source in enumerate(snapshot.cells, start=FIRST_COST_COLUMN):
        target = worksheet.cell(target_row, offset)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format
        target.comment = copy(source.comment)
        target._hyperlink = copy(source.hyperlink)
        if clear_money and MONEY_FIRST_COLUMN <= offset <= MONEY_LAST_COLUMN:
            target.value = None
        else:
            target.value = _translate_formula(source.value, snapshot.source_row, target_row, offset)
    dimension = worksheet.row_dimensions[target_row]
    dimension.height = snapshot.height
    dimension.hidden = snapshot.hidden


def _write_records(workbook, cc_code: str, records: list[dict[str, object]]) -> None:
    _migrate_legacy_manual_metadata(workbook)
    metadata = _metadata_sheet(workbook, create=True)
    metadata.sheet_state = "veryHidden"
    retained = [
        tuple(metadata.cell(row, column).value for column in range(1, len(ROW_ORDER_HEADERS) + 1))
        for row in range(2, int(metadata.max_row or 1) + 1)
        if _cc(metadata.cell(row, 1).value) != cc_code
    ]
    metadata.delete_rows(1, int(metadata.max_row or 1))
    metadata.append(ROW_ORDER_HEADERS)
    for values in retained:
        metadata.append(values)
    for record in records:
        metadata.append(tuple(record[header] for header in ROW_ORDER_HEADERS))


def _read_or_legacy_records(workbook, worksheet, cc_code: str) -> list[dict[str, object]]:
    records = _read_records(workbook, cc_code)
    if records is None:
        return _records_from_legacy(workbook, worksheet, cc_code)
    if any(str(record["sheet_name"]) != worksheet.title for record in records):
        raise OutputCostRowOrderError(f"Metadata thứ tự dòng của CC {cc_code} trỏ tới sheet không đúng.")
    return records


def read_cost_rows(workbook_path: str | Path, cc_code: object) -> list[CostRow]:
    """Return movable cost rows in their saved/current display order."""
    normalized_cc = _cc(cc_code)
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        worksheet = _hub_sheet(workbook)
        records = _read_or_legacy_records(workbook, worksheet, normalized_cc)
        return [
            CostRow(
                str(record["row_id"]), str(record["row_kind"]),
                _normalise(worksheet.cell(int(record["current_row"]), DESCRIPTION_COLUMN).value),
                worksheet.cell(int(record["current_row"]), ACCOUNT_COLUMN).value,
                int(record["current_row"]),
            )
            for record in records
        ]
    finally:
        workbook.close()


def has_saved_cost_row_order(workbook_path: str | Path, cc_code: object) -> bool:
    """Return whether this CC has row-level layout metadata already saved."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return _read_records(workbook, _cc(cc_code)) is not None
    finally:
        workbook.close()


def save_cost_row_order(workbook_path: str | Path, cc_code: object, row_ids: list[str]) -> dict[str, int]:
    """Apply user order to one output workbook and persist row ownership."""
    normalized_cc = _cc(cc_code)
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        worksheet = _hub_sheet(workbook)
        records = _read_or_legacy_records(workbook, worksheet, normalized_cc)
        by_id = {str(record["row_id"]): record for record in records}
        if len(row_ids) != len(records) or set(row_ids) != set(by_id):
            raise OutputCostRowOrderError("Thứ tự dòng lưu không khớp với các dòng chi phí trong workbook.")
        snapshots = {row_id: _snapshot_row(worksheet, by_id[row_id]) for row_id in row_ids}
        last_row = _last_cost_row(worksheet)
        for row in range(FIRST_COST_ROW, last_row + 1):
            _clear_row(worksheet, row)
        written: list[dict[str, object]] = []
        for index, row_id in enumerate(row_ids, start=1):
            record = dict(by_id[row_id])
            target_row = FIRST_COST_ROW + index - 1
            _write_snapshot(worksheet, snapshots[row_id], target_row, clear_money=False)
            record["sort_order"] = index
            record["current_row"] = target_row
            written.append(record)
        _write_records(workbook, normalized_cc, written)
        workbook.save(workbook_path)
    finally:
        workbook.close()
    return {"rows_saved": len(row_ids), "manual_rows": sum(row_id.startswith("manual:") for row_id in row_ids)}


def restore_cost_layout(
    generated_workbook_path: str | Path,
    cc_code: object,
    source_workbook_path: str | Path,
    *,
    source_kind: str,
) -> dict[str, int | str]:
    """Merge fresh common rows with a saved mixed source layout."""
    if source_kind not in {"current_fiscal_year", "previous_fiscal_year"}:
        raise OutputCostRowOrderError(f"Loại nguồn kế thừa không hợp lệ: {source_kind}")
    normalized_cc = _cc(cc_code)
    source = openpyxl.load_workbook(source_workbook_path, data_only=False)
    generated = openpyxl.load_workbook(generated_workbook_path, data_only=False)
    try:
        source_sheet = _hub_sheet(source)
        generated_sheet = _hub_sheet(generated)
        source_records = _read_records(source, normalized_cc)
        if source_records is None:
            raise OutputCostRowOrderError(
                f"Workbook nguồn CC {normalized_cc} chưa có thứ tự dòng đã lưu. Hãy dùng nút Sắp xếp dòng chi phí để lưu thứ tự trước."
            )
        if any(str(record["sheet_name"]) != source_sheet.title for record in source_records):
            raise OutputCostRowOrderError(f"Metadata thứ tự dòng của CC {normalized_cc} trỏ tới sheet không đúng.")
        source_snapshots = {str(record["row_id"]): _snapshot_row(source_sheet, record) for record in source_records}
        common_records = _records_for_generated_common(generated_sheet, normalized_cc)
        common_snapshots = {str(record["row_id"]): _snapshot_row(generated_sheet, record) for record in common_records}

        ordered: list[tuple[_RowSnapshot, bool]] = []
        used_common: set[str] = set()
        for record in source_records:
            row_id = str(record["row_id"])
            if record["row_kind"] == "manual":
                ordered.append((source_snapshots[row_id], source_kind == "previous_fiscal_year"))
            elif row_id in common_snapshots:
                ordered.append((common_snapshots[row_id], False))
                used_common.add(row_id)
        for record in common_records:
            row_id = str(record["row_id"])
            if row_id not in used_common:
                ordered.append((common_snapshots[row_id], False))

        last_row = _last_cost_row(generated_sheet)
        for row in range(FIRST_COST_ROW, last_row + 1):
            _clear_row(generated_sheet, row)
        written: list[dict[str, object]] = []
        for index, (snapshot, clear_money) in enumerate(ordered, start=1):
            target_row = FIRST_COST_ROW + index - 1
            _write_snapshot(generated_sheet, snapshot, target_row, clear_money=clear_money)
            written.append({
                "cc_code": normalized_cc, "sheet_name": generated_sheet.title,
                "row_id": snapshot.row_id, "row_kind": snapshot.row_kind,
                "signature": snapshot.signature, "sort_order": index,
                "current_row": target_row, "schema_version": OUTPUT_COST_ROW_ORDER_SCHEMA_VERSION,
            })
        _write_records(generated, normalized_cc, written)
        generated.save(generated_workbook_path)
    finally:
        source.close()
        generated.close()

    manual_rows = sum(snapshot.row_kind == "manual" for snapshot, _ in ordered)
    return {
        "cc_code": normalized_cc,
        "manual_rows_preserved": manual_rows,
        "new_common_rows": len(common_records) - len(used_common),
        "rows_written": len(ordered),
    }
