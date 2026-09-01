# -*- coding: utf-8 -*-
from __future__ import annotations
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple, Any
import os
import glob
import re
import ast
from decimal import Decimal, ROUND_HALF_UP
from enum import Enum
import warnings

from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

class VarianceStatus(Enum):
    INCREASE = "Tăng"
    DECREASE = "Giảm"
    UNCHANGED = "Không đổi"
    NEW_ITEM = "Mới phát sinh"
    REMOVED = "Cắt giảm hoàn toàn"

@dataclass
class ComparisonContext:
    fiscal_year_base: int
    fiscal_year_current: int
    cost_center_code: str
    base_file_path: str
    current_file_path: str
    threshold_percent: float = 10.0
    threshold_absolute: float = 50000000.0

@dataclass
class CostLineVariance:
    account_code: str
    item_name: str
    base_value: float
    current_value: float
    variance_absolute: float
    variance_percent: Optional[float]
    status: VarianceStatus
    is_alert: bool

@dataclass
class VarianceReport:
    context: ComparisonContext
    lines: List[CostLineVariance] = field(default_factory=list)
    total_base: float = 0.0
    total_current: float = 0.0
    total_variance_absolute: float = 0.0
    total_variance_percent: Optional[float] = 0.0

def _resolve_hub_sheet_name(filepath: str) -> str:
    """
    Xác định tên sheet chi tiết chi phí (hub sheet) trong file MP FORM.
    Ưu tiên sử dụng find_hub_sheet_name() từ hệ thống chính.
    Nếu file không có cấu trúc FORM chuẩn, fallback tìm sheet chứa '内訳'.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        # Thử dùng hệ thống chính trước
        try:
            from src.utils.excel_helpers import find_hub_sheet_name
            name = find_hub_sheet_name(wb)
            return name
        except Exception:
            pass
        # Fallback: tìm sheet có chứa '内訳'
        hub_candidates = [s for s in wb.sheetnames if "内訳" in s]
        if len(hub_candidates) == 1:
            return hub_candidates[0]
        raise ValueError(
            f"Không tìm thấy trang tính chi tiết MP (tên chứa '内訳') trong tệp '{os.path.basename(filepath)}'.\n\n"
            f"Nguyên nhân: Tệp đã chọn không đúng định dạng FORM chuẩn của chương trình hoặc trang tính đã bị đổi tên/xóa.\n"
            f"Cách xử lý:\n"
            f"1. Kiểm tra lại đường dẫn tệp vừa chọn.\n"
            f"2. Đảm bảo tệp là báo cáo MP xuất từ chương trình có chứa trang tính '内訳ﾘｽﾄ(4～3月)'."
        )
    finally:
        wb.close()


# Account codes in MP FORM are numeric and ≥ 7 digits (e.g. 9114120018, 5005246282).
# Rows without such a code in the account column are headers, labels, or layout rows.
import re as _re
_ACCOUNT_CODE_PATTERN = _re.compile(r"^\d{7,}$")


def _is_cost_row(excel_row: int, acc_value: Any) -> bool:
    """
    Kiểm tra xem dòng hiện tại có phải là dòng chi phí hợp lệ theo cấu trúc FORM thật không.
    Dòng chi phí phải nằm trong các dòng input hợp lệ (nhân sự) hoặc từ dòng 38 trở đi,
    VÀ phải có mã tài khoản là số nguyên ≥ 7 chữ số (ví dụ: 9114120018, 5005246282).
    """
    from src.utils.excel_helpers import FORM_TEMPLATE_INPUT_ROWS, FORM_SHARED_COST_START_ROW
    if excel_row not in FORM_TEMPLATE_INPUT_ROWS and excel_row < FORM_SHARED_COST_START_ROW:
        return False

    import pandas as pd
    if acc_value is None or (isinstance(acc_value, float) and pd.isna(acc_value)):
        return False
    s = str(acc_value).strip()
    if not s or s.lower() == "nan":
        return False
    # Cho phép cả số nguyên lẫn float có phần thập phân .0
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            return False
    return bool(_ACCOUNT_CODE_PATTERN.match(s))


def safe_load_mp_form(filepath: str) -> "pd.DataFrame":
    """
    Đọc file MP FORM, tự động xác định đúng sheet chi tiết chi phí (hub sheet).
    Không dùng sheet_name=0 vì sheet đầu (採算表(USD)) là bảng tổng hợp lãi lỗ.
    """
    import openpyxl
    import pandas as pd
    try:
        hub_sheet = _resolve_hub_sheet_name(filepath)
        # openpyxl writes formulas but cannot write Excel's cached formula
        # results. A freshly generated MP workbook therefore has blank values
        # in data_only mode until somebody opens it in Excel. YoY must not
        # depend on that manual step.
        # YoY dereferences formulas repeatedly. Normal workbook mode keeps cell
        # access in memory; read_only mode re-scans worksheet XML per lookup.
        value_book = openpyxl.load_workbook(filepath, data_only=True)
        formula_book = openpyxl.load_workbook(filepath, data_only=False)
        try:
            value_sheet = value_book[hub_sheet]
            formula_sheet = formula_book[hub_sheet]
            max_row = max(int(value_sheet.max_row or 0), int(formula_sheet.max_row or 0))
            max_column = max(18, int(value_sheet.max_column or 0), int(formula_sheet.max_column or 0))
            rows = [
                [value_sheet.cell(row=row, column=column).value for column in range(1, max_column + 1)]
                for row in range(1, max_row + 1)
            ]
            dataframe = pd.DataFrame(rows)
            _hydrate_cost_descriptions(dataframe, formula_sheet, value_sheet)
            _hydrate_formula_totals(dataframe, formula_sheet, value_sheet)
            return dataframe
        finally:
            value_book.close()
            formula_book.close()
    except Exception as e:
        raise RuntimeError(f"Lỗi khi đọc tệp '{os.path.basename(filepath)}':\n{str(e)}")

class _MpFormulaResolver:
    """Evaluate the arithmetic formulas generated by the MP exporter only."""

    _CELL_REF = re.compile(r"(?i)\b([A-Z]{1,3}\d+)\b")
    _RANGE_REF = re.compile(r"(?i)\b([A-Z]{1,3}\d+)\s*:\s*([A-Z]{1,3}\d+)\b")

    def __init__(self, formula_sheet, value_sheet) -> None:
        self.formula_sheet = formula_sheet
        self.value_sheet = value_sheet
        self._cache: dict[tuple[int, int], float] = {}
        self._active: set[tuple[int, int]] = set()

    @staticmethod
    def _as_number(value: object) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip().replace(",", "")
            if text and not text.startswith("="):
                try:
                    return float(text)
                except ValueError:
                    return None
        return None

    @staticmethod
    def _excel_round(value: float, digits: float = 0) -> float:
        places = int(digits)
        quantum = Decimal("1").scaleb(-places)
        return float(Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP))

    def cell_value(self, row: int, column: int) -> float:
        key = (row, column)
        if key in self._cache:
            return self._cache[key]
        if key in self._active:
            raise ValueError(f"C\u00f4ng th\u1ee9c MP b\u1ecb v\u00f2ng l\u1eb7p t\u1ea1i {self.formula_sheet.cell(row, column).coordinate}.")

        cached = self._as_number(self.value_sheet.cell(row, column).value)
        if cached is not None:
            self._cache[key] = cached
            return cached

        formula = self.formula_sheet.cell(row, column).value
        if formula is None or (isinstance(formula, str) and not formula.strip()):
            return 0.0
        literal = self._as_number(formula)
        if literal is not None:
            self._cache[key] = literal
            return literal
        if not isinstance(formula, str) or not formula.lstrip().startswith("="):
            raise ValueError(
                f"Kh\u00f4ng th\u1ec3 \u0111\u1ecdc gi\u00e1 tr\u1ecb s\u1ed1 t\u1ea1i {self.formula_sheet.cell(row, column).coordinate}."
            )

        self._active.add(key)
        try:
            result = self._evaluate_formula(formula)
        except Exception as exc:
            raise ValueError(
                f"Kh\u00f4ng th\u1ec3 t\u1ef1 t\u00ednh c\u00f4ng th\u1ee9c MP t\u1ea1i {self.formula_sheet.cell(row, column).coordinate}: {formula}"
            ) from exc
        finally:
            self._active.remove(key)
        self._cache[key] = result
        return result

    def _range_sum(self, start: str, end: str) -> float:
        start_column, start_row = coordinate_from_string(start)
        end_column, end_row = coordinate_from_string(end)
        left = column_index_from_string(start_column)
        right = column_index_from_string(end_column)
        return sum(
            self.cell_value(row, column)
            for row in range(min(start_row, end_row), max(start_row, end_row) + 1)
            for column in range(min(left, right), max(left, right) + 1)
        )

    def _evaluate_formula(self, formula: str) -> float:
        expression = formula.strip().lstrip("=").replace("$", "").replace("^", "**")
        ranges: list[tuple[str, str]] = []

        def replace_range(match: re.Match[str]) -> str:
            ranges.append((match.group(1).upper(), match.group(2).upper()))
            return f"__MP_RANGE_{len(ranges) - 1}__"

        expression = self._RANGE_REF.sub(replace_range, expression)
        expression = self._CELL_REF.sub(lambda match: f'CELL("{match.group(1).upper()}")', expression)
        for index, (start, end) in enumerate(ranges):
            expression = expression.replace(f"__MP_RANGE_{index}__", f'SUM_RANGE("{start}", "{end}")')

        tree = ast.parse(expression, mode="eval")
        return float(self._evaluate_node(tree.body))

    def _evaluate_node(self, node: ast.AST) -> float:
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = self._evaluate_node(node.operand)
            return value if isinstance(node.op, ast.UAdd) else -value
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow)):
            left = self._evaluate_node(node.left)
            right = self._evaluate_node(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right
            return left ** right
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            name = node.func.id.upper()
            if name == "CELL" and len(node.args) == 1 and isinstance(node.args[0], ast.Constant):
                column, row = coordinate_from_string(str(node.args[0].value))
                return self.cell_value(row, column_index_from_string(column))
            if name == "SUM_RANGE" and len(node.args) == 2 and all(isinstance(arg, ast.Constant) for arg in node.args):
                return self._range_sum(str(node.args[0].value), str(node.args[1].value))
            values = [self._evaluate_node(arg) for arg in node.args]
            if name == "SUM":
                return sum(values)
            if name == "ROUND" and len(values) == 2:
                return self._excel_round(values[0], values[1])
        raise ValueError("C\u00f4ng th\u1ee9c kh\u00f4ng thu\u1ed9c d\u1ea1ng MP c\u00f3 th\u1ec3 t\u1ef1 t\u00ednh.")


def _hydrate_cost_descriptions(dataframe: "pd.DataFrame", formula_sheet, value_sheet) -> None:
    """Use the exporter source-description column for cost-row identity.

    Column C in newly exported MP files is a VLOOKUP formula.  Until Excel has
    opened the file it has no cached result, while column S already contains
    the concrete source description written by the exporter.  YoY needs that
    concrete text to distinguish, for example, land from building costs under
    the same account code.
    """
    account_column = 2
    description_column = 3
    source_description_column = 19
    for row in range(1, int(formula_sheet.max_row or 0) + 1):
        account_value = value_sheet.cell(row, account_column).value
        if not _is_cost_row(row, account_value):
            continue
        source_description = value_sheet.cell(row, source_description_column).value
        if not isinstance(source_description, str) or not source_description.strip():
            continue
        dataframe.iat[row - 1, description_column - 1] = source_description.strip()


def _hydrate_formula_totals(dataframe: "pd.DataFrame", formula_sheet, value_sheet) -> None:
    """Fill column R from MP formulas when Excel has not cached it yet."""
    total_column = 18
    resolver = _MpFormulaResolver(formula_sheet, value_sheet)
    for row in range(1, int(formula_sheet.max_row or 0) + 1):
        account_value = value_sheet.cell(row, 2).value
        if not _is_cost_row(row, account_value):
            continue
        cached_total = resolver._as_number(value_sheet.cell(row, total_column).value)
        formula_total = formula_sheet.cell(row, total_column).value
        if cached_total is not None or not isinstance(formula_total, str) or not formula_total.lstrip().startswith("="):
            continue
        dataframe.iat[row - 1, total_column - 1] = resolver.cell_value(row, total_column)


def calculate_variance(base_val: float, current_val: float) -> tuple[float, float, VarianceStatus]:
    """
    Returns (absolute_variance, percent_variance, status)
    """
    variance_abs = current_val - base_val
    if base_val == 0.0 and current_val == 0.0:
        return 0.0, 0.0, VarianceStatus.UNCHANGED
    elif base_val == 0.0 and current_val > 0.0:
        return variance_abs, 100.0, VarianceStatus.NEW_ITEM
    elif current_val == 0.0 and base_val > 0.0:
        return variance_abs, -100.0, VarianceStatus.REMOVED

    variance_pct = (variance_abs / base_val) * 100.0
    if variance_abs > 0:
        return variance_abs, variance_pct, VarianceStatus.INCREASE
    elif variance_abs < 0:
        return variance_abs, variance_pct, VarianceStatus.DECREASE
    return variance_abs, variance_pct, VarianceStatus.UNCHANGED

def is_variance_alert(variance_abs: float, variance_pct: float, threshold_abs: float, threshold_pct: float) -> bool:
    return abs(variance_pct) >= threshold_pct or abs(variance_abs) >= threshold_abs

def _validate_and_extract_cost_series(
    df: "pd.DataFrame",
    val_col: Any,
    acc_col: Any,
    name_col: Any,
    file_label: str,
    file_path: Optional[str] = None
) -> "pd.Series":
    """
    Kiểm tra cột chi phí trong DataFrame theo nguyên tắc:
    - Ô trống (""), ô chỉ có khoảng trắng ("   "), None hoặc NaN: dừng và báo lỗi rõ ràng.
    - Giá trị số 0 (int 0, float 0.0, string "0", "0.0"): hợp lệ, chuyển thành float 0.0.
    - Giá trị số khác: hợp lệ, chuyển thành float.
    - Giá trị chữ không hợp lệ (không parse được thành float): báo lỗi.
    """
    import pandas as pd

    file_info = file_label
    if file_path:
        file_info = f"{file_label} ('{os.path.basename(file_path)}')"

    cleaned_values = []

    for idx, row in df.iterrows():
        # Xác định dòng Excel (1-based index)
        if isinstance(idx, (int, float)):
            excel_row = int(idx) + 1
        else:
            excel_row = str(idx)

        acc_str = ""
        if acc_col in row and pd.notna(row[acc_col]):
            s = str(row[acc_col]).strip()
            if s and s.lower() != "nan":
                acc_str = s

        name_str = ""
        if name_col in row and pd.notna(row[name_col]):
            s = str(row[name_col]).strip()
            if s and s.lower() != "nan":
                name_str = s

        details = []
        if acc_str:
            details.append(f"Mã tài khoản: '{acc_str}'")
        if name_str:
            details.append(f"Tên khoản mục: '{name_str}'")
        detail_text = f" [{', '.join(details)}]" if details else ""

        val = row.get(val_col) if hasattr(row, 'get') else row[val_col]

        # Kiểm tra ô trống, chỉ khoảng trắng, None, NaN
        is_blank = False
        if val is None:
            is_blank = True
        elif pd.isna(val):
            is_blank = True
        elif isinstance(val, str) and val.strip() == "":
            is_blank = True

        if is_blank:
            fname = os.path.basename(file_path) if file_path else file_label
            raise ValueError(
                f"Tệp {file_info} (dòng Excel {excel_row}){detail_text}: Ô chi phí đang để trống.\n\n"
                f"Nguyên nhân: Ô này chưa có số liệu chi phí hoặc thiếu baseline T3 (chương trình không tự hiểu ô trống là 0).\n"
                f"Cách xử lý:\n"
                f"1. Mở tệp '{fname}'.\n"
                f"2. Kiểm tra dòng Excel {excel_row}{detail_text}.\n"
                f"3. Nếu chi phí thực tế bằng 0, hãy nhập rõ số 0. Nếu chưa hoàn tất số liệu, hãy bổ sung baseline T3 trước khi so sánh.\n"
                f"4. Lưu tệp Excel và thực hiện so sánh lại."
            )

        # Parse giá trị số (cho phép 0, 0.0, số dương, số âm, số có dấu phẩy)
        try:
            if isinstance(val, str):
                num_val = float(val.strip().replace(",", ""))
            else:
                num_val = float(val)
        except (ValueError, TypeError) as e:
            fname = os.path.basename(file_path) if file_path else file_label
            raise ValueError(
                f"Tệp {file_info} (dòng Excel {excel_row}){detail_text}: Dữ liệu tại cột chi phí không phải là số hợp lệ (giá trị: '{val}').\n\n"
                f"Nguyên nhân: Ô chứa ký tự chữ hoặc định dạng không đúng chuẩn số.\n"
                f"Cách xử lý:\n"
                f"1. Mở tệp '{fname}'.\n"
                f"2. Kiểm tra dòng Excel {excel_row}{detail_text}.\n"
                f"3. Sửa lại giá trị thành số hợp lệ.\n"
                f"4. Lưu tệp Excel và thực hiện so sánh lại."
            )

        cleaned_values.append(num_val)

    return pd.Series(cleaned_values, index=df.index, dtype=float)

def map_and_analyze_variances(df_base: "pd.DataFrame", df_current: "pd.DataFrame", ctx: ComparisonContext, acc_col: Any = 1, name_col: Any = 2, val_col: Any = 17) -> VarianceReport:
    """
    Maps rows by Account and Name, then calculates variance.
    Only processes cost rows (rows with valid 7+ digit account codes).
    Header, department, total, and layout rows are silently skipped.
    """
    import pandas as pd

    df_base = df_base.copy()
    df_current = df_current.copy()

    # Filter to cost rows only: rows where account column has a valid 7+ digit numeric code
    base_mask = [ _is_cost_row(idx + 1, val) for idx, val in zip(df_base.index, df_base[acc_col]) ]
    curr_mask = [ _is_cost_row(idx + 1, val) for idx, val in zip(df_current.index, df_current[acc_col]) ]

    df_base = df_base[base_mask].copy()
    df_current = df_current[curr_mask].copy()

    if df_base.empty and df_current.empty:
        raise ValueError(
            f"Không tìm thấy dòng chi phí hợp lệ (mã tài khoản ≥ 7 chữ số) trong cả hai tệp.\n\n"
            f"Nguyên nhân: Tệp không đúng cấu trúc FORM chuẩn hoặc các dòng chi phí đã bị xóa.\n"
            f"Cách xử lý:\n"
            f"1. Kiểm tra lại xem đã chọn đúng tệp báo cáo MP chuẩn chưa.\n"
            f"2. Đảm bảo trang tính chi tiết có chứa các dòng chi phí với mã tài khoản hợp lệ từ dòng 38 trở đi."
        )

    # Normalize account codes to clean integer strings
    df_base[acc_col] = [
        str(int(float(str(v).strip()))) if _is_cost_row(idx + 1, v) else str(v).strip()
        for idx, v in zip(df_base.index, df_base[acc_col])
    ]
    df_current[acc_col] = [
        str(int(float(str(v).strip()))) if _is_cost_row(idx + 1, v) else str(v).strip()
        for idx, v in zip(df_current.index, df_current[acc_col])
    ]

    # Validate and extract numeric cost series (fail-closed on blanks/NaNs/non-numeric)
    df_base[val_col] = _validate_and_extract_cost_series(
        df_base, val_col, acc_col, name_col, "Năm trước", ctx.base_file_path
    )
    df_current[val_col] = _validate_and_extract_cost_series(
        df_current, val_col, acc_col, name_col, "Năm nay", ctx.current_file_path
    )

    # Create combined key
    df_base["_key"] = df_base[acc_col].astype(str).str.strip() + "|" + df_base[name_col].astype(str).str.strip()
    df_current["_key"] = df_current[acc_col].astype(str).str.strip() + "|" + df_current[name_col].astype(str).str.strip()

    # Aggregate duplicate keys to prevent cartesian product inflation
    df_base_agg = df_base.groupby("_key", as_index=False).agg({
        acc_col: "first",
        name_col: "first",
        val_col: "sum"
    })

    df_curr_agg = df_current.groupby("_key", as_index=False).agg({
        acc_col: "first",
        name_col: "first",
        val_col: "sum"
    })

    # Outer join to get all items
    merged = pd.merge(df_base_agg, df_curr_agg, on="_key", how="outer", suffixes=("_base", "_curr"))

    lines = []
    total_base = 0.0
    total_current = 0.0

    for _, row in merged.iterrows():
        # Handle NA from outer join
        b_val_col = f"{val_col}_base"
        c_val_col = f"{val_col}_curr"
        acc_base_col = f"{acc_col}_base"
        acc_curr_col = f"{acc_col}_curr"
        name_base_col = f"{name_col}_base"
        name_curr_col = f"{name_col}_curr"

        b_val = float(row[b_val_col]) if pd.notna(row[b_val_col]) else 0.0
        c_val = float(row[c_val_col]) if pd.notna(row[c_val_col]) else 0.0

        account = str(row[acc_base_col]) if pd.notna(row[acc_base_col]) else str(row[acc_curr_col])
        name = str(row[name_base_col]) if pd.notna(row[name_base_col]) else str(row[name_curr_col])

        # Clean string "nan" just in case
        if account == "nan": account = ""
        if name == "nan": name = ""

        var_abs, var_pct, status = calculate_variance(b_val, c_val)
        is_alert = is_variance_alert(var_abs, var_pct, ctx.threshold_absolute, ctx.threshold_percent)

        lines.append(CostLineVariance(
            account_code=account,
            item_name=name,
            base_value=b_val,
            current_value=c_val,
            variance_absolute=var_abs,
            variance_percent=var_pct,
            status=status,
            is_alert=is_alert
        ))
        total_base += b_val
        total_current += c_val

    tot_abs, tot_pct, _ = calculate_variance(total_base, total_current)

    return VarianceReport(
        context=ctx,
        lines=lines,
        total_base=total_base,
        total_current=total_current,
        total_variance_absolute=tot_abs,
        total_variance_percent=tot_pct
    )

import re

def scan_directories_and_pair_files(base_dir: str, curr_dir: str) -> Tuple[List[Tuple[str, str, str]], List[str]]:
    """
    Scans two directories and pairs Excel files by finding matching substrings (like Cost Center code).
    Returns tuple of (pairs, unmatched_files_list)
    """
    base_files = glob.glob(os.path.join(base_dir, "*.xlsx"))
    curr_files = glob.glob(os.path.join(curr_dir, "*.xlsx"))

    # Extract CC code using regex (assumes 4 or 10 digits explicitly named CC code)
    def extract_cc(filename: str) -> str:
        # Match exactly 10 digits or 4 digits, ignoring surrounding non-digits (like _)
        match = re.search(r'(?<!\d)(\d{10}|\d{4})(?!\d)', filename)
        if match:
            return match.group(1)
        return filename.split(".")[0]

    base_dict = {extract_cc(os.path.basename(f)): f for f in base_files}
    curr_dict = {extract_cc(os.path.basename(f)): f for f in curr_files}

    pairs = []
    unmatched = []
    for cc, b_path in base_dict.items():
        if cc in curr_dict:
            pairs.append((cc, b_path, curr_dict[cc]))
        else:
            unmatched.append(os.path.basename(b_path))

    for cc, c_path in curr_dict.items():
        if cc not in base_dict:
            unmatched.append(os.path.basename(c_path))

    return pairs, unmatched

def batch_analyze_variances(pairs: List[Tuple[str, str, str]], base_fy: int, curr_fy: int, thresh_pct: float, thresh_abs: float) -> Tuple[List[VarianceReport], List[str]]:
    reports = []
    errors = []
    for cc, b_path, c_path in pairs:
        ctx = ComparisonContext(
            fiscal_year_base=base_fy,
            fiscal_year_current=curr_fy,
            cost_center_code=cc,
            base_file_path=b_path,
            current_file_path=c_path,
            threshold_percent=thresh_pct,
            threshold_absolute=thresh_abs
        )
        try:
            df_base = safe_load_mp_form(b_path)
            df_curr = safe_load_mp_form(c_path)
            report = map_and_analyze_variances(df_base, df_curr, ctx)
            reports.append(report)
        except Exception as e:
            errors.append(f"Lỗi file {cc}: {str(e)}")
    return reports, errors
