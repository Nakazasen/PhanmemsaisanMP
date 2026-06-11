"""Fixed-assets reference skeleton writer.

This module writes explicit opt-in, reference-assisted skeleton rows for account
5005026371 from the Phase 42N2E secondary skeleton candidate CSV. Rows written by
this module are not source-derived and must carry provenance in column T.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.engine.column_s_normalizer import cell_has_month_cost, normalize_output_description_column_s

SHEET_NAME = "内訳ﾘｽﾄ(4～3月)"
TARGET_ACCOUNT = "5005026371"
CANDIDATE_CLASSIFICATION = "REFERENCE_ASSISTED_FILL_CANDIDATE"
PROVENANCE_LABEL = (
    "REFERENCE_ASSISTED_SECONDARY_SKELETON; "
    "account=5005026371; scoped-reference-fill; "
    "reason=phase42n2e_reference_assisted_fill_candidate; not source-derived"
)
BUSINESS_CHECK_COLUMNS = (2, 19, *range(6, 18), 20)
MONTH_SAMPLE_COLUMNS = {
    "month_F_sample": 6,
    "month_Q_sample": 17,
}


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _business_row_present(ws, row: int) -> bool:
    return any(_norm(ws.cell(row, col).value) for col in BUSINESS_CHECK_COLUMNS)


def _last_business_row(ws) -> int:
    for row in range(ws.max_row, 0, -1):
        if _business_row_present(ws, row):
            return row
    return 0


def _next_empty_row(ws, start_row: int) -> int:
    row = max(1, start_row)
    while _business_row_present(ws, row):
        row += 1
    return row


def _candidate_key(row: dict[str, str]) -> tuple[str, str, str]:
    return (
        _norm(row.get("account")),
        _norm(row.get("description")),
        _norm(row.get("pattern_signature")),
    )


def load_fixed_assets_skeleton_candidates(csv_path: str | Path) -> list[dict[str, str]]:
    """Load 42N2E fixed-assets candidates only.

    The 42N2E CSV can contain repeated secondary examples. The writer keeps one
    row per account/description/pattern skeleton so repeated department examples
    do not create duplicate workbook rows.
    """
    path = Path(csv_path)
    selected: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if row.get("classification") != CANDIDATE_CLASSIFICATION:
                continue
            if _norm(row.get("account")) != TARGET_ACCOUNT:
                continue
            selected.append(row)
    return selected


def _has_usable_skeleton(candidate: dict[str, str]) -> bool:
    if _norm(candidate.get("account")) != TARGET_ACCOUNT:
        return False
    if not _norm(candidate.get("description")):
        return False
    return any(cell_has_month_cost(candidate.get(column)) for column in MONTH_SAMPLE_COLUMNS)


def _write_candidate(ws, row_index: int, candidate: dict[str, str]) -> None:
    ws.cell(row_index, 2).value = _norm(candidate.get("account"))
    ws.cell(row_index, 19).value = _norm(candidate.get("description"))
    for csv_column, excel_column in MONTH_SAMPLE_COLUMNS.items():
        value = _norm(candidate.get(csv_column))
        if value:
            ws.cell(row_index, excel_column).value = value
    ws.cell(row_index, 20).value = PROVENANCE_LABEL


def apply_fixed_assets_reference_skeleton_to_workbook(
    workbook_path: str | Path,
    csv_path: str | Path,
    start_row: int | None = None,
) -> dict[str, int]:
    """Append fixed-assets reference skeleton rows with explicit provenance."""
    candidates = load_fixed_assets_skeleton_candidates(csv_path)
    wb = load_workbook(workbook_path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {SHEET_NAME}")
        ws = wb[SHEET_NAME]
        resolved_start = int(start_row) if start_row is not None else _last_business_row(ws) + 1
        target_row = _next_empty_row(ws, resolved_start)
        written = 0
        skipped_existing = 0
        skipped_incomplete = 0
        for candidate in candidates:
            if not _has_usable_skeleton(candidate):
                skipped_incomplete += 1
                continue
            if _business_row_present(ws, target_row):
                skipped_existing += 1
                target_row = _next_empty_row(ws, target_row + 1)
            _write_candidate(ws, target_row, candidate)
            written += 1
            target_row = _next_empty_row(ws, target_row + 1)
        normalize_output_description_column_s(ws)
        wb.save(workbook_path)
        return {
            "selected": len(candidates),
            "written": written,
            "skipped_existing": skipped_existing,
            "skipped_incomplete": skipped_incomplete,
            "start_row": resolved_start,
        }
    finally:
        wb.close()
