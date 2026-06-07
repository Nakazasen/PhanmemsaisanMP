"""Read-only Admin consumables file-order preview for MP2027."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ADMIN_CONSUMABLE_ITEMS = (
    ("toilet_paper", "トイレットペーパー", "FY2027予定", "トイレットペーパー"),
    ("hand_soap", "手洗い洗剤", "FY2027予定", "手洗い洗剤"),
    ("alcohol_disinfectant", "アルコール消毒", "Cách tính phân bổ 振替計算", "アルコール消毒"),
)


@dataclass(frozen=True)
class AdminConsumablePreviewItem:
    item_id: str
    display_name: str
    source_workbook: str
    source_sheet: str
    source_row: int | None
    source_description: str | None
    source_month_values_sample: tuple[Any, Any]
    planned_row: int
    formula_policy: str
    confidence: str
    note: str


@dataclass(frozen=True)
class AdminConsumablesFileOrderPreview:
    cost_center: str
    admin_source_path: str
    allocation_source_path: str | None
    planned_start_row: int
    planned_end_row: int
    blank_row_after: int
    items: tuple[AdminConsumablePreviewItem, ...]


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _find_row_containing(worksheet, token: str) -> tuple[int | None, list[Any]]:
    needle = token.lower()
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        joined = " | ".join(_normalize(value) for value in values).lower()
        if needle in joined:
            return row_index, values
    return None, []


def _month_sample(sheet_name: str, values: list[Any]) -> tuple[Any, Any]:
    if not values:
        return None, None
    # FY2027予定 uses B:M as Apr-Mar monthly values.
    if sheet_name == "FY2027予定" and len(values) >= 13:
        return values[1], values[12]
    # Calculation sheet rows use C:N as Apr-Mar totals/unit prices when present.
    if len(values) >= 14:
        return values[2], values[13]
    return None, None


def preview_admin_consumables_file_order(
    admin_source_path: str | Path,
    allocation_source_path: str | Path | None = None,
    cost_center: str | int = "1412000040",
    start_row: int = 207,
) -> AdminConsumablesFileOrderPreview:
    """Build a read-only preview for the Admin consumables mini-batch."""
    admin_source = Path(admin_source_path)
    allocation_source = Path(allocation_source_path) if allocation_source_path else None
    value_wb = load_workbook(admin_source, read_only=True, data_only=True)
    try:
        items: list[AdminConsumablePreviewItem] = []
        for offset, (item_id, display_name, sheet_name, token) in enumerate(ADMIN_CONSUMABLE_ITEMS):
            worksheet = value_wb[sheet_name]
            source_row, row_values = _find_row_containing(worksheet, token)
            description = _normalize(row_values[0]) if row_values else None
            apr, mar = _month_sample(sheet_name, row_values)
            confidence = "HIGH" if source_row is not None else "UNKNOWN"
            formula_policy = "COPY_SOURCE_MONTH_SAMPLE" if apr not in (None, "") or mar not in (None, "") else "UNKNOWN"
            note = "incremental Admin consumables placement after Facility block; not final global file-order"
            if confidence == "UNKNOWN":
                note = "SOURCE_NOT_FOUND_IN_ADMIN_WORKBOOK"
            items.append(
                AdminConsumablePreviewItem(
                    item_id=item_id,
                    display_name=display_name,
                    source_workbook=str(admin_source),
                    source_sheet=sheet_name,
                    source_row=source_row,
                    source_description=description,
                    source_month_values_sample=(apr, mar),
                    planned_row=start_row + offset,
                    formula_policy=formula_policy,
                    confidence=confidence,
                    note=note,
                )
            )
        return AdminConsumablesFileOrderPreview(
            cost_center=str(cost_center),
            admin_source_path=str(admin_source),
            allocation_source_path=str(allocation_source) if allocation_source else None,
            planned_start_row=start_row,
            planned_end_row=start_row + len(items) - 1,
            blank_row_after=start_row + len(items),
            items=tuple(items),
        )
    finally:
        value_wb.close()
